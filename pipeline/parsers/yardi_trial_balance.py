"""
Parser for Yardi Trial Balance export (.xlsx format).

Expected format:
  Row 1: "Revolution Labs Owner, LLC (revlabpm)"
  Row 2: "Trial Balance"
  Row 3: "Period = Apr 2026"
  Row 4: "Book = Accrual ; Tree = ysi_tb"
  Row 5: column group headers (Forward / Ending)
  Row 6: column headers (Balance / Debit / Credit / Balance)
  Row 7+: data rows (account_code, account_name, fwd_balance, debit, credit, ending_balance)
  Last row: Total row (NaN in col A)
"""

import os
import re
from dataclasses import dataclass, field
from typing import List, Optional, Tuple

import openpyxl


@dataclass
class TBAccount:
    account_code: str
    account_name: str
    forward_balance: float
    debit: float
    credit: float
    ending_balance: float

    @property
    def net_activity(self) -> float:
        """Net debit/credit activity for the period."""
        return self.debit - self.credit


@dataclass
class TBMetadata:
    entity_name: str = ''
    period: str = ''
    book: str = ''
    source_file: str = ''


@dataclass
class TBResult:
    metadata: TBMetadata
    accounts: List[TBAccount] = field(default_factory=list)
    total_debits: float = 0.0
    total_credits: float = 0.0
    is_balanced: bool = False

    def get_account(self, code: str) -> Optional[TBAccount]:
        for acct in self.accounts:
            if acct.account_code == code:
                return acct
        return None

    @property
    def account_map(self):
        return {a.account_code: a for a in self.accounts}


def _safe_float(val) -> float:
    if val is None or val == '':
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(',', '').strip())
    except (ValueError, TypeError):
        return 0.0


def _safe_str(val) -> str:
    if val is None:
        return ''
    return str(val).strip()


def parse(filepath: str, sheet_name: str = None) -> TBResult:
    """
    Parse a Yardi Trial Balance export file.

    Args:
        filepath:   Path to .xlsx file.
        sheet_name: Sheet to read (defaults to active sheet).

    Returns:
        TBResult with all accounts and metadata.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    metadata = TBMetadata(source_file=os.path.basename(filepath))

    # Rows 1–4: metadata
    row1 = _safe_str(ws.cell(row=1, column=1).value)
    row3 = _safe_str(ws.cell(row=3, column=1).value)
    row4 = _safe_str(ws.cell(row=4, column=1).value)

    # "Revolution Labs Owner, LLC (revlabpm)"
    match = re.match(r'^(.+?)\s*\((\w+)\)', row1)
    if match:
        metadata.entity_name = match.group(1).strip()
    else:
        metadata.entity_name = row1

    # "Period = Apr 2026"
    if '=' in row3:
        metadata.period = row3.split('=', 1)[1].strip()

    # "Book = Accrual ; Tree = ysi_tb"
    if '=' in row4:
        metadata.book = row4.split('=', 1)[1].split(';')[0].strip()

    accounts: List[TBAccount] = []
    total_debits = 0.0
    total_credits = 0.0

    # Data rows start at row 7 (rows 5-6 are column headers)
    for row_num in range(7, ws.max_row + 1):
        col_a = ws.cell(row=row_num, column=1).value
        col_b = ws.cell(row=row_num, column=2).value

        # Skip blank rows
        if col_a is None and col_b is None:
            continue

        code = _safe_str(col_a)
        name = _safe_str(col_b)

        # Total row: account code is blank/NaN but col_b says "Total"
        if not code or name.lower() == 'total':
            total_debits = _safe_float(ws.cell(row=row_num, column=4).value)
            total_credits = _safe_float(ws.cell(row=row_num, column=5).value)
            continue

        # Skip rows where code is not numeric (headers that slipped through)
        if not code.replace('-', '').replace(' ', '').isdigit():
            continue

        fwd = _safe_float(ws.cell(row=row_num, column=3).value)
        debit = _safe_float(ws.cell(row=row_num, column=4).value)
        credit = _safe_float(ws.cell(row=row_num, column=5).value)
        ending = _safe_float(ws.cell(row=row_num, column=6).value)

        accounts.append(TBAccount(
            account_code=code,
            account_name=name,
            forward_balance=fwd,
            debit=debit,
            credit=credit,
            ending_balance=ending,
        ))

    is_balanced = abs(total_debits - total_credits) < 0.05

    return TBResult(
        metadata=metadata,
        accounts=accounts,
        total_debits=total_debits,
        total_credits=total_credits,
        is_balanced=is_balanced,
    )

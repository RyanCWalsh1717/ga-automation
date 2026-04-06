"""
Yardi Rent Roll Parser

Parses Yardi Rent Roll (Tenancy Schedule) export files with unit-level lease data:
- Row 1: Report title "Tenancy Schedule II"
- Row 2: Property, AsOf date, and notes
- Row 3+: Column headers (spread across multiple rows for multi-level headers)
- Row 7+: Lease and unit data (may have continuation rows for rent steps)
"""

from openpyxl import load_workbook
from datetime import datetime
from typing import List, Dict, Tuple, Optional


def parse(filepath: str) -> List[Dict]:
    """
    Parse a Yardi Rent Roll export file.

    Args:
        filepath: Path to the Excel file

    Returns:
        List of dictionaries representing lease/unit records

    Raises:
        FileNotFoundError: If file does not exist
        ValueError: If file structure is invalid
    """

    def _normalize_header(header):
        if not header:
            return ""
        header = header.strip().lower()
        header = header.replace(' ', '_').replace('/', '_').replace('-', '_')
        while '__' in header:
            header = header.replace('__', '_')
        return header.strip('_')

    def _norm_val(value):
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, str):
            return value.strip()
        return value

    def _get_metadata(ws):
        metadata = {}
        title = ws.cell(1, 1).value
        if title:
            metadata['report_title'] = str(title).strip()
        prop_line = ws.cell(2, 1).value
        if prop_line:
            prop_str = str(prop_line)
            metadata['report_details'] = prop_str
            if "Property:" in prop_str:
                parts = prop_str.split("As of Date:")
                if len(parts) > 1:
                    prop_part = parts[0].split("Property:")[1].strip()
                    metadata['property'] = prop_part.split()[0] if prop_part else None
                    date_part = parts[1].strip().split()[0] if len(parts) > 1 else None
                    if date_part:
                        metadata['as_of_date'] = date_part
        return metadata

    def _get_headers(ws, row):
        headers = []
        for cell in ws[row]:
            value = cell.value
            if value:
                clean_value = str(value).replace('\n', ' ').strip()
                headers.append(clean_value)
            else:
                headers.append(None)
        return headers

    def _build_record(headers, row_values, metadata):
        if not row_values or all(v is None for v in row_values):
            return None
        record = {}
        for i, header in enumerate(headers):
            if header and i < len(row_values):
                value = row_values[i]
                key = _normalize_header(header)
                record[key] = _norm_val(value)
        record.update(metadata)
        return record if any(v is not None for v in record.values()) else None

    def _get_rent_step(headers, row_values):
        rent_step_data = {}
        for i, header in enumerate(headers):
            if header and i < len(row_values):
                value = row_values[i]
                if value is not None:
                    key = _normalize_header(header)
                    rent_step_data[key] = _norm_val(value)
        return rent_step_data if rent_step_data else None

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        raise FileNotFoundError(f"Cannot open file: {filepath}") from e

    ws = wb.active
    data = []

    metadata = _get_metadata(ws)
    headers = _get_headers(ws, row=3)

    if not headers:
        raise ValueError("Cannot extract headers from Rent Roll file")

    current_unit = None

    for row_num in range(7, ws.max_row + 1):
        row = ws[row_num]
        row_values = [cell.value for cell in row]

        if all(v is None for v in row_values):
            continue

        if row_values[0] is not None and str(row_values[0]).startswith("Revolution Labs"):
            record = _build_record(headers, row_values, metadata)
            if record:
                data.append(record)
                current_unit = record
        elif row_values[0] is None and any(v is not None for v in row_values):
            if current_unit is not None:
                rent_step_data = _get_rent_step(headers, row_values)
                if rent_step_data:
                    step_record = dict(current_unit)
                    step_record.update(rent_step_data)
                    data.append(step_record)

    return data


def validate(filepath: str) -> Tuple[bool, List[str]]:
    """
    Validate that a file has the expected Rent Roll structure.

    Args:
        filepath: Path to the Excel file

    Returns:
        Tuple of (is_valid: bool, issues: list of error strings)
    """
    issues = []

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return False, [f"Cannot open file: {e}"]

    ws = wb.active

    if not ws.cell(1, 1).value or "Tenancy Schedule" not in str(ws.cell(1, 1).value):
        issues.append("Row 1 missing 'Tenancy Schedule II' title")

    if not ws.cell(2, 1).value or "Property" not in str(ws.cell(2, 1).value):
        issues.append("Row 2 missing property information")

    headers = []
    for cell in ws[3]:
        value = cell.value
        if value:
            headers.append(str(value).replace('\n', ' ').strip())
        else:
            headers.append(None)

    if not headers:
        issues.append("Cannot extract headers from row 3")
    else:
        if len(headers) < 3:
            issues.append("Expected at least 3 header columns")

    return len(issues) == 0, issues


if __name__ == "__main__":
    import sys
    import json

    if len(sys.argv) < 2:
        print("Usage: python yardi_rent_roll.py <filepath>")
        sys.exit(1)

    filepath = sys.argv[1]

    is_valid, is_issues = validate(filepath)
    if not is_valid:
        print("Validation errors:")
        for issue in is_issues:
            print(f"  - {issue}")
        sys.exit(1)

    data = parse(filepath)
    print(f"Successfully parsed {len(data)} rent roll records")
    print(f"\nSample records (first 2 entries):")
    for i, record in enumerate(data[:2]):
        print(f"\nRecord {i+1}:")
        sample = {k: v for k, v in list(record.items())[:10]}
        print(json.dumps(sample, indent=2, default=str))

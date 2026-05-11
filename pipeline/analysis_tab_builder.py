"""
Analysis Tab Builder — Revolution Labs Monthly Close
=====================================================
Handles the "copy-and-extend" logic for analysis tabs that accumulate
multi-year history from the prior workpaper.

Workflow per tab
----------------
1. Locate the prior period's version in the workbook.  After the carry-forward
   rename (e.g., "Loan Analysis" → "Mar-2026 Loan Analysis") we search by base
   name.
2. Copy all cell values into a new current-period worksheet.
3. Find the insertion point — the "Ending Balance" row that holds the SUM formula.
4. Call ws.insert_rows() to shift that row (and everything below) downward, then
   write the new current-period data rows into the newly vacated space.
5. Update the SUM formula to include the new rows.
6. Rebuild the GL / TB tie-out from live TB data (replaces the JLL VLOOKUP refs
   that break after the carry-forward rename).

Tab coverage
------------
  115200 Escrow RET          — GL 115200 + Berkadia payment_re_taxes
  115300 Escrow Insurance    — GL 115300 + Berkadia insurance escrow
  115600 Restricted Cash     — GL 115600 current-period transactions
  RE Tax Analysis            — GL 115200 + Berkadia escrow deposit
  Insurance Analysis         — Prepaid ledger (column-based, monthly amortization)
  Loan Analysis              — GL 213200 / 801110 + Berkadia per-loan interest data
"""

import re
from copy import copy
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl.utils import column_index_from_string, get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# Month / period helpers
# ─────────────────────────────────────────────────────────────────────────────

_MONTH_ABBR = {
    1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
    7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec',
}
_MONTH_NUM = {v: k for k, v in _MONTH_ABBR.items()}


def _period_to_dt(period: str) -> Optional[datetime]:
    """'Apr-2026' or 'Apr 2026' → datetime(2026, 4, 1)"""
    if not period:
        return None
    for sep in ('-', ' '):
        parts = period.strip().split(sep, 1)
        if len(parts) == 2:
            mon = _MONTH_NUM.get(parts[0][:3].title())
            if mon:
                try:
                    return datetime(int(parts[1].strip()), mon, 1)
                except ValueError:
                    pass
    return None


def _fmt_mmy(period: str) -> str:
    """'Apr-2026' → '04/26'"""
    d = _period_to_dt(period)
    return d.strftime('%m/%y') if d else ''


def _fmt_long(period: str) -> str:
    """'Apr-2026' → 'Apr 2026'"""
    d = _period_to_dt(period)
    return d.strftime('%b %Y') if d else period or ''


def _prior_long(period: str) -> str:
    """'Apr-2026' → 'Mar 2026'"""
    d = _period_to_dt(period)
    if not d:
        return ''
    prev_month = d.month - 1 if d.month > 1 else 12
    prev_year  = d.year if d.month > 1 else d.year - 1
    return f"{_MONTH_ABBR[prev_month]} {prev_year}"


def _quarter_label(period: str) -> str:
    """'Apr-2026' → 'Q2-2026'"""
    d = _period_to_dt(period)
    if not d:
        return ''
    q = (d.month - 1) // 3 + 1
    return f'Q{q}-{d.year}'


def _safe_float(v) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


# ─────────────────────────────────────────────────────────────────────────────
# Prior-tab discovery
# ─────────────────────────────────────────────────────────────────────────────

def _find_prior_tab(wb, base_names: List[str], current_prefix: str):
    """
    Find the prior-period copy of an analysis tab.

    Searches for any sheet whose lowercase name CONTAINS one of the base_names
    (stripped, lowercase) and does NOT start with current_prefix.

    Returns (worksheet, sheet_name) or (None, None).
    """
    bases = [b.lower().strip() for b in base_names]
    for sname in wb.sheetnames:
        if sname.startswith(current_prefix):
            continue
        sl = sname.lower()
        for b in bases:
            if b in sl:
                return wb[sname], sname
    return None, None


# ─────────────────────────────────────────────────────────────────────────────
# Content copying  (values only — avoids broken cross-sheet formula refs)
# ─────────────────────────────────────────────────────────────────────────────

def _copy_tab_values(source_ws, target_ws):
    """
    Copy all cell values, row heights, column widths, and merged ranges
    from source_ws to target_ws.

    Formulas are copied as formula strings (not evaluated values).  After the
    carry-forward rename, cross-sheet refs like ='General '!A5 will show as
    #REF in Excel for historical data — acceptable since those values were
    already finalised.  The current-period tie-out is rebuilt fresh by
    _rebuild_tieout(), which overwrites the stale VLOOKUP cells.
    """
    # Column widths
    for col_letter, cdim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = cdim.width

    # Row heights
    for row_num, rdim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_num].height = rdim.height

    # Cell values and styles
    for row in source_ws.iter_rows():
        for cell in row:
            tc = target_ws.cell(row=cell.row, column=cell.column)
            tc.value = cell.value
            if cell.has_style:
                try:
                    tc.font       = copy(cell.font)
                    tc.fill       = copy(cell.fill)
                    tc.border     = copy(cell.border)
                    tc.alignment  = copy(cell.alignment)
                    tc.number_format = cell.number_format
                except Exception:
                    pass

    # Merged cells
    for merge_range in list(source_ws.merged_cells.ranges):
        try:
            target_ws.merge_cells(str(merge_range))
        except Exception:
            pass

    # Tab colour
    if source_ws.sheet_properties.tabColor:
        target_ws.sheet_properties.tabColor = source_ws.sheet_properties.tabColor


# ─────────────────────────────────────────────────────────────────────────────
# Insertion-point detection
# ─────────────────────────────────────────────────────────────────────────────

def _find_insertion_point(ws, amount_col: int) -> Dict[str, Any]:
    """
    Scan the worksheet to find the structural boundaries.

    Strategy:
      1. Look for a cell in amount_col containing '=SUM(' — that is the
         "Ending Balance" / total row.
      2. As a fallback, look for a cell in any column containing the literal
         string pattern "='General " (the JLL "Ending Balance per GL" label row).

    Handles both simple SUM formulas (=SUM(F8:F37)) and compound ones
    (=SUM(F60:F75)+SUM(F96:F112)).  For compound formulas the insertion
    point is the total row itself; new rows are inserted before it and a
    new SUM component is appended rather than rewriting the entire formula.

    Returns dict with keys:
        insert_before_row   — row to insert new data BEFORE
        sum_start_row       — first row of the LAST SUM range (used for
                              the new appended component)
        sum_end_row         — last row of the LAST SUM range
        total_row           — row holding the SUM formula
        gl_row              — row with 'GL' label (may be None)
        variance_row        — row with 'Variance' label (may be None)
        compound_sum        — True if the formula is a multi-part SUM
        original_formula    — the original formula string (for compound rebuild)
    """
    result: Dict[str, Any] = {
        'insert_before_row': None,
        'sum_start_row':     None,
        'sum_end_row':       None,
        'total_row':         None,
        'gl_row':            None,
        'variance_row':      None,
        'compound_sum':      False,
        'original_formula':  None,
    }

    max_row = ws.max_row or 1

    for r in range(max_row, 0, -1):
        # ── Check adjacent cells for GL / Variance labels ──────────────
        for c in range(1, min(amount_col + 3, 20)):
            adj = ws.cell(row=r, column=c).value
            if isinstance(adj, str):
                al = adj.strip().lower()
                if al == 'variance' and result['variance_row'] is None:
                    result['variance_row'] = r
                elif al in ('gl', 'g/l', 'gl balance') and result['gl_row'] is None:
                    result['gl_row'] = r

        # ── Check amount column for SUM formula ───────────────────────
        cell_val = ws.cell(row=r, column=amount_col).value
        if isinstance(cell_val, str) and 'SUM(' in cell_val.upper():
            result['total_row'] = r
            result['insert_before_row'] = r
            result['original_formula'] = cell_val

            # Find ALL SUM ranges in the formula (handles compound formulas)
            all_ranges = re.findall(
                r'SUM\(\s*[A-Z]+(\d+)\s*:\s*[A-Z]+(\d+)\s*\)',
                cell_val, re.I,
            )
            if all_ranges:
                # Store the LAST range's start/end — new rows go just before
                # the total row, so the new component appends after the last range
                result['sum_start_row'] = int(all_ranges[-1][0])
                result['sum_end_row']   = int(all_ranges[-1][1])
                result['compound_sum']  = len(all_ranges) > 1
            break

        # ── Fallback: JLL "='General '!A10" ending-balance label ───────
        if result['insert_before_row'] is None:
            for c in range(1, 20):
                adj = ws.cell(row=r, column=c).value
                if isinstance(adj, str) and adj.strip().startswith("='General"):
                    result['insert_before_row'] = r
                    break

    return result


# ─────────────────────────────────────────────────────────────────────────────
# Row-insertion + formula update
# ─────────────────────────────────────────────────────────────────────────────

def _insert_rows_and_write(
    ws,
    ip: Dict[str, Any],
    new_rows: List[Dict[str, Any]],
    amount_col: int,
    period: str,
    tb_map: Optional[dict],
    account_code: str,
):
    """
    1. Insert len(new_rows) blank rows at ip['insert_before_row'].
    2. Write new_rows data.
    3. Update the SUM formula at the (now-shifted) total row.
    4. Rebuild the GL / TB tie-out section.

    Each item in new_rows is a dict mapping column letters to values,
    e.g. {'B': 'Description', 'F': 1234.56}.
    """
    n = len(new_rows)
    if n == 0:
        return

    insert_at = ip.get('insert_before_row') or (ws.max_row + 2)

    # ── 1. Insert blank rows ──────────────────────────────────────────
    ws.insert_rows(insert_at, amount=n)

    # ── 2. Write data into the new blank rows ─────────────────────────
    for i, row_data in enumerate(new_rows):
        r = insert_at + i
        for col_letter, val in row_data.items():
            if not col_letter:
                continue
            try:
                col_idx = column_index_from_string(col_letter.strip().upper())
            except Exception:
                continue
            cell = ws.cell(row=r, column=col_idx)
            cell.value = val
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0.00;(#,##0.00);"-"'

    # ── 3. Update SUM formula ─────────────────────────────────────────
    if ip.get('total_row'):
        new_total_row = ip['total_row'] + n
        col_l         = get_column_letter(amount_col)

        if ip.get('compound_sum') and ip.get('original_formula'):
            # Compound formula (e.g. =SUM(F60:F75)+SUM(F96:F112)).
            # Preserve the original ranges exactly — they didn't shift because
            # insertion happened AFTER them (before the total row).
            # Append a new component for the rows we just inserted.
            new_component_start = insert_at          # first new row
            new_component_end   = insert_at + n - 1  # last new row
            new_formula = (
                ip['original_formula'].rstrip()
                + f'+SUM({col_l}{new_component_start}:{col_l}{new_component_end})'
            )
            ws.cell(row=new_total_row, column=amount_col).value = new_formula

        elif ip.get('sum_start_row'):
            # Simple single-range SUM — extend the end row.
            new_sum_end = (ip['sum_end_row'] or ip['total_row'] - 1) + n
            ws.cell(row=new_total_row, column=amount_col).value = (
                f'=SUM({col_l}{ip["sum_start_row"]}:{col_l}{new_sum_end})'
            )

    # ── 4. Rebuild tie-out ────────────────────────────────────────────
    # Locate where the GL / Variance rows ended up after insertion
    gl_row  = (ip.get('gl_row')       or 0) + n if ip.get('gl_row')       else None
    var_row = (ip.get('variance_row') or 0) + n if ip.get('variance_row') else None
    total_row_final = (ip.get('total_row') or 0) + n if ip.get('total_row') else None

    _rebuild_tieout(ws, total_row_final, gl_row, var_row, amount_col, tb_map, account_code)


def _rebuild_tieout(
    ws,
    total_row: Optional[int],
    gl_row: Optional[int],
    var_row: Optional[int],
    amount_col: int,
    tb_map: Optional[dict],
    account_code: str,
):
    """
    Overwrite the GL and Variance cells with fresh values from tb_map.
    This replaces the stale JLL VLOOKUP references that break after the
    carry-forward rename.

    Also updates the account-code cell (used by the VLOOKUP key) to a
    plain integer so it reads cleanly even without the General tab.
    """
    tb_acct   = (tb_map or {}).get(account_code)
    tb_ending = tb_acct.ending_balance if tb_acct else None
    col_l     = get_column_letter(amount_col)

    # Update GL cell
    if gl_row:
        gl_cell = ws.cell(row=gl_row, column=amount_col)
        if tb_ending is not None:
            gl_cell.value = tb_ending
            gl_cell.number_format = '#,##0.00;(#,##0.00);"-"'
        else:
            gl_cell.value = 'Not in TB'

    # Update Variance cell
    if var_row and total_row and gl_row:
        ws.cell(row=var_row, column=amount_col).value = (
            f'={col_l}{gl_row}-{col_l}{total_row}'
        )
        ws.cell(row=var_row, column=amount_col).number_format = '#,##0.00;(#,##0.00);"-"'

    # Fix account-code key cell (row just above GL label row)
    if gl_row and gl_row >= 3:
        # The account-code cell is typically 2 rows above GL in JLL layout
        key_row = gl_row - 2
        for c in range(1, amount_col + 2):
            v = ws.cell(row=key_row, column=c).value
            # If cell holds a formula referencing General tab, overwrite it
            if isinstance(v, str) and ("='General" in v or account_code in str(v)):
                ws.cell(row=key_row, column=amount_col).value = (
                    int(account_code) if account_code.isdigit() else account_code
                )
                break
            elif str(v) == account_code:
                break  # already a plain value


# ─────────────────────────────────────────────────────────────────────────────
# Analysis tab seed data — bootstraps first-run tabs without a prior workpaper
# Source: Analysis Tabs Workpapers.xlsx  (manual workpaper as of Jan-2026)
# ─────────────────────────────────────────────────────────────────────────────

# RE Tax Analysis seed rows
# Format: (description, date_str_or_None, acct_135120, acct_641110)
# None = blank cell  |  Amounts follow GL sign convention
_RET_ANALYSIS_SEED: List[Tuple] = [
    # ── Parcel 00011619 ─────────────────────────────────────────────────────
    ('PARCEL ID:  00011619', None, None, None),
    ('3QTR Payment by client covering Jan thru March', '01/25/2023', 796749.49, None),
    ('Reclass January Tax Exp fr PPD',  '01/25/2023', -265583.00, 265583.00),
    ('Reclass Feb Tax Exp fr PPD',       '02/25/2023', -265583.00, 265583.00),
    ('Reclass Mar Tax Exp fr PPD',       '03/25/2023', -265583.49, 265583.49),
    ('4QTR Payment by client covering Apr thru June', '04/25/2023', 796748.99, None),
    ('Reclass April Tax Exp fr PPD',    '04/25/2023', -265583.00, 265583.00),
    ('Reclass May Tax Exp fr PPD',      '05/25/2023', -265583.00, 265583.00),
    ('Reclass June Tax Exp fr PPD',     '06/25/2023', -265582.99, 265582.99),
    ('FY23/24 1QTR Payment covering July thru Sept', '07/14/2023', None, 433692.41),
    ('Reclass Aug and Sept Exp to PPD', '07/25/2023', 289128.27, -289128.27),
    ('Reclass Aug Tax Exp fr PPD',      '08/25/2023', -144564.14, 144564.14),
    ('Reclass Sept Tax Exp fr PPD',     '09/25/2023', -144564.14, 144564.14),
    ('Accr 9/23 Tax Increase',          '09/25/2023', None, 239204.01),
    ('FY23/24 2QTR Payment covering Oct thru Dec-2023', '10/10/2023', None, 433692.41),
    ('Reclass Nov and Dec -2023 Exp to PPD', '10/25/2023', 289128.27, -289128.27),
    ('Accr 10/23 Tax Increase',         '10/25/2023', None, 86935.84),
    ('Reclass Nov Tax Exp fr PPD',      '11/25/2023', -144564.14, 144564.14),
    ('Accr 11/23 Tax Increase',         '11/25/2023', None, 86935.84),
    ('Reclass Dec Tax Exp fr PPD',      '12/25/2023', -144564.14, 144564.14),
    ('Accr 12/23 Tax Increase',         '12/25/2023', None, 86935.84),
    ('FY 24- Q3 Payment Covering Jan thru March 2024', '01/25/2024', None, 893567.25),
    ('RE Tax Accrual Adjustment - Q3-FY24', None, None, -229943.16),
    ('Reclass Feb -24 & March 24',      None, 595711.50, -595711.50),
    ('RE Tax Accrual Adjustment',       None, -153295.44, 153295.44),
    ('Amortize Feb-24 RET Expense',     None, -297855.75, 297855.75),
    ('Feb-24 Tax Accrual Adjustment',   None, 76647.72, -76647.72),
    ('Amortize March-24 RET Expense',   None, -297855.75, 297855.75),
    ('Reverse Feb accrual adj to zero out', None, 76647.72, -76647.72),
    ('FY 24- Q4 Payment Covering April thru June 2024', '04/05/2024', None, 893567.24),
    ('Reclass May -24 & June 24',       None, 595711.49, -595711.49),
    ('RE Tax Accrual Adjustment - Q4-FY24', None, -229943.16, None),
    ('RE Tax due per 5.10.24 Loan Stmt', None, 76647.72, -76647.72),
    ('Amortize 5.24 RET Expense (a)',   None, -297855.75, 297855.75),
    ('Amortize Q3-FY24 Accrual Adjustment (a)', None, 76647.72, -76647.72),
    ('Amortize 5.24 RET Expense (b)',   None, -297855.74, 297855.74),
    ('Amortize Q3-FY24 Accrual Adjustment (b)', None, 76647.72, -76647.72),
    ('RE Tax Payment FY 2025 (1st Quarter) July-Sept 2024', '07/25/2024', None, 680220.57),
    ('Reclass Aug-Sept 2024 RE Tax',    '07/25/2024', 453480.38, -453480.38),
    ('Amortize 8.24 RET Tax Expense',   '08/25/2024', -226740.19, 226740.19),
    ('Amortize 09.24 RET Tax Expense',  '09/25/2024', -226740.19, 226740.19),
    ('RE Tax Payment FY 2025 (2nd Quarter) Oct-Dec 2024', '10/17/2024', None, 680220.57),
    ('Reclass Nov-Dec 2024 RE Tax',     '10/25/2024', 453480.38, -453480.38),
    ('Amortize 11.24 RET Tax Expense',  '11/25/2024', -226740.19, 226740.19),
    ('Amortize 12.24 RET Tax Expense',  '12/25/2024', -226740.19, 226740.19),
    ('Accr: True-up the RE Tax expense', '12/25/2024', None, 1301023.33),
    ('RE Tax Payment FY 2025 (3rd Quarter) Jan - Mar 2025', '01/16/2025', None, 651404.28),
    ('Reclass Feb-Mar 2025 RE Tax',     '01/25/2025', 434269.52, -434269.52),
    ('Amort: Feb 2025 RE Tax',          '02/25/2025', -217134.76, 217134.76),
    ('Amort: Mar 2025 RE Tax',          '03/25/2025', -217134.76, 217134.76),
    ('RE Tax Payment FY 2025 (4th Quarter) April - June 2025', '04/16/2025', None, 651404.28),
    ('Reclass May-June 2025 RE Tax',    '04/25/2025', 434269.52, -434269.52),
    ('Amort: May 2025 RE Tax',          '05/25/2025', -217134.76, 217134.76),
    ('Amort: June 2025 RE Tax',         '06/25/2025', -217134.76, 217134.76),
    ('RE Tax Payment FY 2025 (4th Quarter) July - Sep 2025', '07/17/2025', None, 682444.67),
    ('Reclass Aug-Sep 2025 RE Tax',     '07/25/2025', 454963.11, -454963.11),
    ('Amort: Aug 2025 RE Tax',          '08/25/2025', -227481.56, 227481.56),
    ('Amort: Sep 2025 RE Tax',          '09/25/2025', -227481.55, 227481.55),
    ('RE Tax Payment FY 2025 (4th Quarter) Oct - Dec 2025', '10/25/2025', None, 682457.73),
    ('Reclass Nov-Dec 2025 RE Tax',     '10/25/2025', 454971.82, -454971.82),
    ('Amort: Nov 2025 RE Tax',          '11/25/2025', -227485.91, 227485.91),
    ('Amort: Dec 2025 RE Tax',          '12/25/2025', -227485.91, 227485.91),
    ('RE Tax Payment FY 2026 (1st Quarter) Jan - March 2026', '01/25/2026', None, 498494.07),
    ('Reclass Feb-March 2026 RE Tax',   '01/25/2026', 332329.38, -332329.38),
    # ── Parcel R0140050002 ──────────────────────────────────────────────────
    ('PARCEL ID:  R0140050002', None, None, None),
    ('4QTR Payment by Client covering Jan thru March', '01/25/2024', None, 206.01),
    ('Reclass Feb & March 2024 to PPD', '01/25/2024', 137.34, -137.34),
    ('Reclass Feb Tax Exp fr PPD (R2)',  '02/25/2023', -68.67, 68.67),
    ('Reclass Mar Tax Exp fr PPD (R2)',  '03/25/2023', -68.67, 68.67),
    ('FY24-Q4 WALTHAM',                 '04/25/2024', None, 206.00),
    ('Reclass May & June 2024 to PPD',  '04/25/2024', 137.33, -137.33),
    ('Amortize May-2024 Exp',           '05/25/2024', -68.67, 68.67),
    ('Amortize June-2024 Exp',          '05/25/2024', -68.67, 68.67),
    ('Massachusetts Appellate Tax Board payment', '06/25/2024', None, 5000.00),
    ('FY25-Q1 WALTHAM',                 '07/25/2024', None, 200.27),
    ('Reclass Aug & Sept 2024 to PPD',  '07/25/2024', 133.51, -133.51),
    ('Amortize 8.2024 RET Tax Expense', '08/25/2024', -66.76, 66.76),
    ('Amortize 09.2024 RET Tax Expense', '09/25/2024', -66.75, 66.75),
    ('FY25-Q2 WALTHAM',                 '10/17/2024', None, 200.27),
    ('Reclass Nov-Dec 2024 to PPD',     '10/25/2024', 133.51, -133.51),
    ('Amortize 11.2024 RET Tax Expense', '11/25/2024', -66.75, 66.75),
    ('Amortize 12.2024 RET Tax Expense', '12/25/2024', -66.76, 66.76),
    ('FY25-Q3 WALTHAM',                 '01/16/2025', None, 226.41),
    ('Reclass Feb-Mar 2025 RE Tax (R2)', '01/25/2025', 150.94, -150.94),
    ('Amort: Feb 2025 RE Tax (R2)',      '02/25/2025', -75.47, 75.47),
    ('Amort: Mar 2025 RE Tax (R2)',      '03/25/2025', -75.47, 75.47),
    ('FY25-Q4 WALTHAM',                 '04/16/2025', None, 226.41),
    ('Reclass May-June 2025 RE Tax (R2)', '04/25/2025', 150.94, -150.94),
    ('Amort: May 2025 RE Tax (R2)',      '05/25/2025', -75.47, 75.47),
    ('Amort: June 2025 RE Tax (R2)',     '06/25/2025', -75.47, 75.47),
    ('Tax Board Filing Fee 2025',        '06/25/2025', None, 5000.00),
    ('RE Tax Payment July - Sep 2025 (R2)', '07/17/2025', None, 226.41),
    ('Reclass Aug-Sep 2025 RE Tax (R2)', '07/25/2025', 150.94, -150.94),
    ('Amort: Aug 2025 RE Tax (R2)',      '08/25/2025', -75.47, 75.47),
    ('Amort: Sep 2025 RE Tax (R2)',      '09/25/2025', -75.47, 75.47),
    ('RE Tax Payment Oct - Dec 2025 (R2)', '10/25/2025', None, 213.34),
    ('Reclass Nov-Dec 2025 RE Tax (R2)', '10/25/2025', 142.23, -142.23),
    ('Amort: Nov 2025 RE Tax (R2)',      '11/25/2025', -71.11, 71.11),
    ('Amort: Dec 2025 RE Tax (R2)',      '12/25/2025', -71.12, 71.12),
    ('RE Tax Payment FY 2026 Jan - March 2026 (R2)', '01/25/2025', None, 256.74),
    ('Reclass Feb-March 2026 RE Tax (R2)', '01/25/2025', 171.16, -171.16),
]
# Ending balance as of Jan-2026 workpaper
_RET_ANALYSIS_ENDING = {'135120': 332500.54, '641110': 166250.27}

# Loan Analysis seed rows
# Format: (loan_num_or_None, description, date_str_or_None,
#           acct_231100_rl, acct_231100_rpm, acct_213200, acct_801110)
# None = blank.  Starting mortgage balance is a single header row.
_LOAN_ANALYSIS_SEED: List[Tuple] = [
    # ── Starting balance 12/31/24 ────────────────────────────────────────
    (None, 'Balance at 12/31/24-Dec interest accrual', None, -92104195.17, None, -701314.29, None),
    # ── Jan 2025 ─────────────────────────────────────────────────────────
    (None,    'Dec 2024 Accrual Reversal',          '01/25', None, None, 701314.29,  -701314.29),
    (1159010, 'Dec Interest Payment',               '01/25', None, None, None,        451630.77),
    (1159011, 'Dec Interest Payment',               '01/25', None, None, None,        224715.17),
    (1159012, 'Dec Interest Payment',               '01/25', None, None, None,         24968.35),
    (1159010, 'Accr Jan Interest Due -11159010',    '01/25', None, None, -440923.66,  440923.66),
    (1159011, 'Accr Jan Interest Due -11159011',    '01/25', None, None, -221503.04,  221503.04),
    (1159012, 'Accr Jan Interest Due -11159012',    '01/25', None, None,  -24611.45,   24611.45),
    # ── Feb 2025 ─────────────────────────────────────────────────────────
    (None,    'Jan 2025 Accrual Reversal',          '02/25', None, None, 687038.15,  -687038.15),
    (1159010, 'Jan Interest Payment',               '02/25', None, None, None,        440923.66),
    (1159011, 'Jan Interest Payment',               '02/25', None, None, None,        221503.04),
    (1159012, 'Jan Interest Payment',               '02/25', None, None, None,         24611.45),
    (1159010, 'Accr Feb Interest Due -11159010',    '02/25', None, None, -398952.08,  398952.08),
    (1159011, 'Accr Feb Interest Due -11159011',    '02/25', None, None, -200276.80,  200276.80),
    (1159012, 'Accr Feb Interest Due -11159012',    '02/25', None, None,  -22252.98,   22252.98),
    # ── Mar 2025 ─────────────────────────────────────────────────────────
    (None,    'Feb 2025 Accrual Reversal',          '03/25', None, None, 621481.86,  -621481.86),
    (1159010, 'Feb Interest Payment',               '03/25', None, None, None,        398952.08),
    (1159011, 'Feb Interest Payment',               '03/25', None, None, None,        200276.80),
    (1159012, 'Feb Interest Payment',               '03/25', None, None, None,         22252.98),
    (1159010, 'Accr Mar Interest Due -11159010',    '03/25', None, None, -442351.27,  442351.27),
    (1159011, 'Accr Mar Interest Due -11159011',    '03/25', None, None, -221931.32,  221931.32),
    (1159012, 'Accr Mar Interest Due -11159012',    '03/25', None, None,  -24659.04,   24659.04),
    # ── Apr 2025 ─────────────────────────────────────────────────────────
    (None,    'Mar 2025 Accrual Reversal',          '04/25', None, None, 688941.63,  -688941.63),
    (1159010, 'Mar Interest Payment',               '04/25', None, None, None,        442351.27),
    (1159011, 'Mar Interest Payment',               '04/25', None, None, None,        221931.32),
    (1159012, 'Mar Interest Payment',               '04/25', None, None, None,         24659.04),
    (1159010, 'Accr April Interest Due -11159010',  '04/25', None, None, -426354.98,  426354.98),
    (1159011, 'Accr April Interest Due -11159011',  '04/25', None, None, -214254.16,  214254.16),
    (1159012, 'Accr April Interest Due -11159012',  '04/25', None, None,  -23806.02,   23806.02),
    # ── May 2025 ─────────────────────────────────────────────────────────
    (None,    'April 2025 Accrual Reversal',        '05/25', None, None, 664415.16,  -664415.16),
    (1159010, 'April Interest Payment',             '05/25', None, None, None,        426354.92),
    (1159011, 'April Interest Payment',             '05/25', None, None, None,        214254.16),
    (1159012, 'April Interest Payment',             '05/25', None, None, None,         23806.02),
    (1159010, 'Accr May Interest Due -11159010',    '05/25', None, None, -443005.60,  443005.60),
    (1159011, 'Accr May Interest Due -11159011',    '05/25', None, None, -222127.62,  222127.62),
    (1159012, 'Accr May Interest Due -11159012',    '05/25', None, None,  -24680.85,   24680.85),
    # ── Jun 2025 ─────────────────────────────────────────────────────────
    (None,    'May 2025 Accrual Reversal',          '06/25', None, None, 689814.07,  -689814.07),
    (1159010, 'May Interest Payment',               '06/25', None, None, None,        443005.60),
    (1159011, 'May Interest Payment',               '06/25', None, None, None,        222127.62),
    (1159012, 'May Interest Payment',               '06/25', None, None, None,         24680.85),
    (None,    'June 2025 Accrual Reversal',         '06/25', None, None, 665950.17,  -665950.17),
    (1159010, 'Accr June Interest Due -11159010',   '06/25', None, None, -427506.22,  427506.22),
    (1159011, 'Accr June Interest Due -11159011',   '06/25', None, None, -214599.55,  214599.55),
    (1159012, 'Accr June Interest Due -11159012',   '06/25', None, None,  -23844.40,   23844.40),
    # ── Jul 2025 ─────────────────────────────────────────────────────────
    (1159010, 'June Interest Payment',              '07/25', None, None, None,        427506.22),
    (1159011, 'June Interest Payment',              '07/25', None, None, None,        214599.55),
    (1159012, 'June Interest Payment',              '07/25', None, None, None,         23844.40),
    (1159010, 'Accr July Interest Due -11159010',   '07/25', None, None, -443362.50,  443362.50),
    (1159011, 'Accr July Interest Due -11159011',   '07/25', None, None, -222234.69,  222234.69),
    (1159012, 'Accr July Interest Due -11159012',   '07/25', None, None,  -24692.74,   24692.74),
    # ── Aug 2025 ─────────────────────────────────────────────────────────
    (None,    'July 2025 Accrual Reversal',         '08/25', None, None, 690289.93,  -690289.93),
    (1159010, 'july Interest Payment',              '08/25', None, None, None,        443362.50),
    (1159011, 'july Interest Payment',              '08/25', None, None, None,        222234.69),
    (1159012, 'july Interest Payment',              '08/25', None, None, None,         24692.74),
    (None,    'August 2025 Accrual Reversal',       '08/25', None, None, 691796.86,  -691796.86),
    (1159010, 'Accr Aug Interest Due -11159010',    '08/25', None, None, -444492.69,  444492.69),
    (1159011, 'Accr Aug Interest Due -11159011',    '08/25', None, None, -222573.75,  222573.75),
    (1159012, 'Accr Aug Interest Due -11159012',    '08/25', None, None,  -24730.42,   24730.42),
    # ── Sep 2025 ─────────────────────────────────────────────────────────
    (1159010, 'Aug Interest Payment',               '09/25', None, None, None,        444492.69),
    (1159011, 'Aug Interest Payment',               '09/25', None, None, None,        222573.75),
    (1159012, 'Aug Interest Payment',               '09/25', None, None, None,         24730.42),
    (1159010, 'Accr Sep Interest Due -11159010',    '09/25', None, None, -420655.98,  420655.98),
    (1159011, 'Accr Sep Interest Due -11159011',    '09/25', None, None, -212544.48,  212544.48),
    (1159012, 'Accr Sep Interest Due -11159012',    '09/25', None, None,  -23616.05,   23616.05),
    # ── Oct 2025 ─────────────────────────────────────────────────────────
    (None,    'September 2025 Accrual Reversal',    '10/25', None, None, 656816.51,  -656816.51),
    (1159010, 'Sep Interest Payment',               '10/25', None, None, None,        420655.98),
    (1159011, 'Sep Interest Payment',               '10/25', None, None, None,        212544.48),
    (1159012, 'Sep Interest Payment',               '10/25', None, None, None,         23616.05),
    (1159010, 'Accr Oct Interest Due -11159010',    '10/25', None, None, -427896.67,  427896.67),
    (1159011, 'Accr Oct Interest Due -11159011',    '10/25', None, None, -217594.94,  217594.94),
    (1159012, 'Accr Oct Interest Due -11159012',    '10/25', None, None,  -24177.22,   24177.22),
    # ── Nov 2025 ─────────────────────────────────────────────────────────
    (None,    'October 2025 Accrual Reversal',      '11/25', None, None, 669668.83,  -669668.83),
    (1159010, 'Nov Interest Payment',               '11/25', None, None, None,        427896.67),
    (1159011, 'Nov Interest Payment',               '11/25', None, None, None,        217594.94),
    (1159012, 'Nov Interest Payment',               '11/25', None, None, None,         24177.22),
    (1159010, 'Accr Nov Interest Due -11159010',    '11/25', None, None, -407934.08,  407934.08),
    (1159011, 'Accr Nov Interest Due -11159011',    '11/25', None, None, -208727.91,  208727.91),
    (1159012, 'Accr Nov Interest Due -11159012',    '11/25', None, None,  -23191.99,   23191.99),
    # ── Dec 2025 ─────────────────────────────────────────────────────────
    (None,    'November 2025 Accrual Reversal',     '12/25', None, None, 639853.98,  -639853.98),
    (1159010, 'Dec Interest Payment',               '12/25', None, None, None,        407934.08),
    (1159011, 'Dec Interest Payment',               '12/25', None, None, None,        208727.91),
    (1159012, 'Dec Interest Payment',               '12/25', None, None, None,         23191.99),
    (1159010, 'Accr Dec Interest Due -11159010',    '12/25', None, None, -409694.58,  409694.58),
    (1159011, 'Accr Dec Interest Due -11159011',    '12/25', None, None, -212134.31,  212134.31),
    (1159012, 'Accr Dec Interest Due -11159012',    '12/25', None, None,  -23570.48,   23570.48),
    # ── Jan 2026 ─────────────────────────────────────────────────────────
    (None,    'December 2025 Accrual Reversal',     '01/26', None, None, 645399.37,  -645399.37),
    (1159010, 'Jan Interest Payment',               '01/26', None, None, None,        409694.58),
    (1159011, 'Jan Interest Payment',               '01/26', None, None, None,        212134.31),
    (1159012, 'Jan Interest Payment',               '01/26', None, None, None,         23570.48),
    (1159010, 'Accr Jan Interest Due -11159010',    '01/26', None, None, -403805.67,  403805.67),
    (1159011, 'Accr Jan Interest Due -11159011',    '01/26', None, None, -210367.64,  210367.64),
    (1159012, 'Accr Jan Interest Due -11159012',    '01/26', None, None,  -23374.18,   23374.18),
]
# Ending balance as of Jan-2026 workpaper
_LOAN_ANALYSIS_ENDING = {
    '231100_rl':  -92104195.17,   # Revlab mortgage payable
    '231100_rpm':          0.00,  # Revlabpm (none)
    '213200':      -637547.49,    # Accrued interest payable
    '801110':       637547.49,    # Interest expense
}


# ─────────────────────────────────────────────────────────────────────────────
# Seed tab writers  (called instead of _write_stub when no prior workpaper)
# ─────────────────────────────────────────────────────────────────────────────

def _write_seed_ret_analysis(ws, period: str, tb_map: Optional[dict]):
    """
    Build a full-history RE Tax Analysis tab from _RET_ANALYSIS_SEED.
    Replaces _write_stub when no prior workpaper is provided.

    Column layout:
      A (1): blank gutter
      B (2): Description
      D (4): Date
      F (6): A/C 135120 Prepaid RE Taxes
      H (8): A/C 641110 RE Tax Expense
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import column_index_from_string as _ci

    _hdr_fill = PatternFill(start_color='375623', end_color='375623', fill_type='solid')
    _hdr_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    _bold     = Font(name='Calibri', size=11, bold=True)
    _std      = Font(name='Calibri', size=11)
    _alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _num_fmt  = '#,##0.00;(#,##0.00);"-"'
    THIN      = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 3
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 3
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 3
    ws.column_dimensions['H'].width = 18

    # Row 1: title
    ws.cell(row=1, column=2).value = 'Revolution Labs — Analysis of Real Estate Taxes'
    ws.cell(row=1, column=2).font  = Font(name='Calibri', size=13, bold=True)
    # Row 2: parcel info
    ws.cell(row=2, column=2).value = 'Lexington, MA  |  Parcel I.D. 00011619 and R0140050002'
    # Row 3: period
    ws.cell(row=3, column=2).value = f'Period: {period}'

    # Row 5: column headers
    HDR = 5
    for col, label in [(2, 'Description'), (4, 'Date'),
                       (6, 'A/C 135120\nPrepaid RE Taxes'),
                       (8, 'A/C 641110\nRE Tax Expense')]:
        c = ws.cell(row=HDR, column=col, value=label)
        c.font      = _hdr_font
        c.fill      = _hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = THIN
    ws.row_dimensions[HDR].height = 30

    # Data rows
    data_start = 7
    row = data_start
    alt_ctr = 0
    for desc, date_str, acct_135, acct_641 in _RET_ANALYSIS_SEED:
        is_section = (acct_135 is None and acct_641 is None and date_str is None)
        if is_section:
            # Section header row (parcel label)
            c = ws.cell(row=row, column=2, value=desc)
            c.font = Font(name='Calibri', size=11, bold=True, underline='single')
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            row += 1
            alt_ctr = 0
            continue
        alt = _alt_fill if alt_ctr % 2 == 1 else None
        alt_ctr += 1

        c_desc = ws.cell(row=row, column=2, value=desc)
        c_desc.font   = _std
        c_desc.border = THIN
        if alt: c_desc.fill = alt

        c_date = ws.cell(row=row, column=4, value=date_str or '')
        c_date.font   = _std
        c_date.border = THIN
        if alt: c_date.fill = alt

        for col, val in [(6, acct_135), (8, acct_641)]:
            c = ws.cell(row=row, column=col)
            c.value        = val if val is not None else None
            c.font         = _std
            c.border       = THIN
            c.alignment    = Alignment(horizontal='right')
            if val is not None:
                c.number_format = _num_fmt
            if alt: c.fill = alt
        row += 1

    data_end = row - 1

    # Ending balance row
    row += 1
    TOTAL_ROW = row
    ws.cell(row=row, column=2).value = f'Ending Balance per GL as of {period}'
    ws.cell(row=row, column=2).font  = _hdr_font
    ws.cell(row=row, column=2).fill  = _hdr_fill
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    for col, key in [(6, '135120'), (8, '641110')]:
        seed_val = _RET_ANALYSIS_ENDING.get(key, 0.0)
        c = ws.cell(row=row, column=col, value=seed_val)
        c.font          = _hdr_font
        c.fill          = _hdr_fill
        c.border        = THIN
        c.number_format = _num_fmt
        c.alignment     = Alignment(horizontal='right')
    row += 2

    # GL / Variance tie-out
    for account_code, col, label in [
        ('135120', 6, 'A/C 135120 Prepaid RE Taxes'),
        ('641110', 8, 'A/C 641110 RE Tax Expense'),
    ]:
        col_l   = get_column_letter(col)
        tb_acct = (tb_map or {}).get(account_code)
        tb_val  = tb_acct.ending_balance if tb_acct else None

        ws.cell(row=row, column=2).value = label
        ws.cell(row=row, column=2).font  = _bold
        ws.cell(row=row, column=col - 1).value = 'GL'
        ws.cell(row=row, column=col - 1).font  = _bold
        gl_cell = ws.cell(row=row, column=col, value=tb_val)
        if tb_val is not None:
            gl_cell.number_format = _num_fmt
        gl_row = row
        row += 1

        ws.cell(row=row, column=col - 1).value = 'Variance'
        ws.cell(row=row, column=col - 1).font  = _bold
        var_cell = ws.cell(row=row, column=col,
                           value=f'={col_l}{gl_row}-{col_l}{TOTAL_ROW}')
        var_cell.number_format = _num_fmt
        row += 2

    ws.freeze_panes = 'B6'


def _write_seed_loan_analysis(ws, period: str, tb_map: Optional[dict]):
    """
    Build a full-history Loan Analysis tab from _LOAN_ANALYSIS_SEED.
    Replaces _write_stub when no prior workpaper is provided.

    Column layout (matches JLL format):
      A (1): Loan #
      B (2): Description
      D (4): Date (MM/YY)
      F (6): 231100 Revlab  mortgage payable
      G (7): 231100 Revlabpm
      I (9): 213200 Accrued Interest Payable
      K (11): 801110 Interest Expense
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _hdr_fill = PatternFill(start_color='375623', end_color='375623', fill_type='solid')
    _hdr_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    _bold     = Font(name='Calibri', size=11, bold=True)
    _std      = Font(name='Calibri', size=11)
    _alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    _num_fmt  = '#,##0.00;(#,##0.00);"-"'
    THIN      = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 46
    ws.column_dimensions['C'].width = 3
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 3
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 3
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 3
    ws.column_dimensions['K'].width = 18

    # Title block
    ws.cell(row=1, column=1).value = 'Revolution Labs'
    ws.cell(row=1, column=1).font  = Font(name='Calibri', size=13, bold=True)
    ws.cell(row=2, column=1).value = 'Accrued Interest Payable (213200)'
    ws.cell(row=3, column=1).value = f'Period: {period}'

    # Column headers row 5
    ws.cell(row=5, column=6).value = 'Revlab';   ws.cell(row=5, column=7).value = 'Revlabpm'
    ws.cell(row=5, column=9).value = 'Accrued Interest Payable'
    ws.cell(row=5, column=11).value = 'Interest Expense'
    HDR = 6
    for col, label in [(1, 'Loan #'), (2, 'Description'), (4, 'Date'),
                       (6, 'A/C 231100\nMortgage Payable'),
                       (7, 'A/C 231100\nMortgage Payable'),
                       (9, 'A/C 213200\nAccrued Int Payable'),
                       (11, 'A/C 801110\nInterest Expense')]:
        c = ws.cell(row=HDR, column=col, value=label)
        c.font      = _hdr_font
        c.fill      = _hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = THIN
    ws.row_dimensions[HDR].height = 30

    # Data rows
    data_start = 8
    row = data_start
    for i, (loan_num, desc, date_str, rl, rpm, acct_213, acct_801) in enumerate(
            _LOAN_ANALYSIS_SEED):
        alt = _alt_fill if i % 2 == 1 else None

        for col, val in [(1, str(loan_num) if loan_num else ''),
                         (2, desc), (4, date_str or '')]:
            c = ws.cell(row=row, column=col, value=val)
            c.font   = _std
            c.border = THIN
            if alt: c.fill = alt

        for col, val in [(6, rl), (7, rpm), (9, acct_213), (11, acct_801)]:
            c = ws.cell(row=row, column=col)
            c.value  = val if val is not None else None
            c.font   = _std
            c.border = THIN
            c.alignment = Alignment(horizontal='right')
            if val is not None:
                c.number_format = _num_fmt
            if alt: c.fill = alt
        row += 1

    data_end = row - 1

    # Ending balance row
    row += 1
    TOTAL_ROW = row
    ws.cell(row=row, column=2).value = f'Ending Balance per GL as of {period}'
    ws.cell(row=row, column=2).font  = _hdr_font
    ws.cell(row=row, column=2).fill  = _hdr_fill
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    for col, val in [(6, _LOAN_ANALYSIS_ENDING['231100_rl']),
                     (7, _LOAN_ANALYSIS_ENDING['231100_rpm']),
                     (9, _LOAN_ANALYSIS_ENDING['213200']),
                     (11, _LOAN_ANALYSIS_ENDING['801110'])]:
        c = ws.cell(row=row, column=col, value=val)
        c.font          = _hdr_font
        c.fill          = _hdr_fill
        c.border        = THIN
        c.number_format = _num_fmt
        c.alignment     = Alignment(horizontal='right')
    row += 2

    # GL / Variance tie-out  (col I for 213200; col K for 801110)
    for account_code, col, label in [
        ('213200', 9,  '213200 Accrued Interest Payable'),
        ('801110', 11, '801110 Interest Expense'),
    ]:
        col_l   = get_column_letter(col)
        tb_acct = (tb_map or {}).get(account_code)
        tb_val  = tb_acct.ending_balance if tb_acct else None

        ws.cell(row=row, column=2).value = label
        ws.cell(row=row, column=2).font  = _bold
        ws.cell(row=row, column=col - 1).value = 'GL'
        ws.cell(row=row, column=col - 1).font  = _bold
        gl_cell = ws.cell(row=row, column=col, value=tb_val)
        if tb_val is not None:
            gl_cell.number_format = _num_fmt
        gl_row = row
        row += 1

        ws.cell(row=row, column=col - 1).value = 'Variance'
        ws.cell(row=row, column=col - 1).font  = _bold
        var_cell = ws.cell(row=row, column=col,
                           value=f'={col_l}{gl_row}-{col_l}{TOTAL_ROW}')
        var_cell.number_format = _num_fmt
        row += 2

    ws.freeze_panes = 'B7'


# ─────────────────────────────────────────────────────────────────────────────
# GL helper
# ─────────────────────────────────────────────────────────────────────────────

def _get_txns(gl_result, account_code: str) -> list:
    if not gl_result:
        return []
    for acct in (gl_result.accounts or []):
        if acct.account_code == account_code:
            return acct.transactions or []
    return []


# ─────────────────────────────────────────────────────────────────────────────
# Stub writer (no prior workpaper)
# ─────────────────────────────────────────────────────────────────────────────

def _write_stub(ws, display_name: str, period: str, new_rows: list,
                amount_col: int, tb_map: Optional[dict], account_code: str):
    """
    Minimal fallback when no prior workpaper was uploaded.
    Writes headers, current-period data rows, a SUM total, and a tie-out.
    """
    col_l = get_column_letter(amount_col)

    ws.cell(row=1, column=2).value = display_name
    ws.cell(row=2, column=2).value = f'Period: {period}'
    ws.cell(row=3, column=2).value = (
        'NOTE: Upload prior workpaper to carry forward historical data.'
    )
    ws.cell(row=5, column=2).value = 'Description'
    ws.cell(row=5, column=4).value = 'Date'
    ws.cell(row=5, column=amount_col).value = 'Amount'

    data_start = 7
    for i, row_data in enumerate(new_rows):
        r = data_start + i
        for col_letter, val in row_data.items():
            if not col_letter:
                continue
            try:
                c = column_index_from_string(col_letter.strip().upper())
            except Exception:
                continue
            ws.cell(row=r, column=c).value = val
            if isinstance(val, (int, float)):
                ws.cell(row=r, column=c).number_format = '#,##0.00;(#,##0.00);"-"'

    total_row = data_start + len(new_rows) + 1
    ws.cell(row=total_row, column=2).value = 'Ending Balance'
    ws.cell(row=total_row, column=amount_col).value = (
        f'=SUM({col_l}{data_start}:{col_l}{total_row - 1})'
    )
    ws.cell(row=total_row, column=amount_col).number_format = '#,##0.00;(#,##0.00);"-"'

    # Account code key
    ws.cell(row=total_row + 2, column=amount_col).value = (
        int(account_code) if account_code.isdigit() else account_code
    )

    tb_acct   = (tb_map or {}).get(account_code)
    tb_ending = tb_acct.ending_balance if tb_acct else None

    ws.cell(row=total_row + 4, column=amount_col - 1).value = 'GL'
    ws.cell(row=total_row + 4, column=amount_col).value = tb_ending
    if tb_ending is not None:
        ws.cell(row=total_row + 4, column=amount_col).number_format = '#,##0.00;(#,##0.00);"-"'

    ws.cell(row=total_row + 5, column=amount_col - 1).value = 'Variance'
    ws.cell(row=total_row + 5, column=amount_col).value = (
        f'={col_l}{total_row + 4}-{col_l}{total_row}'
    )


# ─────────────────────────────────────────────────────────────────────────────
# ── 115200  Escrow RET
# ─────────────────────────────────────────────────────────────────────────────

def build_ret_escrow_tab(
    wb, berkadia_loans, gl_result, period,
    current_prefix, tab_prefix, tb_map=None,
):
    """
    115200 Escrow RET: copy prior tab + append current-period escrow deposit row.

    Data sources (in priority order):
      1. GL account 115200 debit transactions (actual Yardi entries)
      2. Berkadia payment_re_taxes (sum across all loans)

    Column layout (RCW):  B = Description | D = Date | F = Per Stmt amount
    Amount column: F (col 6)
    """
    prior_ws, _ = _find_prior_tab(
        wb,
        ['115200 Escrow RET', '115200', 'Escrow RET', 'Escrow Real Estate Tax'],
        current_prefix,
    )
    tab_name = (tab_prefix + '115200 Escrow RET')[:31]
    if tab_name in wb.sheetnames:
        return

    ws = wb.create_sheet(tab_name)
    if prior_ws:
        _copy_tab_values(prior_ws, ws)

    # Build new rows from GL first
    txns = _get_txns(gl_result, '115200')
    new_rows = []
    for txn in txns:
        date_str = txn.date.strftime('%m/%d/%Y') if getattr(txn, 'date', None) else ''
        net = _safe_float(getattr(txn, 'debit', 0)) - _safe_float(getattr(txn, 'credit', 0))
        new_rows.append({
            'B': (txn.description or '')[:60],
            'D': date_str,
            'F': net,
        })

    # Fallback: Berkadia escrow deposit sum
    if not new_rows and berkadia_loans:
        total = sum(_safe_float(l.get('payment_re_taxes', 0)) for l in berkadia_loans)
        if total:
            new_rows.append({
                'B': f'RET ESCROW Payment {_fmt_long(period)}',
                'D': '',
                'F': total,
            })

    if prior_ws and new_rows:
        ip = _find_insertion_point(ws, 6)
        _insert_rows_and_write(ws, ip, new_rows, 6, period, tb_map, '115200')
    elif not prior_ws:
        _write_stub(ws, '115200 Escrow RET', period, new_rows, 6, tb_map, '115200')


# ─────────────────────────────────────────────────────────────────────────────
# ── 115300  Escrow Insurance
# ─────────────────────────────────────────────────────────────────────────────

def build_insurance_escrow_tab(
    wb, berkadia_loans, gl_result, period,
    current_prefix, tab_prefix, tb_map=None,
):
    """
    115300 Escrow Insurance: copy prior tab + append current-period entry.

    Data sources:
      1. GL account 115300 transactions
      2. Berkadia insurance_escrow_balance change (ending - prior ending)
         If unavailable, use payment_reserves as proxy.

    Amount column: F (col 6)
    """
    prior_ws, _ = _find_prior_tab(
        wb,
        ['115300 Escrow Insurance', '115300', 'Escrow Insurance'],
        current_prefix,
    )
    tab_name = (tab_prefix + '115300 Escrow Insurance')[:31]
    if tab_name in wb.sheetnames:
        return

    ws = wb.create_sheet(tab_name)
    if prior_ws:
        _copy_tab_values(prior_ws, ws)

    txns = _get_txns(gl_result, '115300')
    new_rows = []
    for txn in txns:
        date_str = txn.date.strftime('%m/%d/%Y') if getattr(txn, 'date', None) else ''
        net = _safe_float(getattr(txn, 'debit', 0)) - _safe_float(getattr(txn, 'credit', 0))
        new_rows.append({
            'B': (txn.description or f'Property Insurance per {_fmt_long(period)} stmt due')[:60],
            'D': date_str,
            'F': net,
        })

    if not new_rows and berkadia_loans:
        total = sum(
            _safe_float(l.get('payment_insurance', 0)
                        or l.get('payment_reserves', 0))
            for l in berkadia_loans
        )
        if total:
            new_rows.append({
                'B': f'Property Insurance per {_fmt_long(period)} stmt due',
                'D': '',
                'F': total,
            })

    if prior_ws and new_rows:
        ip = _find_insertion_point(ws, 6)
        _insert_rows_and_write(ws, ip, new_rows, 6, period, tb_map, '115300')
    elif not prior_ws:
        _write_stub(ws, '115300 Escrow Insurance', period, new_rows, 6, tb_map, '115300')


# ─────────────────────────────────────────────────────────────────────────────
# ── 115600  Restricted Cash – Other
# ─────────────────────────────────────────────────────────────────────────────

def build_restricted_cash_tab(
    wb, gl_result, period,
    current_prefix, tab_prefix, tb_map=None,
):
    """
    115600 Restricted Cash: copy prior tab + append current-period GL entries.
    Typical activity: monthly interest income deposits.

    Amount column: F (col 6)
    """
    prior_ws, _ = _find_prior_tab(
        wb,
        ['115600', 'Restricted Cash - Other', 'Restricted Cash'],
        current_prefix,
    )
    tab_name = (tab_prefix + '115600 Restricted Cash')[:31]
    if tab_name in wb.sheetnames:
        return

    ws = wb.create_sheet(tab_name)
    if prior_ws:
        _copy_tab_values(prior_ws, ws)

    txns = _get_txns(gl_result, '115600')
    new_rows = []
    for txn in txns:
        date_str = txn.date.strftime('%m/%d/%Y') if getattr(txn, 'date', None) else ''
        net = _safe_float(getattr(txn, 'debit', 0)) - _safe_float(getattr(txn, 'credit', 0))
        new_rows.append({
            'B': f'Rcd: {_fmt_long(period)} {txn.description or "Interest Income"}'[:60],
            'D': date_str,
            'F': net,
        })

    if prior_ws and new_rows:
        ip = _find_insertion_point(ws, 6)
        _insert_rows_and_write(ws, ip, new_rows, 6, period, tb_map, '115600')
    elif not prior_ws:
        _write_stub(ws, '115600 Restricted Cash - Other', period, new_rows, 6, tb_map, '115600')


# ─────────────────────────────────────────────────────────────────────────────
# ── RE Tax Analysis
# ─────────────────────────────────────────────────────────────────────────────

def build_ret_analysis_tab(
    wb, berkadia_loans, gl_result, period,
    current_prefix, tab_prefix, tb_map=None,
):
    """
    RE Tax Analysis: copy prior tab + append the monthly RE tax movements.

    RCW column layout (matches JLL format):
      B = Description | D = Date
      F = A/C 135120  Prepaid RE Taxes
            DR (positive): payment-month deferral (DR 135120 / CR 641110)
            CR (negative): release-month amortization (DR 641110 / CR 135120)
      H = A/C 641110  RE Tax Expense
            Release months: counterpart of the 135120 CR
            Payment months: Berkadia's quarterly bill (DR 641110 / CR 115200) also shown here

    Data sources (preference order):
      1. GL 135120 transactions — drives the F column; release entries also populate H
      2. GL 641110 transactions — any non-reclass entries fill the H column only
      3. Berkadia payment_re_taxes — fallback when GL has no 135120 activity

    Amount column for tie-out: F (col 6) → 135120 ending balance
    """
    prior_ws, _ = _find_prior_tab(
        wb,
        ['RE Tax Analysis', 'Real Estate Tax Analysis', 'RE Tax'],
        current_prefix,
    )
    tab_name = (tab_prefix + 'RE Tax Analysis')[:31]
    if tab_name in wb.sheetnames:
        return

    ws = wb.create_sheet(tab_name)
    if prior_ws:
        _copy_tab_values(prior_ws, ws)

    # 135120 = Prepaid RE Taxes (the asset account that moves each period)
    # 641110 = RE Tax Expense (debit on release months; net of Berkadia + pipeline JEs)
    txns_135120 = _get_txns(gl_result, '135120')
    txns_641110 = _get_txns(gl_result, '641110')

    new_rows = []

    # Each 135120 transaction drives the F column.
    # DR 135120 (payment month deferral) → positive in F
    # CR 135120 (release month) → negative in F; counterpart expense in H
    for txn in txns_135120:
        debit  = _safe_float(getattr(txn, 'debit',  0))
        credit = _safe_float(getattr(txn, 'credit', 0))
        net    = debit - credit          # positive = added to prepaid, negative = released
        if net == 0:
            continue
        date_str = txn.date.strftime('%m/%d/%Y') if getattr(txn, 'date', None) else ''
        desc     = txn.description or (
            f'Reclass {"Tax Exp to PPD" if net > 0 else "Tax Exp fr PPD"} {_fmt_mmy(period)}'
        )
        row = {'B': desc[:60], 'D': date_str, 'F': net}
        if net < 0:
            # Release: the debit hits 641110 — show in H
            row['H'] = abs(net)
        new_rows.append(row)

    # 641110-only transactions (e.g., the Berkadia quarterly posting DR 641110/CR 115200)
    # that don't have a 135120 counterpart — show in H only so the expense column is complete
    _135120_descs = {(r.get('B') or '').lower() for r in new_rows}
    for txn in txns_641110:
        debit  = _safe_float(getattr(txn, 'debit',  0))
        credit = _safe_float(getattr(txn, 'credit', 0))
        net    = debit - credit
        if net == 0:
            continue
        desc = (txn.description or '').lower()
        # Skip if already captured via the 135120 path (reclass entries appear in both)
        if any(word in desc for word in ('reclass', 'prepaid', 'ppd', '135120')):
            continue
        date_str = txn.date.strftime('%m/%d/%Y') if getattr(txn, 'date', None) else ''
        new_rows.append({
            'B': (txn.description or f'RE Tax {_fmt_mmy(period)}')[:60],
            'D': date_str,
            'H': net,
        })

    # Berkadia fallback when GL has no 135120 or 641110 activity yet
    if not new_rows and berkadia_loans:
        total = sum(_safe_float(l.get('payment_re_taxes', 0)) for l in berkadia_loans)
        if total:
            new_rows.append({
                'B': f'RET ESCROW Payment {_fmt_mmy(period)} - {_quarter_label(period)}',
                'D': '',
                'F': total,
            })

    if prior_ws and new_rows:
        ip = _find_insertion_point(ws, 6)
        _insert_rows_and_write(ws, ip, new_rows, 6, period, tb_map, '135120')
    elif not prior_ws:
        # First-period: write full historical seed data instead of a minimal stub
        _write_seed_ret_analysis(ws, period, tb_map)


# ─────────────────────────────────────────────────────────────────────────────
# ── Insurance Analysis stub writer  (no prior workpaper — first period)
# ─────────────────────────────────────────────────────────────────────────────

def _write_stub_insurance(ws, prepaid_active: list, period: str, tb_map: Optional[dict]):
    """
    Build a fresh Insurance Analysis tab for the first close period (no prior workpaper).

    Column layout (matches JLL format):
      B  Description / Policy Type
      C  Term  (service_start → service_end)
      D  Total Premium
      E  Per Month
      F  Starting Prepaid Balance  (label = "As of {period start}")
      G  Current-period expense    (label = period date, e.g. "01/2026")
      H  Ending Prepaid Balance    (= F − G, formula)
      I  639110 Total YTD Expense  (formula)
      J  639120 Total YTD Expense  (formula)
      K  135110 Ending Prepaid     (formula = H column sum)

    A GL / Variance tie-out block is appended below the data.
    """
    ins_expense_accounts = {'639110', '639120'}
    ins_items = [
        i for i in prepaid_active
        if str(i.get('gl_account_number', '')).strip() in ins_expense_accounts
    ]

    period_dt   = _period_to_dt(period)
    period_lbl  = period_dt.strftime('%m/%Y') if period_dt else period or ''
    start_lbl   = f'As of {period_dt.strftime("%m/01/%Y")}' if period_dt else 'Starting Balance'

    # ── Header rows ───────────────────────────────────────────────────────────
    ws.cell(row=1, column=2).value = 'Revolution Labs — Insurance Analysis'
    ws.cell(row=2, column=2).value = f'Period: {period}'
    ws.cell(row=3, column=2).value = 'A/C 639110 / 639120  Insurance Expense  |  A/C 135110  Prepaid Insurance'

    # Column header row
    HDR = 5
    headers = [
        ('B', 'Policy Type / Description'),
        ('C', 'Term'),
        ('D', 'Total Premium'),
        ('E', 'Per Month'),
        ('F', start_lbl),
        ('G', period_lbl),
        ('H', 'Ending Prepaid'),
        ('I', 'A/C 639110\nExpense'),
        ('J', 'A/C 639120\nExpense'),
        ('K', 'A/C 135110\nPrepaid'),
    ]
    for col_ltr, label in headers:
        from openpyxl.utils import column_index_from_string
        c = ws.cell(row=HDR, column=column_index_from_string(col_ltr), value=label)
        c.font      = _stub_hdr_font()
        c.fill      = _stub_hdr_fill()
        c.alignment = _stub_center_wrap()
    ws.row_dimensions[HDR].height = 30

    # Column widths
    col_widths = {'B': 38, 'C': 24, 'D': 14, 'E': 12,
                  'F': 14, 'G': 12, 'H': 14, 'I': 12, 'J': 12, 'K': 14}
    for col_ltr, w in col_widths.items():
        ws.column_dimensions[col_ltr].width = w
    ws.column_dimensions['A'].width = 2

    # ── Data rows ─────────────────────────────────────────────────────────────
    DATA_START = HDR + 1

    def _fmt_date(v):
        if v and hasattr(v, 'strftime'):
            return v.strftime('%m/%d/%Y')
        return str(v) if v else ''

    row = DATA_START
    for i, item in enumerate(ins_items):
        gl_acct    = str(item.get('gl_account_number', '')).strip()
        desc       = item.get('description') or item.get('vendor') or ''
        svc_start  = item.get('service_start')
        svc_end    = item.get('service_end')
        term       = f'{_fmt_date(svc_start)} – {_fmt_date(svc_end)}' if svc_start else ''
        premium    = _safe_float(item.get('total_amount'))
        per_month  = _safe_float(item.get('monthly_amount'))
        rem_months = int(item.get('remaining_months') or 0)

        # Starting prepaid = balance before this period's release
        #   post-advance remaining + 1 month = (rem_months + 1) × per_month
        starting_prepaid = per_month * (rem_months + 1)
        # Ending prepaid = remaining after this period
        ending_prepaid   = per_month * rem_months

        alt = _stub_alt_fill() if i % 2 == 1 else None

        row_vals = {
            'B': desc,
            'C': term,
            'D': premium,
            'E': per_month,
            'F': starting_prepaid,
            'G': per_month,        # current period expense
            'H': ending_prepaid,   # = F - G (could also write as formula, value is cleaner)
        }
        # Route expense to the correct total column
        if gl_acct == '639110':
            row_vals['I'] = per_month
        elif gl_acct == '639120':
            row_vals['J'] = per_month

        for col_ltr, val in row_vals.items():
            ci = column_index_from_string(col_ltr)
            c  = ws.cell(row=row, column=ci, value=val)
            if isinstance(val, float) and col_ltr in ('D', 'E', 'F', 'G', 'H', 'I', 'J'):
                c.number_format = '#,##0.00;(#,##0.00);"-"'
            if alt:
                c.fill = alt
        row += 1

    DATA_END = row - 1

    # ── Totals row ────────────────────────────────────────────────────────────
    row += 1
    TOTAL_ROW = row
    ws.cell(row=row, column=2).value = 'Total'
    ws.cell(row=row, column=2).font  = _stub_bold_font()
    for col_ltr in ('D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'):
        ci   = column_index_from_string(col_ltr)
        cell = ws.cell(row=row, column=ci)
        cell.value        = f'=SUM({col_ltr}{DATA_START}:{col_ltr}{DATA_END})'
        cell.number_format = '#,##0.00;(#,##0.00);"-"'
        cell.font          = _stub_bold_font()
        cell.fill          = _stub_blue_fill()
    row += 2

    # ── GL / TB tie-out ───────────────────────────────────────────────────────
    for account_code, col_ltr, label in [
        ('639110', 'I', '639110 Insurance Expense'),
        ('639120', 'J', '639120 GL Expense'),
        ('135110', 'K', '135110 Prepaid Insurance'),
    ]:
        ci       = column_index_from_string(col_ltr)
        tb_acct  = (tb_map or {}).get(account_code)
        tb_val   = tb_acct.ending_balance if tb_acct else None

        ws.cell(row=row, column=2).value = label
        ws.cell(row=row, column=2).font  = _stub_bold_font()

        ws.cell(row=row, column=ci - 1).value = 'GL'
        ws.cell(row=row, column=ci - 1).font  = _stub_bold_font()
        gl_cell = ws.cell(row=row, column=ci, value=tb_val)
        if tb_val is not None:
            gl_cell.number_format = '#,##0.00;(#,##0.00);"-"'
        gl_row = row
        row += 1

        ws.cell(row=row, column=ci - 1).value = 'Variance'
        ws.cell(row=row, column=ci - 1).font  = _stub_bold_font()
        var_cell = ws.cell(
            row=row, column=ci,
            value=f'={col_ltr}{gl_row}-{col_ltr}{TOTAL_ROW}',
        )
        var_cell.number_format = '#,##0.00;(#,##0.00);"-"'
        row += 2

    ws.freeze_panes = 'B6'


def _stub_hdr_font():
    from openpyxl.styles import Font
    return Font(name='Calibri', size=10, bold=True, color='FFFFFF')

def _stub_hdr_fill():
    from openpyxl.styles import PatternFill
    return PatternFill(start_color='002060', end_color='002060', fill_type='solid')

def _stub_blue_fill():
    from openpyxl.styles import PatternFill
    return PatternFill(start_color='D6EAE1', end_color='D6EAE1', fill_type='solid')

def _stub_alt_fill():
    from openpyxl.styles import PatternFill
    return PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

def _stub_bold_font():
    from openpyxl.styles import Font
    return Font(name='Calibri', size=10, bold=True)

def _stub_center_wrap():
    from openpyxl.styles import Alignment
    return Alignment(horizontal='center', vertical='center', wrap_text=True)


# ─────────────────────────────────────────────────────────────────────────────
# ── Insurance Analysis  (column-based — one column per month)
# ─────────────────────────────────────────────────────────────────────────────

def build_insurance_analysis_tab(
    wb, prepaid_active, gl_result, period,
    current_prefix, tab_prefix, tb_map=None,
):
    """
    Insurance Analysis: copy prior tab, then find which column corresponds to
    the current period and fill in the monthly amortization amounts.

    Structure (RCW):
      Row 7: date headers starting from a base date, each subsequent column
             is the prior column + 31 days (=I7+31, =J7+31, …).
      Rows 8+: one row per policy, amounts in the matching date column.
      Cols U–X: TOTAL, TOTAL, PREPAID INSURANCE, ACCRUED INSURANCE  (summary cols)

    Strategy:
      1. Copy prior tab.
      2. Scan row 7 for a datetime whose month/year matches the current period.
         If found, that is the target column.
         If not found, find the last datetime column and use the next column,
         writing the date formula continuation (=prev_col+31).
      3. For each data row (8+), match to a prepaid_active item by description
         and write the monthly_amount into the target column.
      4. Update the TOTAL/SUM columns (U onward) if they don't already include
         the new column in their ranges.

    No tie-out rebuild needed — the Insurance Analysis does its own GL lookups
    in the summary columns (U–X), which are rebuilt by the VLOOKUP formulas;
    we update those cells with direct TB values instead.
    """
    prior_ws, _ = _find_prior_tab(
        wb, ['Insurance Analysis'], current_prefix,
    )
    tab_name = (tab_prefix + 'Insurance Analysis')[:31]
    if tab_name in wb.sheetnames:
        return

    ws = wb.create_sheet(tab_name)

    if not prior_ws:
        # First-period fresh build (January 2026 starting point).
        # Build the full JLL-style grid directly from the prepaid ledger.
        _write_stub_insurance(ws, prepaid_active or [], period, tb_map)
        return

    _copy_tab_values(prior_ws, ws)

    if not prepaid_active:
        return

    period_dt = _period_to_dt(period)
    if not period_dt:
        return

    # ── Find the target column ────────────────────────────────────────
    target_col = None
    last_date_col = None

    # Row 7 has the date headers — search up to col 40
    for col in range(1, 41):
        val = ws.cell(row=7, column=col).value
        if isinstance(val, datetime):
            if val.year == period_dt.year and val.month == period_dt.month:
                target_col = col
                break
            last_date_col = col

    if target_col is None and last_date_col is not None:
        # Append the current period as the next column
        target_col = last_date_col + 1
        prev_letter = get_column_letter(last_date_col)
        ws.cell(row=7, column=target_col).value = f'={prev_letter}7+31'

    if target_col is None:
        return

    target_letter = get_column_letter(target_col)

    # ── Insurance prepaid items ───────────────────────────────────────
    # gl_account_number in the ledger is the EXPENSE account (debit side).
    # 639110 = Property Insurance Expense, 639120 = General Liability Expense.
    # 135110 / 213300 are the offsetting BS accounts — NOT stored as gl_account_number.
    ins_expense_accounts = {'639110', '639120'}
    ins_items = [
        i for i in prepaid_active
        if str(i.get('gl_account_number', '')).strip() in ins_expense_accounts
    ]

    # Build a lookup: description keywords → monthly_amount
    item_lookup = {}
    for item in ins_items:
        desc = (
            item.get('description') or item.get('vendor') or ''
        ).lower().strip()
        monthly = float(item.get('monthly_amount') or 0)
        if desc and monthly:
            item_lookup[desc] = monthly

    # ── Fill monthly amounts into matching data rows ──────────────────
    skip_keywords = ('date', 'paid', 'premium', 'per month', 'prepaid', 'term',
                     'accrual', 'total', 'expense', 'insurance', 'account')

    for row_num in range(8, min(ws.max_row, 30)):
        # Find the description for this row (cols B–E area)
        row_desc = ''
        for c in range(2, 8):
            v = ws.cell(row=row_num, column=c).value
            if isinstance(v, str) and v.strip():
                row_desc = v.strip().lower()
                if not any(k in row_desc for k in skip_keywords):
                    break
                row_desc = ''

        if not row_desc:
            continue
        if any(k in row_desc for k in skip_keywords):
            continue

        # Match to a prepaid item
        monthly = 0.0
        for item_desc, amt in item_lookup.items():
            # Check if at least one meaningful word from the item description
            # appears in the row description (ignore short words)
            words = [w for w in item_desc.split() if len(w) > 3]
            if words and any(w in row_desc for w in words[:3]):
                monthly = amt
                break

        if monthly:
            ws.cell(row=row_num, column=target_col).value = round(monthly, 2)
            ws.cell(row=row_num, column=target_col).number_format = '#,##0.00'

    # ── Update summary TOTAL column SUM ranges if needed ─────────────
    # Summary cols U (21), V (22), W (23), X (24) hold =SUM(I:T row) style
    for sum_col in range(21, 26):
        for row_num in range(8, 25):
            cell = ws.cell(row=row_num, column=sum_col)
            val  = cell.value
            if not isinstance(val, str) or not val.upper().startswith('=SUM('):
                continue
            # Check if target_letter is already in the range
            m = re.search(
                r'=SUM\(\s*([A-Z]+)(\d+)\s*:\s*([A-Z]+)(\d+)\s*\)',
                val, re.I,
            )
            if not m:
                continue
            start_letter = m.group(1).upper()
            end_letter   = m.group(3).upper()
            row_ref      = m.group(2)
            start_idx    = column_index_from_string(start_letter)
            end_idx      = column_index_from_string(end_letter)
            if target_col < start_idx or target_col > end_idx + 1:
                continue  # not adjacent, skip
            if target_col > end_idx:
                # Extend the range
                new_end = get_column_letter(target_col)
                cell.value = f'=SUM({start_letter}{row_ref}:{new_end}{row_ref})'

    # ── Rebuild GL tie-out cells (cols U–X, rows ~22–25) ─────────────
    for account_code, sum_col in [('639110', 21), ('639120', 22),
                                   ('135110', 23), ('213300', 24)]:
        tb_acct = (tb_map or {}).get(account_code)
        if tb_acct is None:
            continue
        # Find GL row for this column (scan for 'GL' label nearby)
        for gl_row in range(20, min(ws.max_row, 32)):
            lbl_cell = ws.cell(row=gl_row, column=sum_col - 1).value
            if isinstance(lbl_cell, str) and lbl_cell.strip().lower() in ('gl', 'g/l'):
                ws.cell(row=gl_row, column=sum_col).value = tb_acct.ending_balance
                ws.cell(row=gl_row, column=sum_col).number_format = '#,##0.00;(#,##0.00);"-"'
                # Variance: next row
                vs_cell = ws.cell(row=gl_row + 1, column=sum_col)
                col_l   = get_column_letter(sum_col)
                # Find the total row for this column
                for tot_r in range(gl_row - 1, gl_row - 5, -1):
                    tv = ws.cell(row=tot_r, column=sum_col).value
                    if isinstance(tv, (int, float)) or (isinstance(tv, str) and tv.startswith('=SUM')):
                        vs_cell.value = f'={col_l}{gl_row}-{col_l}{tot_r}'
                        vs_cell.number_format = '#,##0.00;(#,##0.00);"-"'
                        break
                break


# ─────────────────────────────────────────────────────────────────────────────
# ── Loan Analysis
# ─────────────────────────────────────────────────────────────────────────────

def build_loan_analysis_tab(
    wb, berkadia_loans, gl_result, period,
    current_prefix, tab_prefix, tb_map=None,
):
    """
    Loan Analysis: copy prior tab + append current-period interest cycle rows.

    Per loan, per period, the standard three-row pattern is:
      1. {prior_month} Accrual Reversal   → 213200 debit (+), 801110 credit (-)
      2. {loan_id} {prior_month} Interest Payment → 801110 debit (+)
      3. {loan_id} Accr {curr_month} Interest Due → 213200 credit (-), 801110 debit (+)

    Data sources (preference order):
      1. GL account 213200 transactions  — matched to loan by description suffix
         (e.g., description contains "1159010" or "159010")
      2. GL account 801110 transactions  — interest expense counterpart
      3. Berkadia payment_interest       — confirms the payment amount per loan

    RCW column layout:
      A  = Loan number
      B  = Description
      D  = Period string (MM/YY)
      F  = 231100 Revlab     (mortgage payable — blank for interest entries)
      G  = 231100 Revlabpm   (mortgage payable — blank for interest entries)
      I  = 213200 Accrued Interest Payable
      K  = 801110 Interest Expense

    Amount column for SUM / tie-out: I (col 9).
    """
    prior_ws, _ = _find_prior_tab(wb, ['Loan Analysis'], current_prefix)
    tab_name = (tab_prefix + 'Loan Analysis')[:31]
    if tab_name in wb.sheetnames:
        return

    ws = wb.create_sheet(tab_name)
    if prior_ws:
        _copy_tab_values(prior_ws, ws)

    txns_213200 = _get_txns(gl_result, '213200')
    txns_801110 = _get_txns(gl_result, '801110')
    period_str  = _fmt_mmy(period)
    prior_str   = _prior_long(period)

    # ── Identify loan IDs ─────────────────────────────────────────────
    loan_ids = _extract_loan_ids(txns_213200, berkadia_loans or [])

    new_rows = []

    if loan_ids:
        for lid in loan_ids:
            _add_loan_rows(
                new_rows, lid, txns_213200, txns_801110,
                berkadia_loans or [], period_str, prior_str,
            )
    else:
        # Fallback: dump all 213200/801110 transactions without loan grouping
        for txn in txns_213200:
            net = _safe_float(getattr(txn, 'debit', 0)) - _safe_float(getattr(txn, 'credit', 0))
            new_rows.append({
                'B': (txn.description or '')[:50],
                'D': period_str,
                'I': net,
            })
        for txn in txns_801110:
            net = _safe_float(getattr(txn, 'debit', 0)) - _safe_float(getattr(txn, 'credit', 0))
            new_rows.append({
                'B': (txn.description or '')[:50],
                'D': period_str,
                'K': net,
            })

    if prior_ws and new_rows:
        ip = _find_insertion_point(ws, 9)   # col I = 9
        _insert_rows_and_write(ws, ip, new_rows, 9, period, tb_map, '213200')
    elif not prior_ws:
        # First-period: write full historical seed data instead of a minimal stub
        _write_seed_loan_analysis(ws, period, tb_map)


def _extract_loan_ids(txns_213200: list, berkadia_loans: list) -> List[str]:
    """
    Derive the set of loan IDs to process, in order.

    Priority:
      1. Berkadia loan_number fields  (e.g., "1159010")
      2. IDs parsed from GL 213200 transaction descriptions
    """
    seen: set = set()
    ids: List[str] = []

    # From Berkadia
    for loan in berkadia_loans:
        lid = str(loan.get('loan_number') or '').strip()
        if lid and lid not in seen:
            seen.add(lid)
            ids.append(lid)

    # From GL descriptions (pattern: 6-7 digit number starting with 1159)
    if not ids:
        for txn in txns_213200:
            desc = txn.description or ''
            for m in re.finditer(r'\b1?1590\d\d\b', desc):
                raw = m.group()
                # Normalise to 7 digits
                lid = raw if len(raw) == 7 else ('1' + raw if len(raw) == 6 else raw)
                if lid not in seen:
                    seen.add(lid)
                    ids.append(lid)

    return ids


def _add_loan_rows(
    new_rows: list,
    loan_id: str,
    txns_213200: list,
    txns_801110: list,
    berkadia_loans: list,
    period_str: str,
    prior_str: str,
):
    """
    Append the three standard rows for one loan to new_rows.
    """
    suffix = loan_id[-6:]   # e.g., "159010" from "1159010"

    def _matches(txn):
        d = (txn.description or '').lower()
        return suffix in d or loan_id.lower() in d

    loan_213 = [t for t in txns_213200 if _matches(t)]
    loan_811 = [t for t in txns_801110 if _matches(t)]

    berk = next(
        (l for l in berkadia_loans
         if str(l.get('loan_number', '')).endswith(suffix)),
        None,
    )

    # ── Row 1: Prior-month accrual reversal ──────────────────────────
    reversal = next(
        (t for t in loan_213
         if 'reversal' in (t.description or '').lower()
         or 'reversal' in (t.description or '').lower()),
        None,
    )
    if reversal:
        rev_213 = _safe_float(getattr(reversal, 'debit', 0)) - _safe_float(getattr(reversal, 'credit', 0))
        rev_811 = -rev_213
    else:
        rev_213 = rev_811 = 0.0

    new_rows.append({
        'B': f'{prior_str} Accrual Reversal',
        'D': period_str,
        'I': rev_213 if rev_213 != 0 else None,
        'K': rev_811 if rev_811 != 0 else None,
    })

    # ── Row 2: Interest payment ───────────────────────────────────────
    payment_txn = next(
        (t for t in loan_213
         if any(k in (t.description or '').lower() for k in ('payment', 'pytm'))
         and 'reversal' not in (t.description or '').lower()),
        None,
    )
    if payment_txn:
        pay_amt = (
            _safe_float(getattr(payment_txn, 'debit', 0))
            - _safe_float(getattr(payment_txn, 'credit', 0))
        )
    else:
        pay_amt = _safe_float((berk or {}).get('payment_interest', 0))

    new_rows.append({
        'A': loan_id,
        'B': f'{prior_str} Interest Payment',
        'D': period_str,
        'K': pay_amt if pay_amt != 0 else None,
    })

    # ── Row 3: New accrual ────────────────────────────────────────────
    accrual_txn = next(
        (t for t in loan_213
         if any(k in (t.description or '').lower() for k in ('accr', 'accrual'))
         and 'reversal' not in (t.description or '').lower()),
        None,
    )
    if accrual_txn:
        acc_213 = (
            _safe_float(getattr(accrual_txn, 'debit', 0))
            - _safe_float(getattr(accrual_txn, 'credit', 0))
        )
        acc_811 = -acc_213
    else:
        acc_213 = acc_811 = 0.0

    new_rows.append({
        'A': loan_id,
        'B': f'Accr {period_str} Interest Due -{loan_id}',
        'D': period_str,
        'I': acc_213 if acc_213 != 0 else None,
        'K': acc_811 if acc_811 != 0 else None,
    })


# ─────────────────────────────────────────────────────────────────────────────
# ── Main entry point  (called from bs_workpaper_generator)
# ─────────────────────────────────────────────────────────────────────────────

def build_all_analysis_tabs(
    wb,
    period: str,
    current_prefix: str,
    tab_prefix: str,
    gl_result=None,
    tb_map: Optional[dict] = None,
    berkadia_loans: Optional[list] = None,
    prepaid_active: Optional[list] = None,
):
    """
    Build all analysis tabs for the current period.

    Called from bs_workpaper_generator.generate_bs_workpaper() after the
    standard BS account tabs have been written.

    Order matters: simpler / self-contained tabs first, complex last.
    Each builder checks whether its tab already exists and returns early
    if so (idempotent).
    """
    build_ret_escrow_tab(
        wb, berkadia_loans, gl_result, period,
        current_prefix, tab_prefix, tb_map,
    )
    build_insurance_escrow_tab(
        wb, berkadia_loans, gl_result, period,
        current_prefix, tab_prefix, tb_map,
    )
    build_restricted_cash_tab(
        wb, gl_result, period,
        current_prefix, tab_prefix, tb_map,
    )
    build_ret_analysis_tab(
        wb, berkadia_loans, gl_result, period,
        current_prefix, tab_prefix, tb_map,
    )
    build_insurance_analysis_tab(
        wb, prepaid_active, gl_result, period,
        current_prefix, tab_prefix, tb_map,
    )
    build_loan_analysis_tab(
        wb, berkadia_loans, gl_result, period,
        current_prefix, tab_prefix, tb_map,
    )

"""
Close Process Tracker Generator
================================
Produces GA_Close_Tracker.xlsx — a step-by-step record of the full
monthly close lifecycle from JLL handoff through Lauren's package release.

9 steps covering:
  0. JLL Completes Bank Rec & Payments
  1. Pass 1 Files Uploaded & JEs Generated
  2. JEs Uploaded to Yardi
  3. Final Close Run in Yardi
  4. Final Files Re-Exported from Yardi
  5. Pass 2 Files Uploaded
  6. Reports Generated (Pass 2)
  7. QC Review Complete (Ryan / Natasha)
  8. Final Package Released to Lauren

Called from app.py when Step 8 is marked complete.
"""

from __future__ import annotations

from datetime import datetime
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Brand colours ─────────────────────────────────────────────────────────────
_BLACK      = '000000'
_GRP_GREEN  = '2D6F50'
_WHITE      = 'FFFFFF'
_PASS_GREEN = 'E2EFDA'
_PASS_FONT  = '006100'
_PEND_GRAY  = 'F2F2F2'
_PEND_FONT  = '757575'
_AUTO_BLUE  = 'DDEEFF'   # light blue tint for auto-detected rows

_THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def _font(bold=False, italic=False, size=11, color=_BLACK, name='Calibri') -> Font:
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)


# ── Step definitions ──────────────────────────────────────────────────────────
# Each tuple: (index, description, type)
# type = 'auto'   → pipeline sets this automatically
# type = 'manual' → user clicks "Mark Complete"

CLOSE_TRACKER_STEPS: List[Tuple[int, str, str]] = [
    (0, 'JLL Completes Bank Rec & Payments',        'manual'),
    (1, 'Pass 1 Files Uploaded & JEs Generated',    'auto'),
    (2, 'JEs Uploaded to Yardi',                    'manual'),
    (3, 'Final Close Run in Yardi',                 'manual'),
    (4, 'Final Files Re-Exported from Yardi',       'manual'),
    (5, 'Pass 2 Files Uploaded',                    'auto'),
    (6, 'Reports Generated (Pass 2)',               'auto'),
    (7, 'QC Review Complete (Property Accountant / Accounting Manager)', 'auto'),
    (8, 'Final Package Released to CFO',             'manual'),
]


def generate_close_tracker_xlsx(
    output_path: str,
    close_tracker: Dict[int, Dict[str, str]],
    period: str,
    property_name: str,
) -> str:
    """
    Write GA_Close_Tracker.xlsx and return output_path.

    Args:
        output_path:    Destination .xlsx path.
        close_tracker:  {step_index: {'completed_by': str, 'timestamp': str,
                          'auto': bool}}
                        Steps not present are treated as 'Pending'.
        period:         Close period string e.g. 'Jan-2026'.
        property_name:  Property display name e.g. 'Revolution Labs'.

    Returns:
        output_path
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Close Tracker'

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {
        'A': 2,    # left margin
        'B': 6,    # Step #
        'C': 40,   # Description
        'D': 16,   # Status
        'E': 22,   # Completed By
        'F': 22,   # Timestamp
        'G': 14,   # Auto-Detected
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    title = ws.cell(row=1, column=2,
                    value=f'{property_name} — Monthly Close Process Tracker')
    title.font = _font(bold=True, size=14, color=_WHITE)
    title.fill = _fill(_BLACK)
    title.alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=7)
    ws.row_dimensions[1].height = 22

    # ── Row 2: Sub-header ─────────────────────────────────────────────────────
    sub = ws.cell(row=2, column=2,
                  value=f'Period: {period}  |  Generated: '
                        f'{datetime.now().strftime("%m/%d/%Y %H:%M")}')
    sub.font = _font(italic=True, size=10, color=_WHITE)
    sub.fill = _fill(_GRP_GREEN)
    sub.alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
    ws.row_dimensions[2].height = 18

    # ── Row 3: Blank spacer ───────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Row 4: Column headers ─────────────────────────────────────────────────
    col_headers = ['Step', 'Description', 'Status',
                   'Completed By', 'Timestamp', 'Auto-Detected']
    for ci, hdr in enumerate(col_headers):
        cell = ws.cell(row=4, column=ci + 2, value=hdr)
        cell.font = _font(bold=True, size=10, color=_WHITE)
        cell.fill = _fill(_GRP_GREEN)
        cell.border = _THIN
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
    ws.row_dimensions[4].height = 20

    # ── Rows 5–13: One row per step ───────────────────────────────────────────
    for step_idx, description, step_type in CLOSE_TRACKER_STEPS:
        row = step_idx + 5
        ct_entry = close_tracker.get(step_idx)

        is_complete   = bool(ct_entry)
        completed_by  = ct_entry.get('completed_by', '') if ct_entry else ''
        timestamp     = ct_entry.get('timestamp', '')    if ct_entry else ''
        is_auto       = ct_entry.get('auto', False)      if ct_entry else False

        status     = 'Complete' if is_complete else 'Pending'
        auto_label = 'Yes' if is_auto else ('—' if not is_complete else 'No')

        # Row fill: green if complete, light blue tint if auto-detected, gray if pending
        if is_complete and is_auto:
            row_fill_hex = _AUTO_BLUE
        elif is_complete:
            row_fill_hex = _PASS_GREEN
        else:
            row_fill_hex = _PEND_GRAY

        row_fill  = _fill(row_fill_hex)
        stat_font = _font(bold=True, size=10,
                          color=_PASS_FONT if is_complete else _PEND_FONT)
        text_font = _font(size=10)

        vals = [step_idx + 1, description, status, completed_by, timestamp, auto_label]
        for ci, val in enumerate(vals):
            cell = ws.cell(row=row, column=ci + 2, value=val)
            cell.fill   = row_fill
            cell.border = _THIN

            if ci == 0:   # Step #
                cell.font = _font(bold=True, size=10,
                                  color=_PASS_FONT if is_complete else _PEND_FONT)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 2:  # Status
                cell.font = stat_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 5:  # Auto-Detected
                cell.font = _font(size=10,
                                  color=_PASS_FONT if is_auto else _PEND_FONT)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.font = text_font
                cell.alignment = Alignment(vertical='center',
                                           wrap_text=(ci == 1))

        ws.row_dimensions[row].height = 18

    # ── Spacer row ────────────────────────────────────────────────────────────
    spacer_row = len(CLOSE_TRACKER_STEPS) + 5
    ws.row_dimensions[spacer_row].height = 6

    # ── Summary row ───────────────────────────────────────────────────────────
    complete_count = sum(1 for i in range(len(CLOSE_TRACKER_STEPS))
                         if i in close_tracker)
    total_count    = len(CLOSE_TRACKER_STEPS)
    summary_row    = spacer_row + 1

    all_done   = complete_count == total_count
    sc = ws.cell(row=summary_row, column=2,
                 value=f'{complete_count} of {total_count} steps complete')
    sc.font      = _font(bold=True, size=10,
                         color=_PASS_FONT if all_done else '9C0006')
    sc.fill      = _fill(_PASS_GREEN if all_done else 'FFCCCC')
    sc.border    = _THIN
    sc.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=summary_row, start_column=2,
                   end_row=summary_row, end_column=7)
    ws.row_dimensions[summary_row].height = 20

    wb.save(output_path)
    return output_path

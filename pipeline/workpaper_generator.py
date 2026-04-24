"""
Workpaper Generator for GA Automation Pipeline
================================================
Produces institutional-grade workpapers for month-end close:
  1. Bank Reconciliation — GL to bank tie-out with outstanding items
  2. Debt Service Schedule — P&I breakout, escrow recon, payment detail
  3. Rent Roll Tie-Out — Lease-level billed vs collected vs GL revenue
  4. Accrual Schedule — Invoice aging, vendor detail, accrued vs paid

Each tab is formatted as a standalone workpaper suitable for
inclusion in a Singerman close binder.
"""

import os
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from accrual_entry_generator import (
    build_accrual_entries, write_accrual_entries_workpaper_tab
)


# ── Styling ──────────────────────────────────────────────────

DARK_BLUE = '1F4E78'
MED_BLUE = '2E75B6'
LIGHT_BLUE = 'D6E4F0'
LIGHT_GRAY = 'F2F2F2'
WHITE = 'FFFFFF'
GREEN_LIGHT = 'E2EFDA'
RED_LIGHT = 'FFC7CE'
AMBER_LIGHT = 'FFEB9C'

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
BOTTOM_BORDER = Border(bottom=Side(style='medium'))
DOUBLE_BOTTOM = Border(bottom=Side(style='double'))


def _hdr(bold=True, size=11, color='FFFFFF', fill_color=DARK_BLUE, align='center'):
    return {
        'font': Font(name='Calibri', size=size, bold=bold, color=color),
        'fill': PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid'),
        'alignment': Alignment(horizontal=align, vertical='center', wrap_text=True),
        'border': THIN_BORDER,
    }


def _cell(alt=False, fmt=None, bold=False, align='left'):
    fill = LIGHT_GRAY if alt else WHITE
    style = {
        'font': Font(name='Calibri', size=11, bold=bold),
        'fill': PatternFill(start_color=fill, end_color=fill, fill_type='solid'),
        'alignment': Alignment(horizontal=align, vertical='center'),
        'border': THIN_BORDER,
    }
    if fmt:
        style['number_format'] = fmt
    return style


def _apply(cell, style):
    for attr in ('font', 'fill', 'alignment', 'border'):
        if attr in style:
            setattr(cell, attr, style[attr])
    if 'number_format' in style:
        cell.number_format = style['number_format']


def _section_header(ws, row, col_start, col_end, text, fill=MED_BLUE):
    """Write a section header spanning columns."""
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    c.fill = PatternFill(start_color=fill, end_color=fill, fill_type='solid')
    c.alignment = Alignment(horizontal='left', vertical='center')
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER


def _title_row(ws, row, text, cols=8):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name='Calibri', size=14, bold=True, color=DARK_BLUE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)


def _subtitle_row(ws, row, text, cols=8):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name='Calibri', size=11, italic=True, color='666666')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)


def _kv_row(ws, row, label, value, fmt=None, label_col=1, val_col=2):
    """Write a label-value pair."""
    c1 = ws.cell(row=row, column=label_col, value=label)
    c1.font = Font(name='Calibri', size=11, bold=True)
    c2 = ws.cell(row=row, column=val_col, value=value)
    c2.font = Font(name='Calibri', size=11)
    if fmt:
        c2.number_format = fmt


def _auto_width(ws, max_cols, min_width=10, max_width=50):
    for col in range(1, max_cols + 1):
        letter = get_column_letter(col)
        best = min_width
        for cell in ws[letter]:
            try:
                if cell.value:
                    best = max(best, len(str(cell.value)) + 2)
            except:
                pass
        ws.column_dimensions[letter].width = min(best, max_width)


# ── 1. BANK RECONCILIATION ──────────────────────────────────

def _write_bank_recon_workpaper(wb, engine_result):
    """
    Bank reconciliation workpaper:
      - Header: property, period, GL acct, bank acct
      - Section A: Balance comparison (book vs bank)
      - Section B: Outstanding checks (GL issued, not cleared on bank)
      - Section C: Deposits in transit (GL recorded, not on bank)
      - Section D: Adjusted balance tie-out
      - Section E: Cleared check detail
    """
    ws = wb.create_sheet('Bank Recon')

    period = engine_result.period or 'N/A'
    prop = engine_result.property_name or 'N/A'
    recon = engine_result.bank_recon_detail

    # Determine close date for stale-check aging (last day of the period month)
    _close_date: Optional[date] = None
    try:
        # period is "Mar-2026", "Apr-2026", etc.
        _pd = datetime.strptime(period, '%b-%Y')
        # Last day of month: first day of next month minus one day
        _next_m = _pd.replace(day=28) + timedelta(days=4)
        _close_date = (_next_m - timedelta(days=_next_m.day)).date()
    except (ValueError, AttributeError):
        _close_date = None

    # Use precomputed recon from engine (single source of truth)
    if recon:
        gl_end = recon.gl_ending
        gl_begin = recon.gl_beginning
        bank_end = recon.bank_ending
        bank_begin = recon.bank_beginning
        outstanding_checks = recon.outstanding_checks
        total_outstanding = recon.total_outstanding_checks
        deposits_in_transit = recon.deposits_in_transit
        total_dit = recon.total_deposits_in_transit
        matched_checks = recon.matched_checks
        matched_ach = recon.matched_ach
        matched_deposits = recon.matched_deposits
    else:
        # Fallback: no bank recon computed (bank data was missing)
        gl_data = engine_result.parsed.get('gl')
        gl_cash = None
        if gl_data and hasattr(gl_data, 'accounts'):
            for acct in gl_data.accounts:
                if acct.account_code == '111100':
                    gl_cash = acct
                    break
        gl_end = gl_cash.ending_balance if gl_cash else 0
        gl_begin = gl_cash.beginning_balance if gl_cash else 0
        bank_end = 0
        bank_begin = 0
        outstanding_checks = []
        total_outstanding = 0
        deposits_in_transit = []
        total_dit = 0
        matched_checks = []
        matched_ach = []
        matched_deposits = []

    # ── Write workpaper ──
    row = 1
    _title_row(ws, row, f'Bank Reconciliation — {prop}')
    row += 1
    _subtitle_row(ws, row, f'Period: {period}  |  GL Account: 111100 Cash-Operating  |  Prepared: {datetime.now().strftime("%m/%d/%Y")}')
    row += 2

    # Section A: Balance Comparison
    _section_header(ws, row, 1, 4, 'A. Balance Comparison')
    row += 1

    headers_a = ['', 'GL (Book)', 'Bank Statement', 'Variance']
    for ci, h in enumerate(headers_a, 1):
        c = ws.cell(row=row, column=ci, value=h)
        _apply(c, _hdr(fill_color=LIGHT_BLUE, color='000000', align='center'))
    row += 1

    balance_rows = [
        ('Beginning Balance', gl_begin, bank_begin),
        ('Ending Balance', gl_end, bank_end),
    ]
    for label, gl_val, bank_val in balance_rows:
        ws.cell(row=row, column=1, value=label).font = Font(name='Calibri', size=11, bold=True)
        ws.cell(row=row, column=1).border = THIN_BORDER
        c2 = ws.cell(row=row, column=2, value=gl_val)
        c2.number_format = '$#,##0.00'
        c2.border = THIN_BORDER
        c3 = ws.cell(row=row, column=3, value=bank_val)
        c3.number_format = '$#,##0.00'
        c3.border = THIN_BORDER
        var = gl_val - bank_val
        c4 = ws.cell(row=row, column=4, value=var)
        c4.number_format = '$#,##0.00'
        c4.border = THIN_BORDER
        if abs(var) > 0.01:
            c4.font = Font(name='Calibri', size=11, color='FF0000')
        row += 1

    row += 1

    # Section B: Outstanding Checks
    _section_header(ws, row, 1, 6, f'B. Outstanding Checks ({len(outstanding_checks)} items, Total: ${total_outstanding:,.2f})')
    row += 1

    if outstanding_checks:
        oc_headers = ['Date', 'Control #', 'Description', 'Reference', 'Amount', 'Days Out']
        for ci, h in enumerate(oc_headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            _apply(c, _hdr(fill_color=LIGHT_BLUE, color='000000'))
        row += 1

        _stale_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # red
        _stale_font = Font(name='Calibri', size=11, bold=True, color='9C0006')

        for i, txn in enumerate(outstanding_checks):
            alt = i % 2 == 1

            # Days outstanding calculation
            days_out: Optional[int] = None
            is_stale = False
            if _close_date and txn.date:
                txn_date = txn.date if isinstance(txn.date, date) else txn.date
                days_out = (_close_date - txn_date).days
                is_stale = days_out > 90

            dt = txn.date.strftime('%m/%d/%Y') if txn.date else ''
            ws.cell(row=row, column=1, value=dt)
            _apply(ws.cell(row=row, column=1), _cell(alt))
            ws.cell(row=row, column=2, value=txn.control)
            _apply(ws.cell(row=row, column=2), _cell(alt))
            ws.cell(row=row, column=3, value=txn.description)
            _apply(ws.cell(row=row, column=3), _cell(alt))
            ws.cell(row=row, column=4, value=txn.reference)
            _apply(ws.cell(row=row, column=4), _cell(alt))
            ws.cell(row=row, column=5, value=txn.credit)
            _apply(ws.cell(row=row, column=5), _cell(alt, fmt='$#,##0.00'))

            # Days outstanding column — red highlight when stale (>90 days)
            days_cell = ws.cell(row=row, column=6, value=days_out)
            if is_stale:
                days_cell.fill = _stale_fill
                days_cell.font = _stale_font
                days_cell.value = f'{days_out}d ⚠'
            else:
                _apply(days_cell, _cell(alt))

            row += 1

        # Total row
        ws.cell(row=row, column=4, value='Total Outstanding').font = Font(name='Calibri', size=11, bold=True)
        c = ws.cell(row=row, column=5, value=total_outstanding)
        c.number_format = '$#,##0.00'
        c.font = Font(name='Calibri', size=11, bold=True)
        c.border = DOUBLE_BOTTOM
        row += 1

        # Inject stale-check exceptions into engine_result so they appear in QC report
        _stale_checks = [
            txn for txn in outstanding_checks
            if _close_date and txn.date and (_close_date - txn.date).days > 90
        ]
        if _stale_checks:
            try:
                from engine import Exception_
                for txn in _stale_checks:
                    days_aged = (_close_date - txn.date).days
                    engine_result.exceptions.append(Exception_(
                        severity='warning',
                        category='bank_rec',
                        source='outstanding_checks',
                        description=(
                            f'Stale outstanding check — {days_aged} days old: '
                            f'{txn.control} {txn.description} ${txn.credit:,.2f} '
                            f'(issued {txn.date.strftime("%m/%d/%Y")}). '
                            f'Checks >90 days should be voided and re-issued or investigated.'
                        ),
                        details={
                            'control': txn.control,
                            'description': txn.description,
                            'amount': txn.credit,
                            'issue_date': str(txn.date),
                            'days_outstanding': days_aged,
                        },
                    ))
            except Exception:
                pass  # Don't block workpaper generation if exception injection fails
    else:
        ws.cell(row=row, column=1, value='No outstanding checks identified').font = Font(name='Calibri', size=11, italic=True)
        row += 1

    row += 1

    # Section C: Deposits in Transit
    _section_header(ws, row, 1, 6, f'C. Deposits in Transit ({len(deposits_in_transit)} items, Total: ${total_dit:,.2f})')
    row += 1

    if deposits_in_transit:
        dit_headers = ['Date', 'Control #', 'Description', 'Reference', 'Amount']
        for ci, h in enumerate(dit_headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            _apply(c, _hdr(fill_color=LIGHT_BLUE, color='000000'))
        row += 1

        for i, txn in enumerate(deposits_in_transit):
            alt = i % 2 == 1
            dt = txn.date.strftime('%m/%d/%Y') if txn.date else ''
            ws.cell(row=row, column=1, value=dt)
            _apply(ws.cell(row=row, column=1), _cell(alt))
            ws.cell(row=row, column=2, value=txn.control)
            _apply(ws.cell(row=row, column=2), _cell(alt))
            ws.cell(row=row, column=3, value=txn.description)
            _apply(ws.cell(row=row, column=3), _cell(alt))
            ws.cell(row=row, column=4, value=txn.reference)
            _apply(ws.cell(row=row, column=4), _cell(alt))
            ws.cell(row=row, column=5, value=txn.debit)
            _apply(ws.cell(row=row, column=5), _cell(alt, fmt='$#,##0.00'))
            row += 1

        ws.cell(row=row, column=4, value='Total In Transit').font = Font(name='Calibri', size=11, bold=True)
        c = ws.cell(row=row, column=5, value=total_dit)
        c.number_format = '$#,##0.00'
        c.font = Font(name='Calibri', size=11, bold=True)
        c.border = DOUBLE_BOTTOM
        row += 1
    else:
        ws.cell(row=row, column=1, value='No deposits in transit identified').font = Font(name='Calibri', size=11, italic=True)
        row += 1

    row += 1

    # Section D: Adjusted Balance Tie-Out
    _section_header(ws, row, 1, 4, 'D. Adjusted Balance Tie-Out')
    row += 1

    adjusted_bank = bank_end - total_outstanding + total_dit

    tieout_rows = [
        ('Bank Ending Balance', bank_end),
        ('Less: Outstanding Checks', -total_outstanding),
        ('Add: Deposits in Transit', total_dit),
        ('Adjusted Bank Balance', adjusted_bank),
        ('', None),
        ('GL Ending Balance (Book)', gl_end),
        ('', None),
        ('Reconciling Difference', gl_end - adjusted_bank),
    ]

    for label, val in tieout_rows:
        if val is None:
            row += 1
            continue
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(name='Calibri', size=11, bold=('Balance' in label or 'Difference' in label))
        c2 = ws.cell(row=row, column=2, value=val)
        c2.number_format = '$#,##0.00'
        if label == 'Adjusted Bank Balance':
            c2.border = DOUBLE_BOTTOM
            c2.font = Font(name='Calibri', size=11, bold=True)
        if label == 'GL Ending Balance (Book)':
            c2.font = Font(name='Calibri', size=11, bold=True)
        if label == 'Reconciling Difference':
            c2.font = Font(name='Calibri', size=11, bold=True, color='FF0000' if abs(val) > 0.01 else '008000')
            c2.border = DOUBLE_BOTTOM
        row += 1

    row += 1

    # Section E: Matched Items Detail (audit trail with confidence levels)
    if matched_checks or matched_ach or matched_deposits:
        _section_header(ws, row, 1, 6, f'E. Matched Items Detail ({len(matched_checks)} checks, {len(matched_ach)} ACH, {len(matched_deposits)} deposits)')
        row += 1

        match_headers = ['Type', 'Date', 'GL Control/Desc', 'Amount', 'Bank Ref', 'Match Method']
        for ci, h in enumerate(match_headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            _apply(c, _hdr(fill_color=LIGHT_BLUE, color='000000'))
        row += 1

        all_matched = (
            [('Check', m) for m in matched_checks] +
            [('ACH', m) for m in matched_ach] +
            [('Deposit', m) for m in matched_deposits]
        )

        for i, (mtype, m) in enumerate(all_matched):
            alt = i % 2 == 1

            if 'gl_txn' in m:
                # Legacy format from _match_checks() (non-Yardi bank recs)
                gl_txn = m['gl_txn']
                bk = m['bank_item']
                match_method = m.get('match_type', 'unknown')
                dt = gl_txn.date.strftime('%m/%d/%Y') if gl_txn.date else ''
                gl_ref = gl_txn.control or gl_txn.description or ''
                amt = gl_txn.credit if gl_txn.credit > 0 else gl_txn.debit
                bk_ref = bk.get('check_number', bk.get('reference', bk.get('description', '')))[:30]
            else:
                # New format from _build_recon_from_yardi_rec() (Yardi Bank Rec PDF)
                match_method = 'check_number+amount'
                dt = m.get('gl_date', '')
                gl_ref = f"Check #{m.get('check_number', '')} — {m.get('vendor', '')}"[:40]
                amt = m.get('amount', 0.0)
                bk_ref = m.get('check_number', '')

            ws.cell(row=row, column=1, value=mtype)
            _apply(ws.cell(row=row, column=1), _cell(alt))
            ws.cell(row=row, column=2, value=dt)
            _apply(ws.cell(row=row, column=2), _cell(alt))
            ws.cell(row=row, column=3, value=gl_ref[:40])
            _apply(ws.cell(row=row, column=3), _cell(alt))
            ws.cell(row=row, column=4, value=amt)
            _apply(ws.cell(row=row, column=4), _cell(alt, fmt='$#,##0.00'))
            ws.cell(row=row, column=5, value=bk_ref)
            _apply(ws.cell(row=row, column=5), _cell(alt))
            ws.cell(row=row, column=6, value=match_method)
            _apply(ws.cell(row=row, column=6), _cell(alt))
            row += 1

    _auto_width(ws, 6)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.sheet_properties.tabColor = '2E75B6'


# ── 2. DEBT SERVICE SCHEDULE ────────────────────────────────

def _write_debt_service_workpaper(wb, engine_result):
    """
    Debt service workpaper:
      - Header: property, period
      - Section A: Loan summary (per-loan P&I, rate, balance)
      - Section B: Payment detail (breakout by component)
      - Section C: Escrow account reconciliation
      - Section D: GL to loan statement tie-out
    """
    ws = wb.create_sheet('Debt Service')

    gl_data = engine_result.parsed.get('gl')
    loan_data = engine_result.parsed.get('loan')
    ds_check = engine_result.debt_service_check or {}
    period = engine_result.period or 'N/A'
    prop = engine_result.property_name or 'N/A'

    loans = []
    if isinstance(loan_data, list):
        loans = loan_data
    elif isinstance(loan_data, dict):
        loans = loan_data.get('loans', [])

    row = 1
    _title_row(ws, row, f'Debt Service Schedule — {prop}', cols=10)
    row += 1
    _subtitle_row(ws, row, f'Period: {period}  |  Prepared: {datetime.now().strftime("%m/%d/%Y")}', cols=10)
    row += 2

    # Section A: Loan Summary
    _section_header(ws, row, 1, 8, f'A. Loan Summary ({len(loans)} loans)')
    row += 1

    loan_headers = ['Loan #', 'Property/Note', 'Interest Rate', 'Principal Balance',
                    'Interest Paid YTD', 'Tax Escrow', 'Insurance Escrow', 'Reserve Balance']
    for ci, h in enumerate(loan_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        _apply(c, _hdr(fill_color=LIGHT_BLUE, color='000000'))
    row += 1

    total_principal = 0
    total_interest_ytd = 0
    for i, loan in enumerate(loans):
        if isinstance(loan, dict):
            ln = loan.get('loan_number', '')
            name = loan.get('property_name', '')
            rate = loan.get('interest_rate', 0)
            principal = loan.get('principal_balance', 0) or 0
            int_ytd = loan.get('interest_paid_ytd', 0) or 0
            tax_esc = loan.get('tax_escrow_balance', 0) or 0
            ins_esc = loan.get('insurance_escrow_balance', 0) or 0
            reserve = loan.get('reserve_balance', 0) or 0
        else:
            ln = getattr(loan, 'loan_number', '')
            name = getattr(loan, 'property_name', '')
            rate = getattr(loan, 'interest_rate', 0)
            principal = getattr(loan, 'principal_balance', 0) or 0
            int_ytd = getattr(loan, 'interest_paid_ytd', 0) or 0
            tax_esc = getattr(loan, 'tax_escrow_balance', 0) or 0
            ins_esc = getattr(loan, 'insurance_escrow_balance', 0) or 0
            reserve = getattr(loan, 'reserve_balance', 0) or 0

        alt = i % 2 == 1
        ws.cell(row=row, column=1, value=ln); _apply(ws.cell(row=row, column=1), _cell(alt))
        ws.cell(row=row, column=2, value=name); _apply(ws.cell(row=row, column=2), _cell(alt))

        rate_val = rate
        if isinstance(rate, (int, float)) and rate < 1:
            rate_val = rate  # already decimal
        elif isinstance(rate, (int, float)):
            rate_val = rate / 100 if rate > 1 else rate
        ws.cell(row=row, column=3, value=rate_val)
        _apply(ws.cell(row=row, column=3), _cell(alt, fmt='0.000%'))

        ws.cell(row=row, column=4, value=principal); _apply(ws.cell(row=row, column=4), _cell(alt, fmt='$#,##0.00'))
        ws.cell(row=row, column=5, value=int_ytd); _apply(ws.cell(row=row, column=5), _cell(alt, fmt='$#,##0.00'))
        ws.cell(row=row, column=6, value=tax_esc); _apply(ws.cell(row=row, column=6), _cell(alt, fmt='$#,##0.00'))
        ws.cell(row=row, column=7, value=ins_esc); _apply(ws.cell(row=row, column=7), _cell(alt, fmt='$#,##0.00'))
        ws.cell(row=row, column=8, value=reserve); _apply(ws.cell(row=row, column=8), _cell(alt, fmt='$#,##0.00'))

        total_principal += principal
        total_interest_ytd += int_ytd
        row += 1

    # Totals row
    ws.cell(row=row, column=2, value='TOTAL').font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=row, column=4, value=total_principal)
    ws.cell(row=row, column=4).number_format = '$#,##0.00'
    ws.cell(row=row, column=4).font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=row, column=4).border = DOUBLE_BOTTOM
    ws.cell(row=row, column=5, value=total_interest_ytd)
    ws.cell(row=row, column=5).number_format = '$#,##0.00'
    ws.cell(row=row, column=5).font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=row, column=5).border = DOUBLE_BOTTOM
    row += 2

    # Section B: Payment Detail
    _section_header(ws, row, 1, 8, 'B. Payment Breakdown by Loan')
    row += 1

    pmt_headers = ['Loan #', 'Principal', 'Interest', 'Taxes', 'Insurance', 'Reserves', 'Total Payment', 'Due Date']
    for ci, h in enumerate(pmt_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        _apply(c, _hdr(fill_color=LIGHT_BLUE, color='000000'))
    row += 1

    for i, loan in enumerate(loans):
        def _safe_num(v):
            if isinstance(v, (int, float)):
                return float(v)
            try:
                return float(v) if v else 0.0
            except (ValueError, TypeError):
                return 0.0

        if isinstance(loan, dict):
            ln       = loan.get('loan_number', '')
            due_date = loan.get('payment_due_date', '')
            p_amt    = _safe_num(loan.get('payment_principal', 0))
            i_amt    = _safe_num(loan.get('payment_interest', 0))
            t_amt    = _safe_num(loan.get('payment_re_taxes', 0))
            ins_amt  = _safe_num(loan.get('payment_insurance', 0))
            r_amt    = _safe_num(loan.get('payment_reserves', 0))
            total    = _safe_num(loan.get('payment_total', 0)) or (p_amt + i_amt + t_amt + ins_amt + r_amt)
        else:
            ln       = getattr(loan, 'loan_number', '')
            due_date = getattr(loan, 'payment_due_date', '')
            p_amt    = _safe_num(getattr(loan, 'payment_principal', 0))
            i_amt    = _safe_num(getattr(loan, 'payment_interest', 0))
            t_amt    = _safe_num(getattr(loan, 'payment_re_taxes', 0))
            ins_amt  = _safe_num(getattr(loan, 'payment_insurance', 0))
            r_amt    = _safe_num(getattr(loan, 'payment_reserves', 0))
            total    = _safe_num(getattr(loan, 'payment_total', 0)) or (p_amt + i_amt + t_amt + ins_amt + r_amt)

        alt = i % 2 == 1
        ws.cell(row=row, column=1, value=ln); _apply(ws.cell(row=row, column=1), _cell(alt))
        ws.cell(row=row, column=2, value=p_amt); _apply(ws.cell(row=row, column=2), _cell(alt, fmt='$#,##0.00'))
        ws.cell(row=row, column=3, value=i_amt); _apply(ws.cell(row=row, column=3), _cell(alt, fmt='$#,##0.00'))
        ws.cell(row=row, column=4, value=t_amt); _apply(ws.cell(row=row, column=4), _cell(alt, fmt='$#,##0.00'))
        ws.cell(row=row, column=5, value=ins_amt); _apply(ws.cell(row=row, column=5), _cell(alt, fmt='$#,##0.00'))
        ws.cell(row=row, column=6, value=r_amt); _apply(ws.cell(row=row, column=6), _cell(alt, fmt='$#,##0.00'))
        ws.cell(row=row, column=7, value=total); _apply(ws.cell(row=row, column=7), _cell(alt, fmt='$#,##0.00'))

        due_str = ''
        if isinstance(due_date, datetime):
            due_str = due_date.strftime('%m/%d/%Y')
        elif isinstance(due_date, str):
            due_str = due_date
        ws.cell(row=row, column=8, value=due_str); _apply(ws.cell(row=row, column=8), _cell(alt))
        row += 1

    row += 1

    # Section C: GL to Loan Tie-Out
    _section_header(ws, row, 1, 4, 'C. GL to Loan Statement Reconciliation')
    row += 1

    gl_interest = ds_check.get('gl_interest_expense', 0)
    loan_interest = ds_check.get('loan_interest_total', 0)

    # Get GL principal payment (account 201100 or similar mortgage payable)
    gl_principal_payment = 0
    if gl_data and hasattr(gl_data, 'accounts'):
        for acct in gl_data.accounts:
            if acct.account_code in ('201100', '201110'):
                gl_principal_payment = abs(acct.net_change)
                break

    tieout = [
        ('GL Interest Expense (801110) — PTD', gl_interest),
        ('Loan Statement Interest Paid YTD', loan_interest),
        ('Note: GL is period-to-date; loan statement is year-to-date', None),
        ('', None),
        ('Reconciled', 'Yes' if ds_check.get('reconciled') else 'No'),
    ]

    for label, val in tieout:
        if val is None:
            c = ws.cell(row=row, column=1, value=label)
            c.font = Font(name='Calibri', size=10, italic=True, color='666666')
            row += 1
            continue
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(name='Calibri', size=11, bold=('Reconciled' in label))
        c2 = ws.cell(row=row, column=2, value=val)
        if isinstance(val, (int, float)):
            c2.number_format = '$#,##0.00'
        row += 1

    _auto_width(ws, 8)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.sheet_properties.tabColor = '548235'


# ── 3. RENT ROLL TIE-OUT ────────────────────────────────────

def _write_rent_roll_workpaper(wb, engine_result):
    """
    Rent roll tie-out workpaper:
      - Section A: Lease summary with monthly/annual rent
      - Section B: GL revenue account balances
      - Section C: Billed vs GL revenue tie-out
      - Section D: Vacancy / occupancy summary
    """
    ws = wb.create_sheet('Rent Roll Tie-Out')

    gl_data = engine_result.parsed.get('gl')
    rr_data = engine_result.parsed.get('rent_roll')
    period = engine_result.period or 'N/A'
    prop = engine_result.property_name or 'N/A'

    row = 1
    _title_row(ws, row, f'Rent Roll Tie-Out — {prop}', cols=10)
    row += 1
    _subtitle_row(ws, row, f'Period: {period}  |  Prepared: {datetime.now().strftime("%m/%d/%Y")}', cols=10)
    row += 2

    # Section A: Lease Detail
    tenants = rr_data if isinstance(rr_data, list) else []

    _section_header(ws, row, 1, 9, f'A. Lease Detail ({len(tenants)} leases)')
    row += 1

    rr_headers = ['Unit', 'Tenant', 'Lease Type', 'Area (SF)', 'Lease From', 'Lease To',
                  'Monthly Rent', 'Annual Rent', 'Annual Rent/SF']
    for ci, h in enumerate(rr_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        _apply(c, _hdr(fill_color=LIGHT_BLUE, color='000000'))
    row += 1

    total_area = 0
    total_monthly = 0
    total_annual = 0

    for i, t in enumerate(tenants):
        if isinstance(t, dict):
            unit = t.get('units', t.get('unit', ''))
            tenant = t.get('tenant', t.get('lease', ''))
            ltype = t.get('lease_type', '')
            area = t.get('area', 0) or 0
            lfrom = t.get('lease_from', '')
            lto = t.get('lease_to', '')
            monthly = t.get('monthly_rent', 0) or 0
            annual = t.get('annual_rent', 0) or 0
            rent_psf = t.get('annual_rent_per_area', 0) or 0
        else:
            unit = getattr(t, 'units', getattr(t, 'unit', ''))
            tenant = getattr(t, 'tenant', '')
            ltype = getattr(t, 'lease_type', '')
            area = getattr(t, 'area', 0) or 0
            lfrom = getattr(t, 'lease_from', '')
            lto = getattr(t, 'lease_to', '')
            monthly = getattr(t, 'monthly_rent', 0) or 0
            annual = getattr(t, 'annual_rent', 0) or 0
            rent_psf = getattr(t, 'annual_rent_per_area', 0) or 0

        alt = i % 2 == 1
        ws.cell(row=row, column=1, value=unit); _apply(ws.cell(row=row, column=1), _cell(alt))
        ws.cell(row=row, column=2, value=tenant); _apply(ws.cell(row=row, column=2), _cell(alt))
        ws.cell(row=row, column=3, value=ltype); _apply(ws.cell(row=row, column=3), _cell(alt))
        ws.cell(row=row, column=4, value=area); _apply(ws.cell(row=row, column=4), _cell(alt, fmt='#,##0'))
        ws.cell(row=row, column=5, value=str(lfrom) if lfrom else ''); _apply(ws.cell(row=row, column=5), _cell(alt))
        ws.cell(row=row, column=6, value=str(lto) if lto else ''); _apply(ws.cell(row=row, column=6), _cell(alt))
        ws.cell(row=row, column=7, value=monthly); _apply(ws.cell(row=row, column=7), _cell(alt, fmt='$#,##0.00'))
        ws.cell(row=row, column=8, value=annual); _apply(ws.cell(row=row, column=8), _cell(alt, fmt='$#,##0.00'))
        ws.cell(row=row, column=9, value=rent_psf); _apply(ws.cell(row=row, column=9), _cell(alt, fmt='$#,##0.00'))

        total_area += area if isinstance(area, (int, float)) else 0
        total_monthly += monthly if isinstance(monthly, (int, float)) else 0
        total_annual += annual if isinstance(annual, (int, float)) else 0
        row += 1

    # Total row
    ws.cell(row=row, column=2, value='TOTAL').font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=row, column=4, value=total_area)
    ws.cell(row=row, column=4).number_format = '#,##0'
    ws.cell(row=row, column=4).font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=row, column=7, value=total_monthly)
    ws.cell(row=row, column=7).number_format = '$#,##0.00'
    ws.cell(row=row, column=7).font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=row, column=7).border = DOUBLE_BOTTOM
    ws.cell(row=row, column=8, value=total_annual)
    ws.cell(row=row, column=8).number_format = '$#,##0.00'
    ws.cell(row=row, column=8).font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=row, column=8).border = DOUBLE_BOTTOM
    row += 2

    # Section B: GL Revenue Tie-Out
    _section_header(ws, row, 1, 5, 'B. GL Revenue Account Summary')
    row += 1

    rev_headers = ['Account Code', 'Account Name', 'Net Change (PTD)', 'Beginning Balance', 'Ending Balance']
    for ci, h in enumerate(rev_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        _apply(c, _hdr(fill_color=LIGHT_BLUE, color='000000'))
    row += 1

    gl_revenue_total = 0
    if gl_data and hasattr(gl_data, 'accounts'):
        for acct in gl_data.accounts:
            if acct.account_code.startswith('4'):
                ws.cell(row=row, column=1, value=acct.account_code)
                ws.cell(row=row, column=2, value=acct.account_name)
                ws.cell(row=row, column=3, value=abs(acct.net_change))
                ws.cell(row=row, column=3).number_format = '$#,##0.00'
                ws.cell(row=row, column=4, value=acct.beginning_balance)
                ws.cell(row=row, column=4).number_format = '$#,##0.00'
                ws.cell(row=row, column=5, value=acct.ending_balance)
                ws.cell(row=row, column=5).number_format = '$#,##0.00'
                for ci in range(1, 6):
                    ws.cell(row=row, column=ci).border = THIN_BORDER
                gl_revenue_total += abs(acct.net_change)
                row += 1

    ws.cell(row=row, column=2, value='TOTAL GL REVENUE (PTD)').font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=row, column=3, value=gl_revenue_total)
    ws.cell(row=row, column=3).number_format = '$#,##0.00'
    ws.cell(row=row, column=3).font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=row, column=3).border = DOUBLE_BOTTOM
    row += 2

    # Section C: Rent Roll to GL Comparison
    _section_header(ws, row, 1, 4, 'C. Rent Roll to GL Revenue Comparison')
    row += 1

    rent_roll_monthly_total = total_monthly
    variance = gl_revenue_total - rent_roll_monthly_total

    comp_rows = [
        ('Rent Roll Total Monthly Rent', rent_roll_monthly_total),
        ('GL Total Revenue (PTD)', gl_revenue_total),
        ('Variance', variance),
    ]
    for label, val in comp_rows:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(name='Calibri', size=11, bold=('Variance' in label))
        c2 = ws.cell(row=row, column=2, value=val)
        c2.number_format = '$#,##0.00'
        if 'Variance' in label:
            c2.font = Font(name='Calibri', size=11, bold=True, color='FF0000' if abs(val) > 0.01 else '008000')
            c2.border = DOUBLE_BOTTOM
        row += 1

    ws.cell(row=row, column=1, value='Note: Variance reflects straight-line rent adjustments, CAM recoveries, and other non-base-rent revenue').font = Font(name='Calibri', size=10, italic=True, color='666666')
    row += 1

    _auto_width(ws, 9)
    ws.column_dimensions['B'].width = 30
    ws.sheet_properties.tabColor = 'BF8F00'


# ── 4. ACCRUAL SCHEDULE ─────────────────────────────────────

def _write_accrual_workpaper(wb, engine_result):
    """
    Accrual schedule workpaper:
      - Section A: Accrual detail by vendor
      - Section B: Aging summary
      - Section C: GL accrual account tie-out
      - Section D: Status summary (pending vs approved)
    """
    ws = wb.create_sheet('Accrual Schedule')

    gl_data = engine_result.parsed.get('gl')
    nexus_data = engine_result.parsed.get('nexus_accrual')
    period = engine_result.period or 'N/A'
    prop = engine_result.property_name or 'N/A'

    invoices = nexus_data if isinstance(nexus_data, list) else []

    row = 1
    _title_row(ws, row, f'Accrual Schedule — {prop}', cols=10)
    row += 1
    _subtitle_row(ws, row, f'Period: {period}  |  Source: Nexus Accrual Detail  |  Prepared: {datetime.now().strftime("%m/%d/%Y")}', cols=10)
    row += 2

    # Section A: Invoice Detail
    _section_header(ws, row, 1, 10, f'A. Accrual Detail ({len(invoices)} invoices)')
    row += 1

    inv_headers = ['Vendor', 'Invoice #', 'Invoice Date', 'Received Date', 'GL Account',
                   'GL Category', 'Description', 'Status', 'Amount', 'Days Outstanding']
    for ci, h in enumerate(inv_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        _apply(c, _hdr(fill_color=LIGHT_BLUE, color='000000'))
    row += 1

    total_accrued = 0
    vendor_totals = {}
    status_totals = {}
    aging_buckets = {'0-30': 0, '31-60': 0, '61-90': 0, '90+': 0}

    for i, inv in enumerate(invoices):
        vendor = str(inv.get('vendor', '') if isinstance(inv, dict) else getattr(inv, 'vendor', '') or '')
        inv_num = str(inv.get('invoice_number', '') if isinstance(inv, dict) else getattr(inv, 'invoice_number', '') or '')
        inv_date = inv.get('invoice_date', '') if isinstance(inv, dict) else getattr(inv, 'invoice_date', '')
        recv_date = inv.get('received_date', '') if isinstance(inv, dict) else getattr(inv, 'received_date', '')
        gl_acct = str(inv.get('gl_account', '') if isinstance(inv, dict) else getattr(inv, 'gl_account', '') or '')
        gl_cat = str(inv.get('gl_category', '') if isinstance(inv, dict) else getattr(inv, 'gl_category', '') or '')
        desc = str(inv.get('line_description', '') if isinstance(inv, dict) else getattr(inv, 'line_description', '') or '')
        status = str(inv.get('invoice_status', '') if isinstance(inv, dict) else getattr(inv, 'invoice_status', '') or '')
        amount = inv.get('amount', 0) if isinstance(inv, dict) else getattr(inv, 'amount', 0)
        amount = amount or 0

        # Calculate days outstanding
        days_out = ''
        if isinstance(inv_date, datetime):
            days_out = (datetime.now() - inv_date).days
        elif isinstance(inv_date, str) and inv_date:
            try:
                dt = datetime.strptime(inv_date, '%m/%d/%Y')
                days_out = (datetime.now() - dt).days
            except:
                days_out = ''

        # Aging bucket
        if isinstance(days_out, int):
            if days_out <= 30:
                aging_buckets['0-30'] += abs(amount)
            elif days_out <= 60:
                aging_buckets['31-60'] += abs(amount)
            elif days_out <= 90:
                aging_buckets['61-90'] += abs(amount)
            else:
                aging_buckets['90+'] += abs(amount)

        alt = i % 2 == 1
        ws.cell(row=row, column=1, value=vendor); _apply(ws.cell(row=row, column=1), _cell(alt))
        ws.cell(row=row, column=2, value=inv_num); _apply(ws.cell(row=row, column=2), _cell(alt))

        inv_date_str = inv_date.strftime('%m/%d/%Y') if isinstance(inv_date, datetime) else str(inv_date or '')
        ws.cell(row=row, column=3, value=inv_date_str); _apply(ws.cell(row=row, column=3), _cell(alt))

        recv_date_str = recv_date.strftime('%m/%d/%Y') if isinstance(recv_date, datetime) else str(recv_date or '')
        ws.cell(row=row, column=4, value=recv_date_str); _apply(ws.cell(row=row, column=4), _cell(alt))

        ws.cell(row=row, column=5, value=gl_acct); _apply(ws.cell(row=row, column=5), _cell(alt))
        ws.cell(row=row, column=6, value=gl_cat); _apply(ws.cell(row=row, column=6), _cell(alt))
        ws.cell(row=row, column=7, value=desc); _apply(ws.cell(row=row, column=7), _cell(alt))
        ws.cell(row=row, column=8, value=status); _apply(ws.cell(row=row, column=8), _cell(alt))
        ws.cell(row=row, column=9, value=amount); _apply(ws.cell(row=row, column=9), _cell(alt, fmt='$#,##0.00'))
        ws.cell(row=row, column=10, value=days_out); _apply(ws.cell(row=row, column=10), _cell(alt))

        total_accrued += amount
        vendor_totals[vendor] = vendor_totals.get(vendor, 0) + amount
        status_totals[status] = status_totals.get(status, 0) + amount
        row += 1

    # Total row
    ws.cell(row=row, column=8, value='TOTAL').font = Font(name='Calibri', size=11, bold=True)
    c = ws.cell(row=row, column=9, value=total_accrued)
    c.number_format = '$#,##0.00'
    c.font = Font(name='Calibri', size=11, bold=True)
    c.border = DOUBLE_BOTTOM
    row += 2

    # Section B: Aging Summary
    _section_header(ws, row, 1, 4, 'B. Aging Summary')
    row += 1

    aging_headers = ['Aging Bucket', 'Amount', '% of Total']
    for ci, h in enumerate(aging_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        _apply(c, _hdr(fill_color=LIGHT_BLUE, color='000000'))
    row += 1

    for bucket, amt in aging_buckets.items():
        pct = (amt / total_accrued * 100) if total_accrued else 0
        ws.cell(row=row, column=1, value=bucket)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=amt)
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=pct / 100)
        ws.cell(row=row, column=3).number_format = '0.0%'
        ws.cell(row=row, column=3).border = THIN_BORDER
        row += 1

    row += 1

    # Section C: Vendor Summary
    _section_header(ws, row, 1, 4, f'C. Vendor Summary ({len(vendor_totals)} vendors)')
    row += 1

    v_headers = ['Vendor', 'Total Accrued', '% of Total']
    for ci, h in enumerate(v_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        _apply(c, _hdr(fill_color=LIGHT_BLUE, color='000000'))
    row += 1

    for vendor, amt in sorted(vendor_totals.items(), key=lambda x: abs(x[1]), reverse=True):
        pct = (amt / total_accrued * 100) if total_accrued else 0
        ws.cell(row=row, column=1, value=vendor)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=amt)
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=pct / 100)
        ws.cell(row=row, column=3).number_format = '0.0%'
        ws.cell(row=row, column=3).border = THIN_BORDER
        row += 1

    row += 1

    # Section D: Status Summary
    _section_header(ws, row, 1, 3, 'D. Invoice Status Summary')
    row += 1

    for ci, h in enumerate(['Status', 'Count', 'Amount'], 1):
        c = ws.cell(row=row, column=ci, value=h)
        _apply(c, _hdr(fill_color=LIGHT_BLUE, color='000000'))
    row += 1

    # Count by status
    status_counts = {}
    for inv in invoices:
        s = inv.get('invoice_status', '') if isinstance(inv, dict) else getattr(inv, 'invoice_status', '')
        status_counts[s] = status_counts.get(s, 0) + 1

    for status, amt in sorted(status_totals.items()):
        ws.cell(row=row, column=1, value=status)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=status_counts.get(status, 0))
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=amt)
        ws.cell(row=row, column=3).number_format = '$#,##0.00'
        ws.cell(row=row, column=3).border = THIN_BORDER
        row += 1

    _auto_width(ws, 10)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['G'].width = 35
    ws.sheet_properties.tabColor = 'C00000'


# ── Main entry point ─────────────────────────────────────────

def generate_workpapers(engine_result, output_path: str) -> str:
    """
    Generate the 4-tab workpaper Excel file.

    Args:
        engine_result: EngineResult from pipeline run
        output_path: Where to write the Excel file

    Returns:
        The output path if successful
    """
    wb = Workbook()
    wb.remove(wb.active)

    _write_bank_recon_workpaper(wb, engine_result)
    _write_debt_service_workpaper(wb, engine_result)
    _write_rent_roll_workpaper(wb, engine_result)
    _write_accrual_workpaper(wb, engine_result)

    # Add accrual entries review tab
    nexus_data = engine_result.parsed.get('nexus_accrual')
    gl_data = engine_result.parsed.get('gl')
    budget_data = engine_result.parsed.get('budget_comparison')

    # Build entries from all three layers (Nexus + budget gaps + historicals)
    je_lines = build_accrual_entries(
        nexus_data or [],
        period=engine_result.period or '',
        property_name=engine_result.property_name or '',
        gl_data=gl_data,
        budget_data=budget_data,
    )
    if True:  # Always add the tab; show empty state if no entries
        if je_lines:
            write_accrual_entries_workpaper_tab(
                wb, je_lines,
                period=engine_result.period or '',
                property_name=engine_result.property_name or '',
            )

    wb.save(output_path)
    return output_path

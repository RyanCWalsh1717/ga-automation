"""
Management Fee Invoice Generator
=================================
Populates the GRP Excel invoice template and exports to PDF.

Template: pipeline/templates/mgmt_fee_invoice_template.xlsx
  (REv Labs Management fee invoice tempalte.xlsx — exact formatting preserved)

Only 4 cells need to be written each month:
  H7  — Invoice number  (RevLabsPM{MM}{YYYY})
  H8  — Invoice date    (last day of the close month; H9 = =H8 follows automatically)
  E15 — Period label    ('January 2026 Property Management Fee')
  F15 — Collections     (cash received — all fee amounts are Excel formulas off this)

All fee calculations live as Excel formulas in the template:
  H15 = =F15 * G15          (3.00% total fee)
  H16 = =-MAX(F16*G16,5000) (JLL deduction: greater of 1.25% or $5,000)
  H19 = =IF(SUM(H15:H18)>0, SUM(H15:H18), 0)  (Balance Due = GRP's portion)

PDF conversion priority:
  1. win32com  — Windows + Microsoft Excel installed (exact rendering)
  2. LibreOffice headless  — cross-platform if `libreoffice` available
  3. reportlab fallback  — always works, close match to template
"""

from __future__ import annotations

import io
import os
import re
import shutil
import calendar
import tempfile
from datetime import date, datetime
from typing import Optional

# ── Template path ─────────────────────────────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
_TEMPLATE = os.path.join(_HERE, 'templates', 'mgmt_fee_invoice_template.xlsx')

# ── Constants ─────────────────────────────────────────────────────────────────
_JLL_MIN   = 5_000.0   # $5,000 minimum for JLL deduction (per template formula)
_JLL_RATE  = 0.0125
_TOTAL_RATE = 0.0300

_MONTH_MAP = {
    1: 'January', 2: 'February', 3: 'March',    4: 'April',
    5: 'May',     6: 'June',     7: 'July',      8: 'August',
    9: 'September', 10: 'October', 11: 'November', 12: 'December',
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_period(period: str):
    """Return (year, month_int) from 'Jan-2026'. Returns (0, 0) on failure."""
    abbr = {
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,
        'may': 5, 'jun': 6, 'jul': 7, 'aug': 8,
        'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
    }
    m = re.search(r'([A-Za-z]{3})[\s\-](\d{4})', period or '')
    if not m:
        return 0, 0
    return int(m.group(2)), abbr.get(m.group(1).lower(), 0)


def _month_end(year: int, month: int) -> date:
    return date(year, month, calendar.monthrange(year, month)[1])


# ── Excel population ──────────────────────────────────────────────────────────

def _populate_excel(period: str, cash_received: float, out_xlsx: str,
                    invoice_prefix: str = 'RevLabsPM') -> None:
    """
    Copy the invoice template and fill in the 4 dynamic cells.
    Writes the populated workbook to out_xlsx.
    """
    import openpyxl

    if not os.path.exists(_TEMPLATE):
        raise FileNotFoundError(
            f'Invoice template not found: {_TEMPLATE}\n'
            'Place mgmt_fee_invoice_template.xlsx in pipeline/templates/'
        )

    year, month = _parse_period(period)
    if not year or not month:
        raise ValueError(f"Cannot parse period '{period}'")

    inv_date   = _month_end(year, month)
    inv_num    = f'{invoice_prefix}{month:02d}{year}'
    month_lbl  = _MONTH_MAP.get(month, str(month))
    period_lbl = f'{month_lbl} {year} Property Management Fee'

    shutil.copy2(_TEMPLATE, out_xlsx)
    wb = openpyxl.load_workbook(out_xlsx)
    ws = wb.active

    # Only 4 cells need updating — everything else is static or a formula
    ws['H7'] = inv_num
    ws['H8'] = datetime(inv_date.year, inv_date.month, inv_date.day)
    ws['E15'] = period_lbl
    ws['F15'] = round(cash_received, 2)

    wb.save(out_xlsx)


# ── PDF conversion ────────────────────────────────────────────────────────────

def _pdf_via_win32com(xlsx_path: str, pdf_path: str) -> bool:
    """Convert Excel to PDF using Microsoft Excel (Windows only)."""
    try:
        import win32com.client
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
            wb.Worksheets(1).ExportAsFixedFormat(
                0,                          # xlTypePDF
                os.path.abspath(pdf_path),
                1,                          # xlQualityStandard
                True,                       # IncludeDocProperties
                False,                      # IgnorePrintAreas
            )
            wb.Close(False)
            return True
        finally:
            excel.Quit()
    except Exception:
        return False


def _pdf_via_libreoffice(xlsx_path: str, pdf_dir: str) -> bool:
    """Convert Excel to PDF using LibreOffice headless."""
    import subprocess
    try:
        result = subprocess.run(
            ['libreoffice', '--headless', '--convert-to', 'pdf',
             '--outdir', pdf_dir, os.path.abspath(xlsx_path)],
            capture_output=True, timeout=60,
        )
        return result.returncode == 0
    except Exception:
        return False


def _pdf_via_reportlab(period: str, cash_received: float, pdf_path: str,
                       invoice_prefix: str = 'RevLabsPM',
                       payment_ach: dict = None,
                       payment_check: dict = None,
                       jll_rate: float = _JLL_RATE,
                       jll_minimum: float = _JLL_MIN,
                       total_rate: float = _TOTAL_RATE,
                       bill_to: str = '') -> None:  # C-NF-4
    """
    Reportlab fallback — closely matches the template layout.
    Used when neither win32com nor LibreOffice is available.

    Table columns: DATE | DESCRIPTION | COLLECTIONS | RATE | AMOUNT
    (ACTIVITY column removed — description is self-explanatory and the
    combined text no longer overflows into adjacent columns.)
    """
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors

    GREEN_DARK  = colors.HexColor('#2A651C')
    GREEN_LIGHT = colors.HexColor('#D4E0D2')
    BLACK       = colors.black

    # Use payment_ach/payment_check from config if provided, else RevLabs defaults
    _ach   = payment_ach   or {}
    _check = payment_check or {}
    GRP_NAME = _check.get('payable_to', _ach.get('account_name', 'Greatland Realty Partners LLC'))
    GRP_A1   = _check.get('address_line1', 'One Federal Street, 28th Floor')
    GRP_A2   = _check.get('address_line2', 'Boston, MA 02110')
    GRP_WEB  = 'www.greatlandpartners.com'
    GRP_ATTN = _check.get('attention', '')              # C-NF-5: no personal name fallback
    ACH_BANK    = _ach.get('bank_name', '')
    ACH_ACCT    = _ach.get('account_number', '')         # C-NF-5: no real account in source
    ACH_ROUTING = _ach.get('routing_number', '')
    ACH_ADDR    = _ach.get('bank_address', '1 Federal St, Boston, MA 02110')

    year, month = _parse_period(period)
    inv_date   = _month_end(year, month)
    inv_date_s = f'{inv_date.month}/{inv_date.day}/{inv_date.year}'
    inv_num    = f'{invoice_prefix}{month:02d}{year}'
    month_lbl  = _MONTH_MAP.get(month, str(month))

    total_fee  = round(cash_received * total_rate, 2)
    jll_fee    = round(max(cash_received * jll_rate, jll_minimum), 2)
    balance    = round(max(total_fee - jll_fee, 0.0), 2)

    W, H = letter
    buf  = io.BytesIO()
    c    = rl_canvas.Canvas(buf, pagesize=letter)

    def _y(top): return H - top

    L, R = 57.0, 555.0
    # Two left columns: DATE (narrow) | DESCRIPTION (wide)
    DATE_X = 57.0
    DESC_X = 135.0          # starts where ACTIVITY used to — description has room to ~380
    R_COLL = 420.0          # right-aligned: Collections
    R_RATE = 468.0          # right-aligned: Rate
    R_AMT  = 555.0          # right-aligned: Amount

    # ── Company header ────────────────────────────────────────────────────────
    c.setFillColor(BLACK)
    c.setFont('Times-Bold',  10); c.drawString(L, _y(42), GRP_NAME)
    c.setFont('Times-Roman', 10)
    c.drawString(L, _y(56), GRP_A1)
    c.drawString(L, _y(70), GRP_A2)
    c.drawString(L, _y(84), GRP_WEB)

    # ── INVOICE title ─────────────────────────────────────────────────────────
    c.setFillColor(GREEN_DARK)
    c.setFont('Times-Bold', 20)
    c.drawString(L, _y(108), 'INVOICE')
    c.setFillColor(BLACK)

    # ── Meta block (right side) ───────────────────────────────────────────────
    # Rows below here start at top=140, not 112 — the 20pt INVOICE title's cap
    # height reaches up to roughly top=94, so top=112 (only 4pt past its own
    # baseline at 108) put BILL TO/META inside the title's glyph box, visibly
    # overlapping it. +28 clears it with real margin, absorbed by the large
    # (180pt, now ~152pt) gap before PAYMENT INSTRUCTIONS below.
    META = [
        ('INVOICE #', inv_num,          140),
        ('DATE',      inv_date_s,        154),
        ('DUE DATE',  inv_date_s,        168),
        ('TERMS',     'Due on receipt',  182),
    ]
    for lbl, val, top in META:
        c.setFont('Times-Bold',  10); c.drawString(355, _y(top), lbl)
        c.setFont('Times-Roman', 10); c.drawString(460, _y(top), val)

    # ── Bill To (left side, same rows as Meta) ────────────────────────────────
    c.setFont('Times-Bold',  10); c.drawString(L, _y(140), 'BILL TO:')
    c.setFont('Times-Roman', 10); c.drawString(L, _y(154), bill_to)  # C-NF-4: from config

    # ── Divider above table ───────────────────────────────────────────────────
    c.setStrokeColor(BLACK); c.setLineWidth(0.75)
    c.line(L, _y(202), R, _y(202))

    # ── Table header row (green fill) ─────────────────────────────────────────
    c.setFillColor(GREEN_LIGHT); c.setStrokeColor(GREEN_LIGHT)
    c.rect(L, _y(224), R - L, 22, fill=1, stroke=0)
    c.setStrokeColor(BLACK); c.line(L, _y(224), R, _y(224))

    c.setFont('Times-Bold', 9); c.setFillColor(BLACK)
    c.drawString(DATE_X, _y(216), 'DATE')
    c.drawString(DESC_X, _y(216), 'DESCRIPTION')
    c.setFillColor(GREEN_DARK)
    c.drawRightString(R_COLL, _y(216), 'COLLECTIONS')
    c.setFillColor(BLACK)
    c.drawRightString(R_RATE, _y(216), 'RATE')
    c.drawRightString(R_AMT,  _y(216), 'AMOUNT')

    def _money(v, neg=False):
        return f'{"-" if neg else ""}${abs(v):,.2f}'

    # ── Row 1 — Management fee ────────────────────────────────────────────────
    c.setFont('Times-Roman', 10)
    c.drawString(DATE_X, _y(242), inv_date_s)
    c.drawString(DESC_X, _y(242), f'{month_lbl} {year} Property Management Fee')
    c.drawRightString(R_COLL, _y(242), _money(cash_received))
    c.drawRightString(R_RATE, _y(242), f'{total_rate * 100:.2f}%')
    c.drawRightString(R_AMT,  _y(242), _money(total_fee))
    c.setLineWidth(0.5); c.line(L, _y(250), R, _y(250))

    # ── Row 2 — JLL deduction ─────────────────────────────────────────────────
    c.setFont('Times-Roman', 10)
    c.drawString(DATE_X, _y(266), inv_date_s)
    c.drawString(DESC_X, _y(266), 'Less: JLL Portion')
    c.drawRightString(R_COLL, _y(266), _money(cash_received))
    c.drawRightString(R_RATE, _y(266), f'{jll_rate * 100:.2f}%')
    c.drawRightString(R_AMT,  _y(266), _money(jll_fee, neg=True))
    c.setLineWidth(0.5); c.line(L, _y(274), R, _y(274))

    # ── Balance Due ───────────────────────────────────────────────────────────
    c.setLineWidth(0.75); c.line(L, _y(300), R, _y(300))
    c.setFont('Times-Bold', 10)
    c.drawString(L,         _y(312), 'BALANCE DUE')
    c.drawRightString(R_AMT, _y(312), _money(balance))

    # ── Payment instructions ──────────────────────────────────────────────────
    c.setFont('Times-Bold', 9)
    c.drawString(L, _y(464), 'PAYMENT INSTRUCTIONS:')

    # Electronic payment (left column)
    c.drawString(L, _y(480), 'Electronic Payment:')
    ach = [
        ('Account Name:',         GRP_NAME),
        ('Bank Name:',            ACH_BANK),
        ('Bank Account #:',       ACH_ACCT),
        ('Bank Routing (ABA) #:', ACH_ROUTING),
        ('Bank Address:',         ACH_ADDR),
    ]
    for i, (lbl, val) in enumerate(ach):
        ry = 496 + i * 14
        c.setFont('Times-Bold',  9); c.drawString(L,     _y(ry), lbl)
        c.setFont('Times-Roman', 9); c.drawString(170.0, _y(ry), val)

    # Check payment (right column)
    c.setFont('Times-Bold', 9)
    c.drawString(310.0, _y(480), 'Check Payment:')
    chk = [
        ('Payable to:',       GRP_NAME),
        ('Mailing Address:',  GRP_A1),
        ('',                  GRP_A2),
        ('Attention:',        GRP_ATTN),
    ]
    ry = 496
    for lbl, val in chk:
        if lbl:
            c.setFont('Times-Bold',  9); c.drawString(310.0, _y(ry), lbl)
            c.setFont('Times-Roman', 9); c.drawString(400.0, _y(ry), val)
        else:
            c.setFont('Times-Roman', 9); c.drawString(400.0, _y(ry), val)
        ry += 14

    c.showPage(); c.save()
    with open(pdf_path, 'wb') as fh:
        fh.write(buf.getvalue())


# ── Public API ────────────────────────────────────────────────────────────────

def generate_invoice(
    period: str,
    cash_received: float,
    output_path: Optional[str] = None,
    property_config=None,
) -> bytes:
    """
    Generate the management fee invoice PDF by populating the Excel template.

    Conversion priority:
      1. win32com  (Windows + Excel — exact template rendering)
      2. LibreOffice headless  (cross-platform)
      3. reportlab  (always works, close match)

    Args:
        period:        Period string e.g. 'Jan-2026'.
        cash_received: Cash collected this period (management fee basis).
        output_path:   Optional path to write the PDF. Always returns bytes.

    Returns:
        PDF bytes.
    """
    # Resolve invoice prefix, payment instructions, and fee rates from property_config
    cfg_prefix = (
        getattr(property_config, 'invoice_prefix', None) or 'RevLabsPM'
        if property_config else 'RevLabsPM'
    )
    cfg_ach   = (getattr(property_config, 'payment_ach', None)   or {}) if property_config else {}
    cfg_check = (getattr(property_config, 'payment_check', None) or {}) if property_config else {}
    # C-NF-4: bill_to from config — entity_name, display_name, or leave blank
    cfg_bill_to = ''
    if property_config:
        cfg_bill_to = (
            getattr(property_config, 'entity_name', '')
            or getattr(property_config, 'property_display_name', '')
            or getattr(property_config, 'property_name', '')
            or ''
        )

    # Pull fee rates from config so the invoice matches what was actually agreed.
    # Falls back to the module-level defaults (JLL 1.25%, total 3.00%, min $5,000)
    # when no config is supplied or the config has no management_fees defined.
    if property_config and getattr(property_config, 'management_fees', None):
        cfg_jll_rate  = getattr(property_config, 'management_fee_jll_rate', _JLL_RATE)
        cfg_total_rate = getattr(property_config, 'total_management_fee_rate', _TOTAL_RATE)
        # JLL minimum from the JLL fee line (ManagementFeeLineConfig.minimum)
        _jll_line = next(
            (f for f in property_config.management_fees if f.name.upper() == 'JLL'),
            None,
        )
        cfg_jll_min = _jll_line.minimum if _jll_line else _JLL_MIN
    else:
        cfg_jll_rate  = _JLL_RATE
        cfg_total_rate = _TOTAL_RATE
        cfg_jll_min   = _JLL_MIN

    if output_path is None:
        year, month = _parse_period(period)
        output_path = os.path.join(
            tempfile.gettempdir(),
            f'{cfg_prefix}_Invoice_{period.replace("-", "")}.pdf',
        )

    pdf_path  = output_path
    xlsx_path = pdf_path.replace('.pdf', '_populated.xlsx')

    # Step 1 — populate the Excel template
    try:
        _populate_excel(period, cash_received, xlsx_path, invoice_prefix=cfg_prefix)
    except Exception:
        # Template missing or openpyxl issue — skip straight to reportlab
        _pdf_via_reportlab(period, cash_received, pdf_path,
                           invoice_prefix=cfg_prefix,
                           payment_ach=cfg_ach, payment_check=cfg_check,
                           jll_rate=cfg_jll_rate, jll_minimum=cfg_jll_min,
                           total_rate=cfg_total_rate, bill_to=cfg_bill_to)
        with open(pdf_path, 'rb') as fh:
            return fh.read()

    # Step 2 — convert populated Excel → PDF
    pdf_ok = False

    if not pdf_ok:
        pdf_ok = _pdf_via_win32com(xlsx_path, pdf_path)

    if not pdf_ok:
        pdf_dir = os.path.dirname(os.path.abspath(pdf_path))
        if _pdf_via_libreoffice(xlsx_path, pdf_dir):
            lo_name = os.path.splitext(os.path.basename(xlsx_path))[0] + '.pdf'
            lo_path = os.path.join(pdf_dir, lo_name)
            if os.path.exists(lo_path) and lo_path != pdf_path:
                shutil.move(lo_path, pdf_path)
            pdf_ok = os.path.exists(pdf_path)

    if not pdf_ok:
        _pdf_via_reportlab(period, cash_received, pdf_path,
                           invoice_prefix=cfg_prefix,
                           payment_ach=cfg_ach, payment_check=cfg_check,
                           jll_rate=cfg_jll_rate, jll_minimum=cfg_jll_min,
                           total_rate=cfg_total_rate, bill_to=cfg_bill_to)

    try:
        os.remove(xlsx_path)
    except OSError:
        pass

    with open(pdf_path, 'rb') as fh:
        return fh.read()

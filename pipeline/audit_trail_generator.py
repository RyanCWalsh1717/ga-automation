"""
Audit Trail Generator
=====================
Produces a multi-tab Excel workbook documenting everything the pipeline built
during a monthly close cycle.  Intended audience: auditors, Lauren (CFO review),
and GRP principals who want a complete, traceable record of automated entries.

Workbook tabs:
  1.  Summary           — Pipeline run metadata, file inventory, totals, QC status
  2.  JE Log            — Every journal entry generated in Pass 1 (one row per JE pair)
  3.  Management Fee    — Fee calculation detail with basis, rates, invoice reference
  4.  Accrual Check     — Prior-month accruals vs actuals received (Item 6 reconciliation)
  5.  QC Checks         — All QC check results with findings detail
  6.  Yardi ETL CSV     — Exact rows submitted to Yardi, with auditor cross-reference columns
  7.  Methodology       — Standing accounting methodology (materiality floors, amortization
                         conventions, bonus derivation) so an auditor can confirm consistency
                         without reading the codebase
  8.  Cutoff Review     — JE lines dated outside the close period, flagged for manual review
  9.  Reconciling Aging — Outstanding bank-rec items aged as of period end; flags 60+ days
  10. Close & Signoff   — Close tracker completion status and sign-off record, cross-referenced
                         into the audit trail itself

Usage:
    from audit_trail_generator import generate_audit_trail

    path = generate_audit_trail(
        output_path   = '/tmp/GA_Audit_Trail.xlsx',
        period        = 'Mar-2026',
        property_name = 'Revolution Labs Owner, LLC',
        all_je_lines  = [...],          # from pass1_output_files['all_je_lines']
        fee_result    = fee_result,     # MgmtFeeResult dataclass
        qc_report     = qc_report,      # QCReport dataclass (or None)
        prior_accrual_check = [...],    # from check_prior_accrual_vs_actual()
        files_uploaded = {...},         # dict of {label: path} for processed files
    )
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import Any, Dict, List, Optional

from version import get_version

try:
    from close_tracker_generator import CLOSE_TRACKER_STEPS
except ImportError:
    CLOSE_TRACKER_STEPS = []

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter


# ── Palette ──────────────────────────────────────────────────────────────────
_GRP_GREEN      = '1A5C22'    # dark GRP green — main headers
_GRP_GREEN_MID  = '2E7D32'    # mid green — section headers
_GRP_GREEN_LITE = 'E8F5E9'    # very light green — alternating rows / PASS rows
_WHITE          = 'FFFFFF'
_GREY_LITE      = 'F5F5F5'
_GREY_MED       = 'EEEEEE'
_AMBER_LITE     = 'FFF9C4'    # FLAG row background
_RED_LITE       = 'FFEBEE'    # FAIL row background
_BLACK          = '000000'

# Source → (display label, fill hex)
_SOURCE_META: Dict[str, tuple] = {
    'nexus':                  ('Nexus AP',          'BBDEFB'),   # blue
    'invoice_proration':      ('Invoice Proration',  'E1BEE7'),   # purple
    'historical':             ('Historical Pattern', 'FFF8E1'),   # amber
    'prepaid_amortization':   ('Prepaid Amort.',     'E0F2F1'),   # teal
    'prepaid_ledger':         ('Prepaid Release',    'E0F7FA'),   # cyan
    'management_fee':         ('Management Fee',     'C8E6C9'),   # green
    'management_fee_catchup': ('Mgmt Fee Catch-up',  'FFCCBC'),   # orange
    'contract_supplement':    ('One-Off Accrual',    'FFE0B2'),   # light orange
    'tenant_utility_billing': ('Tenant Utility',     'E1F5FE'),   # light cyan
    'bonus_accrual':          ('Bonus Accrual',      'FCE4EC'),   # pink
    'manual':                 ('Manual JE',          'F0F4C3'),   # yellow-green
}

_STATUS_FILL = {
    'PASS': _GRP_GREEN_LITE,
    'FLAG': _AMBER_LITE,
    'FAIL': _RED_LITE,
}


# ── Low-level styling helpers ─────────────────────────────────────────────────

def _font(bold=False, size=10, color=_BLACK, italic=False):
    return Font(name='Calibri', size=size, bold=bold, color=color, italic=italic)


def _fill(hex_color: str):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def _border(style='thin'):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _thin_bottom():
    return Border(bottom=Side(style='thin'))


def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _money(v: float) -> str:
    if v == 0:
        return '$-'
    return f'${abs(v):,.2f}' if v >= 0 else f'(${abs(v):,.2f})'


def _pct(v: float) -> str:
    return f'{v:.2%}'


def _write_header_row(ws, row: int, headers: List[str],
                       fill_hex=_GRP_GREEN, font_size=10):
    """Write a full-width bold header row with white text on dark green."""
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = _font(bold=True, size=font_size, color=_WHITE)
        c.fill      = _fill(fill_hex)
        c.alignment = _align('center', wrap=True)
        c.border    = _border()


def _write_section_header(ws, row: int, text: str, ncols: int):
    """Merged section label on mid-green."""
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=ncols
    )
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _font(bold=True, size=10, color=_WHITE)
    c.fill      = _fill(_GRP_GREEN_MID)
    c.alignment = _align('left')


def _write_kv(ws, row: int, key: str, value, key_bold=True, val_color=_BLACK,
              bg_hex=None):
    """Write a key-value pair spanning two cells."""
    ck = ws.cell(row=row, column=1, value=key)
    ck.font      = _font(bold=key_bold, size=10)
    ck.alignment = _align('left')
    if bg_hex:
        ck.fill = _fill(bg_hex)

    cv = ws.cell(row=row, column=2, value=value)
    cv.font      = _font(size=10, color=val_color)
    cv.alignment = _align('left')
    if bg_hex:
        cv.fill = _fill(bg_hex)
    return row + 1


def _freeze(ws, row=2, col=1):
    ws.freeze_panes = ws.cell(row=row, column=col)


# ── Tab 1: Summary ────────────────────────────────────────────────────────────

def _build_summary(
    ws,
    period: str,
    property_name: str,
    all_je_lines: List[dict],
    fee_result,
    qc_report,
    files_uploaded: Dict[str, str],
    run_ts: str,
    property_config=None,
):
    ws.title = '1 - Summary'
    ws.sheet_properties.tabColor = _GRP_GREEN

    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 18

    # ── Workbook title banner ────────────────────────────────────────────────
    ws.merge_cells('A1:C1')
    c = ws.cell(row=1, column=1,
                value=f'PIPELINE AUDIT TRAIL — {property_name} — {period}')
    c.font      = _font(bold=True, size=13, color=_WHITE)
    c.fill      = _fill(_GRP_GREEN)
    c.alignment = _align('center')
    ws.row_dimensions[1].height = 22

    row = 3
    # Run metadata
    _write_section_header(ws, row, 'RUN METADATA', 3)
    row += 1
    row = _write_kv(ws, row, 'Property',        property_name)
    row = _write_kv(ws, row, 'Close Period',     period)
    row = _write_kv(ws, row, 'Report Generated', run_ts)
    # C-9: derive management company from property config; fallback to generic label
    _mgmt_co = (getattr(property_config, 'management_company', '') or 'GRP')
    _prop_disp = (getattr(property_config, 'property_display_name', '') or property_name or 'GRP Properties')
    row = _write_kv(ws, row, 'Pipeline Version', get_version())
    row = _write_kv(ws, row, 'Management Entity', f'{_mgmt_co} / {_prop_disp}')

    row += 1
    # Files processed
    _write_section_header(ws, row, 'FILES PROCESSED', 3)
    row += 1
    _write_header_row(ws, row, ['File Type', 'Path / Filename', 'Status'],
                      fill_hex=_GRP_GREEN_MID)
    row += 1
    _FILE_LABELS = {
        'gl':                  'Yardi GL (pre-close)',
        'trial_balance':       'Trial Balance',
        'budget_comparison':   'Budget Comparison',
        'kardin_budget':       'Kardin Annual Budget',
        'nexus_accrual':       'Nexus AP Accrual Detail',
        'bank_rec':            'Yardi Bank Rec',
        'daca_bank':           'KeyBank DACA Statement',
        'receivable_detail':   'Yardi Receivable Detail',
        'ar_aging':            'Yardi AR Aging',
        'loan':                'Berkadia Loan Statement',
        'prepaid_ledger':      'Prior Month Prepaid Ledger',
        't12_statement':       '12-Month Income Statement',
    }
    alt = False
    for key, label in _FILE_LABELS.items():
        path = files_uploaded.get(key)
        has  = bool(path and os.path.exists(str(path)))
        bg   = _GRP_GREEN_LITE if alt else _WHITE
        alt  = not alt
        for col, val in enumerate([label,
                                    os.path.basename(str(path)) if has else '—',
                                    '✓ Uploaded' if has else '— Not uploaded'], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = _font(size=9,
                                color='2E7D32' if has else '9E9E9E')
            c.fill      = _fill(bg)
            c.alignment = _align('left')
        row += 1

    row += 1
    # JE summary
    _write_section_header(ws, row, 'PASS 1 — JOURNAL ENTRIES GENERATED', 3)
    row += 1
    _write_header_row(ws, row, ['Source Layer', 'JE Count', 'Total Debits ($)'],
                      fill_hex=_GRP_GREEN_MID)
    row += 1
    je_dr_lines = [l for l in all_je_lines if (l.get('debit') or 0) > 0]
    by_source: Dict[str, Dict] = {}
    for l in je_dr_lines:
        src = l.get('source', 'other')
        if src not in by_source:
            by_source[src] = {'count': 0, 'total': 0.0}
        by_source[src]['count'] += 1
        by_source[src]['total'] += float(l.get('debit') or 0)

    grand_count = 0
    grand_total = 0.0
    alt = False
    for src, meta in _SOURCE_META.items():
        if src not in by_source:
            continue
        label, fill_h = meta
        info = by_source[src]
        bg = fill_h if alt else _WHITE
        alt = not alt
        for col, val in enumerate([label, info['count'], _money(info['total'])], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = _font(size=9)
            c.fill      = _fill(bg)
            c.alignment = _align('right' if col > 1 else 'left')
        grand_count += info['count']
        grand_total += info['total']
        row += 1

    # Grand total row
    for col, val in enumerate(['TOTAL', grand_count, _money(grand_total)], 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = _font(bold=True, size=9)
        c.fill      = _fill(_GREY_MED)
        c.alignment = _align('right' if col > 1 else 'left')
        c.border    = _thin_bottom()
    row += 2

    # Management fee summary
    if fee_result and getattr(fee_result, 'cash_received', 0) > 0:
        _write_section_header(ws, row, 'MANAGEMENT FEE', 3)
        row += 1
        row = _write_kv(ws, row, 'Cash Received Basis',
                        _money(fee_result.cash_received))
        _total_rate = getattr(fee_result, 'total_rate', None)
        _rate_str   = f' ({_total_rate:.2%})' if _total_rate else ''
        row = _write_kv(ws, row, f'Total Fee{_rate_str}',
                        _money(getattr(fee_result, 'total_fee', 0)))
        # C-9: Iterate config-driven fee lines when available; fall back to JLL/GRP labels
        _fee_lines_cfg = getattr(property_config, 'management_fees', None) or []
        if _fee_lines_cfg:
            for _fl in _fee_lines_cfg:
                _fl_label = f'{_fl.name} Fee ({_fl.rate:.2%})'
                _fl_amt   = _money(getattr(fee_result, 'cash_received', 0) * _fl.rate)
                row = _write_kv(ws, row, _fl_label, _fl_amt)
        else:
            # Legacy RevLabs JLL + GRP fallback
            row = _write_kv(ws, row, 'Less JLL Portion (1.25%)',
                            _money(getattr(fee_result, 'jll_fee', 0)))
            row = _write_kv(ws, row, 'GRP Net Fee (1.75%)',
                            _money(getattr(fee_result, 'grp_fee', 0)))
        row += 1

    # QC summary
    if qc_report:
        _write_section_header(ws, row, 'QC CHECK SUMMARY', 3)
        row += 1
        _write_header_row(ws, row, ['Check', 'Status', 'Summary'],
                          fill_hex=_GRP_GREEN_MID)
        row += 1
        for chk in qc_report.checks:
            bg = _STATUS_FILL.get(chk.status, _WHITE)
            for col, val in enumerate(
                [f'{chk.check_id}: {chk.check_name}', chk.status, chk.summary], 1
            ):
                c = ws.cell(row=row, column=col, value=val)
                c.font      = _font(size=9,
                                    bold=(col == 2),
                                    color=('2E7D32' if chk.status == 'PASS'
                                           else 'B71C1C' if chk.status == 'FAIL'
                                           else 'E65100'))
                c.fill      = _fill(bg)
                c.alignment = _align('left', wrap=(col == 3))
            ws.row_dimensions[row].height = 28
            row += 1


# ── Tab 2: JE Log ─────────────────────────────────────────────────────────────

def _build_je_log(ws, all_je_lines: List[dict]):
    ws.title = '2 - JE Log'
    ws.sheet_properties.tabColor = '1565C0'

    # Column widths
    cols_w = {
        'A': 14, 'B': 20, 'C': 13, 'D': 28, 'E': 13, 'F': 28,
        'G': 44, 'H': 28, 'I': 15, 'J': 10, 'K': 12,
    }
    for col, w in cols_w.items():
        ws.column_dimensions[col].width = w

    # Banner
    ws.merge_cells('A1:K1')
    c = ws.cell(row=1, column=1, value='PASS 1 — JOURNAL ENTRY LOG')
    c.font      = _font(bold=True, size=12, color=_WHITE)
    c.fill      = _fill('1565C0')
    c.alignment = _align('center')
    ws.row_dimensions[1].height = 20

    headers = [
        'JE #', 'Source Layer',
        'DR Account', 'DR Account Name',
        'CR Account', 'CR Account Name',
        'Description', 'Vendor / Invoice',
        'Amount ($)', 'Confidence', 'Auto-Rev.',
    ]
    _write_header_row(ws, 2, headers, fill_hex='1565C0', font_size=9)
    _freeze(ws, row=3)

    # Build JE pair dict: je_number → {dr line, cr line}
    je_pairs: Dict[str, Dict] = {}
    for l in all_je_lines:
        jn = l.get('je_number', '')
        if jn not in je_pairs:
            je_pairs[jn] = {'dr': None, 'cr': None, '_src': l.get('source', ''),
                             '_rev': l.get('auto_reverse', True)}
        if (l.get('debit') or 0) > 0:
            je_pairs[jn]['dr'] = l
        elif (l.get('credit') or 0) > 0:
            je_pairs[jn]['cr'] = l

    row = 3
    # Sort: source order then JE number
    _SRC_ORDER = list(_SOURCE_META.keys())
    def _sort_key(item):
        jn, pair = item
        src = pair['_src']
        try:
            si = _SRC_ORDER.index(src)
        except ValueError:
            si = 99
        return (si, jn)

    for jn, pair in sorted(je_pairs.items(), key=_sort_key):
        dr   = pair.get('dr') or {}
        cr   = pair.get('cr') or {}
        src  = pair['_src']
        _, fill_h = _SOURCE_META.get(src, ('Other', _GREY_LITE))

        amount   = float(dr.get('debit') or cr.get('credit') or 0)
        vendor   = str(dr.get('vendor') or '')
        inv_num  = str(dr.get('invoice_number') or '')
        vendor_s = f'{vendor}  {inv_num}'.strip('  ').strip()

        desc     = str(dr.get('description') or cr.get('description') or '')
        conf     = str(dr.get('confidence') or '')
        src_lbl  = _SOURCE_META.get(src, (src, ''))[0]
        auto_rev = '✓' if pair['_rev'] else '—'

        vals = [
            jn,
            src_lbl,
            dr.get('account_code', ''),
            dr.get('account_name', ''),
            cr.get('account_code', ''),
            cr.get('account_name', ''),
            desc,
            vendor_s,
            amount,
            conf,
            auto_rev,
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = _fill(fill_h)
            c.alignment = _align('right' if col == 9 else 'left', wrap=(col == 7))
            c.font      = _font(size=9)
            if col == 9 and isinstance(val, (int, float)):
                c.number_format = '#,##0.00'
        row += 1

    # Legend row
    row += 1
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=11
    )
    c = ws.cell(row=row, column=1, value='Color legend:  ' + '   '.join(
        f'■ {lbl}' for lbl, _ in _SOURCE_META.values()
    ))
    c.font      = _font(size=8, italic=True, color='616161')
    c.alignment = _align('left')


# ── Tab 3: Management Fee ─────────────────────────────────────────────────────

def _build_mgmt_fee(ws, fee_result, period: str, property_name: str, property_config=None):
    ws.title = '3 - Management Fee'
    ws.sheet_properties.tabColor = '2E7D32'

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22

    ws.merge_cells('A1:C1')
    c = ws.cell(row=1, column=1,
                value=f'MANAGEMENT FEE CALCULATION — {period}')
    c.font      = _font(bold=True, size=12, color=_WHITE)
    c.fill      = _fill(_GRP_GREEN)
    c.alignment = _align('center')
    ws.row_dimensions[1].height = 20

    if not fee_result or not getattr(fee_result, 'cash_received', 0):
        ws.cell(row=3, column=1,
                value='No management fee calculated this period.').font = _font(italic=True)
        return

    row = 3
    _write_section_header(ws, row, 'CASH RECEIVED BASIS', 3)
    row += 1
    row = _write_kv(ws, row, 'Cash Received',
                    _money(fee_result.cash_received))
    row = _write_kv(ws, row, 'Source',
                    getattr(fee_result, 'cash_source', '—'))
    prepay = getattr(fee_result, 'prepayment_excluded', 0) or 0
    if prepay:
        row = _write_kv(ws, row, 'Less Prepayments Excluded',
                        f'({_money(prepay)})')
    row += 1

    _write_section_header(ws, row, 'FEE CALCULATION', 3)
    row += 1
    _write_header_row(ws, row, ['Component', 'Rate', 'Amount'],
                      fill_hex=_GRP_GREEN_MID, font_size=9)
    row += 1

    cash = fee_result.cash_received
    total_fee = getattr(fee_result, 'total_fee', round(cash * 0.03, 2))
    jll_fee   = getattr(fee_result, 'jll_fee',   round(cash * 0.0125, 2))
    grp_fee   = getattr(fee_result, 'grp_fee',   round(cash * 0.0175, 2))

    _rows = [
        ('Total Management Fee',   '3.00%',  total_fee, False),
        ('Less JLL Portion',       '1.25%',  -jll_fee,  False),
        ('GRP Net Fee (Balance Due)', '1.75%', grp_fee, True),
    ]
    for lbl, rate, amt, bold in _rows:
        bg = _GRP_GREEN_LITE if bold else _WHITE
        for col, val in enumerate([lbl, rate, _money(amt)], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = _font(bold=bold, size=10)
            c.fill      = _fill(bg)
            c.alignment = _align('right' if col > 1 else 'left')
            if bold:
                c.border = _thin_bottom()
        row += 1

    row += 1
    _write_section_header(ws, row, 'JOURNAL ENTRY', 3)
    row += 1
    row = _write_kv(ws, row, 'Debit',  '637130  Admin-Management Fees')
    row = _write_kv(ws, row, 'Credit', '213100  Accrued Expenses')
    row = _write_kv(ws, row, 'Amount', _money(total_fee))
    row = _write_kv(ws, row, 'JE Reference', 'MGT-001')

    # C-9: resolve per-property strings from config instead of hardcoding RevLabs values.
    _inv_prefix   = (getattr(property_config, 'invoice_prefix', '') or 'RevLabsPM') if property_config else 'RevLabsPM'
    _bill_to      = (getattr(property_config, 'property_name', '') or 'Revolution Labs Owner, LLC') if property_config else 'Revolution Labs Owner, LLC'
    _payable_to   = (getattr(property_config, 'management_company', '') or 'Greatland Realty Partners') if property_config else 'Greatland Realty Partners'
    # Append 'LLC' only when the company name doesn't already contain a suffix.
    if _payable_to and not any(s in _payable_to for s in ('LLC', 'LP', 'Inc', 'Corp')):
        _payable_to = _payable_to + ' LLC'

    inv_num = None
    try:
        import re, calendar
        m = re.search(r'([A-Za-z]{3})[- ](\d{4})', period)
        if m:
            mos = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
                   'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
            mo = mos.get(m.group(1).lower(), 0)
            yr = int(m.group(2))
            if mo:
                inv_num = f'{_inv_prefix}{mo:02d}{yr}'
    except Exception:
        pass
    if inv_num:
        row += 1
        _write_section_header(ws, row, 'INVOICE', 3)
        row += 1
        row = _write_kv(ws, row, 'Invoice Number', inv_num)
        row = _write_kv(ws, row, 'Bill To', _bill_to)
        row = _write_kv(ws, row, 'Payable To', _payable_to)
        row = _write_kv(ws, row, 'GRP Net Due', _money(grp_fee))


# ── Tab 4: Accrual vs Actual ──────────────────────────────────────────────────

def _build_accrual_check(ws, prior_accrual_check: List[dict]):
    ws.title = '4 - Accrual Check'
    ws.sheet_properties.tabColor = 'E65100'

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 14

    ws.merge_cells('A1:G1')
    c = ws.cell(row=1, column=1,
                value='PRIOR MONTH ACCRUAL vs ACTUALS RECEIVED')
    c.font      = _font(bold=True, size=12, color=_WHITE)
    c.fill      = _fill('E65100')
    c.alignment = _align('center')
    ws.row_dimensions[1].height = 20

    row = 3
    ws.merge_cells(f'A{row}:G{row}')
    note = ws.cell(row=row, column=1,
        value=(
            'J-type auto-reversals of last month\'s pipeline accruals compared to actual '
            'invoices received this period.  NOT YET BILLED = invoice not yet received; '
            'pipeline has re-accrued.  MATCHED = actual within 5% of accrual.'
        ))
    note.font      = _font(size=9, italic=True, color='616161')
    note.alignment = _align('left', wrap=True)
    ws.row_dimensions[row].height = 28
    row += 2

    if not prior_accrual_check:
        ws.cell(row=row, column=1,
                value='No prior-month accrual auto-reversals detected in GL.').font = (
                    _font(italic=True, color='9E9E9E'))
        return

    headers = [
        'Account', 'Account Name',
        'Prior Accrual', 'Actual Billed', 'Variance',
        'Status', 'JE Reference',
    ]
    _write_header_row(ws, row, headers, fill_hex='E65100', font_size=9)
    _freeze(ws, row=row + 1)
    row += 1

    _STATUS_BG = {
        'MATCHED':        _GRP_GREEN_LITE,
        'NOT YET BILLED': _AMBER_LITE,
        'PARTIAL':        _AMBER_LITE,
        'OVER INVOICED':  _RED_LITE,
    }
    _STATUS_ICON = {
        'MATCHED':        '✅ MATCHED',
        'NOT YET BILLED': '🔄 NOT YET BILLED',
        'PARTIAL':        '⚠️ PARTIAL',
        'OVER INVOICED':  '⚠️ OVER INVOICED',
    }

    for r in prior_accrual_check:
        bg  = _STATUS_BG.get(r['status'], _WHITE)
        var = r['variance']
        var_s = (f'+${var:,.2f}' if var >= 0 else f'-${abs(var):,.2f}')
        vals = [
            r['account_code'],
            r['account_name'],
            r['reversal_amount'],
            r['actual_amount'],
            var_s,
            _STATUS_ICON.get(r['status'], r['status']),
            r['je_refs'],
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = _fill(bg)
            c.alignment = _align('right' if col in (3, 4) else 'left')
            c.font      = _font(size=9)
            if col in (3, 4) and isinstance(val, (int, float)):
                c.number_format = '#,##0.00'
        row += 1

    # Summary counts below
    row += 1
    counts = {}
    for r in prior_accrual_check:
        counts[r['status']] = counts.get(r['status'], 0) + 1
    for status, cnt in sorted(counts.items()):
        c1 = ws.cell(row=row, column=1, value=f'{cnt}x  {status}')
        c1.font      = _font(size=9, bold=True)
        c1.fill      = _fill(_STATUS_BG.get(status, _WHITE))
        c1.alignment = _align('left')
        row += 1


# ── Tab 5: QC Checks ──────────────────────────────────────────────────────────

def _build_qc_checks(ws, qc_report):
    ws.title = '5 - QC Checks'
    ws.sheet_properties.tabColor = '6A1B9A'

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 16

    ws.merge_cells('A1:F1')
    c = ws.cell(row=1, column=1, value='QC CHECK RESULTS')
    c.font      = _font(bold=True, size=12, color=_WHITE)
    c.fill      = _fill('6A1B9A')
    c.alignment = _align('center')
    ws.row_dimensions[1].height = 20

    if not qc_report:
        ws.cell(row=3, column=1,
                value='QC checks not run this period (Pass 2 not executed).').font = (
                    _font(italic=True, color='9E9E9E'))
        return

    row = 3
    _write_header_row(ws, row,
        ['Check ID', 'Check Name', 'Status', 'Summary', 'Account / Item', 'Amount'],
        fill_hex='6A1B9A', font_size=9)
    _freeze(ws, row=row + 1)
    row += 1

    for chk in qc_report.checks:
        bg       = _STATUS_FILL.get(chk.status, _WHITE)
        txt_col  = ('2E7D32' if chk.status == 'PASS'
                    else 'B71C1C' if chk.status == 'FAIL' else 'E65100')

        if not chk.findings:
            # Single summary row
            for col, val in enumerate(
                [chk.check_id, chk.check_name, chk.status, chk.summary, '', ''], 1
            ):
                c = ws.cell(row=row, column=col, value=val)
                c.fill      = _fill(bg)
                c.font      = _font(size=9, bold=(col == 3), color=txt_col if col == 3 else _BLACK)
                c.alignment = _align('left', wrap=(col == 4))
            ws.row_dimensions[row].height = 30
            row += 1
        else:
            # First row: check header
            for col, val in enumerate(
                [chk.check_id, chk.check_name, chk.status, chk.summary, '', ''], 1
            ):
                c = ws.cell(row=row, column=col, value=val)
                c.fill      = _fill(bg)
                c.font      = _font(size=9, bold=True, color=txt_col if col == 3 else _BLACK)
                c.alignment = _align('left', wrap=(col == 4))
            ws.row_dimensions[row].height = 30
            row += 1

            # Findings detail rows
            for f in chk.findings[:25]:   # cap at 25 per check to keep readable
                diff = getattr(f, 'difference', None)
                diff_s = _money(diff) if diff is not None else ''
                acct_s = (getattr(f, 'account_name', '') or
                           getattr(f, 'account_code', '') or '')
                note_s = getattr(f, 'note', '') or ''
                for col, val in enumerate(
                    ['', '', '', note_s, acct_s, diff_s], 1
                ):
                    c = ws.cell(row=row, column=col, value=val)
                    c.fill      = _fill(_GREY_LITE)
                    c.font      = _font(size=8, italic=True)
                    c.alignment = _align('left', wrap=(col == 4))
                row += 1

        # Separator
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = _thin_bottom()
        row += 1


# ── Tab 6: Yardi ETL CSV (exact import rows) ─────────────────────────────────

def _build_yardi_csv_tab(ws, all_je_lines: List[dict], period: str, property_code: str):
    """
    Reconstruct the exact rows that were written to GA_Accruals_JE.csv and
    display them as a readable Excel grid.

    Shows only the 12 populated ETL columns (TRANNUM, DATE, PROPERTY, ACCOUNT,
    POSTMONTH, BOOKNUM, AMOUNT, REMARK, REF, DESC, DISPLAYTYPE, ReverseNextMonth)
    plus three [AUDITOR REF] columns not sent to Yardi: JE #, Source Layer, Account Name.

    The 53 blank ETL columns are omitted for readability — the header row notes this.
    """
    from calendar import monthrange as _monthrange

    ws.title = '6 - Yardi ETL CSV'
    ws.sheet_properties.tabColor = '37474F'   # dark slate

    _COL_W = {
        'A': 10, 'B': 13, 'C': 12, 'D': 10, 'E': 13,
        'F': 9,  'G': 14, 'H': 44, 'I': 20, 'J': 44,
        'K': 30, 'L': 7,  'M': 14, 'N': 20, 'O': 30,
    }
    for col, w in _COL_W.items():
        ws.column_dimensions[col].width = w

    # ── Banner ───────────────────────────────────────────────────────────────
    ws.merge_cells('A1:O1')
    c = ws.cell(row=1, column=1,
                value='YARDI ETL IMPORT — GA_Accruals_JE.csv  (exact rows submitted to Yardi)')
    c.font      = _font(bold=True, size=12, color=_WHITE)
    c.fill      = _fill('37474F')
    c.alignment = _align('center')
    ws.row_dimensions[1].height = 20

    # ── Row 2: note about blank columns ──────────────────────────────────────
    ws.merge_cells('A2:O2')
    note = ws.cell(row=2, column=1,
                   value=(
                       'ETL format is 65 columns; 53 are always blank and omitted here. '
                       'Columns A–L = exact Yardi ETL fields. '
                       'Columns M–O = [AUDITOR REF] cross-reference only — not sent to Yardi. '
                       'AMOUNT: positive = Debit, negative = Credit.'
                   ))
    note.font      = _font(size=9, italic=True, color='616161')
    note.alignment = _align('left', wrap=True)
    ws.row_dimensions[2].height = 24

    # ── Row 3: Simulated ETL Row 1 "FinJournals" identifier ──────────────────
    ws.merge_cells('A3:O3')
    fi = ws.cell(row=3, column=1,
                 value='← ETL Row 1 FinJournals record-type identifier (col A = "FinJournals", cols B–BM blank)')
    fi.font      = _font(size=9, italic=True, color='9E9E9E')
    fi.fill      = _fill(_GREY_LITE)
    fi.alignment = _align('left')

    # ── Row 4: Column headers ─────────────────────────────────────────────────
    _ETL_COL_HEADERS = [
        'TRANNUM',    'DATE',     'PROPERTY', 'ACCOUNT',
        'POSTMONTH',  'BOOKNUM',  'AMOUNT',   'REMARK',
        'REF',        'DESC',     'DISPLAYTYPE', 'ReverseNextMonth',
    ]
    _AUD_HEADERS = ['[AUD] JE #', '[AUD] Source Layer', '[AUD] Account Name']
    all_headers  = _ETL_COL_HEADERS + _AUD_HEADERS

    _ETL_FILL    = '37474F'   # dark slate — ETL columns
    _AUD_FILL    = '4527A0'   # deep purple — auditor columns

    for col, h in enumerate(_ETL_COL_HEADERS, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font      = _font(bold=True, size=9, color=_WHITE)
        c.fill      = _fill(_ETL_FILL)
        c.alignment = _align('center')
        c.border    = _border()
    for col, h in enumerate(_AUD_HEADERS, len(_ETL_COL_HEADERS) + 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font      = _font(bold=True, size=9, color=_WHITE)
        c.fill      = _fill(_AUD_FILL)
        c.alignment = _align('center')
        c.border    = _border()

    _freeze(ws, row=5)

    # ── Derive period end date (mirrors generate_etl_csv logic) ──────────────
    period_date = ''
    for _pfmt in ('%b-%Y', '%B-%Y', '%b %Y', '%B %Y', '%m-%Y', '%m/%Y'):
        try:
            from datetime import datetime as _dt, date as _date
            _parsed = _dt.strptime(period.strip(), _pfmt)
            _last   = _monthrange(_parsed.year, _parsed.month)[1]
            period_date = _date(_parsed.year, _parsed.month, _last).strftime('%m/%d/%Y')
            break
        except Exception:
            pass

    # ── Build batch map (same deterministic logic as generate_etl_csv) ────────
    batch_map: dict = {}
    batch_counter = 1
    for line in all_je_lines:
        jn = line.get('je_number', '')
        if jn not in batch_map:
            batch_map[jn] = batch_counter
            batch_counter += 1

    # ── Pre-scan: which batches auto-reverse (touch 213100)? ─────────────────
    _batches_213100 = {
        line.get('je_number', '')
        for line in all_je_lines
        if str(line.get('account_code', '') or '').strip() == '213100'
    }

    # ── Data rows ────────────────────────────────────────────────────────────
    row = 5
    alt = False
    for line in all_je_lines:
        je_num  = line.get('je_number', '')
        batch   = batch_map.get(je_num, 1)
        desc    = str(line.get('description', '') or '')[:60]
        gl_acct = str(line.get('account_code', '') or '')
        ref     = str(line.get('reference', '') or je_num)
        debit   = line.get('debit', 0) or 0
        credit  = line.get('credit', 0) or 0
        amount  = round(debit - credit, 2)
        bm_batch = -1 if je_num in _batches_213100 else 0
        bm       = line.get('reverse_next_month', bm_batch)

        # Source-based row shading (reuse _SOURCE_META palette for the auditor cols)
        src = line.get('source', '')
        _, src_fill_h = _SOURCE_META.get(src, ('Other', _GREY_LITE))
        etl_bg = _GREY_LITE if alt else _WHITE
        alt    = not alt

        etl_vals = [
            batch,                              # TRANNUM
            period_date,                        # DATE
            property_code,                      # PROPERTY
            gl_acct,                            # ACCOUNT
            period_date,                        # POSTMONTH
            1,                                  # BOOKNUM
            amount,                             # AMOUNT
            desc,                               # REMARK
            ref,                                # REF
            desc,                               # DESC
            'Standard Journal Display Type',    # DISPLAYTYPE
            bm,                                 # ReverseNextMonth
        ]
        aud_vals = [
            je_num,                                          # [AUD] JE #
            _SOURCE_META.get(src, (src, ''))[0],             # [AUD] Source Layer
            str(line.get('account_name', '') or ''),         # [AUD] Account Name
        ]

        for col, val in enumerate(etl_vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = _fill(etl_bg)
            c.font      = _font(size=9)
            c.alignment = _align('right' if col in (1, 6, 7, 12) else 'left')
            c.border    = _border()
            if col == 7 and isinstance(val, (int, float)):   # AMOUNT
                c.number_format = '#,##0.00;(#,##0.00)'

        for col, val in enumerate(aud_vals, len(etl_vals) + 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = _fill(src_fill_h)   # source-layer color for auditor cols
            c.font      = _font(size=9)
            c.alignment = _align('left')
            c.border    = _border()

        row += 1

    # ── Totals row ────────────────────────────────────────────────────────────
    total_dr = sum(line.get('debit', 0) or 0 for line in all_je_lines)
    total_cr = sum(line.get('credit', 0) or 0 for line in all_je_lines)
    for col in range(1, len(all_headers) + 1):
        c = ws.cell(row=row, column=col)
        c.fill   = _fill(_GREY_MED)
        c.border = _border()
        if col == 1:
            c.value     = 'TOTALS'
            c.font      = _font(bold=True, size=9)
            c.alignment = _align('right')
        elif col == 7:   # AMOUNT col — show net (should be 0 for balanced JEs)
            net = round(total_dr - total_cr, 2)
            c.value         = net
            c.font          = _font(bold=True, size=9,
                                    color='B71C1C' if net != 0 else '2E7D32')
            c.number_format = '#,##0.00;(#,##0.00)'
            c.alignment     = _align('right')
        elif col == 13:  # [AUD] JE # col — show row count
            c.value     = f'{len(all_je_lines)} rows'
            c.font      = _font(bold=True, size=9, color='616161')
            c.alignment = _align('left')
        elif col == 14:  # [AUD] Source col — show DR / CR totals
            c.value     = f'DR ${total_dr:,.2f}  |  CR ${total_cr:,.2f}'
            c.font      = _font(bold=True, size=9, color='616161')
            c.alignment = _align('left')

    row += 2
    # ── Legend ────────────────────────────────────────────────────────────────
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=len(all_headers)
    )
    leg = ws.cell(row=row, column=1,
                  value='[AUD] column shading matches source layer from Tab 2 - JE Log: '
                        + '   '.join(f'■ {lbl}' for lbl, _ in _SOURCE_META.values()))
    leg.font      = _font(size=8, italic=True, color='616161')
    leg.alignment = _align('left')


def _period_bounds(period: str):
    """Return (start_date, end_date) for a period string like 'Mar-2026', or (None, None)."""
    from calendar import monthrange as _monthrange
    from datetime import date as _date
    for _pfmt in ('%b-%Y', '%B-%Y', '%b %Y', '%B %Y', '%m-%Y', '%m/%Y'):
        try:
            _parsed = datetime.strptime(period.strip(), _pfmt)
            _last = _monthrange(_parsed.year, _parsed.month)[1]
            return _date(_parsed.year, _parsed.month, 1), _date(_parsed.year, _parsed.month, _last)
        except Exception:
            continue
    return None, None


# ── Tab 7: Methodology Reference ──────────────────────────────────────────────

def _build_methodology(ws, property_config=None):
    """
    Plain-language reference for the standing accounting methodology behind
    the pipeline's estimates and judgment areas — materiality floors,
    amortization conventions, bonus derivation — so an auditor can confirm
    consistency without reading the codebase.
    """
    ws.title = '7 - Methodology'
    ws.sheet_properties.tabColor = '5D4037'

    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 70

    ws.merge_cells('A1:B1')
    c = ws.cell(row=1, column=1, value='ACCOUNTING METHODOLOGY REFERENCE')
    c.font      = _font(bold=True, size=12, color=_WHITE)
    c.fill      = _fill('5D4037')
    c.alignment = _align('center')
    ws.row_dimensions[1].height = 20

    row = 3
    note = ws.cell(row=row, column=1,
        value=('Standing rules applied consistently every period — not re-derived or '
               'adjusted case-by-case. Per-property overrides are noted where applicable.'))
    ws.merge_cells(f'A{row}:B{row}')
    note.font = _font(size=9, italic=True, color='616161')
    note.alignment = _align('left', wrap=True)
    ws.row_dimensions[row].height = 24
    row += 2

    _qc_thr = getattr(property_config, 'qc_thresholds', None) or {}
    _t1_abs = _qc_thr.get('tier1_abs', 5000.0)
    _t1_pct = _qc_thr.get('tier1_pct', 0.05)
    _t2_min = _qc_thr.get('tier2_min', 2500.0)
    _mom_sw = _qc_thr.get('mom_swing', 10000.0)

    _write_section_header(ws, row, 'ACCRUAL DETECTION (4 LAYERS)', 2)
    row += 1
    row = _write_kv(ws, row, 'Layer 1 — Nexus Open Invoices',
                    'Any open invoice in the Nexus AP export not yet posted to the GL is accrued in full.')
    row = _write_kv(ws, row, 'Layer 2 — Invoice Proration',
                    'Utilities: daily rate x uncovered days. All other recurring services: full prior invoice amount.')
    row = _write_kv(ws, row, 'Layer 3 — Historical Recurring',
                    'Budget Comparison YTD actual / months elapsed. January uses annual budget / 12 as a fallback. '
                    f'Materiality floor: $5,000 — accounts below this are not auto-accrued.')
    row = _write_kv(ws, row, 'Layer 4 — Payroll Bonus',
                    'User-entered annual amount / 12, or Kardin-derived if not entered. Suppressed in the month the '
                    'bonus is actually paid (GL activity already reflects it).')
    row += 1

    _write_section_header(ws, row, 'MATERIALITY / VARIANCE THRESHOLDS', 2)
    row += 1
    row = _write_kv(ws, row, 'Tier 1 Variance Flag', f'>= ${_t1_abs:,.0f} OR >= {_t1_pct:.0%} of budget')
    row = _write_kv(ws, row, 'Tier 2 Variance Flag', f'${_t2_min:,.0f} - ${_t1_abs:,.0f} and >= 5% of budget')
    row = _write_kv(ws, row, 'Month-over-Month Swing Flag', f'>= ${_mom_sw:,.0f}, either direction')
    row += 1

    _write_section_header(ws, row, 'PREPAID / AMORTIZATION CONVENTIONS', 2)
    row += 1
    row = _write_kv(ws, row, 'Endpoint Proration',
                    'First and last month of service prorated by actual days active in that month. '
                    'All middle months amortize the full monthly amount.')
    row = _write_kv(ws, row, 'RE Tax Quarterly Cycle',
                    'Payment month: defer 2/3 of the quarterly bill to prepaid. Each of the next 2 months '
                    'releases 1/3. Applied identically every quarter, no case-by-case adjustment.')
    row = _write_kv(ws, row, 'Insurance Amortization',
                    'Config-driven (named policies) where available; falls back to Kardin-derived or '
                    'Budget-Comparison-derived monthly amounts if no policy config exists for this property.')
    row += 1

    _write_section_header(ws, row, 'MANAGEMENT FEE', 2)
    row += 1
    _fee_lines_cfg = getattr(property_config, 'management_fees', None) or []
    if _fee_lines_cfg:
        for _fl in _fee_lines_cfg:
            row = _write_kv(ws, row, f'{_fl.name} Rate', f'{_fl.rate:.2%} of cash received')
    else:
        row = _write_kv(ws, row, 'Fee Rate', '3.00% of cash received (1.25% + 1.75% split)')
    row = _write_kv(ws, row, 'Cash Received Basis Priority',
                    'Receivable Summary > Receivable Detail + AR Aging > Receivable Detail only > '
                    'DACA additions > GL 111100 debits > Revenue proxy (first available wins)')


# ── Tab 8: Cutoff Review ──────────────────────────────────────────────────────

def _build_cutoff_review(ws, all_je_lines: List[dict], period: str):
    """
    Lists any JE line whose own date falls outside the close period's date
    range — a manual cutoff review aid, not an automated pass/fail. Accrual
    JEs are conventionally dated at period end; a line dated outside the
    period is worth a second look, not necessarily an error (e.g. a Nexus
    invoice's original date may legitimately predate the period).
    """
    ws.title = '8 - Cutoff Review'
    ws.sheet_properties.tabColor = 'AD1457'

    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 34
    ws.column_dimensions['D'].width = 44
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16

    ws.merge_cells('A1:F1')
    c = ws.cell(row=1, column=1, value='CUTOFF REVIEW — JE LINES DATED OUTSIDE THE CLOSE PERIOD')
    c.font      = _font(bold=True, size=12, color=_WHITE)
    c.fill      = _fill('AD1457')
    c.alignment = _align('center')
    ws.row_dimensions[1].height = 20

    row = 3
    note = ws.cell(row=row, column=1,
        value=('Every JE line with a date field is checked against this period\'s date range. '
               'A line outside the range is not automatically wrong — review each one to confirm '
               'the expense/revenue is assigned to the correct period.'))
    ws.merge_cells(f'A{row}:F{row}')
    note.font = _font(size=9, italic=True, color='616161')
    note.alignment = _align('left', wrap=True)
    ws.row_dimensions[row].height = 28
    row += 2

    period_start, period_end = _period_bounds(period)
    if period_start is None:
        ws.cell(row=row, column=1,
                value='Could not parse close period — cutoff check skipped.').font = _font(italic=True, color='9E9E9E')
        return

    from datetime import datetime as _dt, date as _date

    def _coerce(d):
        if d is None:
            return None
        if isinstance(d, _dt):
            return d.date()
        if isinstance(d, _date):
            return d
        if isinstance(d, str) and d.strip():
            for fmt in ('%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d'):
                try:
                    return _dt.strptime(d.strip(), fmt).date()
                except ValueError:
                    continue
        return None

    flagged = []
    for l in all_je_lines:
        d = _coerce(l.get('date'))
        if d is not None and not (period_start <= d <= period_end):
            flagged.append((l, d))

    if not flagged:
        ws.cell(row=row, column=1,
                value=f'All dated JE lines fall within {period_start:%m/%d/%Y} – {period_end:%m/%d/%Y}. '
                      'No cutoff review items.').font = _font(color='2E7D32', bold=True)
        return

    headers = ['JE #', 'Date', 'Account', 'Description', 'Amount', 'Source']
    _write_header_row(ws, row, headers, fill_hex='AD1457', font_size=9)
    row += 1
    for l, d in sorted(flagged, key=lambda x: x[1]):
        src_lbl = _SOURCE_META.get(l.get('source', ''), (l.get('source', ''), ''))[0]
        amt = float(l.get('debit') or l.get('credit') or 0)
        vals = [
            l.get('je_number', ''), d.strftime('%m/%d/%Y'),
            f"{l.get('account_code', '')}  {l.get('account_name', '')}".strip(),
            str(l.get('description', '') or ''), amt, src_lbl,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = _fill(_AMBER_LITE)
            cell.font = _font(size=9)
            cell.alignment = _align('right' if col == 5 else 'left')
            if col == 5:
                cell.number_format = '#,##0.00'
        row += 1


# ── Tab 9: Reconciling Item Aging ──────────────────────────────────────────────

def _build_reconciling_aging(ws, bank_recon_detail, period: str):
    """
    Ages each outstanding bank-rec item as of period end (days since the
    check/item date). Flags anything outstanding 60+ days — the classic
    auditor question of whether reconciling items are aging or clearing.
    """
    ws.title = '9 - Reconciling Aging'
    ws.sheet_properties.tabColor = '00838F'

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 34
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12

    ws.merge_cells('A1:F1')
    c = ws.cell(row=1, column=1, value='BANK RECONCILIATION — OUTSTANDING ITEM AGING')
    c.font      = _font(bold=True, size=12, color=_WHITE)
    c.fill      = _fill('00838F')
    c.alignment = _align('center')
    ws.row_dimensions[1].height = 20

    row = 3
    if not bank_recon_detail or not getattr(bank_recon_detail, 'outstanding_checks', None):
        ws.cell(row=row, column=1,
                value='No outstanding reconciling items this period.').font = _font(italic=True, color='9E9E9E')
        return

    _, period_end = _period_bounds(period)

    headers = ['Date', 'Reference', 'Description', 'Amount', 'Days Outstanding', 'Status']
    _write_header_row(ws, row, headers, fill_hex='00838F', font_size=9)
    row += 1

    items = list(bank_recon_detail.outstanding_checks)
    items_with_age = []
    for it in items:
        d = getattr(it, 'date', None)
        age = (period_end - d).days if (d is not None and period_end is not None) else None
        items_with_age.append((it, d, age))
    items_with_age.sort(key=lambda x: (x[2] is None, -(x[2] or 0)))

    for it, d, age in items_with_age:
        aged = age is not None and age >= 60
        bg = _AMBER_LITE if aged else _WHITE
        status = f'⚠️ AGED ({age}d)' if aged else (f'{age}d' if age is not None else 'no date')
        vals = [
            d.strftime('%m/%d/%Y') if d else '—',
            str(getattr(it, 'reference', '') or getattr(it, 'control', '') or ''),
            str(getattr(it, 'description', '') or ''),
            float(getattr(it, 'credit', 0) or 0),
            age if age is not None else '—',
            status,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = _fill(bg)
            cell.font = _font(size=9, bold=(col == 6 and aged), color='E65100' if (col == 6 and aged) else _BLACK)
            cell.alignment = _align('right' if col in (4, 5) else 'left')
            if col == 4:
                cell.number_format = '#,##0.00'
        row += 1

    row += 1
    aged_count = sum(1 for _, _, age in items_with_age if age is not None and age >= 60)
    if aged_count:
        c2 = ws.cell(row=row, column=1,
                     value=f'{aged_count} item(s) outstanding 60+ days — review for clearing or write-off.')
        c2.font = _font(bold=True, size=9, color='E65100')


# ── Tab 10: Close Tracker & Sign-off Summary ──────────────────────────────────

def _build_close_signoff(ws, close_tracker: Optional[Dict[int, dict]],
                          signoff_state: Optional[Dict[int, dict]],
                          signoff_items: Optional[List[str]]):
    """
    Pulls the close-tracker completion status and sign-off record into the
    audit trail itself, so an auditor reviewing this one file can see who
    did what and when without being handed two more separate files.
    """
    ws.title = '10 - Close & Signoff'
    ws.sheet_properties.tabColor = '283593'

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18

    ws.merge_cells('A1:D1')
    c = ws.cell(row=1, column=1, value='CLOSE TRACKER & SIGN-OFF SUMMARY')
    c.font      = _font(bold=True, size=12, color=_WHITE)
    c.fill      = _fill('283593')
    c.alignment = _align('center')
    ws.row_dimensions[1].height = 20

    row = 3
    _write_section_header(ws, row, 'CLOSE PROCESS TRACKER (9 STEPS)', 4)
    row += 1
    if not CLOSE_TRACKER_STEPS:
        ws.cell(row=row, column=1, value='Close tracker step list unavailable.').font = _font(italic=True)
        row += 1
    else:
        _write_header_row(ws, row, ['#', 'Step', 'Completed By', 'Timestamp'], fill_hex='283593', font_size=9)
        row += 1
        ct = close_tracker or {}
        for idx, label, _kind in CLOSE_TRACKER_STEPS:
            entry = ct.get(idx)
            done = bool(entry)
            bg = _GRP_GREEN_LITE if done else _GREY_LITE
            vals = [idx, label,
                    (entry or {}).get('completed_by', '—') if done else 'Pending',
                    (entry or {}).get('timestamp', '—') if done else '—']
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = _fill(bg)
                cell.font = _font(size=9, color='2E7D32' if done else '9E9E9E')
                cell.alignment = _align('left')
            row += 1

    row += 1
    _write_section_header(ws, row, 'SIGN-OFF CHECKLIST', 4)
    row += 1
    if not signoff_items:
        ws.cell(row=row, column=1, value='Sign-off checklist not available for this run.').font = _font(italic=True)
        return

    _write_header_row(ws, row, ['#', 'Item', 'Signed By', 'Timestamp'], fill_hex='283593', font_size=9)
    row += 1
    so = signoff_state or {}
    for idx, item in enumerate(signoff_items):
        entry = so.get(idx)
        done = bool(entry)
        bg = _GRP_GREEN_LITE if done else _GREY_LITE
        vals = [idx + 1, item,
                (entry or {}).get('signed_by', '—') if done else 'Pending',
                (entry or {}).get('timestamp', '—') if done else '—']
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = _fill(bg)
            cell.font = _font(size=9, color='2E7D32' if done else '9E9E9E')
            cell.alignment = _align('left')
        row += 1


# ── Public entry point ────────────────────────────────────────────────────────

def generate_audit_trail(
    output_path: str,
    period: str,
    property_name: str,
    all_je_lines: Optional[List[dict]] = None,
    fee_result=None,
    qc_report=None,
    prior_accrual_check: Optional[List[dict]] = None,
    files_uploaded: Optional[Dict[str, Any]] = None,
    property_config=None,
    property_code: str = '',
    bank_recon_detail=None,
    close_tracker: Optional[Dict[int, dict]] = None,
    signoff_state: Optional[Dict[int, dict]] = None,
    signoff_items: Optional[List[str]] = None,
) -> str:
    """
    Generate the GA Pipeline Audit Trail workbook.

    Args:
        output_path:         Destination .xlsx path.
        period:              Close period string, e.g. 'Mar-2026'.
        property_name:       Property display name.
        all_je_lines:        All JE lines from pass1_output_files['all_je_lines'].
        fee_result:          MgmtFeeResult from management_fee.py (or None).
        qc_report:           QCReport from qc_engine.py (or None).
        prior_accrual_check: Output of check_prior_accrual_vs_actual() (or None).
        files_uploaded:      Dict {key: path} of uploaded files for inventory tab.
        property_config:     Optional PropertyConfig — per-property values override
                             generic defaults (invoice prefix, entity names).
        property_code:       Yardi property code — used to reproduce the exact ETL
                             PROPERTY column value in Tab 6 (Yardi ETL CSV).
        bank_recon_detail:   BankReconDetail from engine_result.bank_recon_detail —
                             powers Tab 9 (Reconciling Item Aging).
        close_tracker:       st.session_state.close_tracker dict — powers Tab 10.
        signoff_state:       st.session_state.signoff_state dict — powers Tab 10.
        signoff_items:       Ordered list of sign-off checklist item labels — powers Tab 10.

    Returns:
        output_path (for chaining).
    """
    all_je_lines        = all_je_lines        or []
    prior_accrual_check = prior_accrual_check or []
    files_uploaded      = files_uploaded      or {}

    # Resolve property_code from config if not explicitly passed
    if not property_code and property_config:
        property_code = (getattr(property_config, 'yardi_etl_code', '')
                         or getattr(property_config, 'property_code', '') or '')

    run_ts = datetime.now().strftime('%Y-%m-%d  %H:%M')

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    ws1 = wb.create_sheet()
    _build_summary(ws1, period, property_name, all_je_lines,
                   fee_result, qc_report, files_uploaded, run_ts,
                   property_config=property_config)

    ws2 = wb.create_sheet()
    _build_je_log(ws2, all_je_lines)

    ws3 = wb.create_sheet()
    _build_mgmt_fee(ws3, fee_result, period, property_name, property_config=property_config)

    ws4 = wb.create_sheet()
    _build_accrual_check(ws4, prior_accrual_check)

    ws5 = wb.create_sheet()
    _build_qc_checks(ws5, qc_report)

    ws6 = wb.create_sheet()
    _build_yardi_csv_tab(ws6, all_je_lines, period, property_code)

    ws7 = wb.create_sheet()
    _build_methodology(ws7, property_config=property_config)

    ws8 = wb.create_sheet()
    _build_cutoff_review(ws8, all_je_lines, period)

    ws9 = wb.create_sheet()
    _build_reconciling_aging(ws9, bank_recon_detail, period)

    ws10 = wb.create_sheet()
    _build_close_signoff(ws10, close_tracker, signoff_state, signoff_items)

    wb.save(output_path)
    return output_path

"""
Accrual Entry Generator for GA Automation Pipeline
====================================================
Generates journal entries for accruals in priority layer order:

  Layer 0: Manual overrides (user-entered one-off accruals)
  Layer 1: Nexus open invoices (AP-side, deduped against GL)
  Layer 2: Invoice-period proration (utility accounts: daily rate × uncovered
           days; all other services: full prior invoice amount)
  Layer 3: Historical recurring — BC YTD actual ÷ months elapsed
           (accounts with prior-period spend but no current activity)
  Layer 4: Payroll bonus accruals (user-entered annual ÷ 12 or Kardin-derived)

First-layer-wins: each account is claimed by at most one layer.
Multi-layer detection is surfaced as a review flag on the first-layer entry.

Each accrual generates a two-line entry:
  DR  [Expense GL Account]
  CR  213100 Accrued Expenses (standard accrual liability)
"""

import os
import re
import calendar
from collections import defaultdict
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
from typing import List, Dict, Any, Optional
from openpyxl import Workbook

from accounting_utils import _round, _safe_float
from property_config import is_expense_account
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Period formatting ────────────────────────────────────────

_MONTH_ABBR_AEG = {
    'jan': 'January', 'feb': 'February', 'mar': 'March',
    'apr': 'April',   'may': 'May',       'jun': 'June',
    'jul': 'July',    'aug': 'August',    'sep': 'September',
    'oct': 'October', 'nov': 'November',  'dec': 'December',
}

def _fmt_period(period: str) -> str:
    """
    Format an accounting period string as 'Month YYYY' for JE descriptions.

    'Apr-2026' → 'April 2026'   |  '04/2026' → 'April 2026'
    Falls back to the raw string if unrecognised.
    """
    if not period:
        return period
    m = re.match(r'([A-Za-z]{3,})[- ](\d{4})', period.strip())
    if m:
        key = m.group(1)[:3].lower()
        return f"{_MONTH_ABBR_AEG.get(key, m.group(1).capitalize())} {m.group(2)}"
    m = re.match(r'(\d{1,2})[/\-](\d{4})', period.strip())
    if m:
        mon = int(m.group(1))
        names = list(_MONTH_ABBR_AEG.values())
        if 1 <= mon <= 12:
            return f"{names[mon - 1]} {m.group(2)}"
    return period


# ── GL dedup utilities ──────────────────────────────────────

def _normalize_vendor(name: str) -> str:
    """
    Reduce a vendor name to a stable comparison key.

    Lowercases, strips punctuation, collapses whitespace, and takes the first
    24 characters.  This lets "OpenPath Security, Inc." match "Openpath" and
    "Stewart Title" match "Stewart Title Guaranty Co." without false positives
    on short generic words.

    Returns '' if the name is blank or consists only of punctuation/whitespace.
    """
    if not name:
        return ''
    key = re.sub(r'[^a-z0-9 ]', ' ', name.lower())
    key = ' '.join(key.split())   # collapse whitespace
    return key[:24].strip()


def _build_gl_invoice_lookup(gl_data) -> dict:
    """
    Build lookup structures to check if an invoice is already in GL.

    Returns a dict with three keys:
      'by_reference'    : {reference_str: [txns]}   — exact invoice-number match
      'by_control'      : {control_str:   [txns]}   — control-number substring match
      'by_vendor_amount': {(account_code, vendor_key, amount_cents): [txns]}
                          — secondary dedup when invoice number is absent;
                            amount_cents is int(round(debit * 100)) for expense debits
    """
    lookup = {'by_reference': {}, 'by_control': {}, 'by_vendor_amount': {}}
    if not gl_data or not hasattr(gl_data, 'all_transactions'):
        return lookup

    for txn in gl_data.all_transactions:
        ref = (txn.reference or '').strip()
        if ref:
            lookup['by_reference'].setdefault(ref, []).append(txn)
        ctrl = (txn.control or '').strip()
        if ctrl:
            lookup['by_control'].setdefault(ctrl, []).append(txn)

        # Vendor+amount index — only for expense debit postings (debit > 0)
        if txn.debit > 0:
            vendor_key = _normalize_vendor(txn.description or '')
            if vendor_key:
                amount_cents = int(round(txn.debit * 100))
                va_key = (str(txn.account_code).strip(), vendor_key, amount_cents)
                lookup['by_vendor_amount'].setdefault(va_key, []).append(txn)

    return lookup


def _is_invoice_in_gl(invoice_number: str, gl_lookup: dict) -> bool:
    """Check if an invoice number already appears in GL transactions.

    Match strategy (in order):
      1. Exact match against GL reference field (by_reference dict)
      2. Exact match against GL control number (by_control dict)
      3. Case-insensitive exact match after stripping whitespace

    Substring matching is intentionally NOT used — e.g. invoice "123" must not
    match GL control "P-91234", which would silently suppress a real open invoice.
    """
    if not invoice_number:
        return False
    inv = invoice_number.strip()
    if inv in gl_lookup['by_reference']:
        return True
    # Exact control-number match only (no substring)
    if inv in gl_lookup['by_control']:
        return True
    # Case-insensitive fallback
    inv_lower = inv.lower()
    if any(inv_lower == ctrl.lower() for ctrl in gl_lookup['by_control']):
        return True
    return False


def _is_in_gl_by_vendor_amount(
    vendor: str, amount: float, account_code: str, gl_lookup: dict
) -> bool:
    """
    Secondary dedup: check whether an expense posting with matching vendor name
    and amount already exists in the GL for this account.

    Used only when the invoice number is absent (no reference field to match on).
    A $0.02 tolerance band is applied — we check ±2 cents — to absorb rounding
    differences between the AP system and GL.

    Strategy: vendor name normalization via _normalize_vendor(), then amount
    comparison by integer cents to avoid float equality traps.

    Returns True if a plausible match is found; False otherwise.
    """
    if not vendor or not amount or not account_code:
        return False
    vendor_key = _normalize_vendor(vendor)
    if not vendor_key:
        return False
    acct = str(account_code).strip()
    amount_cents = int(round(abs(amount) * 100))
    by_va = gl_lookup.get('by_vendor_amount', {})

    # Check exact match and ±2 cent tolerance
    for delta in (0, 1, -1, 2, -2):
        key = (acct, vendor_key, amount_cents + delta)
        if key in by_va:
            return True
    return False


# ── Constants ────────────────────────────────────────────────

AP_ACCRUAL_ACCOUNT    = '213100'
AP_ACCRUAL_NAME       = 'Accrued Expenses'

# Account-specific CR overrides: when the DR side is an interest expense account,
# credit 213200 Accrued Interest Payable instead of the standard 213100.
_CR_ACCOUNT_OVERRIDES: dict = {
    '801': ('213200', 'Accrued Interest Payable'),   # 801110 Interest Expense → Accrued Interest
}

def _cr_for(dr_account_code: str):
    """Return (cr_account, cr_name) for the given DR account code.
    Falls back to (AP_ACCRUAL_ACCOUNT, AP_ACCRUAL_NAME) for standard expense accounts.
    """
    code = str(dr_account_code or '').strip()
    for prefix, (cr_acct, cr_name) in _CR_ACCOUNT_OVERRIDES.items():
        if code.startswith(prefix):
            return cr_acct, cr_name
    return AP_ACCRUAL_ACCOUNT, AP_ACCRUAL_NAME

# Known periodic-billing contract accounts.
# The pipeline auto-detects the monthly portion via partial-contract coverage,
# but these accounts often carry quarterly or semi-annual billings that won't
# appear in the GL until the invoice arrives.  The UI surfaces these accounts
# with a supplement input so the reviewer can add the periodic amount on top
# of whatever the pipeline detected automatically.
#   billing_cycle: 'monthly' | 'quarterly' | 'semi-annual'
PERIODIC_CONTRACT_ACCOUNTS: dict = {
    '617110': {'label': 'HVAC Contract',       'billing_cycle': 'quarterly'},
    '619120': {'label': 'PPM Water Treatment', 'billing_cycle': 'monthly'},
    '627230': {'label': 'Fire / Life Safety',  'billing_cycle': 'monthly'},
}
# Tenant sub-metered utility recovery accounts.
# Each month the meter read company provides per-tenant consumption data.
# The property manager posts TWO JEs:
#
#   (1) Per-tenant billing (AR / Recovery):
#       DR 133110  Accounts Receivable Billback   (per tenant, electric and gas)
#       CR 440500  Recovery - Electricity         (electric portion)
#       CR 440700  Recovery - Misc Utilities      (gas portion)
#
#   (2) P&L reclassification (aggregate, electric only):
#       DR 613115  Tenant Electric Reimbursement  (total electric billed to tenants)
#       CR 613110  Utilities - Electricity        (offsets main electricity expense)
#
# JE (2) moves the tenant-reimbursable portion from the main electricity line
# to a dedicated sub-account so the P&L shows landlord vs. tenant shares clearly.
#
# If the meter read JE hasn't been posted yet at close, the pipeline accrues
# the budget amount so NOI is not understated. When the actual meter read
# data is available, the sidebar overrides with per-tenant actual amounts.
TENANT_UTILITY_AR_ACCOUNT   = '133110'
TENANT_UTILITY_AR_NAME      = 'Accounts Receivable Billback'
ELEC_EXPENSE_ACCOUNT        = '613110'
ELEC_EXPENSE_NAME           = 'Utilities - Electricity'
GAS_EXPENSE_ACCOUNT         = '613210'   # Utilities - Gas (per-vendor daily-rate proration)
ELEC_TENANT_REIMB_ACCOUNT   = '613115'
# Metered utility accounts: use per-vendor daily-rate proration (separate line per vendor)
# Both electricity (613110) and gas (613210) have multiple vendors / meters per billing period.
# Properties whose gas GL code differs from 613210 should add their code via
# metered_utility_accounts / per_invoice_utility_accounts in config.yaml — no code change needed.
_METERED_UTILITY_ACCOUNTS   = {ELEC_EXPENSE_ACCOUNT, GAS_EXPENSE_ACCOUNT}
# Gas uses per-INVOICE accrual (one line per meter/service, not per vendor).
# Electricity uses per-VENDOR accrual (all invoices from same vendor combined).
_PER_INVOICE_UTILITY_ACCOUNTS = {GAS_EXPENSE_ACCOUNT}
ELEC_TENANT_REIMB_NAME      = 'Tenant Electric Reimbursement'
TENANT_UTILITY_ACCOUNTS: dict = {
    '440500': {'label': 'Tenant Electric Recovery',     'budget_key': '440500'},
    '440700': {'label': 'Tenant Gas Recovery',          'budget_key': '440700'},
}


def _j_credits(gl_acct) -> float:
    """Sum of J-type (Journal) credit amounts on a GLAccount.

    Yardi transaction control codes: C=Charge, R=Receipt, P=Payable,
    K=Check/PCard, J=Journal.  Only J-type entries are pipeline accruals;
    C/R/P/K entries are real billing/payment transactions and must NOT
    suppress the pipeline's monthly accrual JE.
    """
    if gl_acct is None:
        return 0.0
    return sum(
        t.credit for t in getattr(gl_acct, 'transactions', [])
        if t.credit > 0 and str(getattr(t, 'control', '') or '').upper().startswith('J')
    )


def _j_debits(gl_acct) -> float:
    """Sum of J-type (Journal) debit amounts on a GLAccount."""
    if gl_acct is None:
        return 0.0
    return sum(
        t.debit for t in getattr(gl_acct, 'transactions', [])
        if t.debit > 0 and str(getattr(t, 'control', '') or '').upper().startswith('J')
    )


def _net_j_credit(gl_acct) -> float:
    """
    Net J-type credit on a GLAccount: total J-credits minus total J-debits.

    Returns the amount that does NOT offset to zero across all J-entries.
    Example: $130K J-debit + $130K J-credit + $48K J-credit → net = $48K.
    The $130K pair cancels; only the unmatched $48K (the open prior-accrual
    reversal) is returned.  Returns 0 if J-debits exceed J-credits.
    """
    if gl_acct is None:
        return 0.0
    txns = getattr(gl_acct, 'transactions', [])
    j_cr = sum(t.credit for t in txns
               if t.credit > 0 and str(getattr(t, 'control', '') or '').upper().startswith('J'))
    j_dr = sum(t.debit  for t in txns
               if t.debit  > 0 and str(getattr(t, 'control', '') or '').upper().startswith('J'))
    return max(0.0, round(j_cr - j_dr, 2))

PREPAID_ASSET_ACCOUNT = '135150'
PREPAID_ASSET_NAME    = 'Prepaid Other'

# ── Payroll bonus accounts ───────────────────────────────────────────────────
# Bonuses post to the same account codes as regular payroll.  The annual bonus
# is paid once or twice a year (Jan and Jul at Revolution Labs) but Kardin
# budgets it evenly across all 12 months.
#
# Monthly bonus accrual = (Kardin annual budget ÷ 12) − standard_monthly
#   where standard_monthly = the minimum monthly value across M1–M12
#   (i.e., the non-payment months that carry only base payroll).
#
# The accrual posts every month UNLESS the GL net_change for the period already
# equals or exceeds the monthly average (meaning the actual payment hit the GL
# in that period — no separate accrual needed).
#
# 'kardin_keywords' are matched against the Kardin row's 'description' field
# to select only the bonus-inclusive budget row (not the SX/admin-overhead rows).
PAYROLL_BONUS_ACCOUNTS: dict = {
    '615110': {
        'label':            'RM-Pay/Wages',
        'kardin_keywords':  ['bonus', 'payroll', 'ot'],
    },
    '637110': {
        'label':            'Admin-Pay/Wages',
        'kardin_keywords':  ['bonus', 'salary'],
    },
}

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
DOUBLE_BOTTOM = Border(bottom=Side(style='double'))

DARK_BLUE = '1F4E78'
MED_BLUE = '2E75B6'
LIGHT_BLUE = 'D6E4F0'
LIGHT_GRAY = 'F2F2F2'
WHITE = 'FFFFFF'


def _apply(cell, font=None, fill=None, fmt=None, border=None, align=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = border
    if align:
        cell.alignment = align


def _hdr_font():
    return Font(name='Calibri', size=11, bold=True, color='FFFFFF')

def _hdr_fill():
    return PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')

def _subhdr_fill():
    return PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')


# ── Layer 1b: Insurance prepaid amortization ─────────────────

# Insurance prepaid account and expense accounts
_PREPAID_INSURANCE_ACCT = '135110'   # Restricted Insurance / Prepaid Insurance
_INSURANCE_EXPENSE_ACCTS = {'639110', '639120'}

def detect_insurance_amortization(
    gl_data,
    budget_data,
    period: str = '',
    insurance_policies: Optional[List[Dict[str, Any]]] = None,
    kardin_records: Optional[List[Dict[str, Any]]] = None,
) -> List[Dict[str, Any]]:
    """
    Generate monthly insurance expense entries from Prepaid Insurance (135110).

    JLL method: Annual premiums are paid upfront and held in 135110 Prepaid
    Insurance. Each month JLL posts:
        DR 639110  Insurance-Property         (monthly amount)
        DR 639120  Insurance-General Liability (monthly amount)
        CR 135110  Prepaid Insurance           (combined monthly total)

    We generate these only when:
      1. 135110 has a positive ending balance (prepaid exists to draw down)
      2. The insurance expense account has no current-period GL activity
         (i.e., JLL hasn't posted the entry yet — normal for pre-close GL)

    Three modes (first matching wins):
      Mode A — Config-driven (insurance_policies provided):
        Generates ONE JE line per policy entry in the list.  Each policy has
        a name, expense_account, and fixed monthly_amount.  Supports multiple
        policies on the same expense account (e.g. general property + umbrella,
        both on 639110).  Config amounts take precedence over the budget.

      Mode B — Kardin-driven (kardin_records provided, no config):
        Uses individual Kardin budget rows for insurance expense accounts.
        Each row in Kardin (e.g. "Property Insurance", "Umbrella Policy") maps
        to one JE line, using the month-specific budget amount (m1–m12).
        This correctly handles multiple named policies on the same account code.

      Mode C — BC-driven fallback (budget_data provided, no config or Kardin):
        Falls back to the Budget Comparison PTD column — one combined line per
        expense account code.  Matches JLL's calculation within a few cents.

    Returns a list of dicts: one per line, with 'credit_account' / 'credit_name'
    keys so build_accrual_entries() generates CR 135110 instead of 213100.
    """
    results: List[Dict[str, Any]] = []

    # gl_data is always required; budget_data is only needed for Mode C (BC fallback)
    if not gl_data:
        return results

    _period_label = _fmt_period(period)

    # 1. Check that 135110 has a positive balance to amortise from
    prepaid_balance = 0.0
    for acct in (gl_data.accounts if hasattr(gl_data, 'accounts') else []):
        if str(acct.account_code).strip() == _PREPAID_INSURANCE_ACCT:
            prepaid_balance = acct.ending_balance
            break

    if prepaid_balance <= 0:
        return results

    # 2. Find which insurance expense accounts already have net new expense activity.
    # Use net_change > 0 (net debit) — a net credit means prior accrual auto-reversed
    # and the account still needs amortization this period.
    already_posted: set = set()
    for acct in (gl_data.accounts if hasattr(gl_data, 'accounts') else []):
        code = str(acct.account_code).strip()
        if code in _INSURANCE_EXPENSE_ACCTS and acct.net_change > 0.01:
            already_posted.add(code)

    # ── Mode A: Config-driven — one line per named policy ─────────────────────
    if insurance_policies:
        for pol in insurance_policies:
            code    = str(pol.get('expense_account', '639110')).strip()
            name    = str(pol.get('name', '') or code)
            monthly = float(pol.get('monthly_amount', 0.0) or 0.0)
            if monthly < 1.0:
                continue
            if code in already_posted:
                continue  # JLL already posted it this period

            results.append({
                'account_code':   code,
                'account_name':   'Insurance',
                'amount':         _round(monthly),
                'credit_account': _PREPAID_INSURANCE_ACCT,
                'credit_name':    'Prepaid Insurance',
                'source':         'prepaid_amortization',
                'confidence':     'high',
                'description': (
                    f'Accrual {_period_label} — {name} '
                    f'(insurance prepaid amortization ${monthly:,.2f}/mo, '
                    f'prepaid balance ${prepaid_balance:,.2f})'
                ),
            })
        return results

    # ── Mode B: Kardin-driven — one line per named policy row ─────────────────
    # Kardin exports each insurance policy as a separate row under the same
    # account code (e.g. two rows under 639110: "Property Insurance" and
    # "Umbrella Policy"), with monthly amounts in m1–m12.  We generate one JE
    # line per Kardin row so both policies get their own description and amount.
    _MMAP_INS = {
        'jan': 'm1', 'feb': 'm2', 'mar': 'm3', 'apr': 'm4',
        'may': 'm5', 'jun': 'm6', 'jul': 'm7', 'aug': 'm8',
        'sep': 'm9', 'oct': 'm10', 'nov': 'm11', 'dec': 'm12',
    }
    _period_col = next(
        (col for abbr, col in _MMAP_INS.items() if abbr in (period or '').lower()),
        None,
    )
    _kardin_ins_rows = [
        r for r in (kardin_records or [])
        if str(r.get('account_code', '') or '').strip() in _INSURANCE_EXPENSE_ACCTS
    ]
    if _kardin_ins_rows and _period_col:
        for row in _kardin_ins_rows:
            code = str(row.get('account_code', '') or '').strip()
            if code in already_posted:
                continue
            name    = str(row.get('description', '') or row.get('account_name', '') or code).strip()
            monthly = abs(float(row.get(_period_col, 0) or 0))
            if monthly < 1.0:
                continue
            results.append({
                'account_code':   code,
                'account_name':   'Insurance',
                'amount':         _round(monthly),
                'credit_account': _PREPAID_INSURANCE_ACCT,
                'credit_name':    'Prepaid Insurance',
                'source':         'prepaid_amortization',
                'confidence':     'high',
                'description': (
                    f'Accrual {_period_label} — {name} '
                    f'(insurance prepaid amortization ${monthly:,.2f}/mo, '
                    f'prepaid balance ${prepaid_balance:,.2f})'
                ),
            })
        return results

    # ── Mode C: BC-driven fallback — one combined line per expense account code ─
    budget_rows = budget_data if isinstance(budget_data, list) else []
    for row in budget_rows:
        code = str(row.get('account_code', '') or '').strip()
        if code not in _INSURANCE_EXPENSE_ACCTS:
            continue
        if code in already_posted:
            continue  # JLL already posted it this period

        name       = str(row.get('account_name', '') or code)
        monthly    = abs(float(row.get('ptd_budget', 0) or 0))
        if monthly < 1.0:
            continue

        results.append({
            'account_code':   code,
            'account_name':   name,
            'amount':         _round(monthly),
            'credit_account': _PREPAID_INSURANCE_ACCT,
            'credit_name':    'Prepaid Insurance',
            'source':         'prepaid_amortization',
            'confidence':     'high',
            'description': (
                f'Accrual {_period_label} — {name} '
                f'(insurance prepaid amortization ${monthly:,.2f}/mo, '
                f'prepaid balance ${prepaid_balance:,.2f})'
            ),
        })

    return results


# ── Layer 1c: Real Estate Tax amortization ───────────────────

_RETAX_EXPENSE_ACCT    = '641110'   # Real Estate Taxes (income statement)
_RETAX_PREPAID_ACCT    = '135120'   # Prepaid RE Taxes (balance sheet — credit for monthly JE)
_RETAX_ESCROW_ACCT     = '115200'   # RE Tax Escrow — Berkadia-held; ties to loan statement

# Lexington tax bills due Feb, May, Aug, Nov — Berkadia pays the month prior.
# Payment months: Jan (for Feb bill), Apr (for May bill),
#                 Jul (for Aug bill), Oct (for Nov bill).
_RETAX_PAYMENT_MONTHS  = frozenset({1, 4, 7, 10})


def detect_retax_amortization(
    gl_data,
    period: str = '',
    re_tax_bill_amount: float = 0.0,
    re_tax_payment_months=None,
) -> Optional[Dict[str, Any]]:
    """
    RE Tax prepaid deferral / release JE.  Fires every month automatically —
    falls back to GL auto-detection when re_tax_bill_amount is not entered.

    Business rule (quarterly invoice cycle — Jan / Apr / Jul / Oct):
    ─────────────────────────────────────────────────────────────────
    Payment months (Jan / Apr / Jul / Oct):
        Berkadia pays the town from escrow; Yardi auto-posts the full quarterly
        bill via the loan payment entry:
            DR 641110  Real Estate Taxes (full bill — automatic, NOT by pipeline)
            CR 115200  RE Tax Escrow     (full bill — automatic, NOT by pipeline)

        The pipeline DEFERS 2/3 so only 1/3 hits the P&L this month:
            DR 135120  Prepaid RE Taxes  (2/3 of quarterly bill)
            CR 641110  Real Estate Taxes (2/3 of quarterly bill)

        Suppressed if 135120 already carries a net debit > $100 (deferral
        was posted to Yardi before the pipeline ran).

        GL auto-detect: if re_tax_bill_amount not entered, reads net_641110
        from the GL — Berkadia's auto-post (DR 641110 / CR 115200) makes the
        quarterly bill directly visible as a net debit in 641110.

    Release months (all other months):
        Releases 1/3 of the quarterly bill from prepaid back to expense:
            DR 641110  Real Estate Taxes (1/3 of quarterly bill)
            CR 135120  Prepaid RE Taxes  (1/3 of quarterly bill)

        Suppressed if 641110 already carries a net debit > $100 (release
        was posted to Yardi before the pipeline ran).

        GL auto-detect: if re_tax_bill_amount not entered, back-calculates
        the quarterly bill from 135120's beginning_balance:
          • 1st release month (Feb/May/Aug/Nov): beg_balance = 2/3 × bill
            → bill = beg_balance × 1.5
          • 2nd release month (Mar/Jun/Sep/Dec): beg_balance = 1/3 × bill
            → bill = beg_balance × 3.0

    Args:
        gl_data:             GL parse result (.accounts list with .net_change
                             and .beginning_balance).
        period:              Close period string e.g. 'Jan-2026'.
        re_tax_bill_amount:  Quarterly RE tax bill (user-entered); 0 triggers
                             GL auto-detection.

    Returns a JE dict or None.
    """
    if not gl_data or not hasattr(gl_data, 'accounts'):
        return None

    # Use property-specific payment months or fall back to module default.
    # Cast to int to handle YAML parsing returning strings ('1','4','7','10')
    # instead of integers — prevents period_month (int) membership check failing.
    _payment_months = (
        frozenset(int(m) for m in re_tax_payment_months)
        if re_tax_payment_months
        else _RETAX_PAYMENT_MONTHS
    )

    # Parse period month ("Jan-2026" → 1)
    _MMAP = {
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,
        'may': 5, 'jun': 6, 'jul': 7, 'aug': 8,
        'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
    }
    period_month = 0
    for abbr, num in _MMAP.items():
        if abbr in (period or '').lower():
            period_month = num
            break
    if not period_month:
        return None

    _period_label = _fmt_period(period)

    # Collect current-period net changes, beginning balances, and per-account
    # transaction lists for the single-transaction auto-detection logic below.
    net_641110 = 0.0   # Real Estate Taxes expense (positive = net debit)
    net_115200 = 0.0   # RE Tax Escrow (positive = net debit; Berkadia credit → negative)
    net_135120 = 0.0   # Prepaid RE Taxes asset   (positive = net debit)
    beg_135120 = 0.0   # Prepaid RE Taxes beginning balance (for release auto-detect)
    txns_641110: list = []  # individual GLTransaction objects for 641110
    txns_115200: list = []  # individual GLTransaction objects for 115200
    for acct in gl_data.accounts:
        code = str(acct.account_code).strip()
        if code == _RETAX_EXPENSE_ACCT:
            net_641110  = float(getattr(acct, 'net_change', 0) or 0)
            txns_641110 = list(getattr(acct, 'transactions', None) or [])
        elif code == _RETAX_ESCROW_ACCT:
            net_115200  = float(getattr(acct, 'net_change', 0) or 0)
            txns_115200 = list(getattr(acct, 'transactions', None) or [])
        elif code == _RETAX_PREPAID_ACCT:
            net_135120 = float(getattr(acct, 'net_change', 0) or 0)
            beg_135120 = float(getattr(acct, 'beginning_balance', 0) or 0)

    # ── Auto-detect quarterly bill from GL if not user-provided ──────────────
    bill        = re_tax_bill_amount
    auto_source = ''
    if bill <= 0:
        if period_month in _payment_months:
            # Berkadia auto-posts ONE entry: DR 641110 / CR 115200 for the full bill.
            # Net figures are unreliable because:
            #   115200 net: polluted by monthly escrow deposit debits from the loan payment
            #   641110 net: polluted by the prior-period pipeline release JE auto-reversal
            #               (which credits 641110 for the 1/3 release amount)
            #
            # Solution: find the LARGEST SINGLE TRANSACTION in each account.
            # The Berkadia auto-post is one big transaction; the noise entries are
            # smaller and separate.
            #
            # Priority:
            #   1. Largest single CREDIT in 115200 — cleanest; no pipeline JEs touch 115200
            #   2. Largest single DEBIT  in 641110 — also reliable; pipeline releases are
            #      typically much smaller than the quarterly bill
            #   3. Net 641110 — last resort (may understate due to reversal credits)
            _max_115200_cr = max(
                (float(getattr(t, 'credit', 0) or 0) for t in txns_115200),
                default=0.0,
            )
            _max_641110_dr = max(
                (float(getattr(t, 'debit', 0) or 0) for t in txns_641110),
                default=0.0,
            )
            if _max_115200_cr > 10_000:
                bill        = _max_115200_cr
                auto_source = (
                    f'auto-detected from GL 115200 largest single credit '
                    f'${_max_115200_cr:,.2f} (Berkadia tax payment)'
                )
            elif _max_641110_dr > 10_000:
                bill        = _max_641110_dr
                auto_source = (
                    f'auto-detected from GL 641110 largest single debit '
                    f'${_max_641110_dr:,.2f} (Berkadia tax payment)'
                )
            elif net_641110 > 10_000:
                # Last resort: net figure (may be understated if reversal credits present)
                bill        = net_641110
                auto_source = f'auto-detected from GL 641110 net debit ${net_641110:,.2f} (fallback)'
        else:
            # Back-calculate from 135120 beginning balance
            # 1st release months (Feb/May/Aug/Nov): beg = 2/3 × bill → ×1.5
            # 2nd release months (Mar/Jun/Sep/Dec): beg = 1/3 × bill → ×3.0
            if beg_135120 > 100:
                # Compute release months dynamically from the configured payment months
                # 1st release = month after each payment month; 2nd release = 2 months after
                _FIRST_RELEASE = {(m % 12) + 1 for m in _payment_months}
                multiplier = 1.5 if period_month in _FIRST_RELEASE else 3.0
                bill        = _round(beg_135120 * multiplier)
                auto_source = (
                    f'auto-detected from GL 135120 beginning balance '
                    f'${beg_135120:,.2f} × {multiplier}'
                )

    if bill <= 0:
        import warnings as _warnings
        # H-5 / A-6: Surface a meaningful warning so a zero-bill never silently
        # suppresses the RE tax JE without any indication to the user.
        if period_month in _payment_months:
            # Payment month: couldn't find the Berkadia credit or 641110 net debit.
            _warnings.warn(
                f'RE tax bill could not be auto-detected for payment month {period_month} '
                f'(GL 115200 credit = ${-net_115200:,.2f}, GL 641110 net debit = ${net_641110:,.2f}). '
                f'No Berkadia entry found in GL. RE tax deferral JE will be skipped. '
                f'Enter the bill amount manually in the One-Off Accruals table '
                f'(DR 135120 Prepaid RE Taxes / CR 641110 Real Estate Taxes).',
                UserWarning,
                stacklevel=3,
            )
        else:
            # Release month: 135120 beginning balance was ≤ $100, so back-calculation
            # returned $0.  Either the prepaid was already fully released or the
            # GL beginning balance is missing/wrong.
            _warnings.warn(
                f'RE tax release JE skipped for month {period_month}: GL 135120 '
                f'beginning balance is ${beg_135120:,.2f} (expected > $100 for a '
                f'release month).  If a release is expected, check the 135120 '
                f'beginning balance in the GL export or enter the amount manually '
                f'(DR 641110 Real Estate Taxes / CR 135120 Prepaid RE Taxes).',
                UserWarning,
                stacklevel=3,
            )
        return None

    if period_month in _payment_months:
        # ── Payment month: defer 2/3 → DR 135120 / CR 641110 ────────────────
        if net_135120 > 100.0:
            return None   # deferral already posted in Yardi — suppress

        deferred     = _round(bill * 2.0 / 3.0)
        source_note  = f' ({auto_source})' if auto_source else ''
        return {
            'account_code':   _RETAX_PREPAID_ACCT,
            'account_name':   'Prepaid RE Taxes',
            'amount':         deferred,
            'credit_account': _RETAX_EXPENSE_ACCT,
            'credit_name':    'Real Estate Taxes',
            'source':         'prepaid_amortization',
            'confidence':     'high',
            'auto_reverse':   False,
            'description': (
                f'Accrual {_period_label} — Real Estate Taxes '
                f'(prepaid deferral ${deferred:,.2f}, '
                f'quarterly bill ${bill:,.2f} × 2/3; '
                f'Berkadia auto-posts full bill, pipeline defers 2/3'
                f'{source_note})'
            ),
        }
    else:
        # ── Release month: release 1/3 → DR 641110 / CR 135120 ──────────────
        if net_641110 > 100.0:
            return None   # release already posted in Yardi — suppress

        release      = _round(bill / 3.0)
        source_note  = f' ({auto_source})' if auto_source else ''
        return {
            'account_code':   _RETAX_EXPENSE_ACCT,
            'account_name':   'Real Estate Taxes',
            'amount':         release,
            'credit_account': _RETAX_PREPAID_ACCT,
            'credit_name':    'Prepaid RE Taxes',
            'source':         'prepaid_amortization',
            'confidence':     'high',
            'auto_reverse':   False,
            'description': (
                f'Accrual {_period_label} — Real Estate Taxes '
                f'(prepaid release ${release:,.2f}, '
                f'quarterly bill ${bill:,.2f} / 3'
                f'{source_note})'
            ),
        }


# detect_retax_escrow_je() removed May 2026 — retired, never called.
# Berkadia handles DR 641110 / CR 115200 automatically in Yardi.
# All RE tax pipeline entries generated by detect_retax_amortization().


# ── Tenant utility billing detection ────────────────────────

def detect_tenant_utility_billing(gl_data, budget_data) -> List[Dict[str, Any]]:
    """
    Check whether the tenant sub-metered utility billing JE (meter read) has
    been posted this period for 440500 (electric) and 440700 (gas).

    When NOT posted:  returns budget accrual candidates so the income side of
    NOI is not understated while the expense proration is accruing the full
    building bill.

    When already posted: returns nothing (GL already has the income entry).

    The pipeline accrues ONE aggregate line per account (budget amount) as a
    placeholder.  When the sidebar provides per-tenant actual amounts, those
    replace the budget aggregate and generate one JE line per tenant.

    Returns list of dicts:
        account_code, account_name, amount (budget), label,
        source='tenant_utility_billing', confidence='medium'
    """
    results: List[Dict[str, Any]] = []
    if not gl_data or not budget_data:
        return results

    # Build budget amount lookup
    budget_by_code: Dict[str, float] = {}
    rows = budget_data if isinstance(budget_data, list) else getattr(budget_data, 'line_items', [])
    for row in rows:
        code = str((row.get('account_code') if isinstance(row, dict)
                    else getattr(row, 'account_code', '')) or '').strip()
        ptd  = (row.get('ptd_budget') if isinstance(row, dict)
                else getattr(row, 'ptd_budget', 0)) or 0
        budget_by_code[code] = abs(float(ptd))

    # Check each tenant utility account
    gl_accounts_by_code: Dict[str, Any] = {}
    for acct in (gl_data.accounts if hasattr(gl_data, 'accounts') else []):
        gl_accounts_by_code[str(acct.account_code).strip()] = acct

    for code, info in TENANT_UTILITY_ACCOUNTS.items():
        acct = gl_accounts_by_code.get(code)
        # Only J-type (Journal) credits indicate a pipeline accrual already posted.
        # C-type charges, R-type receipts, etc. are real billing transactions
        # and must not suppress the monthly budget accrual candidate.
        if _j_credits(acct) > 0.01:
            continue   # J-accrual already posted this period

        budget_amt = budget_by_code.get(code, 0.0)
        if budget_amt < 1.0:
            continue

        results.append({
            'account_code': code,
            'account_name': info['label'],
            'amount':       _round(budget_amt),
            'label':        info['label'],
            'source':       'tenant_utility_billing',
            'confidence':   'medium',
            'description': (
                f'Tenant utility accrual — {info["label"]}: '
                f'meter read JE not yet posted. '
                f'Accruing budget ${budget_amt:,.2f}. '
                f'Update with actual per-tenant amounts when meter read received.'
            ),
        })

    return results


# ── Layer 2: Invoice-period proration ────────────────────────

# Billing date range: "01.31.26-03.02.26" or "01.31.26 - 03.02.26"
_DATE_RANGE_RE = re.compile(
    r'(\d{2})\.(\d{2})\.(\d{2})\s*-\s*(\d{2})\.(\d{2})\.(\d{2})'
)
# Single date: "03.13.26"
_SINGLE_DATE_RE = re.compile(r'(\d{2})\.(\d{2})\.(\d{2})')

# Account name fragments that indicate a payroll line
_PAYROLL_NAME_KW  = ('pay/wages', 'pay wages', 'payroll')
# Transaction description fragments that confirm a payroll entry
_PAYROLL_DESC_KW  = ('payroll', 'eng payroll', 'admin payroll', 'pay/wages')



def _parse_date_range(text: str):
    """
    Parse 'MM.DD.YY-MM.DD.YY' billing period from a GL description/remarks string.

    Returns (start: date, end: date) or (None, None) if not found.
    Years are assumed 20xx (adequate through 2099).
    """
    m = _DATE_RANGE_RE.search(text or '')
    if not m:
        return None, None
    try:
        start = date(2000 + int(m.group(3)),  int(m.group(1)),  int(m.group(2)))
        end   = date(2000 + int(m.group(6)),  int(m.group(4)),  int(m.group(5)))
        return (start, end) if end >= start else (None, None)
    except ValueError:
        return None, None


def _parse_single_date(text: str) -> Optional[date]:
    """Parse the first 'MM.DD.YY' date in text. Returns None if none found."""
    m = _SINGLE_DATE_RE.search(text or '')
    if not m:
        return None
    try:
        return date(2000 + int(m.group(3)), int(m.group(1)), int(m.group(2)))
    except ValueError:
        return None


def _month_end_from_period(period_str: str) -> Optional[date]:
    """
    Derive the last calendar day of the reporting month from a period string.

    Handles formats:
      'Mar-2026'  →  date(2026, 3, 31)
      'Mar 2026'  →  date(2026, 3, 31)
    """
    _MONTH_MAP = {
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4,  'May': 5,  'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12,
    }
    if not period_str:
        return None
    m = re.match(r'([A-Za-z]{3})[\s\-](\d{4})', period_str.strip())
    if not m:
        return None
    month = _MONTH_MAP.get(m.group(1).capitalize())
    year  = int(m.group(2))
    if not month:
        return None
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, last_day)


def detect_invoice_proration_accruals(
    gl_data,
    period: str = '',
    month_end: Optional[date] = None,
    materiality: float = 2500.0,
    metered_utility_accounts: Optional[List[str]] = None,
    per_invoice_utility_accounts: Optional[List[str]] = None,
    per_invoice_accrual_accounts: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    """
    Layer 2 — Invoice-period accruals.

    Scans GL transactions for billing date-range references in the remarks /
    description field (format ``MM.DD.YY-MM.DD.YY``).  For each expense account
    where the latest invoiced period ends *before* the close of the reporting
    month, accrues the current month's estimated cost.

    Algorithm
    ---------
    For each expense account (6xxxxx, 5xxxxx, …):

    **Electricity (613110 only)**
      1. Parse ``(start, end, amount)`` from every transaction with a date range.
      2. Group by billing end date; identify the *latest* end date.
      3. For the latest group: compute daily rate = total amount / period days.
      4. Uncovered days  = calendar month-end  −  latest billing end.
      5. Accrual = daily rate × uncovered days   (if > materiality threshold).

    **All other accounts** (gas, water, sewer, janitorial, HVAC, security, etc.)
      1. Parse ``(start, end, amount)`` from every transaction with a date range.
      2. Identify the *latest* billing end date.
      3. Accrual = full last invoice amount (flat monthly rate assumption).

    **Payroll accounts** (account name contains "Pay/Wages" or "Payroll")
      1. Identify payroll runs by description keyword.
      2. Determine pay period length from the gap between consecutive run dates.
      3. Sum all charges in the latest pay period (regular + OT, etc.).
      4. Daily rate = period total / pay-period days.
      5. Accrual = daily rate × days from last run to month-end.

    For multi-vendor accounts (e.g., electricity has both Eversource and
    Hudson Energy), invoices sharing the same billing end date are *combined*:
    the daily rate is the sum across all vendors, accurately reflecting the
    total daily cost of service.

    Args:
        gl_data:     GLParseResult (from parsers.yardi_gl.parse_gl)
        period:      Accounting period string, e.g. 'Mar-2026' (used to derive
                     month-end when ``month_end`` is not supplied explicitly)
        month_end:   Override: last day of the reporting month.  If None, derived
                     from ``period`` or from gl_data.metadata.period.
        materiality: Minimum accrual (default $500) — smaller amounts are skipped.

    Returns:
        List of candidate dicts::

            account_code, account_name, accrual_amount, source ('invoice_proration'),
            description, daily_rate, uncovered_days, period_days, invoice_total
    """
    candidates: List[Dict[str, Any]] = []

    if not gl_data or not hasattr(gl_data, 'accounts'):
        return candidates

    _period_label = _fmt_period(period)

    # ── Resolve reporting month-end ────────────────────────────────────────────
    if month_end is None:
        month_end = _month_end_from_period(period)
    if month_end is None:
        # Try GL metadata
        try:
            month_end = _month_end_from_period(gl_data.metadata.period)
        except Exception:
            pass
    if month_end is None:
        return candidates   # can't prorate without knowing when the month ends

    # ── Effective utility account sets (module defaults + per-property overrides) ──
    # Properties can extend the default sets via metered_utility_accounts /
    # per_invoice_utility_accounts in their config.yaml without touching source code.
    _eff_metered     = set(_METERED_UTILITY_ACCOUNTS)
    _eff_per_invoice = set(_PER_INVOICE_UTILITY_ACCOUNTS)
    if metered_utility_accounts:
        _eff_metered.update(str(a) for a in metered_utility_accounts)
    if per_invoice_utility_accounts:
        _eff_per_invoice.update(str(a) for a in per_invoice_utility_accounts)
        # Any per-invoice account is also implicitly metered
        _eff_metered.update(str(a) for a in per_invoice_utility_accounts)

    # Accounts where Pass 2 should emit one candidate *per invoice* rather than
    # one combined candidate (e.g. Casella — distinct service types per line).
    _eff_per_invoice_accrual: set = set()
    if per_invoice_accrual_accounts:
        _eff_per_invoice_accrual.update(str(a) for a in per_invoice_accrual_accounts)

    for acct in gl_data.accounts:
        code = str(acct.account_code).strip()
        if not code or code[0] not in ('5', '6', '7', '8'):
            continue

        # ── VENDOR BILLING-PERIOD PRORATION ───────────────────────────────────
        # Group transactions that carry a billing date range by their end date.
        # Store (start, end, amount, vendor_desc) so electricity can be split
        # by vendor (electric service vs. electric supplier).
        by_end: Dict[date, List[tuple]] = defaultdict(list)
        has_range_txns = False

        for txn in acct.transactions:
            amt = (txn.debit or 0) - (txn.credit or 0)
            if amt <= 0:
                continue
            start, end = _parse_date_range(txn.remarks or '')
            if start is None:
                start, end = _parse_date_range(txn.description or '')
            if start and end:
                # Vendor name for grouping and display.
                # Use description up to the first '(' — this gives the clean
                # vendor name ("Eversource", "Hudson Energy Services LLC") and
                # is stable across invoices from the same vendor.
                _vname = (txn.description or '').split('(')[0].strip()
                # Service description: remarks content after stripping the date-range prefix.
                # e.g. "11.20.25-12.22.25 82953-68006 HVAC Delivery" → "82953-68006 HVAC Delivery"
                # Used by per-invoice accounts (gas) to distinguish individual meters.
                _rem_raw = (txn.remarks or '').strip()
                _service_desc = re.sub(
                    r'^\d{2}\.\d{2}\.\d{2}\s*-\s*\d{2}\.\d{2}\.\d{2}\s*', '', _rem_raw
                ).strip()
                by_end[end].append((start, end, amt, _vname, _service_desc))
                has_range_txns = True

        if has_range_txns:
            latest_end = max(by_end.keys())
            uncovered  = (month_end - latest_end).days

            if uncovered <= 0:
                # Latest invoice already covers the full month
                continue

            # ── Electricity vs. all other accounts ─────────────────────────
            # Electricity (613110 only):
            #   Prorate by day: daily rate × uncovered days.  Electric bills
            #   span a metered cycle that rarely aligns to month-end, so the
            #   exact uncovered days gives the most accurate accrual.
            #   BREAKOUT: generate one candidate per vendor so electric service
            #   (Eversource delivery) and electric supplier (competitive supplier,
            #   e.g. Hudson, Constellation) appear as separate JE lines.
            #   Each vendor uses ITS OWN latest billing end date — avoids
            #   dropping vendors whose billing cycle ends on a different day
            #   from the vendor with the global latest end date.
            #
            # Everything else (water, sewer, janitorial, HVAC, security,
            #   elevator, etc.):
            #   Accrue the full prior invoice amount.  These are flat monthly
            #   service contracts or fixed utility bills — the current month
            #   will cost the same as the most recent invoice.
            #   NOTE: Gas (613210) is now in _METERED_UTILITY_ACCOUNTS and uses
            #   per-vendor daily-rate proration, same as electricity (613110).
            _is_metered_utility = (code in _eff_metered)  # 613110 elec, 613210/613120 gas

            if _is_metered_utility:
                # One accrual per vendor, using only their LATEST billing end date.
                #
                # RevLabs electricity example:
                #   Eversource  → latest end 12/31/25, invoice $24,450, 29-day cycle
                #   Hudson      → latest end 12/31/25, invoice $45,918, 29-day cycle
                #   Hudson old  → end 12/01/25, superseded by 12/31 invoice → skipped
                #
                # RevLabs gas example (per-invoice):
                #   National Grid meter 1 → end 12/22/25, $23,106.69 HVAC Delivery
                #   National Grid meter 2 → end 12/22/25, $204.32    EMGEN
                #   National Grid meter 3 → end 12/22/25, $732.26    TYGEN
                #   NRG Business Marketing → end 12/22/25, $18,951.27 HVAC Supply
                #   → 4 separate accrual lines
                #
                # Gas (613210) uses PER-INVOICE grouping — one line per meter/invoice.
                # Electricity (613110) uses PER-VENDOR grouping — all invoices from the
                # same vendor (e.g. multiple Eversource line items) combined into one line.

                if code in _eff_per_invoice:
                    # ── Gas: one accrual per meter, using each meter's OWN latest end date ──
                    #
                    # Mirrors the per-vendor electricity approach: every distinct
                    # (vendor, service) combination is treated as one meter.  Each meter
                    # finds its own most-recent billing end date independently — so a meter
                    # that hasn't been re-billed this cycle (e.g. EMGEN) is not silently
                    # dropped just because a different meter (e.g. TYGEN) has a newer invoice.
                    #
                    # Step 1: map each meter key → its latest end date
                    _meter_latest_end: Dict[tuple, date] = {}
                    for _ed, _grp in by_end.items():
                        for _g in _grp:
                            _mk = (_g[3], _g[4])  # (vendor_name, service_desc)
                            if _mk not in _meter_latest_end or _ed > _meter_latest_end[_mk]:
                                _meter_latest_end[_mk] = _ed

                    # Step 2: one candidate per meter at its own latest end date
                    for (_m_vname, _m_service), _m_latest_end in _meter_latest_end.items():
                        _m_invs = [
                            g for g in by_end[_m_latest_end]
                            if g[3] == _m_vname and g[4] == _m_service
                        ]
                        if not _m_invs:
                            continue
                        _inv_start   = min(g[0] for g in _m_invs)
                        _inv_end     = _m_latest_end
                        _inv_amt     = sum(g[2] for g in _m_invs)
                        _inv_vname   = _m_vname
                        _inv_service = _m_service

                        _v_uncovered = (month_end - _inv_end).days
                        if _v_uncovered <= 0:
                            continue
                        _vdays = max(1, (_inv_end - _inv_start).days + 1)  # +1: inclusive end date
                        if _v_uncovered > _vdays * 2.0:
                            continue
                        _vrate    = _inv_amt / _vdays
                        _vaccrual = _vrate * _v_uncovered
                        # Per-invoice utility meters: do NOT apply the $500 materiality floor.
                        # Each meter is a known, discrete recurring charge — a $200 EMGEN line
                        # is just as real as a $23K HVAC delivery line.  The materiality floor
                        # belongs in Layer 3 pattern detection, not here.
                        if _vaccrual < 1.0:
                            continue
                        # Build a readable label: "NATIONAL GRID — 82953-68006 HVAC Delivery"
                        _vendor_label = (
                            f'{_inv_vname} — {_inv_service}' if _inv_service
                            else (_inv_vname or acct.account_name)
                        )
                        _vdesc_line = (
                            f'Accrual {_period_label} — {_vendor_label} '
                            f'(gas proration: last invoice '
                            f'{_inv_start.strftime("%m/%d/%y")}'
                            f'-{_inv_end.strftime("%m/%d/%y")}, '
                            f'${_inv_amt:,.0f}/{_vdays}d = '
                            f'${_vrate:,.2f}/day × {_v_uncovered} days uncovered)'
                        )
                        candidates.append({
                            'account_code':   code,
                            'account_name':   acct.account_name,
                            'accrual_amount': _round(_vaccrual),
                            'source':         'invoice_proration',
                            'description':    _vdesc_line,
                            'vendor':         _vendor_label,
                            'daily_rate':     round(_vrate, 4),
                            'uncovered_days': _v_uncovered,
                            'period_days':    _vdays,
                            'invoice_total':  _round(_inv_amt),
                        })

                else:
                    # ── Electricity: one accrual per vendor, using their latest end date ──
                    # Algorithm:
                    #   1. Map vendor → their latest end date across all billing periods.
                    #   2. For each vendor, combine all their invoices at that latest end date.
                    #   3. Generate one candidate per vendor.

                    # Step 1: map vendor → their latest end date
                    _vendor_latest_end: Dict[str, date] = {}
                    for _ed, _grp in by_end.items():
                        for _g in _grp:
                            _vn = _g[3]  # vendor name (index 3)
                            if _vn not in _vendor_latest_end or _ed > _vendor_latest_end[_vn]:
                                _vendor_latest_end[_vn] = _ed

                    # Step 2 & 3: one accrual per vendor at their latest end date
                    for _vn, _v_latest_end in _vendor_latest_end.items():
                        _v_uncovered = (month_end - _v_latest_end).days
                        if _v_uncovered <= 0:
                            continue  # vendor's latest invoice already covers the month

                        # Combine all transactions for this vendor at their latest end date
                        _v_grp  = [g for g in by_end[_v_latest_end] if g[3] == _vn]
                        _vamt   = sum(g[2] for g in _v_grp)
                        _vstart = min(g[0] for g in _v_grp)
                        _vdays  = max(1, (_v_latest_end - _vstart).days + 1)  # +1: inclusive end date

                        # Sanity cap: don't extrapolate more than 2× the billing period
                        if _v_uncovered > _vdays * 2.0:
                            continue

                        _vrate    = _vamt / _vdays
                        _vaccrual = _vrate * _v_uncovered
                        if _vaccrual < materiality:
                            continue

                        _vendor_label = _vn if _vn else acct.account_name
                        _vdesc_line = (
                            f'Accrual {_period_label} — {_vendor_label} '
                            f'(electricity proration: last invoice '
                            f'{_vstart.strftime("%m/%d/%y")}'
                            f'-{_v_latest_end.strftime("%m/%d/%y")}, '
                            f'${_vamt:,.0f}/{_vdays}d = '
                            f'${_vrate:,.2f}/day × {_v_uncovered} days uncovered)'
                        )
                        candidates.append({
                            'account_code':   code,
                            'account_name':   acct.account_name,
                            'accrual_amount': _round(_vaccrual),
                            'source':         'invoice_proration',
                            'description':    _vdesc_line,
                            'vendor':         _vendor_label,
                            'daily_rate':     round(_vrate, 4),
                            'uncovered_days': _v_uncovered,
                            'period_days':    _vdays,
                            'invoice_total':  _round(_vamt),
                        })
            else:
                # All other accounts (water, sewer, HVAC contracts, janitorial, etc.):
                # Use the latest invoice to derive a monthly rate, then compound with
                # any prior-month auto-reversal (J-credit) so the accrued liability
                # builds correctly for semi-annual / quarterly billing cycles.
                #
                # Monthly contracts (billing_months ≈ 1): accrual = monthly_rate (no change).
                # Multi-month contracts (billing_months > 1, e.g. Water/Sewer billed
                #   semi-annually): accrual = J-credit reversal + monthly_rate.
                #   This grows each month until the real invoice arrives and the
                #   account net is non-zero (suppressed by the non-J-net guard below).
                group        = by_end[latest_end]
                total_amount = sum(g[2] for g in group)
                min_start    = min(g[0] for g in group)
                period_days  = max(1, (latest_end - min_start).days)

                # Derive billing period in months (round to nearest whole month).
                billing_months = max(1, round(period_days / 30.44))
                monthly_rate   = total_amount / billing_months

                # Compound logic: add prior-month auto-reversal (J-credit) to monthly rate
                # so the accrued liability builds each month until the real invoice arrives.
                #
                # Guard: if J-debits >= monthly_rate, a prior pipeline JE is already
                # in the GL for this account — skip to avoid double-accrual.
                _p1_j_dr = _j_debits(acct)
                if _p1_j_dr >= monthly_rate:
                    continue

                # Net J-credit = prior-month auto-reversal signal.
                # Always use it — the billing_months > 1 restriction was incorrectly
                # preventing compound on accounts (e.g. Water/Sewer) whose invoiced
                # period happens to parse as 1 month even though the pipeline needs
                # to compound multiple months of accrual between real invoices.
                _p1_j_cr = _net_j_credit(acct)
                accrual_amount = _p1_j_cr + monthly_rate

                _cmpd_note = (
                    f' — cumulative ${accrual_amount:,.0f} '
                    f'(${_p1_j_cr:,.0f} prior reversal + ${monthly_rate:,.0f}/mo)'
                    if _p1_j_cr > 0 else ''
                )
                accrual_desc = (
                    f'Accrual {_period_label} — {acct.account_name} '
                    f'(last invoice {min_start.strftime("%m/%d/%y")}'
                    f'-{latest_end.strftime("%m/%d/%y")}, '
                    f'${total_amount:,.0f} / {billing_months} mo'
                    f' = ${monthly_rate:,.0f}/mo){_cmpd_note}'
                )
                if accrual_amount >= materiality:
                    candidates.append({
                        'account_code':   code,
                        'account_name':   acct.account_name,
                        'accrual_amount': _round(accrual_amount),
                        'source':         'invoice_proration',
                        'description':    accrual_desc,
                        'daily_rate':     0.0,
                        'uncovered_days': uncovered,
                        'period_days':    period_days,
                        'invoice_total':  _round(total_amount),
                    })
            continue   # Don't also run payroll check for this account

        # ── PAYROLL PRORATION ─────────────────────────────────────────────────
        # Only applicable to accounts whose name suggests payroll.
        name_lower = (acct.account_name or '').lower()
        if not any(kw in name_lower for kw in _PAYROLL_NAME_KW):
            continue

        # Collect payroll runs: debit entries where description mentions payroll.
        payroll_runs: List[tuple] = []   # (run_date: date, amount: float)
        for txn in acct.transactions:
            amt = (txn.debit or 0) - (txn.credit or 0)
            if amt <= 0:
                continue
            combined = ((txn.remarks or '') + ' ' + (txn.description or '')).lower()
            if not any(kw in combined for kw in _PAYROLL_DESC_KW):
                continue
            run_date = _parse_single_date(txn.remarks or '')
            if run_date is None:
                run_date = _parse_single_date(txn.description or '')
            if run_date is None:
                # Fall back to the transaction's posted date
                run_date = txn.date if isinstance(txn.date, date) else None
            if run_date:
                payroll_runs.append((run_date, amt))

        if len(payroll_runs) < 2:
            continue   # Need ≥ 2 runs to infer pay period length

        payroll_runs.sort(key=lambda x: x[0])

        # Pay period length = gap between the two most-recent distinct run dates.
        # Group by date and sum amounts so we can identify the "main" payroll
        # runs vs. small off-cycle entries (e.g., a $1,554 mid-cycle run).
        dates_only = sorted({r[0] for r in payroll_runs})
        if len(dates_only) < 2:
            continue

        # Use the last-two-date gap but enforce a 13-day floor.
        # Off-cycle payroll entries (e.g., a small catch-up run mid-cycle)
        # can create 7-day gaps between payroll dates that make the detected
        # period half the true bi-weekly cycle.  13 days is safely below any
        # bi-weekly (14d) or semi-monthly (13-16d) schedule while filtering out
        # the 7-day false periods from off-cycle runs.
        raw_gap = (dates_only[-1] - dates_only[-2]).days
        pay_period_days = max(13, raw_gap)

        # Latest run date and total amount for that run (regular + OT combined).
        latest_run_date = dates_only[-1]
        latest_run_total = sum(amt for rd, amt in payroll_runs if rd == latest_run_date)

        # Days from last run to month-end = uncovered payroll days.
        uncovered = (month_end - latest_run_date).days
        if uncovered <= 0:
            continue

        daily_rate = latest_run_total / pay_period_days
        accrual    = daily_rate * uncovered

        if accrual < materiality:
            continue

        candidates.append({
            'account_code':   code,
            'account_name':   acct.account_name,
            'accrual_amount': _round(accrual),
            'source':         'invoice_proration',
            'description': (
                f'Payroll accrual — {acct.account_name}: '
                f'last run {latest_run_date.strftime("%m/%d/%y")} '
                f'(${latest_run_total:,.2f}/{pay_period_days}d = '
                f'${daily_rate:,.2f}/day x {uncovered} days uncovered)'
            ),
            'daily_rate':     round(daily_rate, 4),
            'uncovered_days': uncovered,
            'period_days':    pay_period_days,
            'invoice_total':  _round(latest_run_total),
        })

        continue   # payroll path handled — skip recurring-vendor check

    # ── PASS 2: Recurring vendor accruals ────────────────────────────────────
    # Detects expense accounts with recurring vendor invoices already in the GL
    # and accrues the current-month unbilled balance at the same rate.
    #
    # Invoices arrive throughout the entire month (not just in the first 5 days),
    # so no date restriction is applied to non-J debit transactions.
    #
    # Detection criteria — ALL must be true:
    #   1. Expense account (6xxx / 8xxx etc.)
    #   2. Prior-period history (beginning_balance >= $1) — confirms recurring spend
    #   3. At least one non-J debit (actual vendor invoice) totalling >= $1 in the period
    #   4. No J-type debits (would mean a prior pipeline JE is already in the GL)
    #   5. Not already handled by an earlier layer
    #
    # Additionally, Pattern B provides a stronger signal for accounts that have
    # been through the pipeline before: J-type credits (Yardi auto-reversals)
    # confirm a prior-month accrual existed. Pattern B is logged separately in
    # the description for auditability but does not change the accrual amount.
    #
    # Suppressed if J-type debits >= $1 (a prior pipeline JE already posted).

    period_month_start = date(month_end.year, month_end.month, 1)

    _already_coded = {c['account_code'] for c in candidates}

    for acct in gl_data.accounts:
        code = str(acct.account_code).strip()
        if not code or code[0] not in ('5', '6', '7', '8'):
            continue
        if code in _already_coded:
            continue   # already handled by date-range or payroll path

        # Must have some GL activity this period
        if not acct.transactions:
            continue

        # NOTE: beginning_balance is intentionally NOT checked here.
        # Yardi GL exports do not include a Balance Forward row for P&L/expense
        # accounts (6xxx/8xxx), so beginning_balance is always $0 for them.
        # The presence of non-J debits (actual vendor invoices) in the period
        # is sufficient signal — the beginning_balance guard permanently
        # blocked all expense accounts and has been removed.

        # Separate J-type (journal entries) from non-J (actual vendor invoices).
        # J credits = individual auto-reversal transactions — kept as a list so
        #   each reversal can generate its own accrual line (like elec/gas breakout).
        # J debits  = prior pipeline JE already in GL → suppress to avoid double-accrual.
        period_debits  = []    # non-J debit amounts — actual vendor invoices
        j_credit_txns  = []    # individual J-type credit txns (auto-reversals)
        j_debit_total  = 0.0   # J-type net debits (prior pipeline JE posted to GL)

        for txn in acct.transactions:
            ctrl_prefix = (txn.control or '').split('-')[0].upper()
            is_j    = (ctrl_prefix == 'J')
            txn_net = (txn.debit or 0) - (txn.credit or 0)

            if is_j:
                if txn_net < 0:
                    j_credit_txns.append(txn)    # credit J entry → auto-reversal
                elif txn_net > 0:
                    j_debit_total += txn_net      # debit J entry → pipeline JE in GL
                continue   # J entries never count as vendor invoices

            # Non-J debit → actual vendor invoice (any day of the month)
            if txn_net <= 0:
                continue
            period_debits.append(txn_net)

        # If a prior pipeline JE already debited this account, skip — don't double-accrue
        if j_debit_total >= 1.0:
            continue

        if not period_debits:
            continue

        invoice_total = sum(period_debits)
        if invoice_total < 1.0:
            continue

        # Require J-credits to match P-entries 1-for-1.
        #
        # Each J-credit is a Yardi auto-reversal of a prior pipeline accrual —
        # it proves the pipeline previously accrued that invoice and the real
        # invoice then arrived and netted out the accrual.  We re-accrue only
        # for those invoices.
        #
        # If P-entries > J-credits: extra P-entries are direct payables already
        # fully captured in the GL — accruing on top would double-count them.
        # Example: Admin Tenant Relations has 4 payments but only 1 J-reversal →
        #   skip entirely; the 3 non-reversed invoices are already in the GL.
        #
        # If J-credits > P-entries: atypical; could mean a reversal with no
        # matching new invoice yet — skip to avoid a phantom accrual.
        #
        # Only when counts match (e.g. Casella: 3P = 3J) does every invoice
        # have a confirmed prior-pipeline pairing → safe to re-accrue.
        if not j_credit_txns:
            continue
        if len(j_credit_txns) != len(period_debits):
            continue

        # Collect all non-J debits for this account.
        _invoice_lines = []
        for txn in acct.transactions:
            _ctrl = (txn.control or '').split('-')[0].upper()
            if _ctrl == 'J':
                continue
            _txn_amt = (txn.debit or 0) - (txn.credit or 0)
            if _txn_amt < 1.0:
                continue
            _inv_desc = (txn.description or '').strip()
            _invoice_lines.append((_txn_amt, _inv_desc))

        if not _invoice_lines:
            continue

        _total_accrual = sum(a for a, _ in _invoice_lines)
        if _total_accrual < 1.0:
            continue

        if code in _eff_per_invoice_accrual:
            # ── Individual mode ──────────────────────────────────────────────
            # One candidate per invoice line — preserves per-service-type detail
            # (e.g. Casella: separate lines for compactor, recycling, trash).
            for _line_amt, _line_desc in _invoice_lines:
                if _line_amt < 1.0:
                    continue
                candidates.append({
                    'account_code':   code,
                    'account_name':   acct.account_name,
                    'accrual_amount': _round(_line_amt),
                    'source':         'invoice_proration',
                    'description': (
                        f'Accrual {_period_label} — {acct.account_name}: '
                        f'${_line_amt:,.2f} — {_line_desc}'
                    ),
                    'daily_rate':     0.0,
                    'uncovered_days': 0,
                    'period_days':    0,
                    'invoice_total':  _round(_line_amt),
                })
        else:
            # ── Combined mode (default) ──────────────────────────────────────
            # One candidate with all invoice amounts itemised in the description
            # (e.g. Verizon: one line per phone collapsed into a single JE row).
            _vendor_header = (_invoice_lines[0][1]).split('(')[0].strip()
            if len(_invoice_lines) == 1:
                _detail = f'${_invoice_lines[0][0]:,.2f} — {_invoice_lines[0][1]}'
            else:
                _parts = [f'${a:,.2f} {d}' for a, d in _invoice_lines]
                _detail = ' | '.join(_parts)

            candidates.append({
                'account_code':   code,
                'account_name':   acct.account_name,
                'accrual_amount': _round(_total_accrual),
                'source':         'invoice_proration',
                'description': (
                    f'Accrual {_period_label} — {_vendor_header} ({acct.account_name}): '
                    f'{_detail}'
                ),
                'daily_rate':     0.0,
                'uncovered_days': 0,
                'period_days':    0,
                'invoice_total':  _round(_total_accrual),
            })

    return candidates


# detect_budget_gaps() removed May 2026 — retired, never called per CLAUDE.md.
# Budget gap accrual logic was removed; accruing to budget is not good practice.

# ── Layer 4 (runs before budget gap): Historical pattern detection ────────────

def detect_historical_recurring(gl_data, budget_data, period: str = '',
                                t12_result=None,
                                fiscal_year_start_month: int = 1,
                                kardin_records: Optional[List[Dict]] = None,
                                materiality: float = 2500.0,
                                layer3_exclude_accounts: Optional[List[str]] = None,
                                ) -> List[Dict[str, Any]]:
    """
    Identify recurring expense patterns using Budget Comparison YTD actual data.

    Primary method: BC YTD actual ÷ months elapsed (months before current period).
    This is more reliable than GL beginning balance because BC YTD actual reflects
    clean closed prior-period activity without current-month noise.

    Falls back to GL beginning_balance ÷ months_elapsed when BC YTD is unavailable
    for an account.

    January fallback (months_elapsed = 0):
      - With T12: uses December actual from the 12-Month Statement (most accurate).
      - Without T12: uses annual_budget ÷ 12 for accounts with annual budget ≥ $60K.
    February and later: uses BC YTD actual ÷ months_elapsed (T12 not used).

    Args:
        t12_result: Optional T12Result from parsers.yardi_t12.  When provided and
                    period is January, December actuals replace the annual÷12 fallback.

    Returns list of dicts: account_code, account_name, estimated_amount, source='historical'
    """
    candidates = []

    if not gl_data or not hasattr(gl_data, 'accounts'):
        return candidates

    _period_label = _fmt_period(period)

    # Determine current month number — prefer explicit period arg, fall back to GL metadata
    period_str = period or (
        getattr(gl_data.metadata, 'period', '') if hasattr(gl_data, 'metadata') else ''
    )
    _MONTH_MAP_H = {
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12,
    }
    month_num = 0
    for abbr, num in _MONTH_MAP_H.items():
        if abbr in period_str:
            month_num = num
            break

    # months_elapsed = full fiscal months closed before the current period.
    # For calendar-year properties (fiscal_year_start_month=1): Jan=0, Feb=1 … Dec=11.
    # For non-January fiscal years: rebase month_num to the fiscal year so that,
    # e.g., a July FY-start property in December has fiscal month 6 → months_elapsed=5.
    _fy_start = int(fiscal_year_start_month or 1)
    if _fy_start < 1 or _fy_start > 12:
        _fy_start = 1
    if month_num > 0:
        fiscal_month_num = (month_num - _fy_start) % 12 + 1
    else:
        fiscal_month_num = 0
    months_elapsed = fiscal_month_num - 1 if fiscal_month_num > 0 else 0

    # Build YTD actual and budget lookups from Budget Comparison data
    ytd_actual_by_code: Dict[str, float] = {}
    budget_by_code: Dict[str, Any] = {}
    if budget_data:
        budget_items = (
            budget_data if isinstance(budget_data, list)
            else getattr(budget_data, 'line_items', [])
        )
        for item in budget_items:
            if isinstance(item, dict):
                bcode = str(item.get('account_code', '') or '').strip()
                ytd_a = abs(float(item.get('ytd_actual', 0) or 0))
            else:
                bcode = str(getattr(item, 'account_code', '') or '').strip()
                ytd_a = abs(float(getattr(item, 'ytd_actual', 0) or 0))
            if bcode:
                ytd_actual_by_code[bcode] = ytd_a
                budget_by_code[bcode] = item

    # Build Kardin annual totals for accounts NOT already in budget_by_code.
    # Kardin records carry m1–m12 monthly columns; annual = sum of all 12 months.
    # This fills the gap for accounts that appear in Kardin but not in the Yardi BC
    # (e.g. semi-annual accounts like 613310 Water/Sewer with no YTD BC activity).
    kardin_annual_by_code: Dict[str, Dict] = {}   # {code: {'annual': float, 'name': str}}
    for _kr in (kardin_records or []):
        _kc = str(_kr.get('account_code', '') or '').strip()
        if not _kc or _kc in budget_by_code:
            continue   # already have BC data for this account — BC takes priority
        _k_annual = sum(
            abs(float(_kr.get(f'm{i}', 0) or 0)) for i in range(1, 13)
        )
        if _k_annual < 1:
            continue
        _k_name = str(_kr.get('description', '') or _kr.get('account_name', '') or _kc).strip()
        # Keep the highest-annual record if the same code appears multiple times
        if _kc not in kardin_annual_by_code or _k_annual > kardin_annual_by_code[_kc]['annual']:
            kardin_annual_by_code[_kc] = {'annual': _k_annual, 'name': _k_name}

    _gl_seen_codes: set    = set()   # every expense account visited in the GL loop
    _gl_handled_codes: set = set()   # subset that actually produced a candidate

    # Build the Layer 3 exclusion set once (account codes to completely skip)
    _l3_excl: set = set(str(c).strip() for c in (layer3_exclude_accounts or []))

    for acct in gl_data.accounts:
        code = str(acct.account_code).strip()
        # Only expense accounts — uses per-property COA config (defaults to 5/6/7/8xxxxx)
        if not is_expense_account(code):
            continue
        _gl_seen_codes.add(code)

        # Skip accounts explicitly excluded from Layer 3 auto-accrual (e.g. discretionary
        # or irregular spend that Layer 3 would mis-classify as a recurring contract).
        if code in _l3_excl:
            _gl_handled_codes.add(code)  # mark as handled so budget-gap doesn't fire either
            continue

        # Partial-coverage detection: if some (but not enough) invoices have already
        # posted this period, don't suppress entirely — compute the expected monthly
        # amount and accrue only the shortfall.
        #
        # _gl_partial_offset > 0 means we'll subtract it from est_monthly below.
        # The coverage threshold is 25%: if GL has ≥ 25% of expected → fully covered,
        # skip.  If GL has < 25% → partial posting (e.g. one of three Casella invoices
        # arrived), generate top-up for the remainder.
        #
        # Strip J-entries that perfectly cancel each other from the partial-coverage
        # signal.  Use K/P/C-only net so a $130K J-debit + $130K J-credit pair
        # doesn't distort whether a real invoice has already posted this period.
        _j_dr_total = _j_debits(acct)
        _j_cr_total = _j_credits(acct)
        _kpc_net    = acct.net_change - (_j_dr_total - _j_cr_total)  # K/P/C only
        _gl_partial_offset = _kpc_net if _kpc_net > 0.01 else 0.0

        # ── Compound accrual: prior-month auto-reversal detected ──────────────
        # _net_j_credit strips paired J-entries (e.g. $130K bill + $130K reversal)
        # so only the unmatched open reversal (e.g. $48K) is used as the base.
        # _non_j_net uses the same K/P/C-only net (_kpc_net) to detect real bills.
        _j_cr = _net_j_credit(acct)
        # Fallback for Yardi auto-reversals that use non-'J' control codes
        # (e.g. 'AJ', 'RJ') or store the reversal as a negative debit instead
        # of a positive credit.  In either case _net_j_credit() returns 0.
        # Guard: _kpc_net (K/P/C-only net, computed above) will also be a large
        # negative on the expense account — that unexplained net credit IS the
        # reversal. Use abs(_kpc_net) so the compound path can fire correctly.
        if _j_cr < 500 and _kpc_net < -500:
            _j_cr = abs(_kpc_net)
        # 8xxxxx accounts (interest, other income/expense) are flat monthly charges
        # handled by Layer 1b (Berkadia) — never compound them.
        if _j_cr > 500 and not code.startswith('8'):
            _non_j_net = _kpc_net  # K/P/C net already computed above

            # ── Parse billing start date from J-credit description ─────────────
            # Done here (once) so it drives both the monthly rate computation
            # (Priority 1 below) and the date-range description string built later.
            import re as _re_c
            import calendar as _cal_c
            _start_date_str = ''
            _start_date_obj = None
            # Find the J-credit transaction whose amount best matches _j_cr.
            # When multiple J-entries exist (e.g. a $130K pair that cancels
            # plus a $48K open balance), the paired entries carry the original
            # billing start date while the net open entry carries the current
            # billing period start.  Taking the first J-credit would pick the
            # older date (e.g. 03/25/24) → huge month count → rate below floor.
            # Matching on amount (diff closest to zero) picks the right entry.
            _best_jt_cr = None
            _best_jt_diff = float('inf')
            for _jt in getattr(acct, 'transactions', []):
                if _jt.credit <= 0:
                    continue
                if not str(getattr(_jt, 'control', '') or '').upper().startswith('J'):
                    continue
                _diff = abs(_jt.credit - _j_cr)
                if _diff < _best_jt_diff:
                    _best_jt_diff = _diff
                    _best_jt_cr = _jt
            if _best_jt_cr is not None:
                _txt = (f"{getattr(_best_jt_cr, 'description', '') or ''} "
                        f"{getattr(_best_jt_cr, 'remarks', '') or ''}")
                _dm = _re_c.search(r'\b(\d{1,2}/\d{1,2}/\d{2,4})\b', _txt)
                if _dm:
                    _start_date_str = _dm.group(1)
                    from datetime import datetime as _dt_c
                    for _fmt in ('%m/%d/%Y', '%m/%d/%y'):
                        try:
                            _start_date_obj = _dt_c.strptime(_start_date_str, _fmt)
                            break
                        except ValueError:
                            pass

            # ── Monthly rate: billing-period actual → BC → Kardin → fallback ──
            # Priority 1: j_cr ÷ full months in the billing period (most accurate).
            # Counts only full months — skips the prorated start month when billing
            # began after the 15th (e.g. a 9/23 start means Oct is the first full
            # month; Oct/Nov/Dec = 3 months → $48,068 ÷ 3 = $16,023/mo).
            # This avoids Kardin "annual" figures that represent a partial contract
            # year (e.g. 9-month annual ÷ 12 = ¾ of the true monthly rate).
            _has_budget_rate = False
            _mthly_rt = 0.0
            if _start_date_obj is not None and month_num > 0:
                _yr_m_rt = _re_c.search(r'\d{4}', period_str)
                if _yr_m_rt:
                    _p_yr_rt = int(_yr_m_rt.group())
                    # Skip the partial start month if billing began after mid-month
                    _cfm = _start_date_obj.month + (1 if _start_date_obj.day > 15 else 0)
                    _cfy = _start_date_obj.year
                    if _cfm > 12:
                        _cfm, _cfy = 1, _cfy + 1
                    # Walk forward counting full months before the current close month
                    _full_months = 0
                    _m_i, _y_i = _cfm, _cfy
                    while (_y_i, _m_i) < (_p_yr_rt, month_num):
                        _full_months += 1
                        _m_i += 1
                        if _m_i > 12:
                            _m_i, _y_i = 1, _y_i + 1
                    if _full_months >= 1:
                        _mthly_rt = _round(_j_cr / _full_months)
                        _has_budget_rate = True  # derived from actual billing data

            # Priority 2: BC annual ÷ 12
            if not _has_budget_rate:
                _bi = budget_by_code.get(code)
                if _bi is not None:
                    _bi_annual = abs(float(
                        (_bi.get('annual', 0) if isinstance(_bi, dict)
                         else getattr(_bi, 'annual', 0)) or 0
                    ))
                    if _bi_annual >= 1:
                        _mthly_rt = _round(_bi_annual / 12)
                        _has_budget_rate = True
                    else:
                        _mthly_rt = _round(_j_cr / max(months_elapsed, 1))
                # Priority 3: Kardin annual ÷ 12
                elif code in kardin_annual_by_code:
                    _k_annual = kardin_annual_by_code[code]['annual']
                    if _k_annual >= 1:
                        _mthly_rt = _round(_k_annual / 12)
                        _has_budget_rate = True
                    else:
                        _mthly_rt = _round(_j_cr / max(months_elapsed, 1))
                # Priority 4: last resort
                else:
                    _mthly_rt = _round(_j_cr / max(months_elapsed, 1))

            # In January (months_elapsed=0) only compound when a reliable rate
            # is available.  Without one the fallback j_cr÷1 = j_cr would double
            # the accrual (compound = 2×j_cr). Fall through to Jan Path A/B.
            _run_compound = months_elapsed >= 1 or _has_budget_rate
            if _run_compound:
                if _mthly_rt >= materiality:  # materiality floor
                    # Determine billing period length from the start date already
                    # parsed from the J-credit description.  >45 days back from
                    # period start = quarterly or longer → compound accrual.
                    # No parseable date → monthly-only (safe default).
                    _multi_period = False
                    if _start_date_obj is not None and month_num:
                        _yr_m_mp = _re_c.search(r'\d{4}', period_str)
                        if _yr_m_mp:
                            from datetime import date as _date_mp
                            _p_yr_mp = int(_yr_m_mp.group())
                            _period_start_mp = _date_mp(_p_yr_mp, month_num, 1)
                            _multi_period = (
                                (_period_start_mp - _start_date_obj.date()).days > 120
                            )

                    if _multi_period:
                        # Quarterly / semi-annual: re-establish accumulated prior
                        # obligation plus current month.  Only suppress when a
                        # payment beyond the reversal signals the prior period is
                        # already covered by an actual invoice.
                        _payment_beyond_reversal = max(0.0, _non_j_net - _j_cr)
                        if _payment_beyond_reversal >= _mthly_rt * 0.25:
                            continue  # prior period settled + new invoice posted
                        _accrual_amt  = _round(_j_cr + _mthly_rt)
                        _accrual_type = 'compound'
                    else:
                        # Monthly billing cycle: prior period clears each month.
                        # Suppress if the actual invoice has already posted.
                        if _non_j_net >= _mthly_rt * 0.25:
                            continue  # invoice already in GL
                        _accrual_amt  = _round(_mthly_rt)
                        _accrual_type = 'monthly'

                    if _accrual_amt >= 250:
                        # Build period-end string for description
                        _period_end_str = ''
                        if month_num:
                            _yr_m = _re_c.search(r'\d{4}', period_str)
                            if _yr_m:
                                _p_yr   = int(_yr_m.group())
                                _p_last = _cal_c.monthrange(_p_yr, month_num)[1]
                                _period_end_str = f'{month_num}/{_p_last}/{_p_yr}'
                        if _accrual_type == 'compound':
                            if _start_date_str and _period_end_str:
                                _entry_desc = (
                                    f'Accrual {_period_label} — {acct.account_name} '
                                    f'({_start_date_str}-{_period_end_str}; '
                                    f'${_j_cr:,.2f} prior accrual reversed + '
                                    f'${_mthly_rt:,.2f}/mo est.)'
                                )
                            else:
                                _entry_desc = (
                                    f'Accrual {_period_label} — {acct.account_name} '
                                    f'(multi-period: ${_j_cr:,.0f} prior accrual '
                                    f'reversed + ${_mthly_rt:,.0f}/mo est.)'
                                )
                        else:
                            _entry_desc = (
                                f'Accrual {_period_label} — {acct.account_name} '
                                f'(${_mthly_rt:,.0f}/mo est.)'
                            )
                        candidates.append({
                            'account_code':     code,
                            'account_name':     acct.account_name,
                            'estimated_amount': _accrual_amt,
                            'ytd_prior':        _j_cr,
                            'months_prior':     months_elapsed,
                            'source':           'historical',
                            'description':      _entry_desc,
                        })
                        _gl_handled_codes.add(code)
                continue  # compound path evaluated — skip BC YTD normal path

        # ── January fallback: no prior-year YTD data available ────────────────
        # Prefer T12 December actual when uploaded; otherwise use annual budget ÷ 12.
        # This prevents the historical layer from going dark in the first month
        # of the fiscal year when BC YTD and GL beginning balance are both zero.
        if months_elapsed < 1:
            # Path A: T12 December actual (more accurate than annual/12)
            # Only use T12 if December shows meaningful activity (≥$5K).
            # For semi-annual / quarterly accounts (e.g. 613310 Water/Sewer,
            # 631110 Elevator) December T12 will be $0 — fall through to Path B
            # so the Kardin annual÷12 estimate is used instead.
            if t12_result is not None and hasattr(t12_result, 'prior_month'):
                dec_actual = abs(t12_result.prior_month(code, 1))  # Dec = prior to Jan
                if dec_actual >= materiality:
                    # Same GL-activity gate as Feb+ normal path and Jan Path B:
                    # if a real K/P/C invoice has already posted this period and
                    # covers ≥ 25% of the December T12 amount, treat as fully
                    # covered and skip.  This prevents T12 from firing on top of
                    # an invoice that is already in the GL.
                    if _gl_partial_offset >= dec_actual * 0.25:
                        _gl_handled_codes.add(code)
                        continue  # real invoice covers ≥ 25% of expected
                    _t12_accrual_amt = _round(max(0.0, dec_actual - _gl_partial_offset))
                    if _t12_accrual_amt < 250:
                        _gl_handled_codes.add(code)
                        continue
                    _t12_partial_note = (
                        f' (partial top-up — ${_gl_partial_offset:,.0f} already in GL)'
                        if _gl_partial_offset > 0 else ''
                    )
                    candidates.append({
                        'account_code': code,
                        'account_name': acct.account_name,
                        'estimated_amount': _t12_accrual_amt,
                        'ytd_prior': dec_actual,
                        'months_prior': 1,
                        'source': 'historical',
                        'description': (
                            f'Accrual {_period_label} — {acct.account_name} '
                            f'(historical — Dec actual ${dec_actual:,.0f} per T12'
                            f'{_t12_partial_note})'
                        ),
                    })
                    _gl_handled_codes.add(code)
                    continue  # T12 gave a good signal — skip annual/12 fallback
                # Dec actual below materiality floor (semi-annual / quarterly billing) — fall through
                # to Path B so the Kardin annual÷12 estimate still fires.

            # Path B: annual budget ÷ 12 (no T12 available)
            if code not in budget_by_code:
                continue
            bi = budget_by_code[code]
            if isinstance(bi, dict):
                # A-11: Try multiple key names — BC parser may export 'annual_budget'
                # or 'ytd_budget' instead of 'annual' depending on export version.
                # Silent 0 here suppresses ALL January Layer 3 accruals for the account.
                bi_annual = abs(float(
                    bi.get('annual') or bi.get('annual_budget') or bi.get('ytd_budget') or 0
                ))
            else:
                bi_annual = abs(float(
                    getattr(bi, 'annual', None) or getattr(bi, 'annual_budget', None) or 0
                ))

            if bi_annual < 1:
                continue

            est_monthly = bi_annual / 12
            if est_monthly < materiality:
                continue

            # Partial-coverage check (Jan path B)
            if _gl_partial_offset >= est_monthly * 0.25:
                continue  # GL has ≥25% of expected — treat as fully covered
            _accrual_amt = est_monthly - _gl_partial_offset
            if _accrual_amt < 250:
                continue
            _partial_note = (f' (partial top-up — ${_gl_partial_offset:,.0f} already in GL)'
                             if _gl_partial_offset > 0 else '')
            candidates.append({
                'account_code': code,
                'account_name': acct.account_name,
                'estimated_amount': _round(_accrual_amt),
                'ytd_prior': 0.0,
                'months_prior': 0,
                'source': 'historical',
                'description': (
                    f'Accrual {_period_label} — {acct.account_name} '
                    f'(historical est. ${est_monthly:,.0f}/mo, '
                    f'annual budget ${bi_annual:,.0f} ÷ 12'
                    f'{_partial_note})'
                ),
            })
            _gl_handled_codes.add(code)
            continue

        # ── Feb+ normal path: BC YTD ÷ months elapsed ─────────────────────────
        # Try BC YTD actual first; fall back to GL beginning balance
        ytd_prior = ytd_actual_by_code.get(code, 0.0)
        use_gl_fallback = False
        if ytd_prior < 100:
            ytd_prior = abs(acct.beginning_balance)
            use_gl_fallback = True

        if ytd_prior < 100:
            continue

        # Cross-reference against budget: zero budget everywhere = likely discontinued
        if code in budget_by_code:
            bi = budget_by_code[code]
            if isinstance(bi, dict):
                bi_budget = bi.get('ptd_budget', 0) or 0
                bi_annual = bi.get('annual', 0) or 0
            else:
                bi_budget = getattr(bi, 'ptd_budget', 0) or 0
                bi_annual = getattr(bi, 'annual', 0) or 0

            if abs(bi_budget) < 1 and abs(bi_annual) < 1:
                continue  # Zero budget everywhere — likely discontinued

        # Estimate monthly amount from YTD ÷ months elapsed
        est_monthly = ytd_prior / months_elapsed

        # Only flag if estimated monthly >= materiality (recurring expense threshold)
        if est_monthly >= materiality:
            # Partial-coverage check: if some invoices already in GL but < 25% of
            # expected monthly, accrue the shortfall instead of the full amount.
            # This catches vendors like Casella where multiple invoices arrive
            # throughout the month and only some post before the close export.
            if _gl_partial_offset >= est_monthly * 0.25:
                continue  # GL has ≥25% of expected — treat as fully covered
            _accrual_amt = est_monthly - _gl_partial_offset
            if _accrual_amt < 250:
                continue
            _partial_note = (f' (partial top-up — ${_gl_partial_offset:,.0f} already in GL)'
                             if _gl_partial_offset > 0 else '')
            source_note = 'BC YTD' if not use_gl_fallback else 'GL YTD (est.)'
            candidates.append({
                'account_code': code,
                'account_name': acct.account_name,
                'estimated_amount': _round(_accrual_amt),
                'ytd_prior': ytd_prior,
                'months_prior': months_elapsed,
                'source': 'historical',
                'description': (
                    f'Accrual {_period_label} — {acct.account_name} '
                    f'(historical avg ${est_monthly:,.0f}/mo, '
                    f'{source_note} ${ytd_prior:,.0f} ÷ {months_elapsed} mo'
                    f'{_partial_note})'
                ),
            })
            _gl_handled_codes.add(code)

    # Budget-only fallback removed (May 2026).
    # Generating accruals for accounts with no GL activity, based solely on
    # Kardin/BC budget figures, is accruing to budget — not good practice.
    # Accounts like water/sewer or elevator that bill semi-annually but have
    # no current-period GL entry should be handled via the one-off accruals
    # table (the operator knows when those bills are due).

    return candidates


# ── Payroll bonus detection ──────────────────────────────────────────────────

def detect_payroll_bonus_accrual(
    gl_data,
    kardin_records: List[Dict],
    period_month: int,
) -> List[Dict[str, Any]]:
    """
    Generate monthly bonus accrual entries for payroll accounts.

    Business rule
    -------------
    The annual engineering and admin bonuses are paid in January and July
    but should be expensed evenly across all 12 months.  Kardin reflects
    this intent — the two payment months carry higher values while the
    remaining months carry only base payroll.

    Monthly bonus accrual = (Kardin annual ÷ 12) − standard_month
      where standard_month = min(M1..M12) for the bonus-inclusive row.

    The accrual is suppressed if the GL net_change for the period already
    equals or exceeds the monthly average (the actual bonus payment is in
    the GL — no separate accrual needed).

    Args:
        gl_data:        GLParseResult from yardi_gl parser
        kardin_records: List of dicts from parsers.kardin_budget.parse()
        period_month:   Integer month of the reporting period (1=Jan … 12=Dec)

    Returns:
        List of candidate dicts (same shape as budget_gap candidates) with
        source='bonus_accrual'.
    """
    results: List[Dict[str, Any]] = []

    if not gl_data or not kardin_records or not period_month:
        return results

    # Build GL net_change lookup for payroll accounts
    gl_net: dict = {}
    for acct in (gl_data.accounts if hasattr(gl_data, 'accounts') else []):
        code = str(acct.account_code).strip()
        if code in PAYROLL_BONUS_ACCOUNTS:
            gl_net[code] = acct.net_change

    for acct_code, config in PAYROLL_BONUS_ACCOUNTS.items():
        keywords = [k.lower() for k in config['kardin_keywords']]

        # Find Kardin rows for this account that include the bonus component
        bonus_rows = [
            r for r in kardin_records
            if str(r.get('account_code', '') or '').strip() == acct_code
            and any(kw in (r.get('description', '') or '').lower() for kw in keywords)
        ]
        if not bonus_rows:
            continue

        # Sum annual and all monthly amounts across matching rows
        annual = sum(float(r.get('m_total', 0) or 0) for r in bonus_rows)
        if annual <= 0:
            continue

        monthly_avg = annual / 12.0

        # Standard month = minimum Kardin monthly value (non-payment months)
        all_monthly: List[float] = []
        for r in bonus_rows:
            for m in range(1, 13):
                val = float(r.get(f'M{m}', 0) or 0)
                if val > 0:
                    all_monthly.append(val)
        if not all_monthly:
            continue
        standard_monthly = min(all_monthly)

        monthly_bonus = monthly_avg - standard_monthly

        # Skip if not material (< $100)
        if monthly_bonus < 100.0:
            continue

        # Check current-period GL activity
        net = gl_net.get(acct_code, 0.0)

        # Suppress in payment months: GL already ≥ monthly average
        # (the actual bonus payment is in the GL — no accrual needed)
        if net >= monthly_avg:
            continue

        results.append({
            'account_code':    acct_code,
            'account_name':    config['label'],
            'estimated_amount': _round(monthly_bonus),
            'source':          'bonus_accrual',
            'confidence':      'high',
            'description': (
                f'Monthly bonus accrual — {config["label"]}: '
                f'Kardin annual ${annual:,.2f} / 12 = ${monthly_avg:,.2f}/mo avg; '
                f'standard month ${standard_monthly:,.2f}; '
                f'bonus component ${monthly_bonus:,.2f}/mo'
            ),
        })

    return results


# ── Build JE lines from all sources ──────────────────────────────────────────

def build_accrual_entries(nexus_data: list, period: str = '',
                          property_name: str = '',
                          status_filter: list = None,
                          gl_data=None, budget_data=None,
                          period_month_end: Optional[date] = None,
                          manual_accruals: Optional[List[Dict]] = None,
                          tenant_utility_rows: Optional[List[Dict]] = None,
                          kardin_records: Optional[List[Dict]] = None,
                          bonus_overrides: Optional[Dict[str, float]] = None,
                          loan_data=None,
                          re_tax_bill_amount: float = 0.0,
                          re_tax_payment_months=None,
                          t12_result=None,
                          gl_activity_log: Optional[List[Dict]] = None,
                          receivable_detail=None,
                          ledger_release_accounts: Optional[set] = None,
                          payroll_accounts: Optional[List[str]] = None,
                          insurance_policies: Optional[List[Dict]] = None,
                          periodic_contract_accounts: Optional[dict] = None,
                          metered_utility_accounts: Optional[List[str]] = None,
                          per_invoice_utility_accounts: Optional[List[str]] = None,
                          per_invoice_accrual_accounts: Optional[List[str]] = None,
                          accrual_materiality_floor: float = 2500.0,
                          fiscal_year_start_month: int = 1,
                          layer3_exclude_accounts: Optional[List[str]] = None,
                          ) -> List[Dict[str, Any]]:
    """
    Build accrual journal entry lines from four sources (in priority order):

      Layer 0: Manual overrides — user-supplied amounts for accounts that
               cannot be auto-calculated (e.g., semi-annual water/sewer bills)
      Layer 1: Nexus pending invoices (AP-side, deduped against GL)
      Layer 2: Invoice-period proration (billing date ranges in GL descriptions).
               For utility accounts (613/614 codes): prorates by day (daily rate
               × uncovered days). For all other services: accrues the full
               invoice amount — assumes current month will match prior billing.
      Layer 3: Historical recurring — BC YTD actual ÷ months elapsed.
               Fires when an expense account had prior-period activity but is
               silent this month. Skipped in January (no prior data).
      Layer 4: Payroll bonus accruals — monthly bonus component for engineering
               and admin payroll accounts. Driven by user-entered annual amounts
               (bonus_overrides) or Kardin-derived amounts as fallback.

    First-layer-wins: an account claimed by an earlier layer is skipped by all
    later layers. If multiple layers would have fired for the same account, a
    review_flag=True / review_sources=[...] is added to the first-layer DR line
    so the reviewer knows additional signals were also detected.

    Manual overrides take absolute priority and suppress all automated layers
    for the same account.

    Args:
        nexus_data:        List of invoice dicts from Nexus parser
        period:            Accounting period string (e.g., 'Mar-2026')
        property_name:     Property name for the JE header
        status_filter:     Invoice statuses to include (default: all)
        gl_data:           GLParseResult — required for Layers 2-4
        budget_data:       BC rows — required for Layer 3 (historical recurring)
        period_month_end:  Override for the last calendar day of the reporting
                           month (date object).  If None, derived from ``period``
                           or gl_data.metadata.period automatically.
        manual_accruals:   List of dicts for user-supplied accrual amounts::

                               [{
                                   'account_code': '613310',
                                   'account_name': 'Utilities-Water/Sewer',
                                   'amount':        16635.75,   # semi-annual invoice / 6
                                   'description':   'Water/sewer semi-annual invoice $99,814.50 / 6 months',
                               }, ...]

                           Amount is the *monthly* accrual to post.  Description
                           should note the invoice amount and divisor so the
                           reviewer can verify.  Accounts in manual_accruals are
                           excluded from all automated layers.

        gl_activity_log:   Optional mutable list.  When provided, one dict is
                           appended for each account the GL-activity gate
                           suppresses (net_change ≥ $500 in the current period).
                           Each dict has keys:
                               account_code, account_name, ptd_amount (abs)
                           Caller can display these as a gut-check list so the
                           user can verify that existing GL postings are correct
                           before uploading the pipeline JEs to Yardi.

    Returns:
        List of JE line dicts with keys:
          je_number, line, date, account_code, account_name,
          description, reference, debit, credit, vendor, invoice_number, source
    """
    invoices = nexus_data if isinstance(nexus_data, list) else []

    # Status filtering — parser already applies _INCLUDE_STATUSES by default
    # when parsing from the full Nexus Invoice Detail export.  This secondary
    # filter catches any records passed in via other paths (e.g. test fixtures)
    # and respects an explicit status_filter override if provided.
    _default_statuses = {'pending', 'in progress', 'pending approval',
                         'in yardi', 'submitted for payment', 'completed'}
    _filter_set = {s.lower() for s in status_filter} if status_filter else _default_statuses
    invoices = [inv for inv in invoices
                if (inv.get('invoice_status', '') or '').strip().lower() in _filter_set
                or not (inv.get('invoice_status', '') or '').strip()]
    # Note: invoices with no status field are passed through (e.g. manual_accruals)

    # Build GL lookup for Layer 1 deduplication
    gl_lookup = _build_gl_invoice_lookup(gl_data) if gl_data else {'by_reference': {}, 'by_control': {}}

    je_lines = []
    je_num = 1

    # ── Layer 0: Manual accrual overrides ──────────────────────────────────────
    # User-supplied amounts for accounts the engine cannot auto-calculate
    # (e.g., semi-annual water/sewer billing where the invoice amount is known
    # to the property manager but cannot be reliably derived from GL data).
    _manual_accounts: set = set()
    for override in (manual_accruals or []):
        acct_code = str(override.get('account_code', '') or '').strip()
        acct_name = str(override.get('account_name', '') or acct_code)
        amount    = float(override.get('amount', 0) or 0)
        desc      = str(override.get('description', '') or
                        f'Manual accrual — {acct_name}')
        if not acct_code:
            continue

        # Register the account as manually handled BEFORE the amount check so
        # that app.py's dedup pattern (amount=0, non-empty account_code) correctly
        # suppresses Layers 1-4 for this account even when no JE is being generated.
        _manual_accounts.add(acct_code)

        if amount <= 0:
            continue  # account registered for dedup; no JE generated

        # ── Compound accrual + real-invoice guard ────────────────────────────
        #
        # Semi-annual / irregular bills (water/sewer, elevator, etc.) accrue a
        # GROWING liability each month until the real invoice arrives:
        #
        #   Month 1: $20K  (1 × monthly_rate)
        #   Month 2: $40K  (prior $20K reversed + new $20K)
        #   Month 3: $60K  (prior $40K reversed + new $20K)
        #   ...
        #   Month 6: real $120K invoice posts → suppress
        #
        # Mechanics: Yardi auto-reverses the prior month's J-type accrual at
        # the start of the new period.  The J-credit that appears in the current
        # GL IS that reversal — its absolute value equals the prior month's
        # accrual.  Compound accrual = j_credits + monthly_rate.
        #
        # Guard: suppress only when non-J-type (K=check, P=payable, C=charge)
        # net activity >= monthly_rate — that signals the real invoice posted.
        # Using non-J-net (not total net_change) avoids false suppression from
        # the auto-reversal credit which is always present in payment months.
        _man_gl_acct = None
        _man_net_change = 0.0
        if gl_data and hasattr(gl_data, 'accounts'):
            for _mga in gl_data.accounts:
                if str(_mga.account_code).strip() == acct_code:
                    _man_gl_acct = _mga
                    _man_net_change = float(getattr(_mga, 'net_change', 0) or 0)
                    break

        _man_j_cr  = _net_j_credit(_man_gl_acct)  # net prior-month accrual reversal (cancels paired J-entries)
        _man_j_dr  = _j_debits(_man_gl_acct)    # any J-debits already posted this period
        _man_non_j = _man_net_change - (_man_j_dr - _man_j_cr)   # K/P/C-type net

        if _man_non_j >= amount:
            # Real invoice posted and covers at least one month — suppress.
            # Account stays in _manual_accounts so Layers 1-4 don't pile on.
            continue

        # Compound: add this month's slice on top of the *net* prior-month reversal.
        # Subtract any partial real-invoice activity (_man_non_j) from the reversal
        # so we don't double-accrue the portion already covered by a posted invoice.
        _net_reversal    = max(0.0, _man_j_cr - _man_non_j)   # reversal gap not yet offset
        _compound_amount = _net_reversal + amount
        _compound_note   = (f' — cumulative ${_compound_amount:,.0f} '
                            f'(${_net_reversal:,.0f} prior net + ${amount:,.0f}/mo)'
                            if _net_reversal > 0 else '')

        je_id    = f'MAN-{je_num:04d}'
        je_desc  = desc + _compound_note
        je_debit = _round(_compound_amount)
        je_lines.append({
            'je_number':      je_id,
            'line':           1,
            'date':           '',
            'account_code':   acct_code,
            'account_name':   acct_name,
            'description':    je_desc,
            'reference':      'MANUAL',
            'debit':          je_debit,
            'credit':         0,
            'vendor':         '[Manual Override]',
            'invoice_number': '',
            'source':         'manual',
            'confidence':     'high',
        })
        je_lines.append({
            'je_number':      je_id,
            'line':           2,
            'date':           '',
            'account_code':   AP_ACCRUAL_ACCOUNT,
            'account_name':   AP_ACCRUAL_NAME,
            'description':    je_desc,
            'reference':      'MANUAL',
            'debit':          0,
            'credit':         je_debit,
            'vendor':         '[Manual Override]',
            'invoice_number': '',
            'source':         'manual',
            'confidence':     'high',
        })
        je_num += 1

    # ── Tenant utility billing (meter read JE) ─────────────────────────────────
    # Revenue side of the utility accrual: ensures NOI is not understated while
    # the expense proration (Layer 2) accrues the full building bill.
    #
    # Two modes:
    #   a) Actual per-tenant amounts (tenant_utility_rows provided by sidebar):
    #      One DR 131100 / CR 440500 line per tenant for electric.
    #      One DR 131100 / CR 440700 line per tenant for gas.
    #   b) Budget aggregate (no rows provided, account has no GL activity):
    #      Single DR 131100 / CR 440500 (electric budget).
    #      Single DR 131100 / CR 440700 (gas budget).
    #
    # When the meter read JE is already in GL (440500/440700 have activity),
    # this block is skipped entirely for that account.
    _tub_accounts: set = set()

    def _post_tub_line(cr_code: str, cr_name: str, amount: float,
                       tenant: str, desc: str) -> None:
        """Append DR 131100 / CR recovery-account pair for one tenant billing."""
        nonlocal je_num
        je_id = f'TUB-{je_num:04d}'
        je_lines.append({
            'je_number':      je_id, 'line': 1, 'date': '',
            'account_code':   TENANT_UTILITY_AR_ACCOUNT,
            'account_name':   TENANT_UTILITY_AR_NAME,
            'description':    desc,
            'reference':      'METER-READ',
            'debit':          _round(amount), 'credit': 0,
            'vendor':         tenant or '[Tenant Billing]',
            'invoice_number': '',
            'source':         'tenant_utility_billing', 'confidence': 'medium',
        })
        je_lines.append({
            'je_number':      je_id, 'line': 2, 'date': '',
            'account_code':   cr_code,
            'account_name':   cr_name,
            'description':    desc,
            'reference':      'METER-READ',
            'debit':          0, 'credit': _round(amount),
            'vendor':         tenant or '[Tenant Billing]',
            'invoice_number': '',
            'source':         'tenant_utility_billing', 'confidence': 'medium',
        })
        je_num += 1

    # Build GL lookup for skip-guard (activity check). Available whenever gl_data
    # is present; used by both Mode (a) and Mode (b).
    _tub_gl: Dict[str, Any] = {}
    if gl_data:
        _tub_gl = {
            str(a.account_code).strip(): a
            for a in (gl_data.accounts if hasattr(gl_data, 'accounts') else [])
        }

    if tenant_utility_rows:
        # ── Mode (a): per-tenant actuals from sidebar ─────────────────────────
        # User explicitly entered meter-read amounts → ALWAYS generate JEs.
        # No GL activity guard here — suppress logic only applies to auto-detect.
        _total_elec_billed = 0.0   # accumulate for aggregate P&L JE below

        for row in tenant_utility_rows:
            tenant_name = str(row.get('tenant', '') or '').strip()
            elec_amt    = float(row.get('electric', 0) or 0)
            gas_amt     = float(row.get('gas',     0) or 0)
            if not tenant_name:
                continue

            if elec_amt > 0:
                _post_tub_line(
                    '440500', 'Recovery - Electricity', elec_amt,
                    tenant_name,
                    f'Tenant electric billing — {tenant_name} '
                    f'(per meter read) ${elec_amt:,.2f}',
                )
                _tub_accounts.add('440500')
                _total_elec_billed += elec_amt

            if gas_amt > 0:
                _post_tub_line(
                    '440700', 'Recovery - Misc Utilities', gas_amt,
                    tenant_name,
                    f'Tenant gas billing — {tenant_name} '
                    f'(per meter read) ${gas_amt:,.2f}',
                )
                _tub_accounts.add('440700')

        # P&L reclassification: DR 613115 Tenant Electric Reimb / CR 613110 Utilities - Electricity
        # Moves the tenant-reimbursable portion off the main electricity expense line.
        # Skip only if 613115 already has GL activity (613110 routinely has activity
        # from the real electricity bill and must NOT suppress the reclassification).
        if _round(_total_elec_billed) > 0:
            _reimb_gl = _tub_gl.get(ELEC_TENANT_REIMB_ACCOUNT)
            # Suppress only when our OWN ELEC-REIMB entry is already in the GL
            # (prevents double-posting on a re-run after Yardi import).
            # JLL manual reclasses (different reference) do NOT suppress — they
            # reduce what the pipeline posts but don't block it entirely.
            _reimb_gl_txns = getattr(_reimb_gl, 'transactions', [])
            _reimb_posted = sum(
                float(t.debit or 0) for t in _reimb_gl_txns
                if float(t.debit or 0) > 0
                and str(getattr(t, 'control',   '') or '').upper().startswith('J')
                and str(getattr(t, 'reference', '') or '').upper() == 'ELEC-REIMB'
            ) >= 0.01
            # Existing non-pipeline J-debits (e.g. JLL manual reclass) —
            # count toward what has already been reclassed this period.
            _existing_reclass = _round(sum(
                float(t.debit or 0) for t in _reimb_gl_txns
                if float(t.debit or 0) > 0
                and str(getattr(t, 'control',   '') or '').upper().startswith('J')
                and str(getattr(t, 'reference', '') or '').upper() != 'ELEC-REIMB'
            ))
            if not _reimb_posted:
                # ── Reclass: DR 613115 / CR 613110 ───────────────────────────────────
                #
                # In Mode (a) the pipeline owns 440500 — it posts exactly
                # _total_elec_billed as new credits this month.  613115 needs to
                # match 440500, so the reclass is simply:
                #
                #   pipeline_reclass = total_elec_billed − existing_JLL_reclass
                #
                # No catch-up formula: when the pipeline controls the TUB entries,
                # the GL gap between 440500 and 613115 is already captured in
                # existing_reclass vs total_elec_billed.  A catch-up term causes
                # double-counting when 440500 has reversal debits but no current-
                # period C-credits yet (pre-close timing), inflating the result.
                _pipeline_reclass = max(0.0, _round(_total_elec_billed - _existing_reclass))
                if _pipeline_reclass < 0.01:
                    je_num += 1
                    pass  # JLL covered it fully — nothing to post
                else:
                    _jll_note = (
                        f' (JLL posted ${_existing_reclass:,.2f}; pipeline posts ${_pipeline_reclass:,.2f} incremental)'
                        if _existing_reclass >= 0.01 else ''
                    )
                    _cmpd_note = ''
                    _elec_je_id = f'TUB-{je_num:04d}'
                    _elec_desc  = (f'Tenant electricity reclassification — '
                                   f'total billed ${_total_elec_billed:,.2f}'
                                   f'{_jll_note} '
                                   f'(DR {ELEC_TENANT_REIMB_ACCOUNT} / CR {ELEC_EXPENSE_ACCOUNT})')
                    je_lines.append({
                        'je_number':      _elec_je_id, 'line': 1, 'date': '',
                        'account_code':   ELEC_TENANT_REIMB_ACCOUNT,
                        'account_name':   ELEC_TENANT_REIMB_NAME,
                        'description':    _elec_desc,
                        'reference':      'ELEC-REIMB',
                        'debit':          _pipeline_reclass, 'credit': 0,
                        'vendor':         '[Tenant Electric Billing]',
                        'invoice_number': '',
                        'source':         'tenant_utility_billing', 'confidence': 'high',
                    })
                    je_lines.append({
                        'je_number':      _elec_je_id, 'line': 2, 'date': '',
                        'account_code':   ELEC_EXPENSE_ACCOUNT,
                        'account_name':   ELEC_EXPENSE_NAME,
                        'description':    _elec_desc,
                        'reference':      'ELEC-REIMB',
                        'debit':          0, 'credit': _pipeline_reclass,
                        'vendor':         '[Tenant Electric Billing]',
                        'invoice_number': '',
                        'source':         'tenant_utility_billing', 'confidence': 'high',
                    })
                    je_num += 1

    elif gl_data:
        # ── Mode (b): no sidebar rows — use Receivable Detail if uploaded, else budget ──
        #
        # Electric (440500):
        #   Priority 1 — Receivable Detail: sum charges_by_code for codes containing
        #                'ELEC' (e.g. ELEC, ELECTRIC, TELECTRIC, ELECRECOV).
        #                Reflects what was actually charged to tenants this month.
        #   Priority 2 — Budget: BC PTD budget for 440500 (legacy fallback).
        #
        # Gas (440700): Receivable Detail UTILI charges first, then budget fallback.
        #
        # GL activity guard: skip each account if already posted to Yardi.

        # ── Determine electric amount (per-tenant from Receivable Detail) ────────
        _elec_by_tenant: Dict[str, float] = {}
        _rec_elec_amt = 0.0
        _elec_source  = 'budget'
        _elec_conf    = 'medium'

        if receivable_detail and hasattr(receivable_detail, 'elec_by_tenant') and receivable_detail.elec_by_tenant:
            _elec_by_tenant = {k: v for k, v in receivable_detail.elec_by_tenant.items() if v > 0.0}
            _rec_elec_amt   = sum(_elec_by_tenant.values())
            if _rec_elec_amt > 0:
                _elec_source = 'receivable_detail'
                _elec_conf   = 'high'
        elif receivable_detail and hasattr(receivable_detail, 'charges_by_code'):
            # Fallback: aggregate from charges_by_code (no per-tenant breakdown)
            for _cc, _amt in receivable_detail.charges_by_code.items():
                if 'ELEC' in _cc.upper():
                    _rec_elec_amt += _amt
            if _rec_elec_amt > 0:
                _elec_source = 'receivable_detail'
                _elec_conf   = 'high'

        # ── Post 440500 electric recovery — one JE per tenant ─────────────────
        _mode_b_elec_total = 0.0
        _440500_gl = _tub_gl.get('440500')
        # Only J-type (journal) credits indicate a pipeline accrual already posted.
        # C-type charges (JLL billing transactions) and R-type receipts are NOT
        # accruals and must not suppress the pipeline's monthly J-entry.
        _440500_j_credits = _j_credits(_440500_gl)
        _440500_already_posted = _440500_j_credits >= 0.01

        # Generate 440500 AR recovery JE whenever Receivable Detail has per-tenant
        # data.  When only budget/GL fallbacks are available we still respect the
        # _440500_already_posted (J-only) guard to avoid double-posting a J-entry.
        if _elec_source == 'receivable_detail' and _rec_elec_amt > 0:
            if _elec_by_tenant:
                # Per-tenant breakout from Receivable Detail elec_by_tenant
                for _tenant_name, _tenant_amt in sorted(_elec_by_tenant.items()):
                    _post_tub_line(
                        '440500', 'Recovery - Electricity', _tenant_amt,
                        _tenant_name,
                        (f'Tenant electric recovery — {_tenant_name} '
                         f'per Receivable Detail '
                         f'(DR {TENANT_UTILITY_AR_ACCOUNT} / CR 440500)'),
                    )
                _tub_accounts.add('440500')
                _mode_b_elec_total = _rec_elec_amt
            else:
                # Aggregate fallback (elec_by_tenant not available)
                _post_tub_line(
                    '440500', 'Recovery - Electricity', _rec_elec_amt,
                    '[Receivable Detail]',
                    (f'Tenant electric recovery accrual — ${_rec_elec_amt:,.2f} '
                     f'per Receivable Detail electric charges '
                     f'(DR {TENANT_UTILITY_AR_ACCOUNT} / CR 440500)'),
                )
                _tub_accounts.add('440500')
                _mode_b_elec_total = _rec_elec_amt
        elif not _440500_already_posted:
            # No Receivable Detail — fall back to budget only when 440500 is not
            # already covered in GL (budget accrual would double-post if JLL posted).
            if budget_data:
                for cand in detect_tenant_utility_billing(gl_data, budget_data):
                    if cand['account_code'] != '440500':
                        continue
                    _post_tub_line(
                        '440500', 'Recovery - Electricity', cand['amount'],
                        '[Budget Accrual]',
                        cand['description'],
                    )
                    _tub_accounts.add('440500')
                    _mode_b_elec_total += cand['amount']
        else:
            # No Receivable Detail and a J-type accrual already in GL for 440500 —
            # read that J-entry amount for the P&L reclass check; no new JE needed.
            _mode_b_elec_total = _440500_j_credits
            _elec_source = 'gl_activity'
            _elec_conf   = 'high'
            if gl_activity_log is not None:
                gl_activity_log.append({
                    'account_code': '440500',
                    'account_name': 'Recovery - Electricity',
                    'reason': (
                        f'No Receivable Detail uploaded — J-type accrual of '
                        f'${_mode_b_elec_total:,.2f} already in GL for 440500. '
                        f'AR recovery JE skipped; P&L reclass (613115/613110) '
                        f'will be checked separately.'
                    ),
                    'suppressed': True,
                })

        # ── Fallback: derive 440500 from 613115 when JLL posted the reclass ──
        # Scenario: JLL posted the P&L reclass (DR 613115 / CR 613110) but
        # NOT the AR recovery (440500). Receivable Detail not uploaded and no
        # PTD budget for 440500 (common — it's a passthrough account in Kardin).
        # 613115 net_change == the electric total billed, so it's a reliable
        # basis for generating the missing 440500 AR recovery JE.
        if _round(_mode_b_elec_total) == 0:
            _613115_fb_gl = _tub_gl.get(ELEC_TENANT_REIMB_ACCOUNT)
            # total_debits: JLL's reclass is DR 613115 — a credit-only entry is
            # just an auto-reversal and must not trigger the fallback.
            if (_613115_fb_gl is not None
                    and _j_debits(_613115_fb_gl) >= 0.01):
                _fb_amt = _j_debits(_613115_fb_gl)
                _post_tub_line(
                    '440500', 'Recovery - Electricity', _fb_amt,
                    '[JLL Reclass Basis]',
                    (f'Tenant electric recovery — derived from JLL 613115 reclass '
                     f'${_fb_amt:,.2f} '
                     f'(DR {TENANT_UTILITY_AR_ACCOUNT} / CR 440500). '
                     f'Upload Receivable Detail for per-tenant breakdown.'),
                )
                _tub_accounts.add('440500')
                _mode_b_elec_total = _fb_amt
                _elec_source = 'gl_613115_basis'
                _elec_conf   = 'medium'

        # ── Last-resort: use 613115 PTD budget as proxy for 440500 ───────────
        # 440500 (Recovery - Electricity) is a passthrough account that is
        # typically not budgeted in Kardin, so the budget fallback above returns
        # $0.  But 613115 (Tenant Electric Reimbursement) IS usually budgeted —
        # and since 440500 and 613115 always pair, the 613115 PTD budget is the
        # best available estimate of the monthly electric recovery amount.
        if _round(_mode_b_elec_total) == 0 and budget_data:
            _budget_rows = (budget_data if isinstance(budget_data, list)
                            else getattr(budget_data, 'line_items', []))
            _613115_ptd_budget = 0.0
            for _br in _budget_rows:
                _br_code = str((_br.get('account_code') if isinstance(_br, dict)
                                else getattr(_br, 'account_code', '')) or '').strip()
                if _br_code == ELEC_TENANT_REIMB_ACCOUNT:
                    _br_ptd = ((_br.get('ptd_budget') if isinstance(_br, dict)
                                else getattr(_br, 'ptd_budget', 0)) or 0)
                    _613115_ptd_budget = abs(float(_br_ptd))
                    break
            if _613115_ptd_budget >= 1.0:
                _post_tub_line(
                    '440500', 'Recovery - Electricity', _613115_ptd_budget,
                    '[Budget Accrual]',
                    (f'Tenant electric recovery accrual — ${_613115_ptd_budget:,.2f} '
                     f'derived from 613115 PTD budget (440500 not separately budgeted). '
                     f'Upload Receivable Detail or enter meter reads for actuals.'),
                )
                _tub_accounts.add('440500')
                _mode_b_elec_total = _613115_ptd_budget
                _elec_source = 'budget'
                _elec_conf   = 'low'

        # ── Post 440700 gas/misc utility recovery ─────────────────────────────
        _440700_gl = _tub_gl.get('440700')
        _440700_j_credits = _j_credits(_440700_gl)
        _440700_already_posted = _440700_j_credits >= 0.01

        # Receivable Detail path: always generate when UTILI charges are present —
        # same rule as 440500: JLL's C-type billing charge must not suppress the
        # pipeline's monthly J accrual.
        _utili_by_tenant: Dict[str, float] = {}
        if (receivable_detail and hasattr(receivable_detail, 'utili_by_tenant')
                and receivable_detail.utili_by_tenant):
            _utili_by_tenant = {k: v for k, v in receivable_detail.utili_by_tenant.items() if v > 0.0}

        if _utili_by_tenant:
            for _ut_name, _ut_amt in sorted(_utili_by_tenant.items()):
                _post_tub_line(
                    '440700', 'Recovery - Misc Utilities', _ut_amt,
                    _ut_name,
                    (f'Tenant misc utility recovery — {_ut_name} '
                     f'per Receivable Detail '
                     f'(DR {TENANT_UTILITY_AR_ACCOUNT} / CR 440700)'),
                )
            _tub_accounts.add('440700')
        elif (receivable_detail and hasattr(receivable_detail, 'charges_by_code')):
            # Aggregate fallback from charges_by_code (no per-tenant breakdown)
            _utili_total = sum(
                v for k, v in receivable_detail.charges_by_code.items()
                if 'UTILI' in k.upper() or k.upper() in ('UTIL', 'GAS', 'GASREC', 'UTILITIES', 'UTILITY')
            )
            if _utili_total >= 1.0:
                _post_tub_line(
                    '440700', 'Recovery - Misc Utilities', _utili_total,
                    '[Receivable Detail]',
                    (f'Tenant misc utility recovery — ${_utili_total:,.2f} '
                     f'per Receivable Detail UTILI charges '
                     f'(DR {TENANT_UTILITY_AR_ACCOUNT} / CR 440700)'),
                )
                _tub_accounts.add('440700')
        elif not _440700_already_posted and budget_data:
            # Budget fallback only when no Receivable Detail and no J-accrual in GL
            for cand in detect_tenant_utility_billing(gl_data, budget_data):
                if cand['account_code'] != '440700':
                    continue
                _post_tub_line(
                    '440700', 'Recovery - Misc Utilities', cand['amount'],
                    '[Budget Accrual]',
                    cand['description'],
                )
                _tub_accounts.add('440700')

        # ── P&L reclassification for Mode (b) electric ───────────────────────
        # One aggregate reclass entry regardless of how many per-tenant AR JEs were posted.
        # Also fires when 440500 was already posted by JLL — in that case
        # _mode_b_elec_total is read from the existing GL balance and no new
        # AR JE is generated, but the 613115/613110 reclass is still needed.
        if _round(_mode_b_elec_total) > 0:
            _reimb_gl = _tub_gl.get(ELEC_TENANT_REIMB_ACCOUNT)
            # Suppress only when our own ELEC-REIMB entry is already in the GL.
            # JLL manual reclasses (different reference) reduce but don't block.
            _reimb_gl_txns_b = getattr(_reimb_gl, 'transactions', [])
            _reimb_b_posted = sum(
                float(t.debit or 0) for t in _reimb_gl_txns_b
                if float(t.debit or 0) > 0
                and str(getattr(t, 'control',   '') or '').upper().startswith('J')
                and str(getattr(t, 'reference', '') or '').upper() == 'ELEC-REIMB'
            ) >= 0.01
            _existing_reclass_b = _round(sum(
                float(t.debit or 0) for t in _reimb_gl_txns_b
                if float(t.debit or 0) > 0
                and str(getattr(t, 'control',   '') or '').upper().startswith('J')
                and str(getattr(t, 'reference', '') or '').upper() != 'ELEC-REIMB'
            ))
            if not _reimb_b_posted:
                # Same catch-up logic as Mode (a): reclass absorbs the shortfall
                # between prior actual billing and prior TUB estimate.
                _440500_gl_obj_b = _tub_gl.get('440500')
                _440500_j_rev_b = _round(sum(
                    float(t.debit or 0)
                    for t in getattr(_440500_gl_obj_b, 'transactions', [])
                    if float(t.debit or 0) > 0
                    and str(getattr(t, 'control', '') or '').upper().startswith('J')
                    and (
                        ':reversal of' in str(t.description or '').lower()
                        or ':reversal of' in str(getattr(t, 'remarks', '') or '').lower()
                    )
                ))
                _440500_c_cr_b = _round(sum(
                    float(t.credit or 0)
                    for t in getattr(_440500_gl_obj_b, 'transactions', [])
                    if float(t.credit or 0) > 0
                    and not str(getattr(t, 'control', '') or '').upper().startswith('J')
                ))
                _prior_actual_b = _440500_c_cr_b if _440500_c_cr_b >= 0.01 else _mode_b_elec_total
                _catch_up_b = (
                    max(0.0, _round(_prior_actual_b - _440500_j_rev_b))
                    if _440500_j_rev_b >= 0.01 else 0.0
                )
                _reimb_b_total     = _round(_mode_b_elec_total + _catch_up_b)
                # Subtract any reclass JLL already posted; post only the delta.
                _pipeline_reclass_b = max(0.0, _round(_reimb_b_total - _existing_reclass_b))
                if _pipeline_reclass_b < 0.01:
                    je_num += 1
                    pass  # JLL covered it fully; fall through without appending
                else:
                    _elec_je_id  = f'TUB-{je_num:04d}'
                    _src_label   = {
                        'receivable_detail': 'Receivable Detail',
                        'gl_activity':       'GL activity — 440500 posted by JLL',
                        'gl_613115_basis':   'GL activity — 613115 reclass posted by JLL',
                        'budget':            'budget',
                    }.get(_elec_source, _elec_source)
                    _n_tenants   = len(_elec_by_tenant) if _elec_by_tenant else 1
                    _tenant_note = (
                        f'{_n_tenants} tenant(s)' if _elec_source == 'receivable_detail'
                        else _src_label
                    )
                    _cmpd_b_note = (
                        f' — total ${_reimb_b_total:,.2f} '
                        f'(${_catch_up_b:,.2f} catch-up + ${_mode_b_elec_total:,.2f} est.)'
                        if _catch_up_b > 0 else ''
                    )
                    _jll_b_note = (
                        f' (JLL posted ${_existing_reclass_b:,.2f}; pipeline posts ${_pipeline_reclass_b:,.2f} incremental)'
                        if _existing_reclass_b >= 0.01 else ''
                    )
                    _elec_desc   = (f'Tenant electricity reclassification — {_tenant_note} — '
                                    f'${_mode_b_elec_total:,.2f}'
                                    f'{_cmpd_b_note}{_jll_b_note} '
                                    f'(DR {ELEC_TENANT_REIMB_ACCOUNT} / CR {ELEC_EXPENSE_ACCOUNT})')
                    _elec_vendor = {
                        'receivable_detail': '[Receivable Detail]',
                        'gl_activity':       '[JLL GL Activity]',
                        'gl_613115_basis':   '[JLL GL Activity — 613115 basis]',
                        'budget':            '[Budget Accrual]',
                    }.get(_elec_source, '[Budget Accrual]')
                    je_lines.append({
                        'je_number':      _elec_je_id, 'line': 1, 'date': '',
                        'account_code':   ELEC_TENANT_REIMB_ACCOUNT,
                        'account_name':   ELEC_TENANT_REIMB_NAME,
                        'description':    _elec_desc,
                        'reference':      'ELEC-REIMB',
                        'debit':          _pipeline_reclass_b, 'credit': 0,
                        'vendor':         _elec_vendor,
                        'invoice_number': '',
                        'source':         'tenant_utility_billing', 'confidence': _elec_conf,
                    })
                    je_lines.append({
                        'je_number':      _elec_je_id, 'line': 2, 'date': '',
                        'account_code':   ELEC_EXPENSE_ACCOUNT,
                        'account_name':   ELEC_EXPENSE_NAME,
                        'description':    _elec_desc,
                        'reference':      'ELEC-REIMB',
                        'debit':          0, 'credit': _pipeline_reclass_b,
                        'vendor':         _elec_vendor,
                        'invoice_number': '',
                        'source':         'tenant_utility_billing', 'confidence': _elec_conf,
                    })
                    je_num += 1

        # ── Electricity expense accrual (Mode b) ─────────────────────────────
        # Accrue the FULL building electricity expense (613110) at the budget
        # amount whenever the bill hasn't been posted yet.  Runs independently
        # of whether the tenant billing (440500) was already posted — the
        # expense accrual is needed regardless.
        #
        # Skip if 613110 already has GL activity (invoice/JE posted).
        # Skip if no PTD budget for 613110 exists in the BC report.
        #
        # Putting 613110 on line=1 with source='tenant_utility_billing' ensures
        # the _covered seeding (see below) blocks budget_gap from double-accruing.
        _elec_exp_gl = _tub_gl.get(ELEC_EXPENSE_ACCOUNT)
        _elec_exp_active = (
            _elec_exp_gl is not None and abs(_elec_exp_gl.net_change) >= 1.0
        )
        if not _elec_exp_active:
            # ── Derive months_elapsed from close period ────────────────────────
            _elec_month_map = {
                'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
            }
            _elec_month_num = 0
            for _ab, _nm in _elec_month_map.items():
                if _ab in (period or '').lower():
                    _elec_month_num = _nm
                    break
            _elec_months_elapsed = max(0, _elec_month_num - 1)

            # ── Look up 613110 from BC: prefer YTD actual ÷ months, fall back to budget ──
            _elec_exp_amt    = 0.0
            _elec_exp_source = 'budget'
            _bc_rows = (budget_data if isinstance(budget_data, list)
                        else getattr(budget_data, 'line_items', []))
            for _bc_row in _bc_rows:
                _bc_code = str(
                    (_bc_row.get('account_code') if isinstance(_bc_row, dict)
                     else getattr(_bc_row, 'account_code', '')) or ''
                ).strip()
                if _bc_code == ELEC_EXPENSE_ACCOUNT:
                    _ytd_actual = abs(float(
                        (_bc_row.get('ytd_actual') if isinstance(_bc_row, dict)
                         else getattr(_bc_row, 'ytd_actual', 0)) or 0
                    ))
                    _ptd_budget = abs(float(
                        (_bc_row.get('ptd_budget') if isinstance(_bc_row, dict)
                         else getattr(_bc_row, 'ptd_budget', 0)) or 0
                    ))
                    if _elec_months_elapsed >= 1 and _ytd_actual > 0:
                        # Feb onward: average of prior closed months' actual bills
                        _elec_exp_amt    = _ytd_actual / _elec_months_elapsed
                        _elec_exp_source = 'prior_actual'
                    else:
                        # January (months_elapsed=0, no prior-year YTD) → fall back to PTD budget
                        _elec_exp_amt    = _ptd_budget
                        _elec_exp_source = 'budget'
                    break

            if _elec_exp_amt > 500:
                _exp_je_id = f'TUB-{je_num:04d}'
                if _elec_exp_source == 'prior_actual':
                    _exp_desc = (
                        f'Electricity expense accrual — '
                        f'{ELEC_EXPENSE_NAME}: ${_elec_exp_amt:,.2f}/mo '
                        f'(avg of {_elec_months_elapsed} prior month(s) actual; '
                        f'update when actual bill received)'
                    )
                else:
                    _exp_desc = (
                        f'Electricity expense accrual (budget) — '
                        f'{ELEC_EXPENSE_NAME}: ${_elec_exp_amt:,.2f}/mo '
                        f'(no invoice in GL; update when actual bill received)'
                    )
                _exp_vendor = '[Prior Actual Avg]' if _elec_exp_source == 'prior_actual' else '[Budget Accrual]'
                _exp_conf   = 'high' if _elec_exp_source == 'prior_actual' else 'medium'
                je_lines.append({
                    'je_number':      _exp_je_id, 'line': 1, 'date': '',
                    'account_code':   ELEC_EXPENSE_ACCOUNT,
                    'account_name':   ELEC_EXPENSE_NAME,
                    'description':    _exp_desc,
                    'reference':      'ELEC-ACCRUAL',
                    'debit':          _round(_elec_exp_amt), 'credit': 0,
                    'vendor':         _exp_vendor,
                    'invoice_number': '',
                    'source':         'tenant_utility_billing',
                    'confidence':     _exp_conf,
                })
                je_lines.append({
                    'je_number':      _exp_je_id, 'line': 2, 'date': '',
                    'account_code':   AP_ACCRUAL_ACCOUNT,
                    'account_name':   AP_ACCRUAL_NAME,
                    'description':    _exp_desc,
                    'reference':      'ELEC-ACCRUAL',
                    'debit':          0, 'credit': _round(_elec_exp_amt),
                    'vendor':         _exp_vendor,
                    'invoice_number': '',
                    'source':         'tenant_utility_billing',
                    'confidence':     _exp_conf,
                })
                je_num += 1

    # ── Layer 0b: Prepaid / escrow amortization ────────────────────────────────
    # Entries that move cost between balance-sheet assets and P&L expense.
    # Do NOT create a new liability (213100) — they reduce an existing BS asset.
    #
    #   Insurance:   DR 639110/639120  /  CR 135110  Prepaid Insurance
    #
    #   RE Taxes (quarterly invoice cycle — Jan/Apr/Jul/Oct):
    #     Payment months: Berkadia/Yardi auto-posts full quarterly bill
    #       (DR 641110 / CR 115200) via the loan payment entry — NOT by pipeline.
    #       Pipeline defers 2/3:  DR 135120 Prepaid RE Taxes / CR 641110 (2/3)
    #       → only 1/3 of the quarterly bill hits P&L in the payment month.
    #     Release months (all other): pipeline releases 1/3 each month:
    #       DR 641110 Real Estate Taxes / CR 135120 Prepaid RE Taxes (1/3)
    #     Net effect: expense spread evenly — 1/3 per month across the quarter.
    _amort_accounts: set = set()

    def _post_amort(entry: Dict[str, Any], prefix: str, ref: str, vendor: str) -> None:
        """Append a DR/CR amortization pair to je_lines and register the account."""
        nonlocal je_num
        acct_code = entry['account_code']
        if acct_code in _manual_accounts:
            return  # user override takes precedence
        je_id  = f'{prefix}-{je_num:04d}'
        amount = entry['amount']
        desc   = entry['description']
        je_lines.append({
            'je_number':         je_id, 'line': 1, 'date': '',
            'account_code':      acct_code,
            'account_name':      entry['account_name'],
            'description':       desc, 'reference': ref,
            'debit':             _round(amount), 'credit': 0,
            'vendor':            vendor, 'invoice_number': '',
            'source':            'prepaid_amortization', 'confidence': 'high',
            'reverse_next_month': 0,  # prepaid movements are permanent — no reversal
        })
        je_lines.append({
            'je_number':         je_id, 'line': 2, 'date': '',
            'account_code':      entry['credit_account'],
            'account_name':      entry['credit_name'],
            'description':       desc, 'reference': ref,
            'debit':             0, 'credit': _round(amount),
            'vendor':            vendor, 'invoice_number': '',
            'source':            'prepaid_amortization', 'confidence': 'high',
            'reverse_next_month': 0,  # prepaid movements are permanent — no reversal
        })
        _amort_accounts.add(acct_code)
        _amort_accounts.add(entry.get('credit_account', ''))  # cover both sides
        je_num += 1

    # Insurance: DR 639110/639120 / CR 135110
    # Skip any account already covered by the prepaid ledger release JEs
    # (ledger releases DR the same expense accounts / CR 135150 — running
    # detect_insurance_amortization() on top would double-count the expense).
    _ledger_ins_covered = (ledger_release_accounts or set()) & _INSURANCE_EXPENSE_ACCTS
    if gl_data and (budget_data or insurance_policies or kardin_records) and not _ledger_ins_covered:
        for ins in detect_insurance_amortization(
            gl_data, budget_data, period=period,
            insurance_policies=insurance_policies or None,
            kardin_records=kardin_records or None,
        ):
            _post_amort(ins, 'INS', 'INS-AMORT', '[Insurance Amortization]')
    elif _ledger_ins_covered:
        # Still mark the accounts as covered so Layer 3 doesn't double-accrue them
        _amort_accounts.update(_ledger_ins_covered)
        _amort_accounts.add(_PREPAID_INSURANCE_ACCT)

    # RE Taxes — all months when re_tax_bill_amount is entered:
    #   Payment months (Jan/Apr/Jul/Oct):
    #       Berkadia/Yardi auto-posts full bill (DR 641110 / CR 115200).
    #       Pipeline defers 2/3:  DR 135120 Prepaid RE Taxes / CR 641110 Real Estate Taxes
    #   Release months (all other):
    #       Pipeline releases 1/3:  DR 641110 Real Estate Taxes / CR 135120 Prepaid RE Taxes
    if gl_data:
        retax = detect_retax_amortization(gl_data, period=period,
                                           re_tax_bill_amount=re_tax_bill_amount,
                                           re_tax_payment_months=re_tax_payment_months)
        if retax:
            _post_amort(retax, 'TAX', 'TAX-AMORT', '[RE Tax Amortization]')
    # Note: detect_retax_escrow_je() (full-bill DR 641110 / CR 115200) is retained
    # in the codebase for reference but is NO LONGER CALLED — Berkadia's Yardi loan
    # payment entries post this automatically each payment month.

    _seen_nexus_inv_nums: set = set()  # intra-batch dedup — Nexus sometimes has duplicate rows

    for inv in invoices:
        vendor = str(inv.get('vendor', '') or '')
        inv_num = str(inv.get('invoice_number', '') or '')
        inv_date = inv.get('invoice_date', '')
        # Use numeric account number if available (e.g. "637370" not "Admin-Computer/Software (637370)")
        gl_account = str(inv.get('gl_account_number', '') or inv.get('gl_account', '') or '')
        gl_category = str(inv.get('gl_category', '') or '')
        description = str(inv.get('line_description', '') or '')
        amount = inv.get('amount', 0) or 0

        if amount == 0:
            continue

        # Guard: blank GL account would produce an invalid JE that Yardi rejects at import
        if not gl_account:
            continue

        # Skip if user has manually specified this account in the One-Off table —
        # their override (with amount > 0) or suppression (amount = 0) takes precedence.
        if gl_account in _manual_accounts:
            continue

        # Skip if Layer 0b amortization (insurance or RE tax) already claimed this
        # account — prevents a Nexus invoice for e.g. 641110 (RE tax) from generating
        # a second JE on top of the TAX-AMORT entry.
        if gl_account in _amort_accounts:
            continue

        # Intra-batch dedup: Nexus sometimes submits the same invoice twice (resubmit,
        # dual approval workflow). Both would clear the GL dedup check and generate
        # duplicate JEs. Deduplicate within this Nexus batch first.
        if inv_num:
            _inv_key = f'{inv_num}|{gl_account}'
            if _inv_key in _seen_nexus_inv_nums:
                continue
            _seen_nexus_inv_nums.add(_inv_key)

        # Dedup against GL — two strategies, first-match wins:
        #   Strategy 1 (exact):     invoice number matches GL reference/control
        #   Strategy 2 (fuzzy):     vendor name + amount already posted to same account
        #                           (fires only when invoice number is absent)
        if inv_num and _is_invoice_in_gl(inv_num, gl_lookup):
            continue
        if not inv_num and _is_in_gl_by_vendor_amount(vendor, amount, gl_account, gl_lookup):
            continue

        # Format date
        if isinstance(inv_date, datetime):
            date_str = inv_date.strftime('%m/%d/%Y')
        elif isinstance(inv_date, str):
            date_str = inv_date
        else:
            date_str = str(inv_date) if inv_date else ''

        # Build description for JE — "Accrual [Month YYYY] — Vendor #INV-NUM"
        _nex_period_label = _fmt_period(period)
        je_desc = f"Accrual {_nex_period_label} — {vendor}"
        if inv_num:
            je_desc += f" #{inv_num}"
        if description:
            je_desc += f" — {description[:40]}"

        # ── Prepaid split: accrue only current-month portion to expense;
        #    remaining future months go to Prepaid Other (135150).
        #    Month 1 of N: DR expense (1/N) + DR 135150 (N-1/N) / CR 213100 (full)
        is_prepaid = inv.get('is_prepaid', False)
        prepaid_months = int(inv.get('prepaid_months', 1) or 1)

        if is_prepaid and prepaid_months > 1:
            monthly_amt = _round(abs(amount) / prepaid_months)
            rounding_adj = _round(abs(amount) - monthly_amt * prepaid_months)
            current_amt = monthly_amt + rounding_adj          # this period's expense
            future_amt  = abs(amount) - current_amt           # prepaid asset to book
        else:
            current_amt = abs(amount)
            future_amt  = 0.0

        je_id = f"ACC-{je_num:04d}"
        acct_name = gl_category or description[:30]

        # DR line: Expense account (current month only)
        # 'source': 'nexus' is REQUIRED — the _covered exclusion set at the
        # bottom of this function filters on source == 'nexus' to prevent
        # Layers 2-4 from generating duplicate entries for these accounts.
        je_lines.append({
            'je_number':      je_id,
            'line':           1,
            'date':           date_str,
            'account_code':   gl_account,
            'account_name':   acct_name,
            'description':    je_desc,
            'reference':      inv_num,
            'debit':          current_amt,
            'credit':         0,
            'vendor':         vendor,
            'invoice_number': inv_num,
            'source':         'nexus',
        })

        # CR line: Accrued Expenses (213100) or Accrued Interest (213200) depending on DR account
        _cr_acct, _cr_name = _cr_for(gl_account)
        je_lines.append({
            'je_number':      je_id,
            'line':           2,
            'date':           date_str,
            'account_code':   _cr_acct,
            'account_name':   _cr_name,
            'description':    je_desc,
            'reference':      inv_num,
            'debit':          0,
            'credit':         current_amt,
            'vendor':         vendor,
            'invoice_number': inv_num,
            'source':         'nexus',
        })

        je_num += 1

        # Second JE: book future months to Prepaid Other (135150)
        if future_amt > 0:
            je_id_ppd = f"ACC-{je_num:04d}"
            ppd_desc = f"Prepaid booking — {vendor} #{inv_num} ({prepaid_months - 1} future mo)"

            je_lines.append({
                'je_number':      je_id_ppd,
                'line':           1,
                'date':           date_str,
                'account_code':   PREPAID_ASSET_ACCOUNT,
                'account_name':   PREPAID_ASSET_NAME,
                'description':    ppd_desc,
                'reference':      inv_num,
                'debit':          future_amt,
                'credit':         0,
                'vendor':         vendor,
                'invoice_number': inv_num,
                'source':         'nexus',
            })
            je_lines.append({
                'je_number':      je_id_ppd,
                'line':           2,
                'date':           date_str,
                'account_code':   AP_ACCRUAL_ACCOUNT,
                'account_name':   AP_ACCRUAL_NAME,
                'description':    ppd_desc,
                'reference':      inv_num,
                'debit':          0,
                'credit':         future_amt,
                'vendor':         vendor,
                'invoice_number': inv_num,
                'source':         'nexus',
            })
            je_num += 1

    # ── Resolve reporting month-end (used by Layers 2 and onward) ──
    _month_end = period_month_end or _month_end_from_period(period)
    if _month_end is None and gl_data:
        try:
            _month_end = _month_end_from_period(gl_data.metadata.period)
        except Exception:
            pass

    # Accounts that are ALWAYS handled by dedicated pipeline modules and must never
    # be touched by any automated accrual layer (2, 3, or budget gap).
    # 637130 — Admin-Management Fees: computed by management_fee.py (cash-received
    #           basis), always generates MGT-001 separately; historical/proration
    #           layers must not add a second accrual.
    _PIPELINE_RESERVED = {'637130'}

    # Collect accounts already covered by Layers 0 (manual), 0b (amortization),
    # 1 (Nexus), and TUB (tenant utility billing). Seeding _covered here prevents
    # later layers from generating duplicate entries for the same account.
    # TUB line=1 accounts include:
    #   133110  — AR Billback (DR side of each tenant billing JE)
    #   613115  — Tenant Electric Reimb (DR side of P&L reclass)
    #   613110  — Electricity Expense (DR side of Mode-b budget expense accrual, when generated)
    # Ensuring 613110 lands on line=1 prevents budget_gap from double-accruing it.
    _covered = _PIPELINE_RESERVED | _manual_accounts | _amort_accounts | set(
        l['account_code'] for l in je_lines
        if l.get('line') == 1 and l.get('source') in ('nexus', 'tenant_utility_billing')
    )

    # Multi-layer review tracking: when a later layer detects an account that
    # was already claimed by an earlier layer, the earlier entry is flagged for
    # reviewer attention (first-layer-wins, but the reviewer knows why).
    _other_claimants: Dict[str, List[str]] = {}

    # ── Layer 1b: Berkadia actual interest expense — per tranche ─────────────
    # Uses the payment_interest field from each parsed Berkadia loan statement
    # instead of the Layer 3 historical average.  One balanced JE is generated
    # per loan tranche so each tranche is visible in the import CSV and workpaper.
    # Only fires when loan_data is provided and 801110 is not already posted in
    # the GL via J-type entries.
    #
    # Defensive guard: if loan_data is missing (PDF parse failed), mark 801110
    # as reserved so Layer 3 historical NEVER auto-accrues interest expense —
    # interest must always come from Berkadia statements or the user's one-off
    # accruals table, never from a historical average.
    if not loan_data:
        _covered.add('801110')
        _covered.add('213200')
    if loan_data:
        _loans = loan_data if isinstance(loan_data, list) else [loan_data]
        # Check GL once — skip all tranches if J-type interest already posted
        _int_gl = next((a for a in (gl_data.accounts if gl_data else [])
                        if str(a.account_code).strip() == '801110'), None)
        _int_already = _j_debits(_int_gl) >= 1.0

        # When Berkadia statements are uploaded, they are the source of truth for
        # interest expense — generate INT- JEs regardless of what J-type entries
        # are already in the GL (e.g. a prior incorrect accrual that was uploaded
        # to Yardi and needs to be replaced).  The _int_already guard is intentionally
        # bypassed here; the user must void the prior GL entry separately in Yardi.
        if '801110' not in _covered:
            _any_interest_posted = False
            for _ln in _loans:
                if isinstance(_ln, dict):
                    _pi       = _safe_float(_ln.get('payment_interest', 0))
                    _loan_num = _ln.get('loan_number') or _ln.get('account_number') or ''
                    _prop     = _ln.get('property_name', '')
                else:
                    _pi       = _safe_float(getattr(_ln, 'payment_interest', 0))
                    _loan_num = getattr(_ln, 'loan_number', '') or getattr(_ln, 'account_number', '')
                    _prop     = getattr(_ln, 'property_name', '')

                if _pi < 1.0:
                    continue   # no interest on this tranche this period

                # Build a readable tranche label: prefer loan_number, fallback to property
                _tranche_label = (f'Loan #{_loan_num}' if _loan_num
                                  else (_prop or 'Tranche'))
                _period_label = _fmt_period(period)
                _int_desc = (
                    f'Accrual {_period_label} — Berkadia {_tranche_label} '
                    f'Mortgage Interest (${_pi:,.2f})'
                )
                je_id = f"INT-{je_num:04d}"
                je_num += 1
                je_lines += [
                    {
                        'je_number': je_id, 'line': 1, 'date': period,
                        'account_code': '801110',
                        'account_name': 'Interest Expense',
                        'description': _int_desc, 'reference': _loan_num,
                        'debit': _round(_pi), 'credit': 0.0,
                        'vendor': 'Berkadia', 'invoice_number': _loan_num,
                        'source': 'berkadia_interest',
                    },
                    {
                        'je_number': je_id, 'line': 2, 'date': period,
                        'account_code': '213200',
                        'account_name': 'Accrued Interest Payable',
                        'description': _int_desc, 'reference': _loan_num,
                        'debit': 0.0, 'credit': _round(_pi),
                        'vendor': 'Berkadia', 'invoice_number': _loan_num,
                        'source': 'berkadia_interest',
                    },
                ]
                _any_interest_posted = True

            if _any_interest_posted:
                _covered.add('801110')
                _covered.add('213200')

        # Always reserve 801110/213200 when loan_data is provided — even if
        # payment_interest extraction returned 0 for all tranches (e.g. PDF
        # parse partial failure).  Layer 3 historical must NEVER auto-accrue
        # interest when Berkadia statements have been uploaded.
        _covered.add('801110')
        _covered.add('213200')

    # ── Layer 2: Invoice-period proration ──
    if gl_data:
        prorations = detect_invoice_proration_accruals(
            gl_data, period=period, month_end=_month_end,
            metered_utility_accounts=metered_utility_accounts,
            per_invoice_utility_accounts=per_invoice_utility_accounts,
            per_invoice_accrual_accounts=per_invoice_accrual_accounts,
        )
        _proration_covered: set = set()   # accounts handled by this layer
        for pro in prorations:
            if pro['account_code'] in _covered:
                _other_claimants.setdefault(pro['account_code'], []).append('invoice_proration')
                continue   # already handled by Nexus or an earlier layer

            je_id   = f"IPR-{je_num:04d}"
            je_desc = pro['description']

            je_lines.append({
                'je_number':      je_id,
                'line':           1,
                'date':           _month_end.strftime('%m/%d/%Y') if _month_end else '',
                'account_code':   pro['account_code'],
                'account_name':   pro['account_name'],
                'description':    je_desc,
                'reference':      'INV-PRORATION',
                'debit':          pro['accrual_amount'],
                'credit':         0,
                'vendor':         '[Invoice Proration]',
                'invoice_number': '',
                'source':         'invoice_proration',
            })
            _cr_acct, _cr_name = _cr_for(pro['account_code'])
            je_lines.append({
                'je_number':      je_id,
                'line':           2,
                'date':           _month_end.strftime('%m/%d/%Y') if _month_end else '',
                'account_code':   _cr_acct,
                'account_name':   _cr_name,
                'description':    je_desc,
                'reference':      'INV-PRORATION',
                'debit':          0,
                'credit':         pro['accrual_amount'],
                'vendor':         '[Invoice Proration]',
                'invoice_number': '',
                'source':         'invoice_proration',
            })
            # Track covered accounts but do NOT add to _covered mid-loop —
            # an account can have multiple proration candidates (e.g. 613110
            # with Eversource + Hudson on different billing cycles). Adding
            # to _covered after the first candidate would silently drop all
            # subsequent candidates for the same account.
            _proration_covered.add(pro['account_code'])
            je_num += 1

        # Now mark all proration-handled accounts as covered so Layer 3
        # (historical / budget) does not also accrue them.
        _covered.update(_proration_covered)

    # ── GL-activity universal gate ──────────────────────────────────────────
    # After Layer 2, look for accounts with Journal Entry (J-type) activity
    # that the pipeline would otherwise try to accrue. K=Check, C=Charge,
    # R=Receipt, P=Payable are routine operational transactions handled by
    # other layers — only J entries indicate a manually posted JE that could
    # double up with what the pipeline generates.
    #
    # $500 floor on J-type net to ignore small rounding/test entries.
    # Expense accounts: only suppress when J-type net is a debit (new expense).
    # A J-type net credit means prior accruals auto-reversed — still accrue.
    # Balance-sheet accounts: use abs() since BS JEs don't auto-reverse.
    _GL_ACTIVITY_FLOOR = 500.0
    if gl_data and hasattr(gl_data, 'accounts'):
        for _gl_acct in gl_data.accounts:
            _gl_code = str(_gl_acct.account_code).strip()
            if _gl_code in _covered:
                continue

            _j_txns = [
                _t for _t in getattr(_gl_acct, 'transactions', [])
                if (_t.control or '').split('-')[0].upper() == 'J'
                and abs(_t.debit - _t.credit) >= 0.01
            ]
            if not _j_txns:
                continue

            _j_net = sum(_t.debit - _t.credit for _t in _j_txns)

            # Only expense accounts — the pipeline never generates accruals for
            # revenue, AR, AP, or other BS accounts, so J entries there are
            # irrelevant to the gut-check.
            if not is_expense_account(_gl_code):
                continue

            _qualifies = _j_net >= _GL_ACTIVITY_FLOOR

            if _qualifies:
                _covered.add(_gl_code)
                if gl_activity_log is not None:
                    gl_activity_log.append({
                        'account_code': _gl_code,
                        'account_name': str(_gl_acct.account_name or _gl_code),
                        'ptd_amount':   abs(_j_net),
                        'transactions': [
                            {
                                'date':        str(_t.date) if _t.date else '',
                                'type':        (_t.control or '').split('-')[0].upper() or '?',
                                'description': _t.description or '',
                                'reference':   _t.reference or '',
                                'debit':       _t.debit,
                                'credit':      _t.credit,
                            }
                            for _t in _j_txns
                        ],
                    })

    # ── Layer 3 (new order): Historical recurring accruals ──────────────────
    # Moved before budget gap: BC YTD ÷ months elapsed is more reliable than
    # budget assumptions for accounts with prior spending history.
    # Fires when an expense account had activity in prior months but is silent
    # this period — the average prior month spend is used as the accrual estimate.
    if gl_data:
        historicals = detect_historical_recurring(gl_data, budget_data, period=period,
                                                    t12_result=t12_result,
                                                    fiscal_year_start_month=fiscal_year_start_month,
                                                    kardin_records=kardin_records,
                                                    materiality=accrual_materiality_floor,
                                                    layer3_exclude_accounts=layer3_exclude_accounts)
        # Build a quick lookup so we can fetch GL accounts by code in the loop below.
        _gl_acct_lookup = {
            str(a.account_code).strip(): a
            for a in getattr(gl_data, 'accounts', [])
        }

        for hist in historicals:
            if hist['account_code'] in _covered:
                _other_claimants.setdefault(hist['account_code'], []).append('historical')
                continue

            # Compound with prior-month J-credits (auto-reversals of pipeline accruals).
            # Accounts like Water/Sewer (semi-annual billing) land here via Layer 3
            # because their GL descriptions lack date-range markers, but they still
            # need cumulative compounding between real invoices.
            # _net_j_credit returns 0 for accounts whose prior accrual was already
            # offset by a real invoice, so monthly contracts are unaffected.
            _hist_gl_acct = _gl_acct_lookup.get(hist['account_code'])
            _hist_j_cr    = _net_j_credit(_hist_gl_acct)
            _hist_base    = hist['estimated_amount']
            _hist_total   = _round(_hist_base + _hist_j_cr)

            if _hist_j_cr > 0:
                je_desc = (
                    f"{hist['description']} — cumulative ${_hist_total:,.0f} "
                    f"(${_hist_j_cr:,.0f} prior reversal + ${_hist_base:,.0f} est.)"
                )
            else:
                je_desc = hist['description']

            je_id = f"REC-{je_num:04d}"

            je_lines.append({
                'je_number':      je_id,
                'line':           1,
                'date':           '',
                'account_code':   hist['account_code'],
                'account_name':   hist['account_name'],
                'description':    je_desc,
                'reference':      'RECURRING',
                'debit':          _hist_total,
                'credit':         0,
                'vendor':         '[Historical Recurring]',
                'invoice_number': '',
                'source':         'historical',
            })
            _cr_acct, _cr_name = _cr_for(hist['account_code'])
            je_lines.append({
                'je_number':      je_id,
                'line':           2,
                'date':           '',
                'account_code':   _cr_acct,
                'account_name':   _cr_name,
                'description':    je_desc,
                'reference':      'RECURRING',
                'debit':          0,
                'credit':         _hist_total,
                'vendor':         '[Historical Recurring]',
                'invoice_number': '',
                'source':         'historical',
            })
            _covered.add(hist['account_code'])
            je_num += 1

    # ── Layer 4: Payroll bonus accruals ────────────────────────────────────────
    # Accrues the monthly bonus component for engineering and admin payroll
    # accounts (615110, 637110).  Two modes:
    #
    #   a) User-entered annual bonus amounts (bonus_overrides provided):
    #      bonus_overrides = {'615110': 48000.0, '637110': 24000.0}
    #      Monthly accrual = annual / 12 per account.
    #
    #   b) Kardin-derived amounts (when kardin_records provided, no override):
    #      Uses detect_payroll_bonus_accrual() — monthly avg minus standard month.
    #
    # In both modes the accrual is suppressed when the GL already shows a net
    # debit ≥ the monthly average (the actual bonus payment hit the GL).
    _bonus_month_map = {
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,
        'may': 5, 'jun': 6, 'jul': 7, 'aug': 8,
        'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
    }
    _period_month_num = 0
    for _ab, _nm in _bonus_month_map.items():
        if _ab in (period or '').lower():
            _period_month_num = _nm
            break

    # Build effective payroll accounts dict — config can override which accounts are
    # monitored for Layer 4 bonus accruals. Unknown codes get a generic label/keywords.
    _effective_payroll: dict = PAYROLL_BONUS_ACCOUNTS
    if payroll_accounts:
        _effective_payroll = {
            code: PAYROLL_BONUS_ACCOUNTS.get(code, {
                'label':           f'Payroll-{code}',
                'kardin_keywords': ['bonus', 'payroll'],
            })
            for code in payroll_accounts
        }

    # C-2: Build effective periodic contract accounts dict — config can override
    # which accounts are treated as periodic (quarterly/semi-annual) service contracts.
    # Falls back to the module-level PERIODIC_CONTRACT_ACCOUNTS constant (RevLabs
    # defaults: 617110, 619120, 627230) when no per-property override is supplied.
    _effective_periodic: dict = (
        periodic_contract_accounts
        if periodic_contract_accounts is not None
        else PERIODIC_CONTRACT_ACCOUNTS
    )

    if gl_data and _period_month_num:
        # Build GL net_change for payroll accounts
        _gl_net_bonus: Dict[str, float] = {}
        for _ba in (gl_data.accounts if hasattr(gl_data, 'accounts') else []):
            _bc = str(_ba.account_code).strip()
            if _bc in _effective_payroll:
                _gl_net_bonus[_bc] = _ba.net_change

        if bonus_overrides:
            # Mode (a): user-entered annual amounts
            for _ba_code, _ba_cfg in _effective_payroll.items():
                if _ba_code in _covered:
                    continue
                _annual = float(bonus_overrides.get(_ba_code, 0) or 0)
                if _annual <= 0:
                    continue
                _monthly = _round(_annual / 12.0)
                if _monthly < 100:
                    continue
                # Note: Mode (a) does NOT suppress based on GL net because the user
                # entered only the BONUS portion of payroll — not total payroll.
                # Comparing GL total payroll to bonus/12 would always suppress incorrectly.
                # The _covered set (checked above) handles the main suppression case.
                _bon_id = f"BON-{je_num:04d}"
                _bon_desc = (
                    f'Accrual {_fmt_period(period)} — {_ba_cfg["label"]} '
                    f'(bonus ${_annual:,.2f}/yr ÷ 12 = ${_monthly:,.2f}/mo)'
                )
                je_lines.append({
                    'je_number': _bon_id, 'line': 1, 'date': '',
                    'account_code': _ba_code,
                    'account_name': _ba_cfg['label'],
                    'description': _bon_desc,
                    'reference': 'BONUS',
                    'debit': _monthly, 'credit': 0,
                    'vendor': '[Bonus Accrual]', 'invoice_number': '',
                    'source': 'bonus_accrual', 'confidence': 'high',
                })
                _cr_acct, _cr_name = _cr_for(_ba_code)
                je_lines.append({
                    'je_number': _bon_id, 'line': 2, 'date': '',
                    'account_code': _cr_acct, 'account_name': _cr_name,
                    'description': _bon_desc,
                    'reference': 'BONUS',
                    'debit': 0, 'credit': _monthly,
                    'vendor': '[Bonus Accrual]', 'invoice_number': '',
                    'source': 'bonus_accrual', 'confidence': 'high',
                })
                _covered.add(_ba_code)
                je_num += 1

        elif kardin_records:
            # Mode (b): Kardin-derived amounts
            for _bon in detect_payroll_bonus_accrual(
                gl_data, kardin_records, _period_month_num
            ):
                if _bon['account_code'] in _covered:
                    continue
                _bon_id = f"BON-{je_num:04d}"
                _bon_desc_b = (
                    f'Accrual {_fmt_period(period)} — {_bon["account_name"]} '
                    f'(bonus: {_bon["description"]})'
                )
                je_lines.append({
                    'je_number': _bon_id, 'line': 1, 'date': '',
                    'account_code': _bon['account_code'],
                    'account_name': _bon['account_name'],
                    'description': _bon_desc_b,
                    'reference': 'BONUS',
                    'debit': _bon['estimated_amount'], 'credit': 0,
                    'vendor': '[Bonus Accrual]', 'invoice_number': '',
                    'source': 'bonus_accrual', 'confidence': 'high',
                })
                _cr_acct, _cr_name = _cr_for(_bon['account_code'])
                je_lines.append({
                    'je_number': _bon_id, 'line': 2, 'date': '',
                    'account_code': _cr_acct, 'account_name': _cr_name,
                    'description': _bon_desc_b,
                    'reference': 'BONUS',
                    'debit': 0, 'credit': _bon['estimated_amount'],
                    'vendor': '[Bonus Accrual]', 'invoice_number': '',
                    'source': 'bonus_accrual', 'confidence': 'high',
                })
                _covered.add(_bon['account_code'])
                je_num += 1

    # ── Apply multi-layer review flags ──────────────────────────────────────
    # When multiple layers detected the same account, we kept only the first-
    # layer entry but want the reviewer to know that additional layers also
    # triggered.  Add review_flag=True and review_sources=[...] to the DR line.
    if _other_claimants:
        for line in je_lines:
            if line.get('line') == 1:
                acct = str(line.get('account_code', '')).strip()
                if acct in _other_claimants:
                    line['review_flag']    = True
                    line['review_sources'] = _other_claimants[acct]

    return je_lines


# ── Prepaid amortization schedule ───────────────────────────

def build_prepaid_amortization(nexus_data: list, close_period: str = '') -> List[Dict[str, Any]]:
    """
    Build a prepaid expense amortization schedule from Nexus invoices whose
    service period spans more than one calendar month.

    For each qualifying invoice, produces one amortization line per month:
      - current period month  → expense account (normal accrual, not prepaid)
      - future months         → prepaid asset to be released in later months

    Args:
        nexus_data: Parsed Nexus records (from nexus_accrual.parse())
        close_period: Accounting period string e.g. 'Mar-2026'

    Returns:
        List of dicts:
          vendor, invoice_number, description, service_start, service_end,
          total_amount, monthly_amount, amort_month (date), period_label,
          gl_account_number, gl_account, is_current_period, month_index
    """
    lines = []

    # Parse close_period to determine current month
    close_month = None
    if close_period:
        month_map = dict(Jan=1, Feb=2, Mar=3, Apr=4, May=5, Jun=6,
                         Jul=7, Aug=8, Sep=9, Oct=10, Nov=11, Dec=12)
        for mn, mv in month_map.items():
            if mn in close_period:
                year_m = None
                import re
                yr = re.search(r'(\d{4})', close_period)
                if yr:
                    year_m = int(yr.group(1))
                if year_m:
                    close_month = date(year_m, mv, 1)
                break

    for inv in nexus_data:
        if not inv.get('is_prepaid'):
            continue

        svc_start = inv.get('service_start')
        svc_end = inv.get('service_end')
        total_months = inv.get('prepaid_months', 1)
        if not svc_start or not svc_end or total_months <= 1:
            continue

        total_amount = inv.get('amount', 0)
        monthly_amount = _round(total_amount / total_months)
        # Distribute any rounding to first month
        rounding_adj = _round(total_amount - monthly_amount * total_months)

        vendor = inv.get('vendor', '')
        inv_num = inv.get('invoice_number', '')
        desc = inv.get('line_description', '')
        gl_acct_num = inv.get('gl_account_number', inv.get('gl_account', ''))
        gl_acct = inv.get('gl_account', '')

        current_month_start = date(svc_start.year, svc_start.month, 1)
        for i in range(total_months):
            amort_month = current_month_start + relativedelta(months=i)
            month_amt = monthly_amount + (rounding_adj if i == 0 else 0)
            period_label = amort_month.strftime('%b-%Y')
            is_current = (close_month is not None and
                          amort_month.year == close_month.year and
                          amort_month.month == close_month.month)

            lines.append({
                'vendor': vendor,
                'invoice_number': inv_num,
                'description': desc,
                'service_start': svc_start,
                'service_end': svc_end,
                'total_amount': total_amount,
                'monthly_amount': month_amt,
                'amort_month': amort_month,
                'period_label': period_label,
                'gl_account_number': gl_acct_num,
                'gl_account': gl_acct,
                'is_current_period': is_current,
                'month_index': i + 1,
                'total_months': total_months,
            })

    return lines


def write_prepaid_amortization_tab(wb: Workbook, amort_lines: List[Dict],
                                   period: str = '', property_name: str = ''):
    """
    Add a 'Prepaid Amortization' tab to an existing workbook.
    Shows one row per invoice per month with current period highlighted.
    """
    ws = wb.create_sheet('Prepaid Amortization')
    AMBER = 'FFF2CC'
    GREEN_LIGHT = 'E2EFDA'

    row = 1
    c = ws.cell(row=row, column=1, value=f'Prepaid Expense Amortization Schedule — {property_name}')
    c.font = Font(name='Calibri', size=14, bold=True, color=DARK_BLUE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1

    c = ws.cell(row=row, column=1,
                value=f'Period: {period}  |  Invoices with service period > 1 month  |  Prepared: {datetime.now().strftime("%m/%d/%Y")}')
    c.font = Font(name='Calibri', size=11, italic=True, color='666666')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 2

    # Column headers
    headers = ['Vendor', 'Invoice #', 'Description', 'GL Account',
               'Total Amount', 'Service Start', 'Service End', 'Total Months',
               'Period', 'Monthly Amount']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        _apply(c, font=_hdr_font(), fill=_hdr_fill(), border=THIN_BORDER,
               align=Alignment(horizontal='center', vertical='center', wrap_text=True))
    row += 1

    # Group lines by invoice, showing all months
    for line in amort_lines:
        is_cur = line.get('is_current_period', False)
        fill_color = AMBER if is_cur else None
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid') if fill_color else None

        vals = [
            line['vendor'],
            line['invoice_number'],
            line['description'],
            f"{line['gl_account_number']} — {line['gl_account'].split('(')[0].strip()}",
            line['total_amount'] if line['month_index'] == 1 else '',  # Only show on first row
            line['service_start'].strftime('%m/%d/%Y') if line['service_start'] else '',
            line['service_end'].strftime('%m/%d/%Y') if line['service_end'] else '',
            line['total_months'] if line['month_index'] == 1 else '',
            line['period_label'] + (' ← CURRENT' if is_cur else ''),
            line['monthly_amount'],
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = THIN_BORDER
            if fill:
                c.fill = fill
            if ci == 5 and v != '':
                c.number_format = '$#,##0.00'
            if ci == 10:
                c.number_format = '$#,##0.00'
                if is_cur:
                    c.font = Font(name='Calibri', size=11, bold=True)
        row += 1

    # Summary: total current period prepaid expense
    current_total = sum(l['monthly_amount'] for l in amort_lines if l.get('is_current_period'))
    future_total = sum(l['monthly_amount'] for l in amort_lines if not l.get('is_current_period'))
    row += 1
    ws.cell(row=row, column=9, value='Current Period Total').font = Font(name='Calibri', size=11, bold=True)
    c = ws.cell(row=row, column=10, value=current_total)
    c.number_format = '$#,##0.00'
    c.font = Font(name='Calibri', size=11, bold=True)
    c.border = DOUBLE_BOTTOM
    row += 1
    ws.cell(row=row, column=9, value='Future Periods (Prepaid Asset)').font = Font(name='Calibri', size=11, italic=True)
    c = ws.cell(row=row, column=10, value=future_total)
    c.number_format = '$#,##0.00'
    c.font = Font(name='Calibri', size=11, italic=True)

    # Note explaining prepaid accounting
    row += 2
    note = (
        'Note: Current period amounts are expensed via accrual JE (DR expense / CR accrued liabilities). '
        'Future period amounts are recorded as prepaid assets (DR prepaid / CR cash) upon payment, '
        'then amortized monthly (DR expense / CR prepaid).'
    )
    c = ws.cell(row=row, column=1, value=note)
    c.font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 30

    # Column widths
    widths = [25, 15, 40, 35, 14, 14, 14, 10, 18, 16]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + ci)].width = w

    ws.sheet_properties.tabColor = 'ED7D31'  # Orange for prepaid


# ── Prepaid release JEs from ledger ─────────────────────────

def build_prepaid_release_je(ledger_amort_lines: List[Dict],
                              period: str = '',
                              je_start: int = 1) -> List[Dict]:
    """
    Convert prepaid ledger amortization lines (month 2+) into JE line dicts.

    Each entry:
      DR  [expense account]         monthly_amount   (releasing prepaid to expense)
      CR  135110 Prepaid Insurance  monthly_amount   (insurance accounts: 639110/639120)
      CR  135150 Prepaid Other      monthly_amount   (all other accounts)

    Args:
        ledger_amort_lines: from prepaid_ledger.get_current_amortization()
        period: close period string
        je_start: starting JE number (to avoid collisions with Nexus JEs)

    Returns list of JE line dicts compatible with generate_yardi_je_import()
    """
    je_lines = []
    je_num = je_start

    for item in ledger_amort_lines:
        vendor      = str(item.get('vendor', '') or '')
        inv_num     = str(item.get('invoice_number', '') or '')
        desc        = str(item.get('description', '') or '')
        gl_acct     = str(item.get('gl_account_number', '') or '')
        amount      = item.get('monthly_amount', 0) or 0
        period_lbl  = item.get('period_label', period)
        month_idx   = item.get('month_index', '')
        total_mo    = item.get('total_months', '')

        if amount == 0:
            continue

        # Insurance expense accounts (639110/639120) offset against 135110 Prepaid
        # Insurance (includes property, GL, and umbrella policies).
        # All other prepaids offset against 135150 Prepaid Other.
        if gl_acct in _INSURANCE_EXPENSE_ACCTS:
            cr_account = _PREPAID_INSURANCE_ACCT   # 135110
            cr_name    = 'Prepaid Insurance'
        else:
            cr_account = PREPAID_ASSET_ACCOUNT     # 135150
            cr_name    = PREPAID_ASSET_NAME

        je_id   = f"PPD-{je_num:04d}"
        je_desc = (
            f"Accrual {_fmt_period(period_lbl)} — {vendor} "
            f"(prepaid amortization #{inv_num}, mo {month_idx}/{total_mo})"
        )

        # DR: Expense account
        je_lines.append({
            'je_number':      je_id,
            'line':           1,
            'date':           period_lbl,
            'account_code':   gl_acct,
            'account_name':   desc[:40],
            'description':    je_desc,
            'reference':      inv_num,
            'debit':          abs(amount),
            'credit':         0,
            'vendor':         vendor,
            'invoice_number': inv_num,
            'source':         'prepaid_ledger',
        })
        # CR: Prepaid asset (135110 for insurance, 135150 for all others)
        je_lines.append({
            'je_number':      je_id,
            'line':           2,
            'date':           period_lbl,
            'account_code':   cr_account,
            'account_name':   cr_name,
            'description':    je_desc,
            'reference':      inv_num,
            'debit':          0,
            'credit':         abs(amount),
            'vendor':         vendor,
            'invoice_number': inv_num,
            'source':         'prepaid_ledger',
        })
        je_num += 1

    return je_lines


# ── Insurance escrow reconciliation JE ───────────────────────

def build_insurance_escrow_je(
    gl_data,
    loan_data,
    period: str = '',
) -> List[Dict]:
    """
    Generate a reconciling JE to bring GL 135110 (Restricted Insurance) in line
    with the Berkadia loan statement insurance escrow balance.

    If the Berkadia statement shows a different balance than the GL ending balance
    for 135110, this entry adjusts the GL to match.

    JE (GL < Berkadia, escrow grew):   DR 135110 Restricted Insurance / CR 115300 Insurance Escrow
    JE (GL > Berkadia, escrow shrank): DR 115300 Insurance Escrow / CR 135110 Restricted Insurance

    Suppressed if |difference| < $1.00 (no material reconciling item).
    For months with no escrow activity (e.g. March) this returns [].

    Args:
        gl_data:   GLParseResult from yardi_gl parser
        loan_data: Parsed Berkadia loan data (list of dicts or single dict)
        period:    Close period string e.g. 'Apr-2026'

    Returns list of JE line dicts compatible with generate_yardi_je_csv()
    """
    if not gl_data or not loan_data:
        return []

    # Get GL ending balance for 135110
    gl_bal = 0.0
    for acct in (gl_data.accounts if hasattr(gl_data, 'accounts') else []):
        if str(acct.account_code).strip() == '135110':
            gl_bal = acct.ending_balance
            break

    # Sum insurance_escrow_balance across all Berkadia loans
    loans = loan_data if isinstance(loan_data, list) else [loan_data]
    berkadia_bal = 0.0
    for ln in loans:
        if isinstance(ln, dict):
            berkadia_bal += _safe_float(ln.get('insurance_escrow_balance', 0))
        else:
            berkadia_bal += _safe_float(getattr(ln, 'insurance_escrow_balance', 0))

    diff = berkadia_bal - gl_bal   # positive = GL needs to go up; negative = GL needs to go down

    if abs(diff) < 1.0:
        return []

    amount = _round(abs(diff))
    je_id  = 'INS-001'
    je_desc = (
        f'Insurance escrow recon — GL 135110 ${gl_bal:,.2f} vs '
        f'Berkadia ${berkadia_bal:,.2f} (adj ${diff:+,.2f})'
    )

    if diff > 0:
        # GL ending balance too low — debit 135110, credit 115300 Insurance Escrow
        dr_acct, dr_name = '135110', 'Restricted Insurance'
        cr_acct, cr_name = '115300', 'Insurance Escrow'
    else:
        # GL ending balance too high — debit 115300 Insurance Escrow, credit 135110
        dr_acct, dr_name = '115300', 'Insurance Escrow'
        cr_acct, cr_name = '135110', 'Restricted Insurance'

    return [
        {
            'je_number':      je_id,
            'line':           1,
            'date':           period,
            'account_code':   dr_acct,
            'account_name':   dr_name,
            'description':    je_desc,
            'reference':      'INS-ESCROW',
            'debit':          amount,
            'credit':         0,
            'source':         'insurance_escrow',
        },
        {
            'je_number':      je_id,
            'line':           2,
            'date':           period,
            'account_code':   cr_acct,
            'account_name':   cr_name,
            'description':    je_desc,
            'reference':      'INS-ESCROW',
            'debit':          0,
            'credit':         amount,
            'source':         'insurance_escrow',
        },
    ]


# ── Generate Yardi JE import file ────────────────────────────

def generate_yardi_je_import(je_lines: List[Dict], output_path: str,
                              period: str = '', property_name: str = '') -> str:
    """
    Generate a Yardi-compatible journal entry import file (Excel).

    Yardi JE import expects columns:
      Property, Journal #, Date, Account, Description, Reference, Debit, Credit

    Args:
        je_lines: List of JE line dicts from build_accrual_entries()
        output_path: Where to write the Excel file
        period: Accounting period
        property_name: Property code/name

    Returns:
        Output path
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Journal Entries'

    # Header row
    headers = ['Property', 'Journal #', 'Date', 'Account', 'Description',
               'Reference', 'Debit', 'Credit']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        _apply(c, font=_hdr_font(), fill=_hdr_fill(), border=THIN_BORDER,
               align=Alignment(horizontal='center', vertical='center'))

    # Data rows
    prop_code = property_name.split()[0] if property_name else 'REVLABS'

    for ri, line in enumerate(je_lines, 2):
        alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid') if ri % 2 == 0 else None

        ws.cell(row=ri, column=1, value=prop_code)
        ws.cell(row=ri, column=2, value=line['je_number'])
        ws.cell(row=ri, column=3, value=line['date'])
        ws.cell(row=ri, column=4, value=line['account_code'])
        ws.cell(row=ri, column=5, value=line['description'])
        ws.cell(row=ri, column=6, value=line['reference'])

        c_dr = ws.cell(row=ri, column=7, value=line['debit'])
        c_dr.number_format = '$#,##0.00'
        c_cr = ws.cell(row=ri, column=8, value=line['credit'])
        c_cr.number_format = '$#,##0.00'

        for ci in range(1, 9):
            ws.cell(row=ri, column=ci).border = THIN_BORDER
            if alt_fill:
                ws.cell(row=ri, column=ci).fill = alt_fill

    # Totals row
    total_row = len(je_lines) + 2
    ws.cell(row=total_row, column=6, value='TOTAL').font = Font(name='Calibri', size=11, bold=True)
    total_dr = sum(l['debit'] for l in je_lines)
    total_cr = sum(l['credit'] for l in je_lines)
    c_dr = ws.cell(row=total_row, column=7, value=total_dr)
    c_dr.number_format = '$#,##0.00'
    c_dr.font = Font(name='Calibri', size=11, bold=True)
    c_dr.border = DOUBLE_BOTTOM
    c_cr = ws.cell(row=total_row, column=8, value=total_cr)
    c_cr.number_format = '$#,##0.00'
    c_cr.font = Font(name='Calibri', size=11, bold=True)
    c_cr.border = DOUBLE_BOTTOM

    # Validation check
    balance_row = total_row + 1
    ws.cell(row=balance_row, column=6, value='Balance Check').font = Font(name='Calibri', size=10, italic=True)
    diff = total_dr - total_cr
    c_bal = ws.cell(row=balance_row, column=7, value=diff)
    c_bal.number_format = '$#,##0.00'
    c_bal.font = Font(name='Calibri', size=10, italic=True,
                      color='008000' if abs(diff) < 0.01 else 'FF0000')

    # Auto column widths
    for col in range(1, 9):
        letter = chr(64 + col)
        best = 12
        for cell in ws[letter]:
            try:
                if cell.value:
                    best = max(best, len(str(cell.value)) + 2)
            except:
                pass
        ws.column_dimensions[letter].width = min(best, 45)

    wb.save(output_path)
    return output_path


# ── Generate Yardi CSV import (exact Yardi format) ────────────

def generate_yardi_je_csv(je_lines: List[Dict], output_path: str,
                           period: str = '', property_code: str = '',
                           book: str = '') -> str:
    """
    Generate a Yardi-compatible journal entry import CSV.

    17-column format confirmed from working Yardi import example:
      J, batch#, , , date, date, , description, property_code, signed_amount,
      gl_account, , , ref_num, , , Standard Journal Display Type

    Col  1: J (transaction type)
    Col  2: batch number (sequential integer per unique JE)
    Col  3: empty (book — blank uses Yardi default)
    Col  4: empty
    Col  5: period end date (MM/DD/YYYY)
    Col  6: period end date (MM/DD/YYYY)
    Col  7: empty
    Col  8: description (max 60 chars)
    Col  9: property code
    Col 10: signed amount (positive = DR, negative = CR)
    Col 11: GL account code
    Col 12: empty
    Col 13: empty
    Col 14: reference number (numeric — same as batch)
    Col 15: empty
    Col 16: empty
    Col 17: Standard Journal Display Type

    Args:
        je_lines:      List of JE line dicts from build_accrual_entries()
        output_path:   Where to write the .csv file
        period:        Accounting period label (e.g. 'Mar-2026') — used to derive date
        property_code: Yardi property code (default '' — must be passed by caller)
        book:          Unused — kept for signature compatibility (Yardi uses blank)

    Returns:
        output_path
    """
    import csv
    from datetime import datetime, date
    from calendar import monthrange

    # Derive period end date from period string (e.g. 'Mar-2026' → 03/31/2026)
    period_date = ''
    try:
        dt = datetime.strptime(period, '%b-%Y')
        last_day = monthrange(dt.year, dt.month)[1]
        period_date = date(dt.year, dt.month, last_day).strftime('%m/%d/%Y')
    except Exception:
        period_date = datetime.now().strftime('%m/%d/%Y')

    # Assign sequential batch numbers per unique JE
    batch_map = {}
    batch_counter = 1
    for line in je_lines:
        je_num = line.get('je_number', '')
        if je_num not in batch_map:
            batch_map[je_num] = batch_counter
            batch_counter += 1

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for line in je_lines:
            je_num  = line.get('je_number', '')
            batch   = batch_map.get(je_num, 1)
            desc    = str(line.get('description', '') or '')[:60]
            gl_acct = str(line.get('account_code', '') or '')
            ref     = str(line.get('reference', '') or je_num)
            debit   = line.get('debit', 0) or 0
            credit  = line.get('credit', 0) or 0
            # Signed amount: positive = DR, negative = CR
            amount  = round(debit - credit, 2)

            writer.writerow([
                'J',                             # col 1:  transaction type
                batch,                           # col 2:  batch number (integer)
                '',                              # col 3:  book (blank = Yardi default)
                '',                              # col 4:  empty
                period_date,                     # col 5:  reference date
                period_date,                     # col 6:  period date
                '',                              # col 7:  empty
                desc,                            # col 8:  description (quoted by csv.writer if contains comma — Yardi accepts this)
                property_code,                   # col 9:  property code
                amount,                          # col 10: signed amount
                gl_acct,                         # col 11: GL account
                '',                              # col 12: empty
                '',                              # col 13: empty
                1000,                            # col 14: fixed numeric reference (Yardi requirement)
                ref,                             # col 15: text reference code (e.g. INS-AMORT, BUDGET-GAP)
                '',                              # col 16: empty
                'Standard Journal Display Type', # col 17: required by Yardi
            ])

    return output_path


# ── Generate ETL FinJournals CSV (Yardi ETL import format) ───────────────────

# Full 65-column header list from ETL_Financial_FinJournals.xls → Sample_CSV tab
_ETL_HEADERS = [
    'TRANNUM', 'DATE', 'PROPERTY', 'ACCOUNT', 'POSTMONTH', 'BOOKNUM', 'AMOUNT',
    'REMARK', 'REF', 'USERDEFINEDFIELD1', 'USERDEFINEDFIELD2', 'USERDEFINEDFIELD3',
    'USERDEFINEDFIELD4', 'USERDEFINEDFIELD5', 'USERDEFINEDFIELD6', 'USERDEFINEDFIELD7',
    'USERDEFINEDFIELD8', 'TAXPOINTDATE', 'DESC', 'DOCUMENTSEQUENCENUMBER', 'TRANAMOUNT',
    'DETAILTRANAMOUNT', 'LEGALENTITYID', 'BASECURRENCYID', 'TRANCURRENCYID', 'EXCHANGERATE',
    'EXCHANGERATEDATE', 'EXCHANGERATE2', 'EXCHANGERATEDATE2', 'AMOUNT2', 'FROMDATE',
    'TODATE', 'SEGMENT1', 'SEGMENT2', 'SEGMENT3', 'SEGMENT4', 'SEGMENT5', 'SEGMENT6',
    'SEGMENT7', 'SEGMENT8', 'SEGMENT9', 'SEGMENT10', 'SEGMENT11', 'SEGMENT12', 'TAXAMOUNT1',
    'TAXAMOUNT2', 'DETAILVATTRANTYPEID', 'DETAILVATRATEID', 'VATTRANTYPEID', 'VATRateId',
    'DISPLAYTYPE', 'JOB', 'CATEGORY', 'CONTRACT', 'COSTCODE', 'NOTES2', 'DETAILFIELD1',
    'DETAILFIELD2', 'DETAILFIELD3', 'DETAILFIELD4', 'DETAILFIELD5', 'DETAILFIELD6',
    'DETAILFIELD7', 'DETAILFIELD8', 'ReverseNextMonth',
]
# Column index map (0-based)
_ETL_IDX = {h: i for i, h in enumerate(_ETL_HEADERS)}


def generate_etl_csv(je_lines: List[Dict], output_path: str,
                     period: str = '', property_code: str = '',
                     book: str = '', auto_reverse: bool = False) -> str:
    """
    Generate a Yardi ETL FinJournals import CSV.

    Format (from ETL_Financial_FinJournals.xls → Sample_CSV):
      Row 1 : 'FinJournals' in col A, rest blank   (required record-type identifier)
      Row 2 : 65 column headers
      Row 3+ : one row per JE line

    Populated columns:
      TRANNUM        (A)  — batch number (integer, same per JE)
      DATE           (B)  — period end date MM/DD/YYYY
      PROPERTY       (C)  — Yardi property code
      ACCOUNT        (D)  — GL account code
      POSTMONTH      (E)  — period end date MM/DD/YYYY
      BOOKNUM        (F)  — 1 (Accrual book)
      AMOUNT         (G)  — signed amount (positive = DR, negative = CR)
      REMARK         (H)  — JE description (optional; also in DESC)
      REF            (I)  — reference number / JE code
      DESC           (S)  — line description
      DISPLAYTYPE    (AY) — 'Standard Journal Display Type'
      ReverseNextMonth(BM)— -1 (auto-reverse next period) or 0 (no reversal)

    Args:
        je_lines:      List of JE line dicts from build_accrual_entries()
        output_path:   Where to write the .csv file
        period:        Accounting period label e.g. 'Jan-2026' — used to derive date
        property_code: Yardi property code (default '' — must be passed by caller)
        book:          Unused — kept for signature compatibility
        auto_reverse:  Deprecated — kept for signature compatibility but ignored.
                       BM is now determined per-batch: -1 if the batch contains any
                       line posting to 213100 (Accrued Expenses); 0 otherwise.
                       This ensures only true accrual entries auto-reverse, while
                       prepaid, management fee, and reclassification batches do not.
                       Per-line override: set 'reverse_next_month' key on the dict.

    Returns:
        output_path
    """
    import csv
    from datetime import datetime, date
    from calendar import monthrange

    # Derive period end date from period string (e.g. 'Jan-2026' → 01/31/2026).
    # Try several format variants before giving up — avoids writing today's date
    # when the period string is in an unexpected but parseable format.
    period_date = ''
    for _pfmt in ('%b-%Y', '%B-%Y', '%b %Y', '%B %Y', '%m-%Y', '%m/%Y'):
        try:
            dt = datetime.strptime(period.strip(), _pfmt)
            last_day = monthrange(dt.year, dt.month)[1]
            period_date = date(dt.year, dt.month, last_day).strftime('%m/%d/%Y')
            break
        except Exception:
            pass
    if not period_date:
        import warnings as _warn_etl
        _warn_etl.warn(
            f"generate_etl_csv: could not parse period string '{period}' — "
            f"DATE and POSTMONTH will use today's date, which will post to the wrong "
            f"accounting period in Yardi. Pass a period in 'Mon-YYYY' format.",
            UserWarning, stacklevel=2,
        )
        period_date = datetime.now().strftime('%m/%d/%Y')

    # Assign sequential batch numbers per unique JE number
    batch_map: dict = {}
    batch_counter = 1
    for line in je_lines:
        je_num = line.get('je_number', '')
        if je_num not in batch_map:
            batch_map[je_num] = batch_counter
            batch_counter += 1

    # Pre-scan: which JE batches contain at least one line posting to 213100?
    # Only those batches are true accruals that should auto-reverse next month.
    # All other batches (prepaid releases, mgmt fee, reclasses, etc.) get BM = 0.
    _batches_with_213100: set = {
        line.get('je_number', '')
        for line in je_lines
        if str(line.get('account_code', '') or '').strip() == '213100'
    }

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Row 1: record-type identifier
        row1 = [''] * 65
        row1[0] = 'FinJournals'
        writer.writerow(row1)

        # Row 2: column headers
        writer.writerow(_ETL_HEADERS)

        # Data rows (one per JE line)
        for line in je_lines:
            je_num  = line.get('je_number', '')
            batch   = batch_map.get(je_num, 1)
            desc    = str(line.get('description', '') or '')[:60]
            gl_acct = str(line.get('account_code', '') or '')
            ref     = str(line.get('reference', '') or je_num)
            debit   = line.get('debit', 0) or 0
            credit  = line.get('credit', 0) or 0
            amount  = round(debit - credit, 2)  # positive = DR, negative = CR

            # BM: -1 only if this batch touches 213100 (Accrued Expenses).
            # Per-line 'reverse_next_month' key overrides batch-level logic.
            bm_batch = -1 if je_num in _batches_with_213100 else 0
            bm = line.get('reverse_next_month', bm_batch)

            row = [''] * 65
            row[_ETL_IDX['TRANNUM']]          = batch
            row[_ETL_IDX['DATE']]             = period_date
            row[_ETL_IDX['PROPERTY']]         = property_code
            row[_ETL_IDX['ACCOUNT']]          = gl_acct
            row[_ETL_IDX['POSTMONTH']]        = period_date
            row[_ETL_IDX['BOOKNUM']]          = 1
            row[_ETL_IDX['AMOUNT']]           = amount
            row[_ETL_IDX['REMARK']]           = desc
            row[_ETL_IDX['REF']]              = ref
            row[_ETL_IDX['DESC']]             = desc
            row[_ETL_IDX['DISPLAYTYPE']]      = 'Standard Journal Display Type'
            row[_ETL_IDX['ReverseNextMonth']] = bm

            writer.writerow(row)

    return output_path


def build_reversing_je_csv(
    source_etl_path: str,
    next_period: str,
    output_path: str,
    property_code: str = '',
) -> str:
    """
    Generate a Yardi ETL import CSV that reverses all entries in an existing
    accrual JE CSV.  Use this to manually post reversals if Yardi auto-reversal
    fails, or to pre-review what Yardi will reverse on the 1st of next month.

    Strategy:
      - Read source ETL CSV (skip rows 1–2: 'FinJournals' header and column header)
      - Flip every AMOUNT sign (positive DR → negative CR, and vice versa)
      - Update DATE / POSTMONTH to the last day of next_period
      - Prefix REMARK / DESC with 'REV - '
      - Append '-REV' to REF so the batch is identifiable in Yardi
      - Set ReverseNextMonth = 0  (reversals don't themselves auto-reverse)
      - Re-sequence TRANNUM from 1

    Args:
        source_etl_path:  Path to an existing GA ETL accruals CSV
        next_period:      Period for the reversals e.g. 'Feb-2026'
        output_path:      Where to write the reversing CSV
        property_code:    Override Yardi property code (default: preserve from source)

    Returns:
        output_path
    """
    import csv as _csv_rev
    from calendar import monthrange as _monthrange_rev

    # Derive last-day date for next_period
    next_date = ''
    try:
        _dt_rev = datetime.strptime(next_period, '%b-%Y')
        _last_rev = _monthrange_rev(_dt_rev.year, _dt_rev.month)[1]
        from datetime import date as _date_rev
        next_date = _date_rev(_dt_rev.year, _dt_rev.month, _last_rev).strftime('%m/%d/%Y')
    except Exception:
        pass

    # Index shortcuts
    _ti  = _ETL_IDX['TRANNUM']
    _di  = _ETL_IDX['DATE']
    _pi  = _ETL_IDX['POSTMONTH']
    _pri = _ETL_IDX['PROPERTY']
    _ai  = _ETL_IDX['AMOUNT']
    _ri  = _ETL_IDX['REMARK']
    _rfi = _ETL_IDX['REF']
    _dsi = _ETL_IDX['DESC']
    _bmi = _ETL_IDX['ReverseNextMonth']

    # Read source, skip the two header rows
    rows_in: List[List] = []
    try:
        with open(source_etl_path, newline='', encoding='utf-8') as _fh_rev:
            for i, row in enumerate(_csv_rev.reader(_fh_rev)):
                if i < 2 or not any(row):
                    continue
                rows_in.append(list(row))
    except Exception:
        rows_in = []

    def _write_header(w):
        r1 = [''] * 65
        r1[0] = 'FinJournals'
        w.writerow(r1)
        w.writerow(_ETL_HEADERS)

    with open(output_path, 'w', newline='', encoding='utf-8') as _fh_out:
        _w = _csv_rev.writer(_fh_out)
        _write_header(_w)

        if not rows_in:
            return output_path

        # Map original TRANNUM values to new sequential batch numbers
        _batch_map: Dict[str, int] = {}
        _bc = 1
        for _row in rows_in:
            _ob = _row[_ti] if len(_row) > _ti else ''
            if _ob not in _batch_map:
                _batch_map[_ob] = _bc
                _bc += 1

        for _row in rows_in:
            r = _row + [''] * max(0, 65 - len(_row))   # pad to 65 cols

            # Re-batch
            r[_ti] = _batch_map.get(r[_ti], 1)

            # Update dates
            if next_date:
                r[_di] = next_date
                r[_pi] = next_date

            # Override property code
            if property_code:
                r[_pri] = property_code

            # Flip amount sign
            try:
                r[_ai] = round(-float(r[_ai]), 2)
            except (ValueError, TypeError):
                pass

            # Prefix descriptions with 'REV - '
            for _idx in (_ri, _dsi):
                try:
                    _v = str(r[_idx])
                    if _v and not _v.startswith('REV - '):
                        r[_idx] = f'REV - {_v}'[:60]
                except (IndexError, TypeError):
                    pass

            # Suffix REF with '-REV'
            try:
                _rv = str(r[_rfi])
                if _rv and not _rv.endswith('-REV'):
                    r[_rfi] = f'{_rv}-REV'[:30]
            except (IndexError, TypeError):
                pass

            # No auto-reversal on the reversals themselves
            r[_bmi] = 0

            _w.writerow(r)

    return output_path


# ── Add review tab to workpapers ─────────────────────────────

def write_accrual_entries_workpaper_tab(wb: Workbook, je_lines: List[Dict],
                                         period: str = '', property_name: str = ''):
    """
    Add an 'Accrual Entries' review tab to an existing workbook.
    Shows JE detail with DR/CR, grouped by vendor, for review before posting.

    Args:
        wb: Existing workbook to add the tab to
        je_lines: List of JE line dicts from build_accrual_entries()
        period: Accounting period
        property_name: Property name
    """
    ws = wb.create_sheet('Accrual Entries')

    # Title
    row = 1
    c = ws.cell(row=row, column=1, value=f'Accrual Journal Entries — {property_name}')
    c.font = Font(name='Calibri', size=14, bold=True, color=DARK_BLUE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1

    c = ws.cell(row=row, column=1,
                value=f'Period: {period}  |  CR Account: {AP_ACCRUAL_ACCOUNT} {AP_ACCRUAL_NAME}  |  Prepared: {datetime.now().strftime("%m/%d/%Y")}')
    c.font = Font(name='Calibri', size=11, italic=True, color='666666')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1

    # Summary
    total_entries = len([l for l in je_lines if l['line'] == 1])
    total_amount = sum(l['debit'] for l in je_lines)
    c = ws.cell(row=row, column=1,
                value=f'Total Entries: {total_entries}  |  Total Amount: ${total_amount:,.2f}')
    c.font = Font(name='Calibri', size=11, bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 2

    # Column headers
    headers = ['JE #', 'Line', 'Vendor', 'Invoice #', 'Date',
               'Account', 'Description', 'Debit', 'Credit']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        _apply(c, font=_hdr_font(), fill=_hdr_fill(), border=THIN_BORDER,
               align=Alignment(horizontal='center', vertical='center', wrap_text=True))
    row += 1

    # Data rows
    current_je = None
    for i, line in enumerate(je_lines):
        alt = (i // 2) % 2 == 1  # Alternate every JE pair
        fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid') if alt else None

        # JE group separator
        if line['je_number'] != current_je:
            current_je = line['je_number']

        ws.cell(row=row, column=1, value=line['je_number'])
        ws.cell(row=row, column=2, value=line['line'])
        ws.cell(row=row, column=3, value=line['vendor'] if line['line'] == 1 else '')
        ws.cell(row=row, column=4, value=line['invoice_number'] if line['line'] == 1 else '')
        ws.cell(row=row, column=5, value=line['date'] if line['line'] == 1 else '')
        ws.cell(row=row, column=6, value=line['account_code'])

        # Shorten description for CR line
        desc = line['description'] if line['line'] == 1 else f"  CR {AP_ACCRUAL_ACCOUNT}"
        ws.cell(row=row, column=7, value=desc)

        c_dr = ws.cell(row=row, column=8, value=line['debit'] if line['debit'] > 0 else '')
        if line['debit'] > 0:
            c_dr.number_format = '$#,##0.00'

        c_cr = ws.cell(row=row, column=9, value=line['credit'] if line['credit'] > 0 else '')
        if line['credit'] > 0:
            c_cr.number_format = '$#,##0.00'

        for ci in range(1, 10):
            ws.cell(row=row, column=ci).border = THIN_BORDER
            if fill:
                ws.cell(row=row, column=ci).fill = fill

        row += 1

    # Totals
    row += 1
    ws.cell(row=row, column=7, value='TOTAL').font = Font(name='Calibri', size=11, bold=True)
    total_dr = sum(l['debit'] for l in je_lines)
    total_cr = sum(l['credit'] for l in je_lines)
    c_dr = ws.cell(row=row, column=8, value=total_dr)
    c_dr.number_format = '$#,##0.00'
    c_dr.font = Font(name='Calibri', size=11, bold=True)
    c_dr.border = DOUBLE_BOTTOM
    c_cr = ws.cell(row=row, column=9, value=total_cr)
    c_cr.number_format = '$#,##0.00'
    c_cr.font = Font(name='Calibri', size=11, bold=True)
    c_cr.border = DOUBLE_BOTTOM

    # Balance check
    row += 1
    diff = total_dr - total_cr
    ws.cell(row=row, column=7, value='Balance Check').font = Font(name='Calibri', size=10, italic=True)
    c = ws.cell(row=row, column=8, value=diff)
    c.number_format = '$#,##0.00'
    c.font = Font(name='Calibri', size=10, bold=True,
                  color='008000' if abs(diff) < 0.01 else 'FF0000')

    # Account summary section
    row += 3
    c = ws.cell(row=row, column=1, value='Account Summary')
    c.font = Font(name='Calibri', size=12, bold=True, color=DARK_BLUE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    sum_headers = ['Account Code', 'Description', 'Total Debit', 'Entry Count']
    for ci, h in enumerate(sum_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        _apply(c, font=Font(name='Calibri', size=11, bold=True, color='000000'),
               fill=_subhdr_fill(), border=THIN_BORDER)
    row += 1

    # Aggregate by GL account (DR side only)
    acct_totals = {}
    for line in je_lines:
        if line['debit'] > 0:
            code = line['account_code']
            if code not in acct_totals:
                acct_totals[code] = {'name': line['account_name'], 'total': 0, 'count': 0}
            acct_totals[code]['total'] += line['debit']
            acct_totals[code]['count'] += 1

    for code, data in sorted(acct_totals.items()):
        ws.cell(row=row, column=1, value=code); ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=data['name']); ws.cell(row=row, column=2).border = THIN_BORDER
        c = ws.cell(row=row, column=3, value=data['total'])
        c.number_format = '$#,##0.00'
        c.border = THIN_BORDER
        ws.cell(row=row, column=4, value=data['count']); ws.cell(row=row, column=4).border = THIN_BORDER
        row += 1

    # Auto-width
    for col in range(1, 10):
        letter = chr(64 + col) if col <= 26 else 'A'
        best = 12
        for cell in ws[letter]:
            try:
                if cell.value:
                    best = max(best, len(str(cell.value)) + 2)
            except:
                pass
        ws.column_dimensions[letter].width = min(best, 50)

    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['G'].width = 45
    ws.sheet_properties.tabColor = '7030A0'  # Purple for accrual entries


# ── Prior-Month Accrual vs Actual Check ──────────────────────────────────────

def check_prior_accrual_vs_actual(gl_data) -> List[Dict[str, Any]]:
    """
    Identify prior-period accruals that auto-reversed into the current period
    and compare them against actual invoices received.

    Detection method:
      Auto-reversals appear in Yardi as J-type journal entries that:
        • Credit an expense account  (reversing the prior DR expense)
        • Debit account 213100       (reversing the prior CR accrued liability)
      Any J-type JE matching that pattern is treated as an auto-reversal.

    For each reversed expense account the function also sums non-J-type debits
    (P-type checks, K-type PCard) as the "actual invoices received" amount.

    Tolerance: max($250, 5% of reversal) — entries within tolerance are MATCHED.

    Returns a list sorted by account code.  Each dict has:
      account_code    str   — e.g. '613110'
      account_name    str   — e.g. 'Electricity'
      reversal_amount float — absolute value of J-type credit(s) on this account
      actual_amount   float — sum of non-J-type debits on this account
      variance        float — actual − reversal  (+ = overbilled, − = underbilled)
      status          str   — 'MATCHED' | 'NOT YET BILLED' | 'PARTIAL' | 'OVER INVOICED'
      je_refs         str   — comma-separated J-type control numbers involved
    """
    if not gl_data or not hasattr(gl_data, 'all_transactions'):
        return []

    # ── Step 1: group all transactions by J-type control number ──────────────
    je_lines_by_ctrl: Dict[str, List[Any]] = {}
    for txn in gl_data.all_transactions:
        ctrl = (txn.control or '').strip().upper()
        if ctrl.startswith('J-'):
            je_lines_by_ctrl.setdefault(ctrl, []).append(txn)

    # ── Step 2: identify reversal JE control numbers ──────────────────────────
    # A reversal JE has at least one line that debits 213100 (accrued expenses).
    reversal_ctrl_nums: set = set()
    for ctrl, lines in je_lines_by_ctrl.items():
        for txn in lines:
            if txn.account_code == '213100' and txn.debit > 0:
                reversal_ctrl_nums.add(ctrl)
                break

    if not reversal_ctrl_nums:
        return []

    # ── Step 3: collect expense-account credits from reversal JEs ────────────
    # key: account_code → {account_name, reversal_amount, je_refs}
    reversal_by_acct: Dict[str, Dict[str, Any]] = {}
    for ctrl in reversal_ctrl_nums:
        for txn in je_lines_by_ctrl[ctrl]:
            if not is_expense_account(txn.account_code):
                continue
            if txn.credit <= 0:
                continue
            acc = txn.account_code
            if acc not in reversal_by_acct:
                reversal_by_acct[acc] = {
                    'account_name':    txn.account_name,
                    'reversal_amount': 0.0,
                    'je_refs':         set(),
                }
            reversal_by_acct[acc]['reversal_amount'] += txn.credit
            reversal_by_acct[acc]['je_refs'].add(ctrl)

    if not reversal_by_acct:
        return []

    # ── Step 4: sum non-J-type debits as "actual invoices received" ──────────
    actual_by_acct: Dict[str, float] = {acc: 0.0 for acc in reversal_by_acct}
    for txn in gl_data.all_transactions:
        if txn.account_code not in actual_by_acct:
            continue
        ctrl = (txn.control or '').strip().upper()
        if ctrl.startswith('J-'):
            continue                     # skip all journal entries
        if txn.debit > 0:
            actual_by_acct[txn.account_code] += txn.debit

    # ── Step 5: build result rows ─────────────────────────────────────────────
    _MATCH_TOL_FLOOR = 250.0    # $250 minimum tolerance
    _MATCH_TOL_PCT   = 0.05     # 5% of reversal amount

    results: List[Dict[str, Any]] = []
    for acc in sorted(reversal_by_acct):
        info     = reversal_by_acct[acc]
        reversal = round(info['reversal_amount'], 2)
        actual   = round(actual_by_acct.get(acc, 0.0), 2)
        variance = round(actual - reversal, 2)

        tol = max(_MATCH_TOL_FLOOR, reversal * _MATCH_TOL_PCT)

        if actual == 0.0:
            status = 'NOT YET BILLED'
        elif abs(variance) <= tol:
            status = 'MATCHED'
        elif variance > 0:
            status = 'OVER INVOICED'
        else:
            status = 'PARTIAL'

        results.append({
            'account_code':    acc,
            'account_name':    info['account_name'],
            'reversal_amount': reversal,
            'actual_amount':   actual,
            'variance':        variance,
            'status':          status,
            'je_refs':         ', '.join(sorted(info['je_refs'])),
        })

    return results

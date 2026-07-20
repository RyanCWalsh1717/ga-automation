"""
May 2026 — Code Architecture Tests
Tests logic of recent changes using synthetic data (no real files needed).

Covers:
  1. BC parser: sub-line item carry-forward (account code inheritance)
  2. Insurance amortization: Mode A (config), Mode B (Kardin), Mode C (BC)
  3. Reversal flags: prepaid amortization must NOT auto-reverse; accruals must
  4. generate_etl_csv: per-line reverse_next_month override respected
  5. PropertyConfig: insurance_policies loaded from config.yaml
"""
import sys, os, traceback, tempfile

sys.path.insert(0, r'C:\Users\RyanCWalsh\.claude\ga-automation\pipeline')
sys.path.insert(0, r'C:\Users\RyanCWalsh\.claude\ga-automation')

PASS = 0; FAIL = 0

def ok(label):
    global PASS; PASS += 1
    print(f'  PASS  {label}')

def fail(label, detail=''):
    global FAIL; FAIL += 1
    print(f'  FAIL  {label}')
    if detail:
        print(f'        {detail}')


# ── 1. BC PARSER: sub-line item carry-forward ────────────────────────────────
print('\n=== 1. BC PARSER: sub-line item carry-forward ===')
try:
    import openpyxl
    from parsers.yardi_budget_comparison import parse as parse_bc

    # Build a synthetic BC workbook with two sub-line items under 639110
    wb = openpyxl.Workbook()
    ws = wb.active

    # Rows 1-4: metadata
    ws.cell(1, 1).value = 'Property = revlabspm  Revolution Labs Owner, LLC'
    ws.cell(2, 1).value = 'Budget Comparison (Accrual)'
    ws.cell(3, 1).value = 'Period = May-2026'
    ws.cell(4, 1).value = 'Book = Accrual; Tree = Standard'

    # Row 5: headers
    headers = ['Account Code', 'Account Name',
               'PTD Actual', 'PTD Budget', 'PTD Variance', 'PTD % Var',
               'YTD Actual', 'YTD Budget', 'YTD Variance', 'YTD % Var', 'Annual']
    for c, h in enumerate(headers, 1):
        ws.cell(5, c).value = h

    # Row 6: parent row — 639110 Insurance
    ws.cell(6, 1).value = '639110'
    ws.cell(6, 2).value = 'Insurance-Property'
    ws.cell(6, 3).value = 4704.68  # PTD Actual (sum of both policies)
    ws.cell(6, 4).value = 4704.68  # PTD Budget

    # Row 7: sub-line — Property Insurance (no account code in col A)
    ws.cell(7, 1).value = None
    ws.cell(7, 2).value = 'Property Insurance'
    ws.cell(7, 3).value = 4037.78
    ws.cell(7, 4).value = 4037.78

    # Row 8: sub-line — Umbrella Policy (no account code in col A)
    ws.cell(8, 1).value = None
    ws.cell(8, 2).value = 'Umbrella Policy'
    ws.cell(8, 3).value = 666.90
    ws.cell(8, 4).value = 666.90

    # Row 9: unrelated account — should NOT inherit 639110
    ws.cell(9, 1).value = '637130'
    ws.cell(9, 2).value = 'Admin-Management Fees'
    ws.cell(9, 3).value = 17500.0
    ws.cell(9, 4).value = 17500.0

    # Row 10: blank spacer — should be skipped
    # (all None)

    # Row 11: total row — should be skipped
    ws.cell(11, 1).value = None
    ws.cell(11, 2).value = 'Total Insurance'
    ws.cell(11, 3).value = 4704.68
    ws.cell(11, 4).value = 4704.68

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    wb.save(tmp.name)

    rows = parse_bc(tmp.name)
    os.unlink(tmp.name)

    codes = [(r['account_code'], r['account_name'], r.get('is_sub_item')) for r in rows]

    # Parent row
    if any(c == '639110' and n == 'Insurance-Property' and not s for c, n, s in codes):
        ok('Parent row 639110 parsed')
    else:
        fail('Parent row 639110 missing', str(codes))

    # Sub-line: Property Insurance inherits 639110
    if any(c == '639110' and n == 'Property Insurance' and s for c, n, s in codes):
        ok('Sub-line "Property Insurance" inherits 639110, is_sub_item=True')
    else:
        fail('Sub-line "Property Insurance" not found or wrong code', str(codes))

    # Sub-line: Umbrella Policy inherits 639110
    if any(c == '639110' and n == 'Umbrella Policy' and s for c, n, s in codes):
        ok('Sub-line "Umbrella Policy" inherits 639110, is_sub_item=True')
    else:
        fail('Sub-line "Umbrella Policy" not found or wrong code', str(codes))

    # Total row excluded
    if not any(n == 'Total Insurance' for _, n, _ in codes):
        ok('"Total Insurance" row correctly excluded')
    else:
        fail('"Total Insurance" row should have been excluded')

    # Unrelated account correct
    if any(c == '637130' and not s for c, n, s in codes):
        ok('637130 Management Fees parsed with own code, is_sub_item=False')
    else:
        fail('637130 missing or wrong', str(codes))

    # 637130 did not inherit a prior code from 639110
    for c, n, s in codes:
        if n == 'Admin-Management Fees' and c != '637130':
            fail(f'637130 row wrongly inherited code {c}')
            break
    else:
        ok('637130 did not bleed into prior code')

except Exception as e:
    fail('BC parser test crashed', traceback.format_exc())


# ── 2. INSURANCE AMORTIZATION: three modes ───────────────────────────────────
print('\n=== 2. INSURANCE AMORTIZATION: Mode A / B / C ===')
try:
    from accrual_entry_generator import detect_insurance_amortization

    # Minimal GL stub with 135110 positive balance and no 639110 activity yet.
    # transactions defaults to a single real (non-pipeline, control='K') entry
    # matching net_change, so _real_net_change() (used by the production
    # already-posted guard) sees the same value as net_change unless a test
    # explicitly overrides it with its own transactions list.
    class FakeTxn:
        def __init__(self, net):
            self.debit = net if net > 0 else 0
            self.credit = -net if net < 0 else 0
            self.control = 'K'
            self.description = ''
            self.remarks = ''
            self.reference = ''

    class FakeAcct:
        def __init__(self, code, ending, net_change=0, transactions=None):
            self.account_code = code
            self.ending_balance = ending
            self.net_change = net_change
            self.transactions = (
                transactions if transactions is not None
                else ([FakeTxn(net_change)] if net_change else [])
            )

    class FakeGL:
        def __init__(self):
            self.accounts = [
                FakeAcct('135110', 55000.0, 0),   # Prepaid Insurance — positive
                FakeAcct('639110', 0.0, 0.0),      # No activity yet
            ]

    gl = FakeGL()

    # Synthetic BC budget_data row (Mode C fallback)
    bc_data = [
        {'account_code': '639110', 'account_name': 'Insurance-Property',
         'ptd_budget': 4704.68, 'ptd_actual': 0},
    ]

    # Synthetic Kardin rows (Mode B). Monthly columns are uppercase M1-M12 in
    # the real Kardin parser (parsers/kardin_budget.py) — must match here or
    # detect_insurance_amortization()'s row.get('M5', 0) lookup silently misses.
    kardin_data = [
        {'account_code': '639110', 'description': 'Property Insurance',
         'M5': 4037.78, 'm_total': 48453.36},
        {'account_code': '639110', 'description': 'Umbrella Policy',
         'M5': 666.90,  'm_total': 8002.80},
    ]

    # Config policies (Mode A)
    config_policies = [
        {'name': 'Property Insurance', 'expense_account': '639110', 'monthly_amount': 4037.78},
        {'name': 'Umbrella Policy',    'expense_account': '639110', 'monthly_amount': 666.90},
    ]

    # ── Mode A: config-driven ────────────────────────────────────
    res_a = detect_insurance_amortization(gl, bc_data, period='May-2026',
                                          insurance_policies=config_policies)
    if len(res_a) == 2:
        ok(f'Mode A: 2 JE lines generated ({[r["amount"] for r in res_a]})')
    else:
        fail(f'Mode A: expected 2 lines, got {len(res_a)}', str(res_a))

    amounts_a = {r['description'].split('—')[1].split('(')[0].strip(): r['amount'] for r in res_a}
    if abs(sum(r['amount'] for r in res_a) - 4704.68) < 0.02:
        ok(f'Mode A: total amount = ${sum(r["amount"] for r in res_a):,.2f}')
    else:
        fail(f'Mode A: wrong total', str([r['amount'] for r in res_a]))

    # ── Mode B: Kardin-driven ────────────────────────────────────
    res_b = detect_insurance_amortization(gl, bc_data, period='May-2026',
                                          kardin_records=kardin_data)
    if len(res_b) == 2:
        ok(f'Mode B: 2 JE lines generated from Kardin ({[r["amount"] for r in res_b]})')
    else:
        fail(f'Mode B: expected 2 lines, got {len(res_b)}', str(res_b))

    if any('Property Insurance' in r['description'] for r in res_b):
        ok('Mode B: "Property Insurance" description preserved from Kardin')
    else:
        fail('Mode B: Kardin description not found in output')

    if any('Umbrella Policy' in r['description'] for r in res_b):
        ok('Mode B: "Umbrella Policy" description preserved from Kardin')
    else:
        fail('Mode B: Umbrella description not found')

    # ── Mode C: BC fallback ──────────────────────────────────────
    res_c = detect_insurance_amortization(gl, bc_data, period='May-2026')
    if len(res_c) == 1 and abs(res_c[0]['amount'] - 4704.68) < 0.02:
        ok(f'Mode C: 1 combined BC line, amount=${res_c[0]["amount"]:,.2f}')
    else:
        fail(f'Mode C: expected 1 line at $4704.68, got {len(res_c)} lines', str(res_c))

    # Mode A takes precedence over B when both provided
    res_ab = detect_insurance_amortization(gl, bc_data, period='May-2026',
                                           insurance_policies=config_policies,
                                           kardin_records=kardin_data)
    if len(res_ab) == 2 and all(
        abs(r['amount'] - exp) < 0.02
        for r, exp in zip(sorted(res_ab, key=lambda x: x['amount'], reverse=True),
                          [4037.78, 666.90])
    ):
        ok('Mode A takes precedence over B when both provided')
    else:
        fail('Mode A/B precedence wrong', str(res_ab))

    # Already-posted guard
    class FakeGLPosted:
        def __init__(self):
            self.accounts = [
                FakeAcct('135110', 55000.0, 0),
                FakeAcct('639110', 4704.68, 4704.68),  # already posted this period
            ]
    gl_posted = FakeGLPosted()
    res_posted = detect_insurance_amortization(gl_posted, bc_data, period='May-2026',
                                               insurance_policies=config_policies)
    if len(res_posted) == 0:
        ok('Already-posted guard: no entries when 639110 has net debit activity')
    else:
        fail('Already-posted guard failed — generated entries when 639110 already posted')

except Exception as e:
    fail('Insurance amortization test crashed', traceback.format_exc())


# ── 3. REVERSAL FLAGS: prepaid never reverses, accruals always do ─────────────
print('\n=== 3. REVERSAL FLAGS: _post_amort stamps reverse_next_month=0 ===')
try:
    from accrual_entry_generator import build_accrual_entries

    class FakeAcct2:
        def __init__(self, code, ending, net_change=0, beg=0):
            self.account_code      = code
            self.account_name      = f'Account {code}'
            self.ending_balance    = ending
            self.net_change        = net_change
            self.beginning_balance = beg
            self.transactions      = []

    class FakeGL2:
        def __init__(self):
            self.accounts = [
                FakeAcct2('135110', 55000.0),   # Prepaid Insurance
                FakeAcct2('639110', 0.0, 0.0),  # No insurance activity
                FakeAcct2('111100', 100000.0),
            ]
            self.metadata = type('M', (), {'property_code': 'revlabspm', 'period': 'May-2026', 'property_name': 'RevLabs'})()
            self.transactions = []

    gl2 = FakeGL2()
    ins_policies = [
        {'name': 'Property Insurance', 'expense_account': '639110', 'monthly_amount': 4037.78},
        {'name': 'Umbrella Policy',    'expense_account': '639110', 'monthly_amount': 666.90},
    ]

    je = build_accrual_entries(
        [],
        period='May-2026',
        property_name='RevLabs',
        gl_data=gl2,
        budget_data=[],
        insurance_policies=ins_policies,
    )

    ins_lines = [l for l in je if l.get('source') == 'prepaid_amortization']
    if ins_lines:
        all_no_reverse = all(l.get('reverse_next_month') == 0 for l in ins_lines)
        if all_no_reverse:
            ok(f'All {len(ins_lines)} prepaid_amortization lines have reverse_next_month=0')
        else:
            bad = [l for l in ins_lines if l.get('reverse_next_month') != 0]
            fail(f'{len(bad)} prepaid lines missing reverse_next_month=0', str(bad[:2]))
    else:
        fail('No prepaid_amortization lines found in JE output — insurance not generated')

except Exception as e:
    fail('Reversal flag test crashed', traceback.format_exc())


# ── 4. generate_etl_csv: per-line reverse_next_month override ────────────────
print('\n=== 4. generate_etl_csv: BM column respects per-line override ===')
try:
    import csv
    from accrual_entry_generator import generate_etl_csv

    lines = [
        # Accrual — should reverse (BM = -1, default)
        {'je_number': 'ACC-0001', 'line': 1, 'account_code': '637150',
         'account_name': 'Tenant Relations', 'description': 'Accrual May-2026',
         'reference': 'ACC-0001', 'debit': 1000.0, 'credit': 0,
         'source': 'historical_recurring'},
        {'je_number': 'ACC-0001', 'line': 2, 'account_code': '213100',
         'account_name': 'Accrued Expenses', 'description': 'Accrual May-2026',
         'reference': 'ACC-0001', 'debit': 0, 'credit': 1000.0,
         'source': 'historical_recurring'},
        # Prepaid amortization — must NOT reverse (BM = 0)
        {'je_number': 'INS-0001', 'line': 1, 'account_code': '639110',
         'account_name': 'Insurance', 'description': 'Insurance amort May-2026',
         'reference': 'INS-AMORT', 'debit': 4037.78, 'credit': 0,
         'source': 'prepaid_amortization', 'reverse_next_month': 0},
        {'je_number': 'INS-0001', 'line': 2, 'account_code': '135110',
         'account_name': 'Prepaid Insurance', 'description': 'Insurance amort May-2026',
         'reference': 'INS-AMORT', 'debit': 0, 'credit': 4037.78,
         'source': 'prepaid_amortization', 'reverse_next_month': 0},
    ]

    tmp_csv = tempfile.NamedTemporaryFile(suffix='.csv', delete=False, mode='w')
    tmp_csv.close()
    generate_etl_csv(lines, tmp_csv.name, period='May-2026',
                     property_code='revlabspm', auto_reverse=True)

    # Find ReverseNextMonth column index
    with open(tmp_csv.name, newline='', encoding='utf-8') as f:
        reader = list(csv.reader(f))
    os.unlink(tmp_csv.name)

    headers_row = reader[1]
    bm_idx = next((i for i, h in enumerate(headers_row) if h == 'ReverseNextMonth'), None)

    if bm_idx is None:
        fail('ReverseNextMonth column not found in CSV')
    else:
        data_rows = reader[2:]
        accrual_bms  = [r[bm_idx] for r in data_rows if r[3] in ('637150', '213100')]
        prepaid_bms  = [r[bm_idx] for r in data_rows if r[3] in ('639110', '135110')]

        if all(str(b) == '-1' for b in accrual_bms):
            ok(f'Accrual lines: BM = -1 (auto-reverse) ok  {accrual_bms}')
        else:
            fail(f'Accrual lines: expected BM=-1, got {accrual_bms}')

        if all(str(b) == '0' for b in prepaid_bms):
            ok(f'Prepaid lines: BM = 0 (no reversal) ok  {prepaid_bms}')
        else:
            fail(f'Prepaid lines: expected BM=0, got {prepaid_bms}')

except Exception as e:
    fail('generate_etl_csv reversal test crashed', traceback.format_exc())


# ── 5. PropertyConfig: insurance_policies from config.yaml ───────────────────
print('\n=== 5. PropertyConfig: insurance_policies loaded from config.yaml ===')
try:
    from property_config import load_property_config

    cfg = load_property_config('revlabspm', data_dir=r'C:\Users\RyanCWalsh\.claude\ga-automation\data')

    if hasattr(cfg, 'insurance_policies') and cfg.insurance_policies:
        ok(f'insurance_policies loaded: {len(cfg.insurance_policies)} entries')
        for pol in cfg.insurance_policies:
            name   = pol.get('name', '?')
            amt    = pol.get('monthly_amount', 0)
            acct   = pol.get('expense_account', '?')
            if amt > 0:
                ok(f'  {name}: ${amt:,.2f}/mo  acct={acct}')
            else:
                fail(f'  {name}: monthly_amount=0 (not loaded correctly)', str(pol))

        # Validate specific amounts
        amounts = {p['name']: p['monthly_amount'] for p in cfg.insurance_policies}
        if abs(amounts.get('Property Insurance', 0) - 4037.78) < 0.01:
            ok('Property Insurance amount correct: $4,037.78')
        else:
            fail(f'Property Insurance wrong: {amounts.get("Property Insurance")}')

        if abs(amounts.get('Umbrella Policy', 0) - 666.90) < 0.01:
            ok('Umbrella Policy amount correct: $666.90')
        else:
            fail(f'Umbrella Policy wrong: {amounts.get("Umbrella Policy")}')

    else:
        fail('insurance_policies not found or empty on PropertyConfig')

except Exception as e:
    fail('PropertyConfig test crashed', traceback.format_exc())


# ── Summary ──────────────────────────────────────────────────────────────────
print(f'\n{"="*55}')
print(f'  RESULTS:  {PASS} passed  |  {FAIL} failed')
print(f'{"="*55}')
if FAIL == 0:
    print('  All checks green -- PASS')
else:
    print(f'  {FAIL} check(s) need attention -- FAIL')

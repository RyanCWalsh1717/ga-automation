"""
File Classifier — Bulk Upload Auto-Detection
=============================================
Inspects uploaded file content to identify which pipeline input slot each file
belongs to, without relying on filenames.

Returns one of the FILE_CONFIG keys defined in app.py:
  gl, trial_balance, budget_comparison, kardin_budget, t12_statement,
  nexus_accrual, bank_rec, receivable_detail, ar_aging, bank_rec_dev,
  daca_bank, loan, prepaid_ledger
  + pass-2 variants: gl_pass2, budget_comparison_pass2, trial_balance_pass2,
                     t12_statement_pass2, loan_pass2, prior_workpaper

Usage:
  from file_classifier import classify_file
  key, confidence, label = classify_file(filename, file_bytes)
"""

from __future__ import annotations

import io
import re
from typing import Tuple

# ── Readable labels for each key (shown in the UI) ──────────────────────────
FILE_LABELS = {
    "gl":                    "Yardi GL Detail",
    "trial_balance":         "Yardi Trial Balance",
    "budget_comparison":     "Yardi Budget Comparison",
    "kardin_budget":         "Kardin Budget",
    "t12_statement":         "12-Month Income Statement",
    "nexus_accrual":         "Nexus Invoice Detail",
    "bank_rec":              "Yardi / PNC Bank Rec",
    "receivable_summary":    "Yardi Receivable Summary",
    "receivable_detail":     "Yardi Receivable Detail",
    "ar_aging":              "Yardi AR Detail Aging",
    "bank_rec_dev":          "BofA Development Statement (PDF)",
    "bank_rec_dev_xlsx":     "Yardi Development Bank Rec (111210)",
    "bank_rec_xlsx":         "Yardi Operating Bank Rec (111100)",
    "daca_bank":             "KeyBank DACA Statement",
    "daca_bank_rec_xlsx":    "Yardi DACA Bank Rec (115100)",
    "loan":                  "Berkadia Loan Statement(s) — due 7th of following month",
    "prepaid_ledger":        "Prior Month Prepaid Ledger",
    "prior_workpaper":       "Prior Month Workpaper",
    # Pass-2 overrides — same classifier routes here via pass2=True
    "gl_pass2":              "Final GL (Pass 2)",
    "budget_comparison_pass2": "Final Budget Comparison (Pass 2)",
    "trial_balance_pass2":   "Final Trial Balance (Pass 2)",
    "t12_statement_pass2":   "Post-Close T12 (Pass 2)",
    "loan_pass2":            "Berkadia Loan Statements (Pass 2) — due 7th of following month",
    "unknown":               "Unknown — select type",
    # Pass-2 labels (shown when bulk-uploading into the Pass 2 slot)
    "bank_rec_pass2":        "Bank Rec (Pass 2)",
}

# Pass-2 key remapping: classify_file returns the base key; caller passes pass2=True
# and we remap using this dict where a pass-2 override exists.
_PASS2_REMAP = {
    "gl":               "gl_pass2",
    "budget_comparison":"budget_comparison_pass2",
    "trial_balance":    "trial_balance_pass2",
    "t12_statement":    "t12_statement_pass2",
    "loan":             "loan_pass2",
    "bank_rec":         "bank_rec_pass2",   # B-F9: was missing — showed raw key as label
}

# Keys that accept multiple files (list of paths, not a single path)
MULTI_FILE_KEYS = {"loan", "loan_pass2"}


def classify_file(filename: str, file_bytes: bytes,
                  pass2: bool = False,
                  property_config=None) -> Tuple[str, float, str]:
    """
    Classify a single file by content inspection.

    Args:
        filename:        Original filename (used for extension detection only)
        file_bytes:      Raw bytes of the file
        pass2:           If True, remap base keys to their pass-2 variants
        property_config: PropertyConfig for the active property. When provided,
                         bank account numbers and property code are read from
                         config instead of hardcoded RevLabs values — enabling
                         correct auto-classification for any property.

    Returns:
        (key, confidence, label)
        key        — FILE_CONFIG slot name, or 'unknown'
        confidence — 0.0–1.0
        label      — Human-readable type description
    """
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    # Build property-specific signals from config (falls back to RevLabs hardcodes)
    signals = _build_signals(property_config)

    if ext in ("xlsx", "xls"):
        key, conf = _classify_excel(filename, file_bytes, ext, signals)
    elif ext == "pdf":
        key, conf = _classify_pdf(filename, file_bytes, signals)
    else:
        key, conf = "unknown", 0.0

    if pass2 and key in _PASS2_REMAP:
        key = _PASS2_REMAP[key]

    return key, conf, FILE_LABELS.get(key, "Unknown — select type")


def _build_signals(property_config) -> dict:
    """
    Extract property-specific matching signals from PropertyConfig.

    Returns a dict with sets of strings to search for in file content,
    keyed by the file type they indicate.  Falls back to hardcoded
    RevLabs values when property_config is None.

    Signal keys:
        'operating_accounts'   → str set — full account numbers for operating bank rec (xlsx)
        'dev_accounts'         → str set — account numbers for development bank rec (xlsx)
        'daca_accounts'        → str set — account numbers for DACA (xlsx + PDF)
        'property_codes'       → str set — Yardi property codes for GL detection
        'operating_banks'      → str set — bank names for operating PDF (e.g. 'pnc')
        'dev_banks'            → str set — bank names for dev PDF (e.g. 'bank of america')
        'daca_banks'           → str set — bank names for DACA PDF (e.g. 'keybank')
    """
    if not property_config:
        # No config — return empty signals; GL detection falls back to "general ledger" text match
        return {
            'operating_accounts': set(),
            'dev_accounts':       set(),
            'daca_accounts':      set(),
            'property_codes':     set(),
            'operating_banks':    set(),
            'dev_banks':          set(),
            'daca_banks':         set(),
        }

    operating_accounts: set = set()
    dev_accounts:       set = set()
    daca_accounts:      set = set()
    operating_banks:    set = set()
    dev_banks:          set = set()
    daca_banks:         set = set()

    for slug, ba in (property_config.bank_accounts or {}).items():
        slug_l    = slug.lower()
        label_l   = (ba.label or '').lower()
        full      = ba.full_account or ''
        last4     = (ba.last4 or '').lstrip('x')
        bank_name = (ba.bank_name or '').lower()  # explicit bank name from config

        # Determine account type from slug/label keywords
        is_daca = 'daca' in slug_l or 'daca' in label_l
        is_dev  = (not is_daca) and ('dev' in slug_l or 'development' in label_l)
        is_op   = (not is_daca and not is_dev) and ('operat' in slug_l or 'operat' in label_l)

        if is_daca:
            if full:      daca_accounts.add(full)
            if last4:     daca_accounts.add(last4); daca_accounts.add(f'x{last4}')
            if bank_name: daca_banks.add(bank_name)
            daca_banks.add('daca')  # always match the word "daca" for DACA accounts
        elif is_dev:
            if full:      dev_accounts.add(full)
            if last4:     dev_accounts.add(last4)
            if bank_name: dev_banks.add(bank_name)
        elif is_op:
            if full:      operating_accounts.add(full)
            if bank_name: operating_banks.add(bank_name)
        else:
            # Unrecognised slug — treat as operating fallback
            if full:      operating_accounts.add(full)
            if bank_name: operating_banks.add(bank_name)

    # daca_banks must always contain 'daca' so DACA PDFs match regardless of config
    if 'daca' not in daca_banks:
        daca_banks.add('daca')

    prop_code = (property_config.property_code or '').lower()

    return {
        'operating_accounts': operating_accounts,
        'dev_accounts':       dev_accounts,
        'daca_accounts':      daca_accounts,
        'property_codes':     {prop_code} if prop_code else set(),
        'operating_banks':    operating_banks,
        'dev_banks':          dev_banks,
        'daca_banks':         daca_banks,
    }


# ── Excel classifier ─────────────────────────────────────────────────────────

def _classify_excel(filename: str, file_bytes: bytes, ext: str,
                    signals: dict = None) -> Tuple[str, float]:
    try:
        if ext == "xls":
            return _classify_xls(file_bytes)
        return _classify_xlsx(file_bytes, signals or {})
    except Exception:
        return "unknown", 0.0


def _classify_xls(file_bytes: bytes) -> Tuple[str, float]:
    """Classify legacy .xls files — only Nexus uses .xls in this pipeline."""
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        # Read first 5 rows as flat text
        text = " ".join(
            str(ws.cell_value(r, c)).lower()
            for r in range(min(5, ws.nrows))
            for c in range(min(15, ws.ncols))
        )
        if "vendor" in text and "invoice" in text:
            return "nexus_accrual", 0.90
        if "nexus" in text:
            return "nexus_accrual", 0.85
    except Exception:
        pass
    return "nexus_accrual", 0.60  # Only .xls file type used is Nexus


def _classify_xlsx(file_bytes: bytes, signals: dict = None) -> Tuple[str, float]:
    """Classify .xlsx files by reading first rows of the active sheet."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

    # ── Check sheet names first (prepaid ledger, workpaper) ──────────────
    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    # Prepaid ledger seed/output file uses "Active" and "Completed" sheet
    # names — this is the real signal. A broader "any sheet name contains
    # 'prepaid'" check used to run here first, but the real ledger file's own
    # sheets ('135150 PPD Other', 'Active', 'Completed') never actually
    # contain that substring, so it never correctly matched its intended
    # target — it only ever caused a false positive: a real 30-tab workpaper
    # has an account-name sheet ("221100 Prepaid Rent - Tenant") that matched
    # it, misclassifying the entire workpaper as the prepaid ledger. Confirmed
    # on a real file: prepaid_ledger.load() then silently returned 0 active
    # items with no error (no 'Active'/'Completed' sheets to find), so
    # nothing carried forward and no warning fired either.
    if "active" in sheet_names_lower and "completed" in sheet_names_lower:
        wb.close()
        return "prepaid_ledger", 0.90
    # The real template-based GA_Workpapers.xlsx output (generate_bs_workpaper_
    # from_template) never actually hits the two checks below — its sheets are
    # named "111100 PNC Cash", "213100 Accr Exp", etc., not "GL vs TB" or
    # "Jan-2026 GL vs TB". Confirmed on a real generated file: neither check
    # matched, so it fell through to "unknown" (safe — forces manual
    # selection — but not auto-detected). 'summary page' + 'trial balance'
    # together are that template's actual distinctive signature; a standalone
    # Yardi Trial Balance export is a single sheet and won't have a
    # 'summary page' sheet alongside it, so this doesn't collide with that
    # upload slot's own detection.
    if "summary page" in sheet_names_lower and "trial balance" in sheet_names_lower:
        wb.close()
        return "prior_workpaper", 0.90
    if any(s in ("gl vs tb", "bank rec", "debt service", "accruals") for s in sheet_names_lower):
        wb.close()
        return "prior_workpaper", 0.90
    # Workpaper sheets follow pattern "Jan-2026 GL vs TB" etc.
    if any(re.search(r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)-\d{4}\b', s)
           for s in sheet_names_lower):
        wb.close()
        return "prior_workpaper", 0.88

    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 12:
            break
        rows.append([str(v or "").strip() for v in (row or [])[:20]])
    wb.close()

    all_text = " ".join(cell for row in rows for cell in row).lower()

    # ── Kardin: distinctive sheet + header columns ────────────────────────
    # Kardin has 'qryExportData' sheet or PropID/M1–M12 headers
    if "propid" in all_text and ("m1" in all_text or "m12" in all_text):
        return "kardin_budget", 0.95
    if "kardin" in all_text:
        return "kardin_budget", 0.90

    # ── T12: "Statement (12 months)" + ≥10 month-label columns ──────────
    month_pattern = re.compile(
        r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+\d{4}\b'
    )
    month_hits = sum(1 for row in rows for cell in row if month_pattern.search(cell.lower()))
    if month_hits >= 10:
        return "t12_statement", 0.95
    if "statement (12 months)" in all_text or ("period" in all_text and month_hits >= 6):
        return "t12_statement", 0.88

    # ── Trial Balance ─────────────────────────────────────────────────────
    if "trial balance" in all_text:
        return "trial_balance", 0.95

    # ── Budget Comparison ─────────────────────────────────────────────────
    if "budget comparison" in all_text:
        return "budget_comparison", 0.95
    if "ptd budget" in all_text and "ptd actual" in all_text:
        return "budget_comparison", 0.88

    # ── Receivable Summary / Detail / AR Aging ────────────────────────────
    # Must come before the revlabspm→GL fallback: Yardi receivable reports
    # carry the property header ("revlabspm") just like the GL export.
    #
    # Receivable Summary: titled "Receivable Summary"; columns are
    #                     Property | Charge Code | Balance Forward | Charge | Receipt | Ending Balance
    #                     No "Control #" column, no aging buckets.
    # Receivable Detail:  titled "Receivable Detail"; has "Control #", "Charge Code",
    #                     "Charges", "Receipts" columns — no aging buckets.
    # AR Aging:           titled "AR Detail Aging"; has aging-bucket columns
    #                     labelled "30 days", "60 days", "90 days".
    #
    # IMPORTANT: do NOT use bare "30" / "60" / "90" substring checks — any dollar
    # amount or date containing those digits triggers a false positive.  Use the
    # full phrase "30 days" / "60 days" / "90 days" which only appears in AR Aging.

    # Receivable Summary — must check FIRST; its title contains "receivable" which
    # would otherwise match the weaker receivable_detail fallback below.
    if "receivable summary" in all_text:
        return "receivable_summary", 0.95

    if ("ar detail aging" in all_text or "ar aging" in all_text
            or "aging detail" in all_text):
        return "ar_aging", 0.92
    # Secondary AR Aging signal: aging-bucket column headers
    if ("aging" in all_text and (
        "charge code" in all_text or "tenant" in all_text
    ) and ("30 days" in all_text or "60 days" in all_text or "90 days" in all_text)):
        return "ar_aging", 0.85
    if "receivable detail" in all_text and (
        "charge code" in all_text or "control" in all_text
    ):
        # Receivable Detail confirmed by title + Control # column.
        # If aging buckets also present, lean toward ar_aging — but title wins.
        if "ar aging" in all_text or "aging detail" in all_text:
            return "ar_aging", 0.85
        return "receivable_detail", 0.90
    # Weaker signal: "receivable" without the full title (fallback)
    if "receivable" in all_text and (
        "charge code" in all_text or "tenant" in all_text
    ):
        if "aging" in all_text or "30 days" in all_text:
            return "ar_aging", 0.80
        return "receivable_detail", 0.80

    # ── Yardi Bank Rec — account-specific matching (config-driven) ───────────
    # Check DACA first (KeyBank), then development (BofA), then operating (PNC).
    # Uses account numbers and keywords from property config (or RevLabs defaults).
    _s = signals or {}
    _dev_accts  = _s.get('dev_accounts', {'466007913132', '3132'})
    _op_accts   = _s.get('operating_accounts', {'1092223993'})
    _daca_accts = _s.get('daca_accounts', {'329681415132', 'x5132', '5132'})

    if "bank reconciliation report" in all_text:
        # This branch only fires for an actual Yardi "Bank Reconciliation
        # Report" export (Excel), never a raw bank statement PDF — so every
        # outcome here must route to the corresponding "_xlsx" workpaper-fill
        # slot (bank_rec_xlsx / daca_bank_rec_xlsx / bank_rec_dev_xlsx), not
        # the raw-PDF slots (bank_rec / daca_bank) those slots' FILE_CONFIG
        # entries describe. Only "Development" did this correctly before —
        # Operating and DACA were silently misfiled into the PDF slots, so
        # the 111100/115100 workpaper tabs never saw this data even when the
        # right file was uploaded (confirmed 2026-08-03).
        # DACA check first — most specific
        if "daca" in all_text or any(a in all_text for a in _daca_accts):
            return "daca_bank_rec_xlsx", 0.95
        # Development check — BofA / dev account
        if ("development" in all_text or "bank of america" in all_text
                or any(a in all_text for a in _dev_accts)):
            return "bank_rec_dev_xlsx", 0.95
        # Operating (PNC or whatever the operating account is)
        if any(a in all_text for a in _op_accts):
            return "bank_rec_xlsx", 0.95
        # Generic fallback — some bank rec but no account match
        return "bank_rec_xlsx", 0.80

    # ── GL: "General Ledger" OR property code with GL-specific column signals ──
    if "general ledger" in all_text:
        return "gl", 0.95
    # Yardi GL header contains the property code in parentheses, e.g. "(revlabspm)".
    # Require at least one GL-specific column signal to avoid misclassifying other
    # Yardi reports (budget comparison, T12, etc.) that also embed the property code
    # in their header — the property code alone is not sufficient (B-5).
    _prop_codes = _s.get('property_codes', set())
    _gl_col_signals = ('debit', 'credit', 'balance', 'control', 'remarks')
    if any(code in all_text for code in _prop_codes) and any(
        sig in all_text for sig in _gl_col_signals
    ):
        return "gl", 0.85
    # Property code present but no GL column signals — return unknown rather
    # than silently misclassifying as GL and feeding garbage to the GL parser.
    if any(code in all_text for code in _prop_codes):
        return "unknown", 0.40

    # ── Prepaid Ledger (content check before Nexus — seed file has Vendor/Invoice cols) ──
    if ("monthly amt" in all_text or "months posted" in all_text
            or "service start" in all_text or "months left" in all_text):
        return "prepaid_ledger", 0.90

    # ── Nexus in xlsx form ────────────────────────────────────────────────
    if "nexus" in all_text or (
        "vendor" in all_text and "invoice" in all_text and "gl account" in all_text
    ):
        return "nexus_accrual", 0.88

    # ── Berkadia (xlsx version) ───────────────────────────────────────────
    if "berkadia" in all_text:
        return "loan", 0.92

    # ── Prepaid Ledger (column-name fallback) ─────────────────────────────
    if "monthly_amount" in all_text or (
        "monthly amt" in all_text and "months_amortized" in all_text
    ):
        return "prepaid_ledger", 0.88

    return "unknown", 0.0


# ── PDF classifier ───────────────────────────────────────────────────────────

def _classify_pdf(filename: str, file_bytes: bytes,
                  signals: dict = None) -> Tuple[str, float]:
    _s          = signals or {}
    _dev_banks  = _s.get('dev_banks',  {'bank of america', 'bofa'})
    _daca_banks = _s.get('daca_banks', {'keybank', 'daca'})
    _op_banks   = _s.get('operating_banks', {'pnc'})
    _daca_accts = _s.get('daca_accounts', {'329681415132', 'x5132', '5132'})

    try:
        import pdfplumber
        text = ""
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages[:3]:
                text += page.extract_text() or ""
    except Exception:
        # Filename heuristics when pdfplumber fails
        fn = filename.lower()
        if any(b in fn for b in _dev_banks) or "bankofamerica" in fn or "bank_of_america" in fn:
            return "bank_rec_dev", 0.60
        if any(b in fn for b in _daca_banks):
            return "daca_bank", 0.60
        if "berkadia" in fn:
            return "loan", 0.60
        return "unknown", 0.0

    tl = text.lower()

    # ── Development bank (BofA or config equivalent) — check FIRST ───────
    # Dev bank PDFs can contain "bank reconciliation report" too, so check
    # bank name before the generic Yardi rec check below.
    if any(b in tl for b in _dev_banks):
        return "bank_rec_dev", 0.95

    # ── DACA bank rec (KeyBank or config equivalent) ──────────────────────
    # Must come before the generic "bank reconciliation report" check.
    if "bank reconciliation report" in tl and (
        any(b in tl for b in _daca_banks)
        or any(a in text for a in _daca_accts)
    ):
        return "daca_bank", 0.97

    # ── Generic Yardi Bank Rec (operating account PDF) ────────────────────
    if "bank reconciliation report" in tl:
        return "bank_rec", 0.97

    # ── DACA standalone statement (no Yardi header) ───────────────────────
    if any(b in tl for b in _daca_banks) and (
        any(a in text for a in _daca_accts) or "daca" in tl
    ):
        return "daca_bank", 0.95

    # ── Berkadia loan statements ──────────────────────────────────────────
    if "berkadia" in tl:
        return "loan", 0.95

    # ── Operating bank standalone statement (PNC or config equivalent) ────
    if any(b in tl for b in _op_banks) and (
        "account summary" in tl or "corporate business" in tl or "ending balance" in tl
    ):
        return "bank_rec", 0.82

    return "unknown", 0.0

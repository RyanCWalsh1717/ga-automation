"""
Sign-off Record Generator
=========================
Produces GA_Signoff_Record.xlsx — the permanent record of who reviewed
and approved each section of the monthly close package.

Called from app.py when the user clicks "Export Sign-off Sheet".
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Brand colours (match bs_workpaper_generator palette) ─────────────────────
_BLACK      = '000000'
_GRP_GREEN  = '2D6F50'
_LIGHT_GREEN = 'D6EAE1'
_WHITE      = 'FFFFFF'
_PASS_GREEN  = 'E2EFDA'
_PASS_FONT   = '006100'
_PEND_GRAY   = 'F2F2F2'
_PEND_FONT   = '757575'

_THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def _font(bold=False, italic=False, size=11, color=_BLACK, name='Calibri') -> Font:
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)


def generate_signoff_xlsx(
    output_path: str,
    signoff_state: Dict[int, Dict[str, str]],
    items: List[str],
    period: str,
    property_name: str,
) -> str:
    """
    Write GA_Signoff_Record.xlsx and return output_path.

    Args:
        output_path:    Destination .xlsx path.
        signoff_state:  {item_index: {'signed_by': str, 'timestamp': str}}
                        Items not present are treated as 'Pending'.
        items:          Ordered list of checklist item names (0-based index).
        period:         Close period string e.g. 'Jan-2026'.
        property_name:  Property display name e.g. 'Revolution Labs'.

    Returns:
        output_path
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sign-off Record'

    # Column widths
    ws.column_dimensions['A'].width = 2     # left margin
    ws.column_dimensions['B'].width = 5     # #
    ws.column_dimensions['C'].width = 42    # Item
    ws.column_dimensions['D'].width = 16    # Status
    ws.column_dimensions['E'].width = 22    # Signed By
    ws.column_dimensions['F'].width = 22    # Timestamp
    ws.column_dimensions['G'].width = 14    # Period
    ws.column_dimensions['H'].width = 24    # Property

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    title = ws.cell(row=1, column=2,
                    value=f'{property_name} — Monthly Close Sign-off Record')
    title.font = _font(bold=True, size=14, color=_WHITE)
    title.fill = _fill(_BLACK)
    title.alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=8)
    ws.row_dimensions[1].height = 22

    # ── Row 2: Sub-header ────────────────────────────────────────────────────
    sub = ws.cell(row=2, column=2,
                  value=f'Period: {period}  |  Generated: {datetime.now().strftime("%m/%d/%Y %H:%M")}')
    sub.font = _font(italic=True, size=10, color=_WHITE)
    sub.fill = _fill(_GRP_GREEN)
    sub.alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=8)
    ws.row_dimensions[2].height = 18

    # ── Row 3: Blank ─────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Row 4: Column headers ────────────────────────────────────────────────
    col_headers = ['#', 'Checklist Item', 'Status', 'Signed By', 'Timestamp',
                   'Period', 'Property']
    for ci, hdr in enumerate(col_headers):
        cell = ws.cell(row=4, column=ci + 2, value=hdr)
        cell.font = _font(bold=True, size=10, color=_WHITE)
        cell.fill = _fill(_GRP_GREEN)
        cell.border = _THIN
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
    ws.row_dimensions[4].height = 20

    # ── Rows 5+: One row per item ─────────────────────────────────────────────
    for idx, item_name in enumerate(items):
        row = idx + 5
        so = signoff_state.get(idx)
        status     = 'Signed Off' if so else 'Pending'
        signed_by  = so.get('signed_by', '')  if so else ''
        timestamp  = so.get('timestamp', '')  if so else ''

        is_signed  = bool(so)
        row_fill   = _fill(_PASS_GREEN) if is_signed else _fill(_PEND_GRAY)
        num_font   = _font(size=10, color=(_PASS_FONT if is_signed else _pend_font_color()))
        text_font  = _font(size=10)
        stat_font  = _font(bold=True, size=10,
                           color=(_PASS_FONT if is_signed else _PEND_FONT))

        vals = [idx + 1, item_name, status, signed_by, timestamp, period, property_name]
        for ci, val in enumerate(vals):
            cell = ws.cell(row=row, column=ci + 2, value=val)
            cell.fill   = row_fill
            cell.border = _THIN
            cell.alignment = Alignment(vertical='center',
                                       wrap_text=(ci == 1))   # wrap item name
            if ci == 0:
                cell.font = _font(bold=True, size=10,
                                  color=_PASS_FONT if is_signed else _PEND_FONT)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 2:
                cell.font = stat_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.font = text_font

        ws.row_dimensions[row].height = 18

    # ── Summary row ───────────────────────────────────────────────────────────
    summary_row = len(items) + 5
    ws.row_dimensions[summary_row].height = 6   # spacer

    signed_count = sum(1 for i in range(len(items)) if i in signoff_state)
    total_count  = len(items)
    summary_row += 1
    sc = ws.cell(row=summary_row, column=2,
                 value=f'{signed_count} of {total_count} items signed off')
    sc.font = _font(bold=True, size=10,
                    color=_PASS_FONT if signed_count == total_count else '9C0006')
    sc.fill = _fill(_PASS_GREEN if signed_count == total_count else 'FFCCCC')
    sc.border = _THIN
    ws.merge_cells(start_row=summary_row, start_column=2,
                   end_row=summary_row, end_column=8)
    sc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[summary_row].height = 20

    wb.save(output_path)
    return output_path


def _pend_font_color() -> str:
    return _PEND_FONT

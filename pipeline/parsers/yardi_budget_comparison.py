"""
Yardi Budget Comparison (Accrual) Parser

Parses Yardi Budget Comparison export files comparing actual vs budgeted amounts:
- Rows 1-4: Meta information (property, report type, period, book/tree)
- Row 5: Column headers with PTD, YTD, and Annual metrics
- Row 6+: Hierarchical account data with actuals and budget comparisons

Expected columns:
  Account Code, Account Name, PTD Actual, PTD Budget, PTD Variance, PTD % Var,
  YTD Actual, YTD Budget, YTD Variance, YTD % Var, Annual

The account hierarchy is indicated by spacing in the Account Name column.

Features:
- Handles hierarchical GL accounts with indentation
- Extracts Period-to-Date (PTD) and Year-to-Date (YTD) comparisons
- Calculates and preserves variance metrics
- Handles 'N/A' values in percentage columns
- Normalizes numeric values
"""

from openpyxl import load_workbook
from datetime import datetime
from typing import List, Dict, Tuple, Optional


def _detect_bc_columns(ws) -> Tuple[Dict[str, int], int]:
    """
    Scan header rows (1-9) for the row that contains both 'actual' and 'budget'
    keywords, then map each field to its 0-based column index dynamically.

    Returns (col_map, data_start_row) where:
      col_map       — field name → 0-based column index
      data_start_row — 1-based row number where account data begins

    Falls back to the original hardcoded defaults when no header row is found.
    """
    _DEFAULTS: Dict[str, int] = {
        'ptd_actual':      2,
        'ptd_budget':      3,
        'ptd_variance':    4,
        'ptd_percent_var': 5,
        'ytd_actual':      6,
        'ytd_budget':      7,
        'ytd_variance':    8,
        'ytd_percent_var': 9,
        'annual':          10,
    }

    max_col = min(getattr(ws, 'max_column', 15) + 2, 20)

    for hrow in range(1, min(10, ws.max_row + 1)):
        hvals = [
            str(ws.cell(row=hrow, column=c).value or '').strip().lower()
            for c in range(1, max_col + 1)
        ]

        has_actual = any('actual' in h for h in hvals)
        has_budget = any('budget' in h for h in hvals)
        if not (has_actual and has_budget):
            continue

        # Columns by keyword occurrence — first = PTD, second = YTD
        actual_cols   = [i for i, h in enumerate(hvals) if 'actual' in h]
        # Exclude "annual budget" from the regular budget list
        budget_cols   = [i for i, h in enumerate(hvals) if 'budget' in h and 'annual' not in h]
        variance_cols = [i for i, h in enumerate(hvals) if 'variance' in h]
        pct_cols      = [i for i, h in enumerate(hvals) if '%' in h or 'percent' in h]
        annual_cols   = [i for i, h in enumerate(hvals) if 'annual' in h]

        col_map = dict(_DEFAULTS)

        if len(actual_cols) >= 1:   col_map['ptd_actual']      = actual_cols[0]
        if len(actual_cols) >= 2:   col_map['ytd_actual']      = actual_cols[1]
        if len(budget_cols) >= 1:   col_map['ptd_budget']      = budget_cols[0]
        if len(budget_cols) >= 2:   col_map['ytd_budget']      = budget_cols[1]
        if len(variance_cols) >= 1: col_map['ptd_variance']    = variance_cols[0]
        if len(variance_cols) >= 2: col_map['ytd_variance']    = variance_cols[1]
        if len(pct_cols) >= 1:      col_map['ptd_percent_var'] = pct_cols[0]
        if len(pct_cols) >= 2:      col_map['ytd_percent_var'] = pct_cols[1]
        if annual_cols:             col_map['annual']          = annual_cols[0]

        return col_map, hrow + 1

    return _DEFAULTS, 6  # default: data starts at row 6


def parse(filepath: str) -> List[Dict]:
    """
    Parse a Yardi Budget Comparison export file.

    Args:
        filepath: Path to the Excel file

    Returns:
        List of dictionaries representing budget comparison line items

    Raises:
        FileNotFoundError: If file does not exist
        ValueError: If file structure is invalid
    """
    try:
        wb = load_workbook(filepath)
    except Exception as e:
        raise FileNotFoundError(f"Cannot open file: {filepath}") from e

    ws = wb.active
    data = []

    # Meta information is in rows 1-4
    metadata = _extract_metadata(ws)

    # Detect column positions dynamically from header rows
    col_map, data_start_row = _detect_bc_columns(ws)

    # Validate that at least one header was found (extract_headers used only for validation)
    headers = _extract_headers(ws, row=5)
    if not headers:
        raise ValueError("Cannot extract headers from Budget Comparison file")

    # Process data rows starting after the detected header row
    _last_code: Optional[str] = None  # carry-forward for sub-line items

    for row_num in range(data_start_row, ws.max_row + 1):
        row = ws[row_num]
        row_values = [cell.value for cell in row]

        # Skip completely empty rows
        if all(v is None for v in row_values):
            continue

        # Account code should be in first column
        account_code = row_values[0]
        account_name = row_values[1] if len(row_values) > 1 else None

        if account_code is None:
            # Potential sub-line item under the most-recently seen account code.
            # Yardi BC exports indent sub-items (e.g. individual insurance policies
            # under 639110) without repeating the account code in column A.
            # We inherit the parent code when the row has a non-empty description
            # AND at least one non-zero numeric value in the data columns — this
            # excludes blank spacers, section headers, and subtotal labels.
            if _last_code is None:
                continue

            # Require a non-empty, non-total description
            name_str = str(account_name or '').strip()
            if not name_str:
                continue
            name_lower = name_str.lower()
            if any(w in name_lower for w in ('total', 'subtotal', 'grand total', 'net ')):
                continue

            # Require at least one non-zero number in data columns (not col 0 or 1)
            has_value = any(
                isinstance(row_values[i], (int, float)) and row_values[i] != 0
                for i in range(2, min(len(row_values), 12))
            )
            if not has_value:
                continue

            # Inherit parent account code; keep the sub-item's own description
            account_code = _last_code
        else:
            # Normal account row — update carry-forward
            _last_code = str(_normalize_value(account_code) or '').strip()

        # Build record — force code/name to str (Excel may return floats)
        record = {
            'account_code': str(_normalize_value(account_code) or '').strip(),
            'account_name': str(_normalize_value(account_name) or '').strip(),
            'is_sub_item':  row_values[0] is None,  # flag for callers
        }

        # Extract values using dynamically detected column positions.
        # Percent-variance fields (ending '_var') may contain 'N/A' — use flexible normalizer.
        for field_name, col_idx in col_map.items():
            if col_idx < len(row_values):
                value = row_values[col_idx]
                if field_name.endswith('_var'):
                    record[field_name] = _normalize_flexible_numeric(value)
                else:
                    record[field_name] = _normalize_numeric(value)

        # Add metadata
        record.update(metadata)

        data.append(record)

    return data


def validate(filepath: str) -> Tuple[bool, List[str]]:
    """
    Validate that a file has the expected Budget Comparison structure.

    Args:
        filepath: Path to the Excel file

    Returns:
        Tuple of (is_valid: bool, issues: list of error strings)
    """
    issues = []

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return False, [f"Cannot open file: {e}"]

    ws = wb.active

    # Check for expected meta rows
    if not ws.cell(1, 1).value or "Property" not in str(ws.cell(1, 1).value):
        issues.append("Row 1 missing 'Property' meta information")

    if not ws.cell(2, 1).value or "Budget Comparison" not in str(ws.cell(2, 1).value):
        issues.append("Row 2 missing 'Budget Comparison' title")

    # Check headers
    headers = _extract_headers(ws, row=5)
    if not headers:
        issues.append("Cannot extract headers from row 5")
    else:
        # Should have at least Account Code and Account Name
        if len(headers) < 2:
            issues.append("Expected at least 2 header columns")

    return len(issues) == 0, issues


def _extract_metadata(ws) -> Dict:
    """Extract metadata from rows 1-4."""
    metadata = {}

    # Row 1: Property
    prop_line = ws.cell(1, 1).value
    if prop_line:
        parts = str(prop_line).split('=')
        if len(parts) > 1:
            metadata['property'] = parts[1].strip()

    # Row 2: Report type
    report_line = ws.cell(2, 1).value
    if report_line:
        metadata['report_type'] = str(report_line).strip()

    # Row 3: Period
    period_line = ws.cell(3, 1).value
    if period_line:
        parts = str(period_line).split('=')
        if len(parts) > 1:
            metadata['period'] = parts[1].strip()

    # Row 4: Book/Tree
    book_line = ws.cell(4, 1).value
    if book_line:
        parts = str(book_line).split(';')
        for part in parts:
            part = part.strip()
            if '=' in part:
                key, val = part.split('=', 1)
                metadata[key.strip().lower()] = val.strip()

    return metadata


def _extract_headers(ws, row: int) -> List[str]:
    """Extract and clean headers from a specific row."""
    headers = []
    for cell in ws[row]:
        value = cell.value
        if value:
            headers.append(str(value).strip())
        else:
            headers.append(None)
    return headers


def _normalize_value(value):
    """Normalize values for consistent output."""
    if value is None:
        return None

    # Convert datetime to ISO format string
    if isinstance(value, datetime):
        return value.isoformat()

    # Handle strings - strip whitespace
    if isinstance(value, str):
        return value.strip()

    return value


def _normalize_numeric(value):
    """Normalize numeric values, handling None and strings."""
    if value is None:
        return None

    # Try to convert to float
    if isinstance(value, (int, float)):
        return value

    if isinstance(value, str):
        try:
            # Try int first
            if '.' not in value:
                return int(value)
            return float(value)
        except (ValueError, AttributeError):
            return None

    return value


def _normalize_flexible_numeric(value):
    """
    Normalize numeric values, but allow 'N/A' and other special strings.
    Used for percentage variance columns that may contain 'N/A'.
    """
    if value is None:
        return None

    # Try to convert to float
    if isinstance(value, (int, float)):
        return value

    if isinstance(value, str):
        value_upper = value.upper().strip()
        # Preserve 'N/A' and similar special values
        if value_upper in ('N/A', 'NA', '#DIV/0!', 'ERROR'):
            return value_upper

        try:
            # Try int first
            if '.' not in value:
                return int(value)
            return float(value)
        except (ValueError, AttributeError):
            return None

    return value


if __name__ == "__main__":
    import sys
    import json

    if len(sys.argv) < 2:
        print("Usage: python yardi_budget_comparison.py <filepath>")
        sys.exit(1)

    filepath = sys.argv[1]

    # Validate
    is_valid, issues = validate(filepath)
    if not is_valid:
        print(f"Validation errors:")
        for issue in issues:
            print(f"  - {issue}")
        sys.exit(1)

    # Parse
    data = parse(filepath)
    print(f"Successfully parsed {len(data)} budget comparison line items")
    print(f"\nSample records (first 2 entries):")
    for i, record in enumerate(data[:2]):
        print(f"\nRecord {i+1}:")
        print(json.dumps(record, indent=2, default=str))

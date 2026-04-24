"""
Report Generator for GA Automation Pipeline
==============================================
Takes parsed data from all parsers (via EngineResult) and produces:
1. The 8-tab Singerman deliverable workbook (main report)
2. An exception/validation report with detailed findings

The main workbook contains:
  - BS: Balance Sheet
  - IS: Income Statement
  - T12: Trailing 12 Months
  - TB-MTD: Trial Balance Month-to-Date
  - TB-YTD: Trial Balance Year-to-Date
  - GL-MTD: General Ledger Month-to-Date
  - GL-YTD: General Ledger Year-to-Date
  - Tenancy Schedule: Lease and unit information

The exception report contains:
  - Summary: Overview of pipeline run
  - Exceptions: All flagged issues with severity/category
  - Budget Variances: Material variances from budget
  - Bank Recon: GL to bank matching results
  - Debt Service: Loan statement reconciliation
"""

import os
from datetime import datetime, date
from typing import Optional, List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import data structures for building tabs dynamically
from parsers.monthly_report_template import ReportLine, TabStructure
from property_config import is_balance_sheet_account, is_income_statement_account


# ── Styling utilities ────────────────────────────────────────

def _header_style():
    """Create dark blue header style with white text."""
    return {
        'font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
    }


def _data_style(alternate: bool = False):
    """Create data row style with optional alternating color."""
    fill_color = 'D9E1F2' if alternate else 'FFFFFF'
    return {
        'font': Font(name='Calibri', size=11),
        'fill': PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
    }


def _currency_style(alternate: bool = False):
    """Create currency-formatted data style."""
    style = _data_style(alternate)
    style['number_format'] = '$#,##0.00'
    return style


def _percent_style(alternate: bool = False):
    """Create percentage-formatted data style."""
    style = _data_style(alternate)
    style['number_format'] = '0.0%'
    return style


def _apply_style(cell, style):
    """Apply a style dict to a cell."""
    if 'font' in style:
        cell.font = style['font']
    if 'fill' in style:
        cell.fill = style['fill']
    if 'alignment' in style:
        cell.alignment = style['alignment']
    if 'border' in style:
        cell.border = style['border']
    if 'number_format' in style:
        cell.number_format = style['number_format']


def _auto_width_columns(ws, columns: int):
    """Auto-fit column widths."""
    for col_num in range(1, columns + 1):
        col_letter = get_column_letter(col_num)
        max_len = 12
        for cell in ws[col_letter]:
            try:
                if cell.value:
                    cell_len = len(str(cell.value))
                    if cell_len > max_len:
                        max_len = cell_len
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


# ── Helper functions to build tabs from GL and IS data ────────

def _build_bs_tab_from_gl(gl_data) -> Optional[TabStructure]:
    """
    Build Balance Sheet tab from GL data.
    Filter to BS account codes (1xxxxx through 3xxxxx — assets, liabilities, equity).
    Columns: Account Code, Account Name, Balance Current Period, Beginning Balance, Net Change.
    """
    if not gl_data or not hasattr(gl_data, 'accounts'):
        return None

    line_items = []
    for account in gl_data.accounts:
        # Filter to balance sheet accounts — uses per-property COA config (defaults 1-3xxxxx)
        if not is_balance_sheet_account(account.account_code):
            continue

        report_line = ReportLine(
            account_code=account.account_code,
            account_name=account.account_name,
            values={
                'Balance Current Period': account.ending_balance,
                'Beginning Balance': account.beginning_balance,
                'Net Change': account.net_change,
            },
            row_number=0,
            is_total=False,
        )
        line_items.append(report_line)

    return TabStructure(
        name='BS',
        columns=['account_code', 'account_name', 'Balance Current Period', 'Beginning Balance', 'Net Change'],
        data_start_row=2,
        row_count=len(line_items),
        line_items=line_items,
    )


def _build_is_tab_from_is_data(is_data) -> Optional[TabStructure]:
    """
    Build Income Statement tab from IS parsed data.
    Columns: Account Code, Account Name, PTD, PTD %, YTD, YTD %.
    is_data is a list of dicts with keys: account_code, account_name, ptd_amount, ptd_percent, ytd_amount, ytd_percent
    """
    if not is_data:
        return None

    line_items = []
    for row_num, item in enumerate(is_data, start=2):
        report_line = ReportLine(
            account_code=item.get('account_code', ''),
            account_name=item.get('account_name', ''),
            values={
                'PTD': item.get('ptd_amount', 0),
                'PTD %': item.get('ptd_percent', 0),
                'YTD': item.get('ytd_amount', 0),
                'YTD %': item.get('ytd_percent', 0),
            },
            row_number=row_num,
            is_total=False,
        )
        line_items.append(report_line)

    return TabStructure(
        name='IS',
        columns=['account_code', 'account_name', 'PTD', 'PTD %', 'YTD', 'YTD %'],
        data_start_row=2,
        row_count=len(line_items),
        line_items=line_items,
    )


def _build_t12_tab_from_gl(gl_data) -> Optional[TabStructure]:
    """
    Build Trailing 12 Months tab from GL data.

    With a single GL file we only have one month of transaction detail.
    For P&L accounts (4-8xxxxx), we use beginning_balance as YTD-prior
    and distribute it evenly across prior months as an estimate.
    The current month gets the actual net_change.
    Future months are left blank (None) rather than showing misleading zeros.
    """
    if not gl_data or not hasattr(gl_data, 'accounts'):
        return None

    # Determine current month and year from period
    period_str = getattr(gl_data.metadata, 'period', '')
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    month_num = 2  # Default
    year = 2026
    if '-' in period_str:
        name_part = period_str.split('-')[0]
        month_map = {m: i + 1 for i, m in enumerate(month_names)}
        month_num = month_map.get(name_part, 2)
        try:
            year = int(period_str.split('-')[1])
        except (IndexError, ValueError):
            pass

    month_headers = [f'{m} {year}' for m in month_names]
    month_headers.append('Total')

    line_items = []
    for account in gl_data.accounts:
        values = {}
        is_pnl = is_income_statement_account(account.account_code)
        prior_months = month_num - 1

        total = 0
        for i, month_header in enumerate(month_headers[:-1]):
            col_month = i + 1  # 1-based month number

            if col_month == month_num:
                # Current month: actual net_change
                values[month_header] = account.net_change
                total += account.net_change
            elif col_month < month_num and is_pnl and prior_months > 0:
                # Prior months for P&L: distribute beginning_balance evenly
                est = account.beginning_balance / prior_months
                values[month_header] = round(est, 2)
                total += round(est, 2)
            else:
                # Future months or BS accounts with no prior data
                values[month_header] = None

        values['Total'] = total if total != 0 else None

        report_line = ReportLine(
            account_code=account.account_code,
            account_name=account.account_name,
            values=values,
            row_number=0,
            is_total=False,
        )
        line_items.append(report_line)

    return TabStructure(
        name='T12',
        columns=['account_code', 'account_name'] + month_headers,
        data_start_row=2,
        row_count=len(line_items),
        line_items=line_items,
    )


def _build_tb_mtd_tab_from_gl(gl_data) -> Optional[TabStructure]:
    """
    Build Trial Balance MTD tab from GL data.
    Columns: Account Code, Account Name, Forward Balance, Debit, Credit, Ending Balance.
    """
    if not gl_data or not hasattr(gl_data, 'accounts'):
        return None

    line_items = []
    for account in gl_data.accounts:
        report_line = ReportLine(
            account_code=account.account_code,
            account_name=account.account_name,
            values={
                'Forward Balance': account.beginning_balance,
                'Debit': account.total_debits,
                'Credit': account.total_credits,
                'Ending Balance': account.ending_balance,
            },
            row_number=0,
            is_total=False,
        )
        line_items.append(report_line)

    return TabStructure(
        name='TB - MTD',
        columns=['account_code', 'account_name', 'Forward Balance', 'Debit', 'Credit', 'Ending Balance'],
        data_start_row=2,
        row_count=len(line_items),
        line_items=line_items,
    )


def _build_tb_ytd_tab_from_gl(gl_data) -> Optional[TabStructure]:
    """
    Build Trial Balance YTD tab from GL data.

    Differs from TB-MTD:
    - Forward Balance: always 0 for P&L accounts (YTD starts fresh each year),
      beginning_balance for BS accounts
    - Debit/Credit: YTD cumulative (beginning_balance activity + current month)
      For P&L: beginning_balance captures prior-month YTD, so YTD debits =
      prior debits (estimated from beginning) + current debits
    - Ending Balance: same as MTD (it's already the YTD position)
    """
    if not gl_data or not hasattr(gl_data, 'accounts'):
        return None

    line_items = []
    for account in gl_data.accounts:
        is_pnl = is_income_statement_account(account.account_code)

        if is_pnl:
            # P&L accounts: YTD forward balance is 0 (resets each year)
            # YTD activity = beginning_balance (prior months' net) + current month
            # We can't perfectly split prior YTD into DR/CR, so show net activity
            ytd_forward = 0
            ytd_debit = account.beginning_balance + account.total_debits if account.beginning_balance > 0 else account.total_debits
            ytd_credit = abs(account.beginning_balance) + account.total_credits if account.beginning_balance < 0 else account.total_credits
        else:
            # BS accounts: forward balance is the beginning of year
            # For simplicity with single-month data, use beginning_balance
            ytd_forward = account.beginning_balance
            ytd_debit = account.total_debits
            ytd_credit = account.total_credits

        report_line = ReportLine(
            account_code=account.account_code,
            account_name=account.account_name,
            values={
                'Forward Balance': ytd_forward,
                'Debit': ytd_debit,
                'Credit': ytd_credit,
                'Ending Balance': account.ending_balance,
            },
            row_number=0,
            is_total=False,
        )
        line_items.append(report_line)

    return TabStructure(
        name='TB - YTD',
        columns=['account_code', 'account_name', 'Forward Balance', 'Debit', 'Credit', 'Ending Balance'],
        data_start_row=2,
        row_count=len(line_items),
        line_items=line_items,
    )


# ── Main report generator ────────────────────────────────────

def generate_report(engine_result, output_path: str) -> str:
    """
    Generate the 8-tab Singerman deliverable workbook.

    Args:
        engine_result: EngineResult object with parsed data
        output_path: Where to write the Excel file

    Returns:
        The output path if successful
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default blank sheet

    # Extract parsed data
    gl_data = engine_result.parsed.get('gl')
    is_data = engine_result.parsed.get('income_statement')
    rent_roll_data = engine_result.parsed.get('rent_roll')

    # --- Tab 1: Balance Sheet (BS) ---
    bs_tab = _build_bs_tab_from_gl(gl_data)
    if bs_tab:
        _write_bs_tab(wb, bs_tab)
    else:
        _write_empty_tab(wb, 'BS')

    # --- Tab 2: Income Statement (IS) ---
    is_tab = _build_is_tab_from_is_data(is_data)
    if is_tab:
        _write_is_tab(wb, is_tab)
    else:
        _write_empty_tab(wb, 'IS')

    # --- Tab 3: Trailing 12 (T12) ---
    t12_tab = _build_t12_tab_from_gl(gl_data)
    if t12_tab:
        _write_t12_tab(wb, t12_tab)
    else:
        _write_empty_tab(wb, 'T12')

    # --- Tab 4: Trial Balance MTD (TB-MTD) ---
    tb_mtd_tab = _build_tb_mtd_tab_from_gl(gl_data)
    if tb_mtd_tab:
        _write_tb_tab(wb, tb_mtd_tab, 'TB - MTD')
    else:
        _write_empty_tab(wb, 'TB - MTD')

    # --- Tab 5: Trial Balance YTD (TB-YTD) ---
    tb_ytd_tab = _build_tb_ytd_tab_from_gl(gl_data)
    if tb_ytd_tab:
        _write_tb_tab(wb, tb_ytd_tab, 'TB - YTD')
    else:
        _write_empty_tab(wb, 'TB - YTD')

    # --- Tab 6: GL MTD ---
    if gl_data:
        _write_gl_tab(wb, gl_data, 'GL - MTD', mtd_only=True)
    else:
        _write_empty_tab(wb, 'GL - MTD')

    # --- Tab 7: GL YTD ---
    if gl_data:
        _write_gl_tab(wb, gl_data, 'GL - YTD', mtd_only=False)
    else:
        _write_empty_tab(wb, 'GL - YTD')

    # --- Tab 8: Tenancy Schedule ---
    if rent_roll_data:
        _write_tenancy_tab(wb, rent_roll_data)
    else:
        _write_empty_tab(wb, 'Tenancy Schedule')

    # Write workbook
    wb.save(output_path)
    return output_path


# ── Tab writers ─────────────────────────────────────────────

def _write_empty_tab(wb: Workbook, tab_name: str):
    """Create an empty tab with proper headers."""
    ws = wb.create_sheet(tab_name)
    ws['A1'] = f'No data available for {tab_name}'
    ws['A1'].font = Font(name='Calibri', size=11, italic=True)


def _write_bs_tab(wb: Workbook, bs_tab):
    """
    Write Balance Sheet tab.
    Columns: account_code, account_name, Balance Current Period, Beginning Balance, Net Change
    """
    ws = wb.create_sheet('BS')

    headers = ['Account Code', 'Account Name', 'Balance Current Period', 'Beginning Balance', 'Net Change']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        _apply_style(cell, _header_style())

    # Write data rows
    for row_num, line_item in enumerate(bs_tab.line_items, start=2):
        alternate = (row_num - 2) % 2 == 1

        ws.cell(row=row_num, column=1, value=line_item.account_code)
        _apply_style(ws.cell(row=row_num, column=1), _data_style(alternate))

        ws.cell(row=row_num, column=2, value=line_item.account_name)
        _apply_style(ws.cell(row=row_num, column=2), _data_style(alternate))

        for col_num, col_name in enumerate(['Balance Current Period', 'Beginning Balance', 'Net Change'], start=3):
            val = line_item.values.get(col_name, 0)
            ws.cell(row=row_num, column=col_num, value=val)
            _apply_style(ws.cell(row=row_num, column=col_num), _currency_style(alternate))

    _auto_width_columns(ws, len(headers))


def _write_is_tab(wb: Workbook, is_tab):
    """
    Write Income Statement tab.
    Columns: account_code, account_name, PTD, PTD %, YTD, YTD %
    """
    ws = wb.create_sheet('IS')

    headers = ['Account Code', 'Account Name', 'PTD', 'PTD %', 'YTD', 'YTD %']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        _apply_style(cell, _header_style())

    # Write data rows
    for row_num, line_item in enumerate(is_tab.line_items, start=2):
        alternate = (row_num - 2) % 2 == 1

        ws.cell(row=row_num, column=1, value=line_item.account_code)
        _apply_style(ws.cell(row=row_num, column=1), _data_style(alternate))

        ws.cell(row=row_num, column=2, value=line_item.account_name)
        _apply_style(ws.cell(row=row_num, column=2), _data_style(alternate))

        ptd_val = line_item.values.get('PTD', 0)
        ws.cell(row=row_num, column=3, value=ptd_val)
        _apply_style(ws.cell(row=row_num, column=3), _currency_style(alternate))

        ptd_pct = line_item.values.get('PTD %', 0)
        ws.cell(row=row_num, column=4, value=ptd_pct)
        _apply_style(ws.cell(row=row_num, column=4), _percent_style(alternate))

        ytd_val = line_item.values.get('YTD', 0)
        ws.cell(row=row_num, column=5, value=ytd_val)
        _apply_style(ws.cell(row=row_num, column=5), _currency_style(alternate))

        ytd_pct = line_item.values.get('YTD %', 0)
        ws.cell(row=row_num, column=6, value=ytd_pct)
        _apply_style(ws.cell(row=row_num, column=6), _percent_style(alternate))

    _auto_width_columns(ws, len(headers))


def _write_t12_tab(wb: Workbook, t12_tab):
    """
    Write Trailing 12 Months tab.
    Columns: account_code, account_name, Jan 2026...Dec 2026, Total.
    Prior months show estimated values from YTD beginning balance.
    Future months show blank (not zero).
    """
    ws = wb.create_sheet('T12')

    # Extract month headers from the tab structure
    months = [col for col in t12_tab.columns if col not in ['account_code', 'account_name']]
    headers = ['Account Code', 'Account Name'] + months

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        _apply_style(cell, _header_style())

    # Write data rows
    for row_num, line_item in enumerate(t12_tab.line_items, start=2):
        alternate = (row_num - 2) % 2 == 1

        ws.cell(row=row_num, column=1, value=line_item.account_code)
        _apply_style(ws.cell(row=row_num, column=1), _data_style(alternate))

        ws.cell(row=row_num, column=2, value=line_item.account_name)
        _apply_style(ws.cell(row=row_num, column=2), _data_style(alternate))

        for col_num, month in enumerate(months, start=3):
            val = line_item.values.get(month)
            if val is not None:
                ws.cell(row=row_num, column=col_num, value=val)
                _apply_style(ws.cell(row=row_num, column=col_num), _currency_style(alternate))
            else:
                # Leave blank for future months / unavailable data
                _apply_style(ws.cell(row=row_num, column=col_num), _data_style(alternate))

    _auto_width_columns(ws, len(headers))


def _write_tb_tab(wb: Workbook, tb_tab, tab_name: str):
    """
    Write Trial Balance tab (MTD or YTD).
    Columns: account_code, account_name, Forward Balance, Debit, Credit, Ending Balance
    """
    ws = wb.create_sheet(tab_name)

    headers = ['Account Code', 'Account Name', 'Forward Balance', 'Debit', 'Credit', 'Ending Balance']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        _apply_style(cell, _header_style())

    # Write data rows
    for row_num, line_item in enumerate(tb_tab.line_items, start=2):
        alternate = (row_num - 2) % 2 == 1

        ws.cell(row=row_num, column=1, value=line_item.account_code)
        _apply_style(ws.cell(row=row_num, column=1), _data_style(alternate))

        ws.cell(row=row_num, column=2, value=line_item.account_name)
        _apply_style(ws.cell(row=row_num, column=2), _data_style(alternate))

        for col_num, col_name in enumerate(['Forward Balance', 'Debit', 'Credit', 'Ending Balance'], start=3):
            val = line_item.values.get(col_name, 0)
            ws.cell(row=row_num, column=col_num, value=val)
            _apply_style(ws.cell(row=row_num, column=col_num), _currency_style(alternate))

    _auto_width_columns(ws, len(headers))


def _write_gl_tab(wb: Workbook, gl_data, tab_name: str, mtd_only: bool = False):
    """
    Write GL tab (MTD or YTD).
    Columns: Property, Property Name, Date, Period, Person/Description, Control, Reference, Debit, Credit, Balance, Remarks
    """
    ws = wb.create_sheet(tab_name)

    headers = ['Property', 'Property Name', 'Date', 'Period', 'Person/Description', 'Control',
               'Reference', 'Debit', 'Credit', 'Balance', 'Remarks']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        _apply_style(cell, _header_style())

    # Filter transactions if MTD
    transactions = gl_data.all_transactions if hasattr(gl_data, 'all_transactions') else []

    # Write data rows
    for row_num, txn in enumerate(transactions, start=2):
        alternate = (row_num - 2) % 2 == 1

        ws.cell(row=row_num, column=1, value=txn.account_code)
        _apply_style(ws.cell(row=row_num, column=1), _data_style(alternate))

        ws.cell(row=row_num, column=2, value=txn.account_name)
        _apply_style(ws.cell(row=row_num, column=2), _data_style(alternate))

        date_val = txn.date.strftime('%m/%d/%Y') if txn.date else ''
        ws.cell(row=row_num, column=3, value=date_val)
        _apply_style(ws.cell(row=row_num, column=3), _data_style(alternate))

        ws.cell(row=row_num, column=4, value=txn.period)
        _apply_style(ws.cell(row=row_num, column=4), _data_style(alternate))

        ws.cell(row=row_num, column=5, value=txn.description)
        _apply_style(ws.cell(row=row_num, column=5), _data_style(alternate))

        ws.cell(row=row_num, column=6, value=txn.control)
        _apply_style(ws.cell(row=row_num, column=6), _data_style(alternate))

        ws.cell(row=row_num, column=7, value=txn.reference)
        _apply_style(ws.cell(row=row_num, column=7), _data_style(alternate))

        ws.cell(row=row_num, column=8, value=txn.debit)
        _apply_style(ws.cell(row=row_num, column=8), _currency_style(alternate))

        ws.cell(row=row_num, column=9, value=txn.credit)
        _apply_style(ws.cell(row=row_num, column=9), _currency_style(alternate))

        ws.cell(row=row_num, column=10, value=txn.balance)
        _apply_style(ws.cell(row=row_num, column=10), _currency_style(alternate))

        ws.cell(row=row_num, column=11, value=txn.remarks)
        _apply_style(ws.cell(row=row_num, column=11), _data_style(alternate))

    _auto_width_columns(ws, len(headers))


def _write_tenancy_tab(wb: Workbook, rent_roll_data: List[Dict]):
    """
    Write Tenancy Schedule tab.
    Columns: Property, Unit(s), Lease, Lease Type, Area, Lease From, Lease To, Term,
             Tenancy Years, Monthly Rent, Monthly Rent/Area, Annual Rent, Annual Rent/Area,
             Annual Rec./Area, Annual Misc/Area, Security Deposit, LOC Amount
    """
    ws = wb.create_sheet('Tenancy Schedule')

    headers = ['Property', 'Unit(s)', 'Lease', 'Lease Type', 'Area', 'Lease From', 'Lease To',
               'Term', 'Tenancy Years', 'Monthly Rent', 'Monthly Rent/Area', 'Annual Rent',
               'Annual Rent/Area', 'Annual Rec./Area', 'Annual Misc/Area', 'Security Deposit', 'LOC Amount']

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        _apply_style(cell, _header_style())

    # Map rent roll data fields to report columns
    field_map = {
        'Property': 'property',
        'Unit(s)': 'units',
        'Lease': 'tenant',
        'Lease Type': 'lease_type',
        'Area': 'area',
        'Lease From': 'lease_from',
        'Lease To': 'lease_to',
        'Term': 'term_months',
        'Tenancy Years': 'tenancy_years',
        'Monthly Rent': 'monthly_rent',
        'Monthly Rent/Area': 'monthly_rent_per_area',
        'Annual Rent': 'annual_rent',
        'Annual Rent/Area': 'annual_rent_per_area',
        'Annual Rec./Area': 'annual_rec_per_area',
        'Annual Misc/Area': 'annual_misc_per_area',
        'Security Deposit': 'security_deposit',
        'LOC Amount': 'loc_amount',
    }

    # Write data rows
    for row_num, item in enumerate(rent_roll_data, start=2):
        alternate = (row_num - 2) % 2 == 1

        for col_num, header in enumerate(headers, start=1):
            field = field_map.get(header, header.lower().replace('/', '_').replace(' ', '_'))
            val = item.get(field) if isinstance(item, dict) else getattr(item, field, None)

            # Format dates
            if isinstance(val, date):
                val = val.strftime('%m/%d/%Y')

            ws.cell(row=row_num, column=col_num, value=val)

            # Apply appropriate style
            if header in ['Area', 'Monthly Rent', 'Monthly Rent/Area', 'Annual Rent',
                          'Annual Rent/Area', 'Annual Rec./Area', 'Annual Misc/Area',
                          'Security Deposit', 'LOC Amount']:
                _apply_style(ws.cell(row=row_num, column=col_num), _currency_style(alternate))
            else:
                _apply_style(ws.cell(row=row_num, column=col_num), _data_style(alternate))

    _auto_width_columns(ws, len(headers))


# ── Exception report generator ───────────────────────────────

def generate_exception_report(engine_result, output_path: str) -> str:
    """
    Generate an exception and validation detail report.

    Args:
        engine_result: EngineResult object with exceptions and validation data
        output_path: Where to write the Excel file

    Returns:
        The output path if successful
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default blank sheet

    # --- Tab 1: Summary ---
    _write_summary_tab(wb, engine_result)

    # --- Tab 2: Exceptions ---
    _write_exceptions_tab(wb, engine_result)

    # --- Tab 3: Budget Variances ---
    _write_budget_variances_tab(wb, engine_result)

    # --- Tab 4: Bank Recon ---
    _write_bank_recon_tab(wb, engine_result)

    # --- Tab 5: Debt Service ---
    _write_debt_service_tab(wb, engine_result)

    # Write workbook
    wb.save(output_path)
    return output_path


def _write_summary_tab(wb: Workbook, engine_result):
    """Write Summary tab with overview statistics."""
    ws = wb.create_sheet('Summary', 0)

    summary_data = engine_result.summary or {}

    # Title
    ws['A1'] = 'GA Automation Pipeline Summary'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells('A1:B1')

    # Run details
    row = 3
    ws[f'A{row}'] = 'Run ID:'
    ws[f'B{row}'] = engine_result.run_id
    row += 1
    ws[f'A{row}'] = 'Run At:'
    ws[f'B{row}'] = engine_result.run_at
    row += 1
    ws[f'A{row}'] = 'Period:'
    ws[f'B{row}'] = engine_result.period
    row += 1
    ws[f'A{row}'] = 'Property:'
    ws[f'B{row}'] = engine_result.property_name
    row += 1
    ws[f'A{row}'] = 'Status:'
    ws[f'B{row}'] = engine_result.status
    row += 2

    # Summary statistics
    ws[f'A{row}'] = 'Summary Statistics'
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True)
    row += 1

    summary_metrics = [
        ('Files Processed', summary_data.get('files_processed', 0)),
        ('Parsers Succeeded', summary_data.get('parsers_succeeded', 0)),
        ('GL Accounts', summary_data.get('gl_accounts', 0)),
        ('GL Transactions', summary_data.get('gl_transactions', 0)),
        ('GL Balanced', summary_data.get('gl_balanced', False)),
        ('Invoice Matches', summary_data.get('invoice_matches', 0)),
        ('Bank Matches', summary_data.get('bank_matches', 0)),
        ('Budget Variances Flagged', summary_data.get('budget_variances_flagged', 0)),
    ]

    for label, value in summary_metrics:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1

    row += 1
    ws[f'A{row}'] = 'Exception Summary'
    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True)
    row += 1

    exception_metrics = [
        ('Errors', engine_result.error_count),
        ('Warnings', engine_result.warning_count),
        ('Total Issues', len(engine_result.exceptions)),
    ]

    for label, value in exception_metrics:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


def _write_exceptions_tab(wb: Workbook, engine_result):
    """Write Exceptions tab with all flagged issues."""
    ws = wb.create_sheet('Exceptions')

    headers = ['Severity', 'Category', 'Source', 'Description', 'Resolved']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        _apply_style(cell, _header_style())

    # Write exception rows
    for row_num, exc in enumerate(engine_result.exceptions, start=2):
        alternate = (row_num - 2) % 2 == 1
        style = _data_style(alternate)

        # Color code by severity
        if exc.severity == 'error':
            style['fill'] = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        elif exc.severity == 'warning':
            style['fill'] = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

        ws.cell(row=row_num, column=1, value=exc.severity)
        _apply_style(ws.cell(row=row_num, column=1), style)

        ws.cell(row=row_num, column=2, value=exc.category)
        _apply_style(ws.cell(row=row_num, column=2), style)

        ws.cell(row=row_num, column=3, value=exc.source)
        _apply_style(ws.cell(row=row_num, column=3), style)

        ws.cell(row=row_num, column=4, value=exc.description)
        _apply_style(ws.cell(row=row_num, column=4), style)

        ws.cell(row=row_num, column=5, value='Yes' if exc.resolved else 'No')
        _apply_style(ws.cell(row=row_num, column=5), style)

    _auto_width_columns(ws, len(headers))


def _write_budget_variances_tab(wb: Workbook, engine_result):
    """Write Budget Variances tab."""
    ws = wb.create_sheet('Budget Variances')

    headers = ['Account Code', 'Account Name', 'PTD Actual', 'PTD Budget', 'Variance', 'Variance %']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        _apply_style(cell, _header_style())

    # Write variance rows
    for row_num, var in enumerate(engine_result.budget_variances, start=2):
        alternate = (row_num - 2) % 2 == 1

        ws.cell(row=row_num, column=1, value=var.get('account_code', ''))
        _apply_style(ws.cell(row=row_num, column=1), _data_style(alternate))

        ws.cell(row=row_num, column=2, value=var.get('account_name', ''))
        _apply_style(ws.cell(row=row_num, column=2), _data_style(alternate))

        ws.cell(row=row_num, column=3, value=var.get('ptd_actual', 0))
        _apply_style(ws.cell(row=row_num, column=3), _currency_style(alternate))

        ws.cell(row=row_num, column=4, value=var.get('ptd_budget', 0))
        _apply_style(ws.cell(row=row_num, column=4), _currency_style(alternate))

        ws.cell(row=row_num, column=5, value=var.get('variance', 0))
        _apply_style(ws.cell(row=row_num, column=5), _currency_style(alternate))

        var_pct = var.get('variance_pct', 0)
        ws.cell(row=row_num, column=6, value=var_pct / 100 if var_pct else 0)
        _apply_style(ws.cell(row=row_num, column=6), _percent_style(alternate))

    if not engine_result.budget_variances:
        ws['A2'] = 'No material variances flagged'
        ws['A2'].font = Font(name='Calibri', size=11, italic=True)

    _auto_width_columns(ws, len(headers))


def _write_bank_recon_tab(wb: Workbook, engine_result):
    """Write Bank Reconciliation tab."""
    ws = wb.create_sheet('Bank Recon')

    headers = ['Source A', 'Source B', 'Key', 'Amount A', 'Amount B', 'Matched', 'Variance', 'Description']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        _apply_style(cell, _header_style())

    # Write match rows
    for row_num, match in enumerate(engine_result.gl_bank_matches, start=2):
        alternate = (row_num - 2) % 2 == 1

        ws.cell(row=row_num, column=1, value=match.source_a)
        _apply_style(ws.cell(row=row_num, column=1), _data_style(alternate))

        ws.cell(row=row_num, column=2, value=match.source_b)
        _apply_style(ws.cell(row=row_num, column=2), _data_style(alternate))

        ws.cell(row=row_num, column=3, value=match.key)
        _apply_style(ws.cell(row=row_num, column=3), _data_style(alternate))

        ws.cell(row=row_num, column=4, value=match.amount_a)
        _apply_style(ws.cell(row=row_num, column=4), _currency_style(alternate))

        ws.cell(row=row_num, column=5, value=match.amount_b)
        _apply_style(ws.cell(row=row_num, column=5), _currency_style(alternate))

        ws.cell(row=row_num, column=6, value='Yes' if match.matched else 'No')
        _apply_style(ws.cell(row=row_num, column=6), _data_style(alternate))

        ws.cell(row=row_num, column=7, value=match.variance)
        _apply_style(ws.cell(row=row_num, column=7), _currency_style(alternate))

        ws.cell(row=row_num, column=8, value=match.description)
        _apply_style(ws.cell(row=row_num, column=8), _data_style(alternate))

    if not engine_result.gl_bank_matches:
        ws['A2'] = 'No bank reconciliation data available'
        ws['A2'].font = Font(name='Calibri', size=11, italic=True)

    _auto_width_columns(ws, len(headers))


def _write_debt_service_tab(wb: Workbook, engine_result):
    """Write Debt Service tab."""
    ws = wb.create_sheet('Debt Service')

    ds_check = engine_result.debt_service_check or {}

    # Title
    ws['A1'] = 'Debt Service Reconciliation'
    ws['A1'].font = Font(name='Calibri', size=12, bold=True)

    # Summary
    row = 3
    ws[f'A{row}'] = 'GL Interest Expense:'
    ws[f'B{row}'] = ds_check.get('gl_interest_expense', 0)
    ws[f'B{row}'].number_format = '$#,##0.00'
    row += 1
    ws[f'A{row}'] = 'Loan Interest Total (YTD):'
    ws[f'B{row}'] = ds_check.get('loan_interest_total', 0)
    ws[f'B{row}'].number_format = '$#,##0.00'
    row += 1
    ws[f'A{row}'] = 'Variance:'
    ws[f'B{row}'] = ds_check.get('variance', 0)
    ws[f'B{row}'].number_format = '$#,##0.00'
    row += 1
    ws[f'A{row}'] = 'Reconciled:'
    ws[f'B{row}'] = 'Yes' if ds_check.get('reconciled', False) else 'No'
    row += 2

    # Loan details
    ws[f'A{row}'] = 'Loan Details'
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
    row += 1

    loan_headers = ['Loan Name', 'Principal Balance', 'Interest Paid YTD']
    for col_num, header in enumerate(loan_headers, start=1):
        cell = ws.cell(row=row, column=col_num, value=header)
        _apply_style(cell, _header_style())
    row += 1

    for loan in ds_check.get('loans', []):
        ws[f'A{row}'] = loan.get('name', '')
        ws[f'B{row}'] = loan.get('principal_balance', 0)
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'C{row}'] = loan.get('interest_paid_ytd', 0)
        ws[f'C{row}'].number_format = '$#,##0.00'
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20


# ── Test/Demo ────────────────────────────────────────────────

if __name__ == "__main__":
    """
    Simple test: create a minimal EngineResult and generate a report.
    """
    from engine import EngineResult, Exception_
    from datetime import datetime

    # Create a minimal engine result
    result = EngineResult(
        run_id="TEST_20260403_120000",
        run_at=datetime.now().isoformat(),
        period="Mar-2026",
        property_name="Test Property",
    )

    # Add some test exceptions
    result.add_exception(
        "error", "balance", "gl_validation",
        "GL is not balanced",
        total_debits=100000.00,
        total_credits=99999.50,
    )
    result.add_exception(
        "warning", "variance", "budget_comparison",
        "Material variance in account 401000",
        variance=5000.00,
        variance_pct=15.5,
    )

    # Add test budget variances
    result.budget_variances = [
        {
            'account_code': '401000',
            'account_name': 'Rental Income',
            'ptd_actual': 50000.00,
            'ptd_budget': 45000.00,
            'variance': 5000.00,
            'variance_pct': 11.1,
        }
    ]

    # Add test GL bank matches
    from engine import MatchResult
    result.gl_bank_matches = [
        MatchResult(
            source_a="GL",
            source_b="Bank",
            key="Ending Balance",
            amount_a=250000.00,
            amount_b=249500.00,
            matched=False,
            variance=500.00,
            description="GL Cash vs Bank Balance",
        )
    ]

    # Set summary
    result.summary = {
        'files_processed': 5,
        'parsers_succeeded': 5,
        'gl_accounts': 75,
        'gl_transactions': 500,
        'gl_balanced': False,
        'invoice_matches': 45,
        'bank_matches': 10,
        'budget_variances_flagged': 3,
        'exceptions_error': 1,
        'exceptions_warning': 5,
        'status': 'WARNINGS',
    }

    # Generate reports
    output_dir = '/tmp'
    main_report_path = generate_report(result, f'{output_dir}/test_main_report.xlsx')
    exception_report_path = generate_exception_report(result, f'{output_dir}/test_exception_report.xlsx')

    print(f"Main report generated: {main_report_path}")
    print(f"Exception report generated: {exception_report_path}")

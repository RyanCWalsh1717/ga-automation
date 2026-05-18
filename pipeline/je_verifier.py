"""
JE Verification — Pass 1 → Pass 2 Posting Check
=================================================
Compares journal entries generated in Pass 1 against J-type transactions in
the final (post-close) GL to confirm that every line actually posted to Yardi.

Matching strategy (in priority order):
  1. Reference + Account Code — most reliable: REF column from the ETL CSV
     becomes the GL ``reference`` field verbatim.  Works for Nexus invoices
     (reference = invoice number), management fee (MGMT-FEE-JLL / MGMT-FEE-GRP),
     and catch-up JEs (MGMT-CATCHUP).
  2. Description + Account Code — fallback when REF = je_number or is generic.
     REMARK column → GL description / remarks field.  Uses substring match so
     minor Yardi truncation (60-char limit) doesn't break the match.
  3. Amount + Account Code among J-type — last resort for JEs with no
     distinguishing reference or description.

Status per JE number:
  VERIFIED        — all DR/CR lines found in GL with matching amounts (±$0.02)
  AMOUNT_MISMATCH — all lines found but at least one amount differs > $0.02
  PARTIAL         — only some lines found in GL
  MISSING         — no lines found at all in GL

Usage:
    from je_verifier import verify_je_posting

    result = verify_je_posting(
        pass1_je_lines = st.session_state['pass1_output_files']['all_je_lines'],
        gl_parsed      = engine_result.parsed['gl'],
    )
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple


# ── Tolerance ─────────────────────────────────────────────────────────────────
_AMOUNT_TOLERANCE = 0.02   # rounding differences accepted as a match


# ── Dataclasses ───────────────────────────────────────────────────────────────

@dataclass
class JELineResult:
    """Verification result for a single JE line (DR or CR)."""
    je_number:    str
    line:         int           # line number within the JE
    account_code: str
    account_name: str
    expected_dr:  float
    expected_cr:  float
    signed_amount: float        # positive = DR, negative = CR
    match_status: str           # 'matched' | 'amount_mismatch' | 'not_found'
    match_method: str           # 'reference' | 'description' | 'amount' | 'none'
    actual_dr:    float = 0.0
    actual_cr:    float = 0.0
    gl_reference: str = ''
    gl_description: str = ''
    note: str = ''


@dataclass
class JEResult:
    """Verification result for one complete journal entry (all its lines)."""
    je_number:    str
    source:       str           # 'nexus', 'historical', 'management_fee', etc.
    line_count:   int
    status:       str           # 'VERIFIED' | 'AMOUNT_MISMATCH' | 'PARTIAL' | 'MISSING'
    lines:        List[JELineResult] = field(default_factory=list)

    @property
    def verified_count(self) -> int:
        return sum(1 for l in self.lines if l.match_status in ('matched', 'amount_mismatch'))

    @property
    def missing_count(self) -> int:
        return sum(1 for l in self.lines if l.match_status == 'not_found')


@dataclass
class JEVerificationResult:
    """Top-level result returned by verify_je_posting()."""
    total_je_count:    int = 0
    verified_count:    int = 0
    mismatch_count:    int = 0
    partial_count:     int = 0
    missing_count:     int = 0
    je_results:        List[JEResult] = field(default_factory=list)
    unmatched_gl_jtxns: int = 0   # J-type GL txns with no Pass 1 counterpart

    @property
    def all_verified(self) -> bool:
        return self.missing_count == 0 and self.partial_count == 0

    @property
    def summary(self) -> str:
        parts = [f"{self.verified_count}/{self.total_je_count} JEs verified"]
        if self.missing_count:
            parts.append(f"{self.missing_count} MISSING")
        if self.partial_count:
            parts.append(f"{self.partial_count} PARTIAL")
        if self.mismatch_count:
            parts.append(f"{self.mismatch_count} AMOUNT MISMATCH")
        return " — ".join(parts)


# ── GL index builder ──────────────────────────────────────────────────────────

def _build_gl_j_index(gl_parsed) -> Dict[str, list]:
    """
    Build an index of all J-type transactions from the final GL.

    Returns:
        {
            'by_ref_acct':   {(reference, account_code): [txns]}
            'by_desc_acct':  {(description_lower, account_code): [txns]}
            'by_amt_acct':   {(signed_amount, account_code): [txns]}
            'all':           [all J-type txns]
        }
    """
    idx: Dict[str, Any] = {
        'by_ref_acct':  {},
        'by_desc_acct': {},
        'by_amt_acct':  {},
        'all':          [],
    }

    if not gl_parsed:
        return idx

    for acct in (gl_parsed.accounts if hasattr(gl_parsed, 'accounts') else []):
        code = str(acct.account_code).strip()
        for txn in getattr(acct, 'transactions', []):
            ctrl = str(getattr(txn, 'control', '') or '').upper()
            if not ctrl.startswith('J'):
                continue

            ref  = str(getattr(txn, 'reference',   '') or '').strip()
            desc = str(getattr(txn, 'description', '') or '').strip().lower()
            rmk  = str(getattr(txn, 'remarks',     '') or '').strip().lower()
            dr   = float(getattr(txn, 'debit',  0) or 0)
            cr   = float(getattr(txn, 'credit', 0) or 0)
            signed = round(dr - cr, 2)

            rec = {
                'account_code': code,
                'reference':    ref,
                'description':  desc,
                'remarks':      rmk,
                'debit':        dr,
                'credit':       cr,
                'signed':       signed,
                '_txn':         txn,
                '_used':        False,   # consumed-flag prevents double-matching
            }
            idx['all'].append(rec)

            # Index 1: reference + account
            if ref:
                key = (ref, code)
                idx['by_ref_acct'].setdefault(key, []).append(rec)

            # Index 2: description + account (both desc and remarks)
            for txt in {desc, rmk}:
                if txt:
                    key2 = (txt, code)
                    idx['by_desc_acct'].setdefault(key2, []).append(rec)

            # Index 3: signed amount + account
            key3 = (signed, code)
            idx['by_amt_acct'].setdefault(key3, []).append(rec)

    return idx


# ── Single-line matcher ───────────────────────────────────────────────────────

def _match_line(pass1_line: dict, gl_idx: dict) -> Tuple[str, str, dict]:
    """
    Try to find a matching GL J-type transaction for one Pass 1 JE line.

    Returns (match_status, match_method, gl_rec | {}).
      match_status: 'matched' | 'amount_mismatch' | 'not_found'
      match_method: 'reference' | 'description' | 'amount' | 'none'
    """
    acct       = str(pass1_line.get('account_code', '') or '').strip()
    ref_key    = str(pass1_line.get('reference', '') or pass1_line.get('je_number', '') or '').strip()
    desc_raw   = str(pass1_line.get('description', '') or '').strip().lower()
    dr         = float(pass1_line.get('debit',  0) or 0)
    cr         = float(pass1_line.get('credit', 0) or 0)
    signed     = round(dr - cr, 2)

    def _amount_ok(gl_rec: dict) -> bool:
        return abs(gl_rec['signed'] - signed) <= _AMOUNT_TOLERANCE

    def _consume(gl_rec: dict, method: str) -> Tuple[str, str, dict]:
        """Mark the GL record as consumed and return the match result."""
        gl_rec['_used'] = True
        if _amount_ok(gl_rec):
            return 'matched', method, gl_rec
        return 'amount_mismatch', method, gl_rec

    # ── Strategy 1: reference + account ──────────────────────────
    if ref_key:
        candidates = [r for r in gl_idx['by_ref_acct'].get((ref_key, acct), [])
                      if not r['_used']]
        if candidates:
            # Prefer exact amount match; fall back to first candidate
            exact = [c for c in candidates if _amount_ok(c)]
            return _consume(exact[0] if exact else candidates[0], 'reference')

    # ── Strategy 2: description + account ────────────────────────
    # Yardi may truncate; check if Pass 1 description is a substring of GL desc/remarks
    if desc_raw:
        # Use first 40 chars of description as the search key (Yardi truncates at 60)
        desc_key = desc_raw[:40]
        for (gl_txt, gl_acct), recs in gl_idx['by_desc_acct'].items():
            if gl_acct != acct:
                continue
            if desc_key in gl_txt or gl_txt[:40] in desc_raw:
                unused = [r for r in recs if not r['_used']]
                if unused:
                    return _consume(unused[0], 'description')

    # ── Strategy 3: signed amount + account ──────────────────────
    candidates = [r for r in gl_idx['by_amt_acct'].get((signed, acct), [])
                  if not r['_used']]
    if candidates:
        return _consume(candidates[0], 'amount')

    return 'not_found', 'none', {}


# ── Main entry point ──────────────────────────────────────────────────────────

def verify_je_posting(
    pass1_je_lines: List[Dict],
    gl_parsed,
) -> JEVerificationResult:
    """
    Verify that all Pass 1 JE lines posted correctly to the final (post-close) GL.

    Args:
        pass1_je_lines: List of JE line dicts from pass1_output_files['all_je_lines'].
                        Each dict must have: je_number, account_code, debit, credit,
                        reference (optional), description (optional), source (optional).
        gl_parsed:      Parsed final GL (GLParseResult from yardi_gl parser).

    Returns:
        JEVerificationResult with per-JE status and full line-level detail.
    """
    result = JEVerificationResult()

    if not pass1_je_lines or not gl_parsed:
        return result

    # Build GL J-type index
    gl_idx = _build_gl_j_index(gl_parsed)

    # Group Pass 1 lines by je_number (preserving order)
    grouped: Dict[str, List[dict]] = {}
    for line in pass1_je_lines:
        je_num = str(line.get('je_number', '') or '').strip()
        if not je_num:
            continue
        grouped.setdefault(je_num, []).append(line)

    result.total_je_count = len(grouped)

    for je_num, lines in grouped.items():
        source = str(lines[0].get('source', '') or '') if lines else ''
        je_result = JEResult(
            je_number  = je_num,
            source     = source,
            line_count = len(lines),
            status     = 'MISSING',   # updated below
        )

        for idx, line in enumerate(lines, 1):
            acct = str(line.get('account_code', '') or '').strip()
            dr   = float(line.get('debit',  0) or 0)
            cr   = float(line.get('credit', 0) or 0)

            match_status, match_method, gl_rec = _match_line(line, gl_idx)

            lr = JELineResult(
                je_number     = je_num,
                line          = idx,
                account_code  = acct,
                account_name  = str(line.get('account_name', '') or ''),
                expected_dr   = dr,
                expected_cr   = cr,
                signed_amount = round(dr - cr, 2),
                match_status  = match_status,
                match_method  = match_method,
                actual_dr     = gl_rec.get('debit', 0.0),
                actual_cr     = gl_rec.get('credit', 0.0),
                gl_reference  = gl_rec.get('reference', ''),
                gl_description= gl_rec.get('description', ''),
            )

            if match_status == 'amount_mismatch':
                diff = abs(gl_rec['signed'] - lr.signed_amount)
                lr.note = f'Amount differs by ${diff:,.2f}'

            je_result.lines.append(lr)

        # Roll up line statuses → JE status
        statuses = {l.match_status for l in je_result.lines}
        if statuses == {'matched'}:
            je_result.status = 'VERIFIED'
            result.verified_count += 1
        elif 'not_found' not in statuses:
            # All lines found but at least one amount differs
            je_result.status = 'AMOUNT_MISMATCH'
            result.mismatch_count += 1
        elif statuses == {'not_found'}:
            je_result.status = 'MISSING'
            result.missing_count += 1
        else:
            # Mix of found and not_found
            je_result.status = 'PARTIAL'
            result.partial_count += 1

        result.je_results.append(je_result)

    # Count unmatched GL J-type transactions (GL has JEs we didn't generate)
    result.unmatched_gl_jtxns = sum(1 for r in gl_idx['all'] if not r['_used'])

    return result


# ── QC Workbook tab builder ───────────────────────────────────────────────────

def write_je_verification_tab(wb, result: JEVerificationResult, period: str = '') -> None:
    """
    Add a 'JE Verification' sheet to an existing openpyxl Workbook.

    Tab layout:
      Row 1:   Header banner
      Rows 3-5: Summary counts
      Row 7:   Column headers
      Row 8+:  One row per JE result (expandable detail via alternating rows)
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # ── Palette ───────────────────────────────────────────────────
    _GRP_GREEN  = '1A5C22'
    _GRP_MID    = '2E7D32'
    _WHITE      = 'FFFFFF'
    _GREEN_LITE = 'E8F5E9'
    _AMBER_LITE = 'FFF9C4'
    _RED_LITE   = 'FFEBEE'
    _GREY_LITE  = 'F5F5F5'
    _BLACK      = '000000'

    def _fnt(bold=False, size=10, color=_BLACK, italic=False):
        return Font(name='Calibri', size=size, bold=bold, color=color, italic=italic)

    def _fill(hex_color):
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

    THIN = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    STATUS_FILL = {
        'VERIFIED':        _GREEN_LITE,
        'AMOUNT_MISMATCH': _AMBER_LITE,
        'PARTIAL':         _AMBER_LITE,
        'MISSING':         _RED_LITE,
    }
    STATUS_ICON = {
        'VERIFIED': 'PASS', 'AMOUNT_MISMATCH': 'FLAG',
        'PARTIAL': 'FLAG', 'MISSING': 'FAIL',
    }

    ws = wb.create_sheet('JE Verification')
    ws.sheet_properties.tabColor = '1A5C22' if result.all_verified else 'FF6D00'

    ws.column_dimensions['A'].width = 14   # JE Number
    ws.column_dimensions['B'].width = 10   # Status
    ws.column_dimensions['C'].width = 22   # Source
    ws.column_dimensions['D'].width = 8    # Lines
    ws.column_dimensions['E'].width = 8    # Verified
    ws.column_dimensions['F'].width = 8    # Missing
    ws.column_dimensions['G'].width = 55   # Notes

    # Row 1: Banner
    ws.merge_cells('A1:G1')
    c = ws.cell(row=1, column=1,
                value=f'JE VERIFICATION — {period}')
    c.font      = _fnt(bold=True, size=12, color=_WHITE)
    c.fill      = _fill(_GRP_GREEN)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 20

    # Row 2: subtitle
    ws.merge_cells('A2:G2')
    sub = ws.cell(row=2, column=1,
                  value='Confirms that every Pass 1 journal entry appears in the final GL as a J-type transaction.')
    sub.font      = _fnt(size=9, italic=True, color='616161')
    sub.alignment = Alignment(horizontal='left')

    # Rows 3-5: summary
    summary_data = [
        ('Total JEs',       result.total_je_count,  _WHITE),
        ('Verified',        result.verified_count,   _GREEN_LITE),
        ('Amount Mismatch', result.mismatch_count,   _AMBER_LITE if result.mismatch_count else _GREEN_LITE),
        ('Partial',         result.partial_count,    _AMBER_LITE if result.partial_count  else _GREEN_LITE),
        ('Missing',         result.missing_count,    _RED_LITE   if result.missing_count  else _GREEN_LITE),
    ]
    for ri, (lbl, val, bg) in enumerate(summary_data, 3):
        lc = ws.cell(row=ri, column=1, value=lbl)
        vc = ws.cell(row=ri, column=2, value=val)
        for c in (lc, vc):
            c.font      = _fnt(bold=(lbl in ('Verified', 'Missing')))
            c.fill      = _fill(bg)
            c.border    = THIN
            c.alignment = Alignment(horizontal='center' if c.column == 2 else 'left')

    # Row 7: column headers
    hdr_row = 7
    headers = ['JE Number', 'Status', 'Source', 'Lines', 'Verified', 'Missing', 'Notes']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=hdr_row, column=ci, value=h)
        c.font      = _fnt(bold=True, color=_WHITE)
        c.fill      = _fill(_GRP_MID)
        c.border    = THIN
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[hdr_row].height = 28

    # Rows 8+: JE results
    row = hdr_row + 1
    for je in sorted(result.je_results,
                     key=lambda j: (0 if j.status == 'MISSING' else
                                    1 if j.status in ('PARTIAL', 'AMOUNT_MISMATCH') else 2,
                                    j.je_number)):
        bg = STATUS_FILL.get(je.status, _WHITE)
        icon = STATUS_ICON.get(je.status, '')

        # Build notes string
        notes_parts = []
        for ln in je.lines:
            if ln.match_status == 'not_found':
                notes_parts.append(
                    f'Line {ln.line} ({ln.account_code}) not found in GL'
                )
            elif ln.match_status == 'amount_mismatch':
                notes_parts.append(
                    f'Line {ln.line} ({ln.account_code}): {ln.note}'
                )
            elif ln.match_method != 'reference':
                notes_parts.append(
                    f'Line {ln.line} matched by {ln.match_method} (ref not found)'
                )
        notes = '; '.join(notes_parts) if notes_parts else ''

        vals = [
            je.je_number,
            f'{icon} {je.status}',
            je.source.replace('_', ' ').title(),
            je.line_count,
            je.verified_count,
            je.missing_count,
            notes,
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill      = _fill(bg)
            c.font      = _fnt(bold=(ci == 2))
            c.border    = THIN
            c.alignment = Alignment(
                horizontal='center' if ci in (2, 4, 5, 6) else 'left',
                wrap_text=(ci == 7),
            )
        ws.row_dimensions[row].height = 16
        row += 1

    # Freeze header row
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

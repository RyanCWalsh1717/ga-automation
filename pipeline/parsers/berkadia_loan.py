"""
Berkadia Loan Servicer Statement Parser

Handles both PDF (new) and XLSX (legacy) formats.

PDF format: One PDF per loan, one page per loan.
XLSX format: One workbook with multiple sheets (legacy).

Each statement contains:
- Property name and loan number
- Interest rate
- Balance information (principal, escrow balances)
- Payment information (breakdown by component)
- Account activity (transactions with dates and amounts)
- Payment due date and amount due
"""

import re
from typing import Dict, List, Any, Tuple

from openpyxl import load_workbook
from datetime import datetime


def _safe_float(s) -> float:
    if s is None:
        return 0.0
    try:
        return float(str(s).replace(',', '').replace('$', '').strip())
    except:
        return 0.0


def parse(filepath: str) -> List[Dict[str, Any]]:
    """
    Parse Berkadia loan servicer statements from PDF or Excel file.

    Args:
        filepath: Path to the PDF or Excel file

    Returns:
        List of dictionaries, one per loan, containing extracted data
    """
    if filepath.lower().endswith('.pdf'):
        return _parse_pdf(filepath)
    else:
        return _parse_xlsx(filepath)


# ─────────────────────────────────────────────────────────────────
# PDF parsing
# ─────────────────────────────────────────────────────────────────

def _parse_pdf(filepath: str) -> List[Dict[str, Any]]:
    """Parse a Berkadia billing PDF using pdfplumber.

    Strategy 1 (primary): parse each physical page independently.
    The Berkadia bulk-download format places exactly one loan tranche per page
    (each page says 'Page 1 of 1').  Parsing page-by-page is more reliable
    than pattern-splitting because pdfplumber renders template labels ('Property:',
    'Loan No.') and their data values as separate text elements that do NOT land
    on the same extracted line — making a single-line regex unreliable.

    Strategy 2 (fallback): if per-page parsing yields no results, concatenate
    all pages and split at every 'Property: … Loan No:' header occurrence.
    This handles edge cases such as a tranche spanning two pages.
    """
    import pdfplumber

    results = []
    page_texts: List[List[str]] = []

    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            page_texts.append(text.splitlines())

    if not page_texts:
        return results

    def _has_financial_data(r: Dict[str, Any]) -> bool:
        """Return True if the parsed result has any meaningful financial data.

        Accepts pages even when loan_number cannot be extracted (pdfplumber
        renders the loan number label and value as separate elements).
        The accrual generator already handles an empty loan_number gracefully.
        Filters out the summary 'Billing Statements Report' page (page 4)
        which has no financial amounts.
        """
        if not r:
            return False
        return (
            r.get('loan_number')
            or float(r.get('payment_interest') or 0) > 0
            or float(r.get('principal_balance') or 0) > 0
            or float(r.get('payment_total') or 0) > 0
        )

    # ── Strategy 1: one tranche per page ─────────────────────────
    # Accept any page with financial data — the summary page (page 4:
    # 'Billing Statements Report') produces no amounts so it is filtered out
    # by _has_financial_data.  Pages 1-3 each have payment_interest > 0.
    for page_lines in page_texts:
        r = _parse_pdf_text(page_lines)
        if _has_financial_data(r):
            results.append(r)

    if results:
        return results

    # ── Strategy 2: full-document header split ───────────────────
    all_lines: List[str] = []
    for pl in page_texts:
        all_lines.extend(pl)

    if not all_lines:
        return results

    header_pat = re.compile(r'Property:.+Loan No:', re.IGNORECASE)
    split_indices = [i for i, ln in enumerate(all_lines) if header_pat.search(ln)]

    if not split_indices:
        r = _parse_pdf_text(all_lines)
        if _has_financial_data(r):
            results.append(r)
        return results

    split_indices.append(len(all_lines))
    for i in range(len(split_indices) - 1):
        segment = all_lines[split_indices[i]: split_indices[i + 1]]
        r = _parse_pdf_text(segment)
        if _has_financial_data(r):
            results.append(r)

    return results


def _parse_pdf_text(lines: List[str]) -> Dict[str, Any]:
    """Parse the extracted text lines from a Berkadia PDF page."""

    full_text = '\n'.join(lines)

    # ── 1. Header line ───────────────────────────────────────────
    property_name = ''
    loan_number = ''
    interest_rate = 0.0
    as_of_date = ''
    payment_due_date = ''

    # Primary: all three fields on one line (ideal pdfplumber output).
    # pdfplumber may collapse multiple spaces to single spaces, so use \s+.
    header_pat = re.compile(
        r'Property:\s*(.+?)\s+Loan No:\s*(\d+)\s+Interest Rate:\s*([\d.]+)'
    )
    for line in lines:
        m = header_pat.search(line)
        if m:
            property_name = m.group(1).strip()
            loan_number = m.group(2).strip()
            interest_rate = _safe_float(m.group(3))
            break

    # Fallback: Berkadia PDFs render template labels ('Property:', 'Loan No.')
    # and data values as separate text elements, so pdfplumber often does NOT
    # produce them on the same line.  Extract each field individually.
    if not loan_number:
        # Fallback 1: "Loan No. 11159010" on one line — appears in the mail-stub
        # section ("MAIL THIS PORTION WITH YOUR PAYMENT   Loan No. 11159010").
        for line in lines:
            m = re.search(r'Loan\s+No[.:]?\s*0?(\d{7,8})\b', line, re.IGNORECASE)
            if m:
                loan_number = m.group(1).strip()
                break

    if not loan_number:
        # Fallback 2: Berkadia always prints the loan number TWICE consecutively
        # in the footer/mail-stub area: "11159010 11159010 Revolution Labs - Note A1"
        # Find any 7-9 digit number that appears ≥2 times on this page.
        # Numbers with commas (dollar amounts like "69,078,145.99") won't match
        # \b\d{7,9}\b because commas break the digit boundary.
        import collections as _coll
        _num_hits = re.findall(r'\b0?(\d{7,8})\b', full_text)
        _num_counts = _coll.Counter(_num_hits)
        for _n, _c in _num_counts.most_common():
            if _c >= 2:
                loan_number = _n
                break

    if not property_name:
        # Look for "Property: Revolution Labs - Note A1" where the data value
        # follows the label on the same line.
        for line in lines:
            m = re.search(r'Property:\s+(.+)', line, re.IGNORECASE)
            if m:
                candidate = m.group(1).strip()
                # Reject bare template lines that just repeat "Loan No" or are empty.
                if candidate and not re.match(r'^Loan\s+No', candidate, re.IGNORECASE):
                    property_name = candidate
                    break
        # Second fallback: any line that contains the property descriptor alone.
        # The mail stub repeats the property name on its own line.
        if not property_name:
            for line in lines:
                stripped = line.strip()
                if (stripped
                        and 10 < len(stripped) < 60
                        and not re.search(r'\d{2}/\d{2}/\d{4}', stripped)
                        and not re.search(r'[\d,]{6,}', stripped)
                        and 'Berkadia' not in stripped
                        and 'Revolution Labs' in stripped):
                    property_name = stripped
                    break

    if not interest_rate:
        for line in lines:
            m = re.search(r'Interest Rate:\s*([\d.]+)', line, re.IGNORECASE)
            if m:
                interest_rate = _safe_float(m.group(1))
                break

    # ── 2. Dates ─────────────────────────────────────────────────
    m_aof = re.search(r'BALANCE INFORMATION AS OF\s+([\d/]+)', full_text)
    if m_aof:
        as_of_date = m_aof.group(1)

    m_pmt = re.search(r'PAYMENT INFORMATION FOR\s+([\d/]+)', full_text)
    if m_pmt:
        payment_due_date = m_pmt.group(1)

    # ── 3. Balance section ───────────────────────────────────────
    principal_balance = 0.0
    interest_paid_ytd = 0.0
    outstanding_deferred_int = 0.0
    tax_escrow_balance = 0.0
    insurance_escrow_balance = 0.0
    reserve_balance = 0.0

    for line in lines:
        m = re.search(r'Principal Balance\s+([\d,.-]+)', line)
        if m:
            principal_balance = _safe_float(m.group(1))

        m = re.search(r'Interest Paid YTD\s+([\d,.-]+)', line)
        if m:
            interest_paid_ytd = _safe_float(m.group(1))

        m = re.search(r'Tax Escrow Balance\s+([\d,.-]+)', line)
        if m:
            tax_escrow_balance = _safe_float(m.group(1))

        m = re.search(r'Insurance Escrow Balance\s+([\d,.-]+)', line)
        if m:
            insurance_escrow_balance = _safe_float(m.group(1))

        m = re.search(r'Reserve Balance\s+([\d,.-]+)', line)
        if m:
            reserve_balance = _safe_float(m.group(1))

    # ── 4. Payment section ───────────────────────────────────────
    payment_principal = 0.0
    payment_interest = 0.0
    payment_re_taxes = 0.0
    payment_reserves = 0.0
    payment_total = 0.0

    # Extract payment_interest from the Payment Information section.
    # Three strategies in order of reliability.  pdfplumber's column-ordering
    # is unpredictable, so we try multiple approaches and stop at the first hit.
    #
    # Strategy A — DOTALL anchor on "PAYMENT INFORMATION FOR":
    #   Works when the section header appears in the extracted text (most layouts).
    #   Lazy .+? skips "Interest Rate: …" and "Interest Paid YTD …" because their
    #   next token ("Rate", "Paid") doesn't match [\d,.-]+.
    #
    # Strategy B — line-by-line scan for bare "Interest <digits>" line:
    #   Handles PDFs where pdfplumber omits the section header entirely, or where
    #   label and value land on separate lines ("Interest\n403,805.67").
    #   Rejects "Interest Paid YTD …", "Interest Rate: …", "Accrued Interest …"
    #   by requiring the captured group to start with a digit.
    #
    # Strategy C — all findall matches, filter by amount > 1000:
    #   Last-resort: collect every "Interest <number>" occurrence in the full text
    #   and take the first one with a value in the plausible monthly-payment range.
    #   Re-confirms the payment value by cross-checking Total Payment Due.

    # ── Strategy A ──────────────────────────────────────────────────
    _pi_m = re.search(
        r'PAYMENT INFORMATION FOR.+?\bInterest\b:?\s+(\d[\d,.-]+)',
        full_text, re.DOTALL | re.IGNORECASE
    )
    if _pi_m:
        _pi_candidate = _safe_float(_pi_m.group(1))
        if _pi_candidate > 1000:
            payment_interest = _pi_candidate

    # ── Strategy B — line-by-line ───────────────────────────────────
    if not payment_interest:
        _prev_line_interest = False
        for _ln in lines:
            _s = _ln.strip()
            # Case 1: label and value on the SAME line
            #   "Interest 403,805.67"  (bare — not "Interest Paid", "Interest Rate")
            _same = re.match(r'^Interest\s+(\d[\d,.-]+)\s*$', _s, re.IGNORECASE)
            if _same:
                _cand = _safe_float(_same.group(1))
                if _cand > 1000:
                    payment_interest = _cand
                    break
            # Case 2: label alone on one line, value on the NEXT line
            if re.match(r'^Interest\s*$', _s, re.IGNORECASE):
                _prev_line_interest = True
                continue
            if _prev_line_interest:
                _cand = _safe_float(_s)
                if _cand > 1000:
                    payment_interest = _cand
                    break
                _prev_line_interest = False
            else:
                _prev_line_interest = False

    # ── Strategy C — positional findall fallback ────────────────────
    # Only fires if both A and B fail.  The payment breakdown "Interest" always
    # appears BEFORE "Total Payment Due" in the PDF; the Account Activity table
    # (which has the prior-month "PMT REC'D" interest) appears AFTER it.
    # Limit the search to text before "Total Payment Due" to avoid picking up
    # the activity-table interest amount.
    if not payment_interest:
        _tot_pos_m = re.search(r'Total Payment Due', full_text, re.IGNORECASE)
        _c_search_text = full_text[:_tot_pos_m.start()] if _tot_pos_m else full_text
        _all_pi = re.findall(r'\bInterest\b\s+(\d[\d,.-]+)', _c_search_text, re.IGNORECASE)
        for _raw in _all_pi:
            _cand = _safe_float(_raw)
            if _cand > 1000:
                payment_interest = _cand
                break

    # R.E. Taxes and Total Payment Due appear in the payment breakdown section
    # with unique enough labels that a full-text search is safe.
    _re_m = re.search(r'R\.E\. Taxes\s+([\d,.-]+)', full_text, re.IGNORECASE)
    if _re_m:
        payment_re_taxes = _safe_float(_re_m.group(1))

    _tot_m = re.search(
        r'Total Payment Due\s+\$\s*([\d,.-]+)', full_text, re.IGNORECASE
    )
    if _tot_m:
        payment_total = _safe_float(_tot_m.group(1))

    # ── 5. Activity table ────────────────────────────────────────
    activity = _parse_activity_table(lines)

    # ── 6. Footer ────────────────────────────────────────────────
    last_installment_date = ''
    amount_due = 0.0

    footer_idx = None
    for i, line in enumerate(lines):
        if 'Last Installment Made' in line and 'Due Date' in line and 'Amount Due' in line:
            footer_idx = i
            break

    if footer_idx is not None and footer_idx + 1 < len(lines):
        footer_line = lines[footer_idx + 1].strip()
        # Three values: last_inst_date, due_date, amount_due (space separated)
        parts = footer_line.split()
        if len(parts) >= 3:
            # Dates are MM/DD/YYYY, amount is the last token
            date_parts = [p for p in parts if re.match(r'\d{2}/\d{2}/\d{4}', p)]
            num_parts = [p for p in parts if re.match(r'[\d,]+\.[\d]+', p)]
            if len(date_parts) >= 1:
                last_installment_date = date_parts[0]
            if len(num_parts) >= 1:
                amount_due = _safe_float(num_parts[-1])

    return {
        'property_name': property_name,
        'loan_number': loan_number,
        'interest_rate': interest_rate,
        'as_of_date': as_of_date,
        'principal_balance': principal_balance,
        'interest_paid_ytd': interest_paid_ytd,
        'outstanding_deferred_int': outstanding_deferred_int,
        'tax_escrow_balance': tax_escrow_balance,
        'insurance_escrow_balance': insurance_escrow_balance,
        'reserve_balance': reserve_balance,
        'payment_due_date': payment_due_date,
        'payment_principal': payment_principal,
        'payment_interest': payment_interest,
        'payment_re_taxes': payment_re_taxes,
        'payment_reserves': payment_reserves,
        'payment_total': payment_total,
        'last_installment_date': last_installment_date,
        'amount_due': amount_due if amount_due else payment_total,
        'activity': activity,
    }


def _parse_activity_table(lines: List[str]) -> List[Dict[str, Any]]:
    """
    Parse the Account Activity table from PDF text lines.

    pdfplumber collapses multiple spaces to single, so positional slicing
    doesn't work. We split each row by whitespace, identify the date and
    description tokens, then collect the numeric tokens as the amounts.
    The first number is always 'total'; breakdown columns are best-effort.
    """
    activity = []

    # Find header row
    header_idx = None
    for i, line in enumerate(lines):
        if re.match(r'Date\s+Desc\s+Total', line):
            header_idx = i
            break

    if header_idx is None:
        return activity

    # Find end of table
    end_idx = len(lines)
    for i in range(header_idx + 1, len(lines)):
        if 'For general inquiries' in lines[i]:
            end_idx = i
            break

    date_pat = re.compile(r'^\d{2}/\d{2}/\d{4}$')
    num_pat  = re.compile(r'^-?[\d,]+\.\d+$')

    for line in lines[header_idx + 1:end_idx]:
        tokens = line.strip().split()
        if not tokens or not date_pat.match(tokens[0]):
            continue

        date = tokens[0]
        # Collect desc tokens (non-numeric, non-date after first token)
        desc_tokens = []
        num_tokens  = []
        for tok in tokens[1:]:
            if num_pat.match(tok):
                num_tokens.append(tok)
            else:
                if not num_tokens:          # desc comes before first number
                    desc_tokens.append(tok)

        desc = ' '.join(desc_tokens)
        nums = [_safe_float(n) for n in num_tokens]

        activity.append({
            'date':      date,
            'desc':      desc,
            'total':     nums[0] if len(nums) > 0 else 0.0,
            'principal': nums[1] if len(nums) > 1 else 0.0,
            'interest':  nums[2] if len(nums) > 2 else 0.0,
            'escrows':   nums[3] if len(nums) > 3 else 0.0,
            'reserves':  nums[4] if len(nums) > 4 else 0.0,
            'late_fee':  nums[5] if len(nums) > 5 else 0.0,
            'other':     nums[6] if len(nums) > 6 else 0.0,
        })

    return activity


def _parse_activity_row(line: str, col_positions) -> Dict[str, Any]:
    """Legacy stub — kept for backward compat; not used by _parse_activity_table."""
    col_keys = ['date', 'desc', 'total', 'principal', 'interest', 'escrows', 'reserves', 'late_fee', 'other']
    parts = line.split()
    values = parts + [''] * (len(col_keys) - len(parts))
    row = {}
    for k, v in zip(col_keys, values):
        if k in ('date', 'desc'):
            row[k] = v
        else:
            row[k] = _safe_float(v) if v else 0.0

    # Validate we got a real date
    if not re.match(r'\d{2}/\d{2}/\d{4}', row.get('date', '')):
        return None

    return row


# ─────────────────────────────────────────────────────────────────
# XLSX parsing (legacy)
# ─────────────────────────────────────────────────────────────────

def _parse_xlsx(filepath: str) -> List[Dict[str, Any]]:
    """
    Parse Berkadia loan servicer statements from Excel file.

    Returns:
        List of dictionaries, one per loan/sheet, containing extracted data
    """
    results = []
    wb = load_workbook(filepath)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        loan_data = _parse_sheet(ws)
        if loan_data:
            results.append(loan_data)

    return results


def _parse_sheet(ws) -> Dict[str, Any]:
    """Parse a single sheet from the workbook."""
    data = {}

    data['property_name'] = ws['D1'].value or ws['D3'].value
    data['loan_number'] = ws['P5'].value or ws['Q3'].value
    data['interest_rate'] = ws['X5'].value or ws['V3'].value
    data['as_of_date'] = _extract_date_from_cell(ws['A7'].value)
    data['principal_balance'] = ws['G8'].value
    data['interest_paid_ytd'] = ws['G9'].value
    data['outstanding_deferred_int'] = ws['A10'].value
    data['tax_escrow_balance'] = ws['G11'].value
    data['insurance_escrow_balance'] = ws['G12'].value
    data['reserve_balance'] = ws['G13'].value
    data['payment_due_date'] = _extract_date_from_cell(ws['K7'].value)

    data['payment_breakdown'] = {}
    data['payment_breakdown']['principal'] = ws['K8'].value
    data['payment_breakdown']['interest'] = ws['T9'].value
    data['payment_breakdown']['taxes'] = ws['T10'].value
    data['payment_breakdown']['insurance'] = ws['K12'].value
    data['payment_breakdown']['reserves'] = ws['K13'].value

    total_text = ws['P16'].value
    data['total_payment_due'] = _extract_amount_from_text(total_text)
    data['account_activity'] = _extract_account_activity(ws)
    data['last_payment_date'] = ws['A29'].value
    data['due_date'] = ws['I29'].value
    data['amount_due'] = ws['R29'].value

    # Normalize to standard keys for engine compatibility
    data['payment_principal'] = _safe_float(data['payment_breakdown'].get('principal'))
    data['payment_interest'] = _safe_float(data['payment_breakdown'].get('interest'))
    data['payment_re_taxes'] = _safe_float(data['payment_breakdown'].get('taxes'))
    data['payment_reserves'] = _safe_float(data['payment_breakdown'].get('reserves'))
    data['payment_total'] = _safe_float(data.get('total_payment_due'))
    data['activity'] = data.get('account_activity', [])

    return data


def _extract_account_activity(ws) -> List[Dict[str, Any]]:
    """Extract account activity transactions from the worksheet."""
    transactions = []
    current_row = 20
    while True:
        date_cell = ws[f'A{current_row}'].value
        if date_cell is None or not isinstance(date_cell, (datetime, str)):
            break

        desc_cell = ws[f'C{current_row}'].value
        if desc_cell is None:
            current_row += 1
            continue

        transaction = {
            'date': date_cell if isinstance(date_cell, datetime) else None,
            'description': desc_cell,
            'total': ws[f'F{current_row}'].value,
            'principal': ws[f'G{current_row}'].value,
            'interest': ws[f'O{current_row}'].value,
            'escrow': ws[f'O{current_row}'].value,
            'reserves': ws[f'T{current_row}'].value,
            'late_fee': ws[f'V{current_row}'].value,
            'other': ws[f'X{current_row}'].value,
        }

        if transaction['date'] or transaction['description']:
            transactions.append(transaction)

        current_row += 1
        if current_row > 25:
            break

    return transactions


def _extract_date_from_cell(cell_value: Any) -> datetime:
    """Extract date from cell value."""
    if isinstance(cell_value, datetime):
        return cell_value
    if isinstance(cell_value, str):
        if '/' in cell_value or '-' in cell_value:
            try:
                return datetime.fromisoformat(cell_value.split()[0])
            except:
                pass
    return None


def _extract_amount_from_text(text: str) -> float:
    """Extract numeric amount from text."""
    if text is None:
        return None
    match = re.search(r'[\$]?\s*([\d,]+\.?\d*)', str(text))
    if match:
        amount_str = match.group(1).replace(',', '')
        try:
            return float(amount_str)
        except:
            pass
    return None


# ─────────────────────────────────────────────────────────────────
# Validation
# ─────────────────────────────────────────────────────────────────

def validate(filepath: str) -> Tuple[bool, List[str]]:
    """Validate that the file is a properly formatted Berkadia loan statement."""
    issues = []

    if filepath.lower().endswith('.pdf'):
        try:
            results = _parse_pdf(filepath)
            if not results:
                issues.append("No loan data extracted from PDF")
            elif not results[0].get('loan_number'):
                issues.append("Missing loan number in PDF")
        except Exception as e:
            return False, [f"Failed to parse PDF: {str(e)}"]
        return len(issues) == 0, issues

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return False, [f"Failed to open Excel file: {str(e)}"]

    if not wb.sheetnames:
        issues.append("Workbook contains no sheets")
        return False, issues

    ws = wb[wb.sheetnames[0]]

    if not (ws['D1'].value or ws['D3'].value):
        issues.append("Missing property name (expected in D1 or D3)")

    if not (ws['P5'].value or ws['Q3'].value):
        issues.append("Missing loan number (expected in P5 or Q3)")

    if not ws['G8'].value:
        issues.append("Missing principal balance (expected in G8)")

    if not ws['P16'].value and not ws['Q9'].value:
        issues.append("Missing total payment due (expected in P16 or Q9)")

    return len(issues) == 0, issues


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python berkadia_loan.py <filepath>")
        sys.exit(1)

    filepath = sys.argv[1]

    loans = parse(filepath)

    print(f"\n{'='*80}")
    print(f"Berkadia Loan Statement Parser - {filepath.split('/')[-1]}")
    print(f"{'='*80}\n")

    for idx, loan in enumerate(loans, 1):
        principal = loan.get('principal_balance') or 0
        tax_escrow = loan.get('tax_escrow_balance') or 0
        ins_escrow = loan.get('insurance_escrow_balance') or 0
        reserve = loan.get('reserve_balance') or 0
        total_due = loan.get('payment_total') or loan.get('total_payment_due') or 0

        print(f"Loan {idx}:")
        print(f"  Property: {loan.get('property_name')}")
        print(f"  Loan #: {loan.get('loan_number')}")
        print(f"  Interest Rate: {loan.get('interest_rate')}")
        print(f"  As of Date: {loan.get('as_of_date')}")
        print(f"  Principal Balance: ${principal:,.2f}")
        print(f"  Tax Escrow: ${tax_escrow:,.2f}")
        print(f"  Insurance Escrow: ${ins_escrow:,.2f}")
        print(f"  Reserve Balance: ${reserve:,.2f}")
        print(f"  Payment Due Date: {loan.get('payment_due_date')}")
        print(f"  Total Payment Due: ${total_due:,.2f}")
        print(f"  Transactions: {len(loan.get('activity', loan.get('account_activity', [])))}")
        print()

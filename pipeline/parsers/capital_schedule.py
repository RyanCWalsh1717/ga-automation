"""
Capital Accounts Schedule Parser
=================================
Parses the 4-tab capital accounts schedule workbook:
  154500  Building Improvements
  181200  Leasing Commissions
  181300  Legal Leasing Costs
  181400  Tenant Improvement

Each sheet has a header row (row 6, 1-indexed) and data rows that follow.
Column positions vary per sheet — detected from header row text.

Returns a dict keyed by account_code with CapitalAccount dataclasses.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Any


@dataclass
class CapitalRow:
    description: str
    entity: str           # 'revlabs', 'revlabpm', or ''
    commencement_date: str  # e.g. '1/4/2023 - 11/30/2031' or '6/24'
    amount: float
    notes: str = ''


@dataclass
class CapitalAccount:
    account_code: str
    account_name: str
    rows: List[CapitalRow]
    ending_balance: float
    as_of_date: str = ''


# ── Tab definitions ────────────────────────────────────────────────────────────

_TABS = [
    ('154500', '154500 Building Improvements'),
    ('181200', '181200 Leasing Commissions'),
    ('181300', '181300 Legal Leasing Costs'),
    ('181400', '181400 Tenant Improvement'),
]

# Fallback account names in case tab isn't found by full name
_ACCT_NAMES = {
    '154500': 'Building Improvements',
    '181200': 'Leasing Commissions',
    '181300': 'Legal Leasing Costs',
    '181400': 'Tenant Improvement',
}


def _safe_float(v) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def _normalize(s) -> str:
    """Lowercase, strip, collapse whitespace."""
    return re.sub(r'\s+', ' ', str(s or '').lower().strip())


def _find_tab(wb, account_code: str):
    """Return the worksheet matching the account code, by partial name match."""
    for name in wb.sheetnames:
        if account_code in name:
            return wb[name]
    # Fallback: try matching by account name keywords
    keywords = {
        '154500': 'building',
        '181200': 'leasing comm',
        '181300': 'legal',
        '181400': 'tenant improv',
    }
    kw = keywords.get(account_code, '')
    for name in wb.sheetnames:
        if kw and kw in name.lower():
            return wb[name]
    return None


def _detect_columns(header_row):
    """
    Detect column indices for Description, Entity/Bldg, Commencement, Amount
    from a header row (list of cell values).

    Returns dict: {'desc': idx, 'entity': idx_or_None, 'comm': idx_or_None, 'amount': idx}
    """
    result = {'desc': None, 'entity': None, 'comm': None, 'amount': None}
    for i, cell in enumerate(header_row):
        n = _normalize(cell)
        if n == '' or cell is None:
            continue
        if result['desc'] is None and 'description' in n:
            result['desc'] = i
        elif result['entity'] is None and ('bldg' in n or 'building' in n or 'entity' in n):
            result['entity'] = i
        elif result['comm'] is None and 'commencement' in n:
            result['comm'] = i
        elif result['amount'] is None and 'amount' in n:
            result['amount'] = i
    return result


def _parse_sheet(ws, account_code: str) -> Optional[CapitalAccount]:
    """Parse a single sheet into a CapitalAccount."""
    # Read all rows into a list (values only)
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return None

    # Header is row 6 (1-indexed) = index 5
    HEADER_IDX = 5
    if len(all_rows) <= HEADER_IDX:
        return None

    header_row = list(all_rows[HEADER_IDX])
    col_map = _detect_columns(header_row)

    # For 154500: no entity/commencement — use fixed layout if detection fails
    # Fixed fallbacks per account code
    if account_code == '154500':
        if col_map['desc'] is None:
            col_map['desc'] = 1    # col B
        if col_map['amount'] is None:
            col_map['amount'] = 5  # col F
        # col_map['entity'] stays None
        # col_map['comm'] = col D (idx 3) for Date
        if col_map['comm'] is None:
            col_map['comm'] = 3    # col D used as Date
    elif account_code in ('181200', '181300'):
        if col_map['desc'] is None:
            col_map['desc'] = 1    # col B
        if col_map['entity'] is None:
            col_map['entity'] = 2  # col C
        if col_map['comm'] is None:
            col_map['comm'] = 3    # col D
        if col_map['amount'] is None:
            col_map['amount'] = 4  # col E
    elif account_code == '181400':
        if col_map['desc'] is None:
            col_map['desc'] = 1    # col B
        if col_map['entity'] is None:
            col_map['entity'] = 3  # col D
        if col_map['comm'] is None:
            col_map['comm'] = 4    # col E
        if col_map['amount'] is None:
            col_map['amount'] = 6  # col G

    desc_col   = col_map['desc']
    entity_col = col_map['entity']
    comm_col   = col_map['comm']
    amount_col = col_map['amount']

    if desc_col is None or amount_col is None:
        return None

    rows: List[CapitalRow] = []
    ending_balance = 0.0
    as_of_date = ''

    # Data rows start at row 11 (1-indexed) = index 10
    DATA_START_IDX = 10

    for raw in all_rows[DATA_START_IDX:]:
        row = list(raw)
        # Pad row to ensure column access is safe
        while len(row) <= max(desc_col, amount_col,
                               entity_col if entity_col is not None else 0,
                               comm_col   if comm_col   is not None else 0):
            row.append(None)

        desc_val   = row[desc_col]
        amount_val = row[amount_col]

        if desc_val is None and amount_val is None:
            continue  # blank row

        desc_str = str(desc_val or '').strip()
        if not desc_str:
            continue

        # Ending balance detection
        desc_lower = desc_str.lower()
        if 'ending balance' in desc_lower:
            ending_balance = _safe_float(amount_val)
            # Try to extract "as of MM/DD/YY" or "as of MM/YYYY"
            m = re.search(
                r'as\s+of\s+(\d{1,2}[/\-]\d{1,2}(?:[/\-]\d{2,4})?)',
                desc_str, re.IGNORECASE
            )
            if m:
                as_of_date = m.group(1)
            continue  # don't add ending balance to rows list

        # Skip rows with no amount
        if amount_val is None:
            continue

        entity_val = ''
        if entity_col is not None and entity_col < len(row):
            entity_val = str(row[entity_col] or '').strip()

        comm_val = ''
        if comm_col is not None and comm_col < len(row):
            raw_comm = row[comm_col]
            if raw_comm is not None:
                # If it's a datetime object, format it
                try:
                    from datetime import datetime, date as _date
                    if isinstance(raw_comm, (_date, datetime)):
                        comm_val = raw_comm.strftime('%-m/%-d/%Y') if hasattr(raw_comm, 'strftime') else str(raw_comm)
                    else:
                        comm_val = str(raw_comm).strip()
                except Exception:
                    comm_val = str(raw_comm).strip()

        rows.append(CapitalRow(
            description=desc_str,
            entity=entity_val,
            commencement_date=comm_val,
            amount=_safe_float(amount_val),
        ))

    return CapitalAccount(
        account_code=account_code,
        account_name=_ACCT_NAMES.get(account_code, ''),
        rows=rows,
        ending_balance=ending_balance,
        as_of_date=as_of_date,
    )


def parse(filepath: str) -> Dict[str, Any]:
    """
    Parse the capital accounts schedule workbook.

    Returns dict keyed by account_code ('154500', '181200', '181300', '181400').
    Each value is a CapitalAccount dataclass, or None if the tab is missing.
    Also includes '_parse_error': None | str.
    """
    result: Dict[str, Any] = {
        '154500': None,
        '181200': None,
        '181300': None,
        '181400': None,
        '_parse_error': None,
    }

    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        for account_code, _ in _TABS:
            ws = _find_tab(wb, account_code)
            if ws is None:
                continue
            try:
                ca = _parse_sheet(ws, account_code)
                result[account_code] = ca
            except Exception as exc:
                # Non-fatal — individual tab parse failure
                pass

    except Exception as exc:
        result['_parse_error'] = str(exc)

    return result

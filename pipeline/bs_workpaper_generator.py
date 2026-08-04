"""
Workpaper Generator — Revolution Labs Monthly Close
====================================================
Generates the monthly close workpaper with:
  - Summary tab:      all BS accounts, GL ending vs TB ending, variance status
  - Trial Balance tab: direct from Yardi TB export
  - One tab per balance sheet account: transactions + GL ending + TB tie-out
  - Prepaid Schedule tab (if ledger data available)
  - Bank Rec tabs (PNC Operating + DACA)

Historical carry-forward
------------------------
If ``prior_workpaper_path`` is supplied the generator loads the prior
month's workpaper and renames every existing sheet with the
``prior_period`` label (e.g. "Feb-2026 Summary").  The current-period
sheets are then appended with the current ``period`` label (e.g.
"Mar-2026 Summary").  Over time the file accumulates a full history:

    Feb-2026 Summary
    Feb-2026 Trial Balance
    Feb-2026 111100
    …
    Mar-2026 Summary   ← current period (most recent tabs at end)
    Mar-2026 Trial Balance
    Mar-2026 111100
    …

Structure mirrors the Hartwell workpaper pattern:
  [transactions / rollforward]
  ─────────────────────────────
  Ending Balance per GL:   $X    ← computed from GL transactions
  TB Ending Balance:       $X    ← from Yardi TB export
  Variance:                $0    ← must equal zero (flags accrual gaps if not)

The Variance will be non-zero for accounts where accrual JEs are in the TB
but not yet in the GL — surfacing exactly what still needs to be posted.
"""

import os
import re
from datetime import datetime, date
from typing import List, Dict, Optional
from openpyxl import Workbook, load_workbook as _load_workbook
from property_config import is_balance_sheet_account
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

try:
    from analysis_tab_builder import build_all_analysis_tabs as _build_analysis_tabs
except ImportError:
    _build_analysis_tabs = None

try:
    from account_tab_builders import CUSTOM_BUILDERS as _CUSTOM_BUILDERS
except ImportError:
    _CUSTOM_BUILDERS = {}

# Regex to detect already-prefixed sheet names like "Mar-2026 Summary"
_PERIOD_PREFIX_RE = re.compile(
    r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{4} '
)

# Tab names (lowercase) that are always carried forward even though they
# don't start with an account code digit.
_PRIOR_TAB_WHITELIST = {
    'xxxxxxx', 'general ', 'general',
    'mgmt fee', 'tb', 'ts',
    'rent roll rec', 'loan analysis',
    're tax analysis', 'insurance analysis',
    '135150 ppd other', 'accrued insurance',
    'bank rec - operating', 'bank rec - daca', 'bank rec - development',
    'prepaid schedule', 'summary', 'trial balance',
}

# Tab names (lowercase) that are never carried forward — JLL working/utility sheets
_PRIOR_TAB_BLOCKLIST = {
    'sheet1', 'sheet2', 'sheet3',
    'instructions', 'upload', 'input',
    'mgmt fee back up', 'rs', 'deposit register',
    'insu calc', 'accrual calc support',
    'stx', 'stx gl', 'electric bb_recon',
    'sq footage', 'sales tax rec',
}


def _should_carry_forward_tab(tab_name: str) -> bool:
    """
    Return True if a prior-workpaper tab should be renamed and kept.

    Keeps:
      • Account-code tabs — name (stripped) starts with a digit, e.g. '111100 PNC Cash'
      • Known analysis / summary tab names (whitelist)
      • Any tab whose name contains a 6-digit account code

    Drops:
      • Explicitly blocked JLL working / utility tabs
      • Any other text-named tab not in the whitelist
    """
    stripped = tab_name.strip()
    lower    = stripped.lower()

    if lower in _PRIOR_TAB_BLOCKLIST:
        return False

    # Account-code tabs (may have a leading space in JLL files)
    if stripped and stripped[0].isdigit():
        return True

    if lower in _PRIOR_TAB_WHITELIST:
        return True

    # Tab name contains a 6-digit account code anywhere (e.g. ' 2220-010')
    if re.search(r'\b\d{6}\b', stripped):
        return True

    return False


# ── Constants ────────────────────────────────────────────────

# Balance sheet account range (assets + liabilities + equity)
BS_ACCOUNT_RANGE = ('100000', '399999')

# Excel sheet names cannot contain: \ / * ? : [ ]
# Any of these characters in account names or period labels will be replaced with '-'.
_EXCEL_INVALID_CHARS = '\\/*?:[]'


def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    """Return *name* with Excel-illegal characters replaced by '-', truncated to max_len."""
    for ch in _EXCEL_INVALID_CHARS:
        name = name.replace(ch, '-')
    return name[:max_len]

# Tab colors
COLOR_SUMMARY    = '002060'   # Greatland dark navy  — summary
COLOR_TB         = '2D6F50'   # Greatland green      — trial balance
COLOR_BS_STD     = '375623'   # dark green — matches BofA Development tab style
COLOR_BS_COMPLEX = '375623'   # dark green — matches BofA Development tab style

COMPLEX_ACCOUNTS = {'213100', '135110', '135150', '213200', '221100'}

# Accounts that use a JLL-style accrual schedule instead of raw GL transaction detail.
# Each accrual line shows: Expense Acct # | Description | Vendor | FROM | TO | Amount | Notes
_ACCRUAL_SCHEDULE_ACCOUNTS = {'211200', '211300', '213100', '201000'}

# Styling helpers
# ── Greatland Brand Palette ────────────────────────────────────────────────────
# Source: Greatland Theme - New.thmx  (accent5=002060, dk2/accent2=2D6F50)
DARK_BLUE  = '002060'   # Greatland dark navy   (was 1F4E78)
MED_BLUE   = '2D6F50'   # Greatland green        (was 2E75B6)
LIGHT_BLUE = 'D6EAE1'   # light green tint       (was D6E4F0)
LIGHT_GRAY = 'F2F2F2'   # alternating row shade  (unchanged)
GREEN_FILL = 'E2EFDA'   # tie-out pass           (unchanged)
RED_FILL   = 'FFCCCC'   # tie-out fail           (unchanged)
AMBER_FILL = 'FFF2CC'
WHITE      = 'FFFFFF'

# Column layout — col A is always blank; all data starts in col B
_A  = 1   # always blank — never write here
_B  = 2   # first data column (Date / Description / first label)
_C  = 3
_D  = 4
_E  = 5
_F  = 6
_G  = 7
_H  = 8
_I  = 9   # last standard data column (Balance / Total)
_NCOLS = 9  # total columns including blank col A

THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
THICK_BOTTOM = Border(bottom=Side(style='medium'))
DOUBLE_BTM   = Border(bottom=Side(style='double'))

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def _font(bold=False, italic=False, size=11, color='000000', name='Calibri'):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

def _hdr_font():
    return Font(name='Calibri', size=11, bold=True, color='FFFFFF')

def _apply(cell, font=None, fill=None, fmt=None, border=None, align=None):
    if font:   cell.font   = font
    if fill:   cell.fill   = fill
    if fmt:    cell.number_format = fmt
    if border: cell.border = border
    if align:  cell.alignment = align


# ── Main entry point ─────────────────────────────────────────

def generate_bs_workpaper(gl_result, tb_result, output_path: str,
                           period: str = '', property_name: str = '',
                           prepaid_ledger_active: list = None,
                           bank_rec_data: dict = None,
                           gl_cash_balance: float = None,
                           daca_bank_data: dict = None,
                           daca_gl_balance: float = None,
                           je_adjustments: Optional[Dict[str, float]] = None,
                           prior_workpaper_path: str = None,
                           prior_period: str = None,
                           berkadia_loans: list = None,
                           dev_bank_rec_data: dict = None,
                           ar_aging_data=None,
                           capital_schedule_data=None,
                           tb_filepath: str = None,
                           ar_aging_filepath: str = None,
                           ap_aging_filepath: str = None,
                           bank_rec_xlsx_filepath: str = None,
                           daca_bank_rec_xlsx_filepath: str = None,
                           dev_bank_rec_xlsx_filepath: str = None,
                           prepared_by: str = '',
                           property_config=None) -> str:
    """
    Generate the monthly close workpaper (GL vs TB tie-out + bank recs).

    Args:
        gl_result:             GLParseResult from parsers.yardi_gl.parse_gl()
        tb_result:             TBResult from parsers.yardi_trial_balance.parse()
        output_path:           Where to write the .xlsx file
        period:                Period label e.g. 'Mar-2026'
        property_name:         Property display name
        prepaid_ledger_active: Active prepaid items from prepaid_ledger.py (optional)
        bank_rec_data:         Parsed Yardi Bank Rec dict
        gl_cash_balance:       GL ending balance for account 111100 (PNC Operating)
        daca_bank_data:        Parsed KeyBank DACA statement dict
        daca_gl_balance:       GL ending balance for account 115100 (DACA)
        prior_workpaper_path:  Path to the prior month's workpaper .xlsx for
                               historical carry-forward.  All existing sheets are
                               renamed with the prior_period prefix so current-period
                               sheets can be appended without name collisions.
        prior_period:          Period label of the prior workpaper, e.g. 'Feb-2026'.
                               Used to prefix the copied sheets.
        berkadia_loans:        List of loan dicts from parsers.berkadia_loan — used to
                               populate Loan Analysis, RE Tax, and Insurance Escrow tabs.

    Returns:
        output_path
    """
    # Pass 2 safety guard — GL is already final; je_adjustments must not be used.
    if je_adjustments is not None:
        raise ValueError(
            "je_adjustments must not be passed to generate_bs_workpaper() in Pass 2. "
            "The GL is already final after the close — read actuals directly from GL."
        )

    # ── Load prior workpaper (if provided) or start fresh ─────────────────────
    # Strategy:
    #   1. Extract historical per-period summary rows from prior account tabs.
    #   2. Keep only analysis tabs (Loan, RE Tax, Insurance) — renamed with the
    #      prior period prefix so analysis_tab_builder can copy-and-extend them.
    #   3. Delete all other prior tabs (account tabs, Summary, TB, Bank Rec) —
    #      they will be regenerated fresh below.
    # Account tabs are rebuilt as rolling tables (one row per period) so the
    # full balance history lives in a single tab per account.
    _account_history: dict = {}   # {account_code: [sorted period row dicts]}
    _prior_full_detail: dict = {} # {account_code: [full transaction row dicts]}

    if prior_workpaper_path and os.path.exists(prior_workpaper_path):
        try:
            _wb_prior = _load_workbook(prior_workpaper_path)

            # Auto-detect prior period label from prefixed tab names.
            if not prior_period:
                for _n in _wb_prior.sheetnames:
                    _m = _PERIOD_PREFIX_RE.match(_n)
                    if _m:
                        prior_period = _m.group(0).strip()
                        break

            # Extract historical balance data from all prior account tabs.
            _account_history = _extract_account_history(_wb_prior)

            # Extract full transaction detail for escrow + capital accounts.
            _prior_full_detail = _extract_prior_full_detail(_wb_prior)

            # Determine which analysis tab names to carry forward (copy-and-extend).
            # These are the only sheets we keep in the working wb.
            _ANALYSIS_NAMES = {
                'loan analysis', 're tax analysis', 'insurance analysis',
                '135150 ppd other', 'accrued insurance',
                'bank rec - operating', 'bank rec - daca', 'bank rec - development',
            }

            # Build wb from analysis tabs only — start fresh then copy them in.
            wb = Workbook()
            _pfx = (prior_period or 'Prior') + ' '
            for _name in _wb_prior.sheetnames:
                _stripped_lower = _name.strip().lower()
                # Already-prefixed analysis tabs carry straight through.
                _already_pfx = _PERIOD_PREFIX_RE.match(_name)
                _bare_lower  = _PERIOD_PREFIX_RE.sub('', _name).strip().lower()

                is_analysis = (
                    _bare_lower in _ANALYSIS_NAMES
                    or _stripped_lower in _ANALYSIS_NAMES
                )
                if not is_analysis:
                    continue  # skip — will be regenerated

                # Copy sheet from prior wb into our working wb
                from openpyxl import copy as _xl_copy
                try:
                    import copy as _copy
                    _src = _wb_prior[_name]
                    _dst = wb.copy_worksheet(_src) if hasattr(wb, 'copy_worksheet') else None
                    if _dst is None:
                        # openpyxl < 2.5 fallback — skip analysis copy
                        continue
                    # Rename with prior period prefix if not already prefixed
                    if _already_pfx:
                        _dst.title = _name[:31]
                    else:
                        _new_name = (_pfx + _name)[:31]
                        _ctr = 1
                        while _new_name in wb.sheetnames and wb[_new_name] is not _dst:
                            _new_name = (_pfx + _name)[:28] + f'_{_ctr}'
                            _ctr += 1
                        _dst.title = _new_name
                except Exception:
                    pass  # if copy fails, analysis tab is skipped — non-fatal

        except Exception:
            wb = Workbook()
    else:
        wb = Workbook()

    # Tab prefix for all current-period sheets — sanitized for Excel.
    _tab_pfx = (_safe_sheet_name(period) + ' ') if period else ''

    # Entity columns for workpaper tabs — read from GL metadata.
    # Multi-entity GL: metadata.entities = ['lexlab-1', 'lexlab-2', ...]
    #                  → one amount column per entity + Total column.
    # Single-entity GL: metadata.entities = [] (or single item)
    #                  → one amount column, labelled with property_code.
    _gl_entities: list = []
    _entity_label = 'Revlabs'
    if gl_result and hasattr(gl_result, 'metadata') and gl_result.metadata:
        _gl_entities = list(getattr(gl_result.metadata, 'entities', []) or [])
        _entity_label = (
            getattr(gl_result.metadata, 'property_code', '') or
            getattr(gl_result.metadata, 'property_name', '') or ''
        ).strip().lower() or ''

    # C-10: Build entity display-name map from property config when available.
    # Falls back to a hardcoded RevLabs mapping only for backward compatibility.
    _ENTITY_DISPLAY: dict = {}
    _cfg_code = (getattr(property_config, 'property_code', '') or '').lower()
    _cfg_name = (getattr(property_config, 'property_display_name', '') or
                 getattr(property_config, 'property_name', '') or '').strip()
    if _cfg_code and _cfg_name:
        _ENTITY_DISPLAY[_cfg_code] = _cfg_name
    else:
        # RevLabs legacy mapping (for sessions where property_config is unavailable)
        _ENTITY_DISPLAY.update({'revlabspm': 'Revlabs', 'revla': 'Revla'})
    _entity_label  = _ENTITY_DISPLAY.get(_entity_label, _entity_label) or property_name or '[Property]'
    _gl_entities   = [_ENTITY_DISPLAY.get(e.lower(), e) for e in _gl_entities]

    # Build TB lookup: account_code -> TBAccount
    tb_map = {}
    if tb_result and hasattr(tb_result, 'accounts'):
        tb_map = {a.account_code: a for a in tb_result.accounts}

    # Identify balance sheet accounts from GL
    bs_accounts = [
        a for a in (gl_result.accounts if gl_result else [])
        if BS_ACCOUNT_RANGE[0] <= a.account_code <= BS_ACCOUNT_RANGE[1]
    ]

    # Pre-compute: journal control → (expense_code, expense_name) for accrual schedules.
    # For each accrual JE, the credit side (211200/211300/213100) and the debit side
    # (a P&L expense account) share the same journal control number.  We scan all
    # expense-range GL accounts to build this lookup so the accrual schedule tab can
    # show which expense account each accrual line offsets.
    _control_to_expense: dict = {}
    if gl_result:
        for _ea in (gl_result.accounts or []):
            _ec = _ea.account_code
            # P&L accounts are 4xxxxx (revenue) through 8xxxxx (expense)
            if _ec and '4' <= _ec[0] <= '8':
                for _et in (_ea.transactions or []):
                    _ctrl = str(getattr(_et, 'control', '') or '').strip()
                    if _ctrl and _ctrl not in _control_to_expense:
                        _control_to_expense[_ctrl] = (_ec, _ea.account_name)

    # ── Identify TB accounts with no current-period GL activity ──────────────
    # These appear in the Trial Balance (balance carried from prior period) but
    # have zero transactions in this period's GL export.  They still need a tab
    # so the workpaper shows the balance that makes up the G/L.
    _gl_bs_codes = {a.account_code for a in bs_accounts}
    _zero_activity_tb = []
    if tb_result and hasattr(tb_result, 'accounts'):
        for _tba in sorted(tb_result.accounts, key=lambda a: a.account_code):
            if (BS_ACCOUNT_RANGE[0] <= _tba.account_code <= BS_ACCOUNT_RANGE[1]
                    and _tba.account_code not in _gl_bs_codes
                    and abs(_tba.ending_balance) > 0.01):
                _zero_activity_tb.append(_tba)

    # ── Raw Yardi report map — accounts whose tabs use a raw file copy ───────────
    # When a filepath is provided for an account, the tab shows the raw Yardi export
    # instead of the generated GL transaction register.  The same file may be used
    # for multiple accounts (e.g. AR Aging covers both 133100 and 221100).
    _raw_report_map: dict = {}   # {account_code: filepath}
    if ar_aging_filepath and os.path.exists(ar_aging_filepath):
        _raw_report_map['133100'] = ar_aging_filepath   # AR Control
        _raw_report_map['221100'] = ar_aging_filepath   # Prepaid Rent (Pre-payments col)
    if ap_aging_filepath and os.path.exists(ap_aging_filepath):
        _raw_report_map['211100'] = ap_aging_filepath   # AP Control
    if bank_rec_xlsx_filepath and os.path.exists(bank_rec_xlsx_filepath):
        _raw_report_map['111100'] = bank_rec_xlsx_filepath   # PNC Operating
    if daca_bank_rec_xlsx_filepath and os.path.exists(daca_bank_rec_xlsx_filepath):
        _raw_report_map['115100'] = daca_bank_rec_xlsx_filepath  # KeyBank DACA
    if dev_bank_rec_xlsx_filepath and os.path.exists(dev_bank_rec_xlsx_filepath):
        _raw_report_map['111210'] = dev_bank_rec_xlsx_filepath   # BofA Development

    # ── Build workpaper tabs ──────────────────────────────────
    # Trial Balance moved to LAST tab — generated below just before wb.save().
    # Summary Page is generated FIRST — it holds the period-end date in C4,
    # which is referenced by DATEDIF formulas in the 135150 PPD Other tab.
    _write_summary_page(wb, period)

    # Flat list of all GL transactions as dicts — consumed by custom tab builders
    # that need cross-account JE context (e.g. 133110 billback, 213100 accruals).
    _all_je_lines: list = []
    if gl_result:
        for _ea in (gl_result.accounts or []):
            for _et in (_ea.transactions or []):
                _all_je_lines.append({
                    'je_number':    str(getattr(_et, 'control',     '') or ''),
                    'account_code': _ea.account_code,
                    'account_name': _ea.account_name,
                    'description':  str(getattr(_et, 'description', '') or ''),
                    'vendor':       str(getattr(_et, 'remarks',     '') or ''),
                    'debit':        float(getattr(_et, 'debit',  0) or 0),
                    'credit':       float(getattr(_et, 'credit', 0) or 0),
                    'source':       '',
                })

    # Accounts covered by dedicated Analysis tabs — suppress individual account tabs
    # so there's no duplication.  The Analysis tabs (Insurance Analysis, RE Tax Analysis)
    # carry the full history and tie-out; a separate generated GL register would be
    # confusing and redundant.
    _ANALYSIS_COVERED = {'135110', '135120'}

    # Current-period-only accounts — history rows are NOT carried forward.
    # These tabs regenerate fresh each month; reviewers only need this period's data.
    #
    # Roll-forward accounts (everything NOT in this set) accumulate history:
    #   115200/115300/115600 escrow accounts, 133100 AR Control, 154xxx capital,
    #   211200 security deposits, 231100 mortgage payable, 3xxxxx equity.
    _CURRENT_PERIOD_ONLY = {
        '111100',   # PNC Operating        — bank rec is the reference
        '115100',   # DACA                 — bank rec is the reference
        '133110',   # AR Billback          — current period activity only
        '135150',   # Prepaids Other       — covered by Prepaid Schedule tab
        '211100',   # AP Control           — AP Aging is the reference
        '213100',   # Accrued Expenses     — only this month's accruals matter
        '213200',   # Accrued Interest     — covered by Loan Analysis tab
        '221100',   # Prepaid Rent         — current period only
    }

    for acct in bs_accounts:
        if acct.account_code in _ANALYSIS_COVERED:
            continue

        # Suppress history for current-period-only accounts
        _hist = (
            []
            if acct.account_code in _CURRENT_PERIOD_ONLY
            else _account_history.get(acct.account_code, [])
        )
        _tab_acct_name = _safe_sheet_name(f'{acct.account_code} {acct.account_name}')

        # ── Raw Yardi report — copy file directly, skip all generated builders ──
        if acct.account_code in _raw_report_map:
            _copy_raw_tb_sheet(_raw_report_map[acct.account_code], wb,
                               tab_name=_tab_acct_name)
            continue

        # ── Custom builder (account-specific layout) ──────────
        _builder = _CUSTOM_BUILDERS.get(acct.account_code)
        if _builder:
            _builder(
                wb,
                period=period,
                property_name=property_name,
                gl_acct=acct,
                tb_entry=tb_map.get(acct.account_code),
                je_lines=_all_je_lines,
                prepaid_ledger=prepaid_ledger_active,
                daca_data=daca_bank_data,
                bank_rec_data=bank_rec_data,
                ar_aging_data=ar_aging_data,
                capital_schedule_data=capital_schedule_data,
                berkadia_loans=berkadia_loans,
                prior_tab_detail=_prior_full_detail,
                property_config=property_config,
            )
        elif acct.account_code in _ACCRUAL_SCHEDULE_ACCOUNTS:
            _write_accrual_schedule_tab(
                wb, acct, tb_map.get(acct.account_code), period, property_name,
                _control_to_expense,
                tab_prefix='',        # no period prefix — rolling table
                history_rows=_hist,
                prepared_by=prepared_by)
        else:
            _write_account_tab(wb, acct, tb_map.get(acct.account_code), period,
                               property_name, je_adjustments,
                               tab_prefix='',   # no period prefix
                               history_rows=_hist,
                               entity_label=_entity_label,
                               entities=_gl_entities,
                               prepared_by=prepared_by)

    # ── Stub tabs for TB accounts with no current-period GL activity ──────────
    for _tba in _zero_activity_tb:
        if _tba.account_code in _ANALYSIS_COVERED:
            continue   # covered by Insurance Analysis / RE Tax Analysis tabs

        _hist = (
            []
            if _tba.account_code in _CURRENT_PERIOD_ONLY
            else _account_history.get(_tba.account_code, [])
        )
        _tab_tba_name = _safe_sheet_name(f'{_tba.account_code} {_tba.account_name}')
        if _tba.account_code in _raw_report_map:
            _copy_raw_tb_sheet(_raw_report_map[_tba.account_code], wb,
                               tab_name=_tab_tba_name)
        else:
            _write_stub_tab(wb, _tba, period, property_name,
                            tab_prefix='', history_rows=_hist,
                            prepared_by=prepared_by)

    # ── Prepaid amortization schedule tab (if ledger data available) ──
    if prepaid_ledger_active:
        _write_prepaid_schedule_tab(wb, prepaid_ledger_active, period,
                                    property_name, tab_prefix=_tab_pfx,
                                    gl_result=gl_result)

    # ── Bank Rec tab (PNC Operating — account 111100) ──────────────────────────
    # Suppressed when an Excel bank rec is provided — raw file is already in the
    # 111100 account tab and a duplicate generated tab would be redundant.
    if bank_rec_data and '111100' not in _raw_report_map:
        # If gl_cash_balance not passed in, try to pull it from the GL accounts
        _gl_cash = gl_cash_balance
        if _gl_cash is None and gl_result:
            for _acct in (gl_result.accounts or []):
                if _acct.account_code == '111100':
                    _gl_cash = _acct.ending_balance
                    break
        _gl_cash = _gl_cash or 0.0
        _write_bank_rec_tab(
            wb, bank_rec_data, _gl_cash, period, property_name,
            account_label='PNC Operating (x3993)',
            gl_account_code='111100',
            tab_prefix=_tab_pfx,
            prepared_by=prepared_by,
        )

    # ── DACA Bank Rec tab (KeyBank x5132 — account 115100) ────────────────────
    # Suppressed when an Excel bank rec is provided — raw file is in the 115100 tab.
    if daca_bank_data is not None and '115100' not in _raw_report_map:
        _gl_daca = daca_gl_balance
        if _gl_daca is None and gl_result:
            for _acct in (gl_result.accounts or []):
                if _acct.account_code == '115100':
                    _gl_daca = _acct.ending_balance
                    break
        _gl_daca = _gl_daca or 0.0
        _write_daca_bank_rec_tab(
            wb, daca_bank_data, _gl_daca, period, property_name,
            tab_prefix=_tab_pfx,
            prepared_by=prepared_by,
        )

    # ── Development Bank Rec tab (revlabs entity — BofA x3132) ───────────────
    if dev_bank_rec_data is not None:
        # GL balance defaults to 0.0 — revlabs has no activity in the revlabspm
        # GL export; the tab shows the BofA statement balance for reference.
        _gl_dev = float(dev_bank_rec_data.get('gl_balance') or 0)
        _write_bank_rec_tab(
            wb, dev_bank_rec_data, _gl_dev, period, 'Rev Labs (revlabs)',
            account_label='Development Account (revlabs)',
            gl_account_code='',
            tab_prefix=_tab_pfx,
            tab_name_override='Bank Rec - Development',
            prepared_by=prepared_by,
        )

    # ── Analysis tabs (Loan, RE Tax, Insurance, Escrow) ──────────────────────
    # Copy-and-extend: copies the prior period's renamed tab, inserts new rows
    # for current-period data, and rebuilds the GL/TB tie-out from live data.
    if _build_analysis_tabs is not None:
        try:
            _build_analysis_tabs(
                wb,
                period=period,
                current_prefix=_tab_pfx,
                tab_prefix=_tab_pfx,
                gl_result=gl_result,
                tb_map=tb_map,
                berkadia_loans=berkadia_loans or [],
                prepaid_active=prepaid_ledger_active or [],
                property_config=property_config,
                property_name=property_name or '',
            )
        except Exception as _atb_exc:
            import traceback
            print(f"[bs_workpaper_generator] Analysis tab build warning: {_atb_exc}")
            traceback.print_exc()

    # ── Trial Balance tab — always LAST ───────────────────────────────────────
    # Raw Yardi export when filepath available; generated fallback otherwise.
    _tb_tab_name = (_tab_pfx + 'Trial Balance')[:31]
    if tb_filepath and os.path.exists(tb_filepath):
        if not _copy_raw_tb_sheet(tb_filepath, wb, tab_name=_tb_tab_name):
            _write_tb_tab(wb, tb_result, period, property_name, tab_prefix=_tab_pfx,
                          prepared_by=prepared_by)
    else:
        _write_tb_tab(wb, tb_result, period, property_name, tab_prefix=_tab_pfx,
                      prepared_by=prepared_by)

    # Remove the blank default sheet openpyxl creates for new workbooks
    for _default in ('Sheet', 'Sheet1'):
        if _default in wb.sheetnames:
            del wb[_default]

    wb.save(output_path)
    return output_path


# ── Summary Page ─────────────────────────────────────────────

def _write_summary_page(wb, period: str) -> None:
    """
    Create a minimal 'Summary Page' tab with the period-end date in cell C4.

    This cell is referenced by DATEDIF formulas in the '135150 PPD Other' tab:
      =J{r}*DATEDIF(G{r},'Summary Page'!$C$4+1,"M")
    which counts months amortized through the period-end date.

    The tab is intentionally sparse — it exists only to provide the anchor date.
    Additional summary content can be added here manually in future.
    """
    import re
    from calendar import monthrange
    from datetime import date as _date

    ws = wb.create_sheet('Summary Page', 0)   # first tab
    ws.sheet_properties.tabColor = '002060'
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16

    # Parse period-end date from "Jan-2026" → date(2026, 1, 31)
    _MONTH_MAP = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
                  'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
    period_end = None
    m = re.search(r'([A-Za-z]{3})[\s\-](\d{4})', period or '')
    if m:
        mon = _MONTH_MAP.get(m.group(1).lower(), 0)
        yr  = int(m.group(2))
        if mon:
            last_day = monthrange(yr, mon)[1]
            period_end = _date(yr, mon, last_day)

    # Header
    hdr = ws.cell(row=1, column=2, value='Summary Page')
    hdr.font  = _font(bold=True, size=13, color='FFFFFF')
    hdr.fill  = _fill(DARK_BLUE)
    hdr.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
    ws.row_dimensions[1].height = 22

    sub = ws.cell(row=2, column=2,
                  value=f'Period: {period}  |  Generated: {datetime.now().strftime("%m/%d/%Y")}')
    sub.font  = _font(italic=True, size=10, color='FFFFFF')
    sub.fill  = _fill(DARK_BLUE)
    sub.alignment = Alignment(horizontal='left')
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=4)

    # Row 4 — period-end date anchor (C4 — referenced by 135150 DATEDIF formulas)
    lbl = ws.cell(row=4, column=2, value='Period End Date')
    lbl.font  = _font(bold=True)
    lbl.fill  = _fill(LIGHT_BLUE)
    lbl.border = THIN
    lbl.alignment = Alignment(horizontal='left')

    val = ws.cell(row=4, column=3, value=period_end)
    val.font          = _font(bold=True)
    val.fill          = _fill(LIGHT_BLUE)
    val.border        = THIN
    val.number_format = 'MM/DD/YYYY'
    val.alignment     = Alignment(horizontal='center')

    note = ws.cell(row=5, column=3,
                   value='← Used by 135150 PPD Other DATEDIF formulas')
    note.font      = _font(italic=True, size=9, color='666666')
    note.alignment = Alignment(horizontal='left')


# ── Raw Yardi TB sheet copy ───────────────────────────────────

def _copy_raw_tb_sheet(source_path: str, dest_wb, tab_name: str = 'Trial Balance') -> bool:
    """
    Copy the active sheet from the Yardi TB .xlsx file directly into dest_wb.

    Preserves cell values, number formats, fonts, fills, borders, alignments,
    merged cell ranges, column widths and row heights exactly as exported from Yardi.
    Returns True on success, False if the file could not be read.
    """
    import copy as _copy_mod
    try:
        _src_wb = _load_workbook(source_path, data_only=True)
        _src_ws = _src_wb.active

        _dst_ws = dest_wb.create_sheet(tab_name[:31])
        _dst_ws.sheet_properties.tabColor = COLOR_TB

        # Column widths
        for _col_ltr, _col_dim in _src_ws.column_dimensions.items():
            _dst_ws.column_dimensions[_col_ltr].width = _col_dim.width or 8

        # Row heights
        for _row_num, _row_dim in _src_ws.row_dimensions.items():
            if _row_dim.height:
                _dst_ws.row_dimensions[_row_num].height = _row_dim.height

        # Merged cells — must be registered before writing cells to avoid conflicts
        for _mr in list(_src_ws.merged_cells.ranges):
            _dst_ws.merge_cells(str(_mr))

        # All cell values and styles
        for _row in _src_ws.iter_rows():
            for _sc in _row:
                _dc = _dst_ws.cell(row=_sc.row, column=_sc.column, value=_sc.value)
                if _sc.has_style:
                    _dc.font          = _copy_mod.copy(_sc.font)
                    _dc.fill          = _copy_mod.copy(_sc.fill)
                    _dc.border        = _copy_mod.copy(_sc.border)
                    _dc.alignment     = _copy_mod.copy(_sc.alignment)
                    _dc.number_format = _sc.number_format

        return True
    except Exception as _e:
        print(f'[bs_workpaper_generator] Raw TB copy failed: {_e}')
        return False


def _write_no_data_placeholder_tab(wb, tab_name: str, missing_label: str, account_code: str = ''):
    """
    Shown when no raw file is available to regenerate a raw-report tab this
    period. Makes the gap explicit and actionable instead of silently
    carrying forward stale content from a prior period.
    """
    ws = wb.create_sheet(tab_name[:31])
    ws.sheet_properties.tabColor = 'FFC000'
    ws.column_dimensions['B'].width = 95

    ws.cell(2, 2, f'{account_code}  —  No data uploaded this period').font = _font(bold=True, size=12)
    ws.cell(4, 2,
            f'No {missing_label} was uploaded for this close period, '
            f'so this tab could not be regenerated.')
    ws.cell(6, 2,
            f'Upload the {missing_label} in the "Workpaper raw report overrides" '
            f'section of Pass 2 and re-run to populate this tab.')
    return ws


def _write_gl_transactions_tab(wb, tab_name: str, account_code: str,
                               gl_transactions: list, period: str,
                               property_name: str) -> object:
    """
    Fallback for a raw-report tab (111100 PNC Cash / 115100 DACA) when no
    Excel bank rec export was uploaded but a Yardi Bank Rec PDF was — its
    GL-detail pages (parsers.yardi_bank_rec.parse()'s 'gl_transactions')
    carry the same per-transaction data an Excel export would, just from a
    different source file. Not a byte-for-byte copy of Yardi's raw sheet
    (that's only possible from the actual Excel export), but real
    transaction-level detail instead of the "no data uploaded" placeholder.
    """
    ws = wb.create_sheet(tab_name[:31])
    ws.sheet_properties.tabColor = 'FFC000'  # same amber as the placeholder —
                                              # this is a fallback source, not the primary one
    ws.column_dimensions['A'].width = 2

    c = ws.cell(1, 2, f'{account_code} — from Yardi Bank Rec PDF (no Excel export uploaded)')
    c.font = _font(bold=True, size=12, color='FFFFFF')
    c.fill = _fill('BF8F00')
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)

    c = ws.cell(2, 2, f'{property_name or "Revolution Labs"}  |  Period: {period}')
    c.font = _font(italic=True, size=10, color='FFFFFF')
    c.fill = _fill('BF8F00')
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=6)

    headers = ['Date', 'Description', 'Vendor', 'Debit', 'Credit']
    widths  = [14, 45, 24, 16, 16]
    for ci, (h, w) in enumerate(zip(headers, widths)):
        col = 2 + ci
        hc = ws.cell(4, col, h)
        _apply(hc, font=_font(bold=True, size=10, color='FFFFFF'),
               fill=_fill('000000'), border=THIN,
               align=Alignment(horizontal='center'))
        ws.column_dimensions[get_column_letter(col)].width = w

    row = 5
    total_debit = total_credit = 0.0
    for txn in (gl_transactions or []):
        d = txn.get('date')
        debit  = float(txn.get('debit', 0) or 0)
        credit = float(txn.get('credit', 0) or 0)
        total_debit  += debit
        total_credit += credit
        ws.cell(row, 2, d.strftime('%m/%d/%Y') if hasattr(d, 'strftime') else (d or ''))
        ws.cell(row, 3, str(txn.get('description', '') or ''))
        ws.cell(row, 4, str(txn.get('vendor', '') or ''))
        c5 = ws.cell(row, 5, debit if debit else None)
        c5.number_format = '$#,##0.00'
        c6 = ws.cell(row, 6, credit if credit else None)
        c6.number_format = '$#,##0.00'
        row += 1

    if not gl_transactions:
        ws.cell(row, 2, 'No GL transaction detail found in the Bank Rec PDF for this period.')
        row += 1

    row += 1
    ws.cell(row, 2, 'Total').font = _font(bold=True)
    tc5 = ws.cell(row, 5, total_debit)
    tc5.number_format = '$#,##0.00'; tc5.font = _font(bold=True)
    tc6 = ws.cell(row, 6, total_credit)
    tc6.number_format = '$#,##0.00'; tc6.font = _font(bold=True)

    return ws


def _daca_fallback_txns(daca_bank_data: dict) -> list:
    """
    parsers.yardi_daca_rec.parse() doesn't extract a 'gl_transactions'
    section the way parsers.yardi_bank_rec.parse() does for the Operating
    account — no sample DACA Bank Rec PDF exists yet to build that
    extraction against. Reshape what it DOES already extract
    (cleared_deposits / cleared_other_items) into the same
    date/description/vendor/debit/credit shape _write_gl_transactions_tab
    expects, so '115100 DACA' gets real transaction-level data instead of
    a placeholder whenever a DACA Bank Rec PDF (but no Excel export) is
    uploaded.
    """
    if not daca_bank_data:
        return []
    gl_txns = daca_bank_data.get('gl_transactions')
    if gl_txns:
        return gl_txns

    combined = []
    for d in (daca_bank_data.get('cleared_deposits') or []):
        combined.append({
            'date': d.get('date'), 'description': d.get('notes', '') or 'Cleared Deposit',
            'vendor': '', 'debit': float(d.get('amount', 0) or 0), 'credit': 0.0,
        })
    for d in (daca_bank_data.get('cleared_other_items') or []):
        amt = float(d.get('amount', 0) or 0)
        combined.append({
            'date': d.get('date'), 'description': d.get('notes', '') or 'Cleared Item',
            'vendor': '', 'debit': max(amt, 0.0), 'credit': max(-amt, 0.0),
        })
    return combined


# ── Summary tab ───────────────────────────────────────────────

def _write_summary_tab(wb, bs_accounts, tb_map, period, property_name,
                       je_adjustments=None, tab_prefix: str = '',
                       zero_activity_tb_accounts: list = None,
                       prepared_by: str = ''):
    _tab_name = (tab_prefix + 'Summary')[:31]
    ws = wb.create_sheet(_tab_name)
    ws.sheet_properties.tabColor = COLOR_SUMMARY

    # Blank col A — narrow
    ws.column_dimensions['A'].width = 2

    row = 1
    # Title block
    c = ws.cell(row=row, column=_B, value=f'{property_name or "Revolution Labs"} — Workpaper')
    c.font = _font(bold=True, size=14, color='FFFFFF')
    c.fill = _fill(DARK_BLUE)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
    row += 1

    c = ws.cell(row=row, column=_B,
                value=f'Period: {period}  |  '
                      f'Prepared by: {prepared_by or "GRP"}  |  '
                      f'{datetime.now().strftime("%m/%d/%Y")}')
    c.font = _font(italic=True, size=11, color='FFFFFF')
    c.fill = _fill(MED_BLUE)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
    row += 2

    # Column headers — show projected label when JE adjustments are applied
    gl_col_label = 'GL Projected Balance' if je_adjustments else 'GL Ending Balance'
    headers = ['Account', 'Account Name', gl_col_label, 'TB Ending Balance',
               'Variance', 'Status']
    widths  = [12, 40, 22, 20, 16, 10]
    for ci, (h, w) in enumerate(zip(headers, widths)):
        col = _B + ci
        c = ws.cell(row=row, column=col, value=h)
        _apply(c, font=_hdr_font(), fill=_fill('000000'), border=THIN,
               align=Alignment(horizontal='center', vertical='center', wrap_text=True))
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 28
    row += 1

    # Asset / Liability / Equity groupings
    groups = [
        ('ASSETS',      lambda c: '100000' <= c <= '199999'),
        ('LIABILITIES', lambda c: '200000' <= c <= '299999'),
        ('EQUITY',      lambda c: '300000' <= c <= '399999'),
    ]

    all_pass = True
    total_gl_end = 0.0
    total_tb_end = 0.0

    # Zero-activity TB accounts keyed by code for quick lookup within groups
    _zero_map = {}
    for _z in (zero_activity_tb_accounts or []):
        _zero_map[_z.account_code] = _z

    for group_name, group_test in groups:
        group_accts = [a for a in bs_accounts if group_test(a.account_code)]
        # Zero-activity TB accounts that fall in this group (not already in bs_accounts)
        group_zero = [a for a in (zero_activity_tb_accounts or []) if group_test(a.account_code)]

        if not group_accts and not group_zero:
            continue

        # Group header
        c = ws.cell(row=row, column=_B, value=group_name)
        c.font = _font(bold=True, size=11, color=DARK_BLUE)
        c.fill = _fill(LIGHT_BLUE)
        ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
        row += 1

        for acct in group_accts:
            tb_acct = tb_map.get(acct.account_code)
            gl_end  = acct.ending_balance + (je_adjustments or {}).get(acct.account_code, 0.0)
            tb_end  = tb_acct.ending_balance if tb_acct else None
            variance = (gl_end - tb_end) if tb_end is not None else None
            status   = '✓' if (variance is not None and abs(variance) < 0.02) else ('⚠' if tb_end is None else '✗')
            if status != '✓':
                all_pass = False

            alt = (row % 2 == 0)
            row_fill = _fill(LIGHT_GRAY) if alt else None

            ws.cell(row=row, column=_B, value=acct.account_code).border = THIN
            ws.cell(row=row, column=_C, value=acct.account_name).border = THIN
            if row_fill:
                ws.cell(row=row, column=_B).fill = row_fill
                ws.cell(row=row, column=_C).fill = row_fill

            c_gl = ws.cell(row=row, column=_D, value=gl_end)
            _apply(c_gl, fmt='#,##0.00;(#,##0.00);"-"', border=THIN)
            if row_fill: c_gl.fill = row_fill

            if tb_end is not None:
                c_tb = ws.cell(row=row, column=_E, value=tb_end)
                _apply(c_tb, fmt='#,##0.00;(#,##0.00);"-"', border=THIN)
                if row_fill: c_tb.fill = row_fill
            else:
                c_na = ws.cell(row=row, column=_E, value='N/A in TB')
                c_na.font = _font(italic=True, color='888888')
                c_na.border = THIN

            if variance is not None:
                var_fill = _fill(GREEN_FILL) if abs(variance) < 0.02 else _fill(RED_FILL)
                c_var = ws.cell(row=row, column=_F, value=variance)
                _apply(c_var, fmt='#,##0.00;(#,##0.00);"-"', border=THIN, fill=var_fill)
                c_var.font = _font(bold=(abs(variance) >= 0.02))
            else:
                ws.cell(row=row, column=_F, value='').border = THIN

            stat_fill = _fill(GREEN_FILL) if status == '✓' else _fill(RED_FILL)
            c_stat = ws.cell(row=row, column=_G, value=status)
            _apply(c_stat, fill=stat_fill, border=THIN,
                   align=Alignment(horizontal='center'))
            c_stat.font = _font(bold=True, color='006100' if status == '✓' else '9C0006')

            total_gl_end += gl_end
            if tb_end is not None:
                total_tb_end += tb_end
            row += 1

        # ── Zero-activity TB accounts in this group ─────────────────────────────
        # GL ending balance = TB ending balance (no current-period activity).
        # Variance is always $0; status is ✓ with a lighter italic style to indicate
        # "no activity" rather than active reconciliation.
        for tb_acct in sorted(group_zero, key=lambda a: a.account_code):
            tb_end  = tb_acct.ending_balance
            gl_end  = tb_end   # no GL activity — balance unchanged from prior period
            variance = 0.0

            alt = (row % 2 == 0)
            row_fill = _fill(LIGHT_GRAY) if alt else None

            for _col, _val in [(_B, tb_acct.account_code), (_C, tb_acct.account_name)]:
                c = ws.cell(row=row, column=_col, value=_val)
                c.font = _font(italic=True, color='595959')
                c.border = THIN
                if row_fill:
                    c.fill = row_fill

            c_gl = ws.cell(row=row, column=_D, value=gl_end)
            _apply(c_gl, fmt='#,##0.00;(#,##0.00);"-"', border=THIN)
            c_gl.font = _font(italic=True, color='595959')
            if row_fill: c_gl.fill = row_fill

            c_tb = ws.cell(row=row, column=_E, value=tb_end)
            _apply(c_tb, fmt='#,##0.00;(#,##0.00);"-"', border=THIN)
            c_tb.font = _font(italic=True, color='595959')
            if row_fill: c_tb.fill = row_fill

            c_var = ws.cell(row=row, column=_F, value=variance)
            _apply(c_var, fmt='#,##0.00;(#,##0.00);"-"', border=THIN, fill=_fill(GREEN_FILL))
            c_var.font = _font(italic=True, color='006100')

            c_stat = ws.cell(row=row, column=_G, value='✓')
            _apply(c_stat, fill=_fill(GREEN_FILL), border=THIN,
                   align=Alignment(horizontal='center'))
            c_stat.font = _font(italic=True, color='006100')

            total_gl_end += gl_end
            total_tb_end += tb_end
            row += 1

        row += 1  # spacer between groups

    # Overall status banner
    status_text  = 'ALL ACCOUNTS TIE — WORKPAPER COMPLETE' if all_pass else 'VARIANCES FOUND — REVIEW REQUIRED'
    status_color = '006100' if all_pass else '9C0006'
    banner_fill  = GREEN_FILL if all_pass else RED_FILL
    c = ws.cell(row=row, column=_B, value=status_text)
    c.font = _font(bold=True, size=12, color=status_color)
    c.fill = _fill(banner_fill)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
    c.alignment = Alignment(horizontal='center')
    row += 2

    # Note about variances
    if je_adjustments:
        note = ('Note: GL Projected Balance = GL ending balance + pipeline JE adjustments (accruals, '
                'management fee, prepaid amortization). Non-zero variances vs TB indicate JEs not yet '
                'posted to Yardi — expected at pre-close. Post all JEs and re-run for final tie-out.')
    else:
        note = ('Note: Non-zero variances indicate accrual journal entries posted in Yardi (visible in TB) '
                'but not yet reflected in the GL detail file. These are expected for period-end accruals.')
    c = ws.cell(row=row, column=_B, value=note)
    c.font = _font(italic=True, size=10, color='595959')
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
    ws.row_dimensions[row].height = 30

    ws.freeze_panes = 'B4'


# ── Trial Balance tab ─────────────────────────────────────────

def _write_tb_tab(wb, tb_result, period, property_name, tab_prefix: str = '',
                  prepared_by: str = ''):
    _tab_name = (tab_prefix + 'Trial Balance')[:31]
    ws = wb.create_sheet(_tab_name)
    ws.sheet_properties.tabColor = COLOR_TB

    # Blank col A — narrow
    ws.column_dimensions['A'].width = 2

    row = 1
    c = ws.cell(row=row, column=_B, value=f'{property_name or "Revolution Labs"} — Trial Balance')
    c.font = _font(bold=True, size=13, color='FFFFFF')
    c.fill = _fill(MED_BLUE)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
    row += 1

    meta_text = period
    if tb_result and tb_result.metadata:
        meta_text = f'Period: {tb_result.metadata.period}  |  Book: {tb_result.metadata.book}'
    c = ws.cell(row=row, column=_B, value=meta_text)
    c.font = _font(italic=True, color='FFFFFF')
    c.fill = _fill(MED_BLUE)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
    row += 2

    # Column headers
    headers = ['Account', 'Account Name', 'Forward Balance', 'Debit', 'Credit', 'Ending Balance']
    widths  = [12, 42, 18, 18, 18, 18]
    for ci, (h, w) in enumerate(zip(headers, widths)):
        col = _B + ci
        c = ws.cell(row=row, column=col, value=h)
        _apply(c, font=_hdr_font(), fill=_fill(DARK_BLUE), border=THIN,
               align=Alignment(horizontal='center', wrap_text=True))
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 28
    row += 1

    if not tb_result:
        ws.cell(row=row, column=_B, value='No TB data available')
        return

    # Section groupings
    sections = [
        ('ASSETS',             '100000', '199999'),
        ('LIABILITIES',        '200000', '299999'),
        ('EQUITY',             '300000', '399999'),
        ('REVENUE',            '400000', '499999'),
        ('OPERATING EXPENSES', '500000', '799999'),
        ('DEBT SERVICE',       '800000', '999999'),
    ]

    section_totals = {}
    for section_name, lo, hi in sections:
        accts = [a for a in tb_result.accounts if lo <= a.account_code <= hi]
        if not accts:
            continue

        # Section header
        c = ws.cell(row=row, column=_B, value=section_name)
        c.font = _font(bold=True, color=DARK_BLUE)
        c.fill = _fill(LIGHT_BLUE)
        ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
        row += 1

        sec_fwd = sec_dr = sec_cr = sec_end = 0.0
        for i, acct in enumerate(accts):
            alt_fill = _fill(LIGHT_GRAY) if i % 2 == 1 else None
            ws.cell(row=row, column=_B, value=acct.account_code).border = THIN
            ws.cell(row=row, column=_C, value=acct.account_name).border = THIN
            if alt_fill:
                ws.cell(row=row, column=_B).fill = alt_fill
                ws.cell(row=row, column=_C).fill = alt_fill

            for ci, val in enumerate([acct.forward_balance, acct.debit,
                                       acct.credit, acct.ending_balance]):
                c = ws.cell(row=row, column=_D + ci, value=val)
                _apply(c, fmt='#,##0.00;(#,##0.00);"-"', border=THIN)
                if alt_fill:
                    c.fill = alt_fill

            sec_fwd += acct.forward_balance
            sec_dr  += acct.debit
            sec_cr  += acct.credit
            sec_end += acct.ending_balance
            row += 1

        # Section subtotal
        ws.cell(row=row, column=_C, value=f'{section_name} TOTAL').font = _font(bold=True, color=DARK_BLUE)
        ws.cell(row=row, column=_C).border = THIN
        ws.cell(row=row, column=_B).border = THIN
        for ci, val in enumerate([sec_fwd, sec_dr, sec_cr, sec_end]):
            c = ws.cell(row=row, column=_D + ci, value=val)
            _apply(c, font=_font(bold=True), fmt='#,##0.00;(#,##0.00);"-"',
                   border=THIN, fill=_fill(LIGHT_BLUE))
        row += 2

    # Grand total
    all_accts = tb_result.accounts
    ws.cell(row=row, column=_C, value='GRAND TOTAL').font = _font(bold=True, size=12)
    ws.cell(row=row, column=_C).border = DOUBLE_BTM
    ws.cell(row=row, column=_B).border = DOUBLE_BTM
    for ci, val in enumerate([
        sum(a.forward_balance for a in all_accts),
        sum(a.debit for a in all_accts),
        sum(a.credit for a in all_accts),
        sum(a.ending_balance for a in all_accts),
    ]):
        c = ws.cell(row=row, column=_D + ci, value=val)
        _apply(c, font=_font(bold=True, size=12),
               fmt='#,##0.00;(#,##0.00);"-"', border=DOUBLE_BTM)

    ws.freeze_panes = 'B5'


# ── History-extraction helpers ───────────────────────────────

def _safe_float(v):
    """Return float(v) or None if v is None/non-numeric."""
    try:
        return float(v) if v is not None else None
    except (ValueError, TypeError):
        return None


def _extract_old_format_row(ws, period_label: str):
    """
    Extract a single-period summary dict from an old-format account tab
    (transaction detail with GL/TB tie-out rows at the bottom).

    Scans for rows whose text contains:
      "ending balance per gl" → GL ending value (col _I)
      "tb balance"            → TB ending value
      "beginning balance"     → beginning balance value
    """
    gl_end   = None
    tb_end   = None
    beg_bal  = None

    for row_vals in ws.iter_rows(values_only=True, max_row=ws.max_row):
        row_str = ' '.join(str(c or '').lower() for c in row_vals)
        if ('ending balance per gl' in row_str
                or ('ending balance' in row_str and 'gl' in row_str
                    and 'tb' not in row_str and 'projected' not in row_str)):
            for c in row_vals:
                v = _safe_float(c)
                if v is not None:
                    gl_end = v
                    break
        elif 'tb balance' in row_str and gl_end is not None:
            for c in row_vals:
                v = _safe_float(c)
                if v is not None:
                    tb_end = v
                    break
        elif 'beginning balance' in row_str and beg_bal is None:
            for c in row_vals:
                v = _safe_float(c)
                if v is not None:
                    beg_bal = v
                    break

    if gl_end is None:
        return None

    tb_val     = tb_end if tb_end is not None else gl_end
    net_change = gl_end - (beg_bal or 0.0)
    return {
        'period':     period_label,
        'beg_bal':    beg_bal or 0.0,
        'net_change': round(net_change, 2),
        'gl_end':     gl_end,
        'tb_end':     tb_val,
        'variance':   round(gl_end - tb_val, 2),
    }


def _extract_new_format_history(ws) -> list:
    """
    Extract all history rows from a new-format (rolling-table) account tab.

    Looks for a header row containing "Period" in column _B; reads subsequent
    rows until a blank period cell is found.

    Columns (B..G): Period | Beg Balance | Net Activity | GL Ending | TB Ending | Variance
    """
    rows = []
    header_row = None
    for r in range(1, min(15, ws.max_row + 1)):
        val = str(ws.cell(r, _B).value or '').strip().lower()
        if val == 'period':
            header_row = r
            break
    if not header_row:
        return rows

    for r in range(header_row + 1, ws.max_row + 1):
        period_val = str(ws.cell(r, _B).value or '').strip()
        if not period_val or not re.match(
                r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{4}',
                period_val):
            break
        beg  = _safe_float(ws.cell(r, _C).value)
        net  = _safe_float(ws.cell(r, _D).value)
        gl_e = _safe_float(ws.cell(r, _E).value)
        tb_e = _safe_float(ws.cell(r, _F).value)
        var  = _safe_float(ws.cell(r, _G).value)
        if gl_e is not None:
            rows.append({
                'period':     period_val,
                'beg_bal':    beg    if beg  is not None else 0.0,
                'net_change': net    if net  is not None else 0.0,
                'gl_end':     gl_e,
                'tb_end':     tb_e   if tb_e is not None else gl_e,
                'variance':   var    if var  is not None else 0.0,
            })
    return rows


_MONTH_ORDER = dict(Jan=1, Feb=2, Mar=3, Apr=4, May=5, Jun=6,
                    Jul=7, Aug=8, Sep=9, Oct=10, Nov=11, Dec=12)


def _period_sort_key(row: dict):
    parts = str(row.get('period', '')).split('-')
    if len(parts) == 2:
        mon = _MONTH_ORDER.get(parts[0], 0)
        yr  = int(parts[1]) if parts[1].isdigit() else 0
        return (yr, mon)
    return (0, 0)


def _extract_account_history(wb_prior) -> dict:
    """
    Extract per-period summary rows from any workpaper (old or new format).

    Old format: tabs named "Jan-2026 111100" — one tab per period per account.
    New format: tabs named "111100 PNC Cash" — one rolling-table tab per account.

    Returns {account_code: [sorted list of period row dicts]}.
    """
    history: dict = {}

    for sheet_name in (wb_prior.sheetnames if wb_prior else []):
        stripped = sheet_name.strip()

        # New format: tab starts with 6-digit account code, no period prefix
        if re.match(r'^\d{6}', stripped) and not _PERIOD_PREFIX_RE.match(stripped):
            acct_code = stripped[:6]
            ws    = wb_prior[sheet_name]
            rows  = _extract_new_format_history(ws)
            if rows:
                existing = history.get(acct_code, [])
                existing_periods = {r['period'] for r in existing}
                history[acct_code] = existing + [
                    r for r in rows if r['period'] not in existing_periods
                ]
            continue

        # Old format: "Period ACCTCODE [name]", e.g. "Jan-2026 111100"
        pfx_m = _PERIOD_PREFIX_RE.match(stripped)
        if pfx_m:
            period_label = pfx_m.group(0).strip()   # "Jan-2026"
            remainder    = stripped[pfx_m.end():].strip()
            code_m       = re.match(r'^(\d{6})', remainder)
            if code_m:
                acct_code = code_m.group(1)
                ws  = wb_prior[sheet_name]
                row = _extract_old_format_row(ws, period_label)
                if row:
                    existing_periods = {r['period'] for r in history.get(acct_code, [])}
                    if period_label not in existing_periods:
                        history.setdefault(acct_code, []).append(row)

    # Sort each account's history chronologically
    for acct_code in history:
        history[acct_code] = sorted(history[acct_code], key=_period_sort_key)

    return history


# ── Full transaction-level carry-forward for escrow + capital accounts ────────

# Accounts that need full detail rows carried forward month-over-month
_FULL_DETAIL_ACCOUNTS = frozenset({
    '115200', '115300', '115600',                               # escrow / reserve
    '152100', '154100', '154500', '171100',                     # capital — simple
    '181200', '181300', '181400',                               # capital — entity/comm
    '311100', '331100', '381100',                               # equity
})

# Capital tab layout info: {account_code: (has_entity, has_commencement)}
_CAPITAL_TAB_LAYOUTS = {
    '152100': (False, False),
    '154100': (False, False),
    '154500': (False, False),
    '171100': (False, False),
    '181200': (True,  True),
    '181300': (True,  True),
    '181400': (True,  True),
}


def _read_escrow_tab_detail(ws) -> list:
    """
    Extract all transaction rows from an escrow/reserve workpaper tab.

    Tab column layout (written by _build_escrow_tab in account_tab_builders.py):
      B = Date string ('M/D/YYYY')
      C = Description
      D = Entity
      E = Amount (numeric value, NOT a formula)
      F = Running Balance (Excel formula — skipped on read-back)

    Skips header rows, the Balance Forward row, and any row without both a
    date-like value in col B and a numeric value in col E.

    Returns list of {date_str, desc, amt} dicts.
    """
    rows = []
    for row_vals in ws.iter_rows(min_row=1, values_only=True):
        if len(row_vals) < 5:
            continue
        col_b = row_vals[1]   # Date
        col_c = row_vals[2]   # Description
        col_e = row_vals[4]   # Amount

        # Col B must look like a date: M/D/YYYY
        if not isinstance(col_b, str):
            continue
        if not re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', col_b.strip()):
            continue
        # Col E must be numeric (not a formula string)
        if not isinstance(col_e, (int, float)):
            continue

        rows.append({
            'date_str': col_b.strip(),
            'desc':     str(col_c or '').strip(),
            'amt':      float(col_e),
        })
    return rows


def _read_capital_tab_detail(ws, has_entity: bool, has_commencement: bool) -> list:
    """
    Extract all item rows from a capital workpaper tab.

    Tab column layouts (written by _build_capital_tab):
      154500 (has_entity=False, has_commencement=False):
        B = Description, C = Date, D = Amount
      181200/181300/181400 (has_entity=True, has_commencement=True):
        B = Description, C = Entity, D = Commencement Date, E = Amount

    Skips header rows, the column header row, and the Ending Balance / tieout rows.
    Detection: col B is a non-empty string AND the last data column is numeric.

    Returns list of dicts matching the column layout.
    """
    _SKIP_LABELS = frozenset({
        'description', 'entity', 'commencement date', 'date', 'amount',
        'ending balance per gl', 'gl balance', 'tb balance', 'difference',
        'trial balance', 'tieout', 'tie-out', 'workpaper',
    })

    rows = []
    for row_vals in ws.iter_rows(min_row=1, values_only=True):
        col_b = row_vals[1] if len(row_vals) > 1 else None
        if not isinstance(col_b, str) or not col_b.strip():
            continue

        col_b_lower = col_b.strip().lower()
        # Skip column-header rows and summary/tieout labels
        if col_b_lower in _SKIP_LABELS:
            continue
        if any(kw in col_b_lower for kw in ('ending balance', 'gl balance',
                                              'tb balance', 'tieout', 'tie-out')):
            break   # reached summary section — stop reading

        if has_entity and has_commencement:
            # 181xxx: B=desc, C=entity, D=commencement, E=amount
            if len(row_vals) < 5:
                continue
            col_e = row_vals[4]
            if not isinstance(col_e, (int, float)):
                continue
            rows.append({
                'description':       col_b.strip(),
                'entity':            str(row_vals[2] or ''),
                'commencement_date': str(row_vals[3] or ''),
                'amount':            float(col_e),
            })
        else:
            # 154500: B=desc, C=date, D=amount
            if len(row_vals) < 4:
                continue
            col_d = row_vals[3]
            if not isinstance(col_d, (int, float)):
                continue
            rows.append({
                'description': col_b.strip(),
                'date':        str(row_vals[2] or ''),
                'amount':      float(col_d),
            })
    return rows


def _read_equity_contributions_tab_detail(ws) -> list:
    """
    Read rows from a 311100 Contributions tab.

    Column layout (written by build_311100_tab):
      B = Date string ('M/D/YYYY')
      C = Description
      D = Amount (numeric)

    Returns list of {date_str, desc, amt} dicts.
    """
    rows = []
    for row_vals in ws.iter_rows(min_row=1, values_only=True):
        if len(row_vals) < 4:
            continue
        col_b = row_vals[1]   # Date
        col_c = row_vals[2]   # Description
        col_d = row_vals[3]   # Amount

        if not isinstance(col_b, str):
            continue
        if not re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', col_b.strip()):
            continue
        if not isinstance(col_d, (int, float)):
            continue

        rows.append({
            'date_str': col_b.strip(),
            'desc':     str(col_c or '').strip(),
            'amt':      float(col_d),
        })
    return rows


def _read_equity_distributions_tab_detail(ws) -> list:
    """
    Read rows from a 331100 Distributions tab.

    Column layout (written by build_331100_tab):
      B = Date string ('M/D/YYYY')
      C = Description
      D = Entity-1 amount (key: 'revlabs'   — RevLabs-specific; positional read)
      E = Entity-2 amount (key: 'revlabspm' — RevLabs-specific; positional read)
      F = Total amount (all numeric)

    C-10: Keys 'revlabs' / 'revlabspm' are internal positional labels for the two
    entity columns; they reflect the RevLabs two-entity structure and should be
    made config-driven if this function is extended to other properties.

    Skips header rows, totals row, and tie-out rows.
    Returns list of {date_str, desc, revlabs, revlabspm, total} dicts.
    """
    _SKIP = frozenset({'date', 'description', 'revlabs', 'revlabspm', 'total',
                       'total distributions', 'ending balance per gl',
                       'ending balance per tb', 'variance'})
    rows = []
    for row_vals in ws.iter_rows(min_row=1, values_only=True):
        if len(row_vals) < 6:
            continue
        col_b = row_vals[1]   # Date
        col_c = row_vals[2]   # Description
        col_d = row_vals[3]   # Revlabs
        col_e = row_vals[4]   # Revlabpm
        col_f = row_vals[5]   # Total

        if not isinstance(col_b, str):
            continue
        if not re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', col_b.strip()):
            continue
        if isinstance(col_c, str) and col_c.strip().lower() in _SKIP:
            continue
        # Need at least one numeric amount
        if not isinstance(col_f, (int, float)):
            continue

        rows.append({
            'date_str': col_b.strip(),
            'desc':     str(col_c or '').strip(),
            'revlabs':  float(col_d) if isinstance(col_d, (int, float)) else 0.0,
            'revlabspm': float(col_e) if isinstance(col_e, (int, float)) else 0.0,
            'total':    float(col_f),
        })
    return rows


def _read_equity_retained_earnings_split(ws) -> dict:
    """
    Read the entity split from a 381100 Retained Earnings tab.

    Column layout (written by build_381100_tab):
      B = Description (look for 'Beginning Balance')
      C = Revlabpm amount
      D = Revlabs amount

    Returns {revlabspm: float, revlabs: float}.
    Falls back to empty dict if the row can't be found.
    """
    for row_vals in ws.iter_rows(min_row=1, values_only=True):
        if len(row_vals) < 4:
            continue
        col_b = row_vals[1]
        col_c = row_vals[2]
        col_d = row_vals[3]
        if isinstance(col_b, str) and 'beginning balance' in col_b.lower():
            if isinstance(col_c, (int, float)) and isinstance(col_d, (int, float)):
                return {'revlabspm': float(col_c), 'revlabs': float(col_d)}
    return {}


def _extract_prior_full_detail(wb_prior) -> dict:
    """
    Extract full transaction-level detail rows from escrow, capital, and equity
    account tabs in the prior workpaper.

    Covers accounts: 115200, 115300, 115600, 154500, 181200, 181300, 181400,
                     311100, 331100, 381100.

    Returns {account_code: [list of row dicts]}.
    For 381100 specifically, returns {account_code: {revlabspm, revlabs}} dict.
    If the same account code appears in multiple tabs (e.g., with a period prefix),
    the last tab encountered is used — in practice there should be only one tab
    per account in any given workpaper.
    """
    detail: dict = {}
    if not wb_prior:
        return detail

    for sheet_name in wb_prior.sheetnames:
        stripped = sheet_name.strip()
        # Strip any period prefix (e.g., "Mar-2026 115200 RET Escrow" → "115200...")
        bare = _PERIOD_PREFIX_RE.sub('', stripped).strip()
        code_m = re.match(r'^(\d{6})', bare)
        if not code_m:
            continue
        acct_code = code_m.group(1)
        if acct_code not in _FULL_DETAIL_ACCOUNTS:
            continue

        ws = wb_prior[sheet_name]
        if acct_code in ('115200', '115300', '115600'):
            rows = _read_escrow_tab_detail(ws)
            if rows:
                detail[acct_code] = rows
        elif acct_code == '311100':
            rows = _read_equity_contributions_tab_detail(ws)
            if rows:
                detail[acct_code] = rows
        elif acct_code == '331100':
            rows = _read_equity_distributions_tab_detail(ws)
            if rows:
                detail[acct_code] = rows
        elif acct_code == '381100':
            split = _read_equity_retained_earnings_split(ws)
            if split:
                detail[acct_code] = split
        else:
            has_entity, has_comm = _CAPITAL_TAB_LAYOUTS[acct_code]
            rows = _read_capital_tab_detail(ws, has_entity, has_comm)
            if rows:
                detail[acct_code] = rows

    return detail


# ── Account reconciliation tab ────────────────────────────────

# Matches Yardi auto-reversal boilerplate at the start of description fields.
# Examples matched:  "Reversal of J-19118: memo"  |  "J-19118: memo"
# Examples NOT matched: "J-2024"  (no separator → not boilerplate, keep it)
# Requires a colon/dash separator + whitespace after the JE number so a bare
# "J-XXXX" that IS the meaningful description is never accidentally stripped.
_REVERSAL_PREFIX_RE = re.compile(
    r'^(reversal\s+of\s+)?[Jj]-\d+\s*[:\-–]\s+', re.IGNORECASE
)

# Broader reversal detector — matches "reversal of J-XXXXX" ANYWHERE in the
# text, not just as a prefix. Real Yardi exports also produce it as a suffix
# after an empty vendor/memo field (e.g. ": Reversal of J-18456", where the
# leading ": " is the blank field's own separator) which _REVERSAL_PREFIX_RE's
# start anchor misses entirely.
_REVERSAL_ANYWHERE_RE = re.compile(r'reversal\s+of\s+[Jj]-\d+', re.IGNORECASE)


def _is_reversal_txn(txn) -> bool:
    """True if a GL transaction's description/remarks mark it as an auto-reversal."""
    desc    = str(getattr(txn, 'description', '') or '')
    remarks = str(getattr(txn, 'remarks', '') or '')
    return bool(_REVERSAL_ANYWHERE_RE.search(desc) or _REVERSAL_ANYWHERE_RE.search(remarks))


_NO_DESC_FLAG = '[!] No Description - Review Required'


def _fmt_txn_desc(t, flag_blank: bool = False) -> str:
    """
    Build a clean workpaper description from a GLTransaction.

    Priority:
      1. remarks  — the actual transaction memo entered by the user.
      2. description — fallback when remarks is blank.

    Strips Yardi auto-reversal boilerplate ('Reversal of J-XXXXX: ', 'J-XXXXX: ')
    and trailing (tXXX)/(vXXX) suffixes.  Control / Reference numbers are never
    included in the output.

    Args:
        flag_blank: if True and the cleaned result is empty, return the reviewer
                    flag string instead of an empty string.
    """
    desc    = (t.description or '').strip()
    remarks = (t.remarks     or '').strip()

    # Strip Yardi boilerplate from both fields
    if desc:
        desc = _REVERSAL_PREFIX_RE.sub('', desc).strip(' :–-')
        desc = re.sub(r'\s*\([tv]\d+\)\s*$', '', desc).strip()
    if remarks:
        remarks = _REVERSAL_PREFIX_RE.sub('', remarks).strip(' :–-')
        remarks = re.sub(r'\s*\([tv]\d+\)\s*$', '', remarks).strip()

    result = remarks or desc
    if not result and flag_blank:
        return _NO_DESC_FLAG
    return result


# Cash account codes — show Deposits / Disbursements instead of Debit / Credit
_CASH_ACCOUNTS = {'111100', '115100', '115200', '115300', '115600'}


def _is_journal_entry(control: str) -> bool:
    """Return True if the GL control number looks like a journal entry (not a bank item)."""
    c = (control or '').strip().upper()
    # JE controls: 'J-12345', 'GJ-001', 'MGT-001', 'SUP-001', etc.
    return bool(c and (c[0] == 'J' or c[:2] in ('GJ', 'MG', 'SU', 'PR', 'TU')))


def _group_cash_txns_by_day(txns):
    """
    For cash accounts, collapse same-day same-direction bank transactions into
    daily summary rows.  Journal entries are always kept individual (they have
    meaningful descriptions).  Single-item days render as-is.

    Returns a list of dicts, each representing one display row:
        {
          'date':    date | str,
          'desc':    str,
          'control': str,
          'ref':     str,
          'debit':   float | None,
          'credit':  float | None,
          'balance': float,
          'is_summary': bool,   # True → use italic style
        }
    """
    from itertools import groupby as _groupby

    # Sort by date then preserve original order within a date
    sorted_txns = sorted(txns, key=lambda t: (t.date or datetime.min.date()))

    rows = []
    for dt_val, day_iter in _groupby(sorted_txns, key=lambda t: t.date):
        day_list = list(day_iter)

        # Journal entries: always individual
        je_txns   = [t for t in day_list if _is_journal_entry(t.control or '')]
        bank_txns = [t for t in day_list if not _is_journal_entry(t.control or '')]

        # End-of-day balance = last bank transaction's balance for this date.
        # All summary rows on the same day share this balance so the reader
        # sees a single consistent closing balance rather than a mid-day jump.
        eod_balance = bank_txns[-1].balance if bank_txns else (je_txns[-1].balance if je_txns else 0.0)

        # JE rows first (they're typically end-of-month postings)
        for t in je_txns:
            rows.append({
                'date': t.date, 'desc': _fmt_txn_desc(t),
                'control': (t.control or '').strip(),
                'ref': (t.reference or '').strip(),
                'debit':  t.debit  if (t.debit  or 0) > 0.005 else None,
                'credit': t.credit if (t.credit or 0) > 0.005 else None,
                'balance': t.balance, 'is_summary': False,
            })

        # Deposits (debits) — group if more than one
        deposits = [t for t in bank_txns if (t.debit or 0) > 0.005]
        if len(deposits) == 1:
            t = deposits[0]
            rows.append({
                'date': t.date, 'desc': _fmt_txn_desc(t),
                'control': (t.control or '').strip(),
                'ref': (t.reference or '').strip(),
                'debit': t.debit, 'credit': None,
                'balance': t.balance, 'is_summary': False,
            })
        elif deposits:
            total_dep = sum(t.debit for t in deposits)
            rows.append({
                'date': dt_val,
                'desc': f'Daily Deposits — {len(deposits)} items',
                'control': '', 'ref': '',
                'debit': total_dep, 'credit': None,
                'balance': eod_balance, 'is_summary': True,
            })

        # Disbursements (credits) — group if more than one
        disb = [t for t in bank_txns if (t.credit or 0) > 0.005]
        if len(disb) == 1:
            t = disb[0]
            rows.append({
                'date': t.date, 'desc': _fmt_txn_desc(t),
                'control': (t.control or '').strip(),
                'ref': (t.reference or '').strip(),
                'debit': None, 'credit': t.credit,
                'balance': t.balance, 'is_summary': False,
            })
        elif disb:
            total_disb = sum(t.credit for t in disb)
            rows.append({
                'date': dt_val,
                'desc': f'Daily Disbursements — {len(disb)} items',
                'control': '', 'ref': '',
                'debit': None, 'credit': total_disb,
                'balance': eod_balance, 'is_summary': True,
            })

    return rows


def _write_account_tab(wb, gl_acct, tb_acct, period, property_name,
                       je_adjustments=None, tab_prefix: str = '',
                       history_rows: list = None,
                       entity_label: str = '',
                       entities: list = None,
                       prepared_by: str = ''):
    """
    One tab per balance sheet account — clean transaction register.

    Layout:
      Row 1: Account# — Account Name                         (dark blue header)
      Row 2: Property | Period | Prepared date               (green header)
      Row 3: blank
      Row 4: Date | Description | [entity_label]             (column headers)
      Row 5: Balance Forward                                  [beginning balance]
      Row 6+: one row per GL transaction                      [date | desc | net amount]
              → blank description → ⚑ No Description flag
      Row N: Ending Balance                                   [bold blue]
      Row N+2: GL Ending Balance  ┐
      Row N+3: TB Ending Balance  ├── tie-out block
      Row N+4: Variance           ┘
      Row N+6+: Prior period history (if available)

    entity_label: column header for the amount column — read from the GL file's
                  property code/name.  Defaults to 'Amount' if not provided.
    """
    acct_label = _safe_sheet_name(f'{gl_acct.account_code} {gl_acct.account_name}')
    ws = wb.create_sheet(acct_label)

    is_complex  = gl_acct.account_code in COMPLEX_ACCOUNTS
    is_cash     = gl_acct.account_code in _CASH_ACCOUNTS
    ws.sheet_properties.tabColor = COLOR_BS_COMPLEX if is_complex else COLOR_BS_STD
    ws.column_dimensions['A'].width = 2

    # Column layout — Date | Description | [entity col(s)] | (Total if multi-entity)
    # Control and Reference columns removed per review standard.
    # entities: ordered list from GL metadata. Single-entity → one amount col.
    #           Multi-entity → one col per entity + Total.
    _entities = entities if (entities and len(entities) > 1) else []
    _multi    = bool(_entities)

    _DT   = _B           # B  Date
    _DSC  = _B + 1       # C  Description
    # Entity cols start at D (col index _B+2)
    if _multi:
        _ent_cols  = {ent: _B + 2 + i for i, ent in enumerate(_entities)}
        _TOT_COL   = _B + 2 + len(_entities)   # Total column
        _LAST_COL  = _TOT_COL
        _ent_hdrs  = [f'Entity ({e})' for e in _entities] + ['Total']
        _ent_widths = [16] * len(_entities) + [18]
    else:
        _ent_cols  = {}
        _AMT       = _B + 2                    # single entity amount col
        _LAST_COL  = _AMT
        _ent_hdrs  = [f'Entity ({entity_label or "[Property]"})']
        _ent_widths = [18]

    # ── Row 1: Account header ──────────────────────────────────────────────
    row = 1
    c = ws.cell(row=row, column=_B,
                value=f'{gl_acct.account_code} — {gl_acct.account_name}')
    c.font = _font(bold=True, size=13, color='FFFFFF')
    c.fill = _fill('375623')
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_LAST_COL)
    row += 1

    # ── Row 2: Property / period sub-header ──────────────────────────────
    _preparer = prepared_by or 'GRP'   # C-18: no personal name default
    c = ws.cell(row=row, column=_B,
                value=f'{property_name or "Revolution Labs"}  |  '
                      f'Period: {period}  |  '
                      f'Prepared by: {_preparer}  |  '
                      f'{datetime.now().strftime("%m/%d/%Y")}')
    c.font = _font(italic=True, color='FFFFFF', size=10)
    c.fill = _fill('375623')
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_LAST_COL)
    row += 2   # blank row 3

    # ── Row 4: Column headers ─────────────────────────────────────────────
    col_hdrs   = ['Date', 'Description'] + _ent_hdrs
    col_widths = [12,      58 if not _multi else 46] + _ent_widths
    for ci, (h, w) in enumerate(zip(col_hdrs, col_widths)):
        col = _B + ci
        c = ws.cell(row=row, column=col, value=h)
        _apply(c, font=_hdr_font(), fill=_fill('000000'), border=THIN,
               align=Alignment(horizontal='center', wrap_text=True))
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.column_dimensions['A'].width = 2
    ws.row_dimensions[row].height = 22
    row += 1

    # ── Row 5: Beginning balance ──────────────────────────────────────────
    _NUM_FMT = '#,##0.00;(#,##0.00);"-"'
    for col in range(_B, _LAST_COL + 1):
        c = ws.cell(row=row, column=col)
        c.font   = _font(bold=True, italic=True, size=9)
        c.fill   = _fill(LIGHT_GRAY)
        c.border = THIN
    ws.cell(row=row, column=_DSC, value='Balance Forward')
    # Show beginning balance in the Total / single-entity column
    beg_cell = ws.cell(row=row, column=_LAST_COL, value=gl_acct.beginning_balance)
    beg_cell.number_format = _NUM_FMT
    row += 1

    # ── Transaction rows ──────────────────────────────────────────────────
    txns = [t for t in (gl_acct.transactions or []) if t.period == period or not t.period]
    if not txns:
        txns = list(gl_acct.transactions or [])

    net_activity = 0.0
    # For multi-entity: track net per entity for Ending Balance row
    _ent_net: dict = {ent: 0.0 for ent in _entities}
    display_idx = 0

    for t in txns:
        net = round(float(t.debit or 0) - float(t.credit or 0), 2)
        if abs(net) < 0.005:
            continue   # skip zero-net rows — no amount, nothing to flag

        net_activity += net
        alt = _fill(LIGHT_GRAY) if display_idx % 2 == 1 else None
        display_idx += 1

        d      = t.date
        t_date = d.strftime('%m/%d/%Y') if hasattr(d, 'strftime') else str(d or '')
        desc   = _fmt_txn_desc(t, flag_blank=True)

        # Blank all cells in this row first (consistent borders)
        for col in range(_B, _LAST_COL + 1):
            c = ws.cell(row=row, column=col)
            c.border = THIN
            if alt:
                c.fill = alt

        # Date cell
        c1 = ws.cell(row=row, column=_DT, value=t_date)
        _apply(c1, font=_font(size=9), fill=alt, border=THIN)

        # Description cell — amber highlight when description is missing but amount exists
        is_flagged = (desc == _NO_DESC_FLAG)
        c2 = ws.cell(row=row, column=_DSC, value=desc)
        _apply(c2,
               font=_font(size=9, italic=is_flagged, color='7F3F00' if is_flagged else '000000'),
               fill=_fill(AMBER_FILL) if is_flagged else alt,
               border=THIN,
               align=Alignment(wrap_text=True))

        if _multi:
            # Route amount to entity column; show total in Total col
            txn_entity = getattr(t, 'entity', '') or ''
            ent_col = _ent_cols.get(txn_entity, _ent_cols.get(txn_entity.lower()))
            if ent_col:
                ce = ws.cell(row=row, column=ent_col, value=net)
                _apply(ce, font=_font(size=9), fill=alt, fmt=_NUM_FMT, border=THIN,
                       align=Alignment(horizontal='right'))
                _ent_net[txn_entity] = round(_ent_net.get(txn_entity, 0.0) + net, 2)
            ct = ws.cell(row=row, column=_TOT_COL, value=net)
            _apply(ct, font=_font(size=9), fill=alt, fmt=_NUM_FMT, border=THIN,
                   align=Alignment(horizontal='right'))
        else:
            c3 = ws.cell(row=row, column=_AMT, value=net)
            _apply(c3, font=_font(size=9), fill=alt, fmt=_NUM_FMT, border=THIN,
                   align=Alignment(horizontal='right'))
        row += 1

    # ── Ending balance row ────────────────────────────────────────────────
    _je_delta = (je_adjustments or {}).get(gl_acct.account_code, 0.0)
    gl_end    = gl_acct.ending_balance + _je_delta
    for col in range(_B, _LAST_COL + 1):
        c = ws.cell(row=row, column=col)
        c.font   = _font(bold=True, size=9)
        c.fill   = _fill(LIGHT_BLUE)
        c.border = THIN
    ws.cell(row=row, column=_DSC, value='Ending Balance')
    if _multi:
        for ent, col in _ent_cols.items():
            ev = round(_ent_net.get(ent, 0.0), 2)
            ec = ws.cell(row=row, column=col, value=ev or None)
            ec.number_format = _NUM_FMT
        tc = ws.cell(row=row, column=_TOT_COL, value=round(gl_acct.beginning_balance + net_activity, 2) or None)
        tc.number_format = _NUM_FMT
    else:
        ec = ws.cell(row=row, column=_AMT, value=round(gl_acct.beginning_balance + net_activity, 2) or None)
        ec.number_format = _NUM_FMT
    row += 2   # blank gap

    # ── Tie-out block ─────────────────────────────────────────────────────
    workpaper_total = round(gl_acct.beginning_balance + net_activity + _je_delta, 2)
    tb_end   = tb_acct.ending_balance if tb_acct else None
    variance = round(workpaper_total - tb_end, 2) if tb_end is not None else None
    _vzero   = variance is not None and abs(variance) < 0.02

    lbl = ws.cell(row=row, column=_DSC, value=gl_acct.account_code)
    lbl.font = _font(bold=True, size=9)
    row += 1

    gl_lbl = ws.cell(row=row, column=_DSC, value='Workpaper Total')
    gl_lbl.font = _font(size=9)
    gl_val = ws.cell(row=row, column=_LAST_COL, value=workpaper_total)
    gl_val.number_format = _NUM_FMT
    gl_val.font = _font(size=9)
    row += 1

    if tb_end is not None:
        tb_lbl = ws.cell(row=row, column=_DSC, value='TB Ending Balance')
        tb_lbl.font = _font(size=9)
        tb_val = ws.cell(row=row, column=_LAST_COL, value=tb_end)
        tb_val.number_format = _NUM_FMT
        tb_val.font = _font(size=9)
        row += 1

        var_lbl = ws.cell(row=row, column=_DSC, value='Variance')
        var_lbl.font = _font(bold=True, size=9)
        var_val = ws.cell(row=row, column=_LAST_COL, value=variance)
        var_val.number_format = _NUM_FMT
        var_val.font  = _font(bold=True, size=9,
                               color='006100' if _vzero else '9C0006')
        var_val.fill  = _fill(GREEN_FILL if _vzero else RED_FILL)
        row += 1

        if not _vzero:
            row += 1
            note = ws.cell(row=row, column=_B,
                           value=f'Variance ${abs(variance):,.2f} — '
                                 f'review accrual entries for {period}.')
            note.font = _font(italic=True, color='9C0006', size=9)
            ws.merge_cells(start_row=row, start_column=_B,
                           end_row=row, end_column=_LAST_COL)
            row += 1

    # ── Prior period history (if carried forward) ─────────────────────────
    if history_rows:
        row += 1
        hdr = ws.cell(row=row, column=_B, value='Prior Period History')
        hdr.font = _font(bold=True, size=9, color='FFFFFF')
        hdr.fill = _fill(DARK_BLUE)
        ws.merge_cells(start_row=row, start_column=_B,
                       end_row=row, end_column=_B + 5)
        row += 1
        hist_hdrs = ['Period', 'Beg Balance', 'Net Activity', 'GL Ending', 'TB Ending', 'Variance']
        for ci, h in enumerate(hist_hdrs):
            c = ws.cell(row=row, column=_B + ci, value=h)
            _apply(c, font=_hdr_font(), fill=_fill(MED_BLUE), border=THIN,
                   align=Alignment(horizontal='center'))
        row += 1
        for i, hist in enumerate(history_rows):
            alt_fill = _fill(LIGHT_GRAY) if i % 2 == 1 else None
            _var   = hist.get('variance', 0.0) or 0.0
            _vz    = abs(_var) < 0.02
            vals   = [hist.get('period', ''), hist.get('beg_bal', 0.0),
                      hist.get('net_change', 0.0), hist.get('gl_end', 0.0),
                      hist.get('tb_end', hist.get('gl_end', 0.0)), _var]
            for ci, val in enumerate(vals):
                c = ws.cell(row=row, column=_B + ci, value=val)
                c.font   = _font(size=9)
                c.border = THIN
                if alt_fill:
                    c.fill = alt_fill
                if isinstance(val, float):
                    c.number_format = _NUM_FMT
            vc = ws.cell(row=row, column=_B + 5)
            vc.fill = _fill(GREEN_FILL if _vz else RED_FILL)
            vc.font = _font(size=9, color='006100' if _vz else '9C0006')
            row += 1

    ws.freeze_panes = 'B5'


def _write_tieout(ws, row, gl_acct, tb_acct, period, je_delta: float = 0.0):
    """Write the GL ending / TB balance / Variance tie-out block (Hartwell inline style)."""

    # Separator line
    for col in range(_B, _I + 1):
        ws.cell(row=row, column=col).border = THICK_BOTTOM
    row += 1

    gl_ending = gl_acct.ending_balance + je_delta   # projected post-close if je_delta != 0
    tb_ending = tb_acct.ending_balance if tb_acct else None
    variance  = (gl_ending - tb_ending) if tb_ending is not None else None

    # Blank separator row (already advanced past separator line above)
    row += 1

    # GL ending balance — Hartwell inline style
    # Label in _D (description col), value in _I (balance col), bold, light blue fill across data cols
    _gl_label = (f'Projected Balance per GL as of {period} (incl. pipeline JEs)'
                 if je_delta != 0.0 else f'Ending Balance per GL as of {period}')
    label_gl = ws.cell(row=row, column=_D, value=_gl_label)
    label_gl.font = _font(bold=True)
    c_gl = ws.cell(row=row, column=_I, value=gl_ending)
    _apply(c_gl, font=_font(bold=True), fmt='#,##0.00;(#,##0.00);"-"',
           fill=_fill(LIGHT_BLUE), border=THICK_BOTTOM)
    for col in range(_B, _I + 1):
        cell = ws.cell(row=row, column=col)
        if not cell.fill or cell.fill.fill_type == 'none':
            cell.fill = _fill(LIGHT_BLUE)
    row += 1

    # TB balance — label in _H, value in _I
    label_tb = ws.cell(row=row, column=_H, value='TB Balance')
    label_tb.font = _font(bold=True)
    if tb_ending is not None:
        c_tb = ws.cell(row=row, column=_I, value=tb_ending)
        _apply(c_tb, font=_font(bold=True), fmt='#,##0.00;(#,##0.00);"-"',
               fill=_fill(LIGHT_BLUE), border=THIN)
    else:
        c_tb = ws.cell(row=row, column=_I, value='Not in TB')
        c_tb.font = _font(italic=True, color='888888')
    row += 1

    # Variance — label in _H, value in _I; green if zero, red if non-zero
    label_var = ws.cell(row=row, column=_H, value='Variance')
    label_var.font = _font(bold=True)
    if variance is not None:
        is_zero = abs(variance) < 0.02
        var_fill = _fill(GREEN_FILL) if is_zero else _fill(RED_FILL)
        var_color = '006100' if is_zero else '9C0006'
        c_var = ws.cell(row=row, column=_I, value=variance)
        _apply(c_var, font=_font(bold=True, color=var_color),
               fmt='#,##0.00;(#,##0.00);"-"', fill=var_fill, border=DOUBLE_BTM)

        if not is_zero:
            note_row = row + 2
            note = ws.cell(row=note_row, column=_B,
                           value=f'Variance of ${abs(variance):,.2f} — likely accrual entries in TB not yet in GL. '
                                 f'Review accrual JEs posted for this account.')
            note.font = _font(italic=True, color='9C0006', size=10)
            note.alignment = Alignment(wrap_text=True)
            ws.merge_cells(start_row=note_row, start_column=_B,
                           end_row=note_row, end_column=_I)
            ws.row_dimensions[note_row].height = 28
    else:
        ws.cell(row=row, column=_I, value='').border = DOUBLE_BTM


# ── Accrual schedule helpers ─────────────────────────────────

def _parse_accrual_txn(desc: str, expense_name: str = '') -> dict:
    """
    Parse a pipeline-generated accrual description into structured fields for
    the JLL-style accrual schedule tab.

    Returns dict with keys:
        acct_desc   — expense account name (from description or expense_name arg)
        vendor      — vendor name if identifiable
        period_from — billing/service period start (string)
        period_to   — billing/service period end (string)
        notes       — short description line (matches JLL "Acc …" style)
    """
    import re as _re
    result = {
        'acct_desc': expense_name or '',
        'vendor': '',
        'period_from': '',
        'period_to': '',
        'notes': (desc or '').strip(),
    }
    if not desc:
        return result

    # "Invoice proration — Account Name: last invoice MM/DD/YY-MM/DD/YY vendor..."
    m = _re.match(
        r'Invoice proration\s*[—\-]+\s*(.+?):\s*last invoice\s+([\d/][\d/\- ]+?)'
        r'(?:\s+(.+?))?$', desc, _re.I)
    if m:
        if not result['acct_desc']:
            result['acct_desc'] = m.group(1).strip()
        dates_str = m.group(2).strip()
        date_parts = _re.split(r'\s*[-–]\s*', dates_str)
        if len(date_parts) >= 2:
            result['period_from'] = date_parts[0].strip()
            result['period_to']   = date_parts[1].strip()
        vendor_extra = (m.group(3) or '').strip()
        if vendor_extra:
            result['vendor'] = vendor_extra[:40]
        result['notes'] = (
            f"Acc {result['period_from']} - {result['period_to']} "
            f"{result['acct_desc']}"
        ).strip()
        return result

    # "Payroll accrual — Account Name: last run MM/DD/YY (…)"
    m = _re.match(
        r'Payroll accrual\s*[—\-]+\s*(.+?):\s*last run\s+(\d{1,2}/\d{1,2}/\d{2,4})',
        desc, _re.I)
    if m:
        if not result['acct_desc']:
            result['acct_desc'] = m.group(1).strip()
        last_run = m.group(2).strip()
        result['period_from'] = last_run
        result['vendor'] = 'Payroll'
        result['notes'] = f"Acc payroll last run {last_run} {result['acct_desc']}"
        return result

    # "Monthly bonus accrual — Account Name: Kardin annual…"
    m = _re.match(r'Monthly bonus accrual\s*[—\-]+\s*(.+?):', desc, _re.I)
    if m:
        if not result['acct_desc']:
            result['acct_desc'] = m.group(1).strip()
        result['vendor'] = 'Bonus accrual'
        result['notes'] = f"Acc bonus per Kardin {result['acct_desc']}"
        return result

    # "Recurring monthly accrual — Account Name: VENDOR"
    m = _re.match(
        r'Recurring monthly accrual\s*[—\-]+\s*(.+?):\s*(.+?)$', desc, _re.I)
    if m:
        if not result['acct_desc']:
            result['acct_desc'] = m.group(1).strip()
        result['vendor'] = m.group(2).strip()[:40]
        result['notes'] = f"Acc {result['vendor']}"
        return result

    # "Budget gap accrual — Account Name: …"
    m = _re.match(r'Budget gap accrual\s*[—\-]+\s*(.+?):\s*(.+?)$', desc, _re.I)
    if m:
        if not result['acct_desc']:
            result['acct_desc'] = m.group(1).strip()
        result['notes'] = m.group(2).strip()[:60]
        return result

    # "REVIEW REQUIRED — Account Name: …" / "REVIEW — …"
    m = _re.match(r'REVIEW(?:\s+REQUIRED)?\s*[—\-]+\s*(.+?):', desc, _re.I)
    if m:
        if not result['acct_desc']:
            result['acct_desc'] = m.group(1).strip()
        result['vendor'] = '⚠ REVIEW'
        result['notes'] = desc.strip()
        return result

    return result


def _write_accrual_schedule_tab(wb, gl_acct, tb_acct, period, property_name,
                                 control_to_expense: dict, tab_prefix: str = '',
                                 history_rows: list = None,
                                 prepared_by: str = ''):
    """
    Write a JLL-style accrual schedule tab for 211200 / 211300 / 213100.

    Layout matches '213100-Accrued Exp' in the JLL workpaper:
      Col B  Account #         (expense account code from GL debit side)
      Col C  Account Desc      (expense account name)
      Col D  Vendor            (parsed from description)
      Col E  FROM              (billing/service period start)
      Col F  TO                (billing/service period end)
      Col G  Accrual           (negative — credit to this liability account)
      Col H  Description       (short note matching "Acc MM/YY Vendor" style)

    Footer: total → GL balance → variance (should be ≤ $0.02 rounding)
    TB tie-out row appended after the GL section.
    """
    acct_label = _safe_sheet_name(f'{gl_acct.account_code} {gl_acct.account_name}')
    ws = wb.create_sheet(acct_label)
    ws.sheet_properties.tabColor = COLOR_BS_COMPLEX  # red — complex account
    ws.column_dimensions['A'].width = 2

    row = 1
    # ── Header ───────────────────────────────────────────────────
    c = ws.cell(row=row, column=_B,
                value=f'{gl_acct.account_code} — {gl_acct.account_name}')
    c.font = _font(bold=True, size=13, color='FFFFFF')
    c.fill = _fill(DARK_BLUE)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_I)
    row += 1

    c = ws.cell(row=row, column=_B,
                value=(f'Period: {period}  |  '
                       f'{property_name or "Revolution Labs"}  |  '
                       f'Prepared by: {prepared_by or "GRP"}  |  '
                       f'{datetime.now().strftime("%m/%d/%Y")}'))
    c.font = _font(italic=True, color='FFFFFF')
    c.fill = _fill(MED_BLUE)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_I)
    row += 3

    # ── Column headers ────────────────────────────────────────────
    col_specs = [
        ('Account #',        10),
        ('Account Description', 28),
        ('Vendor',           24),
        ('FROM',             14),
        ('TO',               14),
        ('Accrual',          14),
        ('Description',      48),
        ('',                  4),
    ]
    for ci, (h, w) in enumerate(col_specs):
        col = _B + ci
        ws.column_dimensions[get_column_letter(col)].width = w
        if not h:
            continue
        c = ws.cell(row=row, column=col, value=h)
        _apply(c, font=_hdr_font(), fill=_fill(DARK_BLUE), border=THIN,
               align=Alignment(horizontal='center', wrap_text=True))
    ws.row_dimensions[row].height = 24
    row += 1

    # ── Data rows ─────────────────────────────────────────────────
    total_accrual = 0.0
    txns = gl_acct.transactions or []

    for i, txn in enumerate(txns):
        credit = float(txn.credit or 0)
        debit  = float(txn.debit or 0)
        # Net credit = how much is accrued into this liability account
        net_credit = credit - debit

        ctrl = str(getattr(txn, 'control', '') or '').strip()
        expense_info = control_to_expense.get(ctrl, ('', ''))
        expense_code = expense_info[0] if expense_info else ''
        expense_name = expense_info[1] if expense_info else ''

        parsed = _parse_accrual_txn(txn.description or '', expense_name)

        alt_fill = _fill(LIGHT_GRAY) if i % 2 == 1 else None
        is_review = parsed['vendor'] == '⚠ REVIEW'

        # Accrual amount stored as negative (matching JLL's sign convention for credits)
        accrual_amount = -net_credit if net_credit != 0 else None

        row_data = [
            (expense_code,                   'left',   False),
            (parsed['acct_desc'] or expense_name or (txn.description or '')[:40],
                                              'left',   False),
            (parsed['vendor'],               'left',   False),
            (parsed['period_from'],          'center',  False),
            (parsed['period_to'],            'center',  False),
            (accrual_amount,                 'right',   True),   # number format
            (parsed['notes'][:65],           'left',   False),
        ]

        for ci, (val, align_h, is_num) in enumerate(row_data):
            col = _B + ci
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(horizontal=align_h,
                                    wrap_text=(ci == 6))
            c.border = THIN
            if alt_fill:
                c.fill = alt_fill
            if is_num and val is not None:
                c.number_format = '#,##0.00;(#,##0.00);"-"'
            if is_review:
                c.font = _font(bold=True, color='9C0006')

        if accrual_amount is not None:
            total_accrual += accrual_amount
        row += 1

    # ── Total row ─────────────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=_D, value='Rounding').font = _font(italic=True, color='888888')
    row += 1
    ws.cell(row=row, column=_D, value='Total').font = _font(bold=True)
    c_tot = ws.cell(row=row, column=_G, value=total_accrual)
    _apply(c_tot, font=_font(bold=True), fmt='#,##0.00;(#,##0.00);"-"', border=THIN)
    row += 2

    # ── GL balance + variance ─────────────────────────────────────
    gl_ending = gl_acct.ending_balance
    ws.cell(row=row, column=_D, value=str(gl_acct.account_code)).font = _font(bold=True, color=DARK_BLUE)
    row += 2

    ws.cell(row=row, column=_E, value='GL').font = _font(bold=True)
    c_gl = ws.cell(row=row, column=_G, value=gl_ending)
    _apply(c_gl, font=_font(bold=True), fmt='#,##0.00;(#,##0.00);"-"',
           fill=_fill(LIGHT_BLUE), border=THIN)
    row += 1

    # Variance between schedule total and GL ending balance
    sched_variance = (total_accrual + abs(gl_ending)) if gl_ending is not None else None
    ws.cell(row=row, column=_E, value='Variance').font = _font(bold=True)
    if sched_variance is not None:
        is_zero = abs(sched_variance) < 0.02
        c_sv = ws.cell(row=row, column=_G, value=sched_variance if not is_zero else 0)
        _apply(c_sv,
               font=_font(bold=True, color='006100' if is_zero else '9C0006'),
               fmt='#,##0.00;(#,##0.00);"-"',
               fill=_fill(GREEN_FILL if is_zero else RED_FILL),
               border=DOUBLE_BTM)
    row += 2

    # ── TB tie-out (below GL section) ─────────────────────────────
    tb_ending = tb_acct.ending_balance if tb_acct else None
    variance  = (gl_ending - tb_ending) if tb_ending is not None else None

    ws.cell(row=row, column=_H, value='TB Balance').font = _font(bold=True)
    if tb_ending is not None:
        c_tb = ws.cell(row=row, column=_I, value=tb_ending)
        _apply(c_tb, font=_font(bold=True), fmt='#,##0.00;(#,##0.00);"-"',
               fill=_fill(LIGHT_BLUE), border=THIN)
    else:
        c_tb = ws.cell(row=row, column=_I, value='Not in TB')
        c_tb.font = _font(italic=True, color='888888')
    row += 1

    ws.cell(row=row, column=_H, value='Variance').font = _font(bold=True)
    if variance is not None:
        is_zero = abs(variance) < 0.02
        c_var = ws.cell(row=row, column=_I, value=variance)
        _apply(c_var,
               font=_font(bold=True, color='006100' if is_zero else '9C0006'),
               fmt='#,##0.00;(#,##0.00);"-"',
               fill=_fill(GREEN_FILL if is_zero else RED_FILL),
               border=DOUBLE_BTM)
    else:
        ws.cell(row=row, column=_I, value='').border = DOUBLE_BTM

    # ── Historical rollforward (below tie-out) ────────────────────────────────
    if history_rows:
        row += 3
        c = ws.cell(row=row, column=_B, value='Historical GL vs TB Rollforward')
        c.font = _font(bold=True, color='FFFFFF')
        c.fill = _fill(DARK_BLUE)
        ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
        row += 1

        hist_hdrs = ['Period', 'Beg Balance', 'Net Activity', 'GL Ending', 'TB Ending', 'Variance']
        hist_wids = [12, 18, 16, 18, 18, 14]
        for ci, (h, w) in enumerate(zip(hist_hdrs, hist_wids)):
            col = _B + ci
            c = ws.cell(row=row, column=col, value=h)
            _apply(c, font=_hdr_font(), fill=_fill(MED_BLUE), border=THIN,
                   align=Alignment(horizontal='center', wrap_text=True))
        row += 1

        for i, hist in enumerate(history_rows):
            alt_fill = _fill(LIGHT_GRAY) if i % 2 == 1 else None
            _var   = hist.get('variance', 0.0) or 0.0
            _vzero = abs(_var) < 0.02
            row_data = [
                (_B,     hist.get('period', ''),      None),
                (_B + 1, hist.get('beg_bal', 0.0),    '#,##0.00;(#,##0.00);"-"'),
                (_B + 2, hist.get('net_change', 0.0), '#,##0.00;(#,##0.00);"-"'),
                (_B + 3, hist.get('gl_end', 0.0),     '#,##0.00;(#,##0.00);"-"'),
                (_B + 4, hist.get('tb_end', 0.0),     '#,##0.00;(#,##0.00);"-"'),
                (_B + 5, _var,                         '#,##0.00;(#,##0.00);"-"'),
            ]
            for col, val, fmt in row_data:
                c = ws.cell(row=row, column=col, value=val)
                if fmt:
                    c.number_format = fmt
                if alt_fill:
                    c.fill = alt_fill
                c.border = THIN
            vc = ws.cell(row=row, column=_B + 5)
            vc.fill = _fill(GREEN_FILL) if _vzero else _fill(RED_FILL)
            vc.font = _font(color='006100' if _vzero else '9C0006')
            row += 1

        # Current period row in blue
        cur_data = [
            (_B,     period,            None),
            (_B + 1, gl_acct.beginning_balance, '#,##0.00;(#,##0.00);"-"'),
            (_B + 2, gl_acct.net_change,         '#,##0.00;(#,##0.00);"-"'),
            (_B + 3, gl_ending,                  '#,##0.00;(#,##0.00);"-"'),
            (_B + 4, tb_ending if tb_ending is not None else '', '#,##0.00;(#,##0.00);"-"' if tb_ending is not None else None),
            (_B + 5, variance  if variance  is not None else '', '#,##0.00;(#,##0.00);"-"' if variance  is not None else None),
        ]
        for col, val, fmt in cur_data:
            c = ws.cell(row=row, column=col, value=val)
            if fmt:
                c.number_format = fmt
            c.fill = _fill(LIGHT_BLUE)
            c.font = _font(bold=True)
            c.border = THIN
        if variance is not None:
            _vzero = abs(variance) < 0.02
            vc = ws.cell(row=row, column=_B + 5)
            vc.fill = _fill(GREEN_FILL) if _vzero else _fill(RED_FILL)
            vc.font = _font(bold=True, color='006100' if _vzero else '9C0006')

    ws.freeze_panes = 'B5'


# ── Stub tab for zero-activity BS accounts ───────────────────

def _write_stub_tab(wb, tb_acct, period: str, property_name: str,
                    tab_prefix: str = '',
                    history_rows: list = None,
                    prepared_by: str = ''):
    """
    Stub tab for a BS account in the TB with no current-period GL transactions.
    Uses the same rolling-table format as _write_account_tab.
    Current-period row: net activity = 0, GL ending = TB forward balance.
    """
    acct_label = _safe_sheet_name(f'{tb_acct.account_code} {tb_acct.account_name}')
    ws = wb.create_sheet(acct_label)
    ws.sheet_properties.tabColor = COLOR_BS_STD
    ws.column_dimensions['A'].width = 2

    row = 1
    c = ws.cell(row=row, column=_B,
                value=f'{tb_acct.account_code} — {tb_acct.account_name}')
    c.font = _font(bold=True, size=13, color='FFFFFF')
    c.fill = _fill(DARK_BLUE)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
    row += 1

    c = ws.cell(row=row, column=_B,
                value=f'{property_name or "Revolution Labs"}  |  '
                      f'Prepared by: {prepared_by or "GRP"}  |  '
                      f'{datetime.now().strftime("%m/%d/%Y")}')
    c.font = _font(italic=True, color='FFFFFF')
    c.fill = _fill(MED_BLUE)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
    row += 2

    tbl_headers = ['Period', 'Beg Balance', 'Net Activity', 'GL Ending', 'TB Ending', 'Variance']
    tbl_widths  = [12, 18, 16, 18, 18, 14]
    for ci, (h, w) in enumerate(zip(tbl_headers, tbl_widths)):
        col = _B + ci
        c = ws.cell(row=row, column=col, value=h)
        _apply(c, font=_hdr_font(), fill=_fill(DARK_BLUE), border=THIN,
               align=Alignment(horizontal='center', wrap_text=True))
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 24
    row += 1

    for i, hist in enumerate(history_rows or []):
        alt_fill = _fill(LIGHT_GRAY) if i % 2 == 1 else None
        _var   = hist.get('variance', 0.0) or 0.0
        _vzero = abs(_var) < 0.02
        row_data = [
            (_B,     hist.get('period', ''),      None),
            (_B + 1, hist.get('beg_bal', 0.0),    '#,##0.00;(#,##0.00);"-"'),
            (_B + 2, hist.get('net_change', 0.0), '#,##0.00;(#,##0.00);"-"'),
            (_B + 3, hist.get('gl_end', 0.0),     '#,##0.00;(#,##0.00);"-"'),
            (_B + 4, hist.get('tb_end', 0.0),     '#,##0.00;(#,##0.00);"-"'),
            (_B + 5, _var,                         '#,##0.00;(#,##0.00);"-"'),
        ]
        for col, val, fmt in row_data:
            c = ws.cell(row=row, column=col, value=val)
            if fmt:
                c.number_format = fmt
            if alt_fill:
                c.fill = alt_fill
            c.border = THIN
        vc = ws.cell(row=row, column=_B + 5)
        vc.fill = _fill(GREEN_FILL) if _vzero else _fill(RED_FILL)
        vc.font = _font(color='006100' if _vzero else '9C0006')
        row += 1

    # Current period: no GL activity → ending = forward balance from TB
    fwd      = getattr(tb_acct, 'forward_balance', None) or tb_acct.ending_balance
    gl_end   = fwd
    tb_end   = tb_acct.ending_balance
    variance = round(gl_end - tb_end, 2)
    _vzero   = abs(variance) < 0.02

    cur_data = [
        (_B,     period,  None),
        (_B + 1, fwd,     '#,##0.00;(#,##0.00);"-"'),
        (_B + 2, 0.0,     '#,##0.00;(#,##0.00);"-"'),
        (_B + 3, gl_end,  '#,##0.00;(#,##0.00);"-"'),
        (_B + 4, tb_end,  '#,##0.00;(#,##0.00);"-"'),
        (_B + 5, variance,'#,##0.00;(#,##0.00);"-"'),
    ]
    for col, val, fmt in cur_data:
        c = ws.cell(row=row, column=col, value=val)
        if fmt:
            c.number_format = fmt
        c.fill = _fill(LIGHT_BLUE)
        c.font = _font(bold=True)
        c.border = THIN
    vc = ws.cell(row=row, column=_B + 5)
    vc.fill = _fill(GREEN_FILL) if _vzero else _fill(RED_FILL)
    vc.font = _font(bold=True, color='006100' if _vzero else '9C0006')

    ws.freeze_panes = 'B5'


# ── Prepaid amortization schedule tab ────────────────────────

def _write_prepaid_schedule_tab(wb, active_items: list, period: str, property_name: str,
                                 tab_prefix: str = '', gl_result=None):
    """
    Adds a 'Prepaid Schedule' tab using the Hartwell 13-column format.
    Tied to accounts 135xxx.

    Columns _B.._N:
      _B  Description
      _C  G/L Account
      _D  Payment Date
      _E  Payment Amount
      _F  Period Covered
      _G  Start Date
      _H  End Date
      _I  # of Mos. Covered
      _J  Exp per Month
      _K  Months Elapsed
      _L  # of Mos Prepaid
      _M  Prepaid Balance
      _N  Expense Balance
    """
    COLOR_PREPAID = 'ED7D31'   # orange tab — matches prepaid ledger convention

    # Extended column constants for this tab
    _J = 10
    _K = 11
    _L = 12
    _M = 13
    _N = 14

    _tab_name = (tab_prefix + 'Prepaid Schedule')[:31]
    ws = wb.create_sheet(_tab_name)
    ws.sheet_properties.tabColor = COLOR_PREPAID

    # Blank col A — narrow
    ws.column_dimensions['A'].width = 2

    row = 1
    c = ws.cell(row=row, column=_B,
                value=f'{property_name or "Revolution Labs"} — Prepaid Expense Schedule')
    c.font = _font(bold=True, size=13, color='FFFFFF')
    c.fill = _fill(COLOR_PREPAID)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_N)
    row += 1

    c = ws.cell(row=row, column=_B,
                value=f'Period: {period}  |  Active items as of close  |  Account 135xxx — Prepaid Expenses')
    c.font = _font(italic=True, color='FFFFFF')
    c.fill = _fill(COLOR_PREPAID)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_N)
    row += 2

    # Column headers
    headers = [
        'Description',       # _B
        'G/L Account',       # _C
        'Payment Date',      # _D
        'Payment Amount',    # _E
        'Period Covered',    # _F
        'Start Date',        # _G
        'End Date',          # _H
        '# of Mos. Covered', # _I
        'Exp per Month',     # _J
        'Months Elapsed',    # _K
        '# of Mos Prepaid',  # _L
        'Prepaid Balance',   # _M
        'Expense Balance',   # _N
    ]
    widths = [32, 13, 13, 16, 20, 13, 13, 15, 15, 15, 15, 16, 16]
    for ci, (h, w) in enumerate(zip(headers, widths)):
        col = _B + ci
        c = ws.cell(row=row, column=col, value=h)
        _apply(c, font=_hdr_font(), fill=_fill(DARK_BLUE), border=THIN,
               align=Alignment(horizontal='center', wrap_text=True))
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 28
    row += 1

    total_prepaid  = 0.0
    total_expense  = 0.0
    total_monthly  = 0.0   # sum of monthly_amount for expense tie-out

    data_start_row = row   # first data row (for Excel SUM formulas)

    for i, item in enumerate(active_items):
        alt_fill = _fill(LIGHT_GRAY) if i % 2 == 1 else None

        months_elapsed  = int(item.get('months_amortized', 0) or 0)
        months_prepaid  = int(item.get('remaining_months', 0) or 0)
        exp_per_month   = float(item.get('monthly_amount', 0) or 0)
        prepaid_balance = exp_per_month * months_prepaid
        expense_balance = exp_per_month * months_elapsed
        total_prepaid  += prepaid_balance
        total_expense  += expense_balance
        total_monthly  += exp_per_month

        total_months = int(item.get('total_months', 0) or 0)
        total_amount = float(item.get('total_amount', 0) or 0)

        # Payment date
        svc_start = item.get('service_start', '')
        if svc_start and hasattr(svc_start, 'strftime'):
            pay_date_str = svc_start.strftime('%m/%d/%Y')
        else:
            pay_date_str = str(svc_start) if svc_start else ''

        # Start / End date strings
        svc_end = item.get('service_end', '')
        if svc_end and hasattr(svc_end, 'strftime'):
            end_date_str = svc_end.strftime('%m/%d/%Y')
        else:
            end_date_str = str(svc_end) if svc_end else ''

        if svc_start and hasattr(svc_start, 'strftime'):
            start_date_str = svc_start.strftime('%m/%d/%Y')
        else:
            start_date_str = str(svc_start) if svc_start else ''

        # Period covered — use field if set, else format from start/end
        period_covered = item.get('period_covered', '')
        if not period_covered and svc_start and svc_end:
            try:
                if hasattr(svc_start, 'strftime') and hasattr(svc_end, 'strftime'):
                    period_covered = (f'{svc_start.strftime("%m.%d.%y")} - '
                                      f'{svc_end.strftime("%m.%d.%y")}')
            except Exception:
                period_covered = ''

        description = item.get('description', '') or item.get('vendor', '')

        row_vals = [
            description,                         # _B
            item.get('gl_account_number', ''),   # _C
            pay_date_str,                         # _D  Payment Date
            total_amount,                         # _E  Payment Amount
            period_covered,                       # _F  Period Covered
            start_date_str,                       # _G  Start Date
            end_date_str,                         # _H  End Date
            total_months,                         # _I  # of Mos. Covered
            exp_per_month,                        # _J  Exp per Month
            months_elapsed,                       # _K  Months Elapsed
            months_prepaid,                       # _L  # of Mos Prepaid
            prepaid_balance,                      # _M  Prepaid Balance
            expense_balance,                      # _N  Expense Balance
        ]

        for ci, val in enumerate(row_vals):
            col = _B + ci
            c = ws.cell(row=row, column=col, value=val)
            c.border = THIN
            if alt_fill:
                c.fill = alt_fill
            # Number formats
            if col == _E:   # Payment Amount
                c.number_format = '#,##0.00;(#,##0.00);"-"'
            elif col in (_J, _M, _N):  # Exp per Month, Prepaid Balance, Expense Balance
                c.number_format = '#,##0.00;(#,##0.00);"-"'
            elif col in (_I, _K, _L):  # integer month counts
                c.alignment = Alignment(horizontal='center')
                # Color-code months prepaid (_L)
                if col == _L:
                    if months_prepaid == 0:
                        c.font = _font(color='FF0000', bold=True)
                    elif months_prepaid == 1:
                        c.font = _font(color='C55A11', bold=True)
        row += 1

    data_end_row = row - 1   # last item data row (for SUM formulas)

    # ── GL lookup for tie-out ──────────────────────────────────────────────────
    # Prepaid BS balance: sum ending_balance for all 135xxx accounts in GL
    # Period expense:     sum net PTD debit for expense accounts in the schedule
    _135_gl_balance  = 0.0
    _expense_gl_ptd  = 0.0
    _expense_gl_ytd  = 0.0   # ending_balance for P&L accounts = YTD since fiscal year resets Jan
    _expense_accts   = {
        str(item.get('gl_account_number', '')).strip()
        for item in active_items
        if item.get('gl_account_number')
    }
    if gl_result:
        for _ga in (gl_result.accounts if hasattr(gl_result, 'accounts') else []):
            _code = str(_ga.account_code)
            if _code.startswith('135'):
                _135_gl_balance += _safe_float(_ga.ending_balance)
            if _code in _expense_accts:
                # PTD: net debit activity this period only
                _expense_gl_ptd += (_safe_float(_ga.total_debits)
                                    - _safe_float(_ga.total_credits))
                # YTD: ending_balance accumulates from Jan (P&L resets each fiscal year)
                _expense_gl_ytd += _safe_float(_ga.ending_balance)

    # Column letters for Excel SUM formulas
    _col_J_ltr = get_column_letter(_J)   # 'J' — Exp per Month
    _col_M_ltr = get_column_letter(_M)   # 'M' — Prepaid Balance
    _col_N_ltr = get_column_letter(_N)   # 'N' — Expense Balance (cumulative)

    # ── Footer tie-out rows ────────────────────────────────────────────────────
    row += 1  # blank separator

    # ── Section 1: Prepaid Balance Tie-out (135xxx) ────────────────────────────
    # Row A: Total prepaid balance per schedule
    r_sched_prepaid = row
    _lbl = ws.cell(row=row, column=_B, value='Prepaid Balance per Schedule')
    _lbl.font = _font(bold=True)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_L)
    # Excel SUM formula — references actual data cells so manual edits flow through
    c_sched_pre = ws.cell(
        row=row, column=_M,
        value=(f'=SUM({_col_M_ltr}{data_start_row}:{_col_M_ltr}{data_end_row})'
               if data_end_row >= data_start_row else total_prepaid)
    )
    _apply(c_sched_pre, font=_font(bold=True), fmt='#,##0.00;(#,##0.00);"-"',
           fill=_fill(LIGHT_BLUE), border=THICK_BOTTOM)
    for col in range(_B, _N + 1):
        _c = ws.cell(row=row, column=col)
        if not _c.fill or _c.fill.fill_type == 'none':
            _c.fill = _fill(LIGHT_BLUE)
    row += 1

    # Row B: GL ending balance for 135xxx accounts
    r_gl_prepaid = row
    _lbl2 = ws.cell(row=row, column=_B,
                    value='GL Ending Balance — Accounts 135xxx')
    _lbl2.font = _font(bold=True)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_L)
    c_gl_pre = ws.cell(row=row, column=_M,
                       value=_135_gl_balance if (_135_gl_balance or gl_result) else None)
    _apply(c_gl_pre, font=_font(bold=True), fmt='#,##0.00;(#,##0.00);"-"',
           fill=_fill(LIGHT_BLUE), border=THIN)
    for col in range(_B, _N + 1):
        _c = ws.cell(row=row, column=col)
        if not _c.fill or _c.fill.fill_type == 'none':
            _c.fill = _fill(LIGHT_BLUE)
    row += 1

    # Row C: Variance — Prepaid (green if zero)
    _lbl3 = ws.cell(row=row, column=_B, value='Variance')
    _lbl3.font = _font(bold=True)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_L)
    c_var_pre = ws.cell(
        row=row, column=_M,
        value=f'={_col_M_ltr}{r_sched_prepaid}-{_col_M_ltr}{r_gl_prepaid}'
    )
    _pre_var_val  = total_prepaid - _135_gl_balance
    _is_zero_pre  = abs(_pre_var_val) < 0.02
    _apply(c_var_pre,
           font=_font(bold=True, color='006100' if _is_zero_pre else '9C0006'),
           fmt='#,##0.00;(#,##0.00);"-"',
           fill=_fill(GREEN_FILL) if _is_zero_pre else _fill(RED_FILL),
           border=DOUBLE_BTM)
    row += 2

    # ── Section 2: Period Expense Tie-out ──────────────────────────────────────
    # Row D: Period expense per schedule (sum of Exp per Month column)
    r_sched_exp = row
    _exp_acct_lbl = ', '.join(sorted(_expense_accts)) if _expense_accts else 'n/a'
    _lbl4 = ws.cell(row=row, column=_B,
                    value=f'Period Expense per Schedule — Accts {_exp_acct_lbl}')
    _lbl4.font = _font(bold=True)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_I)
    c_sched_exp = ws.cell(
        row=row, column=_J,
        value=(f'=SUM({_col_J_ltr}{data_start_row}:{_col_J_ltr}{data_end_row})'
               if data_end_row >= data_start_row else total_monthly)
    )
    _apply(c_sched_exp, font=_font(bold=True), fmt='#,##0.00;(#,##0.00);"-"',
           fill=_fill(LIGHT_BLUE), border=THICK_BOTTOM)
    for col in range(_B, _L + 1):
        _c = ws.cell(row=row, column=col)
        if not _c.fill or _c.fill.fill_type == 'none':
            _c.fill = _fill(LIGHT_BLUE)
    row += 1

    # Row E: GL PTD activity for those expense accounts
    r_gl_exp = row
    _lbl5 = ws.cell(row=row, column=_B,
                    value='PTD Activity per GL — Expense Accounts')
    _lbl5.font = _font(bold=True)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_I)
    c_gl_exp = ws.cell(row=row, column=_J,
                       value=_expense_gl_ptd if (_expense_gl_ptd or gl_result) else None)
    _apply(c_gl_exp, font=_font(bold=True), fmt='#,##0.00;(#,##0.00);"-"',
           fill=_fill(LIGHT_BLUE), border=THIN)
    for col in range(_B, _L + 1):
        _c = ws.cell(row=row, column=col)
        if not _c.fill or _c.fill.fill_type == 'none':
            _c.fill = _fill(LIGHT_BLUE)
    row += 1

    # Row F: Variance — Expense (green if zero)
    _lbl6 = ws.cell(row=row, column=_B, value='Variance')
    _lbl6.font = _font(bold=True)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_I)
    c_var_exp = ws.cell(
        row=row, column=_J,
        value=f'={_col_J_ltr}{r_sched_exp}-{_col_J_ltr}{r_gl_exp}'
    )
    _exp_var_val  = total_monthly - _expense_gl_ptd
    _is_zero_exp  = abs(_exp_var_val) < 0.02
    _apply(c_var_exp,
           font=_font(bold=True, color='006100' if _is_zero_exp else '9C0006'),
           fmt='#,##0.00;(#,##0.00);"-"',
           fill=_fill(GREEN_FILL) if _is_zero_exp else _fill(RED_FILL),
           border=DOUBLE_BTM)
    row += 2

    # ── Section 3: Cumulative Expense Tie-out (YTD) ────────────────────────────
    # Row G: Cumulative expense per schedule (sum of col _N = exp_per_month × months_elapsed)
    r_sched_ytd = row
    _lbl7 = ws.cell(row=row, column=_B,
                    value=f'Cumulative Expense per Schedule (YTD) — Accts {_exp_acct_lbl}')
    _lbl7.font = _font(bold=True)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_M)
    c_sched_ytd = ws.cell(
        row=row, column=_N,
        value=(f'=SUM({_col_N_ltr}{data_start_row}:{_col_N_ltr}{data_end_row})'
               if data_end_row >= data_start_row else total_expense)
    )
    _apply(c_sched_ytd, font=_font(bold=True), fmt='#,##0.00;(#,##0.00);"-"',
           fill=_fill(LIGHT_BLUE), border=THICK_BOTTOM)
    for col in range(_B, _N + 1):
        _c = ws.cell(row=row, column=col)
        if not _c.fill or _c.fill.fill_type == 'none':
            _c.fill = _fill(LIGHT_BLUE)
    row += 1

    # Row H: GL YTD ending balance for expense accounts (P&L resets Jan → ending_balance = YTD)
    r_gl_ytd = row
    _lbl8 = ws.cell(row=row, column=_B,
                    value='GL YTD Ending Balance — Expense Accounts')
    _lbl8.font = _font(bold=True)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_M)
    c_gl_ytd = ws.cell(row=row, column=_N,
                       value=_expense_gl_ytd if (_expense_gl_ytd or gl_result) else None)
    _apply(c_gl_ytd, font=_font(bold=True), fmt='#,##0.00;(#,##0.00);"-"',
           fill=_fill(LIGHT_BLUE), border=THIN)
    for col in range(_B, _N + 1):
        _c = ws.cell(row=row, column=col)
        if not _c.fill or _c.fill.fill_type == 'none':
            _c.fill = _fill(LIGHT_BLUE)
    row += 1

    # Row I: Variance — YTD (green if zero)
    _lbl9 = ws.cell(row=row, column=_B, value='Variance')
    _lbl9.font = _font(bold=True)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_M)
    c_var_ytd = ws.cell(
        row=row, column=_N,
        value=f'={_col_N_ltr}{r_sched_ytd}-{_col_N_ltr}{r_gl_ytd}'
    )
    _ytd_var_val = total_expense - _expense_gl_ytd
    _is_zero_ytd = abs(_ytd_var_val) < 0.02
    _apply(c_var_ytd,
           font=_font(bold=True, color='006100' if _is_zero_ytd else '9C0006'),
           fmt='#,##0.00;(#,##0.00);"-"',
           fill=_fill(GREEN_FILL) if _is_zero_ytd else _fill(RED_FILL),
           border=DOUBLE_BTM)
    row += 2

    # "[Add Row]" placeholder for manual additions
    c_add = ws.cell(row=row, column=_B, value='[Add Row]')
    c_add.font = _font(italic=True, color='888888')
    c_add.border = THIN
    for col in range(_C, _N + 1):
        ws.cell(row=row, column=col).border = THIN
    row += 2

    note = ws.cell(row=row, column=_B,
                   value='Prepaid Balance = Exp per Month × # Mos Prepaid (should agree to 135xxx in TB).  '
                         'Period Expense = monthly amortization amount (should agree to GL PTD for expense accts).  '
                         'GL balances populated from Pass 2 GL upload.')
    note.font = _font(italic=True, size=10, color='595959')
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_N)
    ws.row_dimensions[row].height = 28

    ws.freeze_panes = 'B5'


# ── Bank Rec tab ─────────────────────────────────────────────

COLOR_BANK_REC = '375623'   # dark green tab

def _write_bank_rec_tab(wb, bank_rec_data: dict, gl_acct_balance: float,
                        period: str, property_name: str,
                        account_label: str = 'PNC Operating (x3993)',
                        gl_account_code: str = '111100',
                        tab_prefix: str = '',
                        tab_name_override: str = None,
                        prepared_by: str = ''):
    """
    Writes one Bank Rec tab showing:
      Balance per Bank Statement
      Less: Outstanding Checks
      = Reconciled Bank Balance  →  must equal GL cash account
    Then lists outstanding checks and cleared checks for reference.
    """
    if tab_name_override:
        _base_name = tab_name_override
    else:
        _base_name = f'Bank Rec - {account_label.split("(")[0].strip()[:20]}'
    tab_name = (tab_prefix + _base_name)[:31]
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = COLOR_BANK_REC

    # Blank col A — narrow
    ws.column_dimensions['A'].width = 2

    row = 1
    # Header
    c = ws.cell(row=row, column=_B,
                value=f'{property_name or "Revolution Labs"} — Bank Reconciliation')
    c.font = _font(bold=True, size=13, color='FFFFFF')
    c.fill = _fill(COLOR_BANK_REC)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
    row += 1

    c = ws.cell(row=row, column=_B,
                value=f'Account: {account_label}  |  Period: {period}  |  '
                      f'Prepared by: {prepared_by or "GRP"}  |  {datetime.now().strftime("%m/%d/%Y")}')
    c.font = _font(italic=True, color='FFFFFF')
    c.fill = _fill(COLOR_BANK_REC)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
    row += 2

    # Column widths
    for ci, w in enumerate([18, 15, 45, 18, 6, 6]):
        ws.column_dimensions[get_column_letter(_B + ci)].width = w

    # ── Reconciliation Summary ────────────────────────────────
    bank_bal    = float(bank_rec_data.get('bank_statement_balance', 0) or 0)
    out_total   = float(bank_rec_data.get('total_outstanding_checks', 0) or 0)
    rec_bal     = float(bank_rec_data.get('reconciled_bank_balance', 0) or 0)
    difference  = rec_bal - gl_acct_balance

    def _rec_row(label, value, bold=False, fill_hex=None, border=THIN, fmt='#,##0.00;(#,##0.00);"-"'):
        nonlocal row
        c_lbl = ws.cell(row=row, column=_B, value=label)
        c_lbl.font = _font(bold=bold)
        c_lbl.alignment = Alignment(horizontal='right')
        ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 3)
        c_val = ws.cell(row=row, column=_B + 4, value=value)
        _apply(c_val, font=_font(bold=bold), fmt=fmt, border=border)
        ws.merge_cells(start_row=row, start_column=_B + 4, end_row=row, end_column=_B + 5)
        if fill_hex:
            c_val.fill = _fill(fill_hex)
        row += 1

    _rec_row('Balance Per Bank Statement:', bank_bal)
    _rec_row(f'  Less: Outstanding Checks:', -out_total)
    ws.cell(row=row - 1, column=_B + 4).border = THICK_BOTTOM
    _rec_row('Reconciled Bank Balance:', rec_bal, bold=True, fill_hex=LIGHT_BLUE, border=DOUBLE_BTM)
    row += 1
    _rec_row(f'Balance per GL — {gl_account_code}:', gl_acct_balance, bold=True, fill_hex=LIGHT_BLUE)

    # Variance row
    is_clean = abs(difference) < 0.02
    var_fill  = GREEN_FILL if is_clean else RED_FILL
    var_color = '006100' if is_clean else '9C0006'
    c_lbl = ws.cell(row=row, column=_B, value='Difference:')
    c_lbl.font = _font(bold=True, color=var_color)
    c_lbl.alignment = Alignment(horizontal='right')
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 3)
    c_val = ws.cell(row=row, column=_B + 4, value=difference)
    _apply(c_val, font=_font(bold=True, color=var_color),
           fmt='#,##0.00;(#,##0.00);"-"', fill=_fill(var_fill), border=DOUBLE_BTM)
    ws.merge_cells(start_row=row, start_column=_B + 4, end_row=row, end_column=_B + 5)
    row += 2

    if not is_clean:
        note = ws.cell(row=row, column=_B,
                       value=f'Reconciling difference of ${abs(difference):,.2f} — investigate before close.')
        note.font = _font(italic=True, color='9C0006', size=10)
        note.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
        row += 2

    # ── Outstanding Checks ────────────────────────────────────
    outstanding = bank_rec_data.get('outstanding_checks', [])
    if outstanding:
        c = ws.cell(row=row, column=_B, value='Outstanding Checks')
        c.font = _font(bold=True, size=12, color='FFFFFF')
        c.fill = _fill(COLOR_BANK_REC)
        ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
        row += 1

        hdrs = ['Check Date', 'Check #', 'Payee', 'Amount', '', '']
        for ci, h in enumerate(hdrs[:4]):
            c = ws.cell(row=row, column=_B + ci, value=h)
            _apply(c, font=_hdr_font(), fill=_fill(DARK_BLUE), border=THIN,
                   align=Alignment(horizontal='center'))
        row += 1

        for i, chk in enumerate(outstanding):
            payee = str(chk.get('payee', '')).split(' - ', 1)[-1]  # strip vendor code prefix
            alt   = _fill(LIGHT_GRAY) if i % 2 == 1 else None
            ws.cell(row=row, column=_B, value=chk.get('date', '')).border = THIN
            ws.cell(row=row, column=_C, value=str(chk.get('check_number', ''))).border = THIN
            ws.cell(row=row, column=_D, value=payee).border = THIN
            c_amt = ws.cell(row=row, column=_E, value=float(chk.get('amount', 0)))
            _apply(c_amt, fmt='#,##0.00', border=THIN)
            if alt:
                for col in range(_B, _B + 4):
                    ws.cell(row=row, column=col).fill = alt
            row += 1

        # Outstanding total
        ws.cell(row=row, column=_D, value='Total Outstanding Checks').font = _font(bold=True)
        c_tot = ws.cell(row=row, column=_E, value=out_total)
        _apply(c_tot, font=_font(bold=True), fmt='#,##0.00', fill=_fill(LIGHT_BLUE), border=DOUBLE_BTM)
        row += 2

    # ── Cleared Checks (reference) ────────────────────────────
    cleared = bank_rec_data.get('cleared_checks', [])
    if cleared:
        c = ws.cell(row=row, column=_B, value='Cleared Checks — Reference')
        c.font = _font(bold=True, size=11, color='595959')
        c.fill = _fill(LIGHT_GRAY)
        ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
        row += 1

        hdrs = ['Date', 'Tran #', 'Payee / Notes', 'Amount', 'Date Cleared', '']
        for ci, h in enumerate(hdrs[:5]):
            c = ws.cell(row=row, column=_B + ci, value=h)
            _apply(c, font=_font(bold=True, color='595959'), fill=_fill(LIGHT_GRAY),
                   border=THIN, align=Alignment(horizontal='center'))
        row += 1

        cleared_total = 0.0
        for i, chk in enumerate(cleared):
            payee = str(chk.get('notes', chk.get('payee', ''))).split(' - ', 1)[-1]
            amt   = float(chk.get('amount', 0))
            cleared_total += amt
            alt   = _fill('F9F9F9') if i % 2 == 1 else None
            ws.cell(row=row, column=_B, value=chk.get('date', '')).border = THIN
            ws.cell(row=row, column=_C, value=str(chk.get('tran_number', chk.get('check_number', '')))).border = THIN
            ws.cell(row=row, column=_D, value=payee).border = THIN
            c_amt = ws.cell(row=row, column=_E, value=amt)
            _apply(c_amt, fmt='#,##0.00', border=THIN)
            ws.cell(row=row, column=_F, value=chk.get('date_cleared', '')).border = THIN
            if alt:
                for col in range(_B, _B + 5):
                    ws.cell(row=row, column=col).fill = alt
            row += 1

        ws.cell(row=row, column=_D, value='Total Cleared Checks').font = _font(bold=True, color='595959')
        c_tot = ws.cell(row=row, column=_E, value=cleared_total)
        _apply(c_tot, font=_font(bold=True, color='595959'), fmt='#,##0.00',
               fill=_fill(LIGHT_GRAY), border=DOUBLE_BTM)

    ws.freeze_panes = 'B4'


# ── DACA Bank Rec tab ────────────────────────────────────────

def _write_daca_bank_rec_tab(wb, daca_bank_data: dict, gl_daca_balance: float,
                              period: str, property_name: str,
                              tab_prefix: str = '',
                              prepared_by: str = ''):
    """
    Writes the DACA Bank Rec tab for KeyBank x5132 (GL account 115100).

    DACA accounts are sweep accounts — deposits are collected here and swept
    daily to PNC Operating.  There are typically no outstanding checks;
    the reconciliation is simply:

        Bank Statement Ending Balance
        = GL Account 115100 Ending Balance
        Difference (should be $0.00)

    The tab also shows:
      - Statement period and account info
      - Beginning → Ending balance from bank statement
      - Full transaction detail if available (sweeps, deposits)
    """
    COLOR_DACA = '375623'   # same dark green family as Operating rec

    _tab_name = (tab_prefix + 'Bank Rec - DACA')[:31]
    ws = wb.create_sheet(_tab_name)
    ws.sheet_properties.tabColor = COLOR_DACA

    # Blank col A — narrow
    ws.column_dimensions['A'].width = 2

    ending_bal = float(daca_bank_data.get('ending_balance') or 0)
    beginning_bal = float(daca_bank_data.get('beginning_balance') or 0)
    acct_num = daca_bank_data.get('account_number') or 'x5132'
    period_info = daca_bank_data.get('statement_period') or {}
    parse_error = daca_bank_data.get('_parse_error')

    row = 1
    # Header
    c = ws.cell(row=row, column=_B,
                value=f'{property_name or "Revolution Labs"} — Bank Reconciliation (DACA)')
    c.font = _font(bold=True, size=13, color='FFFFFF')
    c.fill = _fill(COLOR_DACA)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
    row += 1

    period_str = f'{period_info.get("start", "")} — {period_info.get("end", "")}' if period_info else period
    c = ws.cell(row=row, column=_B,
                value=f'Account: KeyBank DACA (x{acct_num.lstrip("x")})  |  '
                      f'Period: {period_str}  |  GL Account: 115100  |  '
                      f'Prepared by: {prepared_by or "GRP"}  |  {datetime.now().strftime("%m/%d/%Y")}')
    c.font = _font(italic=True, color='FFFFFF')
    c.fill = _fill(COLOR_DACA)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
    row += 2

    # Column widths
    for ci, w in enumerate([22, 15, 42, 18, 6, 6]):
        ws.column_dimensions[get_column_letter(_B + ci)].width = w

    # Parse error warning
    if parse_error:
        c = ws.cell(row=row, column=_B,
                    value=f'⚠  Parser note: {parse_error} — verify balances below manually')
        c.font = _font(italic=True, color='9C0006', size=10)
        c.fill = _fill(AMBER_FILL)
        ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
        row += 2

    # ── Reconciliation block ──────────────────────────────────
    def _daca_row(label, value, bold=False, fill_hex=None, border=THIN,
                  fmt='#,##0.00;(#,##0.00);"-"'):
        nonlocal row
        c_lbl = ws.cell(row=row, column=_B, value=label)
        c_lbl.font = _font(bold=bold)
        c_lbl.alignment = Alignment(horizontal='right')
        ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 3)
        c_val = ws.cell(row=row, column=_B + 4, value=value)
        _apply(c_val, font=_font(bold=bold), fmt=fmt, border=border)
        ws.merge_cells(start_row=row, start_column=_B + 4, end_row=row, end_column=_B + 5)
        if fill_hex:
            c_val.fill = _fill(fill_hex)
        row += 1

    _daca_row('Beginning Balance per Bank Statement:', beginning_bal)
    _daca_row('Ending Balance per Bank Statement:', ending_bal, bold=True,
              fill_hex=LIGHT_BLUE, border=DOUBLE_BTM)
    row += 1
    _daca_row('Balance per GL — Account 115100:', gl_daca_balance, bold=True,
              fill_hex=LIGHT_BLUE)

    # Difference
    difference = ending_bal - gl_daca_balance
    is_clean   = abs(difference) < 0.02
    var_fill   = GREEN_FILL if is_clean else RED_FILL
    var_color  = '006100' if is_clean else '9C0006'

    c_lbl = ws.cell(row=row, column=_B, value='Difference:')
    c_lbl.font = _font(bold=True, color=var_color)
    c_lbl.alignment = Alignment(horizontal='right')
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 3)
    c_val = ws.cell(row=row, column=_B + 4, value=difference)
    _apply(c_val, font=_font(bold=True, color=var_color),
           fmt='#,##0.00;(#,##0.00);"-"', fill=_fill(var_fill), border=DOUBLE_BTM)
    ws.merge_cells(start_row=row, start_column=_B + 4, end_row=row, end_column=_B + 5)
    row += 2

    if not is_clean:
        note = ws.cell(row=row, column=_B,
                       value=f'Reconciling difference of ${abs(difference):,.2f} — '
                             f'investigate before close. DACA account should sweep to zero daily.')
        note.font = _font(italic=True, color='9C0006', size=10)
        note.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
        ws.row_dimensions[row].height = 28
        row += 2

    # ── Account Note ──────────────────────────────────────────
    note2 = ws.cell(row=row, column=_B,
                    value='Note: This is a Deposit Account Control Agreement (DACA) — a sweep account. '
                          'Tenant rent deposits collect here and are swept daily to PNC Operating (x3993). '
                          'No outstanding checks are expected. Month-end balance should be minimal.')
    note2.font = _font(italic=True, size=10, color='595959')
    note2.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
    ws.row_dimensions[row].height = 40
    row += 3

    # ── Raw text preview (first 1500 chars) for auditor reference ──
    raw_text = (daca_bank_data.get('_raw_text') or '').strip()
    if raw_text:
        c_hdr = ws.cell(row=row, column=_B, value='Bank Statement — Extracted Text (Reference)')
        c_hdr.font = _font(bold=True, size=11, color='595959')
        c_hdr.fill = _fill(LIGHT_GRAY)
        ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
        row += 1

        # Split into chunks of ~120 chars per cell so it's readable
        preview = raw_text[:3000]
        for chunk_line in preview.split('\n'):
            if not chunk_line.strip():
                continue
            c = ws.cell(row=row, column=_B, value=chunk_line)
            c.font = _font(size=9, name='Courier New')
            c.alignment = Alignment(wrap_text=False)
            ws.merge_cells(start_row=row, start_column=_B, end_row=row, end_column=_B + 5)
            row += 1
            if row > 200:   # cap to avoid massive sheets
                ws.cell(row=row, column=_B, value='... (truncated) ...').font = _font(italic=True, size=9)
                break

    ws.freeze_panes = 'B4'


# ── Convenience function for app.py ──────────────────────────

def generate(gl_result, tb_result, output_path: str,
             period: str = '', property_name: str = '',
             prepaid_ledger_active: list = None,
             bank_rec_data: dict = None,
             gl_cash_balance: float = None,
             daca_bank_data: dict = None,
             daca_gl_balance: float = None,
             je_adjustments: Optional[Dict[str, float]] = None,
             prior_workpaper_path: str = None,
             prior_period: str = None,
             berkadia_loans: list = None,
             dev_bank_rec_data: dict = None,
             ar_aging_data=None,
             capital_schedule_data=None,
             tb_filepath: str = None,
             ar_aging_filepath: str = None,
             ap_aging_filepath: str = None,
             bank_rec_xlsx_filepath: str = None,
             daca_bank_rec_xlsx_filepath: str = None,
             dev_bank_rec_xlsx_filepath: str = None,
             prepared_by: str = '',
             property_config=None) -> str:
    """Alias for generate_bs_workpaper — called from app.py."""
    return generate_bs_workpaper(gl_result, tb_result, output_path, period,
                                  property_name, prepaid_ledger_active,
                                  bank_rec_data, gl_cash_balance,
                                  daca_bank_data, daca_gl_balance,
                                  je_adjustments,
                                  prior_workpaper_path=prior_workpaper_path,
                                  prior_period=prior_period,
                                  berkadia_loans=berkadia_loans,
                                  dev_bank_rec_data=dev_bank_rec_data,
                                  ar_aging_data=ar_aging_data,
                                  capital_schedule_data=capital_schedule_data,
                                  tb_filepath=tb_filepath,
                                  ar_aging_filepath=ar_aging_filepath,
                                  ap_aging_filepath=ap_aging_filepath,
                                  bank_rec_xlsx_filepath=bank_rec_xlsx_filepath,
                                  daca_bank_rec_xlsx_filepath=daca_bank_rec_xlsx_filepath,
                                  dev_bank_rec_xlsx_filepath=dev_bank_rec_xlsx_filepath,
                                  prepared_by=prepared_by,
                                  property_config=property_config)


# ── Workpaper Seed Generator ──────────────────────────────────────────────────

_SEED_MONTH_ORDER = dict(Jan=1, Feb=2, Mar=3, Apr=4, May=5, Jun=6,
                         Jul=7, Aug=8, Sep=9, Oct=10, Nov=11, Dec=12)


def _seed_period_sort(period_str: str):
    """Sort key for period strings like 'Jan-2026'."""
    parts = str(period_str).split('-')
    if len(parts) == 2:
        mon = _SEED_MONTH_ORDER.get(parts[0], 0)
        yr  = int(parts[1]) if parts[1].isdigit() else 0
        return (yr, mon)
    return (0, 0)


def _write_seed_account_tab(wb: 'Workbook', account_code: str, account_name: str,
                             history_rows: list, property_name: str) -> None:
    """
    Write one account tab in the rolling-table format expected by
    _extract_new_format_history().  Tab name: '{account_code} {account_name}'.

    Columns B–G: Period | Beg Balance | Net Activity | GL Ending | TB Ending | Variance
    """
    tab_name = _safe_sheet_name(f'{account_code} {account_name}')
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = COLOR_BS_STD
    ws.column_dimensions['A'].width = 2

    col_widths = {'B': 14, 'C': 16, 'D': 16, 'E': 16, 'F': 16, 'G': 14}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Row 1 — account header
    hdr_val = f'{account_code}  {account_name}'
    c = ws.cell(row=1, column=_B, value=hdr_val)
    _apply(c, font=_font(bold=True, size=13, color='FFFFFF'),
           fill=_fill(DARK_BLUE),
           align=Alignment(horizontal='left', vertical='center'))
    ws.merge_cells(start_row=1, start_column=_B, end_row=1, end_column=_G)
    ws.row_dimensions[1].height = 22

    # Row 2 — property + seed note
    sub = ws.cell(row=2, column=_B,
                  value=f'{property_name}  |  Historical seed — imported prior balances')
    _apply(sub, font=_font(italic=True, size=9, color='FFFFFF'),
           fill=_fill(DARK_BLUE),
           align=Alignment(horizontal='left'))
    ws.merge_cells(start_row=2, start_column=_B, end_row=2, end_column=_G)

    # Row 3 — blank spacer
    ws.row_dimensions[3].height = 6

    # Row 4 — column headers  (col B must contain "Period" exactly — read by extractor)
    hist_hdrs = ['Period', 'Beg Balance', 'Net Activity', 'GL Ending', 'TB Ending', 'Variance']
    for ci, h in enumerate(hist_hdrs):
        c = ws.cell(row=4, column=_B + ci, value=h)
        _apply(c, font=_hdr_font(), fill=_fill(MED_BLUE), border=THIN,
               align=Alignment(horizontal='center', wrap_text=True))
    ws.row_dimensions[4].height = 28

    # Rows 5+ — data
    _NUM_FMT = '#,##0.00;(#,##0.00);"-"'
    for i, hist in enumerate(history_rows):
        row_num = 5 + i
        _var = round((hist.get('gl_end', 0.0) or 0.0) - (hist.get('tb_end', 0.0) or 0.0), 2)
        vals = [
            hist.get('period', ''),
            hist.get('beg_bal', 0.0),
            hist.get('net_change', 0.0),
            hist.get('gl_end', 0.0),
            hist.get('tb_end', hist.get('gl_end', 0.0)),
            _var,
        ]
        alt = _fill(LIGHT_GRAY) if i % 2 == 1 else None
        for ci, val in enumerate(vals):
            c = ws.cell(row=row_num, column=_B + ci, value=val)
            c.border = THIN
            if alt:
                c.fill = alt
            if isinstance(val, float):
                c.number_format = _NUM_FMT
                c.alignment = Alignment(horizontal='right')
            elif ci == 0:  # period
                c.alignment = Alignment(horizontal='center')
        # Variance cell colour
        vc = ws.cell(row=row_num, column=_B + 5)
        _vz = abs(_var) < 0.02
        vc.fill  = _fill(GREEN_FILL if _vz else RED_FILL)
        vc.font  = _font(size=10, color='006100' if _vz else '9C0006')

    ws.freeze_panes = 'B5'


def generate_workpaper_seed(
    entries: list,
    property_name: str = '',
    as_of_period: str = '',
) -> bytes:
    """
    Build a starter GA_Workpapers.xlsx from manually-entered prior-period account balances.

    Intended for onboarding existing GRP properties to the pipeline.  Ryan enters
    prior-period GL ending balances for specific BS accounts (e.g. the last 3–6
    months of history), downloads the seed, and uploads it as the "prior month
    workpaper" on the first pipeline close.  Every subsequent close carries
    forward normally.

    Each entry dict:
        account_code    (str)   — 6-digit GL code, e.g. '111100'
        account_name    (str)   — display name, e.g. 'PNC Operating Cash'
        period          (str)   — 'Jan-2026', 'Feb-2026', …
        gl_ending       (float) — GL ending balance for this period
        tb_ending       (float) — TB ending balance (if omitted, same as gl_ending)

    The function computes:
        beg_bal    = prior period gl_ending (0 for the earliest period per account)
        net_change = gl_ending − beg_bal
        variance   = gl_ending − tb_ending

    The output is readable by _extract_account_history() in bs_workpaper_generator.py,
    meaning the pipeline's first close will carry this history forward automatically.

    Returns the workbook as raw bytes.
    """
    import io
    from collections import defaultdict
    from openpyxl import Workbook as _WB

    # ── Group entries by account ────────────────────────────────
    by_account: dict = defaultdict(list)
    for entry in entries:
        code = str(entry.get('account_code', '') or '').strip()
        if not code:
            continue
        name = str(entry.get('account_name', '') or '').strip()
        period = str(entry.get('period', '') or '').strip()
        if not period:
            continue
        gl_end = float(entry.get('gl_ending', 0) or 0)
        tb_end = float(entry.get('tb_ending', 0) or 0) or gl_end
        by_account[(code, name)].append({
            'period': period,
            'gl_end': gl_end,
            'tb_end': tb_end,
        })

    # ── Sort each account's rows chronologically; compute beg/net ──
    account_history: dict = {}
    for (code, name), rows in by_account.items():
        sorted_rows = sorted(rows, key=lambda r: _seed_period_sort(r['period']))
        history = []
        prev_gl = 0.0
        for row in sorted_rows:
            gl_end     = row['gl_end']
            tb_end     = row['tb_end']
            beg_bal    = prev_gl
            net_change = round(gl_end - beg_bal, 2)
            history.append({
                'period':     row['period'],
                'beg_bal':    beg_bal,
                'net_change': net_change,
                'gl_end':     gl_end,
                'tb_end':     tb_end,
                'variance':   round(gl_end - tb_end, 2),
            })
            prev_gl = gl_end
        account_history[(code, name)] = history

    # ── Build workbook ──────────────────────────────────────────
    wb = _WB()
    wb.remove(wb.active)   # remove default empty sheet

    # Summary Page first (provides period-end anchor for 135150 DATEDIF formulas)
    _write_summary_page(wb, as_of_period or next(
        (r['period'] for rows in account_history.values() for r in rows[-1:]),
        '',
    ))

    # Account tabs — sorted by code for a tidy workbook
    for (code, name), history in sorted(account_history.items(), key=lambda x: x[0][0]):
        _write_seed_account_tab(wb, code, name, history, property_name)

    # Return as bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# Template-based workpaper generator
# ─────────────────────────────────────────────────────────────────────────────

def generate_bs_workpaper_from_template(
    gl_result,
    tb_result,
    output_path: str,
    template_path: str,
    period: str = '',
    property_name: str = '',
    period_end_date=None,
    prepared_by: str = '',   # C-18: no personal name default
    property_code: str = '',
    ar_aging_filepath: str = '',
    ap_aging_filepath: str = '',
    bank_rec_xlsx_filepath: str = '',
    daca_bank_rec_xlsx_filepath: str = '',
    dev_bank_rec_xlsx_filepath: str = '',
    prepaid_ledger_active: list = None,
    bank_rec_data: dict = None,
    daca_bank_data: dict = None,
) -> str:
    """
    Template-based monthly close workpaper generator.

    Copies the property-specific template Excel file (``GA_Workpaper_Template.xlsx``)
    to *output_path* and then:

    1.  Updates ``Summary Page!C4`` with the period-end date so every DATEDIF
        formula on the analysis tabs (Insurance, 135150 PPD Other) recalculates.
    2.  Updates the row-3 header on every applicable tab with the new period
        label and today's prepared date.
    3.  Auto-fills current-period GL tabs (213100, 133110, 133100, 211300) by
        clearing the placeholder data rows and writing the corresponding GL
        transactions for this month.
    4.  Appends new GL transactions to cumulative running-ledger tabs (115200,
        115300, 115600) — existing historical rows are preserved; only net-new
        transactions (dated after the last row already in the template) are
        added.
    5.  Rebuilds the ``Trial Balance`` tab from the parsed ``tb_result`` so
        all ``VLOOKUP(B1,'Trial Balance'!$A:$F,6,0)`` tie-out formulas resolve.
    6.  Regenerates the 6 raw-report tabs (111100 PNC Cash, 115100 DACA,
        131100 AR Aging, 221100 Prepaid Rent, 211100 AP, 111210 BofA Dev)
        every period from whatever raw file is currently uploaded — a plain
        byte-for-byte copy, same as Yardi exported it. 131100 and 221100
        both copy the same AR Aging Detail file (it's one Yardi report
        covering both categories — not split). If no file is uploaded for
        an account this period, the tab shows an explicit "no data uploaded"
        placeholder instead of stale content from a prior period.

    Analysis tabs (RE Tax Analysis, Insurance Analysis, Loan Analysis,
    135150 PPD Other) and static multi-entity ledger tabs (152100 Land,
    154100 Building, etc.) are left completely intact from the template.

    Args:
        gl_result:       GLParseResult from parsers.yardi_gl.parse_gl()
        tb_result:       TBResult from parsers.yardi_trial_balance.parse()
        output_path:     Where to write the populated .xlsx file
        template_path:   Path to GA_Workpaper_Template.xlsx
        period:          Period label e.g. 'Jan-2026'
        property_name:   Property display name
        period_end_date: datetime.date for last day of period (derived from
                         *period* if not supplied)
        prepared_by:     Name for 'Prepared by:' in row-3 headers
        property_code:   Yardi property code (e.g. 'revlabspm')
        bank_rec_data:   Parsed Yardi Bank Rec PDF dict (parsers.yardi_bank_rec.parse())
                         for the Operating account — used to fill '111100 PNC Cash'
                         from its 'gl_transactions' field when no Excel bank rec
                         report is uploaded. The PDF's GL-detail pages carry the
                         same transaction-level data an Excel export would.
        daca_bank_data:  Same, for the DACA account → '115100 DACA'.

    Returns:
        output_path
    """
    import shutil
    import calendar as _calendar
    import re as _re
    from datetime import date as _date, datetime as _dt
    from openpyxl import load_workbook as _lw
    from openpyxl.utils import get_column_letter as _gcl

    # ── 0. Validate & copy template ──────────────────────────────────────────
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"Workpaper template not found at {template_path}. "
            "Upload GA_Workpaper_Template.xlsx via the template-management section."
        )
    shutil.copy2(template_path, output_path)

    # ── 1. Open copied file ───────────────────────────────────────────────────
    wb = _lw(output_path)

    # ── 2. Build GL lookups ───────────────────────────────────────────────────
    # account_code -> GLAccount
    gl_map: dict = {}
    if gl_result and hasattr(gl_result, 'accounts'):
        for _a in gl_result.accounts:
            gl_map[str(getattr(_a, 'account_code', '') or '').strip()] = _a

    # journal-control -> (account_code, account_name) for the debit / P&L side
    # of entries that credit 213100 / 133110 / etc.
    # We look at all P&L accounts (4–8xxxxx) to find the offsetting leg.
    _ctrl_to_acct: dict = {}
    if gl_result and hasattr(gl_result, 'accounts'):
        for _a in gl_result.accounts:
            _ac = str(getattr(_a, 'account_code', '') or '').strip()
            if _ac and _ac[0] in '45678':
                for _t in (getattr(_a, 'transactions', None) or []):
                    _ctrl = str(getattr(_t, 'control', '') or '').strip()
                    if _ctrl and _ctrl not in _ctrl_to_acct:
                        _ctrl_to_acct[_ctrl] = (_ac, getattr(_a, 'account_name', '') or '')

    # ── 3. Period-end date ────────────────────────────────────────────────────
    _period_end: _date | None = period_end_date
    if _period_end is None and period:
        try:
            _pm = _dt.strptime(period, '%b-%Y')
            _last_day = _calendar.monthrange(_pm.year, _pm.month)[1]
            _period_end = _date(_pm.year, _pm.month, _last_day)
        except Exception:
            _period_end = None

    # ── 4. Derived header strings ─────────────────────────────────────────────
    _today_str = _dt.today().strftime('%m/%d/%Y')
    _period_str = period or ''
    _prep_str = prepared_by or 'GRP'   # C-18: no personal name default
    # Short property label used in row-3 headers
    _prop_label = 'revlabs'
    if property_name:
        _pl = property_name.lower()
        # Strip entity suffixes (e.g. "revolution labs owner, llc" → "revlabs")
        for _suffix in (' owner, llc', ' owner,llc', ', llc', ' llc'):
            _pl = _pl.replace(_suffix, '')
        _pl = _pl.replace('revolution labs', 'revlabs').strip()
        if _pl:
            _prop_label = _pl

    # ── 5. Tab configuration ──────────────────────────────────────────────────
    # Tabs the pipeline auto-fills with GL transactions:
    #   account:      GL account code to pull transactions from
    #   data_start:   first data row in the template
    #   amount_col:   1-based column index for the Amount / total column
    #   cumulative:   True = append new rows (running ledger); False = clear & refill
    #   layout:       column-write strategy
    #     'simple'    → B=Date  C=Description          D=Amount
    #     'gl_acct'   → B=Date  C=Description  D=GL Account   E=Amount
    #     'accr_exp'  → B=Date  C=GL Account   D=Description  E=Vendor  F=Amount
    _FILL_TABS: dict = {
        '115200 RET Escrow': {
            'account': '115200', 'data_start': 7, 'amount_col': 4,
            'cumulative': True, 'layout': 'simple',
        },
        '115300 Insurance Escrow': {
            'account': '115300', 'data_start': 7, 'amount_col': 4,
            'cumulative': True, 'layout': 'simple',
        },
        '115600 Loan Reserve': {
            'account': '115600', 'data_start': 7, 'amount_col': 4,
            'cumulative': True, 'layout': 'simple',
        },
        '133100 Accounts Receivable - Ot': {
            'account': '133100', 'data_start': 6, 'amount_col': 5,
            'cumulative': False, 'layout': 'gl_acct',
        },
        '133110 AR Billback': {
            'account': '133110', 'data_start': 6, 'amount_col': 5,
            'cumulative': False, 'layout': 'gl_acct',
        },
        '211300 Accounts Payable - Other': {
            'account': '211300', 'data_start': 6, 'amount_col': 5,
            'cumulative': False, 'layout': 'gl_acct',
        },
        '213100 Accr Exp': {
            'account': '213100', 'data_start': 6, 'amount_col': 6,
            'cumulative': False, 'layout': 'accr_exp',
        },
    }

    # Raw-report tabs — regenerated every period from whatever source is
    # currently available, rather than staying frozen at template-creation
    # content (previously called _PASTED_TABS and never touched at all).
    # If no raw file is available for an account this period, the tab shows
    # an explicit "no file uploaded" placeholder rather than silently
    # carrying forward stale content from a prior period — that ambiguity
    # was the entire reason this regeneration logic was built.
    #   raw_filepath  : path to the raw file for this period, if uploaded.
    #                   131100 and 221100 both point at the same AR Aging
    #                   Detail file — Yardi's own export mixes both
    #                   categories in one report, so both tabs get an
    #                   identical copy rather than a split.
    #   missing_label : human-readable name of the file to upload, shown in
    #                   the placeholder when nothing is available
    _REGEN_TABS: dict = {
        '111100 PNC Cash': {
            'account': '111100', 'raw_filepath': bank_rec_xlsx_filepath,
            'missing_label': 'Bank Reconciliation Excel (PNC Operating)',
            # Fallback when no Excel export is uploaded — the Yardi Bank Rec
            # PDF (already parsed elsewhere for the reconciliation itself)
            # carries the same GL transaction detail on its later pages.
            'pdf_gl_transactions': (bank_rec_data or {}).get('gl_transactions'),
        },
        '115100 DACA': {
            'account': '115100', 'raw_filepath': daca_bank_rec_xlsx_filepath,
            'missing_label': 'DACA Bank Reconciliation Excel',
            # parsers.yardi_daca_rec.parse() doesn't extract a GL-detail
            # section like the Operating parser does (no sample PDF to build
            # that against yet) — fall back to its cleared_deposits/
            # cleared_other_items lists instead, which it DOES already
            # extract, reshaped into the same tab format.
            'pdf_gl_transactions': _daca_fallback_txns(daca_bank_data),
        },
        '131100 AR Aging': {
            'account': '131100', 'raw_filepath': ar_aging_filepath,
            'missing_label': 'AR Aging Detail report',
        },
        '221100 Prepaid Rent - Tenant': {
            'account': '221100', 'raw_filepath': ar_aging_filepath,
            'missing_label': 'AR Aging Detail report',
        },
        '211100 Accounts Payable - Contr': {
            'account': '211100', 'raw_filepath': ap_aging_filepath,
            'missing_label': 'AP Aging Detail report',
        },
        '111210 Cash - Development - Bof': {
            'account': '111210', 'raw_filepath': dev_bank_rec_xlsx_filepath,
            'missing_label': 'Development Bank Statement (BofA)',
        },
    }

    # ── 6. Helper functions ───────────────────────────────────────────────────

    def _find_tieout_row(ws, search_col: int = 2, max_scan: int = 300) -> int | None:
        """Return the row number containing 'ending balance per gl' in *search_col*."""
        for r in range(1, min(ws.max_row + 1, max_scan + 1)):
            v = ws.cell(r, search_col).value
            if v and isinstance(v, str) and 'ending balance per gl' in v.lower():
                return r
        return None

    def _last_nonempty_row(ws, col: int, row_start: int, row_stop: int) -> int:
        """Return the last row in [row_start, row_stop) that has a value in *col*."""
        last = row_start - 1
        for r in range(row_start, row_stop):
            if ws.cell(r, col).value is not None:
                last = r
        return last

    def _update_row3_header(ws) -> None:
        """Replace the period label and prepared date in the row-3 header cell."""
        v = ws.cell(3, 2).value
        if not (v and isinstance(v, str)):
            return
        new_v = _re.sub(
            r'(Period:\s*)[A-Za-z]+-\d{4}',
            lambda m: m.group(1) + _period_str,
            v,
        )
        new_v = _re.sub(r'\d{2}/\d{2}/\d{4}', _today_str, new_v)
        ws.cell(3, 2).value = new_v

    def _copy_row_style(ws, src_row: int, dst_row: int, col_start: int, col_end: int) -> None:
        """
        Copy font/border/fill/number_format from src_row onto dst_row.

        insert_rows() only shifts existing rows and creates blank ones with
        default (unformatted) styling — it doesn't clone the surrounding
        rows' formatting. Every newly-inserted row (i.e. any account with
        more transactions this period than the template had pre-formatted
        placeholder rows for) came out with no currency format/borders/font,
        looking visibly different from the rest of the table.
        """
        from copy import copy as _copy_style
        for _c in range(col_start, col_end + 1):
            src_cell = ws.cell(src_row, _c)
            dst_cell = ws.cell(dst_row, _c)
            dst_cell.font = _copy_style(src_cell.font)
            dst_cell.border = _copy_style(src_cell.border)
            dst_cell.fill = _copy_style(src_cell.fill)
            dst_cell.alignment = _copy_style(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format

    def _coerce_date(d):
        """
        Return a plain datetime.date from whatever the GL — or an existing
        template row — stores.

        Cumulative tabs (115200, 115300, 115600) determine "new since last
        run" by reading the date already written in the last template row.
        That value can be a real date/datetime OR plain text (e.g. a prior
        run of this same code, or manual template editing, can leave a
        string like '01/16/2026'). Without string support here, the "last
        date" lookup silently resolved to None, which made the fill logic
        treat every transaction as new and re-append rows that were
        already there — confirmed duplicate rows on 115200 and 115600.
        """
        if d is None:
            return None
        if hasattr(d, 'date') and callable(d.date):
            return d.date()
        if isinstance(d, _date):
            return d
        if isinstance(d, str):
            s = d.strip()
            for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%m/%d/%y'):
                try:
                    return _dt.strptime(s, fmt).date()
                except ValueError:
                    continue
        return None

    def _rewrite_tieout_formulas(ws, tieout_row: int, data_start: int,
                                  last_written: int, amount_col: int) -> None:
        """
        Rewrite the three tieout rows:
          tieout_row     → =SUM(XN:XM)
          tieout_row + 1 → =VLOOKUP(B1,'Trial Balance'!$A:$F,6,0)
          tieout_row + 2 → =XN-XM  (variance)
        where X is the column letter for *amount_col*.
        """
        col_ltr = _gcl(amount_col)
        ws.cell(tieout_row,     amount_col).value = (
            f'=SUM({col_ltr}{data_start}:{col_ltr}{last_written})'
        )
        ws.cell(tieout_row + 1, amount_col).value = (
            f"=VLOOKUP(B1,'Trial Balance'!$A:$F,6,0)"
        )
        ws.cell(tieout_row + 2, amount_col).value = (
            f'={col_ltr}{tieout_row}-{col_ltr}{tieout_row + 1}'
        )

    def _boundary_prorated_formula(g: str, h: str, i: str, j: str, r: int, asof_ref: str) -> str:
        """
        Cumulative-amortized-to-date formula matching prepaid_ledger.py's
        REAL release schedule (_month_amount() / get_current_amortization())
        exactly — not the simpler continuous day-based straight-line used
        elsewhere (Insurance Analysis originally, before this fix). That
        model is: every month gets the flat monthly rate EXCEPT the first
        and last service months, which are day-prorated within that one
        month only. Confirmed against the real release logic (summing
        _month_amount() month-by-month) for both Greatland and Apex real
        invoices — exact match at every checkpoint.

        g/h/i/j are the column letters for Start Date / End Date / Total /
        Monthly Amount; r is the row number; asof_ref is the absolute cell
        reference holding the "as of" reporting date.

        total_months uses DATEDIF({g},{h},"M")+1 — verified to match
        prepaid_ledger.py's real `_count_months()` (relativedelta-based)
        exactly across 200k random date-pair tests, including the edge case
        below. This is NOT the same as the more obvious-looking
        DATEDIF({g},{h}+1,"M") (shift the end date by a day, then diff) —
        that alternative undercounts by one whole month whenever the end
        day is earlier in its month than the start day is in its month
        (e.g. starts the 10th, ends the 20th of an 11-months-later month),
        which would silently understate the cumulative total near the end
        of an affected item's schedule.
        """
        return (
            f'=IF({asof_ref}<{g}{r},0,'
            f'IF({asof_ref}>={h}{r},{i}{r},'
            f'MIN(DATEDIF(DATE(YEAR({g}{r}),MONTH({g}{r}),1),{asof_ref},"M")+1,'
            f'DATEDIF({g}{r},{h}{r},"M")+1)*{j}{r}'
            f'-IF(DAY({g}{r})>1,{j}{r}*(DAY({g}{r})-1)/DAY(EOMONTH({g}{r},0)),0)'
            f'-IF({asof_ref}>=DATE(YEAR({h}{r}),MONTH({h}{r}),1),'
            f'IF(DAY({h}{r})<DAY(EOMONTH({h}{r},0)),'
            f'{j}{r}*(DAY(EOMONTH({h}{r},0))-DAY({h}{r}))/DAY(EOMONTH({h}{r},0)),0),0)'
            f'))'
        )

    def _write_txn_row(ws, row: int, txn, layout: str,
                       amount_col: int) -> None:
        """Write one GL transaction into the appropriate columns."""
        txn_date = _coerce_date(getattr(txn, 'date', None))
        desc = (str(getattr(txn, 'description', '') or '').strip() or
                str(getattr(txn, 'remarks', '') or '').strip())
        # Strip a leading blank-field separator (e.g. a blank vendor/memo
        # field exported as ": Reversal of J-18456" — the reversal itself is
        # already filtered out upstream, but the same blank-field artifact
        # can appear on other boilerplate text too).
        desc = desc.lstrip(': –-').strip()
        amt = float(getattr(txn, 'net_amount', 0) or 0)
        ctrl = str(getattr(txn, 'control', '') or '').strip()
        offset_code, _ = _ctrl_to_acct.get(ctrl, ('', ''))

        if layout == 'simple':
            # Date(B)  Description(C)  Amount(D)
            ws.cell(row, 2).value = txn_date
            ws.cell(row, 3).value = desc
            ws.cell(row, amount_col).value = amt

        elif layout == 'gl_acct':
            # Date(B)  Description(C)  GL Account(D)  Amount(E)
            ws.cell(row, 2).value = txn_date
            ws.cell(row, 3).value = desc
            ws.cell(row, 4).value = offset_code or ''
            ws.cell(row, amount_col).value = amt

        elif layout == 'accr_exp':
            # Date(B)  GL Account(C)  Description(D)  Vendor(E)  Amount(F)
            remarks = str(getattr(txn, 'remarks', '') or '').strip()
            vendor = remarks if remarks and remarks != desc else desc
            ws.cell(row, 2).value = txn_date
            ws.cell(row, 3).value = offset_code or ''
            ws.cell(row, 4).value = desc
            ws.cell(row, 5).value = vendor
            ws.cell(row, amount_col).value = amt

    def _fill_ppd_other_tab(ws) -> None:
        """
        Rebuild '135150 PPD Other' from the live prepaid ledger every period.

        This tab used to be static — carried over untouched from the
        template, so new invoices discovered by prepaid_ledger.py never
        showed up here even though the ledger itself was current.

        Layout (unlike the generic _FILL_TABS accounts, this schema is
        specific to this tab and keeps the template's own live formulas
        for the amortization math, only the invoice-level facts are
        written from data):
          B Vendor | C Description | D Invoice # | E Invoice Date |
          F GL Account | G Start Date | H End Date | I Total Amount |
          J =I/(DATEDIF(G,H,"M")+1)           (monthly amount, reference only)
          K monthly-bucket, boundary-day-prorated cumulative (amortized to date)
          L =I-K                              (remaining balance)

        J and K's total-months basis (DATEDIF(G,H,"M")+1, not the more
        obvious-looking DATEDIF(G,H+1,"M")) is deliberately chosen to match
        prepaid_ledger.py's real `_count_months()` (relativedelta-based)
        exactly — verified against 200k random date pairs with zero
        mismatches. The two differ only in a rare edge case (end-of-service
        day earlier in its month than the start day is in its month), where
        DATEDIF(G,H+1,"M") undercounts by one whole month; using the wrong
        one here would silently understate K's cumulative total near the
        end of an affected item's schedule even though this tab's own J/K
        pair would still look internally consistent. K itself uses a
        monthly-bucket model (flat monthly amount every month except the
        first/last service months, which are day-prorated within that one
        month only) — not straight-line day proration across the whole
        period, which understates amortized-to-date for any mid-month start
        (confirmed by Ryan 2026-08-03 with a concrete example). Unlike
        Insurance Analysis's version of this same formula (which doesn't
        feed anything downstream), L's SUM here IS the real 135150 GL
        tie-out, so getting both of these right matters.
        Footer: L{tieout}=SUM(L{data_start}:L{last}), L{tieout+1}=VLOOKUP
        ending balance per TB, L{tieout+2}=variance — same B1-anchor
        pattern as every other tab.
        """
        _data_start = 6
        _tieout = _find_tieout_row(ws)
        if _tieout is None:
            return

        # Only unmerge the DATA region (data_start..tieout-1) — a placeholder
        # like "No activity this period" merged across several columns lives
        # there and blocks .value writes. The footer rows (tieout, +1, +2)
        # have their OWN merged label cells (e.g. 'Ending Balance per GL'
        # spanning B:D) that are never written to — only the single amount
        # column is — so unmerging them was unnecessary and, confirmed
        # against the real template, destroyed their intentional formatting
        # on every sheet, every run.
        for _mg in list(ws.merged_cells.ranges):
            if _mg.min_row < _tieout and _mg.max_row >= _data_start:
                ws.unmerge_cells(str(_mg))

        for _r in range(_data_start, _tieout):
            for _c in range(2, 13):
                ws.cell(_r, _c).value = None

        _items = prepaid_ledger_active or []
        if not _items:
            _rewrite_ppd_tieout(ws, _tieout, _data_start, _data_start)
            return

        # +1 reserves a permanent blank row between the last item and the
        # footer — the template's original design (one real "buffer" row
        # before 'Ending Balance per GL'), which the item-count-driven
        # insert below would otherwise consume once there are enough items
        # to fill every available row, butting the footer directly against
        # the last item with no visual break.
        _cleared_rows = _tieout - _data_start
        _needed_rows = len(_items) + 1
        if _needed_rows > _cleared_rows:
            _to_insert = _needed_rows - _cleared_rows
            _insert_at = _tieout
            ws.insert_rows(_insert_at, _to_insert)
            _tieout += _to_insert

            # openpyxl's insert_rows() shifts real cell objects but does NOT
            # update the merged-cell range *definitions* — confirmed against
            # the real template: the footer's own label merges (e.g. 'Ending
            # Balance per GL' spanning B:D, plus its blank E:G/H:J
            # companions) stayed registered at their OLD row number after
            # the insert, so the newly-inserted row (now holding real item
            # data, not the footer) silently blocked writes to every
            # non-anchor cell in those stale ranges — Description/Invoice
            # #/GL Account/Start Date/Total/Monthly Amt all came back None
            # with no error, while the anchor columns (Vendor/Invoice
            # Date/End Date) wrote fine. ws.unmerge_cells() itself isn't safe
            # to call here — it assumes real MergedCell placeholders still
            # sit at those (now-stale) coordinates, which insert_rows leaves
            # empty, raising KeyError. Drop the range registration directly
            # instead, which is all that's needed since there's nothing
            # left to actually unmerge.
            for _mg in list(ws.merged_cells.ranges):
                if _mg.min_row < _tieout and _mg.max_row >= _data_start:
                    ws.merged_cells.remove(_mg)

        # Apply proper data-row formatting (font/border/fill/number-format)
        # to every row that will hold an item, not just rows insert_rows()
        # just created — confirmed against the real template: it reserves
        # one extra blank "spacer" row (data_start + 9) styled as a plain
        # divider (General number format, no currency/date formats, a
        # heavier "medium" border) rather than a real data row, so an item
        # landing there via the 9th-of-9-original-slots path came out with
        # correct values but wrong formatting instead of the missing-value
        # symptom above (the two mangled rows were adjacent for this exact
        # reason: one pre-existing under-styled spacer immediately followed
        # by one newly-inserted, merge-blocked row).
        for _i in range(len(_items)):
            _r = _data_start + _i
            _src = _data_start if (_r - _data_start) % 2 == 0 else _data_start + 1
            _copy_row_style(ws, _src, _r, 2, 12)

        for _i, _item in enumerate(_items):
            _r = _data_start + _i
            ws.cell(_r, 2).value = str(_item.get('vendor', '') or '')
            ws.cell(_r, 3).value = str(_item.get('description', '') or '')
            ws.cell(_r, 4).value = str(_item.get('invoice_number', '') or '')
            ws.cell(_r, 5).value = _coerce_date(_item.get('invoice_date'))
            ws.cell(_r, 6).value = str(_item.get('gl_account_number', '') or '')
            # Prefer the TRUE original service start over 'service_start' —
            # prepaid_ledger.merge_nexus() deliberately rebases that field to
            # the tracking-start month for its own internal release
            # scheduling (confirmed necessary there), which made a
            # late-discovered item's start date shown here silently wrong
            # (e.g. a true Dec-2025 start displayed as Jan-2026). Falls back
            # to 'service_start' for items with no rebasing history (seed/
            # manually-added items, where it already holds the true date).
            ws.cell(_r, 7).value = _coerce_date(
                _item.get('true_service_start') or _item.get('service_start')
            )
            ws.cell(_r, 8).value = _coerce_date(_item.get('service_end'))
            ws.cell(_r, 9).value = float(_item.get('total_amount', 0) or 0)
            ws.cell(_r, 10).value = f'=I{_r}/(DATEDIF(G{_r},H{_r},"M")+1)'
            # IF-guarded: a blank Start/End date (a real possibility per
            # prepaid_ledger.py's own None-guards on these exact fields)
            # would otherwise evaluate incorrectly with NO visible Excel
            # error — silently reporting the item as fully amortized
            # (remaining balance $0) instead of flagging it. Default to 0
            # amortized (full amount held as Remaining) when either date is
            # missing — safe/conservative and visibly wrong via the blank
            # Start/End cells, not silently wrong via a $0 balance nobody
            # would think to question.
            #
            # Monthly-bucket-with-boundary-proration model (matches the REAL
            # release schedule in prepaid_ledger._month_amount(), not a
            # continuous day-based straight-line) — see
            # _boundary_prorated_formula() for the full model explanation.
            _k_formula = _boundary_prorated_formula('G', 'H', 'I', 'J', _r, "'Summary Page'!$C$4")
            ws.cell(_r, 11).value = f'=IF(OR(G{_r}="",H{_r}=""),0,{_k_formula[1:]})'
            ws.cell(_r, 12).value = f'=I{_r}-K{_r}'

        _last_written = _data_start + len(_items) - 1
        _rewrite_ppd_tieout(ws, _tieout, _data_start, _last_written)

    def _rewrite_ppd_tieout(ws, tieout_row: int, data_start: int, last_written: int) -> None:
        ws.cell(tieout_row,     12).value = f'=SUM(L{data_start}:L{last_written})'
        ws.cell(tieout_row + 1, 12).value = "=VLOOKUP(B1,'Trial Balance'!$A:$F,6,0)"
        ws.cell(tieout_row + 2, 12).value = f'=L{tieout_row}-L{tieout_row + 1}'

    # ── 7. Summary Page: update period-end date anchor ────────────────────────
    if 'Summary Page' in wb.sheetnames:
        _ws_s = wb['Summary Page']
        if _period_end is not None:
            _ws_s['C4'] = _period_end

    # ── 7b. Regenerate raw-report tabs from currently available data ─────────
    for _sn, _rc in _REGEN_TABS.items():
        if _sn not in wb.sheetnames:
            continue
        _acct_code   = _rc['account']
        _orig_idx    = wb.sheetnames.index(_sn)
        _missing_lbl = _rc.get('missing_label', 'source file')

        try:
            del wb[_sn]

            _raw_fp = _rc.get('raw_filepath')
            _copied_ok = False
            if _raw_fp and os.path.exists(_raw_fp):
                _copied_ok = _copy_raw_tb_sheet(_raw_fp, wb, tab_name=_sn)
            if not _copied_ok and _rc.get('pdf_gl_transactions'):
                # No Excel export uploaded, but a Bank Rec PDF was — use its
                # GL-detail transactions instead of showing "no data uploaded".
                _write_gl_transactions_tab(
                    wb, _sn, _acct_code, _rc['pdf_gl_transactions'],
                    _period_str, property_name,
                )
                _copied_ok = True
            if not _copied_ok:
                _write_no_data_placeholder_tab(wb, _sn, _missing_lbl, _acct_code)

            if _sn in wb.sheetnames:
                _cur_idx = wb.sheetnames.index(_sn)
                if _cur_idx != _orig_idx:
                    wb.move_sheet(_sn, offset=(_orig_idx - _cur_idx))
        except Exception as _rgex:
            print(f'[bs_workpaper_generator] Tab regeneration failed for {_sn}: {_rgex}')

    # ── 8. Process every sheet ────────────────────────────────────────────────
    for _sn in wb.sheetnames:
        _ws = wb[_sn]

        # --- Skip regenerated tabs (handled above) and the Trial Balance ---
        if _sn in _REGEN_TABS or _sn == 'Summary Page':
            continue
        if _sn == 'Trial Balance':
            continue

        # --- Update row-3 header (period + prepared date) ---
        _update_row3_header(_ws)

        # --- 135150 PPD Other: distinct schema, filled from the live ledger ---
        if _sn == '135150 PPD Other':
            _fill_ppd_other_tab(_ws)
            continue

        # --- Auto-fill if this tab has a GL data config ---
        if _sn not in _FILL_TABS:
            continue

        _cfg = _FILL_TABS[_sn]
        _acct_code  = _cfg['account']
        _data_start = _cfg['data_start']
        _amt_col    = _cfg['amount_col']
        _cumulative = _cfg['cumulative']
        _layout     = _cfg['layout']

        # Locate tieout row
        _tieout = _find_tieout_row(_ws)
        if _tieout is None:
            continue  # template structure unrecognised — skip

        # Unmerge any merged cells inside the DATA region (data_start..
        # tieout-1) before touching it — a template placeholder row (e.g.
        # "No activity this period", which spans multiple columns) is
        # commonly merged, and writing .value on a merged non-anchor cell
        # raises AttributeError ('MergedCell' object attribute 'value' is
        # read-only). Confirmed present in the real template for
        # 133100/133110/211300/213100.
        #
        # Deliberately does NOT extend into the footer rows (tieout, +1, +2)
        # — those have their own merged LABEL cells ('Ending Balance per
        # GL' etc. spanning multiple columns) that are never written to
        # (only the single amount column is, via _rewrite_tieout_formulas).
        # An earlier version of this fix used `_tieout + 3` and unmerged
        # those too — confirmed against the real template that this broke
        # the footer's intentional formatting on every sheet, every run,
        # for no reason (nothing there was actually causing the crash).
        for _mg in list(_ws.merged_cells.ranges):
            if _mg.min_row < _tieout and _mg.max_row >= _data_start:
                _ws.unmerge_cells(str(_mg))

        # GL account for this tab
        _gl_acct = gl_map.get(_acct_code)
        _all_txns = list(getattr(_gl_acct, 'transactions', None) or []) if _gl_acct else []
        # Auto-reversals of last period's own accrual are a mechanical Yardi
        # artifact, not real activity to review — exclude them from every
        # workpaper detail tab, not just this one. Also drop transactions
        # with no date and no description/remarks — these carry nothing
        # displayable and previously produced a blank-looking row with the
        # amount shifted into the wrong column.
        def _is_blank_txn(t) -> bool:
            has_date = _coerce_date(getattr(t, 'date', None)) is not None
            has_text = bool(str(getattr(t, 'description', '') or '').strip()
                             or str(getattr(t, 'remarks', '') or '').strip())
            return not has_date and not has_text
        _all_txns = [t for t in _all_txns if not _is_reversal_txn(t) and not _is_blank_txn(t)]

        # ── CUMULATIVE tab: append new transactions after existing template rows ──
        if _cumulative:
            # Find the last date already in the template (col B) to avoid duplicates
            _last_row = _last_nonempty_row(_ws, 2, _data_start, _tieout)
            _last_date: _date | None = None
            if _last_row >= _data_start:
                _raw_date = _ws.cell(_last_row, 2).value
                _last_date = _coerce_date(_raw_date)

            # Only keep transactions strictly newer than the last template date
            if _last_date:
                _new_txns = [
                    t for t in _all_txns
                    if _coerce_date(getattr(t, 'date', None)) is not None
                    and _coerce_date(getattr(t, 'date', None)) > _last_date
                ]
            else:
                _new_txns = _all_txns

            if not _new_txns:
                # Nothing new — just tidy the SUM formula range
                _eff_last = max(_last_row, _data_start)
                _rewrite_tieout_formulas(_ws, _tieout, _data_start, _eff_last, _amt_col)
                continue

            # Write position: directly after last existing row
            _write_start = _last_row + 1

            # Ensure enough rows exist before _tieout (leave 1 gap row)
            _rows_avail = _tieout - _write_start - 1
            if len(_new_txns) > _rows_avail:
                _to_insert = len(_new_txns) - _rows_avail
                _insert_at = _tieout
                _ws.insert_rows(_insert_at, _to_insert)
                _tieout += _to_insert
                for _r in range(_insert_at, _insert_at + _to_insert):
                    # Preserve alternating row-banding — copying from
                    # _data_start alone flattened every inserted row to
                    # that one row's (unfilled) style regardless of where
                    # it lands in the banding sequence.
                    _src = _data_start if (_r - _data_start) % 2 == 0 else _data_start + 1
                    _copy_row_style(_ws, _src, _r, 2, _amt_col)

            for _i, _t in enumerate(_new_txns):
                _write_txn_row(_ws, _write_start + _i, _t, _layout, _amt_col)

            _last_written = _write_start + len(_new_txns) - 1

        # ── CURRENT-PERIOD tab: clear placeholder rows and write fresh GL data ──
        else:
            # Clear everything from data_start up to (but not including) tieout row
            for _r in range(_data_start, _tieout):
                for _c in range(2, _amt_col + 2):
                    _ws.cell(_r, _c).value = None

            if not _all_txns:
                # No GL activity — leave blank; SUM over an empty-but-valid range
                _rewrite_tieout_formulas(_ws, _tieout, _data_start, _data_start, _amt_col)
                continue

            # Check whether more rows are needed than the cleared region provides
            _cleared_rows = _tieout - _data_start  # rows data_start … tieout-1
            if len(_all_txns) > _cleared_rows:
                _to_insert = len(_all_txns) - _cleared_rows
                _insert_at = _tieout
                _ws.insert_rows(_insert_at, _to_insert)
                _tieout += _to_insert
                for _r in range(_insert_at, _insert_at + _to_insert):
                    _src = _data_start if (_r - _data_start) % 2 == 0 else _data_start + 1
                    _copy_row_style(_ws, _src, _r, 2, _amt_col)

            for _i, _t in enumerate(_all_txns):
                _write_txn_row(_ws, _data_start + _i, _t, _layout, _amt_col)

            _last_written = _data_start + len(_all_txns) - 1

        # Rewrite tieout, VLOOKUP, and variance formula rows
        _rewrite_tieout_formulas(_ws, _tieout, _data_start, _last_written, _amt_col)

    # ── 9. Rebuild Trial Balance tab ──────────────────────────────────────────
    # The per-account tabs use =VLOOKUP(B1,'Trial Balance'!$A:$F,6,0)
    # where col A = account code and col F = ending balance.
    # We reconstruct this tab from the parsed TBResult so the formulas resolve.
    if (tb_result and hasattr(tb_result, 'accounts') and tb_result.accounts
            and 'Trial Balance' in wb.sheetnames):
        _ws_tb = wb['Trial Balance']
        # Unmerge all cells first (merged cells raise AttributeError on value write)
        for _mg in list(_ws_tb.merged_cells.ranges):
            _ws_tb.unmerge_cells(str(_mg))
        # Clear existing content (skip MergedCell slaves — they're now unmerged)
        for _r in range(1, max(_ws_tb.max_row + 1, 250)):
            for _c in range(1, 8):
                try:
                    _ws_tb.cell(_r, _c).value = None
                except AttributeError:
                    pass  # residual merged cell — safely ignorable
        # Yardi-style header rows
        _ws_tb.cell(1, 1).value = (
            f'Property =  {property_code or ""} {property_name or ""}'.strip()
        )
        _ws_tb.cell(3, 1).value = f'Period = {_period_str}'
        _ws_tb.cell(5, 3).value = 'Forward'
        _ws_tb.cell(5, 6).value = 'Ending'
        _ws_tb.cell(6, 3).value = 'Balance'
        _ws_tb.cell(6, 4).value = 'Debit'
        _ws_tb.cell(6, 5).value = 'Credit'
        _ws_tb.cell(6, 6).value = 'Balance'
        # Data rows — account code in col A is the VLOOKUP key
        for _idx, _ta in enumerate(tb_result.accounts):
            _r = 7 + _idx
            _ws_tb.cell(_r, 1).value = str(getattr(_ta, 'account_code', '') or '')
            _ws_tb.cell(_r, 2).value = getattr(_ta, 'account_name', '') or ''
            _ws_tb.cell(_r, 3).value = float(getattr(_ta, 'forward_balance', 0) or 0)
            _ws_tb.cell(_r, 4).value = float(getattr(_ta, 'debit', 0) or 0)
            _ws_tb.cell(_r, 5).value = float(getattr(_ta, 'credit', 0) or 0)
            _ws_tb.cell(_r, 6).value = float(getattr(_ta, 'ending_balance', 0) or 0)

    # ── 10. Save ──────────────────────────────────────────────────────────────
    wb.save(output_path)
    return output_path

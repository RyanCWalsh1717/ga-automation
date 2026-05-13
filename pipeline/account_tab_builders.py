"""
Custom workpaper tab builders — account-by-account logic
=========================================================
Each function builds a single Excel worksheet with the standard layout:

  Standard tabs  : Date | Description | Entity | Amount
                   (TB tie-out block at bottom)

  115100 DACA    : Date | Tenant/Description | Deposits | Disbursements |
                   Adjustments | Ending Balance (running formula)

  115200/115300  : Date | Description | Entity | Amount | Running Balance

Called from bs_workpaper_generator.generate_bs_workpaper() via the
CUSTOM_BUILDERS dispatch dict.
"""
from __future__ import annotations

import re
from datetime import datetime, date
from typing import Optional, List, Dict, Any

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Shared palette (matches bs_workpaper_generator) ──────────────────────────
# ── Greatland Brand Palette ────────────────────────────────────────────────────
# Source: Greatland Theme - New.thmx  (accent5=002060, dk2/accent2=2D6F50)
DARK_BLUE  = '002060'   # Greatland dark navy   (was 1F4E78)
MED_BLUE   = '2D6F50'   # Greatland green        (was 2E75B6)
LIGHT_BLUE = 'D6EAE1'   # light green tint       (was D6E4F0)
LIGHT_GRAY = 'F2F2F2'   # alternating row shade  (unchanged)
GREEN_FILL = 'E2EFDA'   # tie-out pass           (unchanged)
RED_FILL   = 'FFCCCC'   # tie-out fail           (unchanged)
AMBER_FILL = 'FFF2CC'
WHITE      = 'FFFFFF'

THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
THICK_BOTTOM = Border(bottom=Side(style='medium'))
DOUBLE_BTM   = Border(bottom=Side(style='double'))

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def _font(bold=False, italic=False, size=11, color='000000', name='Calibri'):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

def _apply(cell, font=None, fill=None, fmt=None, border=None, align=None):
    if font:   cell.font   = font
    if fill:   cell.fill   = fill
    if fmt:    cell.number_format = fmt
    if border: cell.border = border
    if align:  cell.alignment = align


# ── Legacy seed helpers (retained for reference; no longer used by builders) ──

# 115200 — RET Escrow historical reference
_RET_SEED: List[tuple] = [
    ('3/25/2024',  'RET Escrow - Per statement due 4.10.24',    232339.00),
    ('5/24/2024',  'RET Escrow - Per statement due 5.10.24',    232339.00),
    ('6/25/2024',  'RET Escrow - Per statement due 6.10.24',    232339.00),
    ('7/25/2024',  'RET Escrow - Per statement due 7.10.24',    232339.00),
    ('7/25/2024',  'RET ESCROW Payment 8.1.24-2024',           -680420.84),
    ('8/25/2024',  'RET Escrow - Per statement due 8.10.24',    232339.00),
    ('9/9/2024',   'RET Escrow - Per statement due 09.09.24',   216069.98),
    ('10/7/2024',  'RET Escrow - Per statement due 10.07.24',   216069.98),
    ('10/17/2024', 'RET ESCROW Payment 10.1.24-2024',          -680420.84),
    ('11/7/2024',  'RET Escrow - Per statement due 11.07.24',   216069.98),
    ('12/9/2024',  'RET Escrow - Per statement due 12.09.24',   216069.98),
    ('1/7/2025',   'RET Escrow - Per statement due 01.07.25',   216069.98),
    ('1/16/2025',  'RET ESCROW Payment 01.1.25-Q3-2025',       -651630.69),
    ('2/7/2025',   'RET Escrow - Per statement due 02.07.25',   216069.98),
    ('3/7/2025',   'RET Escrow - Per statement due 03.07.25',   216069.98),
    ('4/7/2025',   'RET Escrow - Per statement due 04.07.25',   216069.98),
    ('4/16/2025',  'RET ESCROW Payment 01.1.25-Q4-2025',       -651630.69),
    ('5/9/2025',   'RET Escrow - Per statement due 05.09.25',   216069.98),
    ('6/9/2025',   'RET Escrow - Per statement due 06.09.25',   216069.98),
    ('7/17/2025',  'RET ESCROW Payment 01.1.25-Q1-2025',       -682671.08),
    ('7/9/2025',   'RET Escrow - Per statement due 07.09.25',   216069.98),
    ('8/7/2025',   'RET Escrow - Per statement due 08.07.25',   216069.98),
    ('9/8/2025',   'RET Escrow - Per statement due 09.07.25',   216069.98),
    ('10/7/2025',  'RET Escrow - Per statement due 10.07.25',   216069.98),
    ('10/17/2025', 'RET ESCROW Payment 10.1.25-Q2-2025',       -682671.07),
    ('11/7/2025',  'RET Escrow - Per statement due 11.07.25',   216069.98),
    ('12/8/2025',  'RET Escrow - Per statement due 12.08.25',   203295.19),
]

# 115300 — Insurance Escrow seed ledger
_INSUR_SEED: List[tuple] = [
    ('4/25/2024',  'Property Insurance per 4.10.24 stmt due',    23431.00),
    ('5/25/2024',  'Property Insurance per 5.10.24 stmt due',    23431.00),
    ('6/25/2024',  'Property Insurance per 6.10.24 stmt due',    23431.00),
    ('7/25/2024',  'Property Insurance per 7.10.24 stmt due',    23431.00),
    ('8/25/2024',  'Property Insurance per 8.10.24 stmt due',    23431.00),
    ('9/9/2024',   'Property Insurance per 9.09.24 stmt due',    19280.11),
    ('10/7/2024',  'Property Insurance per 10.07.24 stmt due',   19280.11),
    ('11/7/2024',  'Property Insurance per 11.07.24 stmt due',   19280.11),
    ('12/9/2024',  'Property Insurance per 12.09.24 stmt due',   19280.11),
    ('1/7/2025',   'Property Insurance per 01.07.25 stmt due',   19280.11),
    ('2/7/2025',   'Property Insurance per 02.07.25 stmt due',   19280.11),
    ('3/7/2025',   'Property Insurance per 03.07.25 stmt due',   19280.11),
    ('4/7/2025',   'Property Insurance per 04.07.25 stmt due',   19280.11),
    ('5/9/2025',   'Property Insurance per 05.09.25 stmt due',   19280.11),
    ('6/9/2025',   'Property Insurance per 06.09.25 stmt due',   19280.11),
    ('7/7/2025',   'Property Insurance per 07.09.25 stmt due',   19280.11),
    ('8/7/2025',   'Property Insurance per 08.09.25 stmt due',   19280.11),
    ('8/15/2025',  'Property Insurance transfer to revlab owner Operator', -66912.85),
    ('9/8/2025',   'Property Insurance per 09.08.25 stmt due',   19280.11),
    ('10/7/2025',  'Property Insurance per 10.08.25 stmt due',   19280.11),
    ('11/7/2025',  'Property Insurance per 11.08.25 stmt due',   19280.11),
]

# 115600 — Loan Reserve seed ledger
_LOAN_RESERVE_SEED: List[tuple] = [
    ('3/25/2025',  'Rcd: Alchemab Lease Settlement Payment',  4733004.36),
    ('3/25/2025',  'Rcd: 2/25 - 3/25 Interest Income',           350.08),
    ('4/25/2025',  'Rcd: 04/25 Interest Income',                 265.48),
    ('5/25/2025',  'Rcd: 05/25 Interest Income',                 256.43),
    ('6/25/2025',  'Rcd: 06/25 Interest Income',                 238.29),
    ('7/25/2025',  'Rcd: 07/25 Interest Income',                 274.62),
    ('8/25/2025',  'Rcd: 08/25 Interest Income',                 238.31),
    ('9/25/2025',  'Rcd: 09/25 Interest Income',                 256.49),
    ('10/25/2025', 'Rcd: 10/25 Interest Income',                 256.50),
    ('11/25/2025', 'Rcd: 11/25 Interest Income',                 220.19),
    ('12/25/2025', 'Rcd: 12/25 Interest Income',                 283.78),
]

_ENTITY = 'Revlabs'

_MONTH_MAP = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,
    'may': 5, 'jun': 6, 'jul': 7, 'aug': 8,
    'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
}

# Seed map used by _get_escrow_seed_rows (below helpers)
_ESCROW_SEED_MAP = {
    '115200': _RET_SEED,
    '115300': _INSUR_SEED,
    '115600': _LOAN_RESERVE_SEED,
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_close_period(period: str):
    """Return (year, month) ints from 'Mar-2026'.  (0,0) on failure."""
    m = re.search(r'([A-Za-z]{3})[\s\-](\d{4})', period or '')
    if not m:
        return 0, 0
    return int(m.group(2)), _MONTH_MAP.get(m.group(1).lower(), 0)

def _parse_date(d: str) -> Optional[date]:
    """Parse M/D/YYYY or M/D/YY date strings."""
    for fmt in ('%m/%d/%Y', '%m/%d/%y'):
        try:
            return datetime.strptime(d.strip(), fmt).date()
        except (ValueError, AttributeError):
            pass
    return None

def _seed_rows_for_period(seed: List[tuple], close_year: int, close_month: int) -> List[tuple]:
    """Return seed rows whose date falls on or before the last day of close month."""
    result = []
    for date_str, desc, amt in seed:
        d = _parse_date(date_str)
        if d is None:
            continue
        if (d.year, d.month) <= (close_year, close_month):
            result.append((d, desc, amt))
    # Sort chronologically
    result.sort(key=lambda r: r[0])
    return result


def _get_escrow_seed_rows(account_code: str, period: str) -> List[dict]:
    """
    Return seed rows for an escrow account that predate the current close period.

    Used as bootstrap history on the first run (when no prior workpaper has been
    uploaded yet).  Returns list of {date_str, desc, amt} dicts, sorted by date.
    Returns [] for unknown account codes or unparseable periods.
    """
    raw_seed = _ESCROW_SEED_MAP.get(account_code, [])
    if not raw_seed:
        return []

    yr, mo = _parse_close_period(period)
    rows: List[dict] = []
    for date_str, desc, amt in raw_seed:
        d = _parse_date(date_str)
        if d is None:
            continue
        # Include only entries strictly before the current close month
        if (d.year, d.month) < (yr, mo):
            rows.append({'date_str': d.strftime('%m/%d/%Y'), 'desc': desc, 'amt': amt,
                         '_sort': d})

    rows.sort(key=lambda r: r['_sort'])
    for r in rows:
        del r['_sort']
    return rows


def _write_tab_header(ws, account_code: str, account_name: str,
                      period: str, property_name: str, ncols: int = 5):
    """Write the standard 2-row title block at the top of every account tab."""
    ws.column_dimensions['A'].width = 2
    row = 1
    c = ws.cell(row=row, column=2, value=f'{account_code}  {account_name}')
    _apply(c, font=_font(bold=True, size=13, color='FFFFFF'), fill=_fill('375623'),
           align=Alignment(horizontal='left', vertical='center'))
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2 + ncols - 1)
    row += 1
    c = ws.cell(row=row, column=2,
                value=f'{property_name}  |  Period: {period}  |  Prepared: {datetime.now().strftime("%m/%d/%Y")}')
    _apply(c, font=_font(italic=True, size=10, color='FFFFFF'), fill=_fill('375623'),
           align=Alignment(horizontal='left', vertical='center'))
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2 + ncols - 1)
    return 3   # next available row

def _write_col_headers(ws, row: int, headers: List[str],
                       col_widths: List[int]) -> int:
    for ci, (h, w) in enumerate(zip(headers, col_widths)):
        col = 2 + ci
        c = ws.cell(row=row, column=col, value=h)
        _apply(c, font=_font(bold=True, color='FFFFFF'),
               fill=_fill('000000'), border=THIN,
               align=Alignment(horizontal='center', vertical='center'))
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 18
    return row + 1

def _write_tb_tieout(ws, row: int, gl_ending: float, tb_ending: float,
                     amount_col: int = 5) -> int:
    """Write GL/TB tie-out block at the bottom of any standard tab."""
    row += 1
    # Separator
    for col in range(2, amount_col + 1):
        ws.cell(row=row, column=col).border = THICK_BOTTOM
    row += 1

    tieout = [
        ('Ending Balance per GL',   gl_ending,  DARK_BLUE, 'FFFFFF'),
        ('Ending Balance per TB',   tb_ending,  MED_BLUE,  'FFFFFF'),
        ('Variance',                round(gl_ending - tb_ending, 2), None, None),
    ]
    for label, val, fill_hex, font_color in tieout:
        c_lbl = ws.cell(row=row, column=2, value=label)
        _apply(c_lbl,
               font=_font(bold=True, color=font_color or '000000'),
               fill=_fill(fill_hex) if fill_hex else None,
               border=THIN,
               align=Alignment(horizontal='left'))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=amount_col - 1)
        c_val = ws.cell(row=row, column=amount_col, value=val)
        variance = round(gl_ending - tb_ending, 2)
        if label == 'Variance':
            ok = abs(variance) < 0.02
            _apply(c_val,
                   font=_font(bold=True, color='006400' if ok else 'CC0000'),
                   fill=_fill(GREEN_FILL if ok else RED_FILL),
                   fmt='$#,##0.00', border=DOUBLE_BTM)
        else:
            _apply(c_val,
                   font=_font(bold=True, color=font_color or '000000'),
                   fill=_fill(fill_hex) if fill_hex else None,
                   fmt='$#,##0.00', border=THIN)
        row += 1
    return row


# ── Standard 4-column tab  (Date | Description | Entity | Amount) ─────────────

def _write_standard_tab(wb, tab_name: str, tab_color: str,
                        account_code: str, account_name: str,
                        period: str, property_name: str,
                        rows: List[Dict],  # each: {date, description, entity, amount}
                        gl_ending: float, tb_ending: float,
                        col_widths=(14, 48, 16, 18)):
    ws = wb.create_sheet(tab_name[:31])
    ws.sheet_properties.tabColor = tab_color

    next_row = _write_tab_header(ws, account_code, account_name,
                                 period, property_name, ncols=4)
    next_row += 1  # blank spacer
    next_row = _write_col_headers(ws, next_row,
                                  ['Date', 'Description', 'Entity', 'Amount'],
                                  list(col_widths))

    data_start = next_row
    for i, r in enumerate(rows):
        alt = i % 2 == 1
        bg = _fill(LIGHT_GRAY) if alt else None

        d = r.get('date')
        if isinstance(d, date):
            d = d.strftime('%m/%d/%Y')

        c1 = ws.cell(row=next_row, column=2, value=d or '')
        _apply(c1, font=_font(), fill=bg, border=THIN)

        c2 = ws.cell(row=next_row, column=3, value=r.get('description', ''))
        _apply(c2, font=_font(), fill=bg, border=THIN,
               align=Alignment(wrap_text=True))

        c3 = ws.cell(row=next_row, column=4, value=r.get('entity', _ENTITY))
        _apply(c3, font=_font(), fill=bg, border=THIN)

        amt = r.get('amount', 0) or 0
        c4 = ws.cell(row=next_row, column=5, value=amt)
        _apply(c4, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))
        next_row += 1

    _write_tb_tieout(ws, next_row, gl_ending, tb_ending, amount_col=5)
    return ws


# ── Shared GL-based escrow / reserve tab builder ─────────────────────────────

def _group_escrow_gl_transactions(gl_acct) -> list:
    """
    Group GL transactions for an escrow/reserve account by (date, clean_desc),
    net debit-credit, and discard near-zero groups (APPLY clearing pairs).

    Same logic as DACA grouping — gives one clean row per real event rather
    than regurgitating every raw GL sub-line.

    Returns list of dicts: {date_str, desc, amt, sort_date} sorted by date.
    """
    from collections import defaultdict
    from datetime import date as _date

    net_by_key: Dict[tuple, float] = defaultdict(float)
    meta:       Dict[tuple, dict]  = {}

    for txn in (getattr(gl_acct, 'transactions', []) or []):
        d = getattr(txn, 'date', None)
        if isinstance(d, _date):
            date_str  = d.strftime('%m/%d/%Y')
            sort_date = d
        else:
            date_str  = str(d or '')
            sort_date = _date(1900, 1, 1)

        desc_raw   = str(getattr(txn, 'description', '') or '')
        desc_clean = re.sub(r'\s*\([tv]\d+\)\s*$', '', desc_raw).strip()

        debit  = float(getattr(txn, 'debit',  0) or 0)
        credit = float(getattr(txn, 'credit', 0) or 0)

        key = (sort_date, desc_clean)
        net_by_key[key] = round(net_by_key[key] + debit - credit, 2)
        if key not in meta:
            meta[key] = {'date_str': date_str, 'sort_date': sort_date,
                         'desc': desc_clean}

    rows = []
    for key in sorted(net_by_key.keys()):
        net = net_by_key[key]
        if abs(net) < 0.01:
            continue   # APPLY pairs — skip
        info = meta[key]
        rows.append({'date_str': info['date_str'], 'desc': info['desc'],
                     'amt': net, 'sort_date': info['sort_date']})
    return rows


def _build_escrow_tab(wb, account_code: str, account_name: str, tab_color: str,
                      period: str, property_name: str,
                      gl_acct, tb_entry,
                      berkadia_balance_key: str = '',
                      berkadia_loans: list = None,
                      prior_rows: list = None) -> Any:
    """
    Generic workpaper tab for escrow and reserve accounts (115200, 115300, 115600).

    Full carry-forward design — each monthly workpaper shows the complete
    transaction history from inception, not just the current period:

      Balance Forward  = $0.00 when prior history is available
                       = gl_beg when no history (first run, no seeds or prior rows)
      [one grouped row per real event — duplicates / APPLY pairs removed]
        • prior_rows: full detail rows read from the prior workpaper tab
        • seed rows:  bootstrap ledger (first run, no prior workpaper uploaded)
        • current GL: grouped transactions for the close period
      Running balance formula: =F_prev + E_cur  (chains through all rows)
      ──────────────────────────────────────────
      Berkadia Reconciliation block (if berkadia_loans provided):
        Balance per Berkadia Statement   $X
        Balance per GL                   $X
        Difference                       $0
      GL / TB tie-out

    Columns: Date | Description | Entity | Amount | Running Balance

    prior_rows: list of {date_str, desc, amt} dicts from _read_escrow_tab_detail();
                pass None when no prior workpaper is available (triggers seed fallback).
    """
    gl_ending = float(getattr(gl_acct, 'ending_balance', 0) or 0)
    gl_beg    = float(getattr(gl_acct, 'beginning_balance', 0) or 0)
    tb_ending = float(getattr(tb_entry, 'ending_balance', 0) or 0) if tb_entry else gl_ending

    AMT_COL = 5   # E
    BAL_COL = 6   # F

    ws = wb.create_sheet(f'{account_code} {account_name}'[:31])
    ws.sheet_properties.tabColor = tab_color

    next_row = _write_tab_header(ws, account_code, account_name,
                                 period, property_name, ncols=5)
    next_row += 1
    next_row = _write_col_headers(
        ws, next_row,
        ['Date', 'Description', 'Entity', 'Amount', 'Running Balance'],
        [14, 52, 14, 18, 18],
    )

    # ── Determine historical rows and starting balance ────────────────────────
    # Priority: prior workpaper rows → seed bootstrap → current-period only
    if prior_rows is not None:
        # Full carry-forward: prior workpaper provided all history from inception
        hist_rows = prior_rows          # list of {date_str, desc, amt}
        starting_balance = 0.0
    else:
        # First run: fall back to seed data for this account
        hist_rows = _get_escrow_seed_rows(account_code, period)
        starting_balance = 0.0 if hist_rows else gl_beg

    # Current period GL entries (grouped, current close month only)
    current_gl_rows = _group_escrow_gl_transactions(gl_acct)

    # Combined list — historical first, then current period
    all_rows = [
        {'date_str': r['date_str'], 'desc': r['desc'], 'amt': r['amt']}
        for r in hist_rows
    ] + [
        {'date_str': r['date_str'], 'desc': r['desc'], 'amt': r['amt']}
        for r in current_gl_rows
    ]

    # ── Balance Forward ───────────────────────────────────────────────────────
    c_beg = ws.cell(row=next_row, column=3, value='Balance Forward')
    _apply(c_beg, font=_font(italic=True, color='444444'), border=THIN,
           align=Alignment(horizontal='left'))
    ws.merge_cells(start_row=next_row, start_column=3,
                   end_row=next_row, end_column=BAL_COL - 1)
    c_beg_bal = ws.cell(row=next_row, column=BAL_COL, value=starting_balance)
    _apply(c_beg_bal, font=_font(bold=True), fill=_fill(LIGHT_BLUE),
           fmt='$#,##0.00', border=THIN, align=Alignment(horizontal='right'))
    next_row += 1

    # ── All transaction rows (history + current period) ───────────────────────
    for i, r in enumerate(all_rows):
        alt = i % 2 == 1
        bg  = _fill(LIGHT_GRAY) if alt else None
        is_neg = r['amt'] < 0

        c1 = ws.cell(row=next_row, column=2, value=r['date_str'])
        _apply(c1, font=_font(), fill=bg, border=THIN)

        c2 = ws.cell(row=next_row, column=3, value=r['desc'])
        _apply(c2, font=_font(bold=is_neg), fill=bg, border=THIN,
               align=Alignment(wrap_text=True))

        c3 = ws.cell(row=next_row, column=4, value=_ENTITY)
        _apply(c3, font=_font(), fill=bg, border=THIN)

        c4 = ws.cell(row=next_row, column=AMT_COL, value=r['amt'])
        _apply(c4, font=_font(bold=is_neg, color='CC0000' if is_neg else '000000'),
               fill=bg, fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))

        prev_ref = f'F{next_row - 1}'
        formula  = f'={prev_ref}+IFERROR(E{next_row},0)'
        c5 = ws.cell(row=next_row, column=BAL_COL, value=formula)
        _apply(c5, font=_font(bold=True), fill=_fill(LIGHT_BLUE),
               fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))

        next_row += 1

    if not all_rows:
        c = ws.cell(row=next_row, column=2, value='No GL activity this period')
        _apply(c, font=_font(italic=True, color='666666'))
        next_row += 1

    # ── Berkadia statement reconciliation block ───────────────────────────────
    berkadia_bal = None
    berkadia_date = ''
    if berkadia_balance_key and berkadia_loans:
        for loan in (berkadia_loans or []):
            val = loan.get(berkadia_balance_key)
            if val:
                berkadia_bal = (berkadia_bal or 0) + float(val)
                berkadia_date = berkadia_date or (loan.get('as_of_date') or '')

    if berkadia_bal is not None:
        next_row += 1
        date_label = f' as of {berkadia_date}' if berkadia_date else ''
        rec_hdr = ws.cell(row=next_row, column=2, value='Berkadia Reconciliation')
        _apply(rec_hdr, font=_font(bold=True, size=10, color='FFFFFF'),
               fill=_fill(DARK_BLUE), border=THIN,
               align=Alignment(horizontal='left'))
        ws.merge_cells(start_row=next_row, start_column=2,
                       end_row=next_row, end_column=BAL_COL)
        next_row += 1

        diff = round(gl_ending - berkadia_bal, 2)
        rec_rows = [
            (f'Balance per Berkadia Statement{date_label}', berkadia_bal, DARK_BLUE, 'FFFFFF'),
            ('Balance per GL',                              gl_ending,    MED_BLUE,  'FFFFFF'),
            ('Difference',                                  diff,         None,      None),
        ]
        for label, val, fill_hex, font_color in rec_rows:
            c_lbl = ws.cell(row=next_row, column=2, value=label)
            _apply(c_lbl,
                   font=_font(bold=True, color=font_color or '000000'),
                   fill=_fill(fill_hex) if fill_hex else None,
                   border=THIN, align=Alignment(horizontal='left'))
            ws.merge_cells(start_row=next_row, start_column=2,
                           end_row=next_row, end_column=BAL_COL - 1)
            c_val = ws.cell(row=next_row, column=BAL_COL, value=val)
            if label == 'Difference':
                ok = abs(val) < 0.02
                _apply(c_val,
                       font=_font(bold=True, color='006400' if ok else 'CC0000'),
                       fill=_fill(GREEN_FILL if ok else RED_FILL),
                       fmt='$#,##0.00', border=DOUBLE_BTM)
            else:
                _apply(c_val,
                       font=_font(bold=True, color=font_color or '000000'),
                       fill=_fill(fill_hex) if fill_hex else None,
                       fmt='$#,##0.00', border=THIN)
            next_row += 1

    # ── GL / TB tie-out ───────────────────────────────────────────────────────
    next_row += 1
    _write_tb_tieout(ws, next_row, gl_ending, tb_ending, amount_col=BAL_COL)
    return ws


# ── 115200 — RET Escrow ───────────────────────────────────────────────────────

def build_115200_tab(wb, period: str, property_name: str,
                     gl_acct=None, tb_entry=None,
                     berkadia_loans: list = None,
                     prior_tab_detail: dict = None, **_):
    return _build_escrow_tab(wb, '115200', 'RET Escrow', '4472C4',
                             period, property_name, gl_acct, tb_entry,
                             berkadia_balance_key='tax_escrow_balance',
                             berkadia_loans=berkadia_loans,
                             prior_rows=(prior_tab_detail or {}).get('115200'))


# ── 115300 — Insurance Escrow ─────────────────────────────────────────────────

def build_115300_tab(wb, period: str, property_name: str,
                     gl_acct=None, tb_entry=None,
                     berkadia_loans: list = None,
                     prior_tab_detail: dict = None, **_):
    return _build_escrow_tab(wb, '115300', 'Insurance Escrow', '4472C4',
                             period, property_name, gl_acct, tb_entry,
                             berkadia_balance_key='insurance_escrow_balance',
                             berkadia_loans=berkadia_loans,
                             prior_rows=(prior_tab_detail or {}).get('115300'))


# ── 115600 — Loan Reserve ────────────────────────────────────────────────────

def build_115600_tab(wb, period: str, property_name: str,
                     gl_acct=None, tb_entry=None,
                     berkadia_loans: list = None,
                     prior_tab_detail: dict = None, **_):
    return _build_escrow_tab(wb, '115600', 'Loan Reserve', '4472C4',
                             period, property_name, gl_acct, tb_entry,
                             berkadia_balance_key='reserve_balance',
                             berkadia_loans=berkadia_loans,
                             prior_rows=(prior_tab_detail or {}).get('115600'))


# ── 133100 — AR Other ────────────────────────────────────────────────────────

_VAGUE_PATTERNS = re.compile(
    r'^(adj|adjustment|misc|miscellaneous|other|entry|je|reclassif|reclass|'
    r'correction|see note|n/a|tbd|unknown|manual)$',
    re.IGNORECASE,
)

def _is_vague(description: str) -> bool:
    if not description:
        return True
    desc = description.strip()
    if len(desc) <= 3:
        return True
    return bool(_VAGUE_PATTERNS.match(desc))

def build_133100_tab(wb, period: str, property_name: str,
                     gl_acct=None, tb_entry=None, **_):
    gl_ending = float(getattr(gl_acct, 'ending_balance', 0) or 0)
    tb_ending = float(getattr(tb_entry, 'ending_balance', 0) or 0) if tb_entry else gl_ending

    txns = list(getattr(gl_acct, 'transactions', []) or [])

    ws = wb.create_sheet('133100 AR Other'[:31])
    ws.sheet_properties.tabColor = 'BF8F00'

    next_row = _write_tab_header(ws, '133100', 'Accounts Receivable - Other',
                                 period, property_name, ncols=5)
    next_row += 1
    next_row = _write_col_headers(
        ws, next_row,
        ['Date', 'Description', f'Entity ({_ENTITY})', 'Amount', 'Flag'],
        [14, 52, 14, 18, 22],
    )

    for i, txn in enumerate(txns):
        alt = i % 2 == 1
        bg = _fill(LIGHT_GRAY) if alt else None

        d = getattr(txn, 'date', None)
        date_str = d.strftime('%m/%d/%Y') if isinstance(d, date) else str(d or '')
        desc = str(getattr(txn, 'description', '') or '')
        amt  = float(getattr(txn, 'debit', 0) or 0) - float(getattr(txn, 'credit', 0) or 0)
        flag = '⚠ Review description' if _is_vague(desc) else ''

        c1 = ws.cell(row=next_row, column=2, value=date_str)
        _apply(c1, font=_font(), fill=bg, border=THIN)
        c2 = ws.cell(row=next_row, column=3, value=desc)
        _apply(c2, font=_font(), fill=bg, border=THIN, align=Alignment(wrap_text=True))
        c3 = ws.cell(row=next_row, column=4, value=_ENTITY)
        _apply(c3, font=_font(), fill=bg, border=THIN)
        c4 = ws.cell(row=next_row, column=5, value=amt)
        _apply(c4, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))
        c5 = ws.cell(row=next_row, column=6, value=flag)
        if flag:
            _apply(c5, font=_font(bold=True, color='9C0006'),
                   fill=_fill('FFC7CE'), border=THIN)
        else:
            _apply(c5, font=_font(), fill=bg, border=THIN)
        next_row += 1

    if not txns:
        c = ws.cell(row=next_row, column=2, value='No activity this period')
        _apply(c, font=_font(italic=True, color='666666'))
        next_row += 1

    _write_tb_tieout(ws, next_row, gl_ending, tb_ending, amount_col=5)
    return ws


# ── 133110 — AR Tenant Billback ───────────────────────────────────────────────

def build_133110_tab(wb, period: str, property_name: str,
                     gl_acct=None, tb_entry=None,
                     je_lines: List[Dict] = None, **_):
    """
    Builds from Pass 1 JEs tagged source='tenant_utility_billing'.
    Description reformatted to: "Accrued: [period] [tenant name]"
    """
    gl_ending = float(getattr(gl_acct, 'ending_balance', 0) or 0)
    tb_ending = float(getattr(tb_entry, 'ending_balance', 0) or 0) if tb_entry else gl_ending

    # Filter to billback JEs touching 133110
    billback_rows = []
    for je in (je_lines or []):
        if str(je.get('account_code', '')).strip() == '133110':
            amt  = float(je.get('debit', 0) or 0) - float(je.get('credit', 0) or 0)
            raw_desc = str(je.get('description', '') or '')
            vendor   = str(je.get('vendor', '') or '')
            # Reformat: "Accrued: Mar-2026 Accent Therapeutics"
            tenant = vendor or raw_desc
            desc = f'Accrued: {period} {tenant}'.strip()
            # Period of accrual = the close period; GL account from the JE
            gl_acct_code = str(je.get('account_code', '133110'))
            billback_rows.append({
                'description': desc,
                'accrual_period': period,
                'gl_account': gl_acct_code,
                'amount': amt,
            })

    ws = wb.create_sheet('133110 AR Billback'[:31])
    ws.sheet_properties.tabColor = 'BF8F00'

    next_row = _write_tab_header(ws, '133110', 'AR - Tenant Billback',
                                 period, property_name, ncols=4)
    next_row += 1
    next_row = _write_col_headers(
        ws, next_row,
        ['Description', 'Accrual Period', 'GL Account', 'Amount'],
        [52, 18, 14, 18],
    )

    for i, r in enumerate(billback_rows):
        alt = i % 2 == 1
        bg = _fill(LIGHT_GRAY) if alt else None

        c1 = ws.cell(row=next_row, column=2, value=r['description'])
        _apply(c1, font=_font(), fill=bg, border=THIN, align=Alignment(wrap_text=True))
        c2 = ws.cell(row=next_row, column=3, value=r['accrual_period'])
        _apply(c2, font=_font(), fill=bg, border=THIN)
        c3 = ws.cell(row=next_row, column=4, value=r['gl_account'])
        _apply(c3, font=_font(), fill=bg, border=THIN)
        c4 = ws.cell(row=next_row, column=5, value=r['amount'])
        _apply(c4, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))
        next_row += 1

    if not billback_rows:
        c = ws.cell(row=next_row, column=2, value='No tenant billback JEs this period')
        _apply(c, font=_font(italic=True, color='666666'))
        next_row += 1

    _write_tb_tieout(ws, next_row, gl_ending, tb_ending, amount_col=5)
    return ws


# ── 213100 — Accrued Expenses ─────────────────────────────────────────────────

def _213100_clean_desc(txn) -> str:
    """Best description from a GL transaction — strips Yardi JE control prefixes."""
    remarks = str(getattr(txn, 'remarks',      '') or '').strip()
    desc    = str(getattr(txn, 'description',  '') or '').strip()
    raw = remarks or desc
    # Strip "Reversal of J-XXXXX: " and bare "J-XXXXX: " auto-fill prefixes
    raw = re.sub(r'^(reversal\s+of\s+)?[Jj]-\d+\s*[:\-–]\s+',
                 '', raw, flags=re.IGNORECASE).strip()
    return raw or desc


def build_213100_tab(wb, period: str, property_name: str,
                     gl_acct=None, tb_entry=None, je_lines=None, **_):
    """
    One row per current-period accrual posted to account 213100.

    Only credit entries are shown (CR 213100 = new accrual created this period).
    Debit entries are prior-month auto-reversals and are intentionally excluded.

    Columns:  Date | GL Account | Description | Vendor | Amount

    GL Account is resolved by matching the transaction's control number against
    the debit side of the same JE in je_lines (cross-account context).

    Header (B1 = '213100' alone so the footer VLOOKUP resolves correctly):
      B1 = '213100'  ← VLOOKUP anchor for tieout
      B2 = 'Accrued Expenses'
      B3 = property | period | prepared

    Footer (live formulas):
      Ending Balance per GL  = =SUM(F6:F{last_data_row})
      Ending Balance per TB  = =VLOOKUP(B1,'Trial Balance'!$B:$G,6,0)
      Variance               = =F{gl_row}-F{tb_row}
    """
    HDR_GREEN = '375623'

    txns = list(getattr(gl_acct, 'transactions', []) or [])

    # ── Build control → expense account / vendor lookup ───────────────────────
    _LIABILITY_ACCOUNTS = {'211100', '211200', '211300', '213100', '213200', '221100'}
    _ctrl_to_expense: Dict[str, str] = {}
    _ctrl_to_vendor:  Dict[str, str] = {}

    for line in (je_lines or []):
        ctrl = str(line.get('je_number') or '').strip()
        if not ctrl:
            continue
        acct = str(line.get('account_code') or '').strip()
        dbt  = float(line.get('debit', 0) or 0)
        if dbt > 0 and acct not in _LIABILITY_ACCOUNTS and ctrl not in _ctrl_to_expense:
            _ctrl_to_expense[ctrl] = acct
            vendor_val = str(line.get('vendor') or line.get('description') or '').strip()
            _ctrl_to_vendor[ctrl] = vendor_val

    # ── Filter to credit-only entries (current-period accruals) ───────────────
    def _txn_sort_key(t):
        d = getattr(t, 'date', None)
        return d if isinstance(d, date) else date(1900, 1, 1)

    accrual_txns = sorted(
        [t for t in txns
         if float(getattr(t, 'credit', 0) or 0) > float(getattr(t, 'debit', 0) or 0)],
        key=_txn_sort_key,
    )

    ws = wb.create_sheet('213100 Accr Exp'[:31])
    ws.sheet_properties.tabColor = 'FF0000'
    ws.column_dimensions['A'].width = 2

    FIRST_COL  = 2   # B
    AMOUNT_COL = 6   # F (5 data cols: B–F)

    col_labels = ['Date', 'GL Account', 'Description', 'Vendor', 'Amount']
    col_widths = [14, 14, 52, 24, 18]

    for ci, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(FIRST_COL + ci)].width = w

    # ── Rows 1-3: header block ────────────────────────────────────────────
    c1 = ws.cell(row=1, column=FIRST_COL, value='213100')
    _apply(c1, font=_font(bold=True, size=13, color='FFFFFF'),
           fill=_fill(HDR_GREEN),
           align=Alignment(horizontal='left', vertical='center'))
    ws.merge_cells(start_row=1, start_column=FIRST_COL, end_row=1, end_column=FIRST_COL + 2)
    ws.row_dimensions[1].height = 20

    c2 = ws.cell(row=2, column=FIRST_COL, value='Accrued Expenses')
    _apply(c2, font=_font(size=11, color='FFFFFF'),
           fill=_fill(HDR_GREEN),
           align=Alignment(horizontal='left', vertical='center'))
    ws.merge_cells(start_row=2, start_column=FIRST_COL, end_row=2, end_column=AMOUNT_COL)

    prop_line = (f'{property_name or "revlabpm"}  |  Period: {period}  |  '
                 f'Prepared: {datetime.now().strftime("%m/%d/%Y")}')
    c3 = ws.cell(row=3, column=FIRST_COL, value=prop_line)
    _apply(c3, font=_font(italic=True, size=10, color='FFFFFF'),
           fill=_fill(HDR_GREEN),
           align=Alignment(horizontal='left', vertical='center'))
    ws.merge_cells(start_row=3, start_column=FIRST_COL, end_row=3, end_column=AMOUNT_COL)

    # Row 4 = blank spacer

    # ── Row 5: column headers ─────────────────────────────────────────────
    for ci, lbl in enumerate(col_labels):
        c = ws.cell(row=5, column=FIRST_COL + ci, value=lbl)
        _apply(c, font=_font(bold=True, size=10, color='FFFFFF'),
               fill=_fill('000000'), border=THIN,
               align=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[5].height = 18

    # ── Data rows from row 6 ──────────────────────────────────────────────
    next_row = 6

    for i, txn in enumerate(accrual_txns):
        alt = i % 2 == 1
        bg  = _fill(LIGHT_GRAY) if alt else None

        d = getattr(txn, 'date', None)
        date_str  = d.strftime('%m/%d/%Y') if isinstance(d, date) else str(d or '')
        ctrl      = str(getattr(txn, 'control', '') or '').strip()
        gl_acct_c = _ctrl_to_expense.get(ctrl, '')
        desc      = _213100_clean_desc(txn)
        vendor    = _ctrl_to_vendor.get(ctrl, '')
        credit    = float(getattr(txn, 'credit', 0) or 0)
        debit     = float(getattr(txn, 'debit',  0) or 0)
        amt       = credit - debit

        for col, val, fmt, wrap in [
            (2, date_str,  None,        False),
            (3, gl_acct_c, None,        False),
            (4, desc,      None,        True),
            (5, vendor,    None,        False),
            (6, amt,       '$#,##0.00', False),
        ]:
            c = ws.cell(row=next_row, column=col, value=val)
            _apply(c, font=_font(size=10), fill=bg, border=THIN,
                   fmt=fmt,
                   align=Alignment(wrap_text=wrap,
                                   horizontal='right' if fmt else 'left'))
        next_row += 1

    last_data_row = max(next_row - 1, 5)

    if not accrual_txns:
        c = ws.cell(row=next_row, column=FIRST_COL, value='No accrual entries this period')
        _apply(c, font=_font(italic=True, size=10, color='666666'), border=THIN)
        ws.merge_cells(start_row=next_row, start_column=FIRST_COL,
                       end_row=next_row, end_column=AMOUNT_COL)
        next_row += 1

    # ── Footer: live formulas ─────────────────────────────────────────────
    next_row += 1   # blank spacer

    for col in range(FIRST_COL, AMOUNT_COL + 1):
        ws.cell(row=next_row, column=col).border = THICK_BOTTOM
    next_row += 1

    F = get_column_letter(AMOUNT_COL)
    gl_row  = next_row
    tb_row  = next_row + 1
    var_row = next_row + 2

    tieout_rows = [
        (gl_row,  'Ending Balance per GL',
         f'=SUM({F}6:{F}{last_data_row})',                          DARK_BLUE, 'FFFFFF'),
        (tb_row,  'Ending Balance per TB',
         f"=VLOOKUP(B1,'Trial Balance'!$B:$G,6,0)",                MED_BLUE,  'FFFFFF'),
        (var_row, 'Variance',
         f'={F}{gl_row}-{F}{tb_row}',                              None,      None),
    ]

    for row, label, formula, fill_hex, font_color in tieout_rows:
        c_lbl = ws.cell(row=row, column=FIRST_COL, value=label)
        _apply(c_lbl,
               font=_font(bold=True, size=10, color=font_color or '000000'),
               fill=_fill(fill_hex) if fill_hex else None, border=THIN,
               align=Alignment(horizontal='left'))
        ws.merge_cells(start_row=row, start_column=FIRST_COL,
                       end_row=row, end_column=AMOUNT_COL - 1)

        c_val = ws.cell(row=row, column=AMOUNT_COL, value=formula)
        _apply(c_val,
               font=_font(bold=True, size=10, color=font_color or '000000'),
               fill=_fill(fill_hex) if fill_hex else None, border=THIN,
               fmt='$#,##0.00',
               align=Alignment(horizontal='right'))

    return ws


# ── 135150 — PPD Other (Prepaid Ledger) ──────────────────────────────────────

def build_135150_tab(wb, period: str, property_name: str,
                     gl_acct=None, tb_entry=None,
                     prepaid_ledger: List[Dict] = None, **_):
    """
    Prepaid Other schedule with live Excel formulas — matching the monthly workpaper template.

    Header (B1 = '135150' alone so the footer VLOOKUP resolves correctly):
      B1 = '135150'           ← VLOOKUP anchor for tieout
      B2 = 'Prepaid - Other'
      B3 = property | period | prepared
      Row 4 = empty spacer
      Row 5 = column headers

    Columns (B–L):
      Vendor | Description | Invoice Number | Invoice Date | G/L Account |
      Start Date | End Date | Total | Monthly Amt | Amt Amort. | Remaining

    Formula columns (live, reference-able):
      J (Monthly Amt)  = =I{r}/DATEDIF(G{r},H{r}+1,"M")
      K (Amt Amort.)   = =J{r}*DATEDIF(G{r},'Summary Page'!$C$4+1,"M")
      L (Remaining)    = =I{r}-K{r}

    Footer (live formulas):
      Ending Balance per GL  = =SUM(L6:L{last_data_row})
      Ending Balance per TB  = =VLOOKUP(B1,'Trial Balance'!$B:$G,6,0)
      Variance               = =L{gl_row}-L{tb_row}

    Requires 'Summary Page'!C4 = period-end date (added by bs_workpaper_generator).
    """
    HDR_GREEN = '375623'

    ws = wb.create_sheet('135150 PPD Other'[:31])
    ws.sheet_properties.tabColor = '70AD47'
    ws.column_dimensions['A'].width = 2

    FIRST_COL = 2   # B
    LAST_COL  = 12  # L

    col_labels = ['Vendor', 'Description', 'Invoice Number', 'Invoice Date',
                  'G/L Account', 'Start Date', 'End Date', 'Total',
                  'Monthly Amt', 'Amt Amort.', 'Remaining']
    col_widths = [24, 32, 16, 14, 14, 12, 12, 16, 14, 14, 16]

    for ci, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(FIRST_COL + ci)].width = w

    # ── Rows 1-3: header block ────────────────────────────────────────────
    # B1 must contain ONLY the account code so the footer VLOOKUP works.
    c1 = ws.cell(row=1, column=FIRST_COL, value='135150')
    _apply(c1, font=_font(bold=True, size=13, color='FFFFFF'),
           fill=_fill(HDR_GREEN),
           align=Alignment(horizontal='left', vertical='center'))
    ws.merge_cells(start_row=1, start_column=FIRST_COL, end_row=1, end_column=FIRST_COL + 2)
    ws.row_dimensions[1].height = 20

    c2 = ws.cell(row=2, column=FIRST_COL, value='Prepaid - Other')
    _apply(c2, font=_font(size=11, color='FFFFFF'),
           fill=_fill(HDR_GREEN),
           align=Alignment(horizontal='left', vertical='center'))
    ws.merge_cells(start_row=2, start_column=FIRST_COL, end_row=2, end_column=LAST_COL)

    prop_line = (f'{property_name or "revlabpm"}  |  Period: {period}  |  '
                 f'Prepared: {datetime.now().strftime("%m/%d/%Y")}')
    c3 = ws.cell(row=3, column=FIRST_COL, value=prop_line)
    _apply(c3, font=_font(italic=True, size=10, color='FFFFFF'),
           fill=_fill(HDR_GREEN),
           align=Alignment(horizontal='left', vertical='center'))
    ws.merge_cells(start_row=3, start_column=FIRST_COL, end_row=3, end_column=LAST_COL)

    # Row 4 = blank spacer

    # ── Row 5: column headers ─────────────────────────────────────────────
    for ci, lbl in enumerate(col_labels):
        c = ws.cell(row=5, column=FIRST_COL + ci, value=lbl)
        _apply(c, font=_font(bold=True, size=10, color='FFFFFF'),
               fill=_fill('000000'), border=THIN,
               align=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[5].height = 18

    # ── Data rows from row 6 ──────────────────────────────────────────────
    def _to_date(val):
        """Coerce various date representations to a Python date (needed for DATEDIF)."""
        if isinstance(val, date):
            return val
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, str) and val:
            for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%Y/%m/%d'):
                try:
                    return datetime.strptime(val.strip(), fmt).date()
                except ValueError:
                    pass
        return None

    ledger   = prepaid_ledger or []
    next_row = 6

    for i, item in enumerate(ledger):
        alt = i % 2 == 1
        bg  = _fill(LIGHT_GRAY) if alt else None

        def _v(key, _item=item):
            return _item.get(key, '') if isinstance(_item, dict) else getattr(_item, key, '')

        vendor     = str(_v('vendor') or _v('description') or '')
        desc       = str(_v('description') or '')
        inv_num    = str(_v('invoice_number') or '')
        inv_date   = _v('invoice_date') or ''
        gl_account = str(_v('gl_account_number') or '')
        total      = float(_v('total_amount') or 0)
        start_d    = _to_date(_v('service_start') or _v('start_date') or _v('first_added_period'))
        end_d      = _to_date(_v('service_end') or _v('end_date'))

        r = next_row

        # Static input columns B–I (values; J/K/L are formula columns)
        for col, val, fmt, wrap in [
            (FIRST_COL + 0, vendor,    None,         True),
            (FIRST_COL + 1, desc,      None,         True),
            (FIRST_COL + 2, inv_num,   None,         False),
            (FIRST_COL + 3, inv_date,  'MM/DD/YYYY', False),
            (FIRST_COL + 4, gl_account, None,        False),
            (FIRST_COL + 5, start_d,   'MM/DD/YYYY', False),
            (FIRST_COL + 6, end_d,     'MM/DD/YYYY', False),
            (FIRST_COL + 7, total,     '$#,##0.00',  False),
        ]:
            c = ws.cell(row=r, column=col, value=val)
            _apply(c, font=_font(size=10), fill=bg, border=THIN,
                   fmt=fmt,
                   align=Alignment(wrap_text=wrap,
                                   horizontal='right' if fmt else 'left'))

        # Formula columns J, K, L
        G = get_column_letter(FIRST_COL + 5)   # Start Date
        H = get_column_letter(FIRST_COL + 6)   # End Date
        I = get_column_letter(FIRST_COL + 7)   # Total
        J = get_column_letter(FIRST_COL + 8)   # Monthly Amt
        K = get_column_letter(FIRST_COL + 9)   # Amt Amort.

        for col, formula in [
            (FIRST_COL + 8, f'={I}{r}/DATEDIF({G}{r},{H}{r}+1,"M")'),
            (FIRST_COL + 9, f'={J}{r}*DATEDIF({G}{r},\'Summary Page\'!$C$4+1,"M")'),
            (FIRST_COL + 10, f'={I}{r}-{K}{r}'),
        ]:
            c = ws.cell(row=r, column=col, value=formula)
            _apply(c, font=_font(size=10), fill=bg, border=THIN,
                   fmt='$#,##0.00',
                   align=Alignment(horizontal='right'))

        next_row += 1

    last_data_row = max(next_row - 1, 5)   # guard: SUM(L6:L5) = 0 when empty

    if not ledger:
        c = ws.cell(row=next_row, column=FIRST_COL, value='No active prepaid items')
        _apply(c, font=_font(italic=True, size=10, color='666666'))
        next_row += 1

    # ── Footer: live formulas ─────────────────────────────────────────────
    next_row += 1   # blank spacer

    for col in range(FIRST_COL, LAST_COL + 1):
        ws.cell(row=next_row, column=col).border = THICK_BOTTOM
    next_row += 1

    L = get_column_letter(LAST_COL)
    gl_row  = next_row
    tb_row  = next_row + 1
    var_row = next_row + 2

    tieout_rows = [
        (gl_row,  'Ending Balance per GL',
         f'=SUM({L}6:{L}{last_data_row})',                          DARK_BLUE, 'FFFFFF'),
        (tb_row,  'Ending Balance per TB',
         f"=VLOOKUP(B1,'Trial Balance'!$B:$G,6,0)",                MED_BLUE,  'FFFFFF'),
        (var_row, 'Variance',
         f'={L}{gl_row}-{L}{tb_row}',                              None,      None),
    ]

    for row, label, formula, fill_hex, font_color in tieout_rows:
        c_lbl = ws.cell(row=row, column=FIRST_COL, value=label)
        _apply(c_lbl,
               font=_font(bold=True, size=10, color=font_color or '000000'),
               fill=_fill(fill_hex) if fill_hex else None, border=THIN,
               align=Alignment(horizontal='left'))
        ws.merge_cells(start_row=row, start_column=FIRST_COL,
                       end_row=row, end_column=LAST_COL - 1)

        c_val = ws.cell(row=row, column=LAST_COL, value=formula)
        _apply(c_val,
               font=_font(bold=True, size=10, color=font_color or '000000'),
               fill=_fill(fill_hex) if fill_hex else None, border=THIN,
               fmt='$#,##0.00',
               align=Alignment(horizontal='right'))

    return ws


# ── 115100 — DACA Restricted Cash ────────────────────────────────────────────

def _group_daca_gl_transactions(gl_acct) -> list:
    """
    Group the Yardi GL transactions for account 115100 by (date, tenant/description)
    and return a list of row dicts ready for the workpaper.

    Yardi posts many individual lines per tenant per day (charge-code breakdown
    plus APPLY clearing entries that net to zero).  We group by (date, clean_desc),
    net the debits and credits, then discard groups that net to zero (APPLY pairs).

    Classification:
      - "Sweep" in description  → Adjustments (negative, labelled "Transfer to PNC")
      - "bank fee" / "service"  → Adjustments (negative, labelled "Bank Fee")
      - net > 0                 → Deposits (tenant name as label)
      - net < 0 (other)         → Disbursements

    Returns list of dicts: {date, desc, deposits, disb, adj, sort_date}
    sorted chronologically.
    """
    from collections import defaultdict
    from datetime import date as _date

    net_by_key: Dict[tuple, float] = defaultdict(float)
    meta: Dict[tuple, dict] = {}   # key -> {date_str, sort_date, desc}

    for txn in (getattr(gl_acct, 'transactions', []) or []):
        d = getattr(txn, 'date', None)
        if isinstance(d, _date):
            date_str  = d.strftime('%m/%d/%Y')
            sort_date = d
        else:
            date_str  = str(d or '')
            sort_date = _date(1900, 1, 1)

        desc_raw   = str(getattr(txn, 'description', '') or '')
        # Strip Yardi tenant codes "(t0000011)" / "(v0000123)"
        desc_clean = re.sub(r'\s*\([tv]\d+\)\s*$', '', desc_raw).strip()

        debit  = float(getattr(txn, 'debit',  0) or 0)
        credit = float(getattr(txn, 'credit', 0) or 0)

        key = (sort_date, desc_clean)
        net_by_key[key] = round(net_by_key[key] + debit - credit, 2)
        if key not in meta:
            meta[key] = {'date_str': date_str, 'sort_date': sort_date,
                         'desc': desc_clean}

    rows = []
    for key in sorted(net_by_key.keys()):
        net  = net_by_key[key]
        info = meta[key]
        desc = info['desc']

        if abs(net) < 0.01:
            continue   # APPLY pairs net to zero — skip

        desc_lower = desc.lower()
        is_sweep   = 'sweep' in desc_lower or 'transfer to pnc' in desc_lower
        is_fee     = any(kw in desc_lower for kw in
                         ('bank fee', 'analysis service', 'service chg', 'monthly bank'))

        if is_sweep:
            rows.append({'date': info['date_str'], 'desc': 'Transfer to PNC',
                         'deposits': 0.0, 'disb': 0.0, 'adj': net,
                         'sort_date': info['sort_date']})
        elif is_fee:
            rows.append({'date': info['date_str'], 'desc': 'Bank Fee',
                         'deposits': 0.0, 'disb': 0.0, 'adj': net,
                         'sort_date': info['sort_date']})
        elif net > 0:
            rows.append({'date': info['date_str'], 'desc': desc,
                         'deposits': net, 'disb': 0.0, 'adj': 0.0,
                         'sort_date': info['sort_date']})
        else:
            rows.append({'date': info['date_str'], 'desc': desc,
                         'deposits': 0.0, 'disb': abs(net), 'adj': 0.0,
                         'sort_date': info['sort_date']})

    return rows


def build_115100_tab(wb, period: str, property_name: str,
                     gl_acct=None, tb_entry=None,
                     daca_data: Dict = None,
                     **_):
    """
    Rolling balance workpaper for DACA (KeyBank x5132 → GL 115100).

    Columns: Date | Tenant / Description | Deposits | Disbursements |
             Adjustments | Ending Balance

    GL transactions are grouped by (date, tenant name) so the workpaper
    matches the hand-prepared format:
      • One row per tenant per day (net of all charge-code sub-lines)
      • Sweeps to PNC → "Transfer to PNC" in the Adjustments column
      • Bank fees     → "Bank Fee" in the Adjustments column
      • Ending Balance is an Excel running-balance formula referencing
        the row above so users can inspect calculations in Excel.

    Reconciliation footer (if daca_data from YardiDACARec / KeyBankDACA is
    provided):
      Bank Statement Balance | GL Ending Balance | Difference
    """
    gl_ending = float(getattr(gl_acct, 'ending_balance', 0) or 0)
    gl_beg    = float(getattr(gl_acct, 'beginning_balance', 0) or 0)
    tb_ending = float(getattr(tb_entry, 'ending_balance', 0) or 0) if tb_entry else gl_ending

    ws = wb.create_sheet('115100 DACA'[:31])
    ws.sheet_properties.tabColor = '2D6F50'

    next_row = _write_tab_header(ws, '115100', 'Restricted Cash - DACA (KeyBank x5132)',
                                 period, property_name, ncols=6)
    next_row += 1
    next_row = _write_col_headers(
        ws, next_row,
        ['Date', 'Tenant / Description', 'Deposits', 'Disbursements', 'Adjustments', 'Ending Balance'],
        [14, 44, 16, 18, 16, 18],
    )

    # ── Beginning balance row ─────────────────────────────────────────────────
    c_beg = ws.cell(row=next_row, column=3, value='Balance Forward')
    _apply(c_beg, font=_font(italic=True, color='444444'), border=THIN,
           align=Alignment(horizontal='left'))
    ws.merge_cells(start_row=next_row, start_column=3,
                   end_row=next_row, end_column=6)
    c_beg_bal = ws.cell(row=next_row, column=7, value=gl_beg)
    _apply(c_beg_bal, font=_font(bold=True), fill=_fill(LIGHT_BLUE),
           fmt='$#,##0.00', border=THIN, align=Alignment(horizontal='right'))
    beg_balance_row = next_row
    next_row += 1

    # ── Build data rows from GL transactions (grouped by date × tenant) ───────
    daca_rows = _group_daca_gl_transactions(gl_acct)

    # Write rows with running Ending Balance formula
    for i, r in enumerate(daca_rows):
        alt = i % 2 == 1
        bg  = _fill(LIGHT_GRAY) if alt else None

        c1 = ws.cell(row=next_row, column=2, value=r['date'])
        _apply(c1, font=_font(), fill=bg, border=THIN)

        c2 = ws.cell(row=next_row, column=3, value=r['desc'])
        _apply(c2, font=_font(), fill=bg, border=THIN)

        c3 = ws.cell(row=next_row, column=4, value=r['deposits'] or None)
        _apply(c3, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))

        c4 = ws.cell(row=next_row, column=5, value=r['disb'] or None)
        _apply(c4, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))

        c5 = ws.cell(row=next_row, column=6, value=r['adj'] or None)
        _apply(c5, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))

        # Ending balance formula: prior G + deposits(D) - disbursements(E) + adjustments(F)
        prev_ref = f'G{next_row - 1}'
        formula  = (f'={prev_ref}'
                    f'+IFERROR(D{next_row},0)'
                    f'-IFERROR(E{next_row},0)'
                    f'+IFERROR(F{next_row},0)')
        c6 = ws.cell(row=next_row, column=7, value=formula)
        _apply(c6, font=_font(bold=True), fill=_fill(LIGHT_BLUE),
               fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))

        next_row += 1

    if not daca_rows:
        c = ws.cell(row=next_row, column=2,
                    value='No DACA GL activity for this period')
        _apply(c, font=_font(italic=True, color='666666'))
        next_row += 1

    # ── Totals row ────────────────────────────────────────────────────────────
    next_row += 1
    c_tot_lbl = ws.cell(row=next_row, column=2, value='Period Totals')
    _apply(c_tot_lbl, font=_font(bold=True), border=THICK_BOTTOM)

    data_start = beg_balance_row + 1
    data_end   = next_row - 2

    for col_letter, col_idx in [('D', 4), ('E', 5), ('F', 6)]:
        if data_end >= data_start:
            tot_formula = f'=SUM({col_letter}{data_start}:{col_letter}{data_end})'
        else:
            tot_formula = 0
        ct = ws.cell(row=next_row, column=col_idx, value=tot_formula)
        _apply(ct, font=_font(bold=True), fmt='$#,##0.00', border=THICK_BOTTOM,
               align=Alignment(horizontal='right'))

    # Ending balance column just shows the last GL ending balance
    ct_end = ws.cell(row=next_row, column=7, value=gl_ending)
    _apply(ct_end, font=_font(bold=True), fill=_fill(LIGHT_BLUE),
           fmt='$#,##0.00', border=THICK_BOTTOM,
           align=Alignment(horizontal='right'))
    next_row += 2

    # ── Bank reconciliation block ─────────────────────────────────────────────
    # Uses daca_data from YardiDACARec or KeyBankDACA parser
    bank_bal = None
    rec_diff = None
    if daca_data and isinstance(daca_data, dict):
        bank_bal = (daca_data.get('reconciled_bank_balance')
                    or daca_data.get('ending_balance'))
        rec_diff = daca_data.get('reconciling_difference')

    rec_rows = [
        ('Bank Statement Balance (Reconciled)',
         bank_bal if bank_bal is not None else '— upload DACA rec to populate —',
         DARK_BLUE, 'FFFFFF'),
        ('Ending Balance per GL',
         gl_ending,
         MED_BLUE,  'FFFFFF'),
        ('Difference',
         round(gl_ending - (bank_bal or gl_ending), 2),
         None, None),
    ]
    c_rec_hdr = ws.cell(row=next_row, column=2, value='Bank Reconciliation')
    _apply(c_rec_hdr, font=_font(bold=True, size=10), border=THIN)
    ws.merge_cells(start_row=next_row, start_column=2,
                   end_row=next_row, end_column=6)
    next_row += 1

    for label, val, fill_hex, font_color in rec_rows:
        c_lbl = ws.cell(row=next_row, column=2, value=label)
        _apply(c_lbl,
               font=_font(bold=True, color=font_color or '000000'),
               fill=_fill(fill_hex) if fill_hex else None,
               border=THIN,
               align=Alignment(horizontal='left'))
        ws.merge_cells(start_row=next_row, start_column=2,
                       end_row=next_row, end_column=6)
        c_val = ws.cell(row=next_row, column=7, value=val)
        if label == 'Difference':
            ok = isinstance(val, (int, float)) and abs(val) < 0.02
            _apply(c_val,
                   font=_font(bold=True, color='006400' if ok else 'CC0000'),
                   fill=_fill(GREEN_FILL if ok else RED_FILL),
                   fmt='$#,##0.00' if isinstance(val, (int, float)) else None,
                   border=DOUBLE_BTM)
        elif isinstance(val, (int, float)):
            _apply(c_val,
                   font=_font(bold=True, color=font_color or '000000'),
                   fill=_fill(fill_hex) if fill_hex else None,
                   fmt='$#,##0.00', border=THIN)
        else:
            _apply(c_val,
                   font=_font(italic=True, color='888888'),
                   border=THIN)
        next_row += 1

    # ── TB tie-out ────────────────────────────────────────────────────────────
    next_row += 1
    _write_tb_tieout(ws, next_row, gl_ending, tb_ending, amount_col=7)
    return ws


# ── 111100 — Operating Cash (PNC Bank Rec) ───────────────────────────────────

def _group_cash_gl_transactions(gl_acct) -> list:
    """
    Group 111100 GL transactions by (date, clean description) and net debit-credit.

    For operating cash:
      - net > 0  → Receipts column  (DR = cash received)
      - net < 0  → Disbursements column (CR = cash paid out; stored as positive)

    Skips groups that net to near-zero (e.g. APPLY clearing pairs in rare cases).
    Strips vendor codes "(v/t NNNNNN)" from descriptions.
    Returns list of dicts: {date, desc, receipts, disb, sort_date}
    sorted chronologically.
    """
    from collections import defaultdict
    from datetime import date as _date

    net_by_key: Dict[tuple, float] = defaultdict(float)
    meta:       Dict[tuple, dict]  = {}

    for txn in (getattr(gl_acct, 'transactions', []) or []):
        d = getattr(txn, 'date', None)
        if isinstance(d, _date):
            date_str  = d.strftime('%m/%d/%Y')
            sort_date = d
        else:
            date_str  = str(d or '')
            sort_date = _date(1900, 1, 1)

        desc_raw   = str(getattr(txn, 'description', '') or '')
        desc_clean = re.sub(r'\s*\([tv]\d+\)\s*$', '', desc_raw).strip()

        debit  = float(getattr(txn, 'debit',  0) or 0)
        credit = float(getattr(txn, 'credit', 0) or 0)

        key = (sort_date, desc_clean)
        net_by_key[key] = round(net_by_key[key] + debit - credit, 2)
        if key not in meta:
            meta[key] = {'date_str': date_str, 'sort_date': sort_date,
                         'desc': desc_clean}

    rows = []
    for key in sorted(net_by_key.keys()):
        net  = net_by_key[key]
        if abs(net) < 0.01:
            continue
        info = meta[key]
        if net > 0:
            rows.append({'date': info['date_str'], 'desc': info['desc'],
                         'receipts': net, 'disb': 0.0,
                         'sort_date': info['sort_date']})
        else:
            rows.append({'date': info['date_str'], 'desc': info['desc'],
                         'receipts': 0.0, 'disb': abs(net),
                         'sort_date': info['sort_date']})

    return rows


def build_111100_tab(wb, period: str, property_name: str,
                     gl_acct=None, tb_entry=None,
                     bank_rec_data: Dict = None,
                     **_):
    """
    Operating Cash (PNC x3993) bank reconciliation workpaper.

    Layout:
      1. GL Activity — Date | Description | Receipts | Disbursements | Running Balance
         (transactions from main Yardi GL export, grouped by date × description)
      2. Bank Reconciliation block — from the Yardi Bank Rec PDF (first 3 pages):
            Balance Per Bank Statement as of {date}
            Less: Outstanding Checks
            Reconciled Bank Balance
            GL Balance (from TB)
            Difference (green = $0)
      3. Outstanding Checks list — date, check#, payee, amount
      4. GL / TB tie-out
    """
    gl_ending = float(getattr(gl_acct, 'ending_balance', 0) or 0)
    gl_beg    = float(getattr(gl_acct, 'beginning_balance', 0) or 0)
    tb_ending = float(getattr(tb_entry, 'ending_balance', 0) or 0) if tb_entry else gl_ending

    ws = wb.create_sheet('111100 PNC Cash'[:31])
    ws.sheet_properties.tabColor = '2D6F50'

    RCPT_COL  = 4   # D — Receipts
    DISB_COL  = 5   # E — Disbursements
    BAL_COL   = 6   # F — Running Balance
    NCOLS     = 5   # B-F

    next_row = _write_tab_header(ws, '111100', 'Cash - Operating Account (PNC x3993)',
                                 period, property_name, ncols=NCOLS)
    next_row += 1
    next_row = _write_col_headers(
        ws, next_row,
        ['Date', 'Description', 'Receipts', 'Disbursements', 'Running Balance'],
        [14, 52, 16, 18, 18],
    )

    # ── Balance Forward row ───────────────────────────────────────────────────
    c_beg = ws.cell(row=next_row, column=3, value='Balance Forward')
    _apply(c_beg, font=_font(italic=True, color='444444'), border=THIN,
           align=Alignment(horizontal='left'))
    ws.merge_cells(start_row=next_row, start_column=3,
                   end_row=next_row, end_column=BAL_COL - 1)
    c_beg_bal = ws.cell(row=next_row, column=BAL_COL, value=gl_beg)
    _apply(c_beg_bal, font=_font(bold=True), fill=_fill(LIGHT_BLUE),
           fmt='$#,##0.00', border=THIN, align=Alignment(horizontal='right'))
    beg_balance_row = next_row
    next_row += 1

    # ── GL transaction rows ───────────────────────────────────────────────────
    cash_rows = _group_cash_gl_transactions(gl_acct)

    for i, r in enumerate(cash_rows):
        alt = i % 2 == 1
        bg  = _fill(LIGHT_GRAY) if alt else None

        c1 = ws.cell(row=next_row, column=2, value=r['date'])
        _apply(c1, font=_font(), fill=bg, border=THIN)

        c2 = ws.cell(row=next_row, column=3, value=r['desc'])
        _apply(c2, font=_font(), fill=bg, border=THIN, align=Alignment(wrap_text=True))

        c3 = ws.cell(row=next_row, column=RCPT_COL, value=r['receipts'] or None)
        _apply(c3, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))

        c4 = ws.cell(row=next_row, column=DISB_COL, value=r['disb'] or None)
        _apply(c4, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))

        # Running balance: prior F + receipts(D) - disbursements(E)
        prev_ref = f'F{next_row - 1}'
        formula  = (f'={prev_ref}'
                    f'+IFERROR(D{next_row},0)'
                    f'-IFERROR(E{next_row},0)')
        c5 = ws.cell(row=next_row, column=BAL_COL, value=formula)
        _apply(c5, font=_font(bold=True), fill=_fill(LIGHT_BLUE),
               fmt='$#,##0.00', border=THIN, align=Alignment(horizontal='right'))

        next_row += 1

    if not cash_rows:
        c = ws.cell(row=next_row, column=2,
                    value='No operating cash GL activity for this period')
        _apply(c, font=_font(italic=True, color='666666'))
        next_row += 1

    # ── Period Totals row ─────────────────────────────────────────────────────
    next_row += 1
    c_tot = ws.cell(row=next_row, column=2, value='Period Totals')
    _apply(c_tot, font=_font(bold=True), border=THICK_BOTTOM)

    data_start = beg_balance_row + 1
    data_end   = next_row - 2
    for col_letter, col_idx in [('D', RCPT_COL), ('E', DISB_COL)]:
        if data_end >= data_start:
            tot_f = f'=SUM({col_letter}{data_start}:{col_letter}{data_end})'
        else:
            tot_f = 0
        ct = ws.cell(row=next_row, column=col_idx, value=tot_f)
        _apply(ct, font=_font(bold=True), fmt='$#,##0.00', border=THICK_BOTTOM,
               align=Alignment(horizontal='right'))

    ct_end = ws.cell(row=next_row, column=BAL_COL, value=gl_ending)
    _apply(ct_end, font=_font(bold=True), fill=_fill(LIGHT_BLUE),
           fmt='$#,##0.00', border=THICK_BOTTOM, align=Alignment(horizontal='right'))
    next_row += 2

    # ── Bank Reconciliation block ─────────────────────────────────────────────
    bank_stmt_bal  = None
    outstanding    = []
    total_oc       = 0.0
    reconciled_bal = None
    stmt_date      = ''
    rec_diff       = None

    if bank_rec_data and isinstance(bank_rec_data, dict):
        bank_stmt_bal  = bank_rec_data.get('bank_statement_balance')
        outstanding    = bank_rec_data.get('outstanding_checks') or []
        total_oc       = bank_rec_data.get('total_outstanding_checks') or 0.0
        reconciled_bal = bank_rec_data.get('reconciled_bank_balance')
        stmt_date      = bank_rec_data.get('statement_date') or ''
        rec_diff       = bank_rec_data.get('reconciling_difference')

    # Reconciliation rows
    date_label = f' as of {stmt_date}' if stmt_date else ''
    rec_blocks = [
        (f'Balance Per Bank Statement{date_label}',
         bank_stmt_bal if bank_stmt_bal is not None else '— upload bank rec to populate —',
         DARK_BLUE, 'FFFFFF'),
        ('Less: Outstanding Checks',
         -total_oc if total_oc else (None if bank_stmt_bal is None else 0.0),
         MED_BLUE, 'FFFFFF'),
        ('Reconciled Bank Balance',
         reconciled_bal if reconciled_bal is not None else (
             round((bank_stmt_bal or 0) - total_oc, 2) if bank_stmt_bal is not None else None),
         DARK_BLUE, 'FFFFFF'),
        ('GL Balance (per Trial Balance)',
         tb_ending,
         MED_BLUE, 'FFFFFF'),
        ('Difference',
         None,  # computed below
         None, None),
    ]

    # Compute difference: reconciled - GL
    _rec = rec_blocks[2][1]
    _rec_float = _rec if isinstance(_rec, (int, float)) else None
    diff_val = round(_rec_float - tb_ending, 2) if _rec_float is not None else None
    # Replace placeholder
    rec_blocks[4] = ('Difference', diff_val, None, None)

    c_rec_hdr = ws.cell(row=next_row, column=2, value='Bank Reconciliation')
    _apply(c_rec_hdr, font=_font(bold=True, size=10, color='FFFFFF'),
           fill=_fill(DARK_BLUE), border=THIN,
           align=Alignment(horizontal='left'))
    ws.merge_cells(start_row=next_row, start_column=2,
                   end_row=next_row, end_column=BAL_COL)
    next_row += 1

    for label, val, fill_hex, font_color in rec_blocks:
        c_lbl = ws.cell(row=next_row, column=2, value=label)
        _apply(c_lbl,
               font=_font(bold=True, color=font_color or '000000'),
               fill=_fill(fill_hex) if fill_hex else None,
               border=THIN,
               align=Alignment(horizontal='left'))
        ws.merge_cells(start_row=next_row, start_column=2,
                       end_row=next_row, end_column=BAL_COL - 1)
        c_val = ws.cell(row=next_row, column=BAL_COL, value=val)
        if label == 'Difference':
            ok = isinstance(val, (int, float)) and abs(val) < 0.02
            _apply(c_val,
                   font=_font(bold=True, color='006400' if ok else 'CC0000'),
                   fill=_fill(GREEN_FILL if ok else RED_FILL),
                   fmt='$#,##0.00' if isinstance(val, (int, float)) else None,
                   border=DOUBLE_BTM)
        elif val is not None and isinstance(val, (int, float)):
            _apply(c_val,
                   font=_font(bold=True, color=font_color or '000000'),
                   fill=_fill(fill_hex) if fill_hex else None,
                   fmt='$#,##0.00', border=THIN)
        else:
            _apply(c_val, font=_font(italic=True, color='888888'), border=THIN)
        next_row += 1

    # ── Outstanding Checks list ───────────────────────────────────────────────
    if outstanding:
        next_row += 1
        c_oc_hdr = ws.cell(row=next_row, column=2, value='Outstanding Checks')
        _apply(c_oc_hdr, font=_font(bold=True, size=10, color='FFFFFF'),
               fill=_fill(MED_BLUE), border=THIN,
               align=Alignment(horizontal='left'))
        ws.merge_cells(start_row=next_row, start_column=2,
                       end_row=next_row, end_column=BAL_COL)
        next_row += 1

        oc_hdrs = ['Date', 'Check #', 'Payee', '', '', 'Amount']
        oc_cols = [2, 3, 4, 5, 6, 7]
        for h, col in zip(oc_hdrs, oc_cols):
            if h:
                c = ws.cell(row=next_row, column=col, value=h)
                _apply(c, font=_font(bold=True, color='FFFFFF'),
                       fill=_fill(DARK_BLUE), border=THIN,
                       align=Alignment(horizontal='center'))
        # Merge payee across cols 4-6
        ws.merge_cells(start_row=next_row, start_column=4,
                       end_row=next_row, end_column=6)
        next_row += 1

        for i, chk in enumerate(outstanding):
            alt = i % 2 == 1
            bg  = _fill(LIGHT_GRAY) if alt else None

            d_val = chk.get('date', '')
            c1 = ws.cell(row=next_row, column=2, value=d_val)
            _apply(c1, font=_font(), fill=bg, border=THIN)

            c2 = ws.cell(row=next_row, column=3, value=chk.get('check_number', ''))
            _apply(c2, font=_font(), fill=bg, border=THIN)

            payee = chk.get('payee', '')
            c3 = ws.cell(row=next_row, column=4, value=payee)
            _apply(c3, font=_font(), fill=bg, border=THIN)
            ws.merge_cells(start_row=next_row, start_column=4,
                           end_row=next_row, end_column=6)

            amt = float(chk.get('amount', 0) or 0)
            c4 = ws.cell(row=next_row, column=7, value=amt)
            _apply(c4, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
                   align=Alignment(horizontal='right'))
            next_row += 1

        # Total Outstanding Checks row
        c_tot_lbl = ws.cell(row=next_row, column=2, value='Total Outstanding Checks')
        _apply(c_tot_lbl, font=_font(bold=True), fill=_fill(LIGHT_GRAY),
               border=THICK_BOTTOM)
        ws.merge_cells(start_row=next_row, start_column=2,
                       end_row=next_row, end_column=6)
        c_tot_val = ws.cell(row=next_row, column=7, value=total_oc)
        _apply(c_tot_val, font=_font(bold=True), fill=_fill(LIGHT_GRAY),
               fmt='$#,##0.00', border=THICK_BOTTOM,
               align=Alignment(horizontal='right'))
        next_row += 1

    # ── GL / TB tie-out ───────────────────────────────────────────────────────
    next_row += 1
    _write_tb_tieout(ws, next_row, gl_ending, tb_ending, amount_col=BAL_COL)
    return ws


# ── 131100 — AR Aging Detail ─────────────────────────────────────────────────

def build_131100_tab(wb, period, property_name, gl_acct=None, tb_entry=None,
                     ar_aging_data=None, **_):
    """AR Aging Detail — shows all charge rows EXCEPT Prepay (those go in 221100)."""
    gl_ending = float(getattr(gl_acct, 'ending_balance', 0) or 0)
    tb_ending = float(getattr(tb_entry, 'ending_balance', 0) or 0) if tb_entry else gl_ending

    ws = wb.create_sheet('131100 AR Aging'[:31])
    ws.sheet_properties.tabColor = 'BF8F00'

    headers   = ['Tenant', 'Charge', 'Tran #', 'Date', 'Current', '0-30', '31-60', '61-90', '90+', 'Total']
    col_widths = [28,       8,        12,       12,     14,        14,     12,      12,      12,    14]
    ncols = len(headers)

    next_row = _write_tab_header(ws, '131100', 'Accounts Receivable - AR Aging',
                                 period, property_name, ncols=ncols)
    next_row += 1
    next_row = _write_col_headers(ws, next_row, headers, col_widths)

    AMOUNT_COL = 11  # col K = column index 11 (B=2 + 9 offset)

    detail_rows = []
    if ar_aging_data is not None:
        try:
            detail_rows = [r for r in (ar_aging_data.detail_rows or [])
                           if not r.is_prepayment]
        except Exception:
            detail_rows = []

    if detail_rows:
        # Group by tenant
        from itertools import groupby
        # Sort by tenant_name first
        sorted_rows = sorted(detail_rows, key=lambda r: r.tenant_name)
        i_row = 0
        for tenant_name, tenant_group in groupby(sorted_rows, key=lambda r: r.tenant_name):
            tenant_rows = list(tenant_group)
            tenant_current_sum = 0.0
            tenant_total_sum   = 0.0
            for dr in tenant_rows:
                alt = i_row % 2 == 1
                bg  = _fill(LIGHT_GRAY) if alt else None

                # Tenant name
                c1 = ws.cell(row=next_row, column=2, value=dr.tenant_name)
                _apply(c1, font=_font(), fill=bg, border=THIN)
                # Charge code
                c2 = ws.cell(row=next_row, column=3, value=dr.charge_code)
                _apply(c2, font=_font(), fill=bg, border=THIN)
                # Tran #
                c3 = ws.cell(row=next_row, column=4, value=dr.tran_number)
                _apply(c3, font=_font(), fill=bg, border=THIN)
                # Date
                date_val = dr.date
                if date_val is not None:
                    try:
                        from datetime import datetime, date as _date
                        if isinstance(date_val, (_date, datetime)):
                            date_str = date_val.strftime('%m/%d/%Y')
                        else:
                            date_str = str(date_val)
                    except Exception:
                        date_str = str(date_val)
                else:
                    date_str = ''
                c4 = ws.cell(row=next_row, column=5, value=date_str)
                _apply(c4, font=_font(), fill=bg, border=THIN)
                # Money columns: Current, 0-30, 31-60, 61-90, 90+, Total
                for ci, val in enumerate([
                    dr.current_owed, dr.owed_0_30, dr.owed_31_60,
                    dr.owed_61_90, dr.owed_over_90, dr.total_owed
                ]):
                    c = ws.cell(row=next_row, column=6 + ci, value=val)
                    _apply(c, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
                           align=Alignment(horizontal='right'))

                tenant_current_sum += dr.current_owed
                tenant_total_sum   += dr.total_owed
                next_row += 1
                i_row    += 1

            # Tenant subtotal row
            c_lbl = ws.cell(row=next_row, column=2,
                            value=f'  Subtotal: {tenant_name}')
            _apply(c_lbl, font=_font(bold=True), fill=_fill(LIGHT_BLUE), border=THIN)
            ws.merge_cells(start_row=next_row, start_column=2,
                           end_row=next_row, end_column=5)
            c_sub_cur = ws.cell(row=next_row, column=6, value=tenant_current_sum)
            _apply(c_sub_cur, font=_font(bold=True), fill=_fill(LIGHT_BLUE),
                   fmt='$#,##0.00', border=THIN, align=Alignment(horizontal='right'))
            # Blank intermediate money cols
            for ci in range(1, 5):
                c = ws.cell(row=next_row, column=6 + ci, value=None)
                _apply(c, fill=_fill(LIGHT_BLUE), border=THIN)
            c_sub_tot = ws.cell(row=next_row, column=11, value=tenant_total_sum)
            _apply(c_sub_tot, font=_font(bold=True), fill=_fill(LIGHT_BLUE),
                   fmt='$#,##0.00', border=THIN, align=Alignment(horizontal='right'))
            next_row += 1

    else:
        # Fallback: GL transactions in standard 4-col format
        txns = list(getattr(gl_acct, 'transactions', []) or [])
        if not txns:
            c = ws.cell(row=next_row, column=2, value='No AR aging data or GL activity')
            _apply(c, font=_font(italic=True, color='666666'))
            next_row += 1
        else:
            for i, txn in enumerate(txns):
                alt = i % 2 == 1
                bg  = _fill(LIGHT_GRAY) if alt else None
                d = getattr(txn, 'date', None)
                from datetime import date as _date
                date_str = d.strftime('%m/%d/%Y') if isinstance(d, _date) else str(d or '')
                desc = str(getattr(txn, 'description', '') or '')
                amt  = float(getattr(txn, 'debit', 0) or 0) - float(getattr(txn, 'credit', 0) or 0)

                c1 = ws.cell(row=next_row, column=2, value=date_str)
                _apply(c1, font=_font(), fill=bg, border=THIN)
                c2 = ws.cell(row=next_row, column=3, value=desc)
                _apply(c2, font=_font(), fill=bg, border=THIN,
                       align=Alignment(wrap_text=True))
                ws.merge_cells(start_row=next_row, start_column=3,
                               end_row=next_row, end_column=10)
                c3 = ws.cell(row=next_row, column=11, value=amt)
                _apply(c3, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
                       align=Alignment(horizontal='right'))
                next_row += 1

    _write_tb_tieout(ws, next_row, gl_ending, tb_ending, amount_col=AMOUNT_COL)
    return ws


# ── 221100 — Prepaid Rent (AR Aging Prepay rows) ──────────────────────────────

def build_221100_tab(wb, period, property_name, gl_acct=None, tb_entry=None,
                     ar_aging_data=None, **_):
    """Prepaid Rent — shows ONLY Prepay rows from AR Aging."""
    gl_ending = float(getattr(gl_acct, 'ending_balance', 0) or 0)
    tb_ending = float(getattr(tb_entry, 'ending_balance', 0) or 0) if tb_entry else gl_ending

    ws = wb.create_sheet('221100 Prepaid Rent'[:31])
    ws.sheet_properties.tabColor = 'FF0000'

    headers    = ['Tenant', 'Tran #', 'Date', 'Pre-payment']
    col_widths = [32,        14,       14,     18]
    ncols = len(headers)

    next_row = _write_tab_header(ws, '221100', 'Prepaid Rent',
                                 period, property_name, ncols=ncols)
    next_row += 1
    next_row = _write_col_headers(ws, next_row, headers, col_widths)

    AMOUNT_COL = 5  # col E = column index 5

    prepay_rows = []
    if ar_aging_data is not None:
        try:
            prepay_rows = [r for r in (ar_aging_data.detail_rows or [])
                           if r.is_prepayment]
        except Exception:
            prepay_rows = []

    if prepay_rows:
        for i, dr in enumerate(prepay_rows):
            alt = i % 2 == 1
            bg  = _fill(LIGHT_GRAY) if alt else None

            c1 = ws.cell(row=next_row, column=2, value=dr.tenant_name)
            _apply(c1, font=_font(), fill=bg, border=THIN)

            c2 = ws.cell(row=next_row, column=3, value=dr.tran_number)
            _apply(c2, font=_font(), fill=bg, border=THIN)

            date_val = dr.date
            if date_val is not None:
                try:
                    from datetime import datetime, date as _date
                    if isinstance(date_val, (_date, datetime)):
                        date_str = date_val.strftime('%m/%d/%Y')
                    else:
                        date_str = str(date_val)
                except Exception:
                    date_str = str(date_val)
            else:
                date_str = ''
            c3 = ws.cell(row=next_row, column=4, value=date_str)
            _apply(c3, font=_font(), fill=bg, border=THIN)

            # total_owed is negative in Yardi for prepayments (liability)
            amt = dr.total_owed
            c4 = ws.cell(row=next_row, column=5, value=amt)
            _apply(c4, font=_font(color='CC0000'), fill=bg,
                   fmt='$#,##0.00', border=THIN,
                   align=Alignment(horizontal='right'))
            next_row += 1

        # Totals row
        total_prepay = sum(r.total_owed for r in prepay_rows)
        c_tot_lbl = ws.cell(row=next_row, column=2, value='Total Prepayments')
        _apply(c_tot_lbl, font=_font(bold=True), fill=_fill(DARK_BLUE),
               border=THIN, align=Alignment(horizontal='left'))
        ws.merge_cells(start_row=next_row, start_column=2,
                       end_row=next_row, end_column=4)
        c_tot_val = ws.cell(row=next_row, column=5, value=total_prepay)
        _apply(c_tot_val, font=_font(bold=True, color='CC0000'),
               fill=_fill(DARK_BLUE), fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))
        next_row += 1

    else:
        # Fallback: GL transactions
        txns = list(getattr(gl_acct, 'transactions', []) or [])
        if not txns:
            c = ws.cell(row=next_row, column=2, value='No prepayment data or GL activity')
            _apply(c, font=_font(italic=True, color='666666'))
            next_row += 1
        else:
            for i, txn in enumerate(txns):
                alt = i % 2 == 1
                bg  = _fill(LIGHT_GRAY) if alt else None
                d = getattr(txn, 'date', None)
                from datetime import date as _date
                date_str = d.strftime('%m/%d/%Y') if isinstance(d, _date) else str(d or '')
                desc = str(getattr(txn, 'description', '') or '')
                amt  = float(getattr(txn, 'debit', 0) or 0) - float(getattr(txn, 'credit', 0) or 0)

                c1 = ws.cell(row=next_row, column=2, value=date_str)
                _apply(c1, font=_font(), fill=bg, border=THIN)
                c2 = ws.cell(row=next_row, column=3, value=desc)
                _apply(c2, font=_font(), fill=bg, border=THIN,
                       align=Alignment(wrap_text=True))
                ws.merge_cells(start_row=next_row, start_column=3,
                               end_row=next_row, end_column=4)
                c3 = ws.cell(row=next_row, column=5, value=amt)
                _apply(c3, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
                       align=Alignment(horizontal='right'))
                next_row += 1

    _write_tb_tieout(ws, next_row, gl_ending, tb_ending, amount_col=AMOUNT_COL)
    return ws


# ── Capital accounts shared helper ────────────────────────────────────────────

def _build_capital_tab(wb, account_code, account_name, tab_color,
                       period, property_name, gl_acct, tb_entry,
                       capital_account,
                       has_entity=True, has_commencement=True,
                       prior_rows: list = None):
    """
    Shared builder for all 4 capital account tabs (154500, 181200, 181300, 181400).

    has_entity=False, has_commencement=False  → 154500 (Description | Date | Amount)
    has_entity=True,  has_commencement=True   → 181200/181300/181400

    Data source priority when capital_schedule_data xlsx is not uploaded:
      1. prior_rows — full detail read from prior workpaper tab (carry-forward)
      2. GL transactions — current period only (last-resort fallback)
    """
    gl_ending = float(getattr(gl_acct, 'ending_balance', 0) or 0)
    tb_ending = float(getattr(tb_entry, 'ending_balance', 0) or 0) if tb_entry else gl_ending

    tab_name = f'{account_code} {account_name}'[:31]
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = tab_color

    if has_entity and has_commencement:
        headers    = ['Description', 'Entity', 'Commencement Date', 'Amount']
        col_widths = [44,             12,       30,                  18]
    else:
        headers    = ['Description', 'Date', 'Amount']
        col_widths = [50,             12,     18]

    ncols = len(headers)
    AMOUNT_COL = 2 + ncols - 1  # last column index (B-based)

    next_row = _write_tab_header(ws, account_code, account_name,
                                 period, property_name, ncols=ncols)
    next_row += 1
    next_row = _write_col_headers(ws, next_row, headers, col_widths)

    has_data = (capital_account is not None and
                capital_account.rows is not None and
                len(capital_account.rows) > 0)

    if not has_data:
        if prior_rows:
            # ── Carry-forward from prior workpaper ────────────────────────────
            for i, row in enumerate(prior_rows):
                alt = i % 2 == 1
                bg  = _fill(LIGHT_GRAY) if alt else None

                if has_entity and has_commencement:
                    # 181xxx: Description | Entity | Commencement Date | Amount
                    c1 = ws.cell(row=next_row, column=2,
                                 value=row.get('description', ''))
                    _apply(c1, font=_font(), fill=bg, border=THIN,
                           align=Alignment(wrap_text=True))
                    c2 = ws.cell(row=next_row, column=3,
                                 value=row.get('entity', ''))
                    _apply(c2, font=_font(), fill=bg, border=THIN)
                    c3 = ws.cell(row=next_row, column=4,
                                 value=row.get('commencement_date', ''))
                    _apply(c3, font=_font(), fill=bg, border=THIN)
                    c4 = ws.cell(row=next_row, column=5,
                                 value=row.get('amount', 0))
                    _apply(c4, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
                           align=Alignment(horizontal='right'))
                else:
                    # 154500: Description | Date | Amount
                    c1 = ws.cell(row=next_row, column=2,
                                 value=row.get('description', ''))
                    _apply(c1, font=_font(), fill=bg, border=THIN,
                           align=Alignment(wrap_text=True))
                    c2 = ws.cell(row=next_row, column=3,
                                 value=row.get('date', ''))
                    _apply(c2, font=_font(), fill=bg, border=THIN)
                    c3 = ws.cell(row=next_row, column=4,
                                 value=row.get('amount', 0))
                    _apply(c3, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
                           align=Alignment(horizontal='right'))

                next_row += 1

            # Ending balance row (same style as capital_schedule_data path)
            c_end_lbl = ws.cell(row=next_row, column=2,
                                value='Ending Balance per GL')
            _apply(c_end_lbl, font=_font(bold=True, color='FFFFFF'),
                   fill=_fill(DARK_BLUE), border=THIN)
            ws.merge_cells(start_row=next_row, start_column=2,
                           end_row=next_row, end_column=AMOUNT_COL - 1)
            c_end_val = ws.cell(row=next_row, column=AMOUNT_COL, value=gl_ending)
            _apply(c_end_val, font=_font(bold=True, color='FFFFFF'),
                   fill=_fill(DARK_BLUE), fmt='$#,##0.00', border=THIN,
                   align=Alignment(horizontal='right'))
            next_row += 1

            _write_tb_tieout(ws, next_row, gl_ending, tb_ending, amount_col=AMOUNT_COL)
            return ws

        # ── No capital schedule and no prior workpaper — fall back to GL ──────
        c_msg = ws.cell(row=next_row, column=2,
                        value='Upload capital schedule for detailed view')
        _apply(c_msg, font=_font(italic=True, color='888888'), border=THIN)
        ws.merge_cells(start_row=next_row, start_column=2,
                       end_row=next_row, end_column=AMOUNT_COL)
        next_row += 2

        # GL transactions fallback
        txns = list(getattr(gl_acct, 'transactions', []) or [])
        for i, txn in enumerate(txns):
            alt = i % 2 == 1
            bg  = _fill(LIGHT_GRAY) if alt else None
            d = getattr(txn, 'date', None)
            from datetime import date as _date
            date_str = d.strftime('%m/%d/%Y') if isinstance(d, _date) else str(d or '')
            desc = str(getattr(txn, 'description', '') or '')
            amt  = float(getattr(txn, 'debit', 0) or 0) - float(getattr(txn, 'credit', 0) or 0)

            c1 = ws.cell(row=next_row, column=2, value=date_str if not has_entity else desc)
            _apply(c1, font=_font(), fill=bg, border=THIN)
            if has_entity:
                c2 = ws.cell(row=next_row, column=3, value='')
                _apply(c2, font=_font(), fill=bg, border=THIN)
                c3 = ws.cell(row=next_row, column=4, value='')
                _apply(c3, font=_font(), fill=bg, border=THIN)
            else:
                c2 = ws.cell(row=next_row, column=3, value=date_str)
                _apply(c2, font=_font(), fill=bg, border=THIN)
            c_amt = ws.cell(row=next_row, column=AMOUNT_COL, value=amt)
            _apply(c_amt, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
                   align=Alignment(horizontal='right'))
            next_row += 1

        _write_tb_tieout(ws, next_row, gl_ending, tb_ending, amount_col=AMOUNT_COL)
        return ws

    # ── Write capital schedule rows ───────────────────────────────────────────
    for i, row in enumerate(capital_account.rows):
        alt = i % 2 == 1
        bg  = _fill(LIGHT_GRAY) if alt else None

        if has_entity and has_commencement:
            c1 = ws.cell(row=next_row, column=2, value=row.description)
            _apply(c1, font=_font(), fill=bg, border=THIN,
                   align=Alignment(wrap_text=True))
            c2 = ws.cell(row=next_row, column=3, value=row.entity)
            _apply(c2, font=_font(), fill=bg, border=THIN)
            c3 = ws.cell(row=next_row, column=4, value=row.commencement_date)
            _apply(c3, font=_font(), fill=bg, border=THIN)
            c4 = ws.cell(row=next_row, column=5, value=row.amount)
            _apply(c4, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
                   align=Alignment(horizontal='right'))
        else:
            # 154500: Description | Date | Amount
            c1 = ws.cell(row=next_row, column=2, value=row.description)
            _apply(c1, font=_font(), fill=bg, border=THIN,
                   align=Alignment(wrap_text=True))
            c2 = ws.cell(row=next_row, column=3, value=row.commencement_date)
            _apply(c2, font=_font(), fill=bg, border=THIN)
            c3 = ws.cell(row=next_row, column=4, value=row.amount)
            _apply(c3, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
                   align=Alignment(horizontal='right'))

        next_row += 1

    # ── Ending balance row ────────────────────────────────────────────────────
    as_of_label = capital_account.as_of_date or ''
    end_label = (f'Ending Balance per GL as of {as_of_label}'
                 if as_of_label else 'Ending Balance per GL')
    c_end_lbl = ws.cell(row=next_row, column=2, value=end_label)
    _apply(c_end_lbl, font=_font(bold=True, color='FFFFFF'),
           fill=_fill(DARK_BLUE), border=THIN)
    ws.merge_cells(start_row=next_row, start_column=2,
                   end_row=next_row, end_column=AMOUNT_COL - 1)
    c_end_val = ws.cell(row=next_row, column=AMOUNT_COL,
                        value=capital_account.ending_balance)
    _apply(c_end_val, font=_font(bold=True, color='FFFFFF'),
           fill=_fill(DARK_BLUE), fmt='$#,##0.00', border=THIN,
           align=Alignment(horizontal='right'))
    next_row += 1

    _write_tb_tieout(ws, next_row, gl_ending, tb_ending, amount_col=AMOUNT_COL)
    return ws


# ── 152100 — Land ────────────────────────────────────────────────────────────

def build_152100_tab(wb, period, property_name, gl_acct=None, tb_entry=None,
                     capital_schedule_data=None, prior_tab_detail: dict = None, **_):
    acct = (capital_schedule_data or {}).get('152100')
    prior_rows = (prior_tab_detail or {}).get('152100') or _CAPITAL_152100_SEED
    return _build_capital_tab(wb, '152100', 'Land', '375623',
                              period, property_name, gl_acct, tb_entry,
                              acct, has_entity=False, has_commencement=False,
                              prior_rows=prior_rows)


# ── 154100 — Building ─────────────────────────────────────────────────────────

def build_154100_tab(wb, period, property_name, gl_acct=None, tb_entry=None,
                     capital_schedule_data=None, prior_tab_detail: dict = None, **_):
    acct = (capital_schedule_data or {}).get('154100')
    prior_rows = (prior_tab_detail or {}).get('154100') or _CAPITAL_154100_SEED
    return _build_capital_tab(wb, '154100', 'Building', '375623',
                              period, property_name, gl_acct, tb_entry,
                              acct, has_entity=False, has_commencement=False,
                              prior_rows=prior_rows)


# ── 154500 — Building Improvements ───────────────────────────────────────────

def build_154500_tab(wb, period, property_name, gl_acct=None, tb_entry=None,
                     capital_schedule_data=None, prior_tab_detail: dict = None, **_):
    acct = (capital_schedule_data or {}).get('154500')
    prior_rows = (prior_tab_detail or {}).get('154500') or _CAPITAL_154500_SEED
    return _build_capital_tab(wb, '154500', 'Building Improvements', '375623',
                              period, property_name, gl_acct, tb_entry,
                              acct, has_entity=False, has_commencement=False,
                              prior_rows=prior_rows)


# ── 171100 — CIP Development ─────────────────────────────────────────────────

def build_171100_tab(wb, period, property_name, gl_acct=None, tb_entry=None,
                     capital_schedule_data=None, prior_tab_detail: dict = None, **_):
    acct = (capital_schedule_data or {}).get('171100')
    # 171100 CIP has a zero balance — no seed rows; falls through to GL transactions
    prior_rows = (prior_tab_detail or {}).get('171100')
    return _build_capital_tab(wb, '171100', 'CIP Development', '375623',
                              period, property_name, gl_acct, tb_entry,
                              acct, has_entity=False, has_commencement=False,
                              prior_rows=prior_rows)


# ── 181200 — Leasing Commissions ─────────────────────────────────────────────

def build_181200_tab(wb, period, property_name, gl_acct=None, tb_entry=None,
                     capital_schedule_data=None, prior_tab_detail: dict = None, **_):
    acct = (capital_schedule_data or {}).get('181200')
    prior_rows = (prior_tab_detail or {}).get('181200') or _CAPITAL_181200_SEED
    return _build_capital_tab(wb, '181200', 'Leasing Commissions', '375623',
                              period, property_name, gl_acct, tb_entry,
                              acct, has_entity=True, has_commencement=True,
                              prior_rows=prior_rows)


# ── 181300 — Legal Leasing Costs ─────────────────────────────────────────────

def build_181300_tab(wb, period, property_name, gl_acct=None, tb_entry=None,
                     capital_schedule_data=None, prior_tab_detail: dict = None, **_):
    acct = (capital_schedule_data or {}).get('181300')
    prior_rows = (prior_tab_detail or {}).get('181300') or _CAPITAL_181300_SEED
    return _build_capital_tab(wb, '181300', 'Legal Leasing Costs', '375623',
                              period, property_name, gl_acct, tb_entry,
                              acct, has_entity=True, has_commencement=True,
                              prior_rows=prior_rows)


# ── 181400 — Tenant Improvement ───────────────────────────────────────────────

def build_181400_tab(wb, period, property_name, gl_acct=None, tb_entry=None,
                     capital_schedule_data=None, prior_tab_detail: dict = None, **_):
    acct = (capital_schedule_data or {}).get('181400')
    prior_rows = (prior_tab_detail or {}).get('181400') or _CAPITAL_181400_SEED
    return _build_capital_tab(wb, '181400', 'Tenant Improvement', '375623',
                              period, property_name, gl_acct, tb_entry,
                              acct, has_entity=True, has_commencement=True,
                              prior_rows=prior_rows)


# ── 187100 — Finance Costs ────────────────────────────────────────────────────

def build_187100_tab(wb, period, property_name, gl_acct=None, tb_entry=None,
                     capital_schedule_data=None, prior_tab_detail: dict = None, **_):
    """
    Finance Costs roll-forward.
    Layout: Date | Description | Revlabs | Revlabpm | Total  (cols B–F)
    Uses _CAPITAL_187100_SEED as bootstrap when no uploaded schedule or prior workpaper.
    """
    gl_ending = float(getattr(gl_acct, 'ending_balance', 0) or 0)
    tb_ending = float(getattr(tb_entry, 'ending_balance', 0) or 0) if tb_entry else gl_ending

    tab_name = '187100 Finance Costs'
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = '375623'

    headers    = ['Date', 'Description', 'Revlabs', 'Revlabpm', 'Total']
    col_widths = [14,      44,            16,         16,          16]
    ncols = len(headers)
    AMOUNT_COL = 6  # column F (Total)

    next_row = _write_tab_header(ws, '187100', 'Finance Costs',
                                 period, property_name, ncols=ncols)
    next_row += 1
    next_row = _write_col_headers(ws, next_row, headers, col_widths)

    # Data source: uploaded schedule → prior workpaper carry-forward → seed
    raw_rows = (prior_tab_detail or {}).get('187100') or _CAPITAL_187100_SEED

    rl_total  = 0.0
    rpm_total = 0.0

    for i, row in enumerate(raw_rows):
        alt = i % 2 == 1
        bg  = _fill(LIGHT_GRAY) if alt else None

        rl_amt  = float(row.get('revlabs',  0) or 0)
        rpm_amt = float(row.get('revlabpm', 0) or 0)
        tot_amt = round(rl_amt + rpm_amt, 2)
        rl_total  = round(rl_total  + rl_amt,  2)
        rpm_total = round(rpm_total + rpm_amt, 2)

        c_date = ws.cell(row=next_row, column=2, value=row.get('date', ''))
        _apply(c_date, font=_font(), fill=bg, border=THIN)
        c_desc = ws.cell(row=next_row, column=3, value=row.get('description', ''))
        _apply(c_desc, font=_font(), fill=bg, border=THIN,
               align=Alignment(wrap_text=True))
        c_rl = ws.cell(row=next_row, column=4, value=rl_amt)
        _apply(c_rl, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))
        c_rpm = ws.cell(row=next_row, column=5, value=rpm_amt)
        _apply(c_rpm, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))
        c_tot = ws.cell(row=next_row, column=6, value=tot_amt)
        _apply(c_tot, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))
        next_row += 1

    # Ending balance row
    grand_total = round(rl_total + rpm_total, 2)
    c_end_lbl = ws.cell(row=next_row, column=2, value='Ending Balance per GL')
    _apply(c_end_lbl, font=_font(bold=True, color='FFFFFF'),
           fill=_fill(DARK_BLUE), border=THIN)
    ws.merge_cells(start_row=next_row, start_column=2,
                   end_row=next_row, end_column=3)
    c_end_rl = ws.cell(row=next_row, column=4, value=rl_total)
    _apply(c_end_rl, font=_font(bold=True, color='FFFFFF'),
           fill=_fill(DARK_BLUE), fmt='$#,##0.00', border=THIN,
           align=Alignment(horizontal='right'))
    c_end_rpm = ws.cell(row=next_row, column=5, value=rpm_total)
    _apply(c_end_rpm, font=_font(bold=True, color='FFFFFF'),
           fill=_fill(DARK_BLUE), fmt='$#,##0.00', border=THIN,
           align=Alignment(horizontal='right'))
    c_end_tot = ws.cell(row=next_row, column=6, value=grand_total)
    _apply(c_end_tot, font=_font(bold=True, color='FFFFFF'),
           fill=_fill(DARK_BLUE), fmt='$#,##0.00', border=THIN,
           align=Alignment(horizontal='right'))
    next_row += 1

    _write_tb_tieout(ws, next_row, gl_ending, tb_ending, amount_col=AMOUNT_COL)
    return ws


# ── Capital account seed data ─────────────────────────────────────────────────
# Source: Capital tabs - Workpapers.xlsx  (manual workpaper as of Jan 2026)
# Used as bootstrap when neither capital_schedule_data nor prior workpaper is available.
# Format mirrors the prior_rows dicts consumed by _build_capital_tab().

# 152100 — Land  (Description | Date | Amount)
_CAPITAL_152100_SEED: List[dict] = [
    {'description': 'Move CIP Cost to FA - 12/14/2022', 'date': '12/31/2022', 'amount': 29000000.00},
]

# 154100 — Building  (single-amount layout; entity noted in description)
_CAPITAL_154100_SEED: List[dict] = [
    {'description': 'Beginning Balance (Revlabpm)', 'date': '01/01/2025', 'amount': 78897071.99},
    {'description': 'Building Allocation (Revlabpm)', 'date': '03/31/2025', 'amount': 520.93},
    {'description': 'Orum Q2 TI Capital (Revlabs)', 'date': '06/26/2025', 'amount': -10220.91},
]

# 154500 — Building Improvements  (Description | Date | Amount)
_CAPITAL_154500_SEED: List[dict] = [
    {'description': 'Penthouse Floor Epoxy-Black Bear Coatings & Concrete', 'date': '06/2024', 'amount': 49570.00},
    {'description': 'Electric Whip for Epoxy - S&W Electrical Contractors Inc.', 'date': '06/2024', 'amount': 705.00},
    {'description': 'Garage Sports Equipment',   'date': '02/2025', 'amount': 12828.10},
    {'description': '2nd Floor HVAC Corrections', 'date': '02/2025', 'amount': 33400.00},
]

# 181200 — Leasing Commissions  (Description | Entity | Commencement Date | Amount)
_CAPITAL_181200_SEED: List[dict] = [
    {'description': 'Keros',    'entity': 'Revlabs', 'commencement_date': '1/4/2023 - 11/30/2031', 'amount': 599121.60},
    {'description': 'Triana',   'entity': 'Revlabs', 'commencement_date': '4/13/2023 - 6/12/2033',  'amount': 638190.00},
    {'description': 'Accent',   'entity': 'Revlabs', 'commencement_date': '4/13/2023 - 6/12/2033',  'amount': 1365063.00},
    {'description': 'Orum',     'entity': 'Revlabs', 'commencement_date': '7/17/2023 - 8/31/2033',  'amount': 448665.00},
    {'description': 'Alchemab', 'entity': 'Revlabs', 'commencement_date': 'Moved-Out',              'amount': 568260.00},
    {'description': 'Rounding', 'entity': '',         'commencement_date': '',                       'amount': 0.40},
]

# 181300 — Legal Leasing Costs  (Description | Entity | Commencement Date | Amount)
_CAPITAL_181300_SEED: List[dict] = [
    {'description': 'Keros',    'entity': 'Revlabs', 'commencement_date': '1/4/2023 - 11/30/2031', 'amount': 48084.48},
    {'description': 'Triana',   'entity': 'Revlabs', 'commencement_date': '4/13/2023 - 6/12/2033',  'amount': 40976.04},
    {'description': 'Accent',   'entity': 'Revlabs', 'commencement_date': '4/13/2023 - 6/12/2033',  'amount': 92969.11},
    {'description': 'Orum',     'entity': 'Revlabs', 'commencement_date': '7/17/2023 - 8/31/2033',  'amount': 28807.27},
    {'description': 'Alchemab', 'entity': 'Revlabs', 'commencement_date': 'Moved-Out',              'amount': 36486.06},
]

# 181400 — Tenant Improvement  (Description | Entity | Commencement Date | Amount)
_CAPITAL_181400_SEED: List[dict] = [
    {'description': 'Keros',    'entity': 'Revlabs', 'commencement_date': '1/4/2023 - 11/30/2031', 'amount': 6775780.00},
    {'description': 'Triana',   'entity': 'Revlabs', 'commencement_date': '4/13/2023 - 6/12/2033',  'amount': 6078000.00},
    {'description': 'Accent',   'entity': 'Revlabs', 'commencement_date': '4/13/2023 - 6/12/2033',  'amount': 12710804.67},
    {'description': 'Orum',     'entity': 'Revlabs', 'commencement_date': '7/17/2023 - 8/31/2033',  'amount': 4273000.00},
    {'description': 'Alchemab', 'entity': 'Revlabs', 'commencement_date': 'Moved-Out',              'amount': 6375751.85},
]

# 187100 — Finance Costs  (Date | Description | Revlabs | Revlabpm | Total)
# Handled by its own builder (build_187100_tab) due to multi-entity column layout.
_CAPITAL_187100_SEED: List[dict] = [
    {'date': '01/01/2024', 'description': 'Beginning Balance',            'revlabs': 0.00,       'revlabpm': 0.00},
    {'date': '10/31/2024', 'description': 'Rev Labs Extension Fee',       'revlabs': 230260.49,  'revlabpm': 0.00},
    {'date': '06/25/2025', 'description': 'Frost Brown Todd LLP Invoice', 'revlabs': 3000.00,    'revlabpm': 0.00},
    {'date': '06/25/2025', 'description': 'Berkadia Commercial Mortgage', 'revlabs': 230260.49,  'revlabpm': 0.00},
    {'date': '06/25/2025', 'description': 'ACORE Capital, LP Invoice',    'revlabs': 1500.00,    'revlabpm': 0.00},
]


# ── Equity account seed data ──────────────────────────────────────────────────
# Source: Rev Labs Equity - Workpapers.xlsx  (manual workpaper as of Jan 2026)

# 311100 — Contributions - Partner A
# Each tuple: (date_str 'MM/DD/YYYY', description, amount)
# Amounts follow GL sign convention: credits to equity are negative
_EQUITY_311100_SEED: List[tuple] = [
    ('01/01/2025', 'Funding', -80942266.06),
]

# 331100 — Distributions - Partner A
# Each tuple: (date_str 'MM/DD/YYYY', description, revlabs_amt, revlabpm_amt)
# Distributions are debits to equity → positive amounts
_EQUITY_331100_SEED: List[tuple] = [
    ('05/01/2021', '331100-Partner Distributions',                              19711731.39, 0.0),
    ('07/01/2024', 'Transfer to Rev Lab Ventures',                              0.0,         1250000.0),
    ('09/01/2024', "Rcd: Distribution_09'24 - Transfer to Rev Lab Ventures",    0.0,         2000000.0),
    ('12/01/2024', "Rcd: Distribution_12'24",                                   0.0,          825000.0),
    ('03/01/2025', "Rcd: Distribution_03'25",                                   0.0,          540000.0),
    ('08/01/2025', "Rcd: Distribution_08'25",                                   0.0,          685000.0),
    ('09/01/2025', "Rcd: Distribution_09'25",                                   0.0,          700000.0),
    ('12/01/2025', "Rcd: Distribution_12'25",                                   0.0,         1025000.0),
]

# 381100 — Retained Earnings - Control
# Entity split for beginning balance (as of Jan 2026 workpaper)
_EQUITY_381100_SEED = {
    'revlabpm': -8184455.73,
    'revlabs':    251436.26,
}


# ── 311100 — Contributions - Partner A ───────────────────────────────────────

def build_311100_tab(wb, period, property_name, gl_acct=None, tb_entry=None,
                     prior_tab_detail: dict = None, **_):
    """
    Contributions - Partner A roll-forward.

    Layout: Date | Description | Amount  (cols B–D)
    Seed:   One-time funding row from Jan 2025; hardcoded first run.
    Carry:  prior_tab_detail['311100'] = list of {date_str, desc, amt}.
    GL:     Any credit activity in current period appended as new rows
            (credits to equity = contributions).
    """
    gl_ending = float(getattr(gl_acct, 'ending_balance', 0) or 0)
    tb_ending = float(getattr(tb_entry, 'ending_balance', 0) or 0) if tb_entry else gl_ending

    yr, mo = _parse_close_period(period)

    ws = wb.create_sheet('311100 Contributions-Partner A'[:31])
    ws.sheet_properties.tabColor = '002060'   # dark navy

    AMOUNT_COL = 4  # col D

    next_row = _write_tab_header(ws, '311100', 'Contributions - Partner A',
                                 period, property_name, ncols=3)
    next_row += 1
    next_row = _write_col_headers(ws, next_row,
                                  ['Date', 'Description', 'Amount'],
                                  [14, 54, 18])

    # ── Determine historical rows ─────────────────────────────────────────────
    prior_rows = (prior_tab_detail or {}).get('311100')
    if prior_rows is not None:
        hist_rows = prior_rows   # [{date_str, desc, amt}]
    else:
        # First run — bootstrap from seed (entries strictly before close period)
        hist_rows = [
            {'date_str': ds, 'desc': desc, 'amt': amt}
            for ds, desc, amt in _EQUITY_311100_SEED
            if _parse_date(ds) and (_parse_date(ds).year, _parse_date(ds).month) < (yr, mo)
        ]

    # ── Current-period GL activity (credits to equity = new contributions) ───
    current_rows = []
    for txn in (getattr(gl_acct, 'transactions', []) or []):
        d = getattr(txn, 'date', None)
        from datetime import date as _date_t
        if isinstance(d, _date_t) and (d.year, d.month) == (yr, mo):
            debit  = float(getattr(txn, 'debit',  0) or 0)
            credit = float(getattr(txn, 'credit', 0) or 0)
            net    = round(debit - credit, 2)   # negative = contribution
            if abs(net) > 0.01:
                desc_raw   = str(getattr(txn, 'description', '') or
                                 getattr(txn, 'remarks', '')     or '')
                desc_clean = re.sub(r'\s*\([tv]\d+\)\s*$', '', desc_raw).strip()
                current_rows.append({
                    'date_str': d.strftime('%m/%d/%Y'),
                    'desc': desc_clean,
                    'amt':  net,
                })

    all_rows = hist_rows + current_rows

    for i, r in enumerate(all_rows):
        alt = i % 2 == 1
        bg  = _fill(LIGHT_GRAY) if alt else None

        c1 = ws.cell(row=next_row, column=2, value=r.get('date_str', ''))
        _apply(c1, font=_font(), fill=bg, border=THIN)

        c2 = ws.cell(row=next_row, column=3, value=r.get('desc', ''))
        _apply(c2, font=_font(), fill=bg, border=THIN, align=Alignment(wrap_text=True))

        c3 = ws.cell(row=next_row, column=4, value=r.get('amt', 0))
        _apply(c3, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))
        next_row += 1

    _write_tb_tieout(ws, next_row, gl_ending, tb_ending, amount_col=AMOUNT_COL)
    return ws


# ── 331100 — Distributions - Partner A ───────────────────────────────────────

def build_331100_tab(wb, period, property_name, gl_acct=None, tb_entry=None,
                     prior_tab_detail: dict = None, **_):
    """
    Distributions - Partner A roll-forward with entity split.

    Layout: Date | Description | Revlabs | Revlabpm | Total  (cols B–F)
    Seed:   8 historical distribution rows; hardcoded first run.
    Carry:  prior_tab_detail['331100'] = list of
            {date_str, desc, revlabs, revlabpm, total}.
    GL:     New debit activity in current period → Revlabpm column.
            (Revlabs distributions are not in the revlabpm GL — seed only.)
    """
    gl_ending = float(getattr(gl_acct, 'ending_balance', 0) or 0)
    tb_ending = float(getattr(tb_entry, 'ending_balance', 0) or 0) if tb_entry else gl_ending

    yr, mo = _parse_close_period(period)

    ws = wb.create_sheet('331100 Distributions - A'[:31])
    ws.sheet_properties.tabColor = '2D6F50'   # Greatland green

    REVLABS_COL  = 4   # col D
    REVLABPM_COL = 5   # col E
    TOTAL_COL    = 6   # col F

    next_row = _write_tab_header(ws, '331100', 'Distributions - Partner A',
                                 period, property_name, ncols=5)
    next_row += 1
    next_row = _write_col_headers(ws, next_row,
                                  ['Date', 'Description', 'Revlabs', 'Revlabpm', 'Total'],
                                  [14, 46, 18, 18, 18])

    # ── Determine historical rows ─────────────────────────────────────────────
    prior_rows = (prior_tab_detail or {}).get('331100')
    if prior_rows is not None:
        hist_rows = prior_rows   # [{date_str, desc, revlabs, revlabpm, total}]
    else:
        # First run — bootstrap from seed (entries strictly before close period)
        hist_rows = []
        for ds, desc, revlabs, revlabpm in _EQUITY_331100_SEED:
            d = _parse_date(ds)
            if d and (d.year, d.month) < (yr, mo):
                hist_rows.append({
                    'date_str': ds, 'desc': desc,
                    'revlabs':  revlabs, 'revlabpm': revlabpm,
                    'total':    round(revlabs + revlabpm, 2),
                })

    # ── Current-period GL activity (debits = new distributions, Revlabpm) ────
    current_rows = []
    for txn in (getattr(gl_acct, 'transactions', []) or []):
        d = getattr(txn, 'date', None)
        from datetime import date as _date_t
        if isinstance(d, _date_t) and (d.year, d.month) == (yr, mo):
            debit  = float(getattr(txn, 'debit',  0) or 0)
            credit = float(getattr(txn, 'credit', 0) or 0)
            net    = round(debit - credit, 2)   # positive = distribution
            if abs(net) > 0.01:
                desc_raw   = str(getattr(txn, 'description', '') or
                                 getattr(txn, 'remarks', '')     or '')
                desc_clean = re.sub(r'\s*\([tv]\d+\)\s*$', '', desc_raw).strip()
                current_rows.append({
                    'date_str': d.strftime('%m/%d/%Y'),
                    'desc':     desc_clean,
                    'revlabs':  0.0,
                    'revlabpm': net,
                    'total':    net,
                })

    all_rows = hist_rows + current_rows

    for i, r in enumerate(all_rows):
        alt = i % 2 == 1
        bg  = _fill(LIGHT_GRAY) if alt else None

        c1 = ws.cell(row=next_row, column=2, value=r.get('date_str', ''))
        _apply(c1, font=_font(), fill=bg, border=THIN)

        c2 = ws.cell(row=next_row, column=3, value=r.get('desc', ''))
        _apply(c2, font=_font(), fill=bg, border=THIN, align=Alignment(wrap_text=True))

        c3 = ws.cell(row=next_row, column=4, value=r.get('revlabs', 0) or 0)
        _apply(c3, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))

        c4 = ws.cell(row=next_row, column=5, value=r.get('revlabpm', 0) or 0)
        _apply(c4, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))

        c5 = ws.cell(row=next_row, column=6, value=r.get('total', 0) or 0)
        _apply(c5, font=_font(), fill=bg, fmt='$#,##0.00', border=THIN,
               align=Alignment(horizontal='right'))
        next_row += 1

    # Totals row
    if all_rows:
        tot_rl  = round(sum(r.get('revlabs',  0) or 0 for r in all_rows), 2)
        tot_rpm = round(sum(r.get('revlabpm', 0) or 0 for r in all_rows), 2)
        tot_all = round(tot_rl + tot_rpm, 2)
        c_lbl = ws.cell(row=next_row, column=2, value='Total Distributions')
        _apply(c_lbl, font=_font(bold=True, color='FFFFFF'), fill=_fill(MED_BLUE),
               border=THIN)
        ws.merge_cells(start_row=next_row, start_column=2,
                       end_row=next_row, end_column=3)
        for col, val in [(4, tot_rl), (5, tot_rpm), (6, tot_all)]:
            c = ws.cell(row=next_row, column=col, value=val)
            _apply(c, font=_font(bold=True, color='FFFFFF'), fill=_fill(MED_BLUE),
                   fmt='$#,##0.00', border=THIN, align=Alignment(horizontal='right'))
        next_row += 1

    _write_tb_tieout(ws, next_row, gl_ending, tb_ending, amount_col=TOTAL_COL)
    return ws


# ── 381100 — Retained Earnings - Control ────────────────────────────────────

def build_381100_tab(wb, period, property_name, gl_acct=None, tb_entry=None,
                     prior_tab_detail: dict = None, **_):
    """
    Retained Earnings - Control snapshot.

    Shows the beginning balance with entity split (Revlabpm | Revlabs | Total).
    Retained earnings does not accumulate new entries intra-year — it represents
    prior-year accumulated net income and is static until year-end close.

    Layout: Description | Revlabpm | Revlabs | Total  (cols B–E)
    Seed:   Entity split from Jan 2026 workpaper; hardcoded first run.
    Carry:  prior_tab_detail['381100'] = {revlabpm, revlabs} for the split.
    GL:     Ending balance drives the GL/TB tie-out; entity split is from seed/prior.
    """
    gl_ending = float(getattr(gl_acct, 'ending_balance', 0) or 0)
    tb_ending = float(getattr(tb_entry, 'ending_balance', 0) or 0) if tb_entry else gl_ending

    ws = wb.create_sheet('381100 Retained Earnings - C'[:31])
    ws.sheet_properties.tabColor = 'ED7D31'   # amber/orange

    TOTAL_COL = 5   # col E

    next_row = _write_tab_header(ws, '381100', 'Retained Earnings - Control',
                                 period, property_name, ncols=4)
    next_row += 1
    next_row = _write_col_headers(ws, next_row,
                                  ['Description', 'Revlabpm', 'Revlabs', 'Total'],
                                  [40, 18, 18, 18])

    # ── Entity split: prior workpaper → seed fallback ────────────────────────
    prior_split = (prior_tab_detail or {}).get('381100')  # {revlabpm, revlabs}
    if prior_split:
        rpm_bal = prior_split.get('revlabpm', _EQUITY_381100_SEED['revlabpm'])
        rl_bal  = prior_split.get('revlabs',  _EQUITY_381100_SEED['revlabs'])
    else:
        rpm_bal = _EQUITY_381100_SEED['revlabpm']
        rl_bal  = _EQUITY_381100_SEED['revlabs']
    total_bal = round(rpm_bal + rl_bal, 2)

    # Beginning Balance row
    c1 = ws.cell(row=next_row, column=2, value='Beginning Balance')
    _apply(c1, font=_font(bold=True), border=THIN)

    c2 = ws.cell(row=next_row, column=3, value=rpm_bal)
    _apply(c2, font=_font(), fmt='$#,##0.00', border=THIN,
           align=Alignment(horizontal='right'))

    c3 = ws.cell(row=next_row, column=4, value=rl_bal)
    _apply(c3, font=_font(), fmt='$#,##0.00', border=THIN,
           align=Alignment(horizontal='right'))

    c4 = ws.cell(row=next_row, column=5, value=total_bal)
    _apply(c4, font=_font(bold=True), fmt='$#,##0.00', border=THIN,
           align=Alignment(horizontal='right'))
    next_row += 1

    # Note row
    c_note = ws.cell(row=next_row, column=2,
                     value='Note: Retained earnings updated at fiscal year-end close only.')
    _apply(c_note, font=_font(italic=True, color='666666', size=9))
    ws.merge_cells(start_row=next_row, start_column=2, end_row=next_row, end_column=5)
    next_row += 1

    _write_tb_tieout(ws, next_row, gl_ending, tb_ending, amount_col=TOTAL_COL)
    return ws


# ── Dispatch table ────────────────────────────────────────────────────────────

CUSTOM_BUILDERS: Dict[str, Any] = {
    '111100': build_111100_tab,
    '115100': build_115100_tab,
    '115200': build_115200_tab,
    '115300': build_115300_tab,
    '115600': build_115600_tab,
    '131100': build_131100_tab,
    '133100': build_133100_tab,
    '133110': build_133110_tab,
    '135150': build_135150_tab,
    '152100': build_152100_tab,
    '154100': build_154100_tab,
    '154500': build_154500_tab,
    '171100': build_171100_tab,
    '181200': build_181200_tab,
    '181300': build_181300_tab,
    '181400': build_181400_tab,
    '187100': build_187100_tab,
    '213100': build_213100_tab,
    '221100': build_221100_tab,
    '311100': build_311100_tab,
    '331100': build_331100_tab,
    '381100': build_381100_tab,
}

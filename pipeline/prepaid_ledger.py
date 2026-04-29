"""
Prepaid Expense Ledger
======================
Persistent tracker for multi-period invoices that need monthly amortization.

Workflow each close:
  1. load(path)          — read existing ledger (or start fresh)
  2. merge_nexus(records, period) — add new prepaids, skip already-tracked
  3. get_current_amortization(period) — JE lines for THIS month's expense
  4. advance_period(period) — increment months_amortized, mark completed items
  5. save(wb, path)      — write updated ledger to Excel for next month's upload

The ledger is a single Excel file with two sheets:
  'Active'    — items still being amortized (uploaded each month)
  'Completed' — items fully amortized (audit trail, never removed)
"""

import re
from calendar import monthrange
from datetime import date, datetime
from dateutil.relativedelta import relativedelta
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Schema ───────────────────────────────────────────────────

ACTIVE_COLS = [
    'vendor', 'invoice_number', 'invoice_date', 'description',
    'gl_account_number', 'gl_account', 'total_amount', 'monthly_amount',
    'service_start', 'service_end', 'total_months',
    'months_amortized', 'remaining_months', 'first_added_period',
    'daily_rate',
]

COMPLETED_COLS = ACTIVE_COLS + ['completed_period']

DARK_BLUE  = '1F4E78'
LIGHT_BLUE = 'D6E4F0'
AMBER      = 'FFF2CC'
GREEN      = 'E2EFDA'
LIGHT_GRAY = 'F2F2F2'
WHITE      = 'FFFFFF'

THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
DOUBLE_BOTTOM = Border(bottom=Side(style='double'))


def _hdr_font():
    return Font(name='Calibri', size=11, bold=True, color='FFFFFF')

def _hdr_fill():
    return PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')

def _apply(cell, font=None, fill=None, fmt=None, border=None, align=None):
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if fmt:    cell.number_format = fmt
    if border: cell.border = border
    if align:  cell.alignment = align


# ── Period helpers ───────────────────────────────────────────

_MONTH_MAP = dict(Jan=1, Feb=2, Mar=3, Apr=4, May=5, Jun=6,
                  Jul=7, Aug=8, Sep=9, Oct=10, Nov=11, Dec=12)


def _period_to_date(period_str: str) -> Optional[date]:
    """Convert 'Mar-2026' → date(2026, 3, 1). Returns None on failure."""
    if not period_str:
        return None
    m = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[- ]?(\d{4})',
                  period_str, re.IGNORECASE)
    if m:
        mon = _MONTH_MAP.get(m.group(1).capitalize(), 0)
        yr  = int(m.group(2))
        if mon:
            return date(yr, mon, 1)
    return None


def _date_to_period(d: date) -> str:
    """Convert date(2026, 3, 1) → 'Mar-2026'."""
    return d.strftime('%b-%Y')


def _ensure_date(val) -> Optional[date]:
    """Coerce Excel cell value to date."""
    if val is None:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, str):
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y'):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                pass
    return None


def _invoice_key(vendor: str, invoice_number: str) -> str:
    return f"{str(vendor).strip().lower()}||{str(invoice_number).strip().lower()}"


# ── Day-based proration helpers ──────────────────────────────

def _is_partial_period(service_start: date, service_end: date) -> bool:
    """True if either the start or end date is mid-month."""
    if service_start is None or service_end is None:
        return False
    _, last_day_start = monthrange(service_start.year, service_start.month)
    _, last_day_end   = monthrange(service_end.year,   service_end.month)
    return service_start.day > 1 or service_end.day != last_day_end


def _calc_daily_rate(total_amount: float, service_start: date, service_end: date) -> float:
    """
    Daily rate for a mid-month service period.

    Convention (matches standard CRE practice):
      First month: days = days_in_month − start_day
        e.g. Oct 20 → 31 − 20 = 11 days
      Last month:  days = end_day
        e.g. Oct 19 → 19 days
      Middle months: full calendar days in that month

    Total service days computed the same way:
      sum of (days in each calendar month spanned by the period)
      using the same partial-month convention above.
    """
    if service_start is None or service_end is None or total_amount == 0:
        return 0.0
    total_days = _count_service_days(service_start, service_end)
    return round(total_amount / total_days, 6) if total_days > 0 else 0.0


def _count_service_days(service_start: date, service_end: date) -> int:
    """
    Count total service days using partial-month convention:
      first month: days_in_month − start_day
      last month:  end_day
      middle months: full days_in_month
    """
    if service_start is None or service_end is None:
        return 0
    total = 0
    cur = date(service_start.year, service_start.month, 1)
    end_month = date(service_end.year, service_end.month, 1)
    while cur <= end_month:
        _, days_in_month = monthrange(cur.year, cur.month)
        if cur.year == service_start.year and cur.month == service_start.month:
            # First month — partial
            total += days_in_month - service_start.day
        elif cur.year == service_end.year and cur.month == service_end.month:
            # Last month — partial
            total += service_end.day
        else:
            total += days_in_month
        cur = (cur + relativedelta(months=1))
    return total


def _month_amount(item: dict, amort_date: date) -> float:
    """
    Return the correct amortization amount for a given calendar month.

    If daily_rate is set (mid-month invoice): compute from daily rate × days.
    Otherwise: return monthly_amount (standard equal-monthly).
    """
    daily_rate = float(item.get('daily_rate') or 0)
    if daily_rate <= 0:
        return float(item.get('monthly_amount', 0) or 0)

    service_start = _ensure_date(item.get('service_start'))
    service_end   = _ensure_date(item.get('service_end'))
    _, days_in_month = monthrange(amort_date.year, amort_date.month)

    # First month of service
    if service_start and amort_date.year == service_start.year \
            and amort_date.month == service_start.month:
        days = days_in_month - service_start.day

    # Last month of service
    elif service_end and amort_date.year == service_end.year \
            and amort_date.month == service_end.month:
        days = service_end.day

    else:
        days = days_in_month

    return round(daily_rate * days, 2)


# ── Load ─────────────────────────────────────────────────────

def load(path: Optional[str]) -> Tuple[List[Dict], List[Dict]]:
    """
    Load existing ledger from Excel file.

    Returns (active_items, completed_items).
    If path is None or file doesn't exist, returns ([], []).
    """
    if not path:
        return [], []
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return [], []

    active    = _read_sheet(wb, 'Active',    ACTIVE_COLS)
    completed = _read_sheet(wb, 'Completed', COMPLETED_COLS)
    return active, completed


_DISPLAY_TO_INTERNAL = {
    'vendor':               'vendor',
    'invoice #':            'invoice_number',
    'invoice date':         'invoice_date',
    'description':          'description',
    'gl account #':         'gl_account_number',
    'gl account name':      'gl_account',
    'total amount':         'total_amount',
    'monthly amt':          'monthly_amount',
    'service start':        'service_start',
    'service end':          'service_end',
    'total months':         'total_months',
    'months posted':        'months_amortized',
    'months left':          'remaining_months',
    'first added':          'first_added_period',
    'daily rate':           'daily_rate',
    'completed period':     'completed_period',
}


def _read_sheet(wb: Workbook, sheet_name: str, expected_cols: List[str]) -> List[Dict]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.values)
    if len(rows) < 2:
        return []

    # Find header row (first row containing 'Vendor' or 'vendor')
    header_row_idx = None
    raw_headers = []
    for i, row in enumerate(rows):
        row_strs = [str(c).strip().lower() if c else '' for c in row]
        if 'vendor' in row_strs:
            header_row_idx = i
            raw_headers = [str(c).strip().lower() if c else '' for c in row]
            break

    if header_row_idx is None:
        return []

    # Map display header names → internal field names
    mapped_headers = [_DISPLAY_TO_INTERNAL.get(h, h) for h in raw_headers]

    records = []
    for row in rows[header_row_idx + 1:]:
        if all(c is None or str(c).strip() == '' for c in row):
            continue
        # Skip totals / summary rows (no vendor value)
        vendor_idx = mapped_headers.index('vendor') if 'vendor' in mapped_headers else 0
        if vendor_idx < len(row) and (row[vendor_idx] is None or str(row[vendor_idx]).strip() == ''):
            continue
        # Skip rows where vendor looks like a label (e.g. 'TOTAL')
        vendor_val = str(row[vendor_idx]).strip() if vendor_idx < len(row) else ''
        if vendor_val.upper() in ('TOTAL', 'GRAND TOTAL', ''):
            continue

        rec = {}
        for ci, h in enumerate(mapped_headers):
            if ci < len(row):
                rec[h] = row[ci]

        # Coerce date fields
        for df in ('invoice_date', 'service_start', 'service_end'):
            rec[df] = _ensure_date(rec.get(df))

        # Coerce numeric fields
        for nf in ('total_amount', 'monthly_amount', 'total_months',
                   'months_amortized', 'remaining_months', 'daily_rate'):
            v = rec.get(nf)
            try:
                rec[nf] = float(v) if v is not None and str(v).strip() != '' else 0.0
            except (TypeError, ValueError):
                rec[nf] = 0.0

        records.append(rec)

    return records


# ── Merge new Nexus prepaids ─────────────────────────────────

def merge_nexus(active: List[Dict], nexus_records: List[Dict],
                close_period: str) -> Tuple[List[Dict], List[str]]:
    """
    Add new prepaid invoices from Nexus that aren't already in the ledger.

    Args:
        active:        Existing active ledger items
        nexus_records: Parsed Nexus records (from nexus_accrual.parse())
        close_period:  Current close period string e.g. 'Mar-2026'

    Returns:
        (updated_active, list_of_new_invoice_numbers_added)
    """
    existing_keys = {_invoice_key(r.get('vendor', ''), r.get('invoice_number', ''))
                     for r in active}
    added = []

    for inv in nexus_records:
        if not inv.get('is_prepaid'):
            continue
        key = _invoice_key(inv.get('vendor', ''), inv.get('invoice_number', ''))
        if key in existing_keys:
            continue

        total_months = inv.get('prepaid_months', 1)
        total_amount = inv.get('amount', 0)
        monthly_amount = round(total_amount / total_months, 2)

        svc_start = _ensure_date(inv.get('service_start'))
        svc_end   = _ensure_date(inv.get('service_end'))

        # Day-based proration: auto-detect mid-month start or end.
        # Sets daily_rate so get_current_amortization() can compute the exact
        # calendar-day amount for each month (partial first, full middle, partial last).
        if svc_start and svc_end and _is_partial_period(svc_start, svc_end):
            daily_rate = _calc_daily_rate(total_amount, svc_start, svc_end)
            # For the first (partial) month the Nexus accrual JE handles the expense;
            # the prepaid ledger tracks months 2+ from the FIRST FULL month onward.
            # Shift first_added_period to the first full month so the anchor is
            # a calendar month boundary (avoids fractional-month anchor arithmetic).
            first_full = date(svc_start.year, svc_start.month, 1) + relativedelta(months=1)
            first_added = _date_to_period(first_full)
        else:
            daily_rate   = 0.0
            first_added  = close_period

        active.append({
            'vendor':            inv.get('vendor', ''),
            'invoice_number':    inv.get('invoice_number', ''),
            'invoice_date':      inv.get('invoice_date'),
            'description':       inv.get('line_description', ''),
            'gl_account_number': inv.get('gl_account_number', ''),
            'gl_account':        inv.get('gl_account', ''),
            'total_amount':      total_amount,
            'monthly_amount':    monthly_amount,
            'service_start':     svc_start,
            'service_end':       svc_end,
            'total_months':      float(total_months),
            'months_amortized':  0.0,
            'remaining_months':  float(total_months),
            'first_added_period': first_added,
            'daily_rate':        daily_rate,
        })
        existing_keys.add(key)
        added.append(inv.get('invoice_number', ''))

    return active, added


# ── Generate current period amortization JE lines ───────────

def get_current_amortization(active: List[Dict], close_period: str) -> List[Dict]:
    """
    Return one amortization record per active ledger item for the current period.

    Each record has enough info to build a JE:
      DR  [gl_account_number]  amount_for_this_month
      CR  135150 Prepaid Other  amount_for_this_month

    Items with months_amortized == 0 are the FIRST month:
      those are expensed via the normal Nexus accrual JE (DR expense / CR 211200)
      and should NOT generate a duplicate here.
    We only generate prepaid-release JEs for months 2+ (months_amortized >= 1).

    Day-based proration: if the item has daily_rate > 0 (set automatically by
    merge_nexus() when service_start is mid-month or service_end is mid-month),
    the amount for each calendar month is computed as:
      first month:  daily_rate × (days_in_month − service_start.day)
      last month:   daily_rate × service_end.day
      middle months: daily_rate × days_in_month
    This matches the convention: 11 days for Oct 20 start (31 − 20 = 11).
    """
    close_date = _period_to_date(close_period)
    if close_date is None:
        # Cannot determine which period to amortize — return nothing rather
        # than silently releasing every active item at once.
        return []

    results = []

    for item in active:
        svc_start = _ensure_date(item.get('service_start'))
        if not svc_start:
            continue

        months_done = int(item.get('months_amortized', 0) or 0)
        remaining   = int(item.get('remaining_months', 0) or 0)

        if remaining <= 0:
            continue

        # Month 0 (first month) is covered by the Nexus accrual JE.
        # Months 1+ are prepaid asset releases.
        if months_done == 0:
            continue

        # Verify this item is due this period.
        # Anchor from first_added_period (not service_start) because invoices
        # are often received after service has already started.
        first_added = _period_to_date(item.get('first_added_period', ''))
        anchor = first_added or date(svc_start.year, svc_start.month, 1)
        amort_month = anchor + relativedelta(months=months_done)
        if close_date and (amort_month.year != close_date.year or
                           amort_month.month != close_date.month):
            continue

        # Amount: day-based if daily_rate set, otherwise fixed monthly_amount
        amount = _month_amount(item, amort_month)

        results.append({
            'vendor':            item.get('vendor', ''),
            'invoice_number':    item.get('invoice_number', ''),
            'description':       item.get('description', ''),
            'gl_account_number': item.get('gl_account_number', ''),
            'gl_account':        item.get('gl_account', ''),
            'monthly_amount':    amount,
            'period_label':      _date_to_period(amort_month),
            'month_index':       months_done + 1,
            'total_months':      int(item.get('total_months', 1) or 1),
            'source':            'prepaid_ledger',
            'is_day_based':      float(item.get('daily_rate') or 0) > 0,
        })

    return results


# ── Advance period ───────────────────────────────────────────

def advance_period(active: List[Dict], completed: List[Dict],
                   close_period: str) -> Tuple[List[Dict], List[Dict]]:
    """
    After the close period JEs are posted:
      - Increment months_amortized by 1 for each active item
      - Move items with remaining_months == 0 to completed list
      - Return (new_active, new_completed)
    """
    new_active = []
    for item in active:
        item = dict(item)  # copy
        months_done = int(item.get('months_amortized', 0) or 0)
        total       = int(item.get('total_months', 1) or 1)

        item['months_amortized'] = months_done + 1
        item['remaining_months'] = max(0, total - item['months_amortized'])

        if item['remaining_months'] <= 0:
            item['completed_period'] = close_period
            completed.append(item)
        else:
            new_active.append(item)

    return new_active, completed


# ── Save to Excel ─────────────────────────────────────────────

def save(active: List[Dict], completed: List[Dict], path: str) -> str:
    """Write the ledger to Excel with 'Active' and 'Completed' tabs."""
    wb = Workbook()

    _write_sheet(wb, 'Active', active, ACTIVE_COLS,
                 title='Prepaid Expense Ledger — Active Items',
                 tab_color='2E75B6')

    _write_sheet(wb, 'Completed', completed, COMPLETED_COLS,
                 title='Prepaid Expense Ledger — Completed Items',
                 tab_color='70AD47')

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    wb.save(path)
    return path


def _write_sheet(wb: Workbook, sheet_name: str, records: List[Dict],
                 cols: List[str], title: str, tab_color: str):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color

    # Title
    ws.cell(row=1, column=1, value=title).font = Font(
        name='Calibri', size=13, bold=True, color=DARK_BLUE)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))

    ws.cell(row=2, column=1,
            value=f'Updated: {datetime.now().strftime("%m/%d/%Y %I:%M %p")}  |  {len(records)} item(s)').font = Font(
        name='Calibri', size=10, italic=True, color='666666')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))

    # Headers
    DISPLAY = {
        'vendor': 'Vendor', 'invoice_number': 'Invoice #',
        'invoice_date': 'Invoice Date', 'description': 'Description',
        'gl_account_number': 'GL Account #', 'gl_account': 'GL Account Name',
        'total_amount': 'Total Amount', 'monthly_amount': 'Monthly Amt',
        'service_start': 'Service Start', 'service_end': 'Service End',
        'total_months': 'Total Months', 'months_amortized': 'Months Posted',
        'remaining_months': 'Months Left', 'first_added_period': 'First Added',
        'daily_rate': 'Daily Rate',
        'completed_period': 'Completed Period',
    }
    hdr_row = 4
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=hdr_row, column=ci, value=DISPLAY.get(col, col))
        _apply(c, font=_hdr_font(), fill=_hdr_fill(), border=THIN,
               align=Alignment(horizontal='center', vertical='center', wrap_text=True))
    ws.row_dimensions[hdr_row].height = 30

    # Data
    for ri, rec in enumerate(records, hdr_row + 1):
        # Highlight active items with remaining months <= 1 (almost done)
        remaining = int(rec.get('remaining_months', 99) or 99)
        row_fill = None
        if sheet_name == 'Active':
            if remaining == 1:
                row_fill = PatternFill(start_color=AMBER, end_color=AMBER, fill_type='solid')
            elif ri % 2 == 0:
                row_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')

        for ci, col in enumerate(cols, 1):
            val = rec.get(col, '')
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = THIN
            if row_fill:
                c.fill = row_fill

            # Formatting by column type
            if col in ('total_amount', 'monthly_amount'):
                c.number_format = '$#,##0.00'
            elif col == 'daily_rate':
                c.number_format = '$#,##0.000000'
            elif col in ('invoice_date', 'service_start', 'service_end'):
                if isinstance(val, date):
                    c.number_format = 'MM/DD/YYYY'
            elif col in ('total_months', 'months_amortized', 'remaining_months'):
                c.number_format = '0'
                c.alignment = Alignment(horizontal='center')
                # Highlight items nearing completion in the Active sheet
                if col == 'remaining_months' and sheet_name == 'Active':
                    if remaining == 0:
                        c.font = Font(name='Calibri', color='FF0000', bold=True)
                    elif remaining == 1:
                        c.font = Font(name='Calibri', color='C55A11', bold=True)

    # Totals row (active sheet only)
    if sheet_name == 'Active' and records:
        total_row = hdr_row + len(records) + 1
        total_col_map = {col: i + 1 for i, col in enumerate(cols)}

        ws.cell(row=total_row,
                column=total_col_map.get('description', 1),
                value='TOTAL').font = Font(name='Calibri', size=11, bold=True)

        for fcol in ('total_amount', 'monthly_amount'):
            ci = total_col_map.get(fcol)
            if ci:
                total = sum(float(r.get(fcol, 0) or 0) for r in records)
                c = ws.cell(row=total_row, column=ci, value=total)
                c.number_format = '$#,##0.00'
                c.font = Font(name='Calibri', size=11, bold=True)
                c.border = DOUBLE_BOTTOM

    # Column widths
    col_widths = {
        'vendor': 28, 'invoice_number': 16, 'invoice_date': 13,
        'description': 38, 'gl_account_number': 14, 'gl_account': 32,
        'total_amount': 14, 'monthly_amount': 13, 'service_start': 13,
        'service_end': 13, 'total_months': 10, 'months_amortized': 12,
        'remaining_months': 12, 'first_added_period': 13, 'daily_rate': 12,
        'completed_period': 14,
    }
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 14)

    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

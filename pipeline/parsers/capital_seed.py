"""
Capital Seed Schedule Parser
============================
Parses the January seed capital schedule (Book3.xlsx) — 7 sheets for all
capital accounts.  This file is uploaded once to bootstrap January 2026 when
no prior workpaper exists.  From February onward the prior workpaper
carry-forward supersedes it.

JLL sheet format
----------------
  Row 1:   Property name
  Row 2:   Account name + code
  Row 3:   Date
  Row 4:   Column headers
  Row 5+:  Data rows (includes "Beginning Balance" carrying balance)
  ...
  "Ending Balance per GL as of [date]" row — ends data section
  Tie-out block (below ending balance, not parsed)

Amount column (0-indexed, col A = 0) per account
-------------------------------------------------
  152100  Land                   col H  index 7
  154100  Building               col I  index 8
  154500  Building Improvements  col F  index 5
  171100  CIP Development        col F  index 5
  181200  Leasing Commissions    col F  index 5
  181300  Legal Leasing Costs    col F  index 5
  181400  Tenant Improvement     col G  index 6

Returns a dict keyed by account_code with CapitalAccount dataclasses that are
type-compatible with parsers/capital_schedule.py's output.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Any


# ── Shared dataclasses (mirror capital_schedule.py for compatibility) ──────────

@dataclass
class CapitalRow:
    description: str
    entity: str = ''
    commencement_date: str = ''    # also used as "Date" for simple accounts
    amount: float = 0.0
    notes: str = ''


@dataclass
class CapitalAccount:
    account_code: str
    account_name: str
    rows: List[CapitalRow]
    ending_balance: float
    as_of_date: str = ''


# ── Account catalogue ──────────────────────────────────────────────────────────

_ACCT_NAMES: Dict[str, str] = {
    '152100': 'Land',
    '154100': 'Building',
    '154500': 'Building Improvements',
    '171100': 'CIP Development',
    '181200': 'Leasing Commissions',
    '181300': 'Legal Leasing Costs',
    '181400': 'Tenant Improvement',
}

# Layout per account:
#   (desc_col, entity_col, comm_col, amount_col, has_entity, has_comm)
# All column indices are 0-based (col A = 0).
_SEED_LAYOUT: Dict[str, tuple] = {
    '152100': (1, None, None, 7, False, False),  # B=desc, H=amount
    '154100': (1, None, None, 8, False, False),  # B=desc, I=amount
    '154500': (1, None, 3,    5, False, False),  # B=desc, D=date, F=amount
    '171100': (1, None, 3,    5, False, False),  # B=desc, D=date, F=amount
    '181200': (1, 2,    3,    5, True,  True),   # B=desc, C=entity, D=comm, F=amount
    '181300': (1, 2,    3,    5, True,  True),
    '181400': (1, 2,    3,    6, True,  True),   # B=desc, C=entity, D=comm, G=amount
}

# Sheet detection keywords (fallback when account code not in sheet name)
_KEYWORDS: Dict[str, str] = {
    '152100': 'land',
    '154100': 'building',
    '154500': 'building improv',
    '171100': 'cip',
    '181200': 'leasing comm',
    '181300': 'legal',
    '181400': 'tenant improv',
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _safe_float(v) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def _find_tab(wb, account_code: str):
    """Return the worksheet whose name contains the account code (exact), or
    fall back to a keyword match on the account name."""
    for name in wb.sheetnames:
        if account_code in name:
            return wb[name]
    kw = _KEYWORDS.get(account_code, '')
    if kw:
        for name in wb.sheetnames:
            if kw in name.lower():
                return wb[name]
    return None


def _fmt_comm(raw) -> str:
    """Format a commencement / date cell value as a string."""
    if raw is None:
        return ''
    try:
        from datetime import datetime, date as _date
        if isinstance(raw, (_date, datetime)):
            try:
                return raw.strftime('%-m/%-d/%Y')   # Linux/macOS
            except ValueError:
                return raw.strftime('%m/%d/%Y')     # Windows
    except Exception:
        pass
    return str(raw).strip()


# ── Sheet parser ───────────────────────────────────────────────────────────────

def _parse_sheet(ws, account_code: str) -> Optional[CapitalAccount]:
    """Parse a single worksheet into a CapitalAccount."""
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return None

    layout = _SEED_LAYOUT.get(account_code)
    if not layout:
        return None

    desc_col, entity_col, comm_col, amount_col, has_entity, has_comm = layout

    # ── Locate the data start row ──────────────────────────────────────────────
    # Scan the first 10 rows for a header row (contains "description" or "amount").
    # Data begins on the row immediately after the header.
    data_start = 4  # default: row 5 (0-indexed 4)
    for scan_idx in range(min(10, len(all_rows))):
        row_text = ' '.join(str(v or '').lower() for v in all_rows[scan_idx])
        if 'description' in row_text or 'amount' in row_text:
            data_start = scan_idx + 1
            break

    # ── Parse data rows ────────────────────────────────────────────────────────
    rows: List[CapitalRow] = []
    ending_balance = 0.0
    as_of_date = ''

    for raw in all_rows[data_start:]:
        row = list(raw)
        # Pad row so column access never raises IndexError
        max_needed = max(desc_col, amount_col,
                         entity_col if entity_col is not None else 0,
                         comm_col   if comm_col   is not None else 0)
        while len(row) <= max_needed:
            row.append(None)

        desc_val   = row[desc_col]
        amount_val = row[amount_col]

        desc_str = str(desc_val or '').strip()
        if not desc_str:
            continue  # blank row — skip

        desc_lower = desc_str.lower()

        # ── Ending balance row: extract value then stop ────────────────────────
        if 'ending balance' in desc_lower:
            ending_balance = _safe_float(amount_val)
            m = re.search(
                r'as\s+of\s+(\d{1,2}[/\-]\d{1,2}(?:[/\-]\d{2,4})?)',
                desc_str, re.IGNORECASE,
            )
            if m:
                as_of_date = m.group(1)
            break   # tie-out block follows — don't read further

        # Skip rows with no numeric amount
        if amount_val is None:
            continue
        try:
            amt = float(amount_val)
        except (ValueError, TypeError):
            continue

        entity_str = ''
        if entity_col is not None:
            entity_str = str(row[entity_col] or '').strip()

        comm_str = ''
        if comm_col is not None:
            comm_str = _fmt_comm(row[comm_col])

        rows.append(CapitalRow(
            description=desc_str,
            entity=entity_str,
            commencement_date=comm_str,
            amount=amt,
        ))

    return CapitalAccount(
        account_code=account_code,
        account_name=_ACCT_NAMES.get(account_code, ''),
        rows=rows,
        ending_balance=ending_balance,
        as_of_date=as_of_date,
    )


# ── Public API ─────────────────────────────────────────────────────────────────

def parse(filepath: str) -> Dict[str, Any]:
    """
    Parse all 7 capital accounts from the seed schedule.

    Returns a dict keyed by account_code (same shape as capital_schedule.parse())::

        {
            '152100': CapitalAccount | None,
            '154100': CapitalAccount | None,
            '154500': CapitalAccount | None,
            '171100': CapitalAccount | None,
            '181200': CapitalAccount | None,
            '181300': CapitalAccount | None,
            '181400': CapitalAccount | None,
            '_parse_error': None | str,
        }
    """
    result: Dict[str, Any] = {code: None for code in _ACCT_NAMES}
    result['_parse_error'] = None

    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        for account_code in _ACCT_NAMES:
            ws = _find_tab(wb, account_code)
            if ws is None:
                continue
            try:
                ca = _parse_sheet(ws, account_code)
                if ca is not None:
                    result[account_code] = ca
            except Exception:
                pass  # non-fatal — individual sheet parse failure
    except Exception as exc:
        result['_parse_error'] = str(exc)

    return result

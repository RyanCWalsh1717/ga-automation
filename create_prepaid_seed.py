"""
create_prepaid_seed.py
======================
One-time script to build the starting GA_Prepaid_Ledger_Updated.xlsx from
the March 2026 RCW workpaper.

Run once:
    python create_prepaid_seed.py

Output:
    GA_Prepaid_Ledger_Seed_Mar2026.xlsx

Upload this file as the "Prior Prepaid Ledger" in Pass 1 for the April close.
After April runs it will be replaced by the pipeline's own updated ledger.

Data sources in the RCW workpaper
----------------------------------
  '135150 PPD Other'   — active / completed non-insurance prepaids
  'Insurance Analysis' — property, GL, umbrella policies (GL 639110/639120/135110)
"""

import sys
import os
import re
import shutil
import tempfile
from datetime import date, datetime
from dateutil.relativedelta import relativedelta

# ── Make sure we can import the pipeline's own save() ──────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_HERE, 'pipeline'))

from prepaid_ledger import save as _ledger_save   # noqa: E402

# ── RCW file path ───────────────────────────────────────────────────────────
RCW_PATH = (
    r'C:\Users\RyanCWalsh\Greatland Realty Partners\Greatland Partners - Documents'
    r'\Portfolio\Revolution Labs\10 - Finance\Claude\March Data\JLL Final Data'
    r'\03.26 Workpapers REVLABS RCW.xlsx'
)

OUT_PATH_MAR = os.path.join(_HERE, 'GA_Prepaid_Ledger_Seed_Mar2026.xlsx')
OUT_PATH_DEC = os.path.join(_HERE, 'GA_Prepaid_Ledger_Seed_Dec2025.xlsx')


# ── Helpers ─────────────────────────────────────────────────────────────────

def _safe_float(v) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _ensure_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%m/%d/%y', '%m-%d-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _parse_date_part(s: str):
    """
    Try to parse a single date token. Handles:
      MM/DD/YYYY  MM/DD/YY  MM/YYYY  MM/YY  MM.DD.YY  MM.DD.YYYY
      Special: 'MM/YY' and 'MM.YY' where YY is a 2-digit year
    Returns date or None.
    """
    s = s.strip().replace('.', '/')   # normalise dots to slashes
    import calendar as _cal
    # Full date formats
    for fmt in ('%m/%d/%Y', '%m/%d/%y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    # Month/Year only (return 1st of that month)
    for fmt in ('%m/%Y', '%m/%y'):
        try:
            d = datetime.strptime(s, fmt).date()
            return d.replace(day=1)
        except ValueError:
            pass
    return None


def _parse_term(term_str: str):
    """
    Parse a service-period string into (start_date, end_date).

    Handles many JLL formats:
      '6/12/2025- 6/12/2026'         slash full dates
      '03/01/25-02/28/26'            slash 2-digit year
      '12.19.25-12.18.26'            dot notation
      '02.10.26 - 02.09.27'          dot with spaces
      '01/26-06/26'                  month/year only
      '12/25/11/26'                  JLL shorthand: MM/YY-MM/YY with no separator char
      '9/2025-9/2026'                month/4-digit year
    Returns (None, None) if unparseable.
    """
    if not term_str:
        return None, None

    raw = str(term_str).strip()

    # ── Normalise spaces around separator ─────────────────────────────────
    # Replace en-dash with regular dash; collapse spaces around dashes
    raw = raw.replace('–', '-').replace('–', '-')

    # ── Special case: JLL 'MM/YY/MM/YY' with no separator (4 slash groups) ──
    # e.g. '12/25/11/26' → start=12/25, end=11/26  (month/YY-month/YY)
    # Detect: exactly 3 slashes, no dash
    if raw.count('/') == 3 and '-' not in raw and ' ' not in raw:
        parts = raw.split('/')
        if all(len(p) <= 2 for p in parts) and len(parts) == 4:
            # Treat as MM/YY - MM/YY
            s_str = f'{parts[0]}/{parts[1]}'
            e_str = f'{parts[2]}/{parts[3]}'
            sd = _parse_date_part(s_str)
            ed = _parse_date_part(e_str)
            if sd and ed:
                return sd, ed

    # ── General: split on the dash that separates start from end ─────────
    # Tricky because dates themselves can contain dashes (they don't here,
    # but dots and slashes can).  Strategy: find all dash positions and
    # try every split point.
    candidates = []
    for i, ch in enumerate(raw):
        if ch == '-':
            s_part = raw[:i].strip()
            e_part = raw[i+1:].strip()
            if s_part and e_part:
                candidates.append((s_part, e_part))

    for s_str, e_str in candidates:
        sd = _parse_date_part(s_str)
        ed = _parse_date_part(e_str)
        if sd and ed and sd <= ed:
            return sd, ed

    return None, None


# ── GL account name lookup ───────────────────────────────────────────────────
_GL_NAMES = {
    '637150': 'Admin-Tenant Relations',
    '637290': 'Admin-Telephone',
    '637330': 'Admin-Printing/Reproduction',
    '637370': 'Admin-Other Admin Expense',
    '680110': 'Management-Professional Fees',
    '680510': 'Management-Software',
    '615320': 'Maintenance-Security',
    '712210': 'Depreciation',
    '639110': 'Insurance-Property',
    '639120': 'Insurance-Liability',
    '135110': 'Prepaid Insurance',
    '213300': 'Accrued Insurance',
    '135120': 'Prepaid RE Tax',
    '641110': 'RE Tax Expense',
    '115200': 'Escrow RE Tax',
}


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════

def _build_for_seed_date(wb, seed_date: date, seed_period: str):
    """
    Build active / completed lists as of a given seed date.
    Items whose service hasn't started yet are excluded from active.
    Items fully amortized by seed_date go to completed.
    """
    from openpyxl import load_workbook as _lw  # already imported above

    all_items = []
    completed = []

    if '135150 PPD Other' in wb.sheetnames:
        raw = _read_135150_raw(wb['135150 PPD Other'], seed_date, seed_period)
        for item in raw:
            if item.get('_skip'):
                continue
            if item['remaining_months'] <= 0 and item['months_amortized'] > 0:
                item['completed_period'] = seed_period
                completed.append(item)
            else:
                all_items.append(item)
        print(f'  135150 PPD Other: {len(all_items)} active, {len(completed)} completed')
    else:
        print('  WARNING: 135150 PPD Other tab not found')

    if 'Insurance Analysis' in wb.sheetnames:
        ins_items = _read_insurance_for(wb['Insurance Analysis'], seed_date, seed_period)
        all_items.extend(ins_items)
        print(f'  Insurance Analysis: {len(ins_items)} active policies')
    else:
        print('  WARNING: Insurance Analysis tab not found')

    return all_items, completed


def _read_135150_raw(ws, seed_date: date, seed_period: str) -> list:
    """
    Read 135150 tab and compute elapsed / remaining as of seed_date.
    Items that haven't started yet get _skip=True.
    """
    items = []
    for row in ws.iter_rows(min_row=9, values_only=True):
        vendor      = row[1]
        gl_acct_raw = row[2]
        pay_date    = row[3]
        total_amt   = row[4]
        period_cov  = row[5]
        tot_months  = row[6]
        monthly_amt = row[7]
        # Ignore JLL's elapsed/remaining — recalculate from dates

        if not vendor or not isinstance(vendor, str):
            continue
        vendor = vendor.strip()
        if not vendor or vendor.lower() in ('vendor', '\\', ''):
            continue
        if str(vendor).startswith('Ending') or str(vendor).startswith('Total'):
            continue
        if not total_amt or not period_cov:
            continue

        # GL account
        if isinstance(gl_acct_raw, (int, float)):
            gl_acct = str(int(gl_acct_raw))
        elif gl_acct_raw:
            s = str(gl_acct_raw).strip()
            gl_acct = s[:-2] if s.endswith('.0') else s
        else:
            continue  # no GL account = note/comment row

        total_months_v  = int(_safe_float(tot_months))  if tot_months  else 0
        monthly_amt_v   = _safe_float(monthly_amt)
        total_amount_v  = _safe_float(total_amt)
        invoice_date_v  = _ensure_date(pay_date)

        service_start, service_end = _parse_term(str(period_cov or ''))
        if not service_start:
            continue

        # Skip items that haven't started yet as of seed_date
        if service_start > seed_date:
            continue

        # Determine daily_rate items
        daily_rate = 0.0
        if service_start and service_end and total_months_v >= 28:
            span_months = (
                (service_end.year - service_start.year) * 12
                + service_end.month - service_start.month
            )
            if span_months <= 1:
                if total_months_v > 0:
                    daily_rate = round(total_amount_v / total_months_v, 6)

        # Compute months/days elapsed as of seed_date
        if daily_rate > 0:
            # Day-based: count days from service_start to seed_date (inclusive)
            days_elapsed = (seed_date - service_start).days + 1
            days_elapsed = min(days_elapsed, total_months_v)
            days_remaining = max(0, total_months_v - days_elapsed)
            months_amort = days_elapsed
            remaining    = days_remaining
        else:
            seed_month_start = date(seed_date.year, seed_date.month, 1)
            svc_month_start  = date(service_start.year, service_start.month, 1)
            # months elapsed = difference in months.
            # If service starts on day 1, the start month is a full amortization
            # period so we add 1 (matches JLL convention).
            # If service starts mid-month, the start month is partial / not counted.
            offset = 1 if service_start.day == 1 else 0
            months_amort = (
                (seed_month_start.year - svc_month_start.year) * 12
                + seed_month_start.month - svc_month_start.month
                + offset
            )
            months_amort = max(0, min(months_amort, total_months_v))
            remaining    = max(0, total_months_v - months_amort)

        item = {
            'vendor':             vendor,
            'invoice_number':     '',
            'invoice_date':       invoice_date_v,
            'description':        f'{vendor} — {period_cov or ""}',
            'gl_account_number':  gl_acct,
            'gl_account':         _GL_NAMES.get(gl_acct, gl_acct),
            'total_amount':       total_amount_v,
            'monthly_amount':     monthly_amt_v,
            'service_start':      service_start,
            'service_end':        service_end,
            'total_months':       total_months_v,
            'months_amortized':   months_amort,
            'remaining_months':   remaining,
            'first_added_period': seed_period,
            'daily_rate':         daily_rate,
            '_skip':              False,
        }
        items.append(item)

    return items


def _read_insurance_for(ws, seed_date: date, seed_period: str) -> list:
    """Insurance Analysis reader parameterised on seed_date."""
    items = []
    last_policy_type = ''

    for r in range(8, 35):
        accrual_marker = ws.cell(row=r, column=3).value
        policy_type    = ws.cell(row=r, column=4).value
        term_str       = ws.cell(row=r, column=5).value
        premium        = ws.cell(row=r, column=6).value
        monthly        = ws.cell(row=r, column=7).value

        if policy_type and isinstance(policy_type, str) and policy_type.strip():
            last_policy_type = policy_type.strip()

        if not premium:
            continue
        if isinstance(accrual_marker, str) and accrual_marker.strip().lower() == 'accr':
            continue
        if not term_str:
            continue
        if not policy_type or not str(policy_type).strip():
            policy_type = last_policy_type

        premium_v = _safe_float(premium)
        monthly_v = _safe_float(monthly)
        if premium_v <= 0 or monthly_v <= 0:
            continue

        service_start, service_end = _parse_term(str(term_str or ''))
        if not service_start or not service_end:
            continue
        if service_start > seed_date:
            continue  # not started yet
        if service_end <= seed_date:
            continue  # already expired

        pt = str(policy_type or '').strip().lower()
        if 'property' in pt:
            gl_acct, gl_name = '639110', 'Insurance-Property'
        elif 'umbrella' in pt:
            gl_acct, gl_name = '639120', 'Insurance-Liability (Umbrella)'
        elif 'general' in pt or 'liability' in pt:
            gl_acct, gl_name = '639120', 'Insurance-General Liability'
        else:
            gl_acct, gl_name = '639120', f'Insurance-{policy_type}'

        seed_mo   = date(seed_date.year, seed_date.month, 1)
        svc_mo    = date(service_start.year, service_start.month, 1)
        months_elapsed = (
            (seed_mo.year - svc_mo.year) * 12
            + seed_mo.month - svc_mo.month
        )
        total_months = (
            (service_end.year - service_start.year) * 12
            + service_end.month - service_start.month
        )
        remaining = max(0, total_months - months_elapsed)

        item = {
            'vendor':             'Insurance',
            'invoice_number':     '',
            'invoice_date':       service_start,
            'description':        str(policy_type or '').strip(),
            'gl_account_number':  gl_acct,
            'gl_account':         gl_name,
            'total_amount':       premium_v,
            'monthly_amount':     round(monthly_v, 2),
            'service_start':      service_start,
            'service_end':        service_end,
            'total_months':       total_months,
            'months_amortized':   months_elapsed,
            'remaining_months':   remaining,
            'first_added_period': seed_period,
            'daily_rate':         0.0,
        }
        items.append(item)

    return items


def _print_summary(label, active, completed):
    print(f'\n{"="*60}')
    print(f'  {label}')
    print(f'{"="*60}')
    print(f'Active items ({len(active)}):')
    for item in active:
        print(f"  {item['vendor'][:38]:38s}  GL {item['gl_account_number']}  "
              f"${item['monthly_amount']:8.2f}/mo  "
              f"{item['months_amortized']:3} elapsed / {item['remaining_months']:3} remaining")
    print(f'\nCompleted items ({len(completed)}):')
    for item in completed:
        print(f"  {item['vendor'][:38]:38s}  GL {item['gl_account_number']}  "
              f"${item['total_amount']:,.2f}")


def main():
    print(f'Reading: {RCW_PATH}')
    if not os.path.exists(RCW_PATH):
        print('ERROR: RCW file not found. Check the path.')
        sys.exit(1)

    tmp = tempfile.mktemp(suffix='.xlsx')
    shutil.copy2(RCW_PATH, tmp)

    try:
        from openpyxl import load_workbook
        wb = load_workbook(tmp, data_only=True)

        # ── Dec 31, 2025  (prior ledger for January 2026 Pass 1) ──────────
        print('\n--- Building Dec 31, 2025 seed ---')
        active_dec, comp_dec = _build_for_seed_date(wb, date(2025, 12, 31), 'Dec-2025')
        _print_summary('Dec 31, 2025', active_dec, comp_dec)
        _ledger_save(active_dec, comp_dec, OUT_PATH_DEC)
        print(f'\nSaved: {OUT_PATH_DEC}')

        # ── Mar 31, 2026  (kept for reference / April use) ────────────────
        print('\n--- Building Mar 31, 2026 seed ---')
        active_mar, comp_mar = _build_for_seed_date(wb, date(2026, 3, 31), 'Mar-2026')
        _print_summary('Mar 31, 2026', active_mar, comp_mar)
        _ledger_save(active_mar, comp_mar, OUT_PATH_MAR)
        print(f'\nSaved: {OUT_PATH_MAR}')

    finally:
        os.unlink(tmp)

    print('\n\nDone. Upload instructions:')
    print('  Jan Pass 1: use GA_Prepaid_Ledger_Seed_Dec2025.xlsx as Prior Prepaid Ledger')
    print('  Feb Pass 1: use the GA_Prepaid_Ledger_Updated.xlsx output from Jan Pass 1')
    print('  Mar Pass 1: use the GA_Prepaid_Ledger_Updated.xlsx output from Feb Pass 1')
    print('  Apr Pass 1: use the GA_Prepaid_Ledger_Updated.xlsx output from Mar Pass 1')


if __name__ == '__main__':
    main()


"""
Yardi Income Statement (Accrual) Parser

Parses Yardi Income Statement export files with hierarchical account structure:
- Rows 1-4: Meta information (property, report type, period, book/tree)
- Row 5: Column headers (Period to Date %, Year to Date %)
- Row 6+: Hierarchical account data with indented descriptions

Expected columns:
  Account Code, Account Name, PTD Amount, PTD %, YTD Amount, YTD %

The account hierarchy is indicated by spacing in the Account Name column.

Features:
- Handles hierarchical GL accounts with indentation
- Extracts period and YTD metrics
- Preserves account hierarchy structure
- Normalizes numeric values and null handling
"""

from openpyxl import load_workbook
from datetime import datetime
from typing import List, Dict, Tuple, Optional


def _detect_is_columns(ws) -> Tuple[int, int, int, int, int]:
    """
    Scan header rows (1-8) to locate the PTD and YTD amount / percent columns.

    Handles single-row headers ("Period to Date", "Year to Date") and two-row
    split headers (group header on one row, "Amount" / "%" on the next).

    Returns (ptd_col, ptd_pct_col, ytd_col, ytd_pct_col, data_start_row)
    where all column indices are 0-based and data_start_row is 1-based.
    Falls back to (2, 3, 4, 5, 6) if no header row is detected.
    """
    max_col = min(getattr(ws, 'max_column', 8) + 2, 12)

    # Accumulate text across all header rows per column (handles split headers).
    combined = [''] * max_col
    last_header_row = 5

    for hrow in range(1, min(9, ws.max_row + 1)):
        hvals = [
            str(ws.cell(row=hrow, column=c).value or '').strip().lower()
            for c in range(1, max_col + 1)
        ]
        # Stop scanning once we hit a data row (col A looks like an account code)
        col_a = str(ws.cell(row=hrow, column=1).value or '').strip()
        if col_a and col_a.replace('-', '').replace(' ', '').isdigit():
            break
        # Only absorb rows that have something interesting beyond cols A/B
        if any(hvals[2:]):
            last_header_row = hrow
            for i, v in enumerate(hvals[:max_col]):
                if v:
                    combined[i] += ' ' + v

    data_start = last_header_row + 1

    # Build candidate lists from combined column text
    period_cols = [i for i, t in enumerate(combined) if i >= 2 and ('period' in t or 'ptd' in t)]
    year_cols   = [i for i, t in enumerate(combined) if i >= 2 and ('year' in t or 'ytd' in t)]
    amount_cols = [i for i, t in enumerate(combined) if i >= 2 and 'amount' in t]
    pct_cols    = [i for i, t in enumerate(combined) if i >= 2 and ('%' in t or 'percent' in t)]

    # PTD column: prefer explicit 'period'/'ptd' keyword, else first 'amount'
    ptd_col = period_cols[0] if period_cols else (amount_cols[0] if amount_cols else 2)

    # YTD column: prefer explicit 'year'/'ytd', else second 'amount'
    ytd_col = year_cols[0] if year_cols else (
        amount_cols[1] if len(amount_cols) >= 2 else ptd_col + 2
    )

    # Percent columns: first/second '%' occurrence after their respective amount cols
    ptd_pct_cols = [c for c in pct_cols if c > ptd_col and c < ytd_col]
    ytd_pct_cols = [c for c in pct_cols if c > ytd_col]
    ptd_pct_col = ptd_pct_cols[0] if ptd_pct_cols else ptd_col + 1
    ytd_pct_col = ytd_pct_cols[0] if ytd_pct_cols else ytd_col + 1

    return ptd_col, ptd_pct_col, ytd_col, ytd_pct_col, data_start


def parse(filepath: str) -> List[Dict]:
    """
    Parse a Yardi Income Statement export file.

    Args:
        filepath: Path to the Excel file

    Returns:
        List of dictionaries representing income statement line items

    Raises:
        FileNotFoundError: If file does not exist
        ValueError: If file structure is invalid
    """
    try:
        wb = load_workbook(filepath)
    except Exception as e:
        raise FileNotFoundError(f"Cannot open file: {filepath}") from e

    ws = wb.active
    data = []

    # Meta information is in rows 1-4
    metadata = _extract_metadata(ws)

    # Detect column positions dynamically from header rows
    ptd_col, ptd_pct_col, ytd_col, ytd_pct_col, data_start_row = _detect_is_columns(ws)

    # Validate headers (extract_headers used only for the None-check)
    headers = _extract_headers(ws, row=5)
    if not headers:
        raise ValueError("Cannot extract headers from Income Statement file")

    # Process data rows starting after the detected header row
    for row_num in range(data_start_row, ws.max_row + 1):
        row = ws[row_num]
        row_values = [cell.value for cell in row]

        # Skip completely empty rows
        if all(v is None for v in row_values):
            continue

        # Account code should be in first column
        account_code = row_values[0]
        account_name = row_values[1] if len(row_values) > 1 else None

        # Skip rows without an account code
        if account_code is None:
            continue

        # Build record — force code/name to str (Excel may return floats)
        record = {
            'account_code': str(_normalize_value(account_code) or '').strip(),
            'account_name': str(_normalize_value(account_name) or '').strip(),
        }

        # Extract numeric values using dynamically detected column positions
        if len(row_values) > ptd_col:
            record['ptd_amount']  = _normalize_numeric(row_values[ptd_col])
        if len(row_values) > ptd_pct_col:
            record['ptd_percent'] = _normalize_numeric(row_values[ptd_pct_col])
        if len(row_values) > ytd_col:
            record['ytd_amount']  = _normalize_numeric(row_values[ytd_col])
        if len(row_values) > ytd_pct_col:
            record['ytd_percent'] = _normalize_numeric(row_values[ytd_pct_col])

        # Add metadata
        record.update(metadata)

        data.append(record)

    return data


def validate(filepath: str) -> Tuple[bool, List[str]]:
    """
    Validate that a file has the expected Income Statement structure.

    Args:
        filepath: Path to the Excel file

    Returns:
        Tuple of (is_valid: bool, issues: list of error strings)
    """
    issues = []

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        return False, [f"Cannot open file: {e}"]

    ws = wb.active

    # Check for expected meta rows
    if not ws.cell(1, 1).value or "Property" not in str(ws.cell(1, 1).value):
        issues.append("Row 1 missing 'Property' meta information")

    if not ws.cell(2, 1).value or "Income Statement" not in str(ws.cell(2, 1).value):
        issues.append("Row 2 missing 'Income Statement' title")

    # Check headers
    headers = _extract_headers(ws, row=5)
    if not headers:
        issues.append("Cannot extract headers from row 5")
    else:
        # Should have at least Account Code and Account Name
        if len(headers) < 2:
            issues.append("Expected at least 2 header columns")

    return len(issues) == 0, issues


def _extract_metadata(ws) -> Dict:
    """Extract metadata from rows 1-4."""
    metadata = {}

    # Row 1: Property
    prop_line = ws.cell(1, 1).value
    if prop_line:
        parts = str(prop_line).split('=')
        if len(parts) > 1:
            metadata['property'] = parts[1].strip()

    # Row 2: Report type
    report_line = ws.cell(2, 1).value
    if report_line:
        metadata['report_type'] = str(report_line).strip()

    # Row 3: Period
    period_line = ws.cell(3, 1).value
    if period_line:
        parts = str(period_line).split('=')
        if len(parts) > 1:
            metadata['period'] = parts[1].strip()

    # Row 4: Book/Tree
    book_line = ws.cell(4, 1).value
    if book_line:
        parts = str(book_line).split(';')
        for part in parts:
            part = part.strip()
            if '=' in part:
                key, val = part.split('=', 1)
                metadata[key.strip().lower()] = val.strip()

    return metadata


def _extract_headers(ws, row: int) -> List[str]:
    """Extract and clean headers from a specific row."""
    headers = []
    for cell in ws[row]:
        value = cell.value
        if value:
            headers.append(str(value).strip())
        else:
            headers.append(None)
    return headers


def _normalize_value(value):
    """Normalize values for consistent output."""
    if value is None:
        return None

    # Convert datetime to ISO format string
    if isinstance(value, datetime):
        return value.isoformat()

    # Handle strings - strip whitespace
    if isinstance(value, str):
        return value.strip()

    return value


def _normalize_numeric(value):
    """Normalize numeric values, handling None and strings."""
    if value is None:
        return None

    # Try to convert to float
    if isinstance(value, (int, float)):
        return value

    if isinstance(value, str):
        try:
            # Try int first
            if '.' not in value:
                return int(value)
            return float(value)
        except (ValueError, AttributeError):
            # If it looks like 'N/A', return as string
            if value.upper() == 'N/A':
                return 'N/A'
            return None

    return value


if __name__ == "__main__":
    import sys
    import json

    if len(sys.argv) < 2:
        print("Usage: python yardi_income_statement.py <filepath>")
        sys.exit(1)

    filepath = sys.argv[1]

    # Validate
    is_valid, issues = validate(filepath)
    if not is_valid:
        print(f"Validation errors:")
        for issue in issues:
            print(f"  - {issue}")
        sys.exit(1)

    # Parse
    data = parse(filepath)
    print(f"Successfully parsed {len(data)} income statement line items")
    print(f"\nSample records (first 2 entries):")
    for i, record in enumerate(data[:2]):
        print(f"\nRecord {i+1}:")
        print(json.dumps(record, indent=2, default=str))

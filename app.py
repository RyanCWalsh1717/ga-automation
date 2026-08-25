"""
GA Automation — Monthly Report Pipeline (Two-Pass)
====================================================
Pass 1 (Pre-Close):  Upload pre-close Yardi GL + supporting files → detect
                     accruals → export 3 JE CSVs for Yardi upload.

Pass 2 (Post-Close): After JEs are posted to Yardi, upload final GL →
                     generate BS workpaper, QC, variance comments, exception
                     report. (Singerman workbook is downloaded directly from Yardi.)
"""

import streamlit as st
import sys
import os
import re
import tempfile
import shutil
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional, Tuple

# ── Setup paths ──────────────────────────────────────────────
pipeline_dir = Path(__file__).parent / "pipeline"
if str(pipeline_dir) not in sys.path:
    sys.path.insert(0, str(pipeline_dir))

from engine import run_pipeline, EngineResult, Exception_
from property_config import is_revenue_account, is_income_statement_account
from report_generator import generate_exception_report
import traceback
from accrual_entry_generator import (
    build_accrual_entries, generate_yardi_je_csv, generate_etl_csv,
    build_prepaid_amortization, build_prepaid_release_je, build_prepaid_reclass_je,
    check_prior_accrual_vs_actual, build_budget_based_accruals, BUDGET_BASED_ACCOUNTS,
)
import prepaid_ledger
import bs_workpaper_generator
from variance_comments import (
    generate_variance_comments,
    generate_variance_comments_grp,
    write_comments_to_budget_comparison,
)
from qc_engine import run_qc, generate_qc_workbook
from management_fee import (
    calculate as calculate_mgmt_fee,
    accrued_fee_from_bc,
    build_management_fee_je,
    detect_prior_period_catchup,
    build_catchup_je,
)
from mgmt_fee_invoice import generate_invoice as generate_mgmt_fee_invoice
from audit_trail_generator import generate_audit_trail


# ── Data directory + property discovery ──────────────────────────────────────
_DATA_DIR = Path(__file__).parent / "data"

# Sign-off checklist items — shared between the Pass 2 sign-off UI and the
# Audit Trail's Close & Signoff tab so both always reference the same list.
_SIGNOFF_ITEMS = [
    "Bank Reconciliation — Operating",
    "Bank Reconciliation — DACA",
    "Management Fee Invoice",
    "GL vs TB Workpaper Tie-out",
    "Variance Commentary",
    "QC Checklist (7-point)",
    "Equity Tabs (311100 / 331100 / 381100)",
    "Exception Report",
]

def _month_abbr(month_num: int) -> str:
    """Return 3-letter month abbreviation from a month number (1=Jan … 12=Dec)."""
    return ['Jan','Feb','Mar','Apr','May','Jun',
            'Jul','Aug','Sep','Oct','Nov','Dec'][(month_num - 1) % 12]


def _committed_path(prop_code: str, filename: str) -> Optional[str]:
    """Return path to a committed reference file if it exists, else None."""
    p = _DATA_DIR / prop_code / filename
    return str(p) if p.exists() else None


def _load_coa_codes(cfg) -> Optional[dict]:
    """
    Return {account_code: account_name} from whichever Chart of Accounts is
    in effect for this property — the shared GRP COA if cfg.uses_grp_coa,
    else this property's own uploaded chart_of_accounts.{xlsx,xls,csv}.
    Returns None if no COA is on file (QC Check 8 skips itself in that case).
    """
    if getattr(cfg, 'uses_grp_coa', False):
        _coa_path = _committed_path('_shared', 'GRP_Chart_of_Accounts.xlsx')
    else:
        _coa_path = None
        for _ext in ('.xlsx', '.xls', '.csv'):
            _coa_path = _committed_path(cfg.property_code, f'chart_of_accounts{_ext}')
            if _coa_path:
                break
    if not _coa_path:
        return None
    try:
        from parsers.yardi_chart_of_accounts import parse as _parse_coa
        _coa_res = _parse_coa(_coa_path)
        return _coa_res.accounts or None
    except Exception:
        return None


def _build_accruals_seed_df(cfg=None):
    """
    Build the one-off accruals seed DataFrame from property config defaults.
    Falls back to RevLabs pre-seeded rows when the config has no default_accruals.
    """
    import pandas as _pd_seed
    if cfg is not None and getattr(cfg, 'default_accruals', None):
        rows = cfg.default_accruals
        _n = len(rows) + 1   # one blank trailing row
        return _pd_seed.DataFrame({
            "Account Code":   [r.get('account_code', '') for r in rows] + [''],
            "Account Name":   [r.get('account_name', '') for r in rows] + [''],
            "Vendor":         [r.get('vendor', '') for r in rows] + [''],
            "Amount ($)":     [0.0] * _n,
            "Prior Accrual ($)": [0.0] * _n,
            "Description":    [''] * _n,
            "Auto-Reverse":   [True] * _n,
            "Split Schedule": [''] * _n,
            # Per-account default from config.yaml — only accounts genuinely
            # billed irregularly (e.g. Water Contract Svc) should compound.
            # Defaults to False (flat monthly) when a row doesn't specify it.
            "Compound":       [bool(r.get('compound', False)) for r in rows] + [False],
        })
    # No default_accruals in config — return a single blank row so the editor renders
    return _pd_seed.DataFrame({
        "Account Code":   [''],
        "Account Name":   [''],
        "Vendor":         [''],
        "Amount ($)":     [0.0],
        "Prior Accrual ($)": [0.0],
        "Description":    [''],
        "Auto-Reverse":   [True],
        "Split Schedule": [''],
        "Compound":       [False],
    })


_OA_COLUMNS = [
    "Account Code", "Account Name", "Vendor", "Amount ($)", "Prior Accrual ($)",
    "Description", "Auto-Reverse", "Split Schedule", "Compound",
]


def _blank_oa_row() -> dict:
    return {
        "Account Code": "", "Account Name": "", "Vendor": "",
        "Amount ($)": 0.0, "Prior Accrual ($)": 0.0, "Description": "",
        "Auto-Reverse": True, "Split Schedule": "", "Compound": False,
    }


def _df_to_oa_rows(df, cfg=None) -> list:
    """
    Convert the manual_accruals_df DataFrame into a list of plain dicts for the
    One-Off Accruals plain-widget row list. Missing columns fall back to
    defaults per-field (rather than mutating the DataFrame first) so an old
    session's DataFrame from before a column was added still converts cleanly.
    """
    if df is None or df.empty:
        return [_blank_oa_row()]
    has_compound_col = "Compound" in df.columns
    # Migration only: sessions saved before the Compound column existed get
    # each row's default from config.yaml (default_accruals[].compound) keyed
    # by account code, rather than blindly defaulting to False for everyone.
    # Confirmed with Ryan 2026-07-28 — only genuinely irregular-billed accounts
    # (e.g. Water Contract Svc) should compound.
    _compound_defaults = {
        str(r.get('account_code', '')).strip(): bool(r.get('compound', False))
        for r in (getattr(cfg, 'default_accruals', None) or [])
    }
    rows = []
    for _, r in df.iterrows():
        _code = str(r.get("Account Code", "") or "").strip()
        rows.append({
            "Account Code": _code,
            "Account Name": str(r.get("Account Name", "") or "").strip(),
            "Vendor": str(r.get("Vendor", "") or "").strip(),
            "Amount ($)": float(r.get("Amount ($)", 0) or 0),
            "Prior Accrual ($)": float(r.get("Prior Accrual ($)", 0) or 0),
            "Description": str(r.get("Description", "") or "").strip(),
            "Auto-Reverse": bool(r.get("Auto-Reverse", True)),
            "Split Schedule": str(r.get("Split Schedule", "") or "").strip(),
            "Compound": bool(r.get("Compound", False)) if has_compound_col
                        else _compound_defaults.get(_code, False),
        })
    return rows or [_blank_oa_row()]


_IC_COLUMNS = ["Leg", "Account", "Account Name", "Credit ($)", "Debit ($)", "Description"]


def _blank_ic_row(leg: str = "") -> dict:
    return {
        "Leg": leg, "Account": "", "Account Name": "",
        "Credit ($)": 0.0, "Debit ($)": 0.0, "Description": "",
    }


def _df_to_ic_rows(df) -> list:
    """Convert interco_recode_df into a list of plain dicts, same pattern as _df_to_oa_rows."""
    if df is None or df.empty:
        return []
    rows = []
    for _, r in df.iterrows():
        rows.append({
            "Leg": str(r.get("Leg", "") or "").strip(),
            "Account": str(r.get("Account", "") or "").strip(),
            "Account Name": str(r.get("Account Name", "") or "").strip(),
            "Credit ($)": float(r.get("Credit ($)", 0) or 0),
            "Debit ($)": float(r.get("Debit ($)", 0) or 0),
            "Description": str(r.get("Description", "") or "").strip(),
        })
    return rows


def _read_interco_df_from_widgets():
    """
    Rebuild the interco_recode_df-shaped DataFrame directly from the live
    plain-widget row state (see the 7xxxxx Intercompany Recode Table block).
    Reads st.session_state[f"ic_*_{rid}"] rather than st.session_state.interco_recode_df
    so callers upstream of that block's own render/write-back (e.g. the Pass 1
    JE-build step, which runs earlier in script order) still see the user's
    latest typed values — plain widgets update their session_state key the
    instant the user interacts, unlike st.data_editor's internal edited_rows
    diff, which previously had to be peeked at separately for this same reason.
    Falls back to the plain DataFrame mirror if the row list hasn't been
    seeded yet this session (e.g. before Pass 1 has ever completed once).
    """
    import pandas as _pd_ic_read
    if "ic_row_ids" not in st.session_state:
        return st.session_state.get("interco_recode_df")
    _rows = [
        {
            "Leg": st.session_state.get(f"ic_leg_{_rid}", ""),
            "Account": st.session_state.get(f"ic_account_{_rid}", ""),
            "Account Name": st.session_state.get(f"ic_name_{_rid}", ""),
            "Credit ($)": float(st.session_state.get(f"ic_credit_{_rid}", 0.0) or 0.0),
            "Debit ($)": float(st.session_state.get(f"ic_debit_{_rid}", 0.0) or 0.0),
            "Description": st.session_state.get(f"ic_desc_{_rid}", ""),
        }
        for _rid in st.session_state.get("ic_row_ids", [])
    ]
    return _pd_ic_read.DataFrame(_rows, columns=_IC_COLUMNS)


def _save_checklist_now() -> None:
    """Persist current close_tracker + custom items to GitHub/local."""
    try:
        from checklist_persistence import save_checklist, session_to_state, period_to_key
        _code = st.session_state.get('active_property_code') or ''
        _pkey = st.session_state.get('checklist_period_key', current_period_key())
        _state = session_to_state(
            st.session_state.close_tracker,
            st.session_state.get('custom_checklist_items', []),
            _code, _pkey,
            locked    = st.session_state.get('checklist_locked', False),
            locked_by = st.session_state.get('checklist_locked_by'),
            locked_at = st.session_state.get('checklist_locked_at'),
        )
        save_checklist(_code, _pkey, _state, str(_DATA_DIR))
    except Exception:
        pass   # persistence is best-effort; never crash the app


def _discover_properties() -> list[dict]:
    """
    Scan data/ for subfolders with a config.yaml.
    Returns list of {'code', 'display_name', 'address', 'cfg'} dicts.
    Returns empty list if none found — UI handles the no-property state.
    """
    from property_config import discover_properties as _disc
    return _disc(str(_DATA_DIR))


# ── Page configuration ───────────────────────────────────────
st.set_page_config(
    page_title="Close Pipeline",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Custom CSS ───────────────────────────────────────────────
# st.html() is the supported API for raw HTML injection in Streamlit ≥ 1.31.
# st.markdown(..., unsafe_allow_html=True) was deprecated in 1.36 and strips
# complex div structures; use st.html() for all HTML/CSS blocks.
st.html("""
<style>
    /* ── Brand palette ── */
    :root {
        --grp-green:     #1A5C22;
        --grp-green-mid: #2E7D32;
        --grp-green-lit: #E8F5E9;
        --success-color: #2E7D32;
        --warning-color: #E65100;
        --error-color:   #B71C1C;
        --info-color:    #1565C0;
        --text-primary:  #212121;
        --text-muted:    #616161;
        --border:        #E0E0E0;
        --bg-alt:        #F9FAFB;
    }

    /* ── Hero banner ── */
    .grp-hero {
        background: linear-gradient(135deg, var(--grp-green) 0%, var(--grp-green-mid) 100%);
        border-radius: 10px;
        padding: 0;
        overflow: hidden;
        margin-bottom: 18px;
        box-shadow: 0 4px 12px rgba(0,0,0,0.18);
        display: flex;
        align-items: stretch;
        min-height: 110px;
    }
    .grp-hero-photo {
        width: 260px;
        min-width: 260px;
        object-fit: cover;
        border-radius: 10px 0 0 10px;
        display: block;
    }
    .grp-hero-body {
        padding: 20px 28px;
        flex: 1;
        display: flex;
        flex-direction: column;
        justify-content: center;
    }
    .grp-hero-title {
        color: #ffffff;
        font-size: 1.55rem;
        font-weight: 700;
        letter-spacing: 0.02em;
        margin: 0 0 4px 0;
        line-height: 1.2;
    }
    .grp-hero-sub {
        color: #C8E6C9;
        font-size: 0.88rem;
        margin: 0 0 10px 0;
        font-weight: 400;
    }
    .grp-hero-badges {
        display: flex;
        gap: 8px;
        flex-wrap: wrap;
        margin-top: 4px;
    }
    .grp-badge {
        background: rgba(255,255,255,0.18);
        border: 1px solid rgba(255,255,255,0.35);
        border-radius: 20px;
        padding: 3px 12px;
        color: #ffffff;
        font-size: 0.75rem;
        font-weight: 500;
        white-space: nowrap;
    }
    .grp-hero-logo {
        padding: 20px 24px;
        display: flex;
        align-items: center;
        justify-content: flex-end;
        min-width: 160px;
    }
    .grp-logo-text {
        color: rgba(255,255,255,0.85);
        font-size: 0.7rem;
        text-align: right;
        line-height: 1.4;
        font-weight: 500;
        letter-spacing: 0.08em;
        text-transform: uppercase;
    }

    /* ── Sidebar property card ── */
    .grp-sidebar-card {
        background: linear-gradient(135deg, var(--grp-green) 0%, var(--grp-green-mid) 100%);
        border-radius: 8px;
        padding: 14px 16px;
        margin-bottom: 12px;
        color: white;
    }
    .grp-sidebar-prop {
        font-size: 0.82rem;
        font-weight: 700;
        color: #ffffff;
        margin: 0 0 2px 0;
        letter-spacing: 0.03em;
    }
    .grp-sidebar-addr {
        font-size: 0.72rem;
        color: #C8E6C9;
        margin: 0;
    }

    /* ── Section headers ── */
    .grp-section {
        color: var(--grp-green);
        font-size: 1.05rem;
        font-weight: 700;
        border-bottom: 2px solid var(--grp-green-lit);
        padding-bottom: 4px;
        margin: 18px 0 10px 0;
    }

    /* ── Status pills ── */
    .status-clean    { color: var(--success-color); font-weight: 600; }
    .status-warnings { color: var(--warning-color); font-weight: 600; }
    .status-errors   { color: var(--error-color);   font-weight: 600; }

    /* ── Exception cards ── */
    .exception-error {
        background-color: #FFEBEE;
        border-left: 4px solid var(--error-color);
        padding: 10px 14px; margin: 8px 0; border-radius: 5px;
    }
    .exception-warning {
        background-color: #FFF8E1;
        border-left: 4px solid var(--warning-color);
        padding: 10px 14px; margin: 8px 0; border-radius: 5px;
    }
    .exception-info {
        background-color: #E3F2FD;
        border-left: 4px solid var(--info-color);
        padding: 10px 14px; margin: 8px 0; border-radius: 5px;
    }

    /* ── Metric cards ── */
    .metric-card {
        background: #ffffff;
        padding: 16px 20px;
        border-radius: 8px;
        border-left: 4px solid var(--grp-green);
        box-shadow: 0 1px 4px rgba(0,0,0,0.08);
    }

    /* ── Streamlit element tweaks ── */
    .stTabs [data-baseweb="tab-list"] {
        gap: 6px;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 6px 6px 0 0;
        font-weight: 500;
    }
    div[data-testid="stSidebarContent"] {
        padding-top: 12px;
    }
</style>
""")


# ── Session state initialization ─────────────────────────────
# Shared
if "uploaded_files" not in st.session_state:
    st.session_state.uploaded_files = {}
if "temp_dir" not in st.session_state:
    st.session_state.temp_dir = tempfile.mkdtemp(prefix="ga_automation_")
if "active_property_code" not in st.session_state:
    # Default to first discovered property
    _init_props = _discover_properties()
    st.session_state.active_property_code = _init_props[0]['code'] if _init_props else None
if "_prev_active_property_code" not in st.session_state:
    st.session_state._prev_active_property_code = st.session_state.active_property_code

# Pass 1 — JE Generation
if "pass1_complete" not in st.session_state:
    st.session_state.pass1_complete = False
if "pass1_engine_result" not in st.session_state:
    st.session_state.pass1_engine_result = None
if "pass1_output_files" not in st.session_state:
    st.session_state.pass1_output_files = {}
if "pass1_run_count" not in st.session_state:
    st.session_state.pass1_run_count = 0
if "upload_key_p1" not in st.session_state:
    st.session_state.upload_key_p1 = 0
if "tub_key" not in st.session_state:
    st.session_state.tub_key = 0

# Pass 2 — Report Generation
if "pass2_complete" not in st.session_state:
    st.session_state.pass2_complete = False
if "pass2_engine_result" not in st.session_state:
    st.session_state.pass2_engine_result = None
if "pass2_output_files" not in st.session_state:
    st.session_state.pass2_output_files = {}
if "upload_key_p2" not in st.session_state:
    st.session_state.upload_key_p2 = 0
if "editor_reset_count" not in st.session_state:
    st.session_state.editor_reset_count = 0
if "je_excluded_jes" not in st.session_state:
    st.session_state.je_excluded_jes = set()
if "je_amount_overrides" not in st.session_state:
    st.session_state.je_amount_overrides = {}   # {je_number: adjusted_amount}

# Audit trail & sign-off
if "prepared_by" not in st.session_state:
    st.session_state.prepared_by = ""   # resolved after _active_cfg loads
if "signoff_state" not in st.session_state:
    st.session_state.signoff_state = {}
if "close_tracker" not in st.session_state:
    st.session_state.close_tracker = {}
if "confirm_reset_all" not in st.session_state:
    st.session_state.confirm_reset_all = False
if "confirm_reset_p2" not in st.session_state:
    st.session_state.confirm_reset_p2 = False

# Dashboard checklist persistence
from checklist_persistence import (
    current_period_key, period_key_to_label, period_to_key,
    load_checklist, save_checklist, session_to_state, state_to_session,
)
if "checklist_period_key" not in st.session_state:
    st.session_state.checklist_period_key = current_period_key()
if "custom_checklist_items" not in st.session_state:
    st.session_state.custom_checklist_items = []
if "checklist_loaded" not in st.session_state:
    st.session_state.checklist_loaded = False
if "checklist_locked" not in st.session_state:
    st.session_state.checklist_locked = False
if "checklist_locked_by" not in st.session_state:
    st.session_state.checklist_locked_by = None
if "checklist_locked_at" not in st.session_state:
    st.session_state.checklist_locked_at = None
if "last_completed_step" not in st.session_state:
    st.session_state.last_completed_step = None
if "bulk_overrides_wp" not in st.session_state:
    st.session_state.bulk_overrides_wp = {}
if "pass1_gl_activity_log" not in st.session_state:
    st.session_state['pass1_gl_activity_log'] = []
if "je_desc_overrides" not in st.session_state:
    st.session_state.je_desc_overrides = {}
if "interco_recode_df" not in st.session_state:
    import pandas as _pd_ic_init
    st.session_state.interco_recode_df = _pd_ic_init.DataFrame({
        "Leg":          _pd_ic_init.Series([], dtype=str),
        "Account":      _pd_ic_init.Series([], dtype=str),
        "Account Name": _pd_ic_init.Series([], dtype=str),
        "Credit ($)":   _pd_ic_init.Series([], dtype=float),
        "Debit ($)":    _pd_ic_init.Series([], dtype=float),
        "Description":  _pd_ic_init.Series([], dtype=str),
    })
# Bumped whenever interco_recode_df is externally replaced (property switch,
# Reset All, Reset Pass 1) or newly-detected 7xxxxx accounts are auto-merged
# into it — the Intercompany Recode plain-widget row list re-seeds from the
# DataFrame only when this counter changes, same pattern as
# _accruals_seed_gen for One-Off Accruals.
if "_interco_seed_gen" not in st.session_state:
    st.session_state._interco_seed_gen = 0

if "post_close_je_df" not in st.session_state:
    import pandas as _pd_init
    st.session_state.post_close_je_df = _pd_init.DataFrame({
        "JE #": ["PC-001", "PC-001"], "Description": ["", ""],
        "Account Code": ["", ""],
        "Debit ($)": [0.0, 0.0], "Credit ($)": [0.0, 0.0],
        "Line Description": ["", ""],
    })

# Tenant list for TUB sidebar inputs. Prefers the LIVE tenant list from an
# uploaded Tenancy Schedule (rent roll) this period — key = Yardi's own
# tenant code (e.g. 't0000017'), so the list never goes stale as leases turn
# over and never relies on anyone typing an arbitrary slug. Falls back to
# this property's static config.yaml tenants list (backward-compat, e.g.
# Revolution Labs, which doesn't upload a rent roll), or empty if neither
# is available. Each entry is (key, display_name).
def _build_tub_tenants(cfg) -> list:
    _ts_path = st.session_state.get('uploaded_files', {}).get('tenancy_schedule')
    if _ts_path and os.path.exists(_ts_path):
        try:
            from parsers.yardi_tenancy_schedule import parse as _parse_ts
            _ts_result = _parse_ts(_ts_path)
        except Exception:
            _ts_result = None
        if _ts_result and not _ts_result._parse_error and _ts_result.tenants:
            _multi_building = len({t.building_code for t in _ts_result.tenants}) > 1
            _seen: dict = {}
            for t in _ts_result.tenants:
                _label = f'{t.tenant_name} ({t.building_name})' if _multi_building else t.tenant_name
                _seen.setdefault(t.tenant_code, _label)
            return list(_seen.items())
    if getattr(cfg, 'tenants', None):
        return [(t['key'], t['name']) for t in cfg.tenants]
    return []

import pandas as pd  # needed for manual_accruals_df init and stale-session reset
if "manual_accruals_df" not in st.session_state:
    # cfg not available yet at this point — use RevLabs fallback; property-aware
    # reset happens after _active_cfg is loaded (see "Rebuild accruals seed" block).
    st.session_state.manual_accruals_df = _build_accruals_seed_df()
    st.session_state._accruals_df_for_property = None
# Bumped every time manual_accruals_df is externally replaced (property switch,
# Reset All, stale-column migration, session-snapshot restore) — the One-Off
# Accruals plain-widget row list re-seeds from the DataFrame only when this
# counter changes, so its own per-render write-back of manual_accruals_df
# (mirroring live widget edits for downstream consumers) never triggers a
# self-inflicted reseed.
if "_accruals_seed_gen" not in st.session_state:
    st.session_state._accruals_seed_gen = 0

# If session has stale columns from an older version, reset the whole table.
# "CR Account" is deprecated; "Auto-Reverse" is now a valid column so is NOT listed here.
if "CR Account" in st.session_state.manual_accruals_df.columns:
    st.session_state.manual_accruals_df = _build_accruals_seed_df()
    st.session_state._accruals_df_for_property = None
    st.session_state._accruals_seed_gen += 1

# Prepaid invoice correction table — lets the user fix a Nexus-parsed prepaid
# invoice (amount, service dates, GL account) directly from the Prepaid
# Expense Amortization panel when the parser misreads something. Keyed by
# the same normalized vendor+invoice-number key prepaid_ledger.py uses, so a
# correction here also fixes the ledger merge and reclass JE on the next run.
if "prepaid_overrides_df" not in st.session_state:
    st.session_state.prepaid_overrides_df = pd.DataFrame({
        "_key": pd.Series([], dtype=str),
        "Vendor": pd.Series([], dtype=str),
        "Invoice #": pd.Series([], dtype=str),
        "GL Account Number": pd.Series([], dtype=str),
        "GL Account Name": pd.Series([], dtype=str),
        "Description": pd.Series([], dtype=str),
        "Total Amount ($)": pd.Series([], dtype=float),
        "Service Start": pd.Series([], dtype=str),
        "Service End": pd.Series([], dtype=str),
        # Hidden snapshot of what Nexus reported for this invoice the FIRST
        # time this row was tracked — used to detect whether the source data
        # has since changed (e.g. a corrected/re-exported Nexus file), so a
        # stale override doesn't silently keep masking updated real data.
        "_orig_amount":        pd.Series([], dtype=float),
        "_orig_service_start": pd.Series([], dtype=str),
        "_orig_service_end":   pd.Series([], dtype=str),
        "_orig_gl_account":    pd.Series([], dtype=str),
    })


def _apply_prepaid_overrides(nexus_records: list, overrides_df) -> list:
    """
    Apply user corrections from the Prepaid Expense Amortization editor onto
    the parsed Nexus records, before they reach merge_nexus() / build_prepaid_
    amortization() / build_accrual_entries(). Matched by the same normalized
    vendor+invoice-number key prepaid_ledger._invoice_key() uses, so a fix
    here stays consistent with ledger matching.
    """
    if overrides_df is None or overrides_df.empty:
        return nexus_records
    from parsers.nexus_accrual import _count_months as _nex_count_months

    _overrides = {}
    for _, _row in overrides_df.iterrows():
        _k = str(_row.get("_key", "") or "").strip()
        if not _k:
            continue
        _overrides[_k] = _row

    if not _overrides:
        return nexus_records

    corrected = []
    for inv in nexus_records:
        if not inv.get('is_prepaid'):
            corrected.append(inv)
            continue
        _key = prepaid_ledger._invoice_key(inv.get('vendor', ''), inv.get('invoice_number', ''))
        _ov = _overrides.get(_key)
        if _ov is None:
            corrected.append(inv)
            continue

        _inv = dict(inv)
        try:
            # No ">0" guard — a user explicitly zeroing this out to void a
            # misparsed invoice must take effect (amount=0 then naturally
            # skips downstream, e.g. Layer 1's "if amount == 0: continue").
            _inv['amount'] = float(_ov.get("Total Amount ($)", 0) or 0)
        except (TypeError, ValueError):
            pass

        _svc_start = pd.to_datetime(_ov.get("Service Start", ""), errors='coerce')
        _svc_end = pd.to_datetime(_ov.get("Service End", ""), errors='coerce')
        if pd.notna(_svc_start):
            _inv['service_start'] = _svc_start.date()
        if pd.notna(_svc_end):
            _inv['service_end'] = _svc_end.date()
        if pd.notna(_svc_start) or pd.notna(_svc_end):
            _inv['prepaid_months'] = _nex_count_months(
                _inv.get('service_start'), _inv.get('service_end')
            )

        _gl_acct = str(_ov.get("GL Account Number", "") or "").strip()
        if _gl_acct:
            _inv['gl_account_number'] = _gl_acct

        corrected.append(_inv)
    return corrected


# ── Image asset loader ────────────────────────────────────────
import base64 as _b64

def _img_b64(fname: str) -> str | None:
    """Load an image from assets/ and return a base64 data-URI, or None."""
    _path = os.path.join(os.path.dirname(__file__), 'assets', fname)
    if os.path.exists(_path):
        with open(_path, 'rb') as _f:
            _raw = _b64.b64encode(_f.read()).decode()
        _ext = fname.rsplit('.', 1)[-1].lower()
        _mime = {'png': 'image/png', 'jpg': 'image/jpeg',
                 'jpeg': 'image/jpeg', 'svg': 'image/svg+xml'}.get(_ext, 'image/png')
        return f'data:{_mime};base64,{_raw}'
    return None

_LOGO_SRC   = _img_b64('grp_logo.png') or _img_b64('grp_logo.svg')

def _prop_hero_src(property_code: str) -> Optional[str]:
    """
    Return a base64 data-URI for the property's hero photo.

    Search order:
      1. data/{property_code}/hero.jpg|jpeg|png|webp  (property data dir)
      2. assets/{property_code}_hero.jpg|...           (exact code match)
      3. assets/{base_code}_hero.jpg|...               (code without trailing 'pm',
         e.g. revlabspm → revlabs_hero.jpg)
    Returns None if no image is found (banner renders text-only).
    """
    import base64 as _b64mod
    _exts = ['.jpg', '.jpeg', '.png', '.webp']

    # 1. Property-specific file in data dir
    for _ext in _exts:
        _p = _DATA_DIR / property_code / f'hero{_ext}'
        if _p.exists():
            _raw = _b64mod.b64encode(_p.read_bytes()).decode()
            _mime = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
                     'png': 'image/png', 'webp': 'image/webp'}.get(_ext.lstrip('.'), 'image/jpeg')
            return f'data:{_mime};base64,{_raw}'

    # 2. Assets folder — exact property code prefix (e.g. revlabspm_hero.jpg)
    for _ext in _exts:
        _src = _img_b64(f'{property_code}_hero{_ext}')
        if _src:
            return _src

    # 3. Assets folder — code without trailing 'pm' suffix
    #    Yardi property codes often end with 'pm' (property management) but
    #    legacy asset files may use the shorter name (revlabspm → revlabs_hero.jpg).
    if property_code.endswith('pm'):
        _base = property_code[:-2]
        for _ext in _exts:
            _src = _img_b64(f'{_base}_hero{_ext}')
            if _src:
                return _src

    return None

# ── Property discovery (used by hero banner + main-page selector + sidebar) ──
# Moved here, before the hero banner: this used to run down near the sidebar
# (after the hero banner), so the banner always rendered with whatever stale
# active_property_code was already in session_state — including right after
# a property is renamed/deleted, when the old code is no longer valid. The
# validation below (resetting to a real property if the current code isn't
# one) needs to happen before anything reads active_property_code, not after.
# A prior fix here tried reading the "🏢 Active Property" selectbox's own
# widget state instead, on the theory that Streamlit applies a widget's new
# value to session_state before the script reruns — true when the WIDGET
# itself changes, but active_property_code is also set directly in several
# other places (initial default, this same validation block, the
# single-property fallback) that never touch the widget's key at all, so it
# went stale independently and showed the WRONG property (confirmed with
# Ryan 2026-08-24 — Hartwell's banner appeared on Rev Labs after Hartwell
# was renamed). Validating active_property_code itself, this early, is the
# actual fix — every downstream reader (hero banner, selector, sidebar,
# tabs) then sees the same correct value.
_all_props   = _discover_properties()
_prop_codes  = [p['code'] for p in _all_props]
_prop_labels = {p['code']: f"{p['display_name']}  ({p['code']})" for p in _all_props}

if not _prop_codes:
    st.warning(
        "⚠️ **No properties configured.** "
        "Go to the **Properties** tab to add your first property.",
        icon="🏗️",
    )
    st.info(
        "Create a `data/{property_code}/config.yaml` file or use the Properties form "
        "to get started. See the How To Use tab for a full walkthrough.",
        icon="ℹ️",
    )
    st.stop()

if st.session_state.active_property_code not in _prop_codes:
    st.session_state.active_property_code = _prop_codes[0]

# Prefer the "🏢 Active Property" selectbox's own live widget value over
# active_property_code when the two disagree. Streamlit applies a widget's
# new value to its own session_state key BEFORE the script starts running,
# regardless of where the widget itself is instantiated (the selectbox
# lives further down, near the sidebar) -- so on the very rerun triggered by
# the user picking a new property, the widget's key already reflects that
# pick while active_property_code (updated by our own code, further down,
# AFTER this point) is still one step behind. Without this, that rerun
# renders the OLD property here, and only self-corrects via THAT code's own
# follow-up st.rerun() a moment later -- a second forced rerun that could
# race with how st.components.v1.html() updates the hero banner's iframe,
# leaving the wrong property's photo on screen even after everything
# settles. Confirmed with Ryan 2026-08-24 (both directions: Hartwell's photo
# stuck showing on Rev Labs, and vice versa). Falls back to
# active_property_code when the widget hasn't rendered yet this session, and
# is validated against _prop_codes either way since a widget value can also
# refer to a property that's since been renamed/deleted.
_selectbox_val = st.session_state.get('active_property_selectbox')
if _selectbox_val in _prop_codes:
    _selected_code = _selectbox_val
else:
    _selected_code = st.session_state.active_property_code
if _selected_code != st.session_state.active_property_code:
    st.session_state.active_property_code = _selected_code

_HERO_SRC = _prop_hero_src(_selected_code)

# ── Hero banner ───────────────────────────────────────────────
_hero_alt = _selected_code or 'Property'
_photo_html = (
    f'<img src="{_HERO_SRC}" class="grp-hero-photo" alt="{_hero_alt}"/>'
    if _HERO_SRC else ''
)
_logo_html = (
    f'<img src="{_LOGO_SRC}" style="max-width:140px;max-height:60px;" alt="GRP Logo"/>'
    if _LOGO_SRC else
    '<div class="grp-logo-text">Greatland<br>Realty<br>Partners</div>'
)

# Load the active property config for the hero banner
# (_active_cfg is set in the sidebar section below, but on first render we
# need it here too — load it again; it's cheap and cached by the YAML file.)
from property_config import load_property_config as _lpc_hero
_hero_cfg = _lpc_hero(_selected_code, str(_DATA_DIR))

_hero_title = f"{_hero_cfg.display()} Monthly Close"
_hero_sub   = ' &nbsp;|&nbsp; '.join(filter(None, [
    _hero_cfg.property_address,
    f"Managed by {_hero_cfg.management_company} for {_hero_cfg.investor_name}"
    if _hero_cfg.management_company and _hero_cfg.investor_name else
    (_hero_cfg.management_company or _hero_cfg.investor_name),
]))
_hero_badges = [f"🏢 {_hero_cfg.property_code}"]
if _hero_cfg.property_size_sf:
    _hero_badges.append(f"📐 ~{_hero_cfg.property_size_sf:,} SF")
if _hero_cfg.property_type:
    _hero_badges.append(f"🔬 {_hero_cfg.property_type}")
_hero_badge_html = " ".join(f'<span class="grp-badge">{b}</span>' for b in _hero_badges)

# Use st.components.v1.html() for the hero banner — it creates a true sandboxed
# iframe that always renders HTML correctly, unlike st.markdown(unsafe_allow_html)
# or st.html() which have both been broken in recent Streamlit versions.
# All hero CSS is inlined here so the component is self-contained.
import streamlit.components.v1 as _stc
_stc.html(f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  html, body {{
    margin: 0; padding: 0; overflow: hidden;
    font-family: 'Segoe UI', Arial, sans-serif;
    background: transparent;
  }}
  .grp-hero {{
    background: linear-gradient(135deg, #1A5C22 0%, #2E7D32 100%);
    border-radius: 10px;
    padding: 0;
    overflow: hidden;
    box-shadow: 0 4px 12px rgba(0,0,0,0.18);
    display: flex;
    align-items: stretch;
    min-height: 130px;
    margin: 0;
  }}
  .grp-hero-photo {{
    width: 240px; min-width: 240px;
    object-fit: cover;
    border-radius: 10px 0 0 10px;
    display: block;
  }}
  .grp-hero-body {{
    padding: 18px 24px;
    flex: 1;
    display: flex;
    flex-direction: column;
    justify-content: center;
  }}
  .grp-hero-title {{
    color: #ffffff;
    font-size: 1.45rem;
    font-weight: 700;
    letter-spacing: 0.02em;
    margin: 0 0 4px 0;
    line-height: 1.2;
  }}
  .grp-hero-sub {{
    color: #C8E6C9;
    font-size: 0.85rem;
    margin: 0 0 8px 0;
    font-weight: 400;
  }}
  .grp-hero-badges {{
    display: flex;
    gap: 7px;
    flex-wrap: wrap;
    margin-top: 3px;
  }}
  .grp-badge {{
    background: rgba(255,255,255,0.18);
    border: 1px solid rgba(255,255,255,0.35);
    border-radius: 20px;
    padding: 2px 11px;
    color: #ffffff;
    font-size: 0.73rem;
    font-weight: 500;
    white-space: nowrap;
  }}
  .grp-hero-logo {{
    padding: 18px 20px;
    display: flex;
    align-items: center;
    justify-content: flex-end;
    min-width: 150px;
  }}
  .grp-logo-text {{
    color: rgba(255,255,255,0.85);
    font-size: 0.68rem;
    text-align: right;
    line-height: 1.4;
    font-weight: 500;
    letter-spacing: 0.08em;
    text-transform: uppercase;
  }}
</style>
</head>
<body>
<div class="grp-hero">
  {_photo_html}
  <div class="grp-hero-body">
    <div class="grp-hero-title">{_hero_title}</div>
    <div class="grp-hero-sub">{_hero_sub}</div>
    <div class="grp-hero-badges">{_hero_badge_html}</div>
  </div>
  <div class="grp-hero-logo">{_logo_html}</div>
</div>
</body>
</html>
""", height=160, scrolling=False)


# ── Sidebar ──────────────────────────────────────────────────────────────────

# Detect property change — reset pipeline state so stale results don't carry over
if st.session_state.get('_prev_active_property_code') != _selected_code:
    st.session_state._prev_active_property_code = _selected_code
    st.session_state.pass1_complete        = False
    st.session_state.pass1_engine_result   = None
    st.session_state.pass1_output_files    = {}
    st.session_state['pass1_gl_activity_log'] = []
    st.session_state.interco_recode_df     = pd.DataFrame({
        "Leg": pd.Series([], dtype=str), "Account": pd.Series([], dtype=str),
        "Account Name": pd.Series([], dtype=str),
        "Credit ($)": pd.Series([], dtype=float), "Debit ($)": pd.Series([], dtype=float),
        "Description": pd.Series([], dtype=str),
    })
    st.session_state._interco_seed_gen     = st.session_state.get('_interco_seed_gen', 0) + 1
    st.session_state.pass2_complete        = False
    st.session_state.pass2_engine_result   = None
    st.session_state.pass2_output_files    = {}
    st.session_state.uploaded_files        = {}
    st.session_state.pass2_manual_prepaids = []   # B-F1: prevent cross-property bleed
    # Reset checklist so the new property's data is loaded
    st.session_state.checklist_loaded    = False
    st.session_state.close_tracker       = {}
    st.session_state.custom_checklist_items = []
    st.session_state.checklist_locked    = False
    st.session_state.checklist_locked_by = None
    st.session_state.checklist_locked_at = None
    st.session_state.last_completed_step = None
    # Clear preparer name so it re-seeds from new property's team
    st.session_state.prepared_by         = ''
    # Clear JE exclusions/overrides so property A's edits don't bleed to property B
    st.session_state.je_excluded_jes     = set()
    st.session_state.je_amount_overrides = {}
    # Clear file-type overrides so assignments from property A don't bleed to B
    st.session_state.bulk_overrides_p1   = {}
    st.session_state.bulk_overrides_p2   = {}
    st.session_state.bulk_overrides_wp   = {}
    # Clear JE description overrides keyed to prior property's run
    st.session_state.je_desc_overrides   = {}
    st.session_state.pop('_je_desc_run', None)
    # Clear post-close JE table and sign-off state so Property A data doesn't bleed into B
    st.session_state.signoff_state        = {}
    st.session_state.post_close_je_df    = pd.DataFrame({
        "JE #": ["PC-001", "PC-001"], "Description": ["", ""],
        "Account Code": ["", ""],
        "Debit ($)": [0.0, 0.0], "Credit ($)": [0.0, 0.0],
        "Line Description": ["", ""],
    })
    st.session_state.pop('_pcje_latest', None)
    # Increment upload widget keys so Streamlit discards Property A's file buffers
    st.session_state.upload_key_p1 = st.session_state.get('upload_key_p1', 0) + 1
    st.session_state.upload_key_p2 = st.session_state.get('upload_key_p2', 0) + 1
    st.session_state.tub_key       = st.session_state.get('tub_key', 0) + 1   # B-13: re-render TUB at $0 on property switch
    # Clear the prior period label so workpaper carry-forward uses the correct month
    st.session_state.pop('prior_period_label_input', None)

# Load config for the selected property
from property_config import load_property_config as _load_prop_cfg
_active_cfg = _load_prop_cfg(_selected_code, str(_DATA_DIR))

# ── Output file prefix helpers (used throughout Pass 1 & Pass 2 for filenames) ─
# _pfx_del  → deliverable prefix for external/ZIP filenames  e.g. 'RevLabs'
# _pfx_int  → internal prefix for individual file downloads  e.g. 'GA'
# _inv_pfx  → invoice prefix                                 e.g. 'RevLabsPM'
# _prop_display → best display name for pipeline fallbacks   e.g. 'Revolution Labs'
_pfx_del      = _active_cfg.deliverable_prefix()            # auto-derived if not set
_pfx_int      = _active_cfg.file_prefix_internal or 'GA'
_inv_pfx      = _active_cfg.invoice_prefix or _pfx_del
_prop_display = _active_cfg.display()
_prop_entity  = _active_cfg.property_name or _prop_display  # full legal entity name

# Resolve prepared_by default on first load or when it's still the empty placeholder
if not st.session_state.get('prepared_by'):
    _team = getattr(_active_cfg, 'team_members', None) or []
    st.session_state.prepared_by = _team[0] if _team else ''

# ── Rebuild accruals seed when active property changes ────────────────────────
# Now that _active_cfg is loaded, rebuild the one-off accruals table if needed.
if st.session_state.get('_accruals_df_for_property') != _selected_code:
    st.session_state.manual_accruals_df = _build_accruals_seed_df(_active_cfg)
    st.session_state._accruals_df_for_property = _selected_code
    st.session_state._accruals_seed_gen = st.session_state.get('_accruals_seed_gen', 0) + 1

# Committed Kardin budget for the active property
_COMMITTED_BUDGET = _committed_path(
    _active_cfg.property_code,
    _active_cfg.kardin_budget_file or 'GA_Kardin_Budget_FY2026.xlsx',
)

# Sidebar property card
_sb_logo = (
    f'<img src="{_LOGO_SRC}" style="max-width:120px;max-height:44px;margin-bottom:8px;display:block;" alt="GRP"/>'
    if _LOGO_SRC else ''
)
st.session_state.prepared_by = st.sidebar.text_input(
    "Prepared by",
    value=st.session_state.prepared_by,
    help="Stamped on every workpaper tab and the run log.",
)

# ── Report an Issue ───────────────────────────────────────────
with st.sidebar.expander("🐛 Report an Issue", expanded=False):
    _fb_reporter  = st.session_state.get('prepared_by', '') or ''
    _fb_prop      = st.session_state.get('active_property_code', '') or ''
    _fb_period    = ''
    try:
        from checklist_persistence import period_key_to_label as _ptl
        _fb_period = _ptl(st.session_state.get('checklist_period_key', ''))
    except Exception:
        pass

    _fb_severity = st.selectbox(
        "Severity",
        ["low", "medium", "high", "critical"],
        index=1,
        key="fb_severity",
        format_func=lambda s: {
            "low": "🟢 Low — cosmetic / minor",
            "medium": "🟡 Medium — something's off",
            "high": "🟠 High — blocking a task",
            "critical": "🔴 Critical — wrong numbers",
        }[s],
    )
    _fb_desc = st.text_area(
        "Describe the issue",
        placeholder="What happened? What did you expect? Which tab/section?",
        key="fb_description",
        height=100,
    )
    if st.button("Submit Issue", use_container_width=True, key="fb_submit"):
        if _fb_desc.strip():
            try:
                import json as _fb_json
                from pathlib import Path as _FbPath
                _fb_log = _FbPath(__file__).parent / 'data' / 'feedback_log.jsonl'
                _fb_log.parent.mkdir(parents=True, exist_ok=True)
                _fb_entry = {
                    'submitted_at':  datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'reporter':      _fb_reporter or 'Unknown',
                    'property_code': _fb_prop,
                    'period':        _fb_period,
                    'severity':      _fb_severity,
                    'description':   _fb_desc.strip(),
                    'status':        'open',
                }
                with open(_fb_log, 'a', encoding='utf-8') as _fb_fh:
                    _fb_fh.write(_fb_json.dumps(_fb_entry) + '\n')
                st.success("Issue logged — Ryan will review on the next health check.")
            except Exception as _fb_err:
                st.error(f"Could not save issue: {_fb_err}")
        else:
            st.warning("Please describe the issue before submitting.")

# ── Session Save / Load ───────────────────────────────────────
with st.sidebar.expander("💾 Save / Load Session", expanded=False):
    st.caption(
        "Save your table inputs (One-Off Accruals, manual JEs) to a file "
        "and reload them in a future session without re-entering data. "
        "File uploads must be re-uploaded separately."
    )
    from session_snapshot import save_snapshot, load_snapshot, restore_snapshot, snapshot_filename as _snap_filename
    _snap_bytes = save_snapshot(st.session_state)
    st.download_button(
        label="⬇️ Save Session Data",
        data=_snap_bytes,
        file_name=_snap_filename(st.session_state),
        mime="application/json",
        use_container_width=True,
        help="Downloads a JSON file with your table inputs and settings.",
    )
    _snap_upload = st.file_uploader(
        "Load session file",
        type=["json"],
        key="session_snapshot_upload",
        label_visibility="collapsed",
    )
    if _snap_upload is not None:
        _snap_data = load_snapshot(_snap_upload.read())
        if _snap_data:
            _restored = restore_snapshot(_snap_data, st.session_state)
            if _restored:
                st.success(f"Restored: {', '.join(_restored)}")
                st.rerun()
            else:
                st.warning("File loaded but nothing to restore — may be an older format.")
        else:
            st.error("Could not parse snapshot file.")

if not st.session_state.confirm_reset_all:
    if st.sidebar.button("🔄 Reset All", use_container_width=True,
                         help="Clear all results and uploaded files"):
        st.session_state.confirm_reset_all = True
        st.rerun()
else:
    st.sidebar.warning("⚠️ This will wipe **all** uploads, results, and sign-offs.")
    _ra_col1, _ra_col2 = st.sidebar.columns(2)
    if _ra_col1.button("✅ Confirm", use_container_width=True, key="confirm_reset_all_btn"):
        st.session_state.confirm_reset_all = False
        st.session_state.pass1_complete = False
        st.session_state.pass1_engine_result = None
        st.session_state.pass1_output_files = {}
        st.session_state['pass1_gl_activity_log'] = []
        st.session_state.pass2_complete = False
        st.session_state.pass2_engine_result = None
        st.session_state.pass2_output_files = {}
        st.session_state.uploaded_files = {}
        st.session_state.bulk_overrides_p1 = {}
        st.session_state.bulk_overrides_p2 = {}
        st.session_state.bulk_overrides_wp = {}
        st.session_state.pass2_manual_prepaids = []   # B-F1: prevent cross-property bleed
        st.session_state.signoff_state = {}
        st.session_state.close_tracker = {}
        st.session_state.upload_key_p1 += 1
        st.session_state.upload_key_p2 += 1
        st.session_state.editor_reset_count = st.session_state.get('editor_reset_count', 0) + 1
        st.session_state.je_excluded_jes     = set()   # clear JE exclusions on full reset
        st.session_state.je_amount_overrides = {}       # clear amount overrides on full reset
        shutil.rmtree(st.session_state.temp_dir, ignore_errors=True)
        st.session_state.temp_dir = tempfile.mkdtemp(prefix="ga_automation_")
        import pandas as _pd
        st.session_state.post_close_je_df = _pd.DataFrame({
            "JE #": ["PC-001", "PC-001"], "Description": ["", ""],
            "Account Code": ["", ""],
            "Debit ($)": [0.0, 0.0], "Credit ($)": [0.0, 0.0],
            "Line Description": ["", ""],
        })
        st.session_state.pop("_pcje_latest", None)
        # Fully rebuild accruals seed — zeroing only the Amount column would
        # leave stale account codes/vendors that suppress automated accruals.
        st.session_state.manual_accruals_df = _build_accruals_seed_df(_active_cfg)
        st.session_state._accruals_df_for_property = _selected_code
        st.session_state._accruals_seed_gen = st.session_state.get('_accruals_seed_gen', 0) + 1
        st.session_state.tub_key += 1   # forces TUB number inputs to re-render at $0
        st.session_state.custom_checklist_items = []
        # Set checklist_loaded = True so the load block does NOT immediately
        # reload from the persistence file — keeps close_tracker empty after reset.
        st.session_state.checklist_loaded    = True
        st.session_state.checklist_locked    = False
        st.session_state.checklist_locked_by = None
        st.session_state.checklist_locked_at = None
        st.session_state.last_completed_step = None
        st.session_state.pass1_run_count     = 0
        st.session_state.bulk_overrides_wp   = {}
        st.session_state['pass1_gl_activity_log'] = []
        st.session_state.je_desc_overrides   = {}
        st.session_state.interco_recode_df   = _pd.DataFrame({
            "Leg": _pd.Series([], dtype=str), "Account": _pd.Series([], dtype=str),
            "Account Name": _pd.Series([], dtype=str),
            "Credit ($)": _pd.Series([], dtype=float), "Debit ($)": _pd.Series([], dtype=float),
            "Description": _pd.Series([], dtype=str),
        })
        st.session_state._interco_seed_gen   = st.session_state.get('_interco_seed_gen', 0) + 1
        # Clear keys missed by prior Reset All logic
        st.session_state.pop('_je_desc_run', None)
        st.session_state.pop('prior_period_label_input', None)
        for _k in [k for k in st.session_state.keys() if k.startswith('je_add_count_')]:
            del st.session_state[_k]
        st.rerun()
    if _ra_col2.button("❌ Cancel", use_container_width=True, key="cancel_reset_all_btn"):
        st.session_state.confirm_reset_all = False
        st.rerun()


FILE_CONFIG = {
    # ── Core ──────────────────────────────────────────────────
    "gl": (
        "Yardi GL Detail (.xlsx)", "xlsx", True, "core",
        "REQUIRED — source of truth for all accounts and transactions.",
    ),
    "trial_balance": (
        "Yardi Trial Balance (.xlsx)", "xlsx", False, "core",
        "Enables GL↔TB tie-out validation on every BS account and QC Check 5. "
        "Without it: BS workpaper generates but all TB columns show 'N/A'.",
    ),
    "budget_comparison": (
        "Yardi Budget Comparison (.xlsx)", "xlsx", False, "core",
        "Enables historical pattern accruals (Layer 3) and variance commentary. "
        "Without it: only Nexus invoice and invoice-proration accruals generated; no variance comments.",
    ),
    "kardin_budget": (
        "Kardin 2026 Budget (.xlsx)", "xlsx", False, "core",
        "Enables QC YTD budget vs Kardin cross-check AND payroll bonus accruals (Layer 4 — "
        "used as fallback when no annual bonus is entered in the Bonus Accrual expander). "
        "Without it: QC budget check and Kardin-derived bonus accruals skipped.",
    ),
    "t12_statement": (
        "12-Month Income Statement (.xlsx)", "xlsx", False, "core",
        "Powers MoM Swings tab (Tab 4) in Pass 2 QC workbook with real prior-month actuals "
        "instead of derived YTD-PTD. Critical for January (prior month = December actuals). "
        "Also improves Layer 3 historical accrual detection in Pass 1.",
    ),
    "nexus_accrual": (
        "Nexus Invoice Detail (.xls / .xlsx)", ["xls", "xlsx"], False, "core",
        "Enables AP accrual detection (Layer 1 — open invoices not yet posted to GL). "
        "Without it: invoice-proration (Layer 2) and historical (Layer 3) still run.",
    ),
    # ── Bank ──────────────────────────────────────────────────
    "bank_rec": (
        "Yardi Bank Rec PDF — Operating (.pdf)", "pdf", False, "bank",
        "PREFERRED bank source. Reads Yardi's pre-computed reconciliation: bank balance, "
        "outstanding checks, reconciled balance, and $0 difference. Enables Operating bank "
        "rec tab in the BS workpaper (PNC x3993 vs GL 111100). Without it: no bank rec tab.",
    ),
    "receivable_summary": (
        "Yardi Receivable Summary (.xlsx)", "xlsx", False, "bank",
        "PRIMARY management fee basis — explicit Prepayment row gives the cleanest prepayment exclusion. "
        "Export from Yardi after bank rec is complete. Takes priority over Receivable Detail when both are uploaded. "
        "Without it: falls back to Receivable Detail (if uploaded) or DACA additions.",
    ),
    "receivable_detail": (
        "Yardi Receivable Detail (.xlsx)", "xlsx", False, "bank",
        "ALTERNATE management fee basis — JLL's exact method. Export from Yardi after bank rec is complete. "
        "Pair with the AR Detail Aging for accurate prepayment exclusion. "
        "Without it: falls back to DACA additions. (Receivable Summary is preferred when available.)",
    ),
    "ar_aging": (
        "Yardi AR Detail Aging (.xlsx)", "xlsx", False, "bank",
        "Prepayment identification for the management fee — the Pre-payments column shows unapplied "
        "tenant credits excluded from the cash-received basis. Upload alongside the Receivable Detail. "
        "Without it: falls back to charge-code scan in the Receivable Detail (less reliable).",
    ),
    "bank_rec_dev": (
        "Development Bank Statement — Bank of America (.pdf)", "pdf", False, "bank",
        "Adds a 'Bank Rec - Development' tab to the Pass 2 workpaper for the revlabs entity. "
        "Upload the BofA Full Analysis Business Checking statement (account x3132). "
        "Without it: development bank rec tab is omitted from the workpaper.",
    ),
    "bank_rec_dev_xlsx": (
        "Yardi Development Bank Rec — 111210 (.xlsx)", "xlsx", False, "bank",
        "Yardi Bank Reconciliation Report for the BofA development account (x3132). "
        "Copies the raw Yardi export directly into the '111210 Cash - Development' tab "
        "in the workpaper — matching the exact format of the 111100 PNC and 115100 DACA tabs. "
        "Without it: tab is generated from the BofA PDF (ending balance only, no detail).",
    ),
    "bank_rec_xlsx": (
        "Yardi Operating Bank Rec — 111100 (.xlsx, optional)", "xlsx", False, "bank",
        "OPTIONAL Excel export for the PNC Operating account — copies the raw Yardi "
        "sheet directly into the '111100 PNC Cash' tab. Not required: if you only have "
        "the 'Yardi Bank Rec PDF — Operating' uploaded above ('bank_rec'), the tab is "
        "filled from that PDF's own GL-detail pages instead — real transaction detail, "
        "just not a byte-for-byte copy of the raw sheet. Without either: placeholder.",
    ),
    "daca_bank_rec_xlsx": (
        "Yardi DACA Bank Rec — 115100 (.xlsx, optional)", "xlsx", False, "bank",
        "OPTIONAL Excel export for the KeyBank DACA account — copies the raw Yardi "
        "sheet directly into the '115100 DACA' tab. Not required: if you only have "
        "the 'KeyBank DACA Statement' uploaded above ('daca_bank'), the tab is filled "
        "from that PDF's cleared deposits/other items instead. Without either: placeholder.",
    ),
    "capital_schedule": (
        "Capital Accounts Schedule (.xlsx)", "xlsx", False, "reference",
        "Capital improvement schedules (154500 Building Improvements, 181200 LC, "
        "181300 Legal, 181400 TI). Upload the monthly Rev Labs capital schedule. "
        "Drives the 4 capital account tabs in the workpaper. "
        "Without it: capital tabs show GL transactions only.",
    ),
    "capital_seed": (
        "Capital Schedule Seed (.xlsx)", "xlsx", False, "reference",
        "January 2026 seed file (Book3.xlsx) — all 7 capital accounts "
        "(152100 Land, 154100 Building, 154500 Bldg Improvements, 171100 CIP, "
        "181200 LC, 181300 Legal, 181400 TI). "
        "Used only when no Capital Accounts Schedule is uploaded. "
        "From February onward the prior workpaper carry-forward supersedes this.",
    ),
    "daca_bank": (
        "DACA Bank Statement — KeyBank x5132 (.pdf)", "pdf", False, "bank",
        "Enables DACA bank rec tab in the BS workpaper (KeyBank x5132 vs GL 115100). "
        "Without it: DACA rec tab is omitted from the workpaper.",
    ),
    # ── Reference ─────────────────────────────────────────────
    "loan": (
        "Berkadia Loan Statements (.pdf)", "pdf", False, "ref",
        "Enables debt service workpaper tab and principal balance tracking. "
        "Without it: debt service section not generated.",
    ),
    "prepaid_ledger": (
        "Prepaid Ledger — prior month (.xlsx)", "xlsx", False, "ref",
        "Carry-forward from prior month for prepaid amortization tracking. "
        "Without it: ledger starts fresh — existing multi-period items won't be carried forward.",
    ),
    "tenancy_schedule": (
        "Tenancy Schedule / Rent Roll (.xlsx)", "xlsx", False, "ref",
        "Current tenant list read live from this period's rent roll — drives the Tenant "
        "Utility Billing table with Yardi's own tenant codes as the identifier, instead of "
        "a static config list that goes stale as leases turn over. "
        "Without it: falls back to this property's config.yaml tenants list, if any.",
    ),
}

file_config = FILE_CONFIG

from file_classifier import classify_file as _classify_file, FILE_LABELS as _FILE_LABELS, MULTI_FILE_KEYS as _MULTI_FILE_KEYS

# ── Session state: persist override selections across reruns ─────────────────
if "bulk_overrides_p1" not in st.session_state:
    st.session_state.bulk_overrides_p1 = {}

# Options shown in the override dropdown for Pass 1
_P1_SLOT_KEYS = [
    "gl", "trial_balance", "budget_comparison", "kardin_budget", "t12_statement",
    "nexus_accrual", "bank_rec", "receivable_summary", "receivable_detail", "ar_aging",
    "bank_rec_dev", "bank_rec_xlsx", "bank_rec_dev_xlsx", "daca_bank_rec_xlsx",
    "capital_schedule", "capital_seed", "daca_bank", "loan",
    "prepaid_ledger", "tenancy_schedule", "unknown",
]
_P1_SLOT_LABELS = [_FILE_LABELS.get(k, k) for k in _P1_SLOT_KEYS]


# ═══════════════════════════════════════════════════════════════
# ── Property selector (main page, above tabs) ─────────────────
# ═══════════════════════════════════════════════════════════════
import pandas as pd

if len(_all_props) > 1:
    _sel_col, _sel_spacer = st.columns([2, 5])
    with _sel_col:
        _selected_code = st.selectbox(
            "🏢 Active Property",
            options=_prop_codes,
            index=_prop_codes.index(st.session_state.active_property_code)
                  if st.session_state.active_property_code in _prop_codes else 0,
            format_func=lambda c: _prop_labels.get(c, c),
            key="active_property_selectbox",
            help="Switching properties resets all pipeline state.",
            label_visibility="visible",
        )
        # Keep session_state in sync (selectbox key ≠ state key here to avoid conflict)
        if _selected_code != st.session_state.active_property_code:
            st.session_state.active_property_code = _selected_code
            st.rerun()
    with _sel_spacer:
        st.markdown(
            f"<div style='padding-top:28px;font-size:0.82rem;color:#616161;'>"
            f"{_prop_labels.get(_selected_code, _selected_code)}&nbsp;&nbsp;|&nbsp;&nbsp;"
            f"{next((p['address'] for p in _all_props if p['code'] == _selected_code), '')}"
            f"</div>",
            unsafe_allow_html=True,
        )
else:
    # Single property — no selector needed, but update state to match
    _selected_code = _prop_codes[0] if _prop_codes else ''
    st.session_state.active_property_code = _selected_code

# ═══════════════════════════════════════════════════════════════
# ── Main content: Two-pass tabs ──────────────────────────────
# ═══════════════════════════════════════════════════════════════

tab3, tab0, tab1, tab2, tab4 = st.tabs([
    "📖  How to Use",
    "🏠  Dashboard",
    "📋  Pass 1 — Generate JEs",
    "📊  Pass 2 — Generate Reports & JEs",
    "⚙️  Properties",
])


# ══════════════════════════════════════════════════════════════
# TAB 0 — DASHBOARD / CLOSE CHECKLIST
# ══════════════════════════════════════════════════════════════
with tab0:

    # ── Load checklist from GitHub once per session / property ────────────────
    _ck_prop   = st.session_state.get('active_property_code', '')
    _ck_pkey   = st.session_state.checklist_period_key
    if not st.session_state.get('checklist_loaded', False):
        try:
            _ck_state = load_checklist(_ck_prop, _ck_pkey, str(_DATA_DIR))
            _ct_loaded, _ci_loaded, _lk, _lk_by, _lk_at = state_to_session(_ck_state)
            for _sk, _sv in _ct_loaded.items():
                if _sk not in st.session_state.close_tracker:
                    st.session_state.close_tracker[_sk] = _sv
            if not st.session_state.custom_checklist_items:
                st.session_state.custom_checklist_items = _ci_loaded
            st.session_state.checklist_locked    = _lk
            st.session_state.checklist_locked_by = _lk_by
            st.session_state.checklist_locked_at = _lk_at
        except Exception:
            pass
        st.session_state.checklist_loaded = True

    from close_tracker_generator import CLOSE_TRACKER_STEPS as _CTS

    # ── Portfolio Overview ─────────────────────────────────────────────────────
    _all_props = _discover_properties()
    if len(_all_props) > 1:
        with st.expander(f"🏢 Portfolio Overview — {len(_all_props)} Properties", expanded=False):
            _po_rows = []
            for _pp in _all_props:
                _pp_code  = _pp['code']
                _pp_name  = _pp.get('display_name', _pp_code)
                _pp_addr  = _pp.get('address', '')
                # Load persisted checklist for this property / current period
                try:
                    _pp_state = load_checklist(_pp_code, _ck_pkey, str(_DATA_DIR))
                    _pp_ct, _pp_ci, _pp_lk, _, _ = state_to_session(_pp_state)
                    _pp_steps = len(_CTS)
                    _pp_done  = sum(1 for i in range(_pp_steps) if i in _pp_ct)
                    _pp_cust  = len(_pp_ci)
                    _pp_cdone = sum(1 for c in _pp_ci if c.get('completed'))
                    _pp_total = _pp_steps + _pp_cust
                    _pp_all   = _pp_done + _pp_cdone
                    _pp_pct   = int(100 * _pp_all / _pp_total) if _pp_total else 0
                    _pp_status = (
                        "🔒 Locked" if _pp_lk else
                        ("✅ Complete" if _pp_pct == 100 else
                         (f"🔄 {_pp_pct}% ({_pp_all}/{_pp_total})" if _pp_all > 0 else "⬜ Not started"))
                    )
                except Exception:
                    _pp_status = "—"
                    _pp_pct    = 0
                _po_rows.append({
                    "Property":   _pp_name,
                    "Address":    _pp_addr,
                    "Period":     period_key_to_label(_ck_pkey),
                    "Status":     _pp_status,
                    "Progress":   f"{_pp_pct}%",
                    "Active":     "⬤" if _pp_code == _ck_prop else "",
                })
            import pandas as _pd_po
            st.dataframe(
                _pd_po.DataFrame(_po_rows),
                use_container_width=True,
                hide_index=True,
            )
            st.caption(
                "Progress reflects persisted checklist data for the selected close period. "
                "Switch properties using the selector in the sidebar."
            )
    # ── Cross-Period Trending ──────────────────────────────────────────────────
    try:
        from period_metrics import load_metrics as _load_metrics
        _trend_all = _load_metrics(str(_DATA_DIR), _ck_prop)
    except Exception:
        _trend_all = []

    if _trend_all:
        # Filter to the current fiscal year only.
        # Fiscal year: starts on fiscal_year_start_month, ends the month before.
        # E.g. FY-start=1 → Jan–Dec of the same calendar year.
        #      FY-start=7 → Jul year N to Jun year N+1.
        from datetime import date as _dt_date
        _fy_sm   = int(getattr(_active_cfg, 'fiscal_year_start_month', 1) or 1)
        _today   = _dt_date.today()
        _cur_mo  = _today.month
        _cur_yr  = _today.year

        # Which fiscal year are we currently in?
        if _fy_sm == 1:
            _fy_start_yr = _cur_yr
        elif _cur_mo >= _fy_sm:
            _fy_start_yr = _cur_yr       # FY started earlier this calendar year
        else:
            _fy_start_yr = _cur_yr - 1  # FY started last calendar year

        # Build the set of (year, month) pairs that belong to this FY
        _fy_months: set = set()
        for _fi in range(12):
            _fm = (_fy_sm - 1 + _fi) % 12 + 1
            _fy = _fy_start_yr if _fm >= _fy_sm else _fy_start_yr + 1
            _fy_months.add((_fy, _fm))

        # Filter records to current FY
        import re as _re_trend
        _mo_map_t = dict(Jan=1,Feb=2,Mar=3,Apr=4,May=5,Jun=6,
                         Jul=7,Aug=8,Sep=9,Oct=10,Nov=11,Dec=12)
        def _in_fy(record: dict) -> bool:
            m = _re_trend.search(
                r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[- ](\d{4})',
                record.get('period', '')
            )
            if not m:
                return False
            return (int(m.group(2)), _mo_map_t.get(m.group(1), 0)) in _fy_months

        _trend_data = [r for r in _trend_all if _in_fy(r)]

        # Build fiscal year label for the expander title
        _fy_end_yr = _fy_start_yr if _fy_sm == 1 else (
            _fy_start_yr + 1 if _fy_sm > 1 else _fy_start_yr
        )
        if _fy_sm == 1:
            _fy_label = f"FY {_fy_start_yr}"
        else:
            _mo_names_t = {v: k for k, v in _mo_map_t.items()}
            _fy_end_mo  = (_fy_sm - 2) % 12 + 1
            _fy_label   = (f"FY {_mo_names_t[_fy_sm]} {_fy_start_yr} – "
                           f"{_mo_names_t[_fy_end_mo]} {_fy_end_yr}")

    if _trend_all and _trend_data:
        # Filter out records where GL extraction failed (all-zero revenue/expenses/noi).
        # These were saved before the period_metrics GL-field fix; fee_amount is
        # populated but the P&L data is absent — hiding them prevents a misleading
        # management-fee bar floating next to an empty NOI chart.
        _trend_data_valid  = [
            r for r in _trend_data
            if not (r.get('revenue', 0) == 0
                    and r.get('expenses', 0) == 0
                    and r.get('noi', 0) == 0)
        ]
        _trend_filtered_ct = len(_trend_data) - len(_trend_data_valid)

        with st.expander(
            f"📈 {_fy_label} Trending — {len(_trend_data_valid)} / 12 period(s)",
            expanded=False,
        ):
            import pandas as _pd_trend

            # Clear-metrics button — wipes metrics.jsonl for the active property
            _clr_met_col, _ = st.columns([1, 3])
            if _clr_met_col.button(
                "🗑️ Clear Trending Data", key="clear_trending_btn",
                help="Remove all saved metrics for this property. "
                     "Data re-populates automatically after each Pass 2 run.",
            ):
                try:
                    from pathlib import Path as _PPath
                    _met_path = (_PPath(str(_DATA_DIR))
                                 / (_active_cfg.property_code or '')
                                 / 'metrics.jsonl')
                    if _met_path.exists():
                        _met_path.unlink()
                    st.success("Trending data cleared. Re-run Pass 2 to rebuild.", icon="🗑️")
                    st.rerun()
                except Exception as _met_err:
                    st.error(f"Could not clear metrics: {_met_err}")

            if _trend_filtered_ct:
                st.caption(
                    f"ℹ️ {_trend_filtered_ct} period(s) hidden — incomplete GL data "
                    f"(saved before a metrics fix). Re-run Pass 2 for those periods to restore them."
                )

            if not _trend_data_valid:
                st.info("No complete period data yet. Run Pass 2 to populate trending.", icon="📊")
            else:
                _periods = [r.get('period', '') for r in _trend_data_valid]

                # ── NOI chart ────────────────────────────────────────────
                st.markdown("##### Net Operating Income")
                _noi_df = _pd_trend.DataFrame({
                    'Period':   _periods,
                    'Revenue':  [r.get('revenue', 0) for r in _trend_data_valid],
                    'Expenses': [r.get('expenses', 0) for r in _trend_data_valid],
                    'NOI':      [r.get('noi', 0) for r in _trend_data_valid],
                }).set_index('Period')
                st.line_chart(_noi_df[['NOI', 'Revenue', 'Expenses']], height=220)

                # ── Management fee + cash row ─────────────────────────────
                _tr_col1, _tr_col2 = st.columns(2)
                with _tr_col1:
                    st.markdown("##### Management Fee")
                    _fee_df = _pd_trend.DataFrame({
                        'Period': _periods,
                        'Fee':    [r.get('fee_amount', 0) for r in _trend_data_valid],
                    }).set_index('Period')
                    st.bar_chart(_fee_df, height=180)

                with _tr_col2:
                    st.markdown("##### Operating Cash Balance")
                    _cash_df = _pd_trend.DataFrame({
                        'Period':    _periods,
                        'Operating': [r.get('operating_cash', 0) for r in _trend_data_valid],
                        'DACA':      [r.get('daca_balance', 0) for r in _trend_data_valid],
                    }).set_index('Period')
                    st.line_chart(_cash_df, height=180)

                # ── QC status summary ─────────────────────────────────────
                st.markdown("##### QC Status by Period")
                _qc_rows = []
                for r in _trend_data_valid:
                    _qc_status = r.get('qc_overall', 'unknown')
                    _qc_icon   = {'pass': '✅', 'warn': '⚠️', 'fail': '❌'}.get(_qc_status, '—')
                    _qc_rows.append({
                        'Period': r.get('period', ''),
                        'Status': f"{_qc_icon} {_qc_status.title()}",
                        'Pass':   r.get('qc_pass', 0),
                        'Warn':   r.get('qc_warn', 0),
                        'Fail':   r.get('qc_fail', 0),
                        'NOI':    f"${r.get('noi', 0):,.0f}",
                        'Fee':    f"${r.get('fee_amount', 0):,.0f}",
                    })
                st.dataframe(
                    _pd_trend.DataFrame(_qc_rows),
                    use_container_width=True,
                    hide_index=True,
                )
                st.caption(
                    f"Showing {_fy_label} only. Metrics saved automatically after each "
                    f"successful Pass 2 run. Resets to a fresh chart when the new fiscal year begins."
                )

    st.divider()

    # ── Top control bar ───────────────────────────────────────────────────────
    _ck_col_name, _ck_col_period, _ck_col_prog = st.columns([2, 2, 4])

    with _ck_col_name:
        _team_names = (_active_cfg.team_members
                       if _active_cfg.team_members
                       else ['[Property Accountant]', '[Property Manager]', '[Accounting Manager/Controller]'])
        _cur_name   = st.session_state.get('prepared_by', _team_names[0])
        _name_idx   = _team_names.index(_cur_name) if _cur_name in _team_names else 0
        _chosen_name = st.selectbox(
            "👤 I am",
            _team_names,
            index=_name_idx,
            key='dashboard_user_name',
        )
        if _chosen_name != st.session_state.prepared_by:
            st.session_state.prepared_by = _chosen_name

    with _ck_col_period:
        # Build a dynamic list of periods: all months for the past year and next year
        from datetime import date
        _today = date.today()
        _fiscal_start = getattr(_active_cfg, 'fiscal_year_start_month', 1) or 1
        _cur_year = _today.year
        # Show all months for the current year plus one year back and one year forward
        _period_options = sorted(set(
            [f'{_cur_year - 1}_{m:02d}' for m in range(1, 13)] +
            [f'{_cur_year}_{m:02d}'     for m in range(1, 13)] +
            [f'{_cur_year + 1}_{m:02d}' for m in range(1, 13)]
        ))
        if _ck_pkey not in _period_options:
            _period_options.append(_ck_pkey)
            _period_options.sort()
        _period_labels_map = {k: period_key_to_label(k) for k in _period_options}
        _period_sel_idx = _period_options.index(_ck_pkey) if _ck_pkey in _period_options else 0
        _period_chosen = st.selectbox(
            "📅 Close Period",
            _period_options,
            index=_period_sel_idx,
            format_func=lambda k: _period_labels_map[k],
            key='dashboard_period_sel',
        )
        if _period_chosen != st.session_state.checklist_period_key:
            st.session_state.checklist_period_key = _period_chosen
            st.session_state.checklist_loaded    = False
            # Full reset to match the property-switch reset pattern above —
            # without clearing close_tracker/custom_checklist_items here, the
            # loader block below only ADDS keys missing from close_tracker
            # (never removes any), so a step completed in the PREVIOUS period
            # stayed marked done after switching to a period where it isn't —
            # and last_completed_step (not period-scoped at all) could still
            # point at a step from the period just left, showing its "Step N
            # Complete" banner for the wrong period entirely.
            st.session_state.close_tracker       = {}
            st.session_state.custom_checklist_items = []
            st.session_state.checklist_locked    = False
            st.session_state.checklist_locked_by = None
            st.session_state.checklist_locked_at = None
            st.session_state.last_completed_step = None
            st.rerun()

    with _ck_col_prog:
        _n_steps   = len(_CTS)
        _n_done    = sum(1 for i in range(_n_steps) if i in st.session_state.close_tracker)
        _n_custom  = len(st.session_state.custom_checklist_items)
        _n_cdone   = sum(1 for c in st.session_state.custom_checklist_items if c.get('completed'))
        _total_all = _n_steps + _n_custom
        _done_all  = _n_done + _n_cdone
        _pct = int(100 * _done_all / _total_all) if _total_all else 0
        _ck_locked = st.session_state.get('checklist_locked', False)

        # Lock badge or progress
        if _ck_locked:
            st.markdown(
                f"<div style='background:#E8F5E9;border:1px solid #2E7D32;border-radius:6px;"
                f"padding:8px 14px;font-size:0.85rem;color:#1B5E20;font-weight:600;'>"
                f"🔒 Period Locked &nbsp;·&nbsp; {period_key_to_label(_ck_pkey)} &nbsp;·&nbsp; "
                f"Locked by {st.session_state.get('checklist_locked_by','—')} "
                f"at {st.session_state.get('checklist_locked_at','—')}"
                f"</div>",
                unsafe_allow_html=True,
            )
        else:
            st.markdown(f"**{period_key_to_label(_ck_pkey)} Close**")
            st.progress(_pct)
            _pct_color = '#2E7D32' if _pct == 100 else '#1565C0'
            st.markdown(
                f"<span style='font-size:0.85rem;color:{_pct_color};font-weight:600;'>"
                f"{_done_all} / {_total_all} tasks complete ({_pct}%)</span>",
                unsafe_allow_html=True,
            )

    # ── Lock / Unlock controls ─────────────────────────────────────────────────
    _ck_locked = st.session_state.get('checklist_locked', False)
    _lock_col, _lock_spacer = st.columns([3, 5])
    with _lock_col:
        if not _ck_locked:
            if _pct == 100:
                # All tasks done — offer lock
                if st.button("🔒 Lock Period", use_container_width=True,
                             help="Locks this period's checklist. All steps are complete. "
                                  "Use Unlock for corrections."):
                    st.session_state.checklist_locked    = True
                    st.session_state.checklist_locked_by = st.session_state.prepared_by
                    st.session_state.checklist_locked_at = datetime.now().strftime('%m/%d/%Y %H:%M')
                    _save_checklist_now()
                    st.rerun()
            else:
                # Not all done — allow locking but warn
                if st.button("🔒 Lock Period Anyway", use_container_width=True,
                             help=f"Lock even though {_total_all - _done_all} task(s) remain. "
                                  "Use this if the close is complete but some steps were skipped."):
                    st.session_state.checklist_locked    = True
                    st.session_state.checklist_locked_by = st.session_state.prepared_by
                    st.session_state.checklist_locked_at = datetime.now().strftime('%m/%d/%Y %H:%M')
                    _save_checklist_now()
                    st.rerun()
        else:
            if st.button("🔓 Unlock Period", use_container_width=True,
                         help="Unlock to make corrections."):
                st.session_state.checklist_locked    = False
                st.session_state.checklist_locked_by = None
                st.session_state.checklist_locked_at = None
                _save_checklist_now()
                st.rerun()

    st.divider()

    # ── Reviewer notification callout ──────────────────────────────────────────
    # Shown after a step is marked complete; cleared by Dismiss button.
    #
    # last_completed_step is a one-shot "you just did this" flag, set at the
    # moment a step is marked complete — nothing clears it if that step later
    # becomes un-done again (the ↩ Undo button below, an auto-step resetting
    # on a later re-run, or switching close periods, none of which touch this
    # variable). Without the close_tracker re-check here, the "Step 9
    # Complete — Close complete!" banner can keep showing even after the
    # live progress count has dropped back below 9/9 — confirmed with Ryan
    # 2026-08-19 (banner said "all 9 confirmed" while the progress bar
    # correctly showed 8/9). Re-validating against the CURRENT close_tracker
    # state makes this self-healing regardless of what caused the step to
    # become un-done.
    _last_step = st.session_state.get('last_completed_step')
    if _last_step is not None and _last_step not in st.session_state.close_tracker:
        _last_step = None
        st.session_state.last_completed_step = None
    if _last_step is not None:
        # Per-step notification config
        _STEP_NOTIFS = {
            0: {
                'emoji': '📋', 'color': '#1565C0', 'bg': '#E3F2FD',
                'headline': 'Next: Start Pass 1',
                'body': ('Bank rec and payments have been confirmed complete for {period}. '
                         'Upload Pass 1 source files and generate the JE CSVs.'),
            },
            1: {
                'emoji': '📤', 'color': '#1565C0', 'bg': '#E3F2FD',
                'headline': 'Next: Upload JE CSVs to Yardi',
                'body': ('Pass 1 JEs have been generated for {period}. '
                         'Download the Accruals, Prepaid, and Manual JE CSVs, upload them to '
                         'Yardi, then run the final close.'),
            },
            2: {
                'emoji': '⚙️', 'color': '#1565C0', 'bg': '#E3F2FD',
                'headline': 'Next: Run final close in Yardi',
                'body': ('JEs have been uploaded to Yardi for {period}. '
                         'Run the final close to post all entries.'),
            },
            3: {
                'emoji': '📥', 'color': '#1565C0', 'bg': '#E3F2FD',
                'headline': 'Next: Re-export final files from Yardi',
                'body': ('Final close has been run in Yardi for {period}. '
                         'Re-export the GL, TB, BC, and Bank Rec — then return to upload Pass 2 files.'),
            },
            4: {
                'emoji': '🗂️', 'color': '#1565C0', 'bg': '#E3F2FD',
                'headline': 'Next: Upload Pass 2 files & generate reports',
                'body': ('Final Yardi files are ready for {period}. '
                         'Upload them to the Pass 2 section and click Generate Reports.'),
            },
            5: {
                'emoji': '📊', 'color': '#1565C0', 'bg': '#E3F2FD',
                'headline': 'Next: Generate Pass 2 reports',
                'body': ('Pass 2 files uploaded for {period}. '
                         'Click Generate Reports in the Pass 2 tab.'),
            },
            6: {
                'emoji': '🔍', 'color': '#5C3317', 'bg': '#FFF8E1',
                'headline': 'Action needed: QC review — Ryan & Natasha',
                'body': ('Pass 2 reports have been generated for {period}. '
                         'Ryan and Natasha: please review the QC workbook, workpapers, '
                         'and exception report before releasing the final package.'),
                'draft': True,
                'draft_label': '📋 Suggested review note for Natasha:',
                'draft_fn': lambda period, prop, team, user: (
                    f"Hi {next((m.split()[0] for m in team if 'natasha' in m.lower()), 'Natasha')},\n\n"
                    f"Pass 2 reports for {period} ({prop}) are ready for QC review.\n\n"
                    f"Please check:\n"
                    f"  1. QC Workbook (7 checks)\n"
                    f"  2. Workpapers (GL vs TB tie-out)\n"
                    f"  3. Budget Comparison with variance comments\n"
                    f"  4. Exception Report\n\n"
                    f"Let me know if anything looks off.\n\n"
                    f"Thanks,\n{user}"
                ),
            },
            7: {
                'emoji': '📦', 'color': '#2D6F50', 'bg': '#E8F5E9',
                'headline': 'Action needed: Release final package to Accounting Manager/Controller',
                'body': ('{period} QC review is complete for {prop}. '
                         'The final package is ready for the Accounting Manager/Controller. '
                         'Upload the deliverables and mark Step 9 complete.'),
                'draft': True,
                'draft_label': '📋 Suggested message for the Accounting Manager/Controller:',
                # team_members convention: [Property Accountant, Property Manager, Accounting Manager/Controller] —
                # Accounting Manager/Controller is the 3rd entry by position, not matched by name.
                'draft_fn': lambda period, prop, team, user: (
                    f"Hi {(team[2].split()[0] if len(team) > 2 and team[2] else 'there')},\n\n"
                    f"The {period} monthly close package for {prop} is ready for your review. "
                    f"All QC checks have passed and the workpapers have been signed off.\n\n"
                    f"Please let us know if you have any questions.\n\n"
                    f"Thank you,\n{user}"
                ),
            },
            8: {
                'emoji': '🎉', 'color': '#2D6F50', 'bg': '#E8F5E9',
                'headline': 'Close complete!',
                'body': ('The {period} monthly close for {prop} is fully complete. '
                         'All 9 steps confirmed. Great work!'),
            },
        }

        _ni = _STEP_NOTIFS.get(_last_step)
        if _ni:
            _period_lbl  = period_key_to_label(_ck_pkey) if '_ck_pkey' in dir() else period_key_to_label(st.session_state.get('checklist_period_key', current_period_key()))
            _team_members = list(getattr(_active_cfg, 'team_members', None) or
                                 ['[Property Accountant]', '[Property Manager]', '[Accounting Manager/Controller]'])
            _prepared_by  = st.session_state.get('prepared_by', 'GRP')
            _body_text = _ni['body'].format(
                period=_period_lbl,
                prop=_prop_display,
            )
            st.markdown(
                f"<div style='background:{_ni['bg']};border-left:4px solid {_ni['color']};"
                f"border-radius:6px;padding:12px 16px;margin-bottom:12px;'>"
                f"<div style='font-weight:700;color:{_ni['color']};font-size:0.92rem;'>"
                f"{_ni['emoji']} Step {_last_step + 1} Complete &nbsp;·&nbsp; {_ni['headline']}"
                f"</div>"
                f"<div style='font-size:0.85rem;color:#424242;margin-top:5px;'>{_body_text}</div>"
                f"</div>",
                unsafe_allow_html=True,
            )
            # Draft message for handoff steps (6 → Natasha, 7 → Lauren)
            if _ni.get('draft'):
                _draft_text = _ni['draft_fn'](
                    _period_lbl, _prop_display, _team_members, _prepared_by
                )
                st.markdown(f"**{_ni['draft_label']}**")
                st.text_area(
                    'draft_msg',
                    value=_draft_text,
                    height=140,
                    key=f'notif_draft_{_last_step}',
                    label_visibility='collapsed',
                )
            _nd_col, _ = st.columns([1, 6])
            with _nd_col:
                if st.button('✕ Dismiss', key='notif_dismiss_btn', use_container_width=True):
                    st.session_state.last_completed_step = None
                    st.rerun()

    # ── Step checklist ─────────────────────────────────────────────────────────
    # Group into Pre-Close / Post-Close
    _PHASE_PRE  = list(range(0, 5))   # steps 0–4
    _PHASE_POST = list(range(5, 9))   # steps 5–8

    _ck_locked = st.session_state.get('checklist_locked', False)

    def _render_phase(phase_label: str, step_indices: list) -> None:
        st.markdown(
            f"<div style='font-size:0.78rem;font-weight:700;text-transform:uppercase;"
            f"letter-spacing:0.08em;color:#757575;margin:8px 0 4px 0;'>{phase_label}</div>",
            unsafe_allow_html=True,
        )
        for _si, _sdesc, _stype in [s for s in _CTS if s[0] in step_indices]:
            _ct_entry = st.session_state.close_tracker.get(_si)
            _is_done  = bool(_ct_entry)
            _icon     = '✅' if _is_done else ('🔄' if _stype == 'auto' else '⬜')
            _by       = _ct_entry.get('completed_by', '') if _ct_entry else '—'
            _ts       = _ct_entry.get('timestamp', '') if _ct_entry else '—'
            _auto_tag = ' *(auto)*' if (_ct_entry and _ct_entry.get('auto')) else ''

            _sc1, _sc2, _sc3, _sc4, _sc5 = st.columns([0.4, 3.5, 1.8, 1.8, 1.5])
            with _sc1:
                st.markdown(f"<div style='font-size:1.1rem;padding-top:4px'>{_icon}</div>",
                            unsafe_allow_html=True)
            with _sc2:
                _color = '#212121' if _is_done else '#616161'
                _weight = '500' if _is_done else '400'
                st.markdown(
                    f"<div style='color:{_color};font-weight:{_weight};"
                    f"font-size:0.88rem;padding-top:6px;'>{_sdesc}{_auto_tag}</div>",
                    unsafe_allow_html=True,
                )
            with _sc3:
                st.markdown(
                    f"<div style='font-size:0.82rem;color:#616161;padding-top:6px;'>{_by}</div>",
                    unsafe_allow_html=True,
                )
            with _sc4:
                st.markdown(
                    f"<div style='font-size:0.82rem;color:#616161;padding-top:6px;'>{_ts}</div>",
                    unsafe_allow_html=True,
                )
            with _sc5:
                if _ck_locked:
                    st.markdown(
                        "<div style='font-size:0.75rem;color:#9E9E9E;padding-top:6px;'>"
                        "🔒 locked</div>",
                        unsafe_allow_html=True,
                    )
                elif _stype == 'manual' and not _is_done:
                    if st.button('Mark Complete', key=f'ck_mark_{_si}',
                                 use_container_width=True):
                        st.session_state.close_tracker[_si] = {
                            'completed_by': st.session_state.prepared_by,
                            'timestamp':    datetime.now().strftime('%m/%d/%Y %H:%M'),
                            'auto':         False,
                        }
                        _save_checklist_now()
                        st.session_state.last_completed_step = _si
                        st.rerun()
                elif _is_done and _stype == 'manual':
                    if st.button('↩ Undo', key=f'ck_undo_{_si}',
                                 use_container_width=True):
                        st.session_state.close_tracker.pop(_si, None)
                        _save_checklist_now()
                        st.rerun()
                else:
                    st.markdown('')   # spacer

    _render_phase('Pre-Close', _PHASE_PRE)
    _render_phase('Post-Close', _PHASE_POST)

    st.divider()

    # ── Custom (one-off) tasks ─────────────────────────────────────────────────
    st.markdown(
        "<div style='font-size:0.78rem;font-weight:700;text-transform:uppercase;"
        "letter-spacing:0.08em;color:#757575;margin:8px 0 4px 0;'>Custom Tasks "
        "This Close</div>",
        unsafe_allow_html=True,
    )

    for _ci_idx, _ci in enumerate(st.session_state.custom_checklist_items):
        _ci_done = _ci.get('completed', False)
        _ci_icon = '✅' if _ci_done else '⬜'
        _ci_by   = _ci.get('completed_by', '—') if _ci_done else _ci.get('created_by', '—')
        _ci_ts   = _ci.get('completed_at', '—') if _ci_done else f"added {_ci.get('created_at','')}"

        _cc1, _cc2, _cc3, _cc4, _cc5 = st.columns([0.4, 3.5, 1.8, 1.8, 1.5])
        with _cc1:
            st.markdown(f"<div style='font-size:1.1rem;padding-top:4px'>{_ci_icon}</div>",
                        unsafe_allow_html=True)
        with _cc2:
            _ci_color = '#212121' if _ci_done else '#616161'
            st.markdown(
                f"<div style='color:{_ci_color};font-size:0.88rem;padding-top:6px;'>"
                f"{_ci['label']}</div>",
                unsafe_allow_html=True,
            )
        with _cc3:
            st.markdown(
                f"<div style='font-size:0.82rem;color:#616161;padding-top:6px;'>{_ci_by}</div>",
                unsafe_allow_html=True,
            )
        with _cc4:
            st.markdown(
                f"<div style='font-size:0.82rem;color:#616161;padding-top:6px;'>{_ci_ts}</div>",
                unsafe_allow_html=True,
            )
        with _cc5:
            if _ck_locked:
                st.markdown(
                    "<div style='font-size:0.75rem;color:#9E9E9E;padding-top:6px;'>"
                    "🔒 locked</div>",
                    unsafe_allow_html=True,
                )
            elif not _ci_done:
                if st.button('Mark Complete', key=f'ci_mark_{_ci_idx}',
                             use_container_width=True):
                    st.session_state.custom_checklist_items[_ci_idx]['completed']    = True
                    st.session_state.custom_checklist_items[_ci_idx]['completed_by'] = (
                        st.session_state.prepared_by)
                    st.session_state.custom_checklist_items[_ci_idx]['completed_at'] = (
                        datetime.now().strftime('%m/%d/%Y %H:%M'))
                    _save_checklist_now()
                    st.rerun()
            else:
                if st.button('↩ Undo', key=f'ci_undo_{_ci_idx}',
                             use_container_width=True):
                    st.session_state.custom_checklist_items[_ci_idx]['completed']    = False
                    st.session_state.custom_checklist_items[_ci_idx]['completed_by'] = None
                    st.session_state.custom_checklist_items[_ci_idx]['completed_at'] = None
                    _save_checklist_now()
                    st.rerun()

    # ── Add custom task (hidden when locked) ──────────────────────────────────
    if not _ck_locked:
        with st.expander('➕  Add a custom task for this close', expanded=False):
            _new_label = st.text_input('Task description', key='new_custom_task_label',
                                       placeholder='e.g. Confirm Berkadia loan pay-off')
            if st.button('Add Task', key='add_custom_task_btn', type='primary'):
                if _new_label.strip():
                    _new_id = f'custom_{len(st.session_state.custom_checklist_items)}'
                    st.session_state.custom_checklist_items.append({
                        'id':           _new_id,
                        'label':        _new_label.strip(),
                        'created_by':   st.session_state.prepared_by,
                        'created_at':   datetime.now().strftime('%m/%d/%Y %H:%M'),
                        'completed':    False,
                        'completed_by': None,
                        'completed_at': None,
                    })
                    _save_checklist_now()
                    st.rerun()
                else:
                    st.warning('Please enter a task description.')


# ──────────────────────────────────────────────────────────────
# TAB 1 — PASS 1: JE GENERATION
# ──────────────────────────────────────────────────────────────
with tab1:
    st.markdown(
        "<div style='background:var(--grp-green-lit,#E8F5E9);border-left:4px solid #1A5C22;"
        "border-radius:5px;padding:8px 16px;margin-bottom:14px;font-size:0.85rem;color:#1A5C22;'>"
        "⬤ &nbsp;<strong>Pass 1 — Pre-Close</strong>&nbsp; Upload source files, review accruals, "
        "download JE CSVs → post to Yardi → run close."
        "</div>",
        unsafe_allow_html=True,
    )

    # ── File Upload ──────────────────────────────────────────────────────────
    st.markdown("### 📥 Upload Pass 1 Files")
    st.caption(
        "Select all source files at once — GL, Budget Comparison, Trial Balance, "
        "Bank Rec, Nexus, DACA, Berkadia, etc. The app auto-detects each file type. "
        "Use the dropdown to correct any mismatches."
    )

    _bulk_p1 = st.file_uploader(
        "Drop Pass 1 files here",
        accept_multiple_files=True,
        type=["xlsx", "xls", "pdf"],
        key=f"bulk_upload_p1_{st.session_state.upload_key_p1}",
        label_visibility="collapsed",
    )

    # Only clear/reclassify Pass 1 slots when the uploaded file set actually
    # changed. Running this on EVERY script rerun (including unrelated
    # interactions elsewhere on the page, like toggling a JE checkbox) meant
    # any transient hiccup in reading the file uploader's state on a given
    # rerun would silently drop files — and everything derived from them —
    # from that run with no error surfaced, requiring a re-run to "settle."
    _bulk_p1_fingerprint = tuple(sorted((f.name, f.size) for f in (_bulk_p1 or [])))
    if st.session_state.get('_bulk_p1_fingerprint') != _bulk_p1_fingerprint:
        st.session_state['_bulk_p1_fingerprint'] = _bulk_p1_fingerprint

        # Clear all Pass 1 slots so stale entries don't persist after a file is removed
        for _clr_key in set(_P1_SLOT_KEYS) - {"unknown"}:
            st.session_state.uploaded_files.pop(_clr_key, None)

    if _bulk_p1:
        _loan_paths_p1: list = []

        for _uf in _bulk_p1:
            _raw = bytes(_uf.getbuffer())
            _det_key, _conf, _det_label = _classify_file(_uf.name, _raw, pass2=False, property_config=_active_cfg)
            # B-6: key by (name, size) so a re-uploaded file with the same name but
            # different content gets a fresh key and the stale override doesn't survive.
            _ovr_key = (_uf.name, _uf.size)
            _eff_key = st.session_state.bulk_overrides_p1.get(_ovr_key, _det_key)

            _ic, _fn_col, _tp_col = st.columns([0.5, 5, 5])
            if _eff_key == "unknown":
                _ic.markdown("⚠️")
            elif _conf >= 0.85:
                _ic.markdown("✅")
            else:
                _ic.markdown("🟡")

            _short = _uf.name if len(_uf.name) <= 40 else _uf.name[:37] + "…"
            _fn_col.caption(_short)

            if _eff_key == "unknown" or _conf < 0.70:
                _cur_idx = (_P1_SLOT_KEYS.index(_eff_key)
                            if _eff_key in _P1_SLOT_KEYS else len(_P1_SLOT_KEYS) - 1)
                _sel_label = _tp_col.selectbox(
                    "type", _P1_SLOT_LABELS, index=_cur_idx,
                    key=f"ovr_p1_{_uf.name}", label_visibility="collapsed",
                )
                _eff_key = _P1_SLOT_KEYS[_P1_SLOT_LABELS.index(_sel_label)]
                st.session_state.bulk_overrides_p1[_ovr_key] = _eff_key
            else:
                _tp_col.caption(_det_label)

            if _eff_key != "unknown":
                _tp = os.path.join(st.session_state.temp_dir, _uf.name)
                if not os.path.exists(_tp) or os.path.getsize(_tp) != _uf.size:
                    with open(_tp, "wb") as _f:
                        _f.write(_raw)
                if _eff_key in _MULTI_FILE_KEYS:
                    _loan_paths_p1.append(_tp)
                else:
                    st.session_state.uploaded_files[_eff_key] = _tp

        if _loan_paths_p1:
            st.session_state.uploaded_files["loan"] = _loan_paths_p1

        # Prune overrides for files no longer in the upload widget (name+size composite key)
        _active_keys_p1 = {(_uf.name, _uf.size) for _uf in _bulk_p1}
        st.session_state.bulk_overrides_p1 = {
            k: v for k, v in st.session_state.bulk_overrides_p1.items()
            if k in _active_keys_p1
        }

    # ── Upload status ─────────────────────────────────────────────────────────
    uploaded_keys = set(st.session_state.uploaded_files.keys())
    missing_impact = []
    if "trial_balance"     not in uploaded_keys: missing_impact.append("No BS tie-out validation (Pass 2)")
    if "budget_comparison" not in uploaded_keys: missing_impact.append("No historical pattern accruals or variance comments")
    if "bank_rec"          not in uploaded_keys: missing_impact.append("No Operating bank rec tab (Pass 2)")
    if "daca_bank"         not in uploaded_keys: missing_impact.append("No DACA bank rec tab (Pass 2)")
    if "loan"              not in uploaded_keys: missing_impact.append("No debt service tab (Pass 2)")

    # Count all Pass 1 uploaded files — exclude Pass 2-only keys so the sidebar
    # count reflects exactly what was dropped in the bulk uploader.
    # List values (e.g. loan with 3 Berkadia PDFs) are unfolded so each PDF counts.
    _P2_ONLY_KEYS = {
        "gl_pass2", "budget_comparison_pass2", "trial_balance_pass2",
        "t12_statement_pass2", "loan_pass2", "bank_rec_pass2",
        "prior_workpaper", "run_log", "ap_aging_pass2",
        "bank_rec_xlsx_pass2", "daca_bank_xlsx_pass2",
        "prepaid_ledger_pass2", "capital_seed_pass2",
    }
    uploaded_count = sum(
        len(v) if isinstance(v, list) else 1
        for k, v in st.session_state.uploaded_files.items()
        if k not in _P2_ONLY_KEYS and k != "unknown" and v is not None
    )
    gl_uploaded = "gl" in uploaded_keys

    _st_c1, _st_c2 = st.columns(2)
    with _st_c1:
        st.caption(f"**{uploaded_count} file(s) uploaded**")
    with _st_c2:
        if missing_impact:
            with st.expander(f"⚠️ {len(missing_impact)} output(s) won't generate", expanded=False):
                for m in missing_impact:
                    st.caption(f"• {m}")

    if not gl_uploaded and uploaded_count > 0:
        st.warning("⚠️ GL Detail is required to run either pass.")

    # ── Committed reference file status ───────────────────────────────────────
    if _COMMITTED_BUDGET:
        _budget_uploaded = "kardin_budget" in uploaded_keys
        if _budget_uploaded:
            st.caption("📊 **Kardin Budget:** Using uploaded file _(overrides committed)_")
        else:
            st.caption("📊 **Kardin Budget:** FY2026 on file — no upload needed")
    else:
        if "kardin_budget" not in uploaded_keys:
            st.caption("📊 **Kardin Budget:** Not uploaded — bonus accruals and QC budget check skipped")

    st.divider()

    # ── Tenant Utility Billing ────────────────────────────────────────────────
    _tenant_utility_rows = []
    _tu_elec_total = 0.0
    _tu_gas_total  = 0.0
    with st.expander("⚡ Tenant Utility Billing — Enter monthly meter read amounts", expanded=False):
        st.caption(
            "Enter electric and gas amounts per tenant from the monthly meter read. "
            "Posts as: DR 133110 / CR 440500 (electric) and CR 440700 (gas). "
            "Leave at $0 to skip — pipeline auto-accrues budget amounts."
        )
        _ts_uploaded = bool(st.session_state.uploaded_files.get('tenancy_schedule'))
        _tub_tenants = _build_tub_tenants(_active_cfg)
        if _ts_uploaded and _tub_tenants:
            st.success(
                f"📋 {len(_tub_tenants)} tenant(s) auto-detected from the uploaded Rent Roll — "
                f"confirm each is correct below. Uncheck any that shouldn't be billed this period.",
                icon="📋",
            )
        _tub_cols = st.columns(max(len(_tub_tenants), 1))
        for (_tkey, _tname), _tcol in zip(_tub_tenants, _tub_cols):
            with _tcol:
                if _ts_uploaded:
                    _tincluded = st.checkbox(
                        _tname, value=True,
                        key=f"tub_incl_{_tkey}_{st.session_state.tub_key}",
                        help="Confirms this tenant from the Rent Roll should be billed this "
                             "period. Uncheck to exclude — e.g. a lease that just ended.",
                    )
                else:
                    st.caption(f"**{_tname}**")
                    _tincluded = True
                _telec = st.number_input(
                    "Electric ($)", min_value=0.0, value=0.0, step=1.0, format="%.2f",
                    key=f"tub_elec_{_tkey}_{st.session_state.tub_key}",
                    disabled=not _tincluded,
                )
                _tgas = st.number_input(
                    "Gas ($)", min_value=0.0, value=0.0, step=1.0, format="%.2f",
                    key=f"tub_gas_{_tkey}_{st.session_state.tub_key}",
                    disabled=not _tincluded,
                )
            if _tincluded and (_telec > 0 or _tgas > 0):
                _tenant_utility_rows.append({'tenant': _tname, 'electric': _telec, 'gas': _tgas})
                _tu_elec_total += _telec
                _tu_gas_total  += _tgas
        if _tenant_utility_rows:
            st.caption(f"✓ {len(_tenant_utility_rows)} tenant(s) — Electric ${_tu_elec_total:,.2f} / Gas ${_tu_gas_total:,.2f}")
        else:
            _rd_loaded = bool(st.session_state.uploaded_files.get("receivable_detail"))
            if _rd_loaded:
                st.caption("↳ No entries — pipeline will read electric amounts from Receivable Detail (per-tenant)")
            elif bool(st.session_state.uploaded_files.get("receivable_summary")):
                st.caption("↳ No entries — Receivable Summary uploaded; upload Receivable Detail too for per-tenant electric breakdown")
            else:
                st.caption(
                    "↳ No entries — upload **Yardi Receivable Detail** in the sidebar for automatic "
                    "per-tenant electric amounts, or enter amounts above manually"
                )

    # ── Payroll Bonus Accrual ─────────────────────────────────────────────────
    with st.expander("💰 Payroll Bonus Accrual — Monthly (optional)", expanded=False):
        st.caption(
            "Enter the annual bonus budget for engineering and/or admin payroll. "
            "The pipeline accrues 1/12 each month and suppresses automatically "
            "in months when the actual bonus payment hits the GL."
        )
        _bonus_col1, _bonus_col2 = st.columns(2)
        with _bonus_col1:
            _bonus_rm = st.number_input(
                "RM-Pay/Wages (615110) — Annual Bonus ($)",
                min_value=0.0, value=0.0, step=1000.0, format="%.2f",
                key="widget_bonus_rm",
                help="Engineering/RM annual bonus budget. Monthly accrual = annual ÷ 12.",
            )
        with _bonus_col2:
            _bonus_admin = st.number_input(
                "Admin-Pay/Wages (637110) — Annual Bonus ($)",
                min_value=0.0, value=0.0, step=1000.0, format="%.2f",
                key="widget_bonus_admin",
                help="Administrative/office annual bonus budget. Monthly accrual = annual ÷ 12.",
            )
        _bonus_overrides: dict = {}
        if _bonus_rm > 0:
            _bonus_overrides['615110'] = float(_bonus_rm)
        if _bonus_admin > 0:
            _bonus_overrides['637110'] = float(_bonus_admin)
        if _bonus_overrides:
            _bonus_monthly_display = {k: f"${v/12:,.2f}/mo" for k, v in _bonus_overrides.items()}
            st.caption(f"✓ Monthly accruals: {' | '.join(f'{k}: {v}' for k, v in _bonus_monthly_display.items())}")
        else:
            st.caption("↳ No bonus entered — pipeline will use Kardin data if uploaded, otherwise skip")

    # ── RE Tax Bill ───────────────────────────────────────────────────────────
    with st.expander("🏛️ RE Tax Bill — Enter every month", expanded=False):
        st.caption(
            "Enter the quarterly RE Tax bill amount **every month** (same amount for all 3 months in each cycle). "
            + (lambda _m: (
                f"**Payment months ({', '.join(_month_abbr(_x) for _x in sorted(_m))}):** "
                f"Lender/escrow agent auto-posts the full bill to Yardi. "
            ))(getattr(_active_cfg, 're_tax_payment_months', None) or [1, 4, 7, 10]) +
            "Pipeline defers 2/3 → DR 135120 Prepaid RE Taxes / CR 641110 Real Estate Taxes. "
            "**Release months (all other):** Pipeline releases 1/3 → DR 641110 Real Estate Taxes / CR 135120 Prepaid RE Taxes. "
            "Net result: expense is spread evenly — 1/3 of the quarterly bill per month."
        )
        _re_tax_bill_amount = st.number_input(
            "Quarterly RE Tax Bill ($)",
            min_value=0.0, value=0.0, step=1000.0, format="%.2f",
            key="widget_re_tax_bill",
        )
        _re_tax_bill_amount = _re_tax_bill_amount if _re_tax_bill_amount > 0 else 0.0

    st.divider()

    st.markdown("""
    **What this does:** Reads your pre-close Yardi GL and detects every accrual entry needed
    to complete the month-end close — invoices in Nexus not yet posted, utility proration,
    historical recurring patterns, management fee, prepaid amortization, bonus accruals,
    and one-off items you enter below. Exports two Yardi-import files.

    **Next step after this tab:** Upload the CSVs to Yardi → run final close → switch to **Pass 2**.
    """)
    st.divider()

    # ── One-Off Accruals Table ────────────────────────────────────────────────
    # Plain widgets (text_input/number_input/checkbox) in a dynamic row list —
    # not st.data_editor. The grid's canvas-based editing (glide-data-grid)
    # could lose an in-progress "Account Code" edit when the native "+" add-row
    # button triggered its own rerun (confirmed with Ryan 2026-08-06 even after
    # removing an unrelated forced st.rerun() elsewhere in this block — the
    # fragility is in the grid itself). Each row here is its own set of
    # independently-keyed widgets, so adding/removing a row never touches any
    # other row's key or in-progress value. Same fix already applied to
    # "Add Missed Entries" below.
    with st.expander("🧾 One-Off Accruals  (DR expense → CR 213100 Accrued Expenses)", expanded=False):
        st.caption(
            "Use this for known invoices not yet in Nexus or Yardi — quarterly contracts, "
            "seasonal items, recurring retainers, semi-annual billings, etc. "
            "All entries debit the expense account and credit **213100 Accrued Expenses** — "
            "they auto-reverse next period. "
            "**Leave Amount at $0** to suppress automated detection for that account without generating a JE — "
            "use this when a JE has already been posted to Yardi to prevent double-counting."
        )

        _split_sch_help = (
            "Which buildings this line's dollar amount gets pro-rated across. "
            "'(use property default)' applies the property's default split "
            "schedule (see ⚙️ Properties tab). 'No Split' posts the full "
            "amount to the parent property code regardless of default."
        ) if _active_cfg.is_multi_building else (
            "Only applies to multi-building properties. "
            "Configure building splits in the ⚙️ Properties tab."
        )
        _SPLIT_DEFAULT_SENTINEL = "(use property default)"
        _split_sch_options = (
            [_SPLIT_DEFAULT_SENTINEL] + sorted(_active_cfg.allocation_schedules.keys()) + ["No Split"]
            if _active_cfg.is_multi_building else [_SPLIT_DEFAULT_SENTINEL]
        )

        _OA_IDS_KEY  = "oa_row_ids"
        _OA_NEXT_KEY = "oa_next_id"
        _OA_GEN_KEY  = "oa_rows_seed_gen"

        def _oa_seed_widget(_rid: int, _seed: dict) -> None:
            st.session_state[f"oa_code_{_rid}"]     = _seed["Account Code"]
            st.session_state[f"oa_name_{_rid}"]     = _seed["Account Name"]
            st.session_state[f"oa_vendor_{_rid}"]   = _seed["Vendor"]
            st.session_state[f"oa_amt_{_rid}"]      = _seed["Amount ($)"]
            st.session_state[f"oa_prior_{_rid}"]    = _seed["Prior Accrual ($)"]
            st.session_state[f"oa_desc_{_rid}"]     = _seed["Description"]
            st.session_state[f"oa_autorev_{_rid}"]  = _seed["Auto-Reverse"]
            st.session_state[f"oa_split_{_rid}"]    = _seed["Split Schedule"] or _SPLIT_DEFAULT_SENTINEL
            st.session_state[f"oa_compound_{_rid}"] = _seed["Compound"]

        # (Re)seed the row list only when manual_accruals_df was just replaced
        # out from under us (property switch, Reset All, session restore) —
        # see _accruals_seed_gen. Never on an ordinary rerun, since this
        # block's own write-back of manual_accruals_df (below) must not
        # trigger re-seeding itself and clobber in-progress edits.
        if (_OA_IDS_KEY not in st.session_state
                or st.session_state.get(_OA_GEN_KEY) != st.session_state.get("_accruals_seed_gen", 0)):
            _oa_seed_rows = _df_to_oa_rows(st.session_state.manual_accruals_df, _active_cfg)
            _oa_new_ids = list(range(len(_oa_seed_rows)))
            for _rid, _seed in zip(_oa_new_ids, _oa_seed_rows):
                _oa_seed_widget(_rid, _seed)
            st.session_state[_OA_IDS_KEY]  = _oa_new_ids
            st.session_state[_OA_NEXT_KEY] = len(_oa_seed_rows)
            st.session_state[_OA_GEN_KEY]  = st.session_state.get("_accruals_seed_gen", 0)

        # ── Account name auto-populate ────────────────────────────────────────
        # When the user types an account code in a row and leaves the Account
        # Name blank, look it up from (1) property config defaults, (2) the
        # parsed GL, (3) Budget Comparison, and fill it in. Must run BEFORE
        # this run's widgets are instantiated below — Streamlit forbids
        # writing to a widget's session_state key after that widget has
        # already rendered in the same script run.
        _acct_name_lookup: dict = {}

        # Source 1: property config default_accruals
        for _da in (getattr(_active_cfg, 'default_accruals', None) or []):
            _da_code = str(_da.get('account_code', '') or '').strip()
            _da_name = str(_da.get('account_name', '') or '').strip()
            if _da_code and _da_name:
                _acct_name_lookup[_da_code] = _da_name

        # Source 2: previously parsed GL (populated after the first Generate JEs run)
        _p1_er_for_lookup = st.session_state.get('pass1_engine_result')
        if _p1_er_for_lookup:
            try:
                _gl_for_lookup = getattr(_p1_er_for_lookup, 'parsed', {}).get('gl')
                if _gl_for_lookup and hasattr(_gl_for_lookup, 'accounts'):
                    for _gla in _gl_for_lookup.accounts:
                        _gla_code = str(getattr(_gla, 'account_code', '') or '').strip()
                        _gla_name = str(getattr(_gla, 'account_name', '') or '').strip()
                        if _gla_code and _gla_name:
                            _acct_name_lookup[_gla_code] = _gla_name
            except Exception:
                pass

        # Source 3: Budget Comparison — same fallback the 7xxxxx Recode table
        # uses below, added here too so an account with no GL activity yet
        # this period (but a budget line) still auto-populates a name.
        if _p1_er_for_lookup:
            try:
                _bc_for_lookup_oa = getattr(_p1_er_for_lookup, 'parsed', {}).get('budget_comparison')
                if _bc_for_lookup_oa and hasattr(_bc_for_lookup_oa, 'line_items'):
                    for _bcl_oa in _bc_for_lookup_oa.line_items:
                        _bcl_oa_code = str(getattr(_bcl_oa, 'account_code', '') or '').strip()
                        _bcl_oa_name = str(getattr(_bcl_oa, 'account_name', '') or '').strip()
                        if _bcl_oa_code and _bcl_oa_name and _bcl_oa_code not in _acct_name_lookup:
                            _acct_name_lookup[_bcl_oa_code] = _bcl_oa_name
            except Exception:
                pass

        for _rid in st.session_state[_OA_IDS_KEY]:
            _code_val = str(st.session_state.get(f"oa_code_{_rid}", "") or "").strip()
            _name_val = str(st.session_state.get(f"oa_name_{_rid}", "") or "").strip()
            if _code_val and not _name_val and _code_val in _acct_name_lookup:
                st.session_state[f"oa_name_{_rid}"] = _acct_name_lookup[_code_val]

        # Widen the Split Schedule dropdown with any value already stored on
        # a row (e.g. a schedule that was renamed/removed since this row was
        # seeded) so the selectbox never errors on a value outside its
        # current options list. Runs after seeding so it sees every row's
        # final value for this run.
        for _rid in st.session_state[_OA_IDS_KEY]:
            _cur_split = st.session_state.get(f"oa_split_{_rid}")
            if _cur_split and _cur_split not in _split_sch_options:
                _split_sch_options.append(_cur_split)

        # ── Row widgets ──────────────────────────────────────────────────────
        for _row_i, _rid in enumerate(st.session_state[_OA_IDS_KEY]):
            _lbl = "visible" if _row_i == 0 else "collapsed"
            _top = st.columns([1.3, 2.0, 1.6, 1.3, 1.3, 0.5])
            with _top[0]:
                st.text_input("DR Account", key=f"oa_code_{_rid}", label_visibility=_lbl,
                              placeholder="e.g. 613310",
                              help="6-digit Yardi GL account code (e.g. 613310)")
            with _top[1]:
                st.text_input("Account Name", key=f"oa_name_{_rid}", label_visibility=_lbl)
            with _top[2]:
                st.text_input("Vendor", key=f"oa_vendor_{_rid}", label_visibility=_lbl)
            with _top[3]:
                st.number_input("Amount ($)", key=f"oa_amt_{_rid}", label_visibility=_lbl,
                                min_value=0.0, step=100.0, format="%.2f",
                                help="Monthly accrual amount — debit to expense account")
            with _top[4]:
                st.number_input("Prior Accrual ($)", key=f"oa_prior_{_rid}", label_visibility=_lbl,
                                min_value=0.0, step=100.0, format="%.2f",
                                help="For semi-annual / irregular accounts (e.g. Water/Sewer): "
                                     "enter the cumulative prior-month accrual balance to seed compounding "
                                     "on the first pipeline run. Leave $0 once prior pipeline auto-reversals "
                                     "appear in the GL — the pipeline picks them up automatically.")
            with _top[5]:
                if _row_i == 0:
                    st.write("")   # align delete button with inputs, not their labels
                if st.button("🗑️", key=f"oa_del_{_rid}", help="Remove this row"):
                    st.session_state[_OA_IDS_KEY] = [
                        _i for _i in st.session_state[_OA_IDS_KEY] if _i != _rid
                    ]
                    st.rerun()

            _bot = st.columns([2.8, 1.0, 1.5, 1.0])
            with _bot[0]:
                st.text_input("Description", key=f"oa_desc_{_rid}", label_visibility=_lbl,
                              help="Description for the Yardi JE line")
            with _bot[1]:
                st.checkbox("Auto-Rev", key=f"oa_autorev_{_rid}",
                           help="✅ Checked = entry auto-reverses next month (ReverseNextMonth = -1). "
                                "Uncheck for permanent JEs that should NOT reverse (ReverseNextMonth = 0).")
            with _bot[2]:
                st.selectbox("Split Schedule", options=_split_sch_options, key=f"oa_split_{_rid}",
                             label_visibility=_lbl, help=_split_sch_help,
                             disabled=not _active_cfg.is_multi_building)
            with _bot[3]:
                st.checkbox("Compound", key=f"oa_compound_{_rid}",
                           help="❌ Unchecked (default for new rows) = flat monthly amount — correct "
                                "for a normal monthly-billed account. "
                                "✅ Check ONLY for accounts billed irregularly (e.g. Water/Sewer "
                                "every 6 months), where this month's JE = last month's auto-reversed "
                                "amount + this month's Amount ($), so the accrued liability keeps "
                                "building until a real invoice clears it. Checking this for a normal "
                                "monthly account double-counts against the real invoice landing "
                                "that month.")
            if _row_i < len(st.session_state[_OA_IDS_KEY]) - 1:
                st.markdown("<hr style='margin:4px 0;opacity:0.15;'>", unsafe_allow_html=True)

        if st.button("➕ Add Row", key="oa_add_row_btn"):
            _new_rid = st.session_state[_OA_NEXT_KEY]
            st.session_state[_OA_NEXT_KEY] += 1
            _oa_seed_widget(_new_rid, _blank_oa_row())
            st.session_state[_OA_IDS_KEY] = st.session_state[_OA_IDS_KEY] + [_new_rid]
            st.rerun()

        # ── Write back to manual_accruals_df for downstream consumers ────────
        # (accrual_entry_generator exclusion list, JE building, session
        # snapshot save) — same DataFrame shape/columns as before, just
        # rebuilt from the live widget values instead of a data_editor return.
        _oa_out_rows = [
            {
                "Account Code": st.session_state.get(f"oa_code_{_rid}", ""),
                "Account Name": st.session_state.get(f"oa_name_{_rid}", ""),
                "Vendor": st.session_state.get(f"oa_vendor_{_rid}", ""),
                "Amount ($)": float(st.session_state.get(f"oa_amt_{_rid}", 0.0) or 0.0),
                "Prior Accrual ($)": float(st.session_state.get(f"oa_prior_{_rid}", 0.0) or 0.0),
                "Description": st.session_state.get(f"oa_desc_{_rid}", ""),
                "Auto-Reverse": bool(st.session_state.get(f"oa_autorev_{_rid}", True)),
                "Split Schedule": st.session_state.get(f"oa_split_{_rid}", ""),
                "Compound": bool(st.session_state.get(f"oa_compound_{_rid}", False)),
            }
            for _rid in st.session_state[_OA_IDS_KEY]
        ]
        accruals_edited_df = pd.DataFrame(_oa_out_rows, columns=_OA_COLUMNS)
        st.session_state.manual_accruals_df = accruals_edited_df

        _accrual_active = accruals_edited_df[
            accruals_edited_df["Account Code"].fillna("").str.strip().astype(bool) &
            (accruals_edited_df["Amount ($)"].fillna(0) > 0)
        ]
        if not _accrual_active.empty:
            st.success(
                f"✅ {len(_accrual_active)} accrual(s) queued — "
                f"${_accrual_active['Amount ($)'].sum():,.2f} total debits",
                icon="✅",
            )

    st.divider()

    # ── Pass 1 Run Button ─────────────────────────────────────────────────────
    # Warn if Pass 1 has already been run — re-running after uploading JEs to
    # Yardi would generate duplicate journal entries.
    if st.session_state.pass1_complete:
        _rerun_run_count = st.session_state.get('pass1_run_count', 1)
        st.warning(
            f"⚠️ **Pass 1 has already been run** ({_rerun_run_count}× this session). "
            f"If you've already imported the JE CSVs into Yardi, re-running will generate "
            f"**duplicate journal entries**. Only continue if you have NOT yet uploaded to Yardi "
            f"or if you've reversed/deleted the prior batch.",
            icon="⚠️",
        )

    # ── Year-end / fiscal year-end banner ────────────────────────
    # Detect if close period is the last month of the fiscal year and surface
    # a reminder about year-end-specific items (bonus true-ups, layer 3 reset).
    try:
        _p1_close_period_raw = st.session_state.uploaded_files.get('gl', '')
        # Prefer parsed GL period (from engine result if available) else fall back
        # to the period embedded in any prior pass1 result
        _fy_close_period = ''
        if st.session_state.pass1_engine_result:
            _fy_close_period = getattr(st.session_state.pass1_engine_result, 'period', '') or ''
        _fy_start_mo = int(getattr(_active_cfg, 'fiscal_year_start_month', 1) or 1)
        _fy_end_mo   = (_fy_start_mo - 2) % 12 + 1   # last month of fiscal year
        _fy_mo_map   = dict(Jan=1,Feb=2,Mar=3,Apr=4,May=5,Jun=6,
                            Jul=7,Aug=8,Sep=9,Oct=10,Nov=11,Dec=12)
        _fy_match    = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[- ](\d{4})',
                                 _fy_close_period)
        if _fy_match and _fy_mo_map.get(_fy_match.group(1)) == _fy_end_mo:
            _fy_label = 'December' if _fy_end_mo == 12 else _fy_match.group(1)
            st.info(
                f"🗓️ **Year-End Close ({_fy_close_period})** — this is the last month of the fiscal year. "
                f"Key reminders: (1) Layer 3 historical averages cover all 12 fiscal months — "
                f"review accruals carefully before posting. "
                f"(2) Confirm all bonus true-ups are reflected in the One-Off Accruals table. "
                f"(3) Verify prepaid items expiring this month get their final release.",
                icon="📅",
            )
    except Exception:
        pass

    col_p1a, col_p1b = st.columns([3, 1])
    with col_p1a:
        pass1_button = st.button(
            "🚀 Generate JEs" if not st.session_state.pass1_complete else "🔁 Re-run Pass 1",
            disabled=not gl_uploaded,
            use_container_width=True,
            key="pass1_run_btn",
            help="Parse pre-close GL and generate all accrual JE CSVs for Yardi upload",
            type="secondary" if st.session_state.pass1_complete else "primary",
        )
    with col_p1b:
        if st.button("🔄 Reset Pass 1", use_container_width=True, key="reset_pass1"):
            st.session_state.pass1_complete = False
            st.session_state.pass1_engine_result = None
            st.session_state.pass1_output_files = {}
            st.session_state['pass1_gl_activity_log'] = []
            st.session_state.bulk_overrides_p1 = {}
            st.session_state.upload_key_p1 += 1
            st.session_state.interco_recode_df = pd.DataFrame({
                "Leg": pd.Series([], dtype=str), "Account": pd.Series([], dtype=str),
                "Account Name": pd.Series([], dtype=str),
                "Credit ($)": pd.Series([], dtype=float), "Debit ($)": pd.Series([], dtype=float),
                "Description": pd.Series([], dtype=str),
            })
            st.session_state._interco_seed_gen = st.session_state.get('_interco_seed_gen', 0) + 1
            for _clr in list(st.session_state.uploaded_files.keys()):
                if _clr not in ("gl_pass2", "budget_comparison_pass2",
                                "trial_balance_pass2", "loan_pass2",
                                "prior_workpaper", "t12_statement_pass2"):
                    st.session_state.uploaded_files.pop(_clr, None)
            # Clear Pass 1 close tracker step (step 1 = JEs generated)
            st.session_state.close_tracker.pop(1, None)
            st.session_state.pass1_run_count = 0
            st.rerun()

    # ── Pass 1 Processing ─────────────────────────────────────────────────────
    if pass1_button or st.session_state.pop('_trigger_pass1_rerun', False):
        with st.spinner("Building accrual entries..."):
            try:
                files_dict = {key: st.session_state.uploaded_files.get(key)
                              for key in file_config.keys()}

                # Auto-load committed Kardin budget if not uploaded this session
                if not files_dict.get("kardin_budget") and _COMMITTED_BUDGET:
                    files_dict["kardin_budget"] = _COMMITTED_BUDGET

                progress_bar = st.progress(0)
                status_text  = st.empty()

                # Step 1: Parse files (pre-close GL)
                status_text.text("Step 1/6: Parsing files...")
                progress_bar.progress(10)
                engine_result = run_pipeline(files_dict)
                st.session_state.pass1_engine_result = engine_result

                gl_parsed  = engine_result.parsed.get('gl')
                bc_parsed  = engine_result.parsed.get('budget_comparison') or []
                nexus_data = engine_result.parsed.get('nexus_accrual')
                close_period = engine_result.period or ''

                # Apply any user corrections made in the Prepaid Expense Amortization
                # editor (prior run) before this data reaches merge_nexus() / the
                # amortization schedule / Layer 1 — so a fixed misread actually
                # changes the ledger and JEs, not just the preview table.
                nexus_data = _apply_prepaid_overrides(
                    nexus_data or [], st.session_state.get("prepaid_overrides_df")
                )

                # Guard: if the GL didn't yield a period we cannot label outputs.
                if not close_period:
                    st.error(
                        "⚠️ **Could not determine close period from GL.**  "
                        "The period label (e.g. 'Jan-2026') is required for all "
                        "output file names. Check that the GL file is a valid Yardi "
                        "export and that the header row contains a recognisable date.",
                        icon="❌",
                    )
                    st.session_state.pass1_complete = False
                    st.stop()

                # ── Loan statement date validation ────────────────────────────
                # Interest paid on the 7th covers the PRIOR month.  For the
                # January close, the correct statement is the one due 2/7/2026.
                # Flag a warning when the uploaded statement's payment_due_date
                # month doesn't equal close_period month + 1.
                _loan_stmts = engine_result.parsed.get('loan') or []
                if not isinstance(_loan_stmts, list):
                    _loan_stmts = [_loan_stmts]
                _cp_month_map = dict(Jan=1,Feb=2,Mar=3,Apr=4,May=5,Jun=6,
                                     Jul=7,Aug=8,Sep=9,Oct=10,Nov=11,Dec=12)
                _cp_mo = next(
                    (v for k, v in _cp_month_map.items() if k in close_period), 0
                )
                _expected_due_mo = (_cp_mo % 12) + 1  # Jan→2, Dec→1
                _cp_yr = re.search(r'\d{4}', close_period)
                _cp_yr = int(_cp_yr.group()) if _cp_yr else 0
                _expected_due_yr = _cp_yr + 1 if _cp_mo == 12 else _cp_yr
                for _ls in _loan_stmts:
                    if not isinstance(_ls, dict):
                        continue
                    _due = str(_ls.get('payment_due_date') or '')
                    if not _due:
                        continue
                    _due_parts = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', _due)
                    if not _due_parts:
                        continue
                    _due_mo, _due_yr = int(_due_parts.group(1)), int(_due_parts.group(3))
                    if _due_mo != _expected_due_mo or _due_yr != _expected_due_yr:
                        _loan_num = _ls.get('loan_number', 'unknown')
                        engine_result.exceptions.append(Exception_(
                            severity='warning',
                            category='loan',
                            source='berkadia_loan',
                            description=(
                                f"Loan {_loan_num}: statement due date {_due} does not match "
                                f"the {close_period} close. For {close_period}, upload the "
                                f"statement due {_expected_due_mo:02d}/07/{_expected_due_yr} "
                                f"(interest accrued in {close_period} is paid on the 7th of "
                                f"the following month)."
                            ),
                            details={
                                'loan_number': _loan_num,
                                'uploaded_due_date': _due,
                                'expected_due_month': f'{_expected_due_mo:02d}/{_expected_due_yr}',
                            },
                        ))

                # Step 2: Detect accrual entries (4-layer)
                status_text.text("Step 2/6: Detecting accrual entries (4 layers)...")
                progress_bar.progress(25)

                # Build manual exclusion list (account codes with entries in the One-Off table)
                # so Layers 2-4 don't double-accrue those accounts.
                #
                # IMPORTANT: rows with Amount ($) = $0 are treated as pure exclusion flags —
                # they suppress automated layers for that account WITHOUT generating a JE.
                # Use $0 when a JE has already been posted to Yardi (e.g., RE Taxes posted
                # manually before the pipeline ran) to prevent double-counting.
                _accruals_tbl_early = st.session_state.get("manual_accruals_df")
                _manual_accruals_input = []
                if _accruals_tbl_early is not None and not _accruals_tbl_early.empty:
                    _manual_accruals_input = [
                        {
                            'account_code': str(r["Account Code"]).strip(),
                            'account_name': str(r.get("Account Name", "") or "").strip(),
                            # Force amount=0 so Layer 0 registers the account for dedup
                            # suppression WITHOUT generating a MAN-XXXX JE entry.
                            # The actual SUP-XXXX JEs (with custom CR account support)
                            # are built separately via _supplement_je_lines below.
                            'amount': 0,
                            'prior_accrual': float(r.get("Prior Accrual ($)", 0) or 0),
                            'description': str(r.get("Description", "") or "").strip(),
                        }
                        for _, r in _accruals_tbl_early.iterrows()
                        if str(r.get("Account Code", "") or "").strip()
                        # Only suppress auto-detection (add to _covered) when the user
                        # has actually entered an amount. Pre-seeded rows with Amount=$0
                        # should NOT block Layer 2 from auto-accruing the account.
                        and float(r.get("Amount ($)", 0) or 0) > 0
                    ]

                # Parse T12 for Pass 1 (improves Layer 3 January historical accrual accuracy)
                _t12_file_p1 = st.session_state.uploaded_files.get("t12_statement")
                _t12_result_p1 = None
                if _t12_file_p1 and os.path.exists(_t12_file_p1):
                    try:
                        from parsers.yardi_t12 import parse as parse_t12
                        _t12_result_p1 = parse_t12(_t12_file_p1)
                    except Exception as _e:
                        st.warning(f"Could not parse 12-Month Statement for Pass 1: {_e}")

                # Parse Receivable Summary — highest-priority management fee source.
                # Provides explicit Prepayment row for clean prepayment exclusion.
                _rs_file = st.session_state.uploaded_files.get("receivable_summary")
                _rs_parsed = None
                if _rs_file and os.path.exists(_rs_file):
                    try:
                        from parsers.yardi_receivable_summary import parse as _parse_rs
                        _rs_parsed = _parse_rs(_rs_file)
                    except Exception as _e:
                        _rs_parsed = None
                        st.warning(f"⚠️ Receivable Summary parse failed — management fee will fall back to a less precise source: {_e}")

                # Parse Receivable Detail — alternate management fee source, also used
                # by the accrual engine (Mode b per-tenant electric breakdown).
                _rd_file = st.session_state.uploaded_files.get("receivable_detail")
                _rd_parsed = None
                if _rd_file and os.path.exists(_rd_file):
                    try:
                        from parsers.yardi_receivable_detail import parse as _parse_rd
                        _rd_parsed = _parse_rd(_rd_file)
                    except Exception as _e:
                        _rd_parsed = None
                        st.warning(f"⚠️ Receivable Detail parse failed — management fee will fall back to a less precise source: {_e}")

                # Step 3: Prepaid ledger — load → merge → release lines
                # Ledger contains Prepaid Other (135150) items only.
                # Insurance (135110) and RE Tax (135120) are excluded from the ledger
                # and handled by dedicated amortization functions in build_accrual_entries().
                status_text.text("Step 3/6: Processing prepaid ledger...")
                progress_bar.progress(45)

                ledger_path = st.session_state.uploaded_files.get("prepaid_ledger")
                if not ledger_path:
                    st.warning(
                        "⚠️ **No Prepaid Ledger uploaded** — prepaid amortization releases will be skipped. "
                        "Upload the prior-month `GA_Prepaid_Ledger_Updated.xlsx` to carry forward.",
                        icon=None,
                    )
                ledger_active, ledger_completed, _ledger_load_err = prepaid_ledger.load(ledger_path)
                if _ledger_load_err:
                    st.warning(
                        f"⚠️ **Prior Prepaid Ledger failed to load** — {_ledger_load_err} "
                        f"Every prepaid release JE for this period will be skipped until a "
                        f"readable ledger is uploaded.",
                        icon=None,
                    )

                # Merge Nexus Invoice Detail into ledger
                ledger_active, newly_added = prepaid_ledger.merge_nexus(
                    ledger_active, nexus_data or [], close_period
                )

                # Apply the 7xxxxx Intercompany Recode table to ledger items whose
                # gl_account_number was just recoded — otherwise every future
                # release keeps debiting the original (miscoded) account even
                # though this period's recode JE already moved the balance
                # elsewhere. Built early (a lightweight source→target map only,
                # not the actual JE lines — those are still built later, from
                # the same table) so it lands before merge_nexus()'s items are
                # scheduled and before _ledger_release_accounts below reads
                # gl_account_number. Confirmed on a real invoice: a Nexus item
                # picked up as 712210 (Software-AP) had been manually recoded
                # to a different account this period, but the ledger kept
                # citing 712210 for every future release.
                _early_recode_map = {}
                # Reads the live per-row widget state directly (see
                # _read_interco_df_from_widgets) rather than st.session_state.
                # interco_recode_df, since plain widgets already hold the
                # user's latest typed value the instant they interact —
                # no need to separately peek at pending edits.
                _rc_tbl_early = _read_interco_df_from_widgets()
                if _rc_tbl_early is not None and not _rc_tbl_early.empty:
                    _rc_pending_cr = None
                    for _, _rc_row in _rc_tbl_early.iterrows():
                        _rc_leg = str(_rc_row.get("Leg", "") or "").strip()
                        if _rc_leg == "CR":
                            _rc_pending_cr = _rc_row
                        elif _rc_leg == "DR" and _rc_pending_cr is not None:
                            _rc_dr_acct = str(_rc_row.get("Account", "") or "").strip()
                            _rc_cr_acct = str(_rc_pending_cr.get("Account", "") or "").strip()
                            if _rc_dr_acct and _rc_cr_acct:
                                _rc_dr_name = (str(_rc_row.get("Account Name", "") or "").strip()
                                               or _rc_dr_acct)
                                _early_recode_map[_rc_cr_acct] = (_rc_dr_acct, _rc_dr_name)
                            _rc_pending_cr = None
                if _early_recode_map:
                    _recoded_items = 0
                    for _item in ledger_active:
                        _cur_acct = str(_item.get('gl_account_number', '') or '').strip()
                        if _cur_acct in _early_recode_map:
                            _new_acct, _new_name = _early_recode_map[_cur_acct]
                            _item['gl_account_number'] = _new_acct
                            _item['gl_account']         = _new_name
                            _recoded_items += 1
                    if _recoded_items:
                        st.info(
                            f"↳ Prepaid ledger: {_recoded_items} item(s) updated to the recoded "
                            f"GL account per the Intercompany Recode table — future releases will "
                            f"debit the corrected account.",
                            icon="ℹ️",
                        )

                # Build visual amortization schedule
                amort_lines = build_prepaid_amortization(nexus_data or [], close_period=close_period)

                # Phase-1 release scan — used only to build _ledger_release_accounts
                # for build_accrual_entries().  This call does NOT yet know which
                # newly-added Nexus invoices were suppressed, so it conservatively
                # skips month-1 for all new items.  The definitive ledger_release_lines
                # (used for JE generation) is computed in Phase 2 below, after we
                # know which Nexus JEs were actually emitted.
                _pre_release_lines = prepaid_ledger.get_current_amortization(ledger_active, close_period)
                _ledger_release_accounts = {
                    str(item.get('gl_account_number', '')).strip()
                    for item in _pre_release_lines
                    if item.get('gl_account_number')
                }

                _gl_activity_log = []
                import warnings as _warnings_mod
                with _warnings_mod.catch_warnings(record=True) as _captured_warnings:
                    _warnings_mod.simplefilter("always")
                    je_lines = build_accrual_entries(
                        nexus_data or [],
                        period=close_period,
                        property_name=engine_result.property_name or '',
                        gl_data=gl_parsed,
                        budget_data=bc_parsed,
                        manual_accruals=_manual_accruals_input or [],
                        tenant_utility_rows=_tenant_utility_rows or None,
                        loan_data=engine_result.parsed.get('loan'),
                        re_tax_bill_amount=_re_tax_bill_amount,
                        re_tax_payment_months=getattr(_active_cfg, 're_tax_payment_months', None) or [1, 4, 7, 10],
                        bonus_overrides=_bonus_overrides or None,
                        kardin_records=engine_result.parsed.get('kardin_budget') or None,
                        t12_result=_t12_result_p1,
                        gl_activity_log=_gl_activity_log,
                        receivable_detail=_rd_parsed,
                        ledger_release_accounts=_ledger_release_accounts,
                        payroll_accounts=getattr(_active_cfg, 'payroll_accounts', None) or None,
                        insurance_policies=getattr(_active_cfg, 'insurance_policies', None) or None,
                        periodic_contract_accounts=getattr(_active_cfg, 'periodic_contract_accounts', None) or None,
                        metered_utility_accounts=getattr(_active_cfg, 'metered_utility_accounts', None) or None,
                        per_invoice_utility_accounts=getattr(_active_cfg, 'per_invoice_utility_accounts', None) or None,
                        per_invoice_accrual_accounts=getattr(_active_cfg, 'per_invoice_accrual_accounts', None) or None,
                        accrual_materiality_floor=getattr(_active_cfg, 'accrual_materiality_floor', 2500.0),
                        fiscal_year_start_month=getattr(_active_cfg, 'fiscal_year_start_month', 1) or 1,
                        layer3_exclude_accounts=getattr(_active_cfg, 'layer3_exclude_accounts', None) or None,
                    )
                # Surface any pipeline UserWarnings (e.g. missing Berkadia RE tax entry) in the UI
                for _w in _captured_warnings:
                    if issubclass(_w.category, UserWarning):
                        st.warning(str(_w.message), icon="⚠️")
                st.session_state['pass1_gl_activity_log'] = _gl_activity_log

                # Budget-based accruals for HVAC / Fire Life Safety / Snow & Ice —
                # these no longer live in the One-Off Accruals table (see config.yaml
                # comment). Skips any account already claimed by an earlier layer
                # (a real invoice came through Nexus/GL normally) or already covered
                # by real GL activity; otherwise accrues from the Kardin monthly
                # budget. HVAC/Fire Life Safety additionally split a separate
                # quarterly-service-invoice line in quarter-end months.
                budget_based_je, _budget_review_flags = build_budget_based_accruals(
                    je_lines, gl_parsed, engine_result.parsed.get('kardin_budget') or [],
                    period=close_period, je_start=1,
                )
                st.session_state['pass1_budget_review_flags'] = _budget_review_flags

                # Phase-2 release scan — now that we know which Nexus JEs fired,
                # determine which newly-added prepaid invoice numbers were suppressed
                # (expense already in GL, invoice deduplicated).  For those items,
                # month-1 must be emitted by the prepaid ledger instead of being
                # silently skipped, otherwise the expense is permanently lost.
                _emitted_nexus_invs = {
                    str(l.get('invoice_number', '') or '').strip().lower()
                    for l in je_lines
                    if l.get('source') == 'nexus' and l.get('invoice_number')
                }
                _suppressed_prepaid_invs = {
                    inv.strip().lower()
                    for inv in (newly_added or [])
                    if inv.strip().lower() and inv.strip().lower() not in _emitted_nexus_invs
                } or None

                # Reclass newly-discovered multi-month prepaid invoices: the full
                # invoice amount hits the expense account when Nexus processes it,
                # and needs a one-time DR Prepaid / CR expense JE moving everything
                # except the amount already correctly expensed. Skips generating a
                # duplicate when the accountant already posted a manual reclass —
                # detected by searching the GL for a real (non-pipeline) "reclass"
                # entry that references this item's own description.
                prepaid_reclass_je, _reclassed_invs = build_prepaid_reclass_je(
                    ledger_active, newly_added or [], gl_parsed,
                    period=close_period, je_start=len(je_lines) // 2 + 1,
                )
                if _suppressed_prepaid_invs:
                    _suppressed_prepaid_invs = _suppressed_prepaid_invs - _reclassed_invs
                if prepaid_reclass_je:
                    st.info(
                        f"↳ Reclassed {len(prepaid_reclass_je) // 2} newly-discovered prepaid "
                        f"invoice(s) to Prepaid Other — verify the accounts/amounts before uploading.",
                        icon="ℹ️",
                    )
                _already_reclassed_invs = _reclassed_invs - {
                    str(l.get('invoice_number', '') or '').strip().lower()
                    for l in prepaid_reclass_je
                }
                if _already_reclassed_invs:
                    st.info(
                        f"↳ {len(_already_reclassed_invs)} newly-discovered prepaid invoice(s) "
                        f"already have a manual reclass posted in the GL — no reclass JE generated "
                        f"for those; still included in the prepaid ledger export.",
                        icon="ℹ️",
                    )

                import warnings as _warnings_prepaid
                with _warnings_prepaid.catch_warnings(record=True) as _prepaid_warns:
                    _warnings_prepaid.simplefilter("always")
                    ledger_release_lines = prepaid_ledger.get_current_amortization(
                        ledger_active, close_period,
                        suppressed_invoice_numbers=_suppressed_prepaid_invs,
                        gl_data=gl_parsed,
                    )
                for _pw in _prepaid_warns:
                    if issubclass(_pw.category, UserWarning):
                        st.warning(str(_pw.message), icon="⚠️")

                # Build prepaid release JEs after je_lines/reclass so JE numbers are sequential
                prepaid_release_je = build_prepaid_release_je(
                    ledger_release_lines,
                    period=close_period,
                    je_start=len(je_lines) // 2 + len(prepaid_reclass_je) // 2 + 1,
                )

                # Advance ledger (increment months_amortized, expire completed)
                ledger_active, ledger_completed = prepaid_ledger.advance_period(
                    ledger_active, ledger_completed, close_period
                )

                updated_ledger_path = os.path.join(
                    st.session_state.temp_dir, f"{_pfx_int}_Prepaid_Ledger_Updated.xlsx"
                )
                prepaid_ledger.save(ledger_active, ledger_completed, updated_ledger_path,
                                    period=close_period, property_name=_prop_display)

                # Step 4: Management fee (JE included in accruals CSV)
                status_text.text("Step 4/6: Calculating management fee...")
                progress_bar.progress(60)

                _daca_file = st.session_state.uploaded_files.get("daca_bank")
                _daca_parsed = None
                if _daca_file and os.path.exists(_daca_file):
                    try:
                        from parsers.yardi_daca_rec import (
                            is_yardi_daca_rec as _is_yardi_daca,
                            parse as _parse_yardi_daca,
                        )
                        from parsers.keybank_daca import parse as _parse_daca
                        if _is_yardi_daca(_daca_file):
                            _daca_parsed = _parse_yardi_daca(_daca_file)
                        else:
                            _daca_parsed = _parse_daca(_daca_file)
                    except Exception:
                        _daca_parsed = None

                # _rd_parsed already parsed above (before build_accrual_entries)

                _ar_aging_file = st.session_state.uploaded_files.get("ar_aging")
                _ar_aging_parsed = None
                if _ar_aging_file and os.path.exists(_ar_aging_file):
                    try:
                        from parsers.yardi_ar_aging import parse as _parse_ar_aging
                        _ar_aging_parsed = _parse_ar_aging(_ar_aging_file)
                    except Exception:
                        _ar_aging_parsed = None

                import warnings as _warnings_fee
                with _warnings_fee.catch_warnings(record=True) as _fee_warns:
                    _warnings_fee.simplefilter("always")
                    fee_result = calculate_mgmt_fee(
                        gl_parsed=gl_parsed,
                        budget_rows=bc_parsed or [],
                        daca_parsed=_daca_parsed,
                        receivable_summary=_rs_parsed,
                        receivable_detail=_rd_parsed,
                        ar_aging=_ar_aging_parsed,
                    )
                    fee_je = build_management_fee_je(
                        fee_result,
                        period=close_period,
                        property_code=engine_result.property_name or _active_cfg.property_code,
                        je_number=f'MGT-{len(je_lines)//2 + 1:03d}',
                        property_config=_active_cfg,
                    )
                for _fw in _fee_warns:
                    if issubclass(_fw.category, UserWarning):
                        st.warning(str(_fw.message), icon="⚠️")

                _catchup_amount = detect_prior_period_catchup(
                    gl_parsed,
                    mgmt_fee_account=_active_cfg.gl_account('mgmt_fee_expense', '637130') if _active_cfg else '637130',
                )
                _catchup_je = []
                if _catchup_amount and _catchup_amount > 0:
                    _catchup_je = build_catchup_je(
                        _catchup_amount,
                        period=close_period,
                        property_code=engine_result.property_name or _active_cfg.property_code,
                        je_number=f'MGT-{len(je_lines)//2 + 2:03d}',
                        property_config=_active_cfg,
                    )

                # Step 5: One-Off Accrual JEs
                status_text.text("Step 5/6: Building one-off accrual entries...")
                progress_bar.progress(75)

                # One-Off Accruals → DR expense / CR 213100 (or custom CR Account if specified)
                _supplement_je_lines = []
                _periodic_supplement_rows = []
                _sup_base = (len(je_lines) // 2 + len(prepaid_reclass_je) // 2
                             + len(prepaid_release_je) // 2 + len(fee_je) // 2
                             + len(budget_based_je) // 2)

                _accruals_tbl = st.session_state.get("manual_accruals_df")
                if _accruals_tbl is not None and not _accruals_tbl.empty:
                    _active_accruals = _accruals_tbl[
                        _accruals_tbl["Account Code"].fillna("").str.strip().astype(bool) &
                        (_accruals_tbl["Amount ($)"].fillna(0) > 0)
                    ]
                    _CR_ACCT_NAMES = {
                        '115200': 'RE Tax Escrow',
                        '115300': 'Insurance Escrow',
                        '133110': 'Tenant AR Billback',
                        '135150': 'Prepaids',
                        '213100': 'Accrued Expenses',
                    }
                    for _, _row in _active_accruals.iterrows():
                        _row_acct_code = str(_row["Account Code"]).strip()
                        if _row_acct_code in BUDGET_BASED_ACCOUNTS:
                            st.warning(
                                f"Account {_row_acct_code} ({BUDGET_BASED_ACCOUNTS[_row_acct_code]['label']}) "
                                f"is now accrued automatically from budget — see the Budget-Based Accrual "
                                f"section below. Remove it from the One-Off Accruals table to avoid a "
                                f"potential duplicate JE.",
                                icon="⚠️",
                            )
                            continue
                        _vendor = str(_row.get("Vendor", "") or "").strip()
                        _desc   = str(_row.get("Description", "") or "").strip()
                        _split_sch_override = str(_row.get("Split Schedule", "") or "").strip()
                        _row_auto_rev = bool(_row.get("Auto-Reverse", True))
                        _row_compound = bool(_row.get("Compound", True))
                        _periodic_supplement_rows.append({
                            'account_code':    str(_row["Account Code"]).strip(),
                            'account_name':    str(_row.get("Account Name", "") or "").strip()
                                               or str(_row["Account Code"]).strip(),
                            'amount':          float(_row["Amount ($)"]),
                            'prior_accrual':   float(_row.get("Prior Accrual ($)", 0) or 0),
                            'compound':        _row_compound,
                            'description':     _desc or _vendor or 'one-off accrual',
                            'vendor':          _vendor,
                            'auto_reverse':    _row_auto_rev,
                            'cr_account':      '213100',
                            'cr_account_name': 'Accrued Expenses',
                            '_split_schedule': _split_sch_override,  # '' = use property default
                        })

                # Build a GL account lookup for compound accrual logic.
                # Stores the GL account object per account_code so we can read
                # J-type credits (prior-month auto-reversal) and non-J net change
                # (real K/P/C invoice activity) for each one-off accrual account.
                _sup_gl_accts: dict = {}
                if gl_parsed and hasattr(gl_parsed, 'accounts'):
                    for _sga in gl_parsed.accounts:
                        _sup_gl_accts[str(_sga.account_code).strip()] = _sga

                def _sup_j_credits(acct_obj) -> float:
                    """J-type credit total (auto-reversals of prior pipeline JEs)."""
                    if acct_obj is None:
                        return 0.0
                    return sum(
                        t.credit for t in getattr(acct_obj, 'transactions', [])
                        if t.credit > 0
                        and str(getattr(t, 'control', '') or '').upper().startswith('J')
                    )

                def _sup_j_debits(acct_obj) -> float:
                    """J-type debit total (pipeline JEs already posted this period)."""
                    if acct_obj is None:
                        return 0.0
                    return sum(
                        t.debit for t in getattr(acct_obj, 'transactions', [])
                        if t.debit > 0
                        and str(getattr(t, 'control', '') or '').upper().startswith('J')
                    )

                _sup_counter = 0
                for _sup in _periodic_supplement_rows:
                    _sup_acct_code = _sup['account_code']
                    _sup_monthly   = round(float(_sup['amount']), 2)   # user-entered monthly rate

                    # ── Compound accrual logic ────────────────────────────────────
                    # If the prior month's pipeline accrual auto-reversed (J-credit in
                    # current GL), compound it with the user-entered monthly amount so
                    # the liability builds correctly: cumulative = reversal + new month.
                    #
                    # The old "real-invoice guard" that suppressed user-entered accruals
                    # when GL non-J activity >= monthly amount has been REMOVED.  The user
                    # explicitly chose to accrue these items — silently dropping them caused
                    # missed accruals.  If a real invoice is already in the GL the user
                    # can simply leave the Amount at $0 to suppress without generating a JE.
                    _sga_obj        = _sup_gl_accts.get(_sup_acct_code)
                    _sga_j_cr       = _sup_j_credits(_sga_obj)
                    _sup_prior_seed = float(_sup.get('prior_accrual', 0) or 0)
                    _sup_do_compound = bool(_sup.get('compound', True))

                    # If no J-credits from a prior pipeline auto-reversal (e.g. first
                    # pipeline run), fall back to the user-entered Prior Accrual ($) to
                    # seed compounding. Once pipeline auto-reversals appear in the GL,
                    # _sga_j_cr will exceed zero and the manual seed is no longer needed.
                    _sga_effective_j_cr = _sga_j_cr if _sga_j_cr > 0 else _sup_prior_seed
                    _prior_seed_note    = ' [prior accrual seeded manually]' if (_sga_j_cr == 0 and _sup_prior_seed > 0) else ''

                    # Compounding only makes sense for accounts billed irregularly
                    # (e.g. Water/Sewer every 6 months), where the accrued liability
                    # needs to keep building until a real invoice finally clears it —
                    # the prior reversal nets against it so the true monthly P&L hit
                    # stays flat. For a normally monthly-billed account, the auto-
                    # reversal is just the expected mechanical reversal of last
                    # month's own JE — adding it back on top double-counts against
                    # the real invoice landing that same month. Uncheck "Compound"
                    # for those accounts to keep this row a flat monthly amount.
                    if _sup_do_compound:
                        _sup_compound   = _sga_effective_j_cr + _sup_monthly
                        _sup_cmpd_note  = (f' — cumulative ${_sup_compound:,.0f} '
                                           f'(${_sga_effective_j_cr:,.0f} prior reversal + ${_sup_monthly:,.0f}/mo)'
                                           f'{_prior_seed_note}'
                                           if _sga_effective_j_cr > 0 else '')
                    else:
                        _sup_compound  = _sup_monthly
                        _sup_cmpd_note = ''

                    _sje_id  = f'SUP-{_sup_base + _sup_counter + 1:04d}'
                    _sup_counter += 1
                    _sup_desc   = (_sup.get('description') or f"{_sup['account_name']} — one-off accrual") + _sup_cmpd_note
                    _sup_vendor = _sup.get('vendor') or _sup['account_name']
                    _sup_cr_acct = _sup.get('cr_account', '213100')
                    _sup_cr_name = _sup.get('cr_account_name', 'Accrued Expenses')
                    _sup_amt     = round(_sup_compound, 2)
                    _sup_split_sch = _sup.get('_split_schedule', '')
                    # Auto-reverse: -1 = Yardi auto-reverses next month; 0 = permanent JE
                    _sup_rev_flag = -1 if _sup.get('auto_reverse', True) else 0
                    _supplement_je_lines.extend([
                        {
                            'je_number': _sje_id, 'line': 1, 'date': close_period,
                            'account_code': _sup['account_code'], 'account_name': _sup['account_name'],
                            'description': _sup_desc, 'reference': 'ONE-OFF-ACCRUAL',
                            'debit': _sup_amt, 'credit': 0, 'vendor': _sup_vendor,
                            'invoice_number': '', 'source': 'contract_supplement', 'confidence': 'high',
                            '_split_schedule': _sup_split_sch,
                            'reverse_next_month': _sup_rev_flag,
                        },
                        {
                            'je_number': _sje_id, 'line': 2, 'date': close_period,
                            'account_code': _sup_cr_acct, 'account_name': _sup_cr_name,
                            'description': _sup_desc, 'reference': 'ONE-OFF-ACCRUAL',
                            'debit': 0, 'credit': _sup_amt, 'vendor': _sup_vendor,
                            'invoice_number': '', 'source': 'contract_supplement', 'confidence': 'high',
                            '_split_schedule': _sup_split_sch,
                            'reverse_next_month': _sup_rev_flag,
                        },
                    ])

                # ── 7xxxxx Intercompany Recode JEs ────────────────────────────
                # Walks the recode table in order: CR row → DR row pairs.
                # DR [target 6/8xxxxx expense account] / CR [7xxxxx account]
                # Permanent (no auto-reverse) — the recode is a permanent reclassification.
                _recode_je_lines = []
                # See _read_interco_df_from_widgets — reads live widget state
                # directly so a DR account typed just before clicking Re-run
                # (without tabbing out first) is still picked up.
                _recode_tbl = _read_interco_df_from_widgets()
                if _recode_tbl is not None and not _recode_tbl.empty:
                    _recode_base = _sup_base + _sup_counter
                    _recode_ri   = 0
                    _pending_cr  = None   # CR row waiting for its DR partner
                    for _, _row in _recode_tbl.iterrows():
                        _row_leg = str(_row.get("Leg", "") or "").strip()
                        if _row_leg == "CR":
                            _pending_cr = _row   # store CR; will pair with next DR
                        elif _row_leg == "DR" and _pending_cr is not None:
                            _dr_acct = str(_row.get("Account", "") or "").strip()
                            if not _dr_acct:
                                _pending_cr = None
                                continue     # DR account not filled in yet — skip
                            _cr_acct = str(_pending_cr.get("Account", "") or "").strip()
                            _cr_name = str(_pending_cr.get("Account Name", "") or "").strip() or _cr_acct
                            _cr_amt  = float(_pending_cr.get("Credit ($)", 0) or 0)
                            _dr_amt  = float(_row.get("Debit ($)", 0) or 0) or _cr_amt
                            _r_desc  = (str(_row.get("Description", "") or "").strip()
                                        or f"Recode {_cr_acct} → {_dr_acct}")
                            _r_je_id = f'REC-{_recode_base + _recode_ri + 1:04d}'
                            _dr_name = (_acct_name_lookup.get(_dr_acct)
                                        or _sup_gl_accts.get(_dr_acct, {})
                                        and getattr(_sup_gl_accts.get(_dr_acct), 'account_name', '')
                                        or _dr_acct)
                            _recode_je_lines.extend([
                                {
                                    'je_number': _r_je_id, 'line': 1, 'date': close_period,
                                    'account_code': _dr_acct, 'account_name': _dr_name,
                                    'description': _r_desc, 'reference': 'INTERCO-RECODE',
                                    'debit': _dr_amt, 'credit': 0, 'vendor': '',
                                    'invoice_number': '', 'source': 'interco_recode',
                                    'confidence': 'high', 'reverse_next_month': 0,
                                },
                                {
                                    'je_number': _r_je_id, 'line': 2, 'date': close_period,
                                    'account_code': _cr_acct, 'account_name': _cr_name,
                                    'description': _r_desc, 'reference': 'INTERCO-RECODE',
                                    'debit': 0, 'credit': _cr_amt, 'vendor': '',
                                    'invoice_number': '', 'source': 'interco_recode',
                                    'confidence': 'high', 'reverse_next_month': 0,
                                },
                            ])
                            _recode_ri  += 1
                            _pending_cr  = None   # consumed

                # Step 6: Assemble all JEs, apply building splits, export 3 CSVs
                status_text.text("Step 6/6: Exporting JE CSVs...")
                progress_bar.progress(88)

                all_je_lines = (
                    je_lines
                    + prepaid_reclass_je
                    + prepaid_release_je
                    + fee_je
                    + _catchup_je
                    + _supplement_je_lines
                    + _recode_je_lines
                    + budget_based_je
                )

                # Apply pro-rata building splits for multi-building properties
                if _active_cfg.is_multi_building:
                    from building_splits_engine import apply_building_splits as _apply_splits
                    all_je_lines = _apply_splits(all_je_lines, _active_cfg)

                # Preserve manually-added JEs ("Add a Missed Entry") across this
                # re-run — they aren't regenerated by the pipeline, so without
                # this a re-run would silently discard them every time.
                _prior_manual_adds = [
                    _l for _l in st.session_state.pass1_output_files.get("all_je_lines", [])
                    if _l.get('source') == 'manual_addition'
                ]
                if _prior_manual_adds:
                    all_je_lines = all_je_lines + _prior_manual_adds

                _accrual_csv_path = None

                _prop_code = (engine_result.parsed.get('gl') and
                              engine_result.parsed['gl'].metadata.property_code) or _active_cfg.property_code
                # ETL PROPERTY field: max 8 chars. Use yardi_etl_code if configured.
                _etl_code = (getattr(_active_cfg, 'yardi_etl_code', '') or _prop_code)[:8]

                if all_je_lines:
                    _accrual_csv_path = os.path.join(st.session_state.temp_dir, f"{_pfx_int}_Accruals_JE.csv")
                    generate_etl_csv(all_je_lines, _accrual_csv_path,
                                     period=close_period, property_code=_etl_code,
                                     auto_reverse=True)

                # Persist Pass 1 outputs
                p1 = st.session_state.pass1_output_files
                p1["all_je_lines"]          = all_je_lines
                # Serialize to JSON so Pass 2 can reload it in a fresh browser session
                try:
                    import json as _json
                    _je_cache_path = os.path.join(st.session_state.temp_dir, f"{_pfx_int}_JE_Cache.json")
                    with open(_je_cache_path, 'w') as _jcf:
                        _json.dump(all_je_lines, _jcf)
                    p1["je_lines_cache"] = _je_cache_path
                except Exception:
                    p1["je_lines_cache"] = None
                p1["accrual_je_csv"]        = _accrual_csv_path
                p1["fee_result"]            = fee_result
                p1["rd_prepayment_amount"]  = getattr(fee_result, 'prepayment_excluded', 0.0)
                p1["catchup_amount"]        = _catchup_amount
                p1["amort_lines"]           = amort_lines
                p1["ledger_active"]         = ledger_active
                p1["ledger_completed"]      = ledger_completed
                p1["newly_added_prepaids"]  = newly_added
                p1["prepaid_ledger_updated"]= updated_ledger_path
                p1["prepaid_released_count"]= len(prepaid_release_je) // 2
                p1["prepaid_release_lines"] = ledger_release_lines
                p1["budget_review_flags"]   = _budget_review_flags
                p1["close_period"]          = close_period

                progress_bar.progress(100)
                status_text.text("✓ JEs ready for Yardi upload!")
                st.session_state.pass1_complete = True
                st.session_state.pass1_run_count = st.session_state.get('pass1_run_count', 0) + 1

                # ── Auto-detect Close Tracker Step 1 ─────────────────────────
                _ct = st.session_state.close_tracker
                if 1 not in _ct:
                    _ct[1] = {
                        "completed_by": st.session_state.get('prepared_by', 'GRP'),
                        "timestamp":    datetime.now().strftime("%m/%d/%Y %H:%M"),
                        "auto":         True,
                    }
                    _save_checklist_now()
                    st.session_state.last_completed_step = 1

                # ── Pass 1 Run Log ────────────────────────────────────────────
                try:
                    from run_log import append_run_log_pass1 as _append_p1_log
                    _p1_rl_path  = os.path.join(st.session_state.temp_dir, "GA_Run_Log.csv")
                    _p1_rl_prior = st.session_state.uploaded_files.get("run_log")
                    _p1_je_count = len(all_je_lines) // 2 if all_je_lines else 0
                    _p1_je_total = sum(float(l.get('debit', 0)) for l in (all_je_lines or []))
                    _append_p1_log(
                        output_path         = _p1_rl_path,
                        prior_log_path      = _p1_rl_prior,
                        timestamp           = datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        prepared_by         = st.session_state.get('prepared_by', 'GRP'),
                        property_name       = engine_result.property_name or _prop_display,
                        period              = close_period,
                        je_count            = _p1_je_count,
                        je_total_dollars    = _p1_je_total,
                        close_tracker_complete = (
                            len(st.session_state.close_tracker) == 9
                        ),
                    )
                    st.session_state.pass1_output_files["run_log"] = _p1_rl_path
                except Exception as _p1_rle:
                    pass   # run log is non-critical

                st.success("Pass 1 complete! Download the JE CSVs below.", icon="✅")

            except Exception as e:
                tb = traceback.format_exc()
                st.error(f"Pass 1 error: {str(e)}", icon="❌")
                st.code(tb, language="python")
                st.session_state.pass1_complete = False

    # ── Pass 1 Results Dashboard ──────────────────────────────────────────────
    if st.session_state.pass1_complete and st.session_state.pass1_engine_result:
        result = st.session_state.pass1_engine_result
        p1     = st.session_state.pass1_output_files
        all_je_lines = p1.get("all_je_lines", [])
        fee_result   = p1.get("fee_result")

        st.divider()
        st.markdown(f"## Pass 1 Results — {result.period}  |  {result.property_name}")


        # ── GL Activity Gut-Check ──────────────────────────────────────────
        # Show a compact warning listing accounts where the pipeline detected
        # an existing JE in the GL and suppressed automated accruals.
        _gl_log = st.session_state.get('pass1_gl_activity_log') or []
        _gl_log_real = [r for r in _gl_log if r.get('account_code') != 'TUB-MODE-B']
        if _gl_log_real:
            _gl_log_sorted = sorted(_gl_log_real, key=lambda x: x['account_code'])
            _acct_list = ', '.join(
                f"{r['account_code']} {r['account_name']}"
                for r in _gl_log_sorted
            )
            st.warning(
                f"⚠️  **Existing GL postings — accruals suppressed for "
                f"{len(_gl_log_sorted)} account{'s' if len(_gl_log_sorted) != 1 else ''}:** "
                f"{_acct_list}. "
                f"Confirm each posting is intentional before uploading JE CSVs to Yardi. "
                f"If a posting is incorrect, delete it from Yardi and re-run Pass 1."
            )
        # ── Management Fee ─────────────────────────────────────────────────
        if fee_result and fee_result.cash_received > 0:
            st.markdown("### Management Fee JE")
            _src_labels = {
                'receivable_summary':          'Receivable Summary (ex-Prepayments)',
                'receivable_detail+ar_aging':  'Receivable Detail (ex-Prepayments via AR Aging)',
                'receivable_detail':           'Receivable Detail (ex-Prepayments)',
                'daca_additions':              'DACA Additions',
                'gl_cash_account':             'GL 111100 Debits',
                'revenue_proxy':               'Revenue Proxy',
                'manual_override':   'Manual Override',
            }
            _src_label = _src_labels.get(fee_result.cash_source, fee_result.cash_source)
            col_f1, col_f2, col_f3, col_f4 = st.columns(4)
            with col_f1:
                st.metric("Cash Received", f"${fee_result.cash_received:,.0f}",
                          help=f"Source: {_src_label}")
            with col_f2:
                st.metric(f"JLL ({fee_result.jll_rate:.2%})", f"${fee_result.jll_fee:,.0f}")
            with col_f3:
                st.metric(f"GRP ({fee_result.grp_rate:.2%})", f"${fee_result.grp_fee:,.0f}")
            with col_f4:
                st.metric(f"Total ({fee_result.total_rate:.2%})", f"${fee_result.total_fee:,.0f}")
            _prepay_amt = p1.get("rd_prepayment_amount", 0.0) or 0.0
            st.caption(f"Basis: {_src_label}"
                       + (f"  ·  Prepayments/Billbacks excluded: ${_prepay_amt:,.2f}" if _prepay_amt > 0 else ""))

            _catchup_amt = p1.get("catchup_amount")
            if _catchup_amt and _catchup_amt > 0:
                st.warning(
                    f"**Management Fee Catch-up Detected — ${_catchup_amt:,.2f}**\n\n"
                    f"Account 637130 shows a net credit (auto-reversal of prior month accrual with "
                    f"no matching invoice). A catch-up entry **(MGT-002)** has been included in the "
                    f"Accruals CSV. **Verify before posting:** confirm the prior month check is still "
                    f"outstanding in AP (213100) before uploading.",
                )
            st.divider()

        # ── JE Preview — Grouped by Credit (BS) Account ───────────────────
        dr_lines = [l for l in all_je_lines if (l.get('debit') or 0) > 0]
        cr_lines = [l for l in all_je_lines if (l.get('credit') or 0) > 0]

        if dr_lines:
            st.markdown("### Accrual Journal Entries")
            st.caption(
                "JE reference prefixes: **IPR** = Invoice Proration · **REC** = Historical Recurring · "
                "**MGT** = Management Fee · **TUB** = Tenant Utility Billing · **SUP** = One-Off Accrual · "
                "**BNS** = Bonus Accrual. "
                "These codes are required for the Yardi import. "
                "✏️ **Description is editable** — click any cell to update the text. "
                "Edited descriptions are written to the downloaded CSV."
            )

            # ── Source → uploaded file label map ────────────────────────────
            _SOURCE_FILE_LABEL = {
                'nexus':                  'Nexus Invoice Detail',
                'invoice_proration':      'Yardi GL Detail',
                'historical':             'Yardi Budget Comparison',
                'prepaid_amortization':   'Prior Month Prepaid Ledger',
                'prepaid_ledger':         'Prior Month Prepaid Ledger',
                'management_fee':         'Management Fee',
                'management_fee_catchup': 'Management Fee (Catch-up)',
                'contract_supplement':    'One-Off Accrual Table',
                'tenant_utility_billing': 'Tenant Utility Billing',
                'bonus_accrual':          'Kardin Budget',
                'manual_addition':        'Manually Added',
            }

            # ── Description cleaner — strip amounts / verbose phrases ────────
            import re as _re_jed
            def _clean_je_desc(raw: str) -> str:
                s = str(raw)
                s = _re_jed.sub(r'\$[\d,]+(?:\.\d+)?(?:/\w+)?', '', s)   # $1,234.56/mo
                s = _re_jed.sub(r'\b[\d,]+\.\d+/\w+', '', s)             # 500.00/day
                s = _re_jed.sub(r'\([A-Za-z]+-?\s*\d{4}\)', '', s)       # (Jan-2026)
                s = _re_jed.sub(r'Historical recurring\s*[—-]\s*', '', s)
                s = _re_jed.sub(r'[Pp]roration\s*[—-]\s*', '', s)
                s = _re_jed.sub(r',?\s*no activity this period.*', '', s, flags=_re_jed.IGNORECASE)
                s = _re_jed.sub(r',?\s*no T12 uploaded.*', '', s, flags=_re_jed.IGNORECASE)
                s = _re_jed.sub(r',?\s*upload for.*', '', s, flags=_re_jed.IGNORECASE)
                s = _re_jed.sub(r'\(annual budget[^)]*\)', '', s, flags=_re_jed.IGNORECASE)
                s = _re_jed.sub(r'×\s*\d+\s*days?', '', s)
                s = _re_jed.sub(r'÷\s*\d+', '', s)
                s = _re_jed.sub(r'\s+', ' ', s).strip().strip('—:,- ').strip()
                return s[:120] if len(s) > 120 else s

            # ── Build CR lookup: je_number → cr_account_code ────────────────
            _cr_lookup: dict = {}
            for _cl in cr_lines:
                _je = _cl.get('je_number', '')
                if _je and _je not in _cr_lookup:
                    _cr_lookup[_je] = {
                        'code': str(_cl.get('account_code', '') or '').strip(),
                    }

            # Friendly labels for well-known CR accounts
            _CR_LABELS = {
                '115200': 'RE Tax Escrow',
                '115300': 'Insurance Escrow',
                '133110': 'Tenant AR Billback',
                '135110': 'Prepaid Insurance',
                '135120': 'Prepaid RE Taxes',
                '135150': 'Prepaids',
                '213100': 'Accrued Expenses',
                '213200': 'Accrued Interest Payable',
                '440500': 'Recovery - Electricity (Tenant Billing)',
                '440700': 'Recovery - Misc Utilities (Tenant Billing)',
                '613110': 'Utilities - Electricity (P&L Reclass)',
                '641110': 'Real Estate Taxes (deferral)',
            }
            def _cr_section_label(code: str) -> str:
                if code in _CR_LABELS:
                    return f"{code} — {_CR_LABELS[code]}"
                if code.startswith('115'):  return f"{code} — Escrow"
                if code.startswith('133'):  return f"{code} — Tenant AR Billback"
                if code.startswith('135'):  return f"{code} — Prepaids"
                if code.startswith('213'):  return f"{code} — Accrued"
                if code.startswith('44'):   return f"{code} — Revenue Recovery"
                if code.startswith('61'):   return f"{code} — Expense Reclassification"
                return f"{code}"

            # ── Group DR lines by CR account ────────────────────────────────
            _groups: dict = {}
            for _dl in dr_lines:
                _cr_code = _cr_lookup.get(_dl.get('je_number', ''), {}).get('code', 'unknown')
                _groups.setdefault(_cr_code, []).append(_dl)

            _sorted_cr_codes = sorted(_groups.keys(), key=lambda c: c.zfill(10))

            _run_key = st.session_state.get('pass1_run_count', 0)

            # ── Summary metrics row ──────────────────────────────────────────
            _excl_set_disp  = st.session_state.get('je_excluded_jes', set())
            _amt_ovr_disp   = st.session_state.get('je_amount_overrides', {})
            _dr_incl = [_l for _l in dr_lines if _l.get('je_number', '') not in _excl_set_disp]
            _src_totals: dict = {}
            for _l in _dr_incl:
                _s   = _l.get('source', 'other')
                _amt = _amt_ovr_disp.get(_l.get('je_number', ''), _l.get('debit') or 0)
                _src_totals[_s] = _src_totals.get(_s, 0) + _amt
            _total_je_count = len(set(_l.get('je_number', '') for _l in _dr_incl))
            _total_amount   = sum(
                _amt_ovr_disp.get(_l.get('je_number', ''), _l.get('debit') or 0)
                for _l in _dr_incl
            )
            _excl_je_count  = len(_excl_set_disp)
            _amt_ovr_count  = len(_amt_ovr_disp)
            _metric_items = [('Included JEs', str(_total_je_count)),
                             ('Included Amount', f"${_total_amount:,.0f}")] + \
                            [(_SOURCE_FILE_LABEL.get(s, s), f"${t:,.0f}")
                             for s, t in _src_totals.items()]
            _n_cols = min(len(_metric_items), 6)
            _metric_cols = st.columns(_n_cols)
            for _mi, (_lbl, _val) in enumerate(_metric_items[:_n_cols]):
                with _metric_cols[_mi]:
                    st.metric(_lbl, _val)
            _status_notes = []
            if _excl_je_count:
                _status_notes.append(
                    f"**{_excl_je_count} JE{'s' if _excl_je_count != 1 else ''} excluded** — "
                    "uncheck → re-check to restore"
                )
            if _amt_ovr_count:
                _status_notes.append(
                    f"**{_amt_ovr_count} amount{'s' if _amt_ovr_count != 1 else ''} overridden** — "
                    "re-run Pass 1 to reset all edits"
                )
            if _status_notes:
                st.info("ℹ️ " + "  ·  ".join(_status_notes), icon=None)

            st.write("")

            # ── Description override state — keyed by run so fresh run resets ─
            if st.session_state.get('_je_desc_run') != _run_key:
                st.session_state.je_desc_overrides    = {}
                st.session_state.je_excluded_jes      = set()   # reset exclusions on new run
                st.session_state.je_amount_overrides  = {}       # reset amount edits on new run
                st.session_state.je_account_overrides = {}       # reset account-code edits on new run
                st.session_state._je_desc_run = _run_key
            _all_desc_edits: dict = {}   # (je_num, acct_code) → edited description
            _all_amount_edits: dict = {}  # je_number → adjusted debit amount
            _all_account_edits: dict = {}  # (je_num, acct_code) → corrected account code
            _all_excl_je_nums: set = set()  # je_numbers unchecked across all expanders

            # ── One expander per CR account ──────────────────────────────────
            for _cr_code in _sorted_cr_codes:
                _group_lines = sorted(
                    _groups[_cr_code],
                    key=lambda _l: str(_l.get('account_code') or ''),
                )
                _group_total = sum(_l.get('debit') or 0 for _l in _group_lines)
                _group_count = len(set(_l.get('je_number', '') for _l in _group_lines))
                _expander_label = (
                    f"CR {_cr_section_label(_cr_code)}  ·  "
                    f"{_group_count} JE{'s' if _group_count != 1 else ''}  ·  "
                    f"${_group_total:,.0f}"
                )

                with st.expander(_expander_label, expanded=True):
                    st.caption(f"Credit account: **{_cr_code}** — all entries below post to this account")

                    _excl_set    = st.session_state.get('je_excluded_jes', set())
                    _amt_ovr_set = st.session_state.get('je_amount_overrides', {})
                    _acct_ovr_set = st.session_state.get('je_account_overrides', {})
                    _rows = []
                    for _l in _group_lines:
                        _okey = (_l.get('je_number', ''), _l.get('account_code', ''))
                        _je_num = _l.get('je_number', '')
                        _desc = (st.session_state.je_desc_overrides.get(_okey)
                                 or _clean_je_desc(_l.get('description') or ''))
                        # Corrected account code (if edited) takes priority; name
                        # always re-derived from the current code, not stored text,
                        # so it stays in sync whichever code ends up in effect.
                        _acct_num  = _acct_ovr_set.get(_okey, _l.get('account_code', ''))
                        _acct_name = (_acct_name_lookup.get(str(_acct_num or '').strip(), '')
                                      or _l.get('account_name') or '')
                        # Use amount override if present, else original debit
                        _amt = _amt_ovr_set.get(_je_num, _l.get('debit') or 0)
                        _rows.append({
                            "Include":            _je_num not in _excl_set,
                            "JE #":               _je_num,
                            "File Source":        _SOURCE_FILE_LABEL.get(_l.get('source', ''), _l.get('source', '')),
                            "GL Account Number":  _acct_num,
                            "GL Account Name":    _acct_name,
                            "Description":        _desc,
                            "Amount":             _amt,
                        })

                    _edited = st.data_editor(
                        _rows,
                        num_rows="fixed",
                        use_container_width=True,
                        column_config={
                            "Include":            st.column_config.CheckboxColumn(
                                               "Include", width="small",
                                               help="Uncheck to exclude this JE from the CSV upload. Re-check to restore."),
                            "JE #":               st.column_config.TextColumn(width="small",  disabled=True),
                            "File Source":        st.column_config.TextColumn(width="medium", disabled=True),
                            "GL Account Number":  st.column_config.TextColumn(
                                               width="small",
                                               help="Edit to correct a miscoded account (e.g. a 7xxxxx "
                                                    "corporate code that should be a 6xxxxx property "
                                                    "expense). GL Account Name updates automatically."),
                            "GL Account Name":    st.column_config.TextColumn(width="medium", disabled=True),
                            "Description":        st.column_config.TextColumn(width="large"),   # ← editable
                            "Amount":             st.column_config.NumberColumn(
                                               format="$%,.2f", width="small", min_value=0,
                                               help="Edit to override the pipeline's computed amount. "
                                                    "Both the DR and CR legs update automatically."),
                        },
                        hide_index=True,
                        key=f"je_ed_{_cr_code}_{_run_key}",
                    )

                    # Collect description edits, amount overrides, and exclusions
                    import pandas as _pd_jed
                    _edit_rows = (_edited.to_dict('records')
                                  if isinstance(_edited, _pd_jed.DataFrame) else list(_edited))
                    for _orig, _edit in zip(_rows, _edit_rows):
                        _k = (
                            _orig['JE #'],
                            str(_orig['GL Account Number']),
                        )
                        if _edit.get('Description', '') != _orig.get('Description', ''):
                            _all_desc_edits[_k] = _edit['Description']
                        _new_acct = str(_edit.get('GL Account Number', '') or '').strip()
                        if _new_acct and _new_acct != _k[1]:
                            _all_account_edits[_k] = _new_acct
                        _new_amt = _edit.get('Amount')
                        _orig_amt = _orig.get('Amount', 0)
                        if _new_amt is not None and abs(float(_new_amt) - float(_orig_amt)) > 0.001:
                            _all_amount_edits[_orig['JE #']] = float(_new_amt)
                        if not _edit.get('Include', True):
                            _all_excl_je_nums.add(_orig['JE #'])

                    # Subtotal — only included rows
                    _group_incl_total = sum(
                        _r.get('Amount', 0) for _r in _edit_rows if _r.get('Include', True)
                    )
                    _group_excl_count = sum(
                        1 for _r in _edit_rows if not _r.get('Include', True)
                    )
                    _sub_cols = st.columns([4, 1])
                    with _sub_cols[1]:
                        _sub_label = f"Subtotal: ${_group_incl_total:,.2f}"
                        if _group_excl_count:
                            _sub_label += f" ({_group_excl_count} excluded)"
                        st.markdown(
                            f"<div style='text-align:right;font-weight:bold;padding-top:4px'>"
                            f"{_sub_label}</div>",
                            unsafe_allow_html=True,
                        )

            # ── Apply description/account edits, amount overrides, exclusions → update CSV ──
            _excl_changed   = _all_excl_je_nums   != st.session_state.get('je_excluded_jes', set())
            _amt_changed    = _all_amount_edits   != st.session_state.get('je_amount_overrides', {})
            _acct_changed   = _all_account_edits  != st.session_state.get('je_account_overrides', {})
            if _all_desc_edits or _excl_changed or _amt_changed or _acct_changed:
                if _all_desc_edits:
                    st.session_state.je_desc_overrides = _all_desc_edits
                if _excl_changed:
                    st.session_state.je_excluded_jes = _all_excl_je_nums
                if _amt_changed:
                    st.session_state.je_amount_overrides = _all_amount_edits
                if _acct_changed:
                    st.session_state.je_account_overrides = _all_account_edits
                # Apply description, account-code, and amount overrides to all_je_lines.
                # Amount override updates both the DR leg (debit) and the CR leg (credit)
                # so the JE remains balanced — both legs carry the same je_number. Account-code
                # correction only touches the DR (expense) leg — the CR leg is a different
                # account by design and isn't part of what the user is correcting here.
                _updated_lines = []
                for _l in p1.get("all_je_lines", []):
                    _je_n = _l.get('je_number', '')
                    _k    = (_je_n, _l.get('account_code', ''))
                    _l    = dict(_l)   # shallow copy so we don't mutate the original
                    if _k in _all_account_edits and (_l.get('debit') or 0) > 0:
                        _new_code = _all_account_edits[_k]
                        _l['account_code'] = _new_code
                        _l['account_name'] = _acct_name_lookup.get(_new_code, '')
                    if _k in _all_desc_edits and (_l.get('debit') or 0) > 0:
                        _l['description'] = _all_desc_edits[_k]
                    if _je_n in _all_amount_edits:
                        _new_a = _all_amount_edits[_je_n]
                        if (_l.get('debit') or 0) > 0:
                            _l['debit']  = _new_a
                        elif (_l.get('credit') or 0) > 0:
                            _l['credit'] = _new_a
                    _updated_lines.append(_l)
                p1["all_je_lines"] = _updated_lines
                # CSV export: exclude both DR and CR legs of excluded JE numbers
                _excl_for_csv = st.session_state.get('je_excluded_jes', set())
                _csv_lines = [_l for _l in _updated_lines
                              if _l.get('je_number', '') not in _excl_for_csv]
                _p1_er = st.session_state.pass1_engine_result
                _p1_prop = (
                    (_p1_er.parsed.get('gl') and _p1_er.parsed['gl'].metadata.property_code)
                    if _p1_er else None
                ) or _active_cfg.property_code
                try:
                    from accrual_entry_generator import generate_etl_csv as _gen_etl_ed
                    _ed_csv = os.path.join(st.session_state.temp_dir, f"{_pfx_int}_Accruals_JE.csv")
                    _p1_etl_code = (getattr(_active_cfg, 'yardi_etl_code', '') or _p1_prop)[:8]
                    _gen_etl_ed(_csv_lines, _ed_csv,
                                period=result.period, property_code=_p1_etl_code,
                                auto_reverse=True)
                    p1["accrual_je_csv"] = _ed_csv
                except Exception:
                    pass

            # ── Add Missed Entry(ies) ──────────────────────────────────────────
            # Lets you append DR/CR pairs to the Accruals CSV after JEs are
            # generated — e.g. forgotten one-off accrual entries. A dynamic
            # list of plain widget rows (text_input/number_input/checkbox),
            # not a data_editor grid — the grid's canvas-based cell editing
            # (glide-data-grid) needed a second click to enter a cell and
            # could lose in-progress text on Description specifically.
            # Confirmed with Ryan 2026-08-06. Plain widgets don't have that
            # class of issue at all, at the cost of needing a row-count
            # control instead of the grid's built-in "+" button.
            st.markdown("#### ➕  Add Missed Entries")

            # Keep the expander open after a successful add (st.rerun() collapses
            # expanders that use a hardcoded expanded=False default).
            _add_expanded_key = f"je_add_expanded_{_run_key}"
            if _add_expanded_key not in st.session_state:
                st.session_state[_add_expanded_key] = False

            # How many rows to show, and a "generation" counter that busts
            # every row widget's key on reset so fields actually go blank
            # after a submit instead of the widgets remembering their last
            # typed value (same reason the original single-row form bumped
            # a counter into its own widget keys).
            _add_rows_key = f"je_add_rows_{_run_key}"
            if _add_rows_key not in st.session_state:
                st.session_state[_add_rows_key] = 1
            _add_gen_key = f"je_add_gen_{_run_key}"
            if _add_gen_key not in st.session_state:
                st.session_state[_add_gen_key] = 0
            _gen = st.session_state[_add_gen_key]

            with st.expander("Add entries to Accruals CSV",
                             expanded=st.session_state[_add_expanded_key]):
                st.caption(
                    "Fill in as many rows as you need — click **+ Add Row** for more — then "
                    "click **Add All Entries** once to submit them together. Blank rows are "
                    "skipped."
                )
                for _r_i in range(st.session_state[_add_rows_key]):
                    _rc1, _rc2, _rc3, _rc4, _rc5 = st.columns([1.3, 1.3, 3.6, 1.5, 1])
                    with _rc1:
                        st.text_input(
                            "DR Account", placeholder="e.g. 637150",
                            key=f"add_dr_{_run_key}_{_gen}_{_r_i}",
                            label_visibility="visible" if _r_i == 0 else "collapsed",
                        )
                    with _rc2:
                        st.text_input(
                            "CR Account", value="213100",
                            key=f"add_cr_{_run_key}_{_gen}_{_r_i}",
                            label_visibility="visible" if _r_i == 0 else "collapsed",
                        )
                    with _rc3:
                        st.text_input(
                            "Description", placeholder="e.g. Tenant Relations accrual",
                            key=f"add_desc_{_run_key}_{_gen}_{_r_i}",
                            label_visibility="visible" if _r_i == 0 else "collapsed",
                        )
                    with _rc4:
                        st.number_input(
                            "Amount ($)", min_value=0.0, step=100.0, format="%.2f",
                            key=f"add_amt_{_run_key}_{_gen}_{_r_i}",
                            label_visibility="visible" if _r_i == 0 else "collapsed",
                        )
                    with _rc5:
                        st.checkbox(
                            "Auto-Rev", value=True,
                            key=f"add_autorev_{_run_key}_{_gen}_{_r_i}",
                            help="✅ Checked = auto-reverses next month (ReverseNextMonth = -1). "
                                 "Uncheck for permanent JEs (ReverseNextMonth = 0).",
                        )

                _row_add_col, _submit_col = st.columns([1, 3])
                with _row_add_col:
                    if st.button("➕ Add Row", key=f"add_row_btn_{_run_key}_{_gen}"):
                        st.session_state[_add_rows_key] += 1
                        st.session_state[_add_expanded_key] = True
                        st.rerun()
                with _submit_col:
                    _add_submit = st.button("Add All Entries", key=f"add_btn_{_run_key}_{_gen}",
                                           type="primary")

                if _add_submit:
                    # Determine the next ADD-XXXX number once, then increment
                    # across every new row so a single submission with several
                    # rows gets sequential JE numbers instead of colliding.
                    _prev_adds = [
                        l for l in p1.get("all_je_lines", [])
                        if str(l.get('je_number', '')).startswith('ADD-')
                    ]
                    _next_add_num = (len(_prev_adds) // 2) + 1

                    _new_je_lines = []
                    _added_summaries = []
                    _skipped_rows = 0
                    for _r_i in range(st.session_state[_add_rows_key]):
                        _dr = (st.session_state.get(f"add_dr_{_run_key}_{_gen}_{_r_i}", '') or '').strip()
                        _cr = (st.session_state.get(f"add_cr_{_run_key}_{_gen}_{_r_i}", '') or '213100').strip()
                        _desc = (st.session_state.get(f"add_desc_{_run_key}_{_gen}_{_r_i}", '') or '').strip()
                        _amt = float(st.session_state.get(f"add_amt_{_run_key}_{_gen}_{_r_i}", 0) or 0)
                        if not _dr or not _desc or _amt <= 0:
                            # A fully blank extra row is normal — only count it
                            # as "skipped" if the user partially filled it in,
                            # so the warning below doesn't fire on every
                            # ordinary single-row submission.
                            if _dr or _desc or _amt:
                                _skipped_rows += 1
                            continue

                        _new_je_id = f"ADD-{_next_add_num:04d}"
                        _next_add_num += 1
                        _add_rev_flag = -1 if bool(st.session_state.get(
                            f"add_autorev_{_run_key}_{_gen}_{_r_i}", True)) else 0
                        _new_je_lines.extend([
                            {
                                'je_number':          _new_je_id, 'line': 1, 'date': '',
                                'account_code':       _dr,
                                'account_name':       '',
                                'description':        _desc,
                                'reference':          'MANUAL-ADD',
                                'debit':              round(_amt, 2), 'credit': 0,
                                'vendor':             '[Manual Addition]',
                                'invoice_number':     '',
                                'source':             'manual_addition',
                                'confidence':         'high',
                                'reverse_next_month': _add_rev_flag,
                            },
                            {
                                'je_number':          _new_je_id, 'line': 2, 'date': '',
                                'account_code':       _cr,
                                'account_name':       '',
                                'description':        _desc,
                                'reference':          'MANUAL-ADD',
                                'debit':              0, 'credit': round(_amt, 2),
                                'vendor':             '[Manual Addition]',
                                'invoice_number':     '',
                                'source':             'manual_addition',
                                'confidence':         'high',
                                'reverse_next_month': _add_rev_flag,
                            },
                        ])
                        _added_summaries.append(
                            f"**{_new_je_id}** — DR {_dr} / CR {_cr}  ${_amt:,.2f}  ·  {_desc}"
                        )

                    if _skipped_rows:
                        st.warning(
                            f"⚠️ Skipped {_skipped_rows} row(s) missing a DR Account, "
                            f"Description, or a positive Amount.",
                            icon="⚠️",
                        )

                    if not _new_je_lines:
                        st.warning(
                            "No complete rows to add — fill in DR Account, Description, "
                            "and Amount for at least one row.", icon="⚠️",
                        )
                    else:
                        _updated_all = p1.get("all_je_lines", []) + _new_je_lines
                        p1["all_je_lines"] = _updated_all

                        # Regenerate CSV once for the whole batch
                        _p1_er_add = st.session_state.pass1_engine_result
                        _p1_prop_add = (
                            (_p1_er_add.parsed.get('gl') and
                             _p1_er_add.parsed['gl'].metadata.property_code)
                            if _p1_er_add else None
                        ) or _active_cfg.property_code
                        try:
                            from accrual_entry_generator import generate_etl_csv as _gen_etl_add
                            _add_csv_path = os.path.join(
                                st.session_state.temp_dir, f"{_pfx_int}_Accruals_JE.csv"
                            )
                            _p1_etl_code_add = (getattr(_active_cfg, 'yardi_etl_code', '') or _p1_prop_add)[:8]
                            _gen_etl_add(
                                _updated_all, _add_csv_path,
                                period=result.period, property_code=_p1_etl_code_add,
                                auto_reverse=True
                            )
                            p1["accrual_je_csv"] = _add_csv_path
                            st.success(
                                f"✅  Added {len(_added_summaries)} entr"
                                f"{'y' if len(_added_summaries) == 1 else 'ies'} — CSV updated.\n\n"
                                + "\n\n".join(_added_summaries),
                                icon="✅",
                            )
                        except Exception as _add_ex:
                            st.warning(
                                f"Entries added to session but CSV regeneration failed: {_add_ex}",
                                icon="⚠️",
                            )

                        # Reset to a single blank row and keep the expander
                        # open so the user can immediately add more. Bumping
                        # the generation counter changes every row widget's
                        # key, which is what actually clears them — resetting
                        # _add_rows_key alone wouldn't, since row 0's widgets
                        # would keep their same keys (and therefore their old
                        # typed values) across the rerun.
                        st.session_state[_add_rows_key] = 1
                        st.session_state[_add_gen_key] += 1
                        st.session_state[_add_expanded_key] = True
                        st.rerun()

                # ── Previously added entries ──────────────────────────────────
                _manual_adds = [
                    l for l in p1.get("all_je_lines", [])
                    if l.get('source') == 'manual_addition' and (l.get('debit') or 0) > 0
                ]
                if _manual_adds:
                    st.caption(
                        f"**{len(_manual_adds)} manually added "
                        f"entr{'ies' if len(_manual_adds) != 1 else 'y'} in this session:**"
                    )
                    for _ma in _manual_adds:
                        _ma_cr = next(
                            (l['account_code'] for l in p1.get("all_je_lines", [])
                             if l.get('je_number') == _ma['je_number'] and (l.get('credit') or 0) > 0),
                            '?'
                        )
                        st.caption(
                            f"• **{_ma['je_number']}** · "
                            f"DR {_ma['account_code']} / CR {_ma_cr} · "
                            f"${_ma.get('debit', 0):,.2f} · "
                            f"{_ma.get('description', '')}"
                        )

            st.divider()
        else:
            st.info("No accrual entries generated. Upload a Nexus file, Budget Comparison, "
                    "or Prepaid Ledger to enable additional accrual detection layers.", icon="ℹ️")

        # ── 7xxxxx Intercompany Recode Table ─────────────────────────────────
        # Shown after the accruals/missed-entry section because the GL must be
        # parsed first to detect 7xxxxx accounts.  Fill in the DR account and
        # re-run Generate JEs to include the recode entries in the CSV.

        _p1_er_ic = st.session_state.get('pass1_engine_result')
        # Detect 7xxx accounts directly from the parsed GL — do NOT rely on
        # engine.py's corp_7xxx_accounts summary key, which may be stale if
        # Streamlit Cloud cached the old engine module bytecode between deploys.
        _interco_detected = []
        _interco_seen_codes: set = set()
        if _p1_er_ic:
            _ic_gl = _p1_er_ic.parsed.get('gl')
            if _ic_gl and hasattr(_ic_gl, 'accounts'):
                for _ic_acct in _ic_gl.accounts:
                    _ic_code = str(getattr(_ic_acct, 'account_code', '') or '').strip()
                    _ic_nc   = float(getattr(_ic_acct, 'net_change', 0) or 0)
                    _ic_eb   = float(getattr(_ic_acct, 'ending_balance', 0) or 0)
                    _ic_sig  = _ic_eb if abs(_ic_eb) >= 0.01 else _ic_nc
                    if abs(_ic_sig) < 0.01:
                        continue
                    if _ic_code.startswith('7'):
                        _interco_detected.append({
                            'account_code': _ic_code,
                            'account_name': str(getattr(_ic_acct, 'account_name', '') or '').strip(),
                            'net_amount':   _ic_sig,
                        })
                        _interco_seen_codes.add(_ic_code)

            # A 7xxxxx account whose activity nets to $0 this period — invoiced
            # in, then reclassed straight out to 135150 Prepaid Other, all in
            # the same month — is invisible to the net/ending-balance check
            # above, even though real (potentially miscoded) money flowed
            # through it. Confirmed on a real Kardin Budgeting Software
            # invoice: coded to 712210 (a 7xxxxx corporate account) on entry,
            # reclassed out to 135150 three weeks later, net change exactly
            # $0.00 — so it never appeared here for Ryan to correct, and the
            # prepaid ledger kept citing 712210 for every future release with
            # no way to catch it. Pair GL transactions by control (JE) number
            # instead of by net: any 7xxxxx leg that moved money into 135150
            # this period is worth surfacing for review regardless of net,
            # since whatever account it reclassed FROM is exactly what the
            # prepaid item will keep citing unless corrected here. Confirmed
            # with Ryan 2026-08-20.
            if _ic_gl:
                _ic_by_control: dict = {}
                for _ic_t in (getattr(_ic_gl, 'all_transactions', None) or []):
                    _ic_ctrl = str(getattr(_ic_t, 'control', '') or '').strip()
                    if _ic_ctrl:
                        _ic_by_control.setdefault(_ic_ctrl, []).append(_ic_t)
                for _ic_legs in _ic_by_control.values():
                    _ic_135150_amt = next(
                        (float(getattr(_lt, 'debit', 0) or 0) for _lt in _ic_legs
                         if str(getattr(_lt, 'account_code', '') or '').strip() == '135150'
                         and float(getattr(_lt, 'debit', 0) or 0) > 0.01),
                        0.0,
                    )
                    if _ic_135150_amt < 0.01:
                        continue
                    for _lt in _ic_legs:
                        _lt_code = str(getattr(_lt, 'account_code', '') or '').strip()
                        if not _lt_code.startswith('7') or _lt_code in _interco_seen_codes:
                            continue
                        _lt_credit = float(getattr(_lt, 'credit', 0) or 0)
                        if abs(_lt_credit - _ic_135150_amt) < 0.01:
                            _interco_detected.append({
                                'account_code': _lt_code,
                                'account_name': str(getattr(_lt, 'account_name', '') or '').strip(),
                                'net_amount':   _lt_credit,
                            })
                            _interco_seen_codes.add(_lt_code)

        # ── 5xxxxx company revenue warning ───────────────────────────────────
        _co_rev_detected = (_p1_er_ic.summary.get('co_rev_5xxx_accounts', [])
                            if _p1_er_ic else [])
        if _co_rev_detected:
            _co_rev_lines = ', '.join(
                f"{a['account_code']} ({a['account_name']}) ${abs(a['net_amount']):,.0f}"
                for a in _co_rev_detected
            )
            st.warning(
                f"⚠️ **5xxxxx Company Revenue on Property GL** — {len(_co_rev_detected)} account(s) detected: "
                f"{_co_rev_lines}. "
                f"5xxxxx is entity-level revenue and should not appear on the property GL. "
                f"Review and recode via the Manual JEs table if needed.",
                icon="⚠️",
            )

        # Auto-merge newly detected accounts into the recode table (idempotent).
        # Dedup by Leg=="CR" + Account.
        #
        # Appends directly to the LIVE row-widget list (when it already
        # exists) instead of mutating interco_recode_df and bumping
        # _interco_seed_gen. Bumping the gen forces a full reseed of every
        # row from interco_recode_df — which discards any value the user
        # just typed this same run (e.g. a DR account typed right before
        # clicking Generate JEs) if that keystroke hadn't yet been captured
        # into interco_recode_df by the render block's own write-back, since
        # the write-back runs AFTER this auto-merge in script order. A direct
        # append only adds the new rows' ids/widget state and never touches
        # any existing row's key, so nothing in-progress can be clobbered.
        # Confirmed with Ryan 2026-08-12.
        #
        # Only run once per actual "Generate JEs"/"Re-run Pass 1" click, not
        # on every incidental rerun this section's OWN persistent
        # pass1_complete gate lets through (e.g. adding a One-Off Accrual
        # row elsewhere on the page, or clicking this table's own 🗑️ delete
        # button) — otherwise deleting a row just triggers an immediate
        # re-detect-and-re-add on the very same rerun (the account is still
        # sitting in the GL, so it looks "missing" the instant its row is
        # gone), and deleting only one leg of a CR/DR pair adds a second
        # fresh pair on top of the now-orphaned leftover leg. Confirmed with
        # Ryan 2026-08-13 — "when I try and remove a row it adds rows".
        _interco_run_key = st.session_state.get('pass1_run_count', 0)
        if _interco_detected and st.session_state.get('_interco_last_merged_run') != _interco_run_key:
            st.session_state['_interco_last_merged_run'] = _interco_run_key
            _ic_row_ids_key_early = "ic_row_ids"
            # The live-append path is only safe when the row list is FRESH
            # for the current _interco_seed_gen. Right after Reset Pass 1 /
            # a property switch, ic_row_ids can still exist (stale, from
            # before the reset) while ic_rows_seed_gen no longer matches
            # _interco_seed_gen — appending live in that state is pointless,
            # since the render block's own reseed check (which also fires
            # this run, because the gen still mismatches) would immediately
            # overwrite ic_row_ids from interco_recode_df anyway, discarding
            # whatever was just appended. Route through the DataFrame in
            # that case instead, so the upcoming reseed picks up the merge
            # too. Confirmed with Ryan 2026-08-12.
            _ic_rows_are_fresh = (
                _ic_row_ids_key_early in st.session_state
                and st.session_state.get("ic_rows_seed_gen") == st.session_state.get("_interco_seed_gen", 0)
            )
            if _ic_rows_are_fresh:
                _existing_cr_accts = {
                    str(st.session_state.get(f"ic_account_{_rid}", "") or "").strip()
                    for _rid in st.session_state[_ic_row_ids_key_early]
                    if str(st.session_state.get(f"ic_leg_{_rid}", "") or "") == "CR"
                }
            else:
                _ic_df_cur = st.session_state.interco_recode_df.copy()
                _existing_cr_accts = set()
                if "Leg" in _ic_df_cur.columns and "Account" in _ic_df_cur.columns:
                    _cr_mask = _ic_df_cur["Leg"].fillna("") == "CR"
                    _existing_cr_accts = set(_ic_df_cur.loc[_cr_mask, "Account"].fillna("").str.strip())

            _new_ic_pairs = []
            for _ic in _interco_detected:
                _ic_code = str(_ic.get('account_code', '')).strip()
                if _ic_code and _ic_code not in _existing_cr_accts:
                    _ic_amt  = abs(float(_ic.get('net_amount', 0)))
                    _ic_name = str(_ic.get('account_name', ''))
                    _ic_desc = f"Recode {_ic_code} to expense account"
                    _new_ic_pairs.append((
                        {"Leg": "CR", "Account": _ic_code, "Account Name": _ic_name,
                         "Credit ($)": _ic_amt, "Debit ($)": 0.0, "Description": _ic_desc},
                        {"Leg": "DR", "Account": "", "Account Name": "",
                         "Credit ($)": 0.0, "Debit ($)": _ic_amt, "Description": _ic_desc},
                    ))

            if _new_ic_pairs and _ic_rows_are_fresh:
                # Row list already live and current — append directly,
                # touching nothing else.
                for _cr_row, _dr_row in _new_ic_pairs:
                    for _new_row in (_cr_row, _dr_row):
                        _new_rid = st.session_state["ic_next_id"]
                        st.session_state["ic_next_id"] += 1
                        st.session_state[f"ic_leg_{_new_rid}"]     = _new_row["Leg"]
                        st.session_state[f"ic_account_{_new_rid}"] = _new_row["Account"]
                        st.session_state[f"ic_name_{_new_rid}"]    = _new_row["Account Name"]
                        st.session_state[f"ic_credit_{_new_rid}"]  = _new_row["Credit ($)"]
                        st.session_state[f"ic_debit_{_new_rid}"]   = _new_row["Debit ($)"]
                        st.session_state[f"ic_desc_{_new_rid}"]    = _new_row["Description"]
                        st.session_state["ic_row_ids"].append(_new_rid)
            elif _new_ic_pairs:
                # Row list doesn't exist yet, or is stale relative to the
                # current _interco_seed_gen (e.g. right after Reset Pass 1) —
                # nothing live to safely preserve, so merge into the
                # DataFrame and bump the gen; the reseed check below (which
                # already needs to run in this case) will pick up both the
                # reset and this merge in one pass.
                _ic_df_cur = st.session_state.interco_recode_df.copy()
                _flat_new_rows = [r for pair in _new_ic_pairs for r in pair]
                st.session_state.interco_recode_df = pd.concat(
                    [_ic_df_cur, pd.DataFrame(_flat_new_rows)], ignore_index=True
                )
                st.session_state._interco_seed_gen = st.session_state.get('_interco_seed_gen', 0) + 1

        # Enrich _acct_name_lookup with budget comparison accounts so the recode
        # DR row Account Name can auto-populate even for 6xxx accounts with no
        # direct GL activity this period (the target account often has zero GL
        # hits because all charges were miscoded to 7xxx instead).
        _p1_er_bc = st.session_state.get('pass1_engine_result')
        if _p1_er_bc:
            try:
                _bc_for_lookup = getattr(_p1_er_bc, 'parsed', {}).get('budget_comparison')
                if _bc_for_lookup and hasattr(_bc_for_lookup, 'line_items'):
                    for _bcl in _bc_for_lookup.line_items:
                        _bcl_code = str(getattr(_bcl, 'account_code', '') or '').strip()
                        _bcl_name = str(getattr(_bcl, 'account_name', '') or '').strip()
                        if _bcl_code and _bcl_name and _bcl_code not in _acct_name_lookup:
                            _acct_name_lookup[_bcl_code] = _bcl_name
            except Exception:
                pass

        _ic_badge = (f"  ⚠️ {len(_interco_detected)} account(s) detected"
                     if _interco_detected else "")
        with st.expander(
            f"🔄 7xxxxx Intercompany Recode  (DR expense → CR 7xxxxx){_ic_badge}",
            expanded=bool(_interco_detected),
        ):
            st.caption(
                "7xxxxx accounts are **corporate expenses** (non-property) — they should not remain on the property GL. "
                "Each detected account auto-populates as a **DR / CR pair**: the CR row is pre-filled with the 7xxxxx account; "
                "enter the **6xxxxx or 8xxxxx** target expense account on the **DR row**. "
                "The Account Name fills automatically once you tab out of the Account field. "
                "The pipeline generates **DR [expense account] / CR [7xxx account]** (permanent — no auto-reverse). "
                "Edit the Amount on the DR row if you're only recoding a partial amount. "
                "**Re-run Generate JEs after filling in target accounts** to include the recode JEs in the CSV."
            )

            # Plain widgets (text_input/number_input) instead of st.data_editor —
            # same fix as One-Off Accruals and Add Missed Entries. This table had
            # an additional, distinct data_editor bug: the auto-populated Account
            # Name never actually displayed, because Streamlit widgets (data_editor
            # included) ignore a freshly-computed `value=`/`data=` on any rerun
            # after the widget's `key` already has established state — so writing
            # the looked-up name into st.session_state.interco_recode_df (a
            # different, non-widget key) and waiting for "the next natural rerun"
            # never actually reached the grid, since the grid kept displaying its
            # OWN internal state under interco_recode_editor_<run_count> instead.
            # Confirmed with Ryan 2026-08-06 (name not populating running Feb after
            # Jan). Plain widgets don't have this failure mode: writing directly to
            # a row's own f"ic_name_{rid}" key before that widget is re-instantiated
            # this run takes effect immediately.
            _IC_IDS_KEY  = "ic_row_ids"
            _IC_NEXT_KEY = "ic_next_id"
            _IC_GEN_KEY  = "ic_rows_seed_gen"

            def _ic_seed_widget(_rid: int, _seed: dict) -> None:
                st.session_state[f"ic_leg_{_rid}"]     = _seed["Leg"]
                st.session_state[f"ic_account_{_rid}"] = _seed["Account"]
                st.session_state[f"ic_name_{_rid}"]    = _seed["Account Name"]
                st.session_state[f"ic_credit_{_rid}"]  = _seed["Credit ($)"]
                st.session_state[f"ic_debit_{_rid}"]   = _seed["Debit ($)"]
                st.session_state[f"ic_desc_{_rid}"]    = _seed["Description"]

            if (_IC_IDS_KEY not in st.session_state
                    or st.session_state.get(_IC_GEN_KEY) != st.session_state.get("_interco_seed_gen", 0)):
                _ic_seed_rows = _df_to_ic_rows(st.session_state.interco_recode_df)
                _ic_new_ids = list(range(len(_ic_seed_rows)))
                for _rid, _seed in zip(_ic_new_ids, _ic_seed_rows):
                    _ic_seed_widget(_rid, _seed)
                st.session_state[_IC_IDS_KEY]  = _ic_new_ids
                st.session_state[_IC_NEXT_KEY] = len(_ic_seed_rows)
                st.session_state[_IC_GEN_KEY]  = st.session_state.get("_interco_seed_gen", 0)

            # Auto-populate Account Name for any row (CR or DR) where Account is
            # filled but Account Name is blank — covers both the DR target
            # account AND a CR-leg 7xxxxx code the user edits manually
            # (auto-detected CR rows get their name pre-filled at detection
            # time above, but that doesn't cover a manually-edited CR account).
            # Must run BEFORE this run's widgets are instantiated below —
            # Streamlit forbids writing to a widget's session_state key after
            # that widget has already rendered in the same script run.
            for _rid in st.session_state[_IC_IDS_KEY]:
                _ic_n_acct = str(st.session_state.get(f"ic_account_{_rid}", "") or "").strip()
                _ic_n_name = str(st.session_state.get(f"ic_name_{_rid}", "") or "").strip()
                if _ic_n_acct and not _ic_n_name:
                    _ic_looked_up = _acct_name_lookup.get(_ic_n_acct, "")
                    if _ic_looked_up:
                        st.session_state[f"ic_name_{_rid}"] = _ic_looked_up

            if not st.session_state[_IC_IDS_KEY]:
                st.caption("No 7xxxxx activity detected this period.")

            for _ic_row_i, _rid in enumerate(st.session_state[_IC_IDS_KEY]):
                _ic_lbl = "visible" if _ic_row_i == 0 else "collapsed"
                _ic_cols = st.columns([0.7, 1.1, 1.8, 1.1, 1.1, 2.2, 0.5])
                with _ic_cols[0]:
                    st.text_input("Leg", key=f"ic_leg_{_rid}", label_visibility=_ic_lbl,
                                  disabled=True,
                                  help="CR = the 7xxxxx being credited out | DR = the expense account to debit")
                with _ic_cols[1]:
                    st.text_input("Account", key=f"ic_account_{_rid}", label_visibility=_ic_lbl,
                                  help="CR row: 7xxxxx account (pre-filled). DR row: enter the 6xxxxx or 8xxxxx target expense account.")
                with _ic_cols[2]:
                    st.text_input("Account Name", key=f"ic_name_{_rid}", label_visibility=_ic_lbl,
                                  disabled=True)
                with _ic_cols[3]:
                    st.number_input("Credit ($)", key=f"ic_credit_{_rid}", label_visibility=_ic_lbl,
                                    min_value=0.0, step=100.0, format="%.2f",
                                    help="Pre-filled on the CR row from the GL. Edit if recoding a partial amount.")
                with _ic_cols[4]:
                    st.number_input("Debit ($)", key=f"ic_debit_{_rid}", label_visibility=_ic_lbl,
                                    min_value=0.0, step=100.0, format="%.2f",
                                    help="Pre-filled on the DR row to match the CR. Edit if recoding a partial amount.")
                with _ic_cols[5]:
                    st.text_input("Description", key=f"ic_desc_{_rid}", label_visibility=_ic_lbl)
                with _ic_cols[6]:
                    if _ic_row_i == 0:
                        st.write("")   # align delete button with inputs, not their labels
                    if st.button("🗑️", key=f"ic_del_{_rid}", help="Remove this row"):
                        st.session_state[_IC_IDS_KEY] = [
                            _i for _i in st.session_state[_IC_IDS_KEY] if _i != _rid
                        ]
                        st.rerun()

            # ── Write back to interco_recode_df for downstream consumers ────────
            # (early prepaid-ledger recode map, JE building — both now read the
            # live widget state directly via _read_interco_df_from_widgets, but
            # this mirror is still what re-seeds the row list on a reset and what
            # the auto-merge-new-accounts dedup check reads against.)
            _ic_out_rows = [
                {
                    "Leg": st.session_state.get(f"ic_leg_{_rid}", ""),
                    "Account": st.session_state.get(f"ic_account_{_rid}", ""),
                    "Account Name": st.session_state.get(f"ic_name_{_rid}", ""),
                    "Credit ($)": float(st.session_state.get(f"ic_credit_{_rid}", 0.0) or 0.0),
                    "Debit ($)": float(st.session_state.get(f"ic_debit_{_rid}", 0.0) or 0.0),
                    "Description": st.session_state.get(f"ic_desc_{_rid}", ""),
                }
                for _rid in st.session_state[_IC_IDS_KEY]
            ]
            _ic_recode_edited = pd.DataFrame(_ic_out_rows, columns=_IC_COLUMNS)
            st.session_state.interco_recode_df = _ic_recode_edited

            # Active = DR rows that have a non-blank target Account and a Debit amount
            _ic_dr_active = _ic_recode_edited[
                (_ic_recode_edited["Leg"].fillna("") == "DR") &
                _ic_recode_edited["Account"].fillna("").str.strip().astype(bool) &
                (_ic_recode_edited["Debit ($)"].fillna(0) > 0)
            ]
            if not _ic_dr_active.empty:
                st.success(
                    f"✅ {len(_ic_dr_active)} recode JE(s) queued — "
                    f"${_ic_dr_active['Debit ($)'].sum():,.2f} total. "
                    f"Re-run Generate JEs to include in the CSV.",
                    icon="✅",
                )
            elif _interco_detected:
                st.warning(
                    "⚠️ 7xxxxx accounts detected — enter the 6xxxxx or 8xxxxx target on each DR row, "
                    "then re-run Generate JEs to include recode JEs in the CSV.",
                    icon="⚠️",
                )

        st.divider()
        if st.button(
            "🔁 Re-run Pass 1",
            key="pass1_rerun_btn_bottom",
            use_container_width=True,
            help="Re-run Pass 1 with your recode and One-Off Accrual table edits included — "
                 "same action as the Re-run Pass 1 button above.",
        ):
            st.session_state['_trigger_pass1_rerun'] = True
            st.rerun()

        # ── Prepaid Amortization Panel ─────────────────────────────────────
        amort_lines = p1.get("amort_lines", [])
        if amort_lines:
            with st.expander("Prepaid Expense Amortization", expanded=False):
                cur_lines = [l for l in amort_lines if l.get('is_current_period')]
                fut_lines = [l for l in amort_lines if not l.get('is_current_period')]
                col_p1, col_p2 = st.columns(2)
                with col_p1:
                    st.metric("Current Period Expense", f"${sum(l['monthly_amount'] for l in cur_lines):,.2f}")
                with col_p2:
                    st.metric("Future Periods (Prepaid Asset)", f"${sum(l['monthly_amount'] for l in fut_lines):,.2f}")

                # One row per unique invoice (first occurrence) — editable, so a
                # misread invoice (wrong amount, service dates, or GL account) can
                # be corrected here. Corrections apply on the NEXT "Generate JEs" /
                # "Re-run Pass 1" click, same as the One-Off Accruals table.
                st.markdown("**Correct a misread invoice**")
                st.caption(
                    "Edit GL Account Number, Total Amount, or Service Start/End below, then "
                    "click **Re-run Pass 1** to apply the correction to the ledger and JEs — "
                    "this is exactly how a Kardin-style miscoding (e.g. a corporate 7xxxxx "
                    "account instead of the right property expense account) gets fixed at the "
                    "source, before it ever reaches the ledger. GL Account Name is looked up "
                    "automatically from the number, same as the Intercompany Recode table, so "
                    "a typo'd account reads as an obviously wrong name instead of just a "
                    "number. Vendor / Invoice # are read-only — they're the match key."
                )
                _seen_keys = set()
                _invoice_rows = []
                for l in amort_lines:
                    _k = prepaid_ledger._invoice_key(l['vendor'], l['invoice_number'])
                    if _k in _seen_keys:
                        continue
                    _seen_keys.add(_k)
                    _fresh_start = str(l['service_start']) if l.get('service_start') else ''
                    _fresh_end   = str(l['service_end']) if l.get('service_end') else ''
                    _fresh_gl    = l['gl_account_number']
                    _fresh_amt   = l['total_amount']
                    _invoice_rows.append({
                        "_key": _k,
                        "Vendor": l['vendor'],
                        "Invoice #": l['invoice_number'],
                        "GL Account Number": _fresh_gl,
                        "GL Account Name": _acct_name_lookup.get(_fresh_gl, ''),
                        "Description": l.get('description', ''),
                        "Total Amount ($)": _fresh_amt,
                        "Service Start": _fresh_start,
                        "Service End": _fresh_end,
                        "_orig_amount": _fresh_amt,
                        "_orig_service_start": _fresh_start,
                        "_orig_service_end": _fresh_end,
                        "_orig_gl_account": _fresh_gl,
                    })
                # Preserve a prior run's edit ONLY if the underlying Nexus data
                # hasn't changed since this invoice was first tracked (compare
                # the stored _orig_* snapshot against this run's fresh values).
                # Otherwise a corrected/re-exported Nexus file (a real workflow
                # this session used repeatedly) would have its new numbers
                # silently overridden by a stale first-seen value forever.
                _prior_ov = st.session_state.prepaid_overrides_df
                _prior_by_key = {r["_key"]: r for _, r in _prior_ov.iterrows()} if not _prior_ov.empty else {}
                _stale_refreshed = []
                _merged_rows = []
                for r in _invoice_rows:
                    _prior_row = _prior_by_key.get(r["_key"])
                    if _prior_row is None:
                        _merged_rows.append(r)
                        continue
                    _source_changed = (
                        float(_prior_row.get("_orig_amount", 0) or 0) != r["_orig_amount"]
                        or str(_prior_row.get("_orig_service_start", "") or "") != r["_orig_service_start"]
                        or str(_prior_row.get("_orig_service_end", "") or "") != r["_orig_service_end"]
                        or str(_prior_row.get("_orig_gl_account", "") or "") != r["_orig_gl_account"]
                    )
                    if _source_changed:
                        _merged_rows.append(r)
                        _stale_refreshed.append(r["Invoice #"])
                    else:
                        _merged_rows.append(_prior_row)
                if _stale_refreshed:
                    st.info(
                        f"↳ Nexus data changed since last tracked for invoice(s) "
                        f"{', '.join(_stale_refreshed)} — override reset to the new "
                        f"values below. Re-verify before re-running.",
                        icon="ℹ️",
                    )
                _base_ov_df = pd.DataFrame(_merged_rows) if _merged_rows else st.session_state.prepaid_overrides_df

                _ov_edited = st.data_editor(
                    _base_ov_df,
                    use_container_width=True,
                    hide_index=True,
                    # GL Account Name is read-only/derived — editing the Number
                    # is what actually corrects the account; the Name column
                    # exists only so a typo'd or miscoded number is instantly
                    # recognizable (e.g. "725070 (Advertising & Marketing)"
                    # next to a Kardin invoice reads as obviously wrong,
                    # whereas the bare number alone didn't).
                    disabled=["Vendor", "Invoice #", "GL Account Name"],
                    column_order=["Vendor", "Invoice #", "GL Account Number", "GL Account Name",
                                  "Description", "Total Amount ($)", "Service Start", "Service End"],
                    column_config={
                        "Vendor": st.column_config.TextColumn(width="medium"),
                        "Invoice #": st.column_config.TextColumn(width="small"),
                        "GL Account Number": st.column_config.TextColumn(width="small"),
                        "GL Account Name": st.column_config.TextColumn(width="medium"),
                        "Description": st.column_config.TextColumn(width="large"),
                        "Total Amount ($)": st.column_config.NumberColumn(format="$%,.2f", min_value=0.0),
                        "Service Start": st.column_config.TextColumn(width="small", help="YYYY-MM-DD"),
                        "Service End": st.column_config.TextColumn(width="small", help="YYYY-MM-DD"),
                    },
                    key="prepaid_overrides_editor",
                )
                # Re-derive GL Account Name from whatever Number the user just
                # typed, so it never shows a stale name next to an edited
                # number — the editor's own row only knows what was there
                # when this render started.
                if "GL Account Number" in _ov_edited.columns:
                    _ov_edited["GL Account Name"] = _ov_edited["GL Account Number"].apply(
                        lambda _c: _acct_name_lookup.get(str(_c or '').strip(), '')
                    )
                st.session_state.prepaid_overrides_df = _ov_edited

                st.markdown("**Full amortization schedule**")
                amort_rows = [{
                    "Vendor": l['vendor'], "Invoice #": l['invoice_number'],
                    "Period": l['period_label'], "Month": f"{l['month_index']}/{l['total_months']}",
                    "Monthly Amount": l['monthly_amount'], "GL Account": l['gl_account_number'],
                    "Current Period": "Yes" if l.get('is_current_period') else "",
                } for l in amort_lines]
                st.dataframe(amort_rows, use_container_width=True, hide_index=True,
                             column_config={
                                 "Vendor": st.column_config.TextColumn(width="medium"),
                                 "Invoice #": st.column_config.TextColumn(width="small"),
                                 "Period": st.column_config.TextColumn(width="small"),
                                 "Month": st.column_config.TextColumn(width="small"),
                                 "Monthly Amount": st.column_config.NumberColumn(format="$%,.2f"),
                                 "GL Account": st.column_config.TextColumn(width="small"),
                                 "Current Period": st.column_config.TextColumn(width="small"),
                             })
            st.divider()

        # ── Prepaid Ledger Status ──────────────────────────────────────────
        ledger_active    = p1.get("ledger_active", [])
        ledger_completed = p1.get("ledger_completed", [])
        newly_added      = p1.get("newly_added_prepaids", [])
        released_count   = p1.get("prepaid_released_count", 0)
        release_lines    = p1.get("prepaid_release_lines", [])
        # close_period itself is a local variable set inside the "Generate JEs"
        # button handler above — it only exists in the SAME script run that
        # button click triggered. This results block renders on every rerun
        # while p1 has persisted data (e.g. the user just clicking something
        # else on the page after Pass 1 already completed), so it must read
        # the period back from p1 rather than reference that local variable,
        # which would raise NameError on any later rerun.
        close_period     = p1.get("close_period", "")
        if ledger_active or ledger_completed or newly_added:
            st.markdown("### Prepaid Ledger")
            col_l1, col_l2, col_l3, col_l4 = st.columns(4)
            with col_l1:
                st.metric("Active Prepaid Items", len(ledger_active))
            with col_l2:
                st.metric("Released This Month", released_count)
            with col_l3:
                st.metric("New This Month", len(newly_added))
            with col_l4:
                st.metric("Completed This Month", len(ledger_completed))

            # Diagnostic: active items but nothing released → period mismatch.
            # False-positive guard: if EVERY active item was newly discovered
            # this period, 0 releases is expected by design — merge_nexus()
            # always skips month-1 for a brand-new item (assumed covered by
            # Nexus's own accrual/reclass instead), with the first real
            # release starting next period. Only warn when at least one
            # active item is a carry-forward (not newly added) — that's the
            # case this diagnostic actually exists to catch: an old item that
            # should be mid-schedule but shows no release, signaling a stale
            # ledger upload.
            _newly_added_invs = {str(n or '').strip() for n in newly_added}
            _carry_forward_items = [
                _it for _it in ledger_active
                if str(_it.get('invoice_number', '') or '').strip() not in _newly_added_invs
            ]
            if _carry_forward_items and released_count == 0:
                from dateutil.relativedelta import relativedelta as _rdelta
                import re as _re
                next_fires = []
                for _item in _carry_forward_items:
                    _fap = _item.get('first_added_period', '')
                    _ma  = int(_item.get('months_amortized', 0) or 0)
                    _rem = int(_item.get('remaining_months', 0) or 0)
                    if _rem > 0 and _fap:
                        _m = _re.search(
                            r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[- ]?(\d{4})',
                            _fap, _re.IGNORECASE)
                        if _m:
                            _mo = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
                                   'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}.get(
                                _m.group(1).lower(), 0)
                            if _mo:
                                from datetime import date as _date
                                _anchor = _date(int(_m.group(2)), _mo, 1)
                                _nf = _anchor + _rdelta(months=_ma)
                                next_fires.append(
                                    f"{_item.get('vendor','?')} — next: {_nf.strftime('%b-%Y')}"
                                )
                st.warning(
                    f"⚠️ **{len(_carry_forward_items)} carried-forward prepaid item(s) but 0 "
                    f"released for {close_period}.** The ledger's `months_amortized` values "
                    f"don't match the current period — the uploaded ledger may be from a prior "
                    f"month. Upload the **updated Prepaid Ledger** from the previous close.\n\n"
                    + ("\n".join(f"• {f}" for f in next_fires[:8]) if next_fires else "")
                )
            if ledger_active:
                ledger_rows = [{
                    "Vendor": item.get('vendor', ''),
                    "Invoice #": item.get('invoice_number', ''),
                    "GL Account": item.get('gl_account_number', ''),
                    "Monthly Amt": item.get('monthly_amount', 0),
                    "Months Left": int(item.get('remaining_months', 0) or 0),
                    "Service End": str(item.get('service_end')) if item.get('service_end') else '',
                    "First Added": item.get('first_added_period', ''),
                } for item in ledger_active]
                st.dataframe(ledger_rows, use_container_width=True, hide_index=True,
                             column_config={
                                 "Vendor": st.column_config.TextColumn(width="medium"),
                                 "Invoice #": st.column_config.TextColumn(width="small"),
                                 "GL Account": st.column_config.TextColumn(width="small"),
                                 "Monthly Amt": st.column_config.NumberColumn(format="$%,.2f"),
                                 "Months Left": st.column_config.NumberColumn(width="small"),
                                 "Service End": st.column_config.TextColumn(width="small"),
                                 "First Added": st.column_config.TextColumn(width="small"),
                             })
            st.divider()

        # ── Budget-Based Accrual Review Flags ───────────────────────────────
        # HVAC / Fire Life Safety / Snow & Ice are accrued automatically from
        # the Kardin budget (see build_budget_based_accruals) instead of a
        # manually-entered One-Off Accruals amount. Snow & Ice in particular
        # always surfaces a flag here each active month (Nov-Mar) so a PM
        # confirms the estimate — or the real invoice, if one landed — is right.
        _budget_review_flags = p1.get("budget_review_flags", [])
        if _budget_review_flags:
            st.markdown("### ⚠️ Budget-Based Accrual — PM Review Needed")
            for _brf in _budget_review_flags:
                st.warning(_brf.get('message', ''), icon="⚠️")
            st.divider()

        # ── Download Section ───────────────────────────────────────────────
        st.markdown("### Download JE Files")
        st.caption("Upload these CSVs to Yardi, run the final close, then switch to **Pass 2** to generate reports.")

        # ── Accruals CSV content breakdown ──────────────────────────────
        _accrual_lines_display = [l for l in all_je_lines
                                  if l.get('source') in {
                                      'nexus', 'historical', 'management_fee',
                                      'management_fee_catchup', 'invoice_proration',
                                      'prepaid_amortization', 'contract_supplement',
                                      'tenant_utility_billing', 'bonus_accrual', 'prepaid_ledger',
                                      'prepaid_reclass', 'manual_addition', 'budget_based_accrual',
                                  }]
        _src_label_map = {
            'nexus':                  'Nexus AP',
            'invoice_proration':      'Invoice Proration',
            'historical':             'Historical Pattern',
            'prepaid_amortization':   'Prepaid Amort.',
            'prepaid_ledger':         'Prepaid Release',
            'prepaid_reclass':        'Prepaid Reclass',
            'management_fee':         'Management Fee',
            'management_fee_catchup': 'Mgmt Fee Catch-up',
            'contract_supplement':    'One-Off Accrual',
            'tenant_utility_billing': 'Tenant Utility',
            'bonus_accrual':          'Bonus Accrual',
            'manual_addition':        'Manually Added',
            'budget_based_accrual':   'Budget-Based Accrual',
        }
        # Count unique JEs (not lines) per source — DR lines only to avoid double-count
        _src_je_counts = {}
        for _l in _accrual_lines_display:
            if (_l.get('debit') or 0) > 0:
                _src = _l.get('source', 'other')
                _je_id = _l.get('je_number', '')
                _src_je_counts.setdefault(_src, set()).add(_je_id)
        _src_je_summary = {_src_label_map.get(s, s): len(ids)
                           for s, ids in _src_je_counts.items() if ids}

        if _src_je_summary:
            _ppd_count = _src_je_counts.get('prepaid_ledger', set())
            _total_je_in_csv = sum(len(v) for v in _src_je_counts.values())
            _breakdown_parts = [f"{lbl}: {cnt}" for lbl, cnt in _src_je_summary.items()]
            _breakdown_str = "  ·  ".join(_breakdown_parts)
            if _ppd_count:
                st.success(
                    f"**Accruals CSV contains {_total_je_in_csv} JEs** — {_breakdown_str}",
                    icon="✅",
                )
            else:
                st.info(
                    f"**Accruals CSV contains {_total_je_in_csv} JEs** — {_breakdown_str}  "
                    f"*(No prepaid releases — upload prior-month ledger and re-run if expected)*",
                    icon="📄",
                )

        # ── Prior Month Accrual vs Actuals ────────────────────────────────────
        _gl_for_check = result.parsed.get('gl') if result.parsed else None
        _prior_check  = check_prior_accrual_vs_actual(_gl_for_check) if _gl_for_check else []

        if _prior_check:
            _not_billed   = [r for r in _prior_check if r['status'] == 'NOT YET BILLED']
            _needs_review = [r for r in _prior_check if r['status'] in ('PARTIAL', 'OVER INVOICED')]
            _matched      = [r for r in _prior_check if r['status'] == 'MATCHED']

            _status_icon = '⚠️' if (_not_billed or _needs_review) else '✅'
            _review_note = ''
            if _not_billed:
                _review_note += f' · {len(_not_billed)} not yet billed (re-accrued)'
            if _needs_review:
                _review_note += f' · {len(_needs_review)} variance(s) to review'

            _check_label = (
                f"{_status_icon} Prior Month Accrual vs Actuals — "
                f"{len(_prior_check)} accounts{_review_note}"
            )
            with st.expander(_check_label, expanded=bool(_not_billed or _needs_review)):
                st.caption(
                    "Auto-reversals of last month's pipeline accruals compared to actual invoices "
                    "received this period. **NOT YET BILLED** = invoice hasn't arrived; pipeline "
                    "has re-accrued it. **MATCHED** = actual within 5% of accrual. "
                    "**PARTIAL / OVER INVOICED** = material variance — review before close."
                )

                _STATUS_EMOJI = {
                    'MATCHED':        '✅ MATCHED',
                    'NOT YET BILLED': '🔄 NOT YET BILLED',
                    'PARTIAL':        '⚠️ PARTIAL',
                    'OVER INVOICED':  '⚠️ OVER INVOICED',
                }
                _check_rows = [{
                    'Account':        f"{r['account_code']}  {r['account_name']}",
                    'Prior Accrual':  f"${r['reversal_amount']:,.2f}",
                    'Actual Billed':  f"${r['actual_amount']:,.2f}",
                    'Variance':       (f"+${r['variance']:,.2f}" if r['variance'] >= 0
                                       else f"-${abs(r['variance']):,.2f}"),
                    'Status':         _STATUS_EMOJI.get(r['status'], r['status']),
                    'JE Ref':         r['je_refs'],
                } for r in _prior_check]

                st.dataframe(
                    _check_rows,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        'Account':       st.column_config.TextColumn('Account',       width='medium'),
                        'Prior Accrual': st.column_config.TextColumn('Prior Accrual', width='small'),
                        'Actual Billed': st.column_config.TextColumn('Actual Billed', width='small'),
                        'Variance':      st.column_config.TextColumn('Variance',      width='small'),
                        'Status':        st.column_config.TextColumn('Status',        width='medium'),
                        'JE Ref':        st.column_config.TextColumn('JE Ref',        width='small'),
                    },
                )

                if _not_billed:
                    _nb_accts = ', '.join(
                        f"{r['account_code']} {r['account_name']}"
                        for r in _not_billed
                    )
                    st.warning(
                        f"**{len(_not_billed)} account(s) not yet billed:** {_nb_accts}. "
                        f"The pipeline has included re-accruals in the Accruals JE CSV. "
                        f"Confirm each was also re-accrued last month before posting.",
                        icon="🔄",
                    )
                if _needs_review:
                    _rv_parts = [
                        f"{r['account_code']} ({r['status']}: "
                        f"accrued ${r['reversal_amount']:,.0f}, "
                        f"billed ${r['actual_amount']:,.0f})"
                        for r in _needs_review
                    ]
                    st.warning(
                        f"**{len(_needs_review)} account(s) with material variance:** "
                        + ', '.join(_rv_parts) + '. Review before close.',
                        icon="⚠️",
                    )

        st.divider()

        # Zip of all 3 CSVs + updated ledger
        import zipfile, io
        _run_key = st.session_state.get('pass1_run_count', 0)
        period_label = (result.period or 'Period').replace('-', '_')
        p1_zip_files = {
            f"{_pfx_del}_{period_label}_Accruals_JE.csv":      p1.get("accrual_je_csv"),
            f"{_pfx_del}_{period_label}_Prepaid_Ledger.xlsx":  p1.get("prepaid_ledger_updated"),
            f"{_pfx_del}_{period_label}_JE_Cache.json":        p1.get("je_lines_cache"),
        }
        p1_zip_files = {k: v for k, v in p1_zip_files.items() if v and os.path.exists(v)}
        if p1_zip_files:
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                for fname, fpath in p1_zip_files.items():
                    zf.write(fpath, fname)
            zip_buf.seek(0)
            st.download_button(
                label=f"📦 Download All JE Files ({len(p1_zip_files)} files)",
                data=zip_buf,
                file_name=f"{_pfx_del}_{period_label}_JE_Package_{datetime.now().strftime('%Y%m%d')}.zip",
                mime="application/zip",
                key=f"dl_zip_{_run_key}",
                use_container_width=True,
            )

        col_d1, col_d2 = st.columns(2)
        _ts_p1 = datetime.now().strftime('%Y%m%d')
        for col, key, label, fname in [
            (col_d1, "accrual_je_csv",        "📄 Accruals JE",    f"{_pfx_int}_Accruals_JE_{_ts_p1}.csv"),
            (col_d2, "prepaid_ledger_updated", "📊 Prepaid Ledger", f"{_pfx_int}_Prepaid_Ledger_{_ts_p1}.xlsx"),
        ]:
            fpath = p1.get(key)
            if fpath and os.path.exists(fpath):
                with col:
                    with open(fpath, "rb") as f:
                        st.download_button(
                            label=label,
                            data=f.read(),
                            file_name=fname,
                            key=f"dl_{key}_{_run_key}",
                            use_container_width=True,
                        )

        st.divider()
        st.info(
            "📌 **Next step:** Upload the JE CSVs to Yardi and run the final close. "
            "Then switch to the **Pass 2 — Generate Reports** tab to produce the "
            "BS workpaper, QC checklist, and variance comments.\n\n"
            "💾 **If you plan to run Pass 2 in a new browser session**, save the "
            "**JE Cache (.json)** from the zip above — upload it in the Pass 2 upload "
            "section to enable full audit trail detail and JE posting verification.",
            icon="➡️",
        )


# ──────────────────────────────────────────────────────────────
# TAB 2 — PASS 2: REPORT GENERATION
# ──────────────────────────────────────────────────────────────
with tab2:
    # Recompute upload state (defined in tab1, available via session_state)
    uploaded_keys = set(st.session_state.uploaded_files.keys())
    gl_uploaded = "gl" in uploaded_keys

    st.markdown(
        "<div style='background:#E3F2FD;border-left:4px solid #1565C0;"
        "border-radius:5px;padding:8px 16px;margin-bottom:14px;font-size:0.85rem;color:#1565C0;'>"
        "⬤ &nbsp;<strong>Pass 2 — Post-Close</strong>&nbsp; Upload final Yardi exports, "
        "generate QC workbook, workpaper, variance comments, audit trail, and management fee invoice."
        "</div>",
        unsafe_allow_html=True,
    )

    st.markdown("""
    **What this does:** Reads the final post-close Yardi GL (after all JEs have been posted)
    and generates the GRP review deliverables — Balance Sheet workpaper, institutional workpapers,
    QC checklist, variance comments, and exception report.

    *(The Singerman monthly report is downloaded directly from Yardi — no need to generate it here.)*
    """)

    if not st.session_state.pass1_complete:
        st.info(
            "ℹ️ Pass 1 hasn't been run yet in this session. If you've already uploaded the "
            "JE CSVs to Yardi and have the final GL ready, you can still run Pass 2 independently.",
            icon="ℹ️",
        )

    # ── Close Status Banner ───────────────────────────────────────────────────
    # Full checklist lives on the Dashboard tab. Pass 2 shows a compact status
    # strip so the team can see where they stand without leaving the work surface.
    from close_tracker_generator import CLOSE_TRACKER_STEPS as _CT_STEPS

    _ct = st.session_state.close_tracker
    _ct_complete_count = sum(1 for i in range(len(_CT_STEPS)) if i in _ct)
    _ct_total = len(_CT_STEPS)
    _ct_pct   = int(100 * _ct_complete_count / _ct_total) if _ct_total else 0

    # Find the last completed step for the banner label
    _ct_last_done = max((i for i in range(_ct_total) if i in _ct), default=None)
    _ct_last_label = (
        _CT_STEPS[_ct_last_done][1] if _ct_last_done is not None else None
    )

    # Choose banner colour based on completion
    if _ct_pct == 100:
        _cts_bg, _cts_border, _cts_color = '#E8F5E9', '#2E7D32', '#1B5E20'
        _cts_icon = '✅'
    elif _ct_pct >= 50:
        _cts_bg, _cts_border, _cts_color = '#E3F2FD', '#1565C0', '#0D47A1'
        _cts_icon = '🔄'
    else:
        _cts_bg, _cts_border, _cts_color = '#FFF8E1', '#E65100', '#BF360C'
        _cts_icon = '⏳'

    _cts_last_html = (
        f"&nbsp;&nbsp;·&nbsp;&nbsp;Last: <em>{_ct_last_label}</em>"
        if _ct_last_label else ""
    )
    _cts_period_label = period_key_to_label(
        st.session_state.get('checklist_period_key', current_period_key())
    )

    st.markdown(
        f"<div style='background:{_cts_bg};border-left:4px solid {_cts_border};"
        f"border-radius:5px;padding:10px 16px;margin-bottom:10px;"
        f"font-size:0.85rem;color:{_cts_color};display:flex;"
        f"align-items:center;gap:12px;'>"
        f"{_cts_icon}&nbsp;&nbsp;"
        f"<strong>{_cts_period_label} Close</strong>&nbsp;&nbsp;·&nbsp;&nbsp;"
        f"{_ct_complete_count} / {_ct_total} steps complete ({_ct_pct}%)"
        f"{_cts_last_html}"
        f"&nbsp;&nbsp;·&nbsp;&nbsp;"
        f"<span style='opacity:0.75;'>Full checklist → Dashboard tab</span>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # Export / download Close Tracker XLSX (functional action — stays here)
    _ct_exp_col, _ct_dl_col, _ = st.columns([2, 2, 3])
    with _ct_exp_col:
        if st.button("📄 Export Close Tracker", use_container_width=True,
                     help=f"Generates {_pfx_int}_Close_Tracker.xlsx and adds it to the ZIP"):
            try:
                from close_tracker_generator import generate_close_tracker_xlsx as _gen_ct2
                _ct_xlsx_path2 = os.path.join(
                    st.session_state.temp_dir, f"{_pfx_int}_Close_Tracker.xlsx"
                )
                _p2r = st.session_state.pass2_engine_result
                _ct_period2 = (_p2r.period if _p2r
                               else st.session_state.get('close_period_input', 'Period'))
                _ct_prop2   = (_p2r.property_name if _p2r else _prop_display)
                _gen_ct2(
                    output_path   = _ct_xlsx_path2,
                    close_tracker = st.session_state.close_tracker,
                    period        = _ct_period2,
                    property_name = _ct_prop2,
                )
                st.session_state.pass2_output_files["close_tracker"] = _ct_xlsx_path2
                st.success("Close Tracker exported — included in the ZIP package.", icon="✅")
                st.rerun()
            except Exception as _ct_e2:
                st.error(f"Close Tracker export failed: {_ct_e2}")

    with _ct_dl_col:
        _ct_dl_path = st.session_state.pass2_output_files.get("close_tracker")
        if _ct_dl_path and os.path.exists(_ct_dl_path):
            with open(_ct_dl_path, "rb") as _ct_f:
                st.download_button(
                    label="⬇️ Download Close Tracker",
                    data=_ct_f.read(),
                    file_name=f"{_pfx_int}_Close_Tracker.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

    st.divider()

    # ── Pass 2: Bulk file upload ──────────────────────────────────────────────
    st.markdown("#### Upload Pass 2 Files")
    st.caption(
        "Drop all post-close files here at once — Final GL, Budget Comparison, "
        "Trial Balance, T12, Berkadia loan statements, and prior workpaper. "
        "The app auto-detects each file type."
    )

    # ── Pass 2 bulk uploader session state ───────────────────────────────────
    if "bulk_overrides_p2" not in st.session_state:
        st.session_state.bulk_overrides_p2 = {}

    _P2_SLOT_KEYS = [
        "gl_pass2", "budget_comparison_pass2", "trial_balance_pass2",
        "t12_statement_pass2", "loan_pass2", "bank_rec_pass2",
        "prior_workpaper", "bank_rec_dev_xlsx", "unknown",
    ]
    _P2_SLOT_LABELS = [_FILE_LABELS.get(k, k) for k in _P2_SLOT_KEYS]

    _bulk_p2 = st.file_uploader(
        "Drop all Pass 2 files here",
        accept_multiple_files=True,
        type=["xlsx", "xls", "pdf"],
        key=f"bulk_upload_p2_{st.session_state.upload_key_p2}",
    )

    # Only clear/reclassify Pass 2 slots when the uploaded file set actually
    # changed — see the matching Pass 1 fingerprint gate for why running this
    # unconditionally on every rerun risks silently dropping files.
    _bulk_p2_fingerprint = tuple(sorted((f.name, f.size) for f in (_bulk_p2 or [])))
    if st.session_state.get('_bulk_p2_fingerprint') != _bulk_p2_fingerprint:
        st.session_state['_bulk_p2_fingerprint'] = _bulk_p2_fingerprint

        # Clear Pass 2 slots so stale entries don't persist after file removal
        for _clr_k2 in set(_P2_SLOT_KEYS) - {"unknown"}:
            st.session_state.uploaded_files.pop(_clr_k2, None)

    if _bulk_p2:
        _loan_paths_p2: list = []

        for _uf2 in _bulk_p2:
            _raw2 = bytes(_uf2.getbuffer())
            _det_key2, _conf2, _det_label2 = _classify_file(_uf2.name, _raw2, pass2=True, property_config=_active_cfg)

            # If classifier returned a base key that isn't in the P2 slot list,
            # keep it as-is (e.g. prior_workpaper has no remap)
            if _det_key2 not in _P2_SLOT_KEYS:
                _det_key2 = "unknown"

            # B-6: key by (name, size) so re-uploading the same filename with different
            # content gets a fresh key and the stale manual override doesn't survive.
            _ovr_key2 = (_uf2.name, _uf2.size)
            _eff_key2 = st.session_state.bulk_overrides_p2.get(_ovr_key2, _det_key2)

            _ic2, _fn2, _tp2 = st.columns([1, 3, 4])
            if _eff_key2 == "unknown":
                _ic2.markdown("⚠️")
            elif _conf2 >= 0.85:
                _ic2.markdown("✅")
            else:
                _ic2.markdown("🟡")

            _short2 = _uf2.name if len(_uf2.name) <= 22 else _uf2.name[:19] + "…"
            _fn2.caption(_short2)

            if _eff_key2 == "unknown" or _conf2 < 0.70:
                _cur_idx2 = (_P2_SLOT_KEYS.index(_eff_key2)
                             if _eff_key2 in _P2_SLOT_KEYS else len(_P2_SLOT_KEYS) - 1)
                _sel_label2 = _tp2.selectbox(
                    "type", _P2_SLOT_LABELS, index=_cur_idx2,
                    key=f"ovr_p2_{_uf2.name}", label_visibility="collapsed",
                )
                _eff_key2 = _P2_SLOT_KEYS[_P2_SLOT_LABELS.index(_sel_label2)]
                st.session_state.bulk_overrides_p2[_ovr_key2] = _eff_key2
            else:
                _tp2.caption(_det_label2)

            if _eff_key2 != "unknown":
                _tp2_path = os.path.join(st.session_state.temp_dir, f"p2_{_uf2.name}")
                if not os.path.exists(_tp2_path) or os.path.getsize(_tp2_path) != _uf2.size:
                    with open(_tp2_path, "wb") as _f2:
                        _f2.write(_raw2)
                if _eff_key2 in _MULTI_FILE_KEYS:
                    _loan_paths_p2.append(_tp2_path)
                else:
                    st.session_state.uploaded_files[_eff_key2] = _tp2_path

        if _loan_paths_p2:
            st.session_state.uploaded_files["loan_pass2"] = _loan_paths_p2

        # Prune stale overrides for files no longer present (name+size composite key)
        _active_keys_p2 = {(_uf2.name, _uf2.size) for _uf2 in _bulk_p2}
        st.session_state.bulk_overrides_p2 = {
            k: v for k, v in st.session_state.bulk_overrides_p2.items()
            if k in _active_keys_p2
        }

        # Show fallback notes for optional slots not yet uploaded
        _p2_loaded = set(st.session_state.uploaded_files.keys())
        if "budget_comparison_pass2" not in _p2_loaded and "budget_comparison" in _p2_loaded:
            st.caption("↳ BC: using Pass 1 version — upload post-close BC above to override")
        if "trial_balance_pass2" not in _p2_loaded and "trial_balance" in _p2_loaded:
            st.caption("↳ TB: using Pass 1 version — upload post-close TB above to override")
        if "t12_statement_pass2" not in _p2_loaded and "t12_statement" in _p2_loaded:
            st.caption("↳ T12: using Pass 1 version — upload post-close T12 above to override")
        if "loan_pass2" not in _p2_loaded and "loan" in _p2_loaded:
            st.caption("↳ Loan: using Pass 1 version — upload post-close PDFs above to override")

        # File count for Pass 2 — list slots (e.g. loan_pass2 with multiple PDFs)
        # are unfolded so 3 Berkadia PDFs count as 3, not 1.
        _p2_count_keys = set(_P2_SLOT_KEYS) - {"unknown"}
        _p2_file_count = sum(
            len(v) if isinstance(v, list) else 1
            for k, v in st.session_state.uploaded_files.items()
            if k in _p2_count_keys and v is not None
        )
        if _p2_file_count > 0:
            st.caption(f"**{_p2_file_count} file(s) uploaded**")

    st.text_input(
        "Prior period label",
        placeholder="e.g. Mar-2026  (leave blank to auto-detect)",
        help=(
            "The period that the uploaded prior workpaper covers. "
            "Required when uploading a workpaper that doesn't have period-prefixed tabs. "
            "Format: Mon-YYYY (e.g. Mar-2026, Feb-2026)."
        ),
        key="prior_period_label_input",
    )

    _rl_upload = st.file_uploader(
        "Prior Run Log (optional)",
        type=["csv"],
        key=f"run_log_upload_{st.session_state.upload_key_p2}",
        help="Upload the GA_Run_Log.csv from last month to carry forward the history.",
    )
    if _rl_upload:
        _rl_save = os.path.join(st.session_state.temp_dir, "prior_run_log.csv")
        with open(_rl_save, "wb") as _f:
            _f.write(_rl_upload.read())
        st.session_state.uploaded_files["run_log"] = _rl_save

    _jec_upload = st.file_uploader(
        "Pass 1 JE Cache (optional)",
        type=["json"],
        key=f"je_cache_upload_{st.session_state.upload_key_p2}",
        help=(
            "Upload the JE_Cache.json from this month's Pass 1 package. "
            "Required only when Pass 2 is running in a different browser session from Pass 1 — "
            "enables full audit trail detail (Tab 2 & Tab 6) and JE posting verification."
        ),
    )
    if _jec_upload:
        _jec_save = os.path.join(st.session_state.temp_dir, "je_lines_cache.json")
        with open(_jec_save, "wb") as _f:
            _f.write(_jec_upload.read())
        st.session_state.uploaded_files["je_lines_cache"] = _jec_save

    # ── Additional workpaper source files (raw Yardi reports per account) ──────
    # These replace the generated GL transaction tabs for the named accounts with
    # the raw Yardi export pasted directly into the workpaper.
    with st.expander("📋 Workpaper raw report overrides (optional)", expanded=False):
        st.caption(
            "Drop raw Yardi exports here — each file is assigned a type and copied "
            "directly into the matching workpaper tab. AR Aging and Capital Schedule "
            "auto-source from Pass 1 when available. Bank Rec also accepts a PDF if "
            "you don't have access to the Excel export — it's parsed the same way as "
            "the main Bank Rec PDF upload. The other report types need the Excel "
            "version; there's no PDF parser for them yet."
        )

        # ── WP bulk uploader session state ────────────────────────────────────
        if "bulk_overrides_wp" not in st.session_state:
            st.session_state.bulk_overrides_wp = {}

        _WP_SLOT_KEYS = [
            "ar_aging_pass2", "ap_aging", "bank_rec_xlsx", "daca_bank_rec_xlsx",
            "capital_schedule_pass2", "prepaid_ledger_p2", "capital_seed", "unknown",
        ]
        _WP_SLOT_LABELS = [
            "AR Aging Detail — 133100 AR Control",
            "AP Aging Detail — 211300 AP Control",
            "Bank Rec Excel — 111100 PNC Operating",
            "Bank Rec Excel — 115100 DACA",
            "Capital Accounts Schedule",
            "Prepaid Ledger Updated",
            "Capital Schedule Seed",
            "Unknown — select type",
        ]
        # PDF is only usable for the two bank rec slots — those are the only
        # report types here with an existing PDF parser (yardi_bank_rec.py /
        # yardi_daca_rec.py, the same ones the main Bank Rec PDF upload uses).
        # AR/AP Aging, Capital Schedule, and Prepaid Ledger have no PDF parser
        # at all, so a PDF assigned to one of those would just be silently
        # unusable — narrow the dropdown instead of letting that happen.
        _WP_PDF_SLOT_KEYS = ["bank_rec_xlsx", "daca_bank_rec_xlsx", "unknown"]
        _WP_PDF_SLOT_LABELS = [
            "Bank Rec PDF — 111100 PNC Operating",
            "Bank Rec PDF — 115100 DACA",
            "Unknown — select type",
        ]

        _bulk_wp = st.file_uploader(
            "Drop all workpaper override files here",
            accept_multiple_files=True,
            type=["xlsx", "xls", "pdf"],
            key=f"bulk_upload_wp_{st.session_state.get('upload_key_p2', 0)}",
        )

        # Only clear WP slots when the uploaded file set actually changed —
        # see the matching Pass 1 fingerprint gate for why running this
        # unconditionally on every rerun risks silently dropping files.
        _bulk_wp_fingerprint = tuple(sorted((f.name, f.size) for f in (_bulk_wp or [])))
        if st.session_state.get('_bulk_wp_fingerprint') != _bulk_wp_fingerprint:
            st.session_state['_bulk_wp_fingerprint'] = _bulk_wp_fingerprint
            for _clr_kwp in set(_WP_SLOT_KEYS) - {"unknown"}:
                st.session_state.uploaded_files.pop(_clr_kwp, None)
            st.session_state.pop('wp_override_bank_rec_data', None)
            st.session_state.pop('wp_override_daca_rec_data', None)

        if _bulk_wp:
            for _ufwp in _bulk_wp:
                _raw_wp = bytes(_ufwp.getbuffer())
                _is_pdf_wp = _ufwp.name.lower().endswith('.pdf')
                # These are custom Yardi reports — auto-detection is unreliable;
                # always show the type selectbox for explicit user assignment.
                _ovr_key_wp = (_ufwp.name, _ufwp.size)  # B-F4: composite key prevents stale classifications
                _eff_key_wp = st.session_state.bulk_overrides_wp.get(_ovr_key_wp, "unknown")

                _slot_keys_wp   = _WP_PDF_SLOT_KEYS if _is_pdf_wp else _WP_SLOT_KEYS
                _slot_labels_wp = _WP_PDF_SLOT_LABELS if _is_pdf_wp else _WP_SLOT_LABELS
                if _is_pdf_wp and _eff_key_wp not in _slot_keys_wp:
                    _eff_key_wp = "unknown"  # a prior Excel assignment doesn't carry over to a PDF re-upload

                _ic_wp, _fn_wp, _tp_wp = st.columns([1, 3, 4])
                _ic_wp.markdown("✅" if _eff_key_wp != "unknown" else "⚠️")

                _short_wp = _ufwp.name if len(_ufwp.name) <= 22 else _ufwp.name[:19] + "…"
                _fn_wp.caption(_short_wp)

                _cur_idx_wp = (
                    _slot_keys_wp.index(_eff_key_wp)
                    if _eff_key_wp in _slot_keys_wp
                    else len(_slot_keys_wp) - 1
                )
                _sel_label_wp = _tp_wp.selectbox(
                    "type", _slot_labels_wp, index=_cur_idx_wp,
                    key=f"ovr_wp_{_ufwp.name}_{_ufwp.size}", label_visibility="collapsed",
                )
                _eff_key_wp = _slot_keys_wp[_slot_labels_wp.index(_sel_label_wp)]
                st.session_state.bulk_overrides_wp[_ovr_key_wp] = _eff_key_wp

                if _is_pdf_wp and _eff_key_wp in ("bank_rec_xlsx", "daca_bank_rec_xlsx"):
                    # Parse right here rather than saving a filepath — the raw_filepath
                    # slots downstream expect a raw-copyable Excel sheet, which a PDF
                    # isn't. Route the parsed result into the same bank_rec_data/
                    # daca_bank_data shape the main PDF upload already produces, so the
                    # existing pdf_gl_transactions fallback tier in
                    # bs_workpaper_generator.py renders it without any new code there.
                    _wp_pdf_path = os.path.join(st.session_state.temp_dir, f"wp_{_ufwp.name}")
                    if not os.path.exists(_wp_pdf_path) or os.path.getsize(_wp_pdf_path) != _ufwp.size:
                        with open(_wp_pdf_path, "wb") as _f_wp_pdf:
                            _f_wp_pdf.write(_raw_wp)
                    try:
                        if _eff_key_wp == "bank_rec_xlsx":
                            from parsers.yardi_bank_rec import parse as _parse_bank_rec_wp
                            st.session_state['wp_override_bank_rec_data'] = _parse_bank_rec_wp(_wp_pdf_path)
                        else:
                            from parsers.yardi_daca_rec import parse as _parse_daca_rec_wp
                            st.session_state['wp_override_daca_rec_data'] = _parse_daca_rec_wp(_wp_pdf_path)
                    except Exception as _pdf_wp_err:
                        st.warning(f"Could not parse {_ufwp.name} as a Bank Rec PDF: {_pdf_wp_err}")
                elif _eff_key_wp != "unknown":
                    _wp_path = os.path.join(st.session_state.temp_dir, f"wp_{_ufwp.name}")
                    if not os.path.exists(_wp_path) or os.path.getsize(_wp_path) != _ufwp.size:
                        with open(_wp_path, "wb") as _f_wp:
                            _f_wp.write(_raw_wp)
                    st.session_state.uploaded_files[_eff_key_wp] = _wp_path

            # Auto-source status for optional slots not yet uploaded
            _wp_loaded = set(st.session_state.uploaded_files.keys())
            if "ar_aging_pass2" not in _wp_loaded:
                _p1_ar = st.session_state.uploaded_files.get("ar_aging")
                if _p1_ar and os.path.exists(_p1_ar):
                    st.caption("↳ AR Aging: auto-sourced from Pass 1")
                else:
                    st.caption("⚠️ AR Aging not uploaded — 133100 tab will show GL transactions")
            if "capital_schedule_pass2" not in _wp_loaded:
                _p1_cap = st.session_state.uploaded_files.get("capital_schedule")
                if _p1_cap and os.path.exists(_p1_cap):
                    st.caption("↳ Capital Schedule: auto-sourced from Pass 1")
                else:
                    st.caption("⚠️ Capital Schedule not uploaded — capital tabs show GL transactions")

            # File count
            _wp_count = sum(
                1 for k in _WP_SLOT_KEYS
                if k != "unknown" and st.session_state.uploaded_files.get(k)
            )
            if _wp_count > 0:
                st.caption(f"**{_wp_count} file(s) assigned**")

            # Clean up overrides for files no longer in uploader (composite key)
            _active_wp = {(_ufwp.name, _ufwp.size) for _ufwp in _bulk_wp}
            st.session_state.bulk_overrides_wp = {
                k: v for k, v in st.session_state.bulk_overrides_wp.items()
                if k in _active_wp
            }

        st.caption(
            "AR Aging auto-sources from Pass 1 in the same session — only re-upload above "
            "if starting Pass 2 in a new browser session.  "
            "Prepaid Ledger falls back to same-session Pass 1 data if not re-uploaded.  "
            "Capital Seed (Book3.xlsx) bootstraps January capital tabs — ignored once a "
            "prior workpaper or current-period Capital Schedule is uploaded."
        )

    # ── Manual Prepaid Add Form ───────────────────────────────────────────────
    # Allows adding prepaid items directly in Pass 2 without re-uploading the ledger.
    # Items are appended to _prepaid_active before the workpaper is generated.
    _EXCLUDED_PREPAID_GL = {'639110', '639120', '641110'}  # Insurance + RE Tax (handled separately)

    if 'pass2_manual_prepaids' not in st.session_state:
        st.session_state.pass2_manual_prepaids = []

    with st.expander("➕ Manually add prepaid items (optional)", expanded=False):
        st.caption(
            "Add items that are not in the uploaded prepaid ledger — e.g. new contracts "
            "signed mid-month. Each item is appended to the active ledger for this run only. "
            "GL accounts 639110 (Insurance), 639120 (Insurance), and 641110 (RE Tax) are "
            "excluded — those are managed by the prepaid insurance and RE-tax modules."
        )

        # Show existing manual items with remove buttons
        if st.session_state.pass2_manual_prepaids:
            st.markdown("**Items added this session:**")
            _remove_idx = None
            for _mi, _mitem in enumerate(st.session_state.pass2_manual_prepaids):
                _mc1, _mc2 = st.columns([8, 1])
                _mc1.markdown(
                    f"**{_mitem['vendor']}** — {_mitem['gl_account_number']} "
                    f"| ${_mitem['monthly_amount']:,.2f}/mo × {_mitem['total_months']} mo "
                    f"| Start: {_mitem['service_start']} "
                    f"| Amortized: {_mitem['months_amortized']}"
                )
                if _mc2.button("✕", key=f"remove_manual_prepaid_{_mi}", help="Remove this item"):
                    _remove_idx = _mi
            if _remove_idx is not None:
                st.session_state.pass2_manual_prepaids.pop(_remove_idx)
                st.rerun()
            st.divider()

        with st.form("manual_prepaid_form_p2", clear_on_submit=True):
            st.markdown("**Add new prepaid item**")

            _fp_c1, _fp_c2 = st.columns(2)
            _fp_vendor = _fp_c1.text_input("Vendor *", placeholder="e.g. Acme Insurance Co.")
            _fp_invoice_number = _fp_c2.text_input("Invoice # (optional)", placeholder="e.g. INV-2026-001")

            _fp_c3, _fp_c4 = st.columns(2)
            _fp_gl_number = _fp_c3.text_input(
                "GL Account # *", placeholder="e.g. 635110",
                help="6-digit GL account code. 639110, 639120, 641110 are excluded.",
            )
            _fp_gl_label = _fp_c4.text_input(
                "GL Account Name", placeholder="e.g. Repairs & Maintenance",
                help="Human-readable account label (optional).",
            )

            _fp_desc = st.text_input(
                "Description *", placeholder="e.g. Annual maintenance contract — Jan–Dec 2026",
            )

            _fp_c5, _fp_c6, _fp_c7 = st.columns(3)
            _fp_start = _fp_c5.date_input("Service Start *", value=None)
            _fp_end = _fp_c6.date_input("Service End *", value=None)
            _fp_monthly = _fp_c7.number_input(
                "Monthly Amount ($) *", min_value=0.01, value=None,
                format="%.2f",
                help="Amount to release per month (mid-month starts are prorated automatically).",
            )

            _fp_c8, _fp_c9 = st.columns(2)
            _fp_amortized = _fp_c8.number_input(
                "Months Already Amortized",
                min_value=0, value=0, step=1,
                help="Enter 0 for a brand-new item. Enter N if N months have already been released.",
            )
            _fp_c9.markdown("")  # spacer

            _fp_submitted = st.form_submit_button("Add to Ledger", use_container_width=True)

        if _fp_submitted:
            # --- Validation ---
            _fp_errors = []
            if not _fp_vendor.strip():
                _fp_errors.append("Vendor is required.")
            if not _fp_desc.strip():
                _fp_errors.append("Description is required.")
            if not _fp_gl_number.strip():
                _fp_errors.append("GL Account # is required.")
            elif not _fp_gl_number.strip().isdigit():
                _fp_errors.append("GL Account # must be numeric (e.g. 635110).")
            elif _fp_gl_number.strip() in _EXCLUDED_PREPAID_GL:
                _fp_errors.append(
                    f"GL account {_fp_gl_number.strip()} is managed by a separate module "
                    "(insurance / RE tax). Use the One-Off Accruals table instead."
                )
            if _fp_monthly is None or _fp_monthly <= 0:
                _fp_errors.append("Monthly Amount must be greater than $0.")
            if _fp_start is None:
                _fp_errors.append("Service Start date is required.")
            if _fp_end is None:
                _fp_errors.append("Service End date is required.")
            if _fp_start and _fp_end and _fp_end <= _fp_start:
                _fp_errors.append("Service End must be after Service Start.")

            if _fp_errors:
                for _fe in _fp_errors:
                    st.error(_fe)
            else:
                from datetime import date as _date_cls
                from dateutil.relativedelta import relativedelta as _rdelta
                import calendar as _cal

                # Compute derived fields
                _fp_start_d = _fp_start if isinstance(_fp_start, _date_cls) else _fp_start
                _fp_end_d   = _fp_end   if isinstance(_fp_end, _date_cls)   else _fp_end

                # Total months: inclusive count (same as prepaid_ledger._count_months)
                _fp_rd = _rdelta(_fp_end_d, _fp_start_d)
                _fp_total_months = _fp_rd.years * 12 + _fp_rd.months + 1

                _fp_total_amount    = round(_fp_monthly * _fp_total_months, 2)
                _fp_remaining       = max(0, _fp_total_months - _fp_amortized)
                _fp_daily_rate      = round(_fp_monthly / _cal.monthrange(_fp_start_d.year, _fp_start_d.month)[1], 4)

                # first_added_period = current close period (YYYY-MM)
                # Try Pass 2 engine result → Pass 1 engine result → today
                import datetime as _dt_mod
                _close_period_str = ''
                _p2r_for_period = st.session_state.get('pass2_engine_result')
                if _p2r_for_period and getattr(_p2r_for_period, 'period', None):
                    _close_period_str = _p2r_for_period.period
                if not _close_period_str:
                    _p1r_for_period = st.session_state.get('pass1_engine_result')
                    if _p1r_for_period and getattr(_p1r_for_period, 'period', None):
                        _close_period_str = _p1r_for_period.period
                if not _close_period_str:
                    _close_period_str = _dt_mod.date.today().strftime('%Y-%m')

                _new_prepaid_item = {
                    'vendor':             _fp_vendor.strip(),
                    'invoice_number':     _fp_invoice_number.strip(),
                    'invoice_date':       None,
                    'description':        _fp_desc.strip(),
                    'gl_account_number':  _fp_gl_number.strip(),
                    'gl_account':         (
                        f"{_fp_gl_label.strip()} ({_fp_gl_number.strip()})"
                        if _fp_gl_label.strip()
                        else _fp_gl_number.strip()
                    ),
                    'total_amount':       _fp_total_amount,
                    'monthly_amount':     round(_fp_monthly, 2),
                    'service_start':      _fp_start_d,
                    'service_end':        _fp_end_d,
                    'total_months':       _fp_total_months,
                    'months_amortized':   _fp_amortized,
                    'remaining_months':   _fp_remaining,
                    'first_added_period': _close_period_str,
                    'daily_rate':         _fp_daily_rate,
                }
                st.session_state.pass2_manual_prepaids.append(_new_prepaid_item)
                st.success(
                    f"✅ Added: **{_fp_vendor.strip()}** — "
                    f"${_fp_monthly:,.2f}/mo × {_fp_total_months} months "
                    f"(GL {_fp_gl_number.strip()})"
                )
                st.rerun()

    # Pass 2 requires either a dedicated post-close GL or at minimum the sidebar GL
    _p2_gl_ready = (
        "gl_pass2" in st.session_state.uploaded_files
        or gl_uploaded
    )

    st.divider()

    # ── Pass 2 Run Button ─────────────────────────────────────────────────────
    col_p2a, col_p2b = st.columns([3, 1])
    with col_p2a:
        pass2_button = st.button(
            "📊 Generate Reports",
            disabled=not _p2_gl_ready,
            use_container_width=True,
            key="pass2_run_btn",
            help="Parse final post-close GL and generate all workpapers and reports",
        )
    with col_p2b:
        if not st.session_state.confirm_reset_p2:
            if st.button("🔄 Reset Pass 2", use_container_width=True, key="reset_pass2"):
                st.session_state.confirm_reset_p2 = True
                st.rerun()
        else:
            st.warning("⚠️ Clears results — uploaded files stay loaded.")
            _rp2_col1, _rp2_col2 = st.columns(2)
            if _rp2_col1.button("✅ Confirm", use_container_width=True, key="confirm_reset_p2_btn"):
                st.session_state.confirm_reset_p2 = False
                st.session_state.pass2_complete = False
                st.session_state.pass2_engine_result = None
                st.session_state.pass2_output_files = {}
                st.session_state.pass2_manual_prepaids = []   # B-F1: prevent cross-property bleed
                st.session_state.signoff_state = {}
                import pandas as _pd_r2
                st.session_state.post_close_je_df = _pd_r2.DataFrame({
                    "JE #": ["PC-001", "PC-001"], "Description": ["", ""],
                    "Account Code": ["", ""],
                    "Debit ($)": [0.0, 0.0], "Credit ($)": [0.0, 0.0],
                    "Line Description": ["", ""],
                })
                st.session_state.pop("_pcje_latest", None)
                # Clear Pass 2 close tracker steps (5=files uploaded, 6=reports generated,
                # 7=QC review complete, 8=package released). Steps 0-4 are pre-Pass-2 and
                # should be preserved since the work already happened.
                for _ct_step in (5, 6, 7, 8):
                    st.session_state.close_tracker.pop(_ct_step, None)
                # NOTE: upload_key_p2 is NOT incremented — uploaded files stay in
                # the uploader widgets so Generate Reports can be re-run immediately
                # without re-dropping any files.
                st.rerun()
            if _rp2_col2.button("❌ Cancel", use_container_width=True, key="cancel_reset_p2_btn"):
                st.session_state.confirm_reset_p2 = False
                st.rerun()

    # ── Pass 2 Processing ─────────────────────────────────────────────────────
    if pass2_button:
        with st.spinner("Generating reports from final GL..."):
            try:
                # Build files dict from shared sidebar uploads, then override
                # with any Pass 2-specific files that were uploaded above.
                files_dict = {key: st.session_state.uploaded_files.get(key)
                              for key in file_config.keys()}
                if st.session_state.uploaded_files.get("gl_pass2"):
                    files_dict["gl"] = st.session_state.uploaded_files["gl_pass2"]
                if st.session_state.uploaded_files.get("budget_comparison_pass2"):
                    files_dict["budget_comparison"] = st.session_state.uploaded_files["budget_comparison_pass2"]
                if st.session_state.uploaded_files.get("trial_balance_pass2"):
                    files_dict["trial_balance"] = st.session_state.uploaded_files["trial_balance_pass2"]
                if st.session_state.uploaded_files.get("loan_pass2"):
                    files_dict["loan"] = st.session_state.uploaded_files["loan_pass2"]
                if st.session_state.uploaded_files.get("bank_rec_pass2"):
                    files_dict["bank_rec"] = st.session_state.uploaded_files["bank_rec_pass2"]
                if st.session_state.uploaded_files.get("t12_statement_pass2"):
                    files_dict["t12_statement"] = st.session_state.uploaded_files["t12_statement_pass2"]

                # Auto-load committed Kardin budget if not uploaded this session
                if not files_dict.get("kardin_budget") and _COMMITTED_BUDGET:
                    files_dict["kardin_budget"] = _COMMITTED_BUDGET

                progress_bar = st.progress(0)
                status_text  = st.empty()

                # Step 1: Parse final (post-close) GL
                status_text.text("Step 1/6: Parsing final GL...")
                progress_bar.progress(10)
                engine_result = run_pipeline(files_dict)
                st.session_state.pass2_engine_result = engine_result

                gl_parsed    = engine_result.parsed.get('gl')
                bc_parsed    = engine_result.parsed.get('budget_comparison') or []
                close_period = engine_result.period or ''

                # Guard: period must be detected before we can label any outputs.
                if not close_period:
                    st.error(
                        "⚠️ **Could not determine close period from final GL.**  "
                        "The period label is required for all output file names. "
                        "Ensure you have uploaded the final (post-close) Yardi GL export.",
                        icon="❌",
                    )
                    st.session_state.pass2_complete = False
                    st.stop()

                # Initialise variables that may be assigned later in conditional
                # parse blocks — ensures they are always defined if those blocks
                # are skipped due to missing files or upstream exceptions.
                _ar_aging_parsed_p2   = None
                _capital_schedule_data = None
                tb_result              = None
                t12_result             = None
                bank_rec_data          = None
                daca_bank_data         = None
                dev_bank_rec_data      = None
                gl_cash_balance        = None
                daca_gl_balance        = None

                # ── Loan statement date validation (Pass 2) ───────────────────
                _loan_stmts_p2 = engine_result.parsed.get('loan') or []
                if not isinstance(_loan_stmts_p2, list):
                    _loan_stmts_p2 = [_loan_stmts_p2]
                _cp_mo_p2 = next(
                    (v for k, v in dict(Jan=1,Feb=2,Mar=3,Apr=4,May=5,Jun=6,
                                        Jul=7,Aug=8,Sep=9,Oct=10,Nov=11,Dec=12).items()
                     if k in close_period), 0
                )
                _exp_due_mo_p2 = (_cp_mo_p2 % 12) + 1
                _cp_yr_p2 = re.search(r'\d{4}', close_period)
                _cp_yr_p2 = int(_cp_yr_p2.group()) if _cp_yr_p2 else 0
                _exp_due_yr_p2 = _cp_yr_p2 + 1 if _cp_mo_p2 == 12 else _cp_yr_p2
                for _ls2 in _loan_stmts_p2:
                    if not isinstance(_ls2, dict):
                        continue
                    _due2 = str(_ls2.get('payment_due_date') or '')
                    if not _due2:
                        continue
                    _due_parts2 = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', _due2)
                    if not _due_parts2:
                        continue
                    _due_mo2, _due_yr2 = int(_due_parts2.group(1)), int(_due_parts2.group(3))
                    if _due_mo2 != _exp_due_mo_p2 or _due_yr2 != _exp_due_yr_p2:
                        _ln2 = _ls2.get('loan_number', 'unknown')
                        engine_result.exceptions.append(Exception_(
                            severity='warning',
                            category='loan',
                            source='berkadia_loan',
                            description=(
                                f"Loan {_ln2}: statement due date {_due2} does not match "
                                f"the {close_period} close. For {close_period}, upload the "
                                f"statement due {_exp_due_mo_p2:02d}/07/{_exp_due_yr_p2} "
                                f"(interest accrued in {close_period} is paid on the 7th of "
                                f"the following month)."
                            ),
                            details={
                                'loan_number': _ln2,
                                'uploaded_due_date': _due2,
                                'expected_due_month': f'{_exp_due_mo_p2:02d}/{_exp_due_yr_p2}',
                            },
                        ))

                # Step 2: Parse trial balance + BS workpaper (no je_adjustments — GL is final)
                # Note: Singerman monthly report is downloaded directly from Yardi by Ryan.
                status_text.text("Step 2/6: Generating BS workpaper...")
                progress_bar.progress(25)

                tb_result = None
                # Resolve TB path: Pass 2 upload takes priority over sidebar upload
                _tb_file = (st.session_state.uploaded_files.get("trial_balance_pass2")
                            or st.session_state.uploaded_files.get("trial_balance"))
                if _tb_file:
                    try:
                        from parsers.yardi_trial_balance import parse as parse_tb
                        tb_result = parse_tb(_tb_file)
                    except Exception as _e:
                        st.warning(f"Could not parse Trial Balance: {_e}")

                # Parse 12-Month Statement — Pass 2 upload takes priority over sidebar
                # Post-close T12 has the current period's JEs baked in, so MoM prior month
                # is sourced from T12 and current month from BC PTD Actual (post-close GL).
                t12_result = None
                _t12_file = (st.session_state.uploaded_files.get("t12_statement_pass2")
                             or st.session_state.uploaded_files.get("t12_statement"))
                if _t12_file and os.path.exists(_t12_file):
                    try:
                        from parsers.yardi_t12 import parse as parse_t12
                        t12_result = parse_t12(_t12_file)
                    except Exception as _e:
                        st.warning(f"Could not parse 12-Month Statement: {_e}")

                # Parse bank statements for BS workpaper
                _gl_result = engine_result.parsed.get('gl')
                bank_rec_data = engine_result.parsed.get("bank_rec")
                gl_cash_balance = None
                if bank_rec_data and _gl_result:
                    for _a in (_gl_result.accounts or []):
                        if _a.account_code == '111100':
                            gl_cash_balance = _a.ending_balance
                            break

                _daca_file = st.session_state.uploaded_files.get("daca_bank")
                daca_bank_data = None
                daca_gl_balance = None
                if _daca_file and os.path.exists(_daca_file):
                    try:
                        from parsers.yardi_daca_rec import (
                            is_yardi_daca_rec as _is_yardi_daca2,
                            parse as _parse_yardi_daca2,
                        )
                        from parsers.keybank_daca import parse as _parse_daca2
                        if _is_yardi_daca2(_daca_file):
                            daca_bank_data = _parse_yardi_daca2(_daca_file)
                        else:
                            daca_bank_data = _parse_daca2(_daca_file)
                    except Exception:
                        daca_bank_data = None
                if daca_bank_data and _gl_result:
                    for _a in (_gl_result.accounts or []):
                        if _a.account_code == '115100':
                            daca_gl_balance = _a.ending_balance
                            break

                _dev_rec_file = st.session_state.uploaded_files.get("bank_rec_dev")
                dev_bank_rec_data = None
                if _dev_rec_file and os.path.exists(_dev_rec_file):
                    try:
                        from parsers.bofa_statement import parse as _parse_bofa
                        dev_bank_rec_data = _parse_bofa(_dev_rec_file)
                    except Exception:
                        dev_bank_rec_data = None

                # Parse AR Aging and Capital Schedule before workpaper generation
                # (both are passed into bs_workpaper_generator.generate())
                # Pass 2-specific upload takes priority; falls back to Pass 1 sidebar file.
                _ar_aging_file_p2 = (
                    st.session_state.uploaded_files.get("ar_aging_pass2")
                    or st.session_state.uploaded_files.get("ar_aging")
                )
                _ar_aging_parsed_p2 = None
                if _ar_aging_file_p2 and os.path.exists(_ar_aging_file_p2):
                    try:
                        from parsers.yardi_ar_aging import parse as _parse_ar_aging2
                        _ar_aging_parsed_p2 = _parse_ar_aging2(_ar_aging_file_p2)
                    except Exception:
                        _ar_aging_parsed_p2 = None

                _capital_file = (
                    st.session_state.uploaded_files.get("capital_schedule_pass2")
                    or st.session_state.uploaded_files.get("capital_schedule")
                )
                _capital_schedule_data = None
                if _capital_file and os.path.exists(_capital_file):
                    try:
                        from parsers.capital_schedule import parse as _parse_capital
                        _capital_schedule_data = _parse_capital(_capital_file)
                    except Exception:
                        _capital_schedule_data = None

                # Seed fallback: use Book3.xlsx when no current-period schedule uploaded
                # (January only — prior workpaper carry-forward supersedes this from Feb onward)
                if not _capital_schedule_data:
                    _capital_seed_file = st.session_state.uploaded_files.get("capital_seed")
                    if _capital_seed_file and os.path.exists(_capital_seed_file):
                        try:
                            from parsers.capital_seed import parse as _parse_capital_seed
                            _capital_schedule_data = _parse_capital_seed(_capital_seed_file)
                        except Exception:
                            pass

                if gl_parsed and getattr(gl_parsed, 'accounts', None):
                    # tb_result is optional — generator writes "No TB data" in the TB tab
                    # when None; never block the whole workpaper just because TB is absent.
                    if not tb_result:
                        st.info(
                            "Trial Balance not uploaded — workpaper TB tab will show 'No TB data available'. "
                            "Upload a Trial Balance in the Pass 2 section to enable the full tie-out.",
                            icon="ℹ️",
                        )
                    try:
                        bs_wp_path = os.path.join(st.session_state.temp_dir, f"{_pfx_int}_Workpapers.xlsx")
                        # GL is final — no je_adjustments needed. The GL already reflects
                        # all posted JEs from Pass 1, so the workpaper ties clean.
                        _prior_wp_path = st.session_state.uploaded_files.get("prior_workpaper")
                        # Warn (rather than silently starting fresh) if this property has
                        # completed prior closes but no prior workpaper was uploaded this
                        # run — otherwise the rolling multi-month file quietly resets with
                        # no indication anything was skipped.
                        if not _prior_wp_path:
                            try:
                                from period_metrics import load_metrics as _load_metrics_wp
                                _prior_periods_wp = _load_metrics_wp(str(_DATA_DIR), _selected_code)
                            except Exception:
                                _prior_periods_wp = []
                            if _prior_periods_wp:
                                st.warning(
                                    f"⚠️ No Prior Month Workpaper uploaded — this property has "
                                    f"{len(_prior_periods_wp)} prior close(s) on record, but the "
                                    f"workpaper will start fresh this month instead of continuing "
                                    f"the rolling file. Upload last month's GA_Workpapers.xlsx as "
                                    f"the 'Prior Month Workpaper' if that wasn't intentional.",
                                    icon="⚠️",
                                )
                        # Prior period label: user override → auto-infer from close_period
                        _user_label = st.session_state.get("prior_period_label_input", "").strip()
                        _prior_period = _user_label if _user_label else None
                        if not _prior_period:
                            try:
                                if close_period:
                                    _mo_map = dict(Jan=1,Feb=2,Mar=3,Apr=4,May=5,Jun=6,
                                                   Jul=7,Aug=8,Sep=9,Oct=10,Nov=11,Dec=12)
                                    _m2 = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[- ](\d{4})', close_period)
                                    if _m2:
                                        _mo = _mo_map[_m2.group(1)]
                                        _yr = int(_m2.group(2))
                                        _prev_mo = _mo - 1 if _mo > 1 else 12
                                        _prev_yr = _yr if _mo > 1 else _yr - 1
                                        _mo_names = {v: k for k, v in _mo_map.items()}
                                        _prior_period = f"{_mo_names[_prev_mo]}-{_prev_yr}"
                            except Exception:
                                _prior_period = None
                        _berkadia_loan_data = engine_result.parsed.get('loan')
                        _berkadia_loans = []
                        if isinstance(_berkadia_loan_data, list):
                            _berkadia_loans = _berkadia_loan_data
                        elif isinstance(_berkadia_loan_data, dict):
                            _berkadia_loans = _berkadia_loan_data.get('loans', [])
                        elif hasattr(_berkadia_loan_data, 'loans'):
                            _berkadia_loans = _berkadia_loan_data.loans

                        # ── Prepaid ledger active items for workpaper ─────────────
                        # Priority: Pass 2 upload → same-session Pass 1 output
                        _prepaid_active = []
                        _prepaid_p2_path = st.session_state.uploaded_files.get("prepaid_ledger_p2")
                        if _prepaid_p2_path and os.path.exists(_prepaid_p2_path):
                            try:
                                _prepaid_active, _, _prepaid_load_err = prepaid_ledger.load(_prepaid_p2_path)
                                if _prepaid_load_err:
                                    st.caption(f"⚠️ Could not read Pass 2 prepaid ledger: {_prepaid_load_err}")
                            except Exception as _pe:
                                st.caption(f"⚠️ Could not read Pass 2 prepaid ledger: {_pe}")
                        if not _prepaid_active:
                            # Fall back to same-session Pass 1 data (already post-advance)
                            _p1_data = st.session_state.get('pass1_output_files', {})
                            _prepaid_active = _p1_data.get('ledger_active', [])

                        # ── Merge manually-added prepaid items ────────────────
                        _manual_prepaids = st.session_state.get('pass2_manual_prepaids', [])
                        if _manual_prepaids:
                            _prepaid_active = list(_prepaid_active) + _manual_prepaids
                            st.caption(
                                f"↳ Prepaid ledger: {len(_prepaid_active) - len(_manual_prepaids)} "
                                f"loaded + {len(_manual_prepaids)} manually added"
                            )

                        # A PDF uploaded in the "Workpaper raw report overrides" section
                        # for either bank rec slot takes priority over the main Bank Rec
                        # PDF upload's own parsed data — it was uploaded specifically to
                        # override this period's workpaper tabs.
                        _effective_bank_rec_data = (
                            st.session_state.get('wp_override_bank_rec_data') or bank_rec_data
                        )
                        _effective_daca_bank_data = (
                            st.session_state.get('wp_override_daca_rec_data') or daca_bank_data
                        )

                        # ── Template-based workpaper (preferred) ─────────────
                        # If a GA_Workpaper_Template.xlsx is committed for this
                        # property, use the template-based generator which preserves
                        # all existing Excel formulas (VLOOKUP, DATEDIF, SUM) and
                        # only updates data rows + date anchors.
                        # Fall back to the legacy from-scratch generator otherwise.
                        _wp_template_path = _committed_path(
                            _selected_code, 'GA_Workpaper_Template.xlsx'
                        )
                        # Roll forward from the uploaded Prior Month Workpaper when
                        # available, instead of always restarting from the pristine
                        # committed template. The committed template is a one-time
                        # bootstrap (frozen at whatever month it was first created —
                        # confirmed still frozen at Jan-2026 for revlabspm); every
                        # month after that is supposed to build on the PREVIOUS
                        # month's own output (that's the entire point of the
                        # "_Workpaper_UPLOAD_NEXT_CLOSE.xlsx" hand-off file and its
                        # warning below). Without this, generate_bs_workpaper_from_
                        # template() silently ignored the uploaded prior workpaper
                        # every single run, re-deriving "what's new" against the
                        # SAME stale baseline each month — so e.g. March's run
                        # never saw February's data at all, landed its own new rows
                        # in the same row range February's had used, and produced a
                        # workpaper that looked like "March replaced February"
                        # instead of "February then March". Confirmed with Ryan
                        # 2026-08-20. Only trusted when it actually looks like a
                        # workpaper (has the tabs this generator expects) — a
                        # wrong file in that upload slot falls back to the
                        # committed template rather than producing garbage.
                        if _prior_wp_path and os.path.exists(_prior_wp_path):
                            try:
                                from openpyxl import load_workbook as _lw_check
                                _wp_check = _lw_check(_prior_wp_path, read_only=True)
                                _wp_check_sheets = set(_wp_check.sheetnames)
                                _wp_check.close()
                                if {'Summary Page', 'Trial Balance'} <= _wp_check_sheets:
                                    _wp_template_path = _prior_wp_path
                                else:
                                    st.warning(
                                        "⚠️ The uploaded Prior Month Workpaper doesn't look like a "
                                        "GA workpaper (missing expected tabs) — using the committed "
                                        "template as the baseline instead. Double-check the file "
                                        "uploaded in that slot.",
                                        icon="⚠️",
                                    )
                            except Exception as _wp_check_err:
                                st.warning(
                                    f"⚠️ Could not read the uploaded Prior Month Workpaper "
                                    f"({_wp_check_err}) — using the committed template as the "
                                    f"baseline instead.",
                                    icon="⚠️",
                                )
                        if _wp_template_path:
                            # Template path preserves existing Excel formulas (VLOOKUP,
                            # DATEDIF, SUM) and only updates data rows + date anchors.
                            # Raw-report tabs (PNC Cash, DACA, AR Aging, Prepaid Rent, AP,
                            # BofA Dev) are regenerated each period from whatever source is
                            # currently uploaded — falling back to a computed builder, then
                            # the generic GL register, if no fresh file is available.
                            bs_workpaper_generator.generate_bs_workpaper_from_template(
                                gl_result=gl_parsed,
                                tb_result=tb_result,
                                output_path=bs_wp_path,
                                template_path=_wp_template_path,
                                period=close_period,
                                property_name=engine_result.property_name or _prop_display,
                                prepared_by=st.session_state.get("prepared_by", "GRP"),
                                property_code=_selected_code,
                                ar_aging_filepath=_ar_aging_file_p2,
                                ap_aging_filepath=st.session_state.uploaded_files.get("ap_aging"),
                                bank_rec_xlsx_filepath=st.session_state.uploaded_files.get("bank_rec_xlsx"),
                                daca_bank_rec_xlsx_filepath=st.session_state.uploaded_files.get("daca_bank_rec_xlsx"),
                                dev_bank_rec_xlsx_filepath=st.session_state.uploaded_files.get("bank_rec_dev_xlsx"),
                                prepaid_ledger_active=_prepaid_active,
                                bank_rec_data=_effective_bank_rec_data,
                                daca_bank_data=_effective_daca_bank_data,
                            )
                            st.caption(
                                "↳ Workpaper: generated from template — PNC Cash, DACA, AR Aging, "
                                "Prepaid Rent, AP, and BofA Dev tabs are refreshed each period from "
                                "whatever raw file is uploaded in the Workpaper raw report overrides "
                                "section; a tab shows a placeholder if its file wasn't uploaded this period."
                            )
                        else:
                            bs_workpaper_generator.generate(
                                gl_result=gl_parsed,
                                tb_result=tb_result,
                                output_path=bs_wp_path,
                                period=close_period,
                                property_name=engine_result.property_name or _prop_display,
                                prepaid_ledger_active=_prepaid_active,
                                bank_rec_data=_effective_bank_rec_data,
                                gl_cash_balance=gl_cash_balance,
                                daca_bank_data=_effective_daca_bank_data,
                                daca_gl_balance=daca_gl_balance,
                                prior_workpaper_path=_prior_wp_path,
                                prior_period=_prior_period,
                                berkadia_loans=_berkadia_loans,
                                dev_bank_rec_data=dev_bank_rec_data,
                                ar_aging_data=_ar_aging_parsed_p2,
                                capital_schedule_data=_capital_schedule_data,
                                tb_filepath=_tb_file,
                                ar_aging_filepath=_ar_aging_file_p2,
                                ap_aging_filepath=st.session_state.uploaded_files.get("ap_aging"),
                                bank_rec_xlsx_filepath=st.session_state.uploaded_files.get("bank_rec_xlsx"),
                                daca_bank_rec_xlsx_filepath=st.session_state.uploaded_files.get("daca_bank_rec_xlsx"),
                                dev_bank_rec_xlsx_filepath=st.session_state.uploaded_files.get("bank_rec_dev_xlsx"),
                                prepared_by=st.session_state.get("prepared_by", "GRP"),
                                property_config=_active_cfg,
                            )
                        st.session_state.pass2_output_files["bs_workpaper"] = bs_wp_path
                    except Exception as _e:
                        import traceback as _tb
                        st.warning(f"Workpaper generation skipped: {_e}")
                        st.code(_tb.format_exc(), language="text")
                elif gl_parsed:
                    # GL object exists but has no accounts — file was readable but empty
                    st.warning(
                        "Workpaper skipped — GL file parsed but contains no account data. "
                        "Verify the uploaded GL is a valid Yardi export with at least one account row.",
                        icon="⚠️",
                    )
                else:
                    st.warning("Workpaper skipped — no GL parsed. Upload a GL in Pass 2.", icon="⚠️")

                # Step 3: (Institutional workpapers removed — not needed)

                # Step 4: Management fee (informational — already in GL)
                status_text.text("Step 4/6: Verifying management fee...")
                progress_bar.progress(58)
                try:
                    _rs_file_p2 = st.session_state.uploaded_files.get("receivable_summary")
                    _rs_parsed_p2 = None
                    if _rs_file_p2 and os.path.exists(_rs_file_p2):
                        try:
                            from parsers.yardi_receivable_summary import parse as _parse_rs2
                            _rs_parsed_p2 = _parse_rs2(_rs_file_p2)
                        except Exception:
                            _rs_parsed_p2 = None

                    _rd_file_p2 = st.session_state.uploaded_files.get("receivable_detail")
                    _rd_parsed_p2 = None
                    if _rd_file_p2 and os.path.exists(_rd_file_p2):
                        try:
                            from parsers.yardi_receivable_detail import parse as _parse_rd2
                            _rd_parsed_p2 = _parse_rd2(_rd_file_p2)
                        except Exception:
                            _rd_parsed_p2 = None

                    import warnings as _warnings_fee_p2
                    with _warnings_fee_p2.catch_warnings(record=True) as _fee_warns_p2:
                        _warnings_fee_p2.simplefilter("always")
                        fee_result = calculate_mgmt_fee(
                            gl_parsed=gl_parsed,
                            budget_rows=bc_parsed or [],
                            daca_parsed=daca_bank_data,
                            receivable_summary=_rs_parsed_p2,
                            receivable_detail=_rd_parsed_p2,
                            ar_aging=_ar_aging_parsed_p2,
                        )
                    for _fw_p2 in _fee_warns_p2:
                        if issubclass(_fw_p2.category, UserWarning):
                            st.warning(str(_fw_p2.message), icon="⚠️")
                    st.session_state.pass2_output_files["fee_result"] = fee_result

                    # Generate management fee invoice PDF
                    if fee_result and fee_result.cash_received > 0:
                        try:
                            _inv_path = os.path.join(
                                st.session_state.temp_dir,
                                f"GA_MgmtFee_Invoice_{close_period.replace('-', '')}.pdf",
                            )
                            generate_mgmt_fee_invoice(
                                period=close_period,
                                cash_received=fee_result.cash_received,
                                output_path=_inv_path,
                                property_config=_active_cfg,
                            )
                            st.session_state.pass2_output_files["fee_invoice"] = _inv_path
                        except Exception as _inv_e:
                            st.warning(f"Management fee invoice skipped: {_inv_e}")
                            st.session_state.pass2_output_files["fee_invoice"] = None
                    else:
                        st.session_state.pass2_output_files["fee_invoice"] = None

                except Exception:
                    fee_result = None
                    st.session_state.pass2_output_files["fee_result"] = None
                    st.session_state.pass2_output_files["fee_invoice"] = None

                # Step 5: QC engine
                status_text.text("Step 5/6: Running QC checks...")
                progress_bar.progress(72)

                kardin_records = engine_result.parsed.get("kardin_budget") or []
                _period_month = 1
                try:
                    _m = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)', close_period)
                    _month_map = dict(Jan=1,Feb=2,Mar=3,Apr=4,May=5,Jun=6,Jul=7,Aug=8,Sep=9,Oct=10,Nov=11,Dec=12)
                    if _m:
                        _period_month = _month_map.get(_m.group(1), 1)
                except Exception:
                    pass

                try:
                    qc_report = run_qc(
                        budget_rows=bc_parsed or [],
                        tb_result=tb_result,
                        gl_parsed=gl_parsed,
                        kardin_records=kardin_records,
                        accrual_entries=[],   # JEs already posted — don't re-evaluate
                        period=close_period,
                        property_name=engine_result.property_name or _prop_entity,
                        period_month=_period_month,
                        cash_received=fee_result.cash_received if fee_result and fee_result.cash_received > 0 else None,
                        loan_data=engine_result.parsed.get('loan'),
                        property_config=_active_cfg,
                        t12_result=t12_result,
                        coa_codes=_load_coa_codes(_active_cfg),
                    )
                    st.session_state.pass2_output_files["qc_report"] = qc_report
                    qc_path = os.path.join(st.session_state.temp_dir, f"{_pfx_int}_QC_Workbook.xlsx")
                    generate_qc_workbook(
                        qc_report, qc_path,
                        tb_result=tb_result,
                        budget_rows=bc_parsed or [],
                        gl_parsed=gl_parsed,
                        loan_data=engine_result.parsed.get('loan'),
                        period_month=_period_month,
                        t12_result=t12_result,
                    )
                    st.session_state.pass2_output_files["qc_workbook"] = qc_path
                except Exception as _e:
                    st.warning(f"QC engine skipped: {_e}")
                    st.session_state.pass2_output_files["qc_report"] = None

                # Step 6: Variance comments + annotated BC
                # No je_adjustments — the final GL already reflects all posted JEs.
                status_text.text("Step 6/6: Generating variance comments...")
                progress_bar.progress(87)

                api_key = None
                try:
                    api_key = st.secrets.get("ANTHROPIC_API_KEY")
                except Exception:
                    pass

                _api_fallback_reason = None
                if bc_parsed:
                    try:
                        comments_map = generate_variance_comments_grp(
                            budget_rows=bc_parsed,
                            gl_parsed=gl_parsed,
                            kardin_records=kardin_records,
                            period=close_period,
                            property_name=engine_result.property_name or _prop_entity,
                            api_key=api_key,
                            investor_name=getattr(_active_cfg, 'investor_name', '') or '',  # C-NF-9: no hardcoded Singerman default
                            firm_name=getattr(_active_cfg, 'firm_name', '') or 'Greatland Realty Partners (GRP)',
                            ai_account_context=getattr(_active_cfg, 'ai_account_context', None) or None,
                            qc_thresholds=getattr(_active_cfg, 'qc_thresholds', None) or None,
                            # No je_adjustments — GL is the final source of truth
                        )
                        _fallback_reasons = {
                            entry.get('_api_fallback')
                            for entry in comments_map.values()
                            if entry.get('_api_fallback')
                        }
                        _api_fallback_reason = next(iter(_fallback_reasons), None)
                        if api_key and _api_fallback_reason:
                            st.warning(
                                f"⚠️ **Variance commentary fallback:** API was requested but failed. "
                                f"Comments were generated from data-driven templates, not AI. "
                                f"**Do not sign off on commentary until this is resolved.**\n\n"
                                f"Reason: {_api_fallback_reason}"
                            )
                        method = 'api' if (api_key and not _api_fallback_reason) else 'data-driven'

                        var_comments = [
                            {
                                'account_code': code,
                                'account_name': entry['account_name'],
                                'comment': entry.get('mtd_comment', ''),
                                'ytd_comment': entry.get('ytd_comment', ''),
                                'mtd_tier': entry.get('mtd_tier', 'tier_3'),
                                'ytd_tier': entry.get('ytd_tier', 'tier_3'),
                                'method': method,
                            }
                            for code, entry in comments_map.items()
                            if entry.get('mtd_tier') != 'tier_3' or entry.get('ytd_tier') != 'tier_3'
                        ]
                        st.session_state.pass2_output_files["variance_comments"] = var_comments

                        # Annotated BC (GRP internal) — prefer Pass 2 final-close BC over sidebar
                        _bc_file = (
                            st.session_state.uploaded_files.get("budget_comparison_pass2")
                            or st.session_state.uploaded_files.get("budget_comparison")
                        )
                        if _bc_file and os.path.exists(_bc_file):
                            _annotated_bc_path = os.path.join(
                                st.session_state.temp_dir, f"{_pfx_int}_Budget_Comparison_Internal.xlsx"
                            )
                            write_comments_to_budget_comparison(
                                input_path=_bc_file,
                                output_path=_annotated_bc_path,
                                comments=comments_map,
                            )
                            st.session_state.pass2_output_files["annotated_bc"] = _annotated_bc_path
                    except Exception as _e:
                        st.warning(f"Variance comments skipped: {_e}")
                        st.session_state.pass2_output_files["variance_comments"] = []
                else:
                    st.session_state.pass2_output_files["variance_comments"] = []

                # Exception report
                if api_key and _api_fallback_reason:
                    engine_result.exceptions.append(Exception_(
                        severity='warning',
                        category='commentary',
                        source='variance_comments',
                        description=(
                            'Variance commentary API fallback: AI commentary was requested but '
                            f'the API call failed. Reason: {_api_fallback_reason}'
                        ),
                        details={'api_fallback_reason': _api_fallback_reason},
                    ))

                exception_path = os.path.join(st.session_state.temp_dir, f"{_pfx_int}_Exceptions_Report.xlsx")
                try:
                    generate_exception_report(engine_result, exception_path)
                    st.session_state.pass2_output_files["exception_report"] = exception_path
                except Exception as _e:
                    st.warning(f"Exception report skipped: {_e}")

                # Store auxiliary pass2 data for dashboard
                st.session_state.pass2_output_files["tb_result"]         = tb_result
                st.session_state.pass2_output_files["bank_rec_data"]     = bank_rec_data
                st.session_state.pass2_output_files["daca_bank_data"]    = daca_bank_data
                st.session_state.pass2_output_files["gl_cash_balance"]   = gl_cash_balance
                st.session_state.pass2_output_files["daca_gl_balance"]   = daca_gl_balance
                st.session_state.pass2_output_files["dev_bank_rec_data"] = dev_bank_rec_data

                # Audit Trail — comprehensive pass-1+pass-2 record for auditors
                _at_qc = None   # B-F5: init before try so except/finally can reference it safely
                try:
                    _at_path = os.path.join(
                        st.session_state.temp_dir,
                        f"{_pfx_int}_Audit_Trail_{close_period.replace('-', '')}.xlsx",
                    )
                    # Pull Pass 1 JE lines — session state first, uploaded cache as fallback
                    _p1_out = st.session_state.get('pass1_output_files', {})
                    _at_je_lines = _p1_out.get('all_je_lines') or []
                    if not _at_je_lines:
                        _jec_path = st.session_state.uploaded_files.get('je_lines_cache')
                        if _jec_path and os.path.exists(_jec_path):
                            try:
                                import json as _json
                                with open(_jec_path) as _jcf:
                                    _at_je_lines = _json.load(_jcf)
                            except Exception:
                                pass
                    _at_fee      = fee_result   # Pass 2 fee verification result
                    _at_qc       = st.session_state.pass2_output_files.get('qc_report')

                    # Prior-month accrual check against the PASS 2 (final) GL
                    _gl_for_at = engine_result.parsed.get('gl') if engine_result.parsed else None
                    _at_prior  = check_prior_accrual_vs_actual(_gl_for_at) if _gl_for_at else []

                    generate_audit_trail(
                        output_path         = _at_path,
                        period              = close_period,
                        property_name       = engine_result.property_name or _prop_display,
                        all_je_lines        = _at_je_lines,
                        fee_result          = _at_fee,
                        qc_report           = _at_qc,
                        prior_accrual_check = _at_prior,
                        files_uploaded      = st.session_state.uploaded_files,
                        property_config     = _active_cfg,
                        property_code       = (getattr(_active_cfg, 'yardi_etl_code', '') or
                                               getattr(_active_cfg, 'property_code', '') or ''),
                        bank_recon_detail   = engine_result.bank_recon_detail,
                        close_tracker       = st.session_state.get('close_tracker', {}),
                        signoff_state       = st.session_state.get('signoff_state', {}),
                        signoff_items       = _SIGNOFF_ITEMS,
                    )
                    st.session_state.pass2_output_files["audit_trail"] = _at_path
                except Exception as _ate:
                    st.warning(f"Audit trail skipped: {_ate}")
                    st.session_state.pass2_output_files["audit_trail"] = None

                # ── Step 8: JE Verification ───────────────────────────────────
                # Compare Pass 1 JE lines against J-type transactions in the
                # final GL to confirm every entry actually posted to Yardi.
                status_text.text("Step 8/8: Verifying JE posting...")
                progress_bar.progress(98)
                try:
                    from je_verifier import verify_je_posting, write_je_verification_tab
                    _p1_je_lines_v = _at_je_lines   # already resolved above (session + cache fallback)
                    _gl_for_v = engine_result.parsed.get('gl') if engine_result.parsed else None

                    if _p1_je_lines_v and _gl_for_v:
                        _je_ver_result = verify_je_posting(_p1_je_lines_v, _gl_for_v)
                        st.session_state.pass2_output_files['je_verification'] = _je_ver_result

                        # Append the verification tab to the existing QC workbook
                        _qc_wb_path = st.session_state.pass2_output_files.get('qc_workbook')
                        if _qc_wb_path and os.path.exists(_qc_wb_path):
                            try:
                                from openpyxl import load_workbook as _load_qc_wb
                                _qc_wb = _load_qc_wb(_qc_wb_path)
                                write_je_verification_tab(_qc_wb, _je_ver_result, period=close_period)
                                _qc_wb.save(_qc_wb_path)
                            except Exception as _qc_tab_err:
                                st.warning(f"JE Verification tab could not be added to QC workbook: {_qc_tab_err}")
                    else:
                        st.session_state.pass2_output_files['je_verification'] = None
                        if not _p1_je_lines_v:
                            st.info(
                                "JE Verification skipped — Pass 1 JE data not found. "
                                "Upload the **JE Cache (.json)** from the Pass 1 download package "
                                "in the Pass 2 upload section above to enable verification.",
                                icon="ℹ️",
                            )
                except Exception as _jve:
                    st.warning(f"JE Verification skipped: {_jve}")
                    st.session_state.pass2_output_files['je_verification'] = None

                # ── Auto-detect Close Tracker Steps 5 & 6 ────────────────────
                _ct = st.session_state.close_tracker
                _p2_ts = datetime.now().strftime("%m/%d/%Y %H:%M")
                _p2_by = st.session_state.get('prepared_by', 'GRP')
                _ck_changed = False
                if 5 not in _ct:
                    _ct[5] = {"completed_by": _p2_by, "timestamp": _p2_ts, "auto": True}
                    _ck_changed = True
                if 6 not in _ct:
                    _ct[6] = {"completed_by": _p2_by, "timestamp": _p2_ts, "auto": True}
                    _ck_changed = True
                if _ck_changed:
                    _save_checklist_now()
                    # Notify at step 6 (highest: QC review ready for team)
                    st.session_state.last_completed_step = 6

                # ── Run Log ───────────────────────────────────────────────────
                try:
                    from run_log import append_run_log as _append_run_log
                    _rl_path  = os.path.join(st.session_state.temp_dir, "GA_Run_Log.csv")
                    # Use Pass 1 run log as prior if it exists, else uploaded prior
                    _rl_prior = (
                        st.session_state.pass1_output_files.get("run_log")
                        or st.session_state.uploaded_files.get("run_log")
                    )
                    _rl_qc    = _at_qc   # always defined — initialized at line 4927 (B-F5)
                    _rl_pass  = sum(1 for c in (_rl_qc.checks if _rl_qc else []) if c.status == 'PASS')
                    _rl_fail  = sum(1 for c in (_rl_qc.checks if _rl_qc else []) if c.status in ('FLAG', 'FAIL'))
                    _rl_files = [k for k, v in st.session_state.pass2_output_files.items() if v]
                    _append_run_log(
                        output_path            = _rl_path,
                        prior_log_path         = _rl_prior,
                        timestamp              = datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        prepared_by            = st.session_state.get('prepared_by', 'GRP'),
                        property_name          = engine_result.property_name or _prop_display,
                        period                 = close_period,
                        files_generated        = _rl_files,
                        qc_checks_passed       = _rl_pass,
                        qc_checks_failed       = _rl_fail,
                        close_tracker_complete = (len(st.session_state.close_tracker) == 9),
                    )
                    st.session_state.pass2_output_files["run_log"] = _rl_path
                except Exception as _rle:
                    pass   # run log is non-critical; never block report generation

                progress_bar.progress(100)
                status_text.text("✓ Reports complete!")
                st.session_state.pass2_complete = True
                st.success("Pass 2 complete! Download reports below.", icon="✅")

                # ── Persist period metrics for cross-period trending ──────
                try:
                    from period_metrics import save_metrics as _save_metrics
                    _pm_gl   = engine_result.parsed.get('gl') if engine_result.parsed else None
                    _pm_qc   = st.session_state.pass2_output_files.get('qc_report')
                    _pm_fee  = st.session_state.pass2_output_files.get('fee_result')
                    _save_metrics(
                        data_dir      = str(_DATA_DIR),
                        property_code = _active_cfg.property_code or '',
                        period        = close_period,
                        property_name = engine_result.property_name or _prop_display,
                        gl_data       = _pm_gl,
                        qc_report     = _pm_qc,
                        fee_result    = _pm_fee,
                    )
                except Exception:
                    pass  # metrics are non-critical; never block report delivery

            except Exception as e:
                tb = traceback.format_exc()
                st.error(f"Pass 2 error: {str(e)}", icon="❌")
                st.code(tb, language="python")
                st.session_state.pass2_complete = False

    # ── Pass 2 Results Dashboard ──────────────────────────────────────────────
    if st.session_state.pass2_complete and st.session_state.pass2_engine_result:
        result = st.session_state.pass2_engine_result
        p2     = st.session_state.pass2_output_files

        st.divider()
        st.markdown(f"## Pass 2 Results — {result.period}  |  {result.property_name}")

        # Period-state indicator (shown first — informational, not an error)
        _ps = result.period_state
        if _ps and _ps.get('state') != 'unknown':
            _state_labels = {
                'pre_close':  ('🟡 Pre-Close', '#f39c12',
                               'GL is for the current period — books still open.'),
                'at_close':   ('🟢 At Close', '#2ecc71', 'Month-end close window.'),
                'post_close': ('🔵 Post-Close', '#3498db',
                               'Running on a closed period. Reports are final.'),
            }
            _label, _color, _desc = _state_labels.get(_ps['state'], ('⚪ Unknown', '#95a5a6', ''))
            _days = _ps.get('days_since_close', 0)
            _day_str = (f'{abs(_days)} day{"s" if abs(_days) != 1 else ""} until close'
                        if _days < 0 else
                        f'{_days} day{"s" if _days != 1 else ""} since close'
                        if _days > 0 else 'closes today')
            _promo = ' _(promoted by 213100 GL signal)_' if _ps.get('promoted') else ''
            with st.expander(f"{_label} — {_day_str}{_promo}", expanded=False):
                st.markdown(_desc)
                if _ps.get('gl_signal_detected'):
                    st.info(
                        f"213100 GL signal detected: net credit ${_ps['gl_signal_amount']:,.2f} "
                        f"— prior-period auto-reversals have posted."
                    )

        # Status banner (CLEAN / WARNINGS / ERRORS from engine validation)
        status = result.status
        status_color = {"CLEAN": "#2ecc71", "WARNINGS": "#f39c12"}.get(status, "#e74c3c")
        status_text_str = {"CLEAN": "✅ CLEAN", "WARNINGS": "⚠️ WARNINGS"}.get(status, "❌ ERRORS")
        st.markdown(f"""
        <div style="background-color: {status_color}20; border-left: 5px solid {status_color};
             padding: 15px; border-radius: 5px; margin: 15px 0;">
            <h3 style="color: {status_color}; margin: 0;">{status_text_str}</h3>
        </div>
        """, unsafe_allow_html=True)

        # ── JE Verification Panel ──────────────────────────────────────────
        _je_ver = p2.get('je_verification')
        if _je_ver is not None:
            st.markdown("### JE Posting Verification")
            _jv_color = '#2ecc71' if _je_ver.all_verified else ('#f39c12' if _je_ver.missing_count == 0 else '#e74c3c')
            _jv_icon  = '✅' if _je_ver.all_verified else ('⚠️' if _je_ver.missing_count == 0 else '❌')
            st.markdown(f"""
            <div style="background:{_jv_color}20;border-left:5px solid {_jv_color};
                 padding:10px 15px;border-radius:5px;margin:10px 0 5px;">
                <strong style="color:{_jv_color};">{_jv_icon} {_je_ver.summary}</strong>
            </div>""", unsafe_allow_html=True)

            if not _je_ver.all_verified:
                # Show problem JEs immediately — reviewer needs to act
                _problem_jes = [j for j in _je_ver.je_results
                                if j.status in ('MISSING', 'PARTIAL', 'AMOUNT_MISMATCH')]
                if _problem_jes:
                    _jv_rows = []
                    for _pj in _problem_jes:
                        _missing_accts = [f"{l.account_code}" for l in _pj.lines if l.match_status == 'not_found']
                        _mismatch_notes = [l.note for l in _pj.lines if l.match_status == 'amount_mismatch']
                        _note = ''
                        if _missing_accts:
                            _note = f"Accts not found in GL: {', '.join(_missing_accts)}"
                        if _mismatch_notes:
                            _note += (' | ' if _note else '') + '; '.join(_mismatch_notes)
                        _jv_rows.append({
                            'JE #':    _pj.je_number,
                            'Status':  _pj.status,
                            'Source':  _pj.source.replace('_', ' ').title(),
                            'Lines':   _pj.line_count,
                            'Found':   _pj.verified_count,
                            'Missing': _pj.missing_count,
                            'Note':    _note,
                        })
                    import pandas as _pd_jv
                    st.dataframe(
                        _pd_jv.DataFrame(_jv_rows),
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            'Status': st.column_config.TextColumn(width='small'),
                            'Lines':  st.column_config.NumberColumn(width='small'),
                            'Found':  st.column_config.NumberColumn(width='small'),
                            'Missing':st.column_config.NumberColumn(width='small'),
                        }
                    )
                    if _je_ver.missing_count > 0:
                        st.warning(
                            f"⚠️ **{_je_ver.missing_count} JE(s) not found in the final GL.** "
                            f"This means they may not have imported into Yardi. "
                            f"Check the Yardi import log, re-upload the CSV, and regenerate Pass 2 to re-verify. "
                            f"See the **JE Verification** tab in the QC Workbook for full line-level detail.",
                            icon="⚠️",
                        )

            with st.expander(f"All JEs ({_je_ver.total_je_count} total)", expanded=False):
                _all_jv_rows = []
                for _aj in _je_ver.je_results:
                    _all_jv_rows.append({
                        'JE #':    _aj.je_number,
                        'Status':  _aj.status,
                        'Source':  _aj.source.replace('_', ' ').title(),
                        'Lines':   _aj.line_count,
                        'Found':   _aj.verified_count,
                    })
                import pandas as _pd_jv2
                st.dataframe(_pd_jv2.DataFrame(_all_jv_rows), use_container_width=True, hide_index=True)

        elif p2.get('je_verification') is None and st.session_state.get('pass1_complete'):
            st.info(
                "JE Verification is available when Pass 1 and Pass 2 are run in the same session. "
                "Re-run Pass 1 then immediately run Pass 2 to enable it.",
                icon="ℹ️",
            )

        # ── QC Summary Panel ───────────────────────────────────────────────
        qc_report = p2.get("qc_report")
        if qc_report:
            st.markdown("### QC Checks")
            qc_overall = qc_report.overall_status
            qc_color = {'PASS': '#2ecc71', 'FLAG': '#f39c12', 'FAIL': '#e74c3c'}.get(qc_overall, '#95a5a6')
            qc_icon  = {'PASS': '✅', 'FLAG': '⚠️', 'FAIL': '❌'}.get(qc_overall, 'ℹ️')
            st.markdown(f"""
            <div style="background:{qc_color}20;border-left:5px solid {qc_color};
                 padding:10px 15px;border-radius:5px;margin:10px 0 5px;">
                <strong style="color:{qc_color};">{qc_icon} QC Overall: {qc_overall}</strong>
                &nbsp;&nbsp;{sum(1 for c in qc_report.checks if c.status=='PASS')} PASS &nbsp;
                {sum(1 for c in qc_report.checks if c.status=='FLAG')} FLAG &nbsp;
                {sum(1 for c in qc_report.checks if c.status=='FAIL')} FAIL
            </div>
            """, unsafe_allow_html=True)
            qc_rows = []
            for chk in qc_report.checks:
                chk_icon = {'PASS': '✅', 'FLAG': '⚠️', 'FAIL': '❌'}.get(chk.status, '')
                qc_rows.append({
                    "Check":    chk.check_id.replace('CHECK_', '') + ' — ' + chk.check_name,
                    "Status":   f"{chk_icon} {chk.status}",
                    "Findings": chk.flag_count,
                    "Summary":  chk.summary,
                })
            st.dataframe(qc_rows, use_container_width=True, hide_index=True,
                         column_config={
                             "Check":    st.column_config.TextColumn(width="medium"),
                             "Status":   st.column_config.TextColumn(width="small"),
                             "Findings": st.column_config.NumberColumn(width="small"),
                             "Summary":  st.column_config.TextColumn(width="large"),
                         })
            st.divider()

        # ── Management Fee Panel ───────────────────────────────────────────
        fee_result = p2.get("fee_result")
        if fee_result and fee_result.cash_received > 0:
            st.markdown("### Management Fee Verification")
            st.caption("Computed from the final GL — used to verify the posted JE is correct.")
            col_f1, col_f2, col_f3, col_f4 = st.columns(4)
            with col_f1:
                st.metric("Cash Received", f"${fee_result.cash_received:,.0f}",
                          help=f"Source: {fee_result.cash_source}")
            with col_f2:
                st.metric(f"JLL ({fee_result.jll_rate:.2%})", f"${fee_result.jll_fee:,.0f}")
            with col_f3:
                st.metric(f"GRP ({fee_result.grp_rate:.2%})", f"${fee_result.grp_fee:,.0f}")
            with col_f4:
                st.metric(f"Total ({fee_result.total_rate:.2%})", f"${fee_result.total_fee:,.0f}")
            bc_for_fee = result.parsed.get('budget_comparison') or []
            accrued = accrued_fee_from_bc(bc_for_fee)
            if accrued > 0:
                diff = accrued - fee_result.total_fee
                diff_str = f"${abs(diff):,.0f} {'over' if diff > 0 else 'under'} calculated"
                st.caption(f"BC accrued (637130): ${accrued:,.2f} — {diff_str}")
            st.divider()

        # ── Budget Variances with Comments ─────────────────────────────────
        if result.budget_variances:
            st.markdown("### Budget Variances (Flagged)")
            var_comments = p2.get("variance_comments", [])
            comments_map_disp = {vc['account_code']: vc.get('comment', '') for vc in var_comments}
            comment_method = var_comments[0].get('method', 'none') if var_comments else 'none'
            if comment_method == 'api':
                st.caption("Variance comments generated via Claude API")
            elif comment_method == 'data-driven':
                st.caption(
                    "Variance comments generated from GL transaction detail "
                    "(configure ANTHROPIC_API_KEY in Streamlit secrets for narrative comments)"
                )
            variance_data = []
            for var in result.budget_variances:
                code = var.get("account_code", "")
                variance_data.append({
                    "Account Code": code,
                    "Account Name": var.get("account_name", ""),
                    "Actual":       var.get("ptd_actual", 0),
                    "Budget":       var.get("ptd_budget", 0),
                    "Variance":     var.get("variance", 0),
                    "Variance %":   f"{var.get('variance_pct', 0):.1f}%",
                    "Comment":      comments_map_disp.get(code, ''),
                })
            st.dataframe(variance_data, use_container_width=True, hide_index=True,
                         column_config={
                             "Account Code": st.column_config.TextColumn(width="small"),
                             "Account Name": st.column_config.TextColumn(width="medium"),
                             "Actual":       st.column_config.NumberColumn(format="$%,.2f"),
                             "Budget":       st.column_config.NumberColumn(format="$%,.2f"),
                             "Variance":     st.column_config.NumberColumn(format="$%,.2f"),
                             "Variance %":   st.column_config.TextColumn(width="small"),
                             "Comment":      st.column_config.TextColumn(width="large"),
                         })
            st.divider()

        # ── Bank Rec Summary Panel ─────────────────────────────────────────
        _bank_rec  = p2.get("bank_rec_data")
        _daca_data = p2.get("daca_bank_data")
        _dev_rec   = p2.get("dev_bank_rec_data")
        _gl_111    = float(p2.get("gl_cash_balance") or 0)
        _daca_gl   = float(p2.get("daca_gl_balance") or 0)
        if _bank_rec or _daca_data or _dev_rec:
            st.markdown("### Bank Reconciliation Summary")
            _rec_cols = st.columns(3)
            # Derive bank account labels from property config
            _ba_cfg = getattr(_active_cfg, 'bank_accounts', None) or {}
            _gl_acc = getattr(_active_cfg, 'gl_accounts', None) or {}
            def _ba_gl(v):
                """Get gl_account from either a BankAccountConfig dataclass or a plain dict."""
                if isinstance(v, dict):
                    return str(v.get('gl_account', ''))
                return str(getattr(v, 'gl_account', '') or '')
            def _ba_attr(v, key, default=''):
                """Get any attribute from either a BankAccountConfig dataclass or a plain dict."""
                if isinstance(v, dict):
                    return v.get(key, default) or default
                return getattr(v, key, default) or default
            _op_slug   = next((k for k, v in _ba_cfg.items() if _ba_gl(v).strip() == str(_gl_acc.get('cash_operating','111100')).strip()), None)
            _daca_slug = next((k for k, v in _ba_cfg.items() if _ba_gl(v).strip() == str(_gl_acc.get('daca','115100')).strip()), None)
            _dev_slugs = [k for k in _ba_cfg if k not in (_op_slug, _daca_slug)] if _ba_cfg else []
            def _ba_label(slug):
                if not slug or slug not in _ba_cfg:
                    return slug or 'Operating'
                ba = _ba_cfg[slug]
                lbl  = _ba_attr(ba, 'label', slug)
                last4 = _ba_attr(ba, 'last4', '')
                gl   = _ba_gl(ba)
                return f"{lbl}{' (' + last4 + ')' if last4 else ''}{' — GL ' + gl if gl else ''}"
            _op_lbl   = _ba_label(_op_slug)   if _op_slug   else f"Operating — GL {_gl_acc.get('cash_operating','111100')}"
            _daca_lbl = _ba_label(_daca_slug) if _daca_slug else f"DACA — GL {_gl_acc.get('daca','115100')}"
            _op_gl_code   = str(_gl_acc.get('cash_operating', '111100'))
            _daca_gl_code = str(_gl_acc.get('daca', '115100'))

            with _rec_cols[0]:
                if _bank_rec:
                    _bank_bal  = float(_bank_rec.get('bank_statement_balance') or 0)
                    _out_total = float(_bank_rec.get('total_outstanding_checks') or 0)
                    _rec_bal   = float(_bank_rec.get('reconciled_bank_balance') or 0)
                    _diff_111  = _rec_bal - _gl_111
                    _icon_111  = "✅" if abs(_diff_111) < 0.02 else "❌"
                    st.markdown(f"""
**{_op_lbl}** {_icon_111}
| | |
|---|---:|
| Bank Statement Balance | ${_bank_bal:,.2f} |
| Less: Outstanding Checks ({len(_bank_rec.get('outstanding_checks') or [])}) | (${_out_total:,.2f}) |
| Reconciled Bank Balance | **${_rec_bal:,.2f}** |
| GL Balance ({_op_gl_code}) | ${_gl_111:,.2f} |
| **Difference** | **${_diff_111:+,.2f}** |
""")
                else:
                    st.caption("Upload Yardi Bank Rec PDF to see Operating account rec summary")
            with _rec_cols[1]:
                if _daca_data:
                    _daca_end  = float(_daca_data.get('ending_balance') or 0)
                    _diff_daca = _daca_end - _daca_gl
                    _icon_daca = "✅" if abs(_diff_daca) < 0.02 else "❌"
                    st.markdown(f"""
**{_daca_lbl}** {_icon_daca}
| | |
|---|---:|
| Bank Statement Ending Balance | ${_daca_end:,.2f} |
| GL Balance ({_daca_gl_code}) | ${_daca_gl:,.2f} |
| **Difference** | **${_diff_daca:+,.2f}** |
""")
                else:
                    st.caption("Upload DACA Bank Statement to see DACA account rec summary")
            with _rec_cols[2]:
                if _dev_rec:
                    _dev_bank_bal  = float(_dev_rec.get('bank_statement_balance') or 0)
                    _dev_out_total = float(_dev_rec.get('total_outstanding_checks') or 0)
                    _dev_rec_bal   = float(_dev_rec.get('reconciled_bank_balance') or 0)
                    _dev_gl_bal    = float(_dev_rec.get('gl_balance') or 0)
                    _dev_diff      = _dev_rec_bal - _dev_gl_bal
                    _dev_icon      = "✅" if abs(_dev_diff) < 0.02 else "❌"
                    _dev_lbl = _ba_label(_dev_slugs[0]) if _dev_slugs else "Secondary Account"
                    st.markdown(f"""
**{_dev_lbl}** {_dev_icon}
| | |
|---|---:|
| Bank Statement Balance | ${_dev_bank_bal:,.2f} |
| Less: Outstanding Checks ({len(_dev_rec.get('outstanding_checks') or [])}) | (${_dev_out_total:,.2f}) |
| Reconciled Bank Balance | **${_dev_rec_bal:,.2f}** |
| GL Balance (per Yardi rec) | ${_dev_gl_bal:,.2f} |
| **Difference** | **${_dev_diff:+,.2f}** |
""")
                else:
                    st.caption("Upload secondary bank statement to see Development account rec summary")
            st.divider()

        # ── Engine Bank Match Detail (collapsible) ─────────────────────────
        if result.gl_bank_matches:
            with st.expander("Engine Bank Match Detail"):
                recon_data = [{
                    "Description": match.description,
                    "GL Amount":   match.amount_a,
                    "Bank Amount": match.amount_b,
                    "Matched":     "✅" if match.matched else "⚠️",
                    "Variance":    abs(match.variance),
                } for match in result.gl_bank_matches]
                st.dataframe(recon_data, use_container_width=True, hide_index=True,
                             column_config={
                                 "Description": st.column_config.TextColumn(),
                                 "GL Amount":   st.column_config.NumberColumn(format="$%,.2f"),
                                 "Bank Amount": st.column_config.NumberColumn(format="$%,.2f"),
                                 "Matched":     st.column_config.TextColumn(),
                                 "Variance":    st.column_config.NumberColumn(format="$%,.2f"),
                             })
            st.divider()

        # ── Debt Service ───────────────────────────────────────────────────
        if result.debt_service_check and result.debt_service_check.get("loans"):
            st.markdown("### Debt Service Summary")
            debt_data = [{
                "Loan":               loan.get("name", "Unknown"),
                "Principal Balance":  loan.get("principal_balance", 0),
                "Interest Paid YTD":  loan.get("interest_paid_ytd", 0),
            } for loan in result.debt_service_check["loans"]]
            st.dataframe(debt_data, use_container_width=True, hide_index=True,
                         column_config={
                             "Loan":              st.column_config.TextColumn(),
                             "Principal Balance": st.column_config.NumberColumn(format="$%,.2f"),
                             "Interest Paid YTD": st.column_config.NumberColumn(format="$%,.2f"),
                         })
            st.divider()

        # ── Summary Metrics ────────────────────────────────────────────────
        st.markdown("### Summary Metrics")
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.metric("Files Processed", result.summary.get("files_processed", 0))
        with col2:
            st.metric("GL Accounts", result.summary.get("gl_accounts", 0))
        with col3:
            st.metric("GL Transactions", result.summary.get("gl_transactions", 0))
        with col4:
            st.metric("GL Balanced", "Yes" if result.summary.get("gl_balanced") else "No")
        with col5:
            st.metric("Exceptions",
                      f"{result.summary.get('exceptions_error', 0)}E / "
                      f"{result.summary.get('exceptions_warning', 0)}W")
        st.divider()

        # ── Parser Status ──────────────────────────────────────────────────
        st.markdown("### Parser Status")
        parser_data = [{"Parser": k.replace("_", " ").title(), "Status": "✅ Success"}
                       for k in result.parsed.keys()]
        if parser_data:
            st.dataframe(parser_data, use_container_width=True, hide_index=True,
                         column_config={
                             "Parser": st.column_config.TextColumn(),
                             "Status": st.column_config.TextColumn(),
                         })
        st.divider()

        # ── Exceptions ─────────────────────────────────────────────────────
        if result.exceptions:
            st.markdown("### Exceptions & Findings")
            for exc in result.exceptions:
                severity_badge = {
                    "error":   "🔴 ERROR",
                    "warning": "🟡 WARNING",
                    "info":    "🔵 INFO",
                }.get(exc.severity, "ℹ️ INFO")
                with st.expander(f"{severity_badge} — {exc.description}"):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**Category:** {exc.category}")
                        st.write(f"**Source:** {exc.source}")
                    with col2:
                        if exc.details:
                            st.write("**Details:**")
                            for key, val in exc.details.items():
                                st.write(f"- {key}: {val}")
        else:
            st.success("No exceptions found! Pipeline validation passed.", icon="✅")

        st.divider()

        # ── Download Section ───────────────────────────────────────────────
        st.markdown("### Download Reports")

        import zipfile, io
        period_label = (result.period or 'Period').replace('-', '_')

        p2_zip_files = {
            f"{_pfx_del}_{period_label}_Workpapers.xlsx":      p2.get("bs_workpaper"),
            f"{_pfx_del}_{period_label}_QC_Workbook.xlsx":     p2.get("qc_workbook"),
            f"{_pfx_del}_{period_label}_Exceptions.xlsx":      p2.get("exception_report"),
            f"{_pfx_del}_{period_label}_BC_Internal.xlsx":     p2.get("annotated_bc"),
            f"{_pfx_del}_{period_label}_Audit_Trail.xlsx":     p2.get("audit_trail"),
            f"{_inv_pfx}_Invoice_{period_label}.pdf":          p2.get("fee_invoice"),
            f"{_pfx_del}_{period_label}_Run_Log.csv":          p2.get("run_log"),
            f"{_pfx_del}_{period_label}_Signoff_Record.xlsx":  p2.get("signoff_record"),
            f"{_pfx_del}_{period_label}_Close_Tracker.xlsx":   p2.get("close_tracker"),
        }
        p2_zip_files = {k: v for k, v in p2_zip_files.items() if v and os.path.exists(v)}

        if p2_zip_files:
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                for fname, fpath in p2_zip_files.items():
                    zf.write(fpath, fname)
            zip_buf.seek(0)
            st.download_button(
                label=f"📦 Download Full Report Package ({len(p2_zip_files)} files)",
                data=zip_buf,
                file_name=f"{_pfx_del}_{period_label}_Reports_{datetime.now().strftime('%Y%m%d')}.zip",
                mime="application/zip",
                use_container_width=True,
                help="Workpapers, QC Workbook, Exception Report, Annotated BC",
            )

        # ── Carry-Forward Package — the 2 files needed to start next month ────
        _p1_out_cf = st.session_state.get('pass1_output_files', {}) or {}
        _cf_ledger_path = _p1_out_cf.get('prepaid_ledger_updated')
        _cf_workpaper_path = p2.get('bs_workpaper')
        _cf_files = {
            f"{_pfx_del}_{period_label}_Prepaid_Ledger_UPLOAD_NEXT_CLOSE.xlsx": _cf_ledger_path,
            f"{_pfx_del}_{period_label}_Workpaper_UPLOAD_NEXT_CLOSE.xlsx":      _cf_workpaper_path,
        }
        _cf_files = {k: v for k, v in _cf_files.items() if v and os.path.exists(v)}
        if _cf_files:
            _cf_zip_buf = io.BytesIO()
            with zipfile.ZipFile(_cf_zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                for fname, fpath in _cf_files.items():
                    zf.write(fpath, fname)
            _cf_zip_buf.seek(0)
            st.download_button(
                label=f"➡️ Download Carry-Forward Package for Next Month ({len(_cf_files)} files)",
                data=_cf_zip_buf,
                file_name=f"{_pfx_del}_{period_label}_CarryForward_{datetime.now().strftime('%Y%m%d')}.zip",
                mime="application/zip",
                use_container_width=True,
                help="Save this. Next month: upload the Prepaid Ledger as the prior-month ledger "
                     "in Pass 1, and the Workpaper as the prior workpaper in Pass 2.",
            )
        elif not _cf_ledger_path:
            st.caption(
                "ℹ️ Carry-Forward Package unavailable — Pass 1 wasn't run in this session, "
                "so the updated Prepaid Ledger isn't available to bundle. "
                "Download it separately from the Pass 1 tab if you still have that session open."
            )

        st.divider()
        st.markdown("##### Individual Downloads")
        _dc1, _dc2, _dc3 = st.columns(3)
        _dl_cols = [_dc1, _dc2, _dc3]

        _ts_p2 = datetime.now().strftime('%Y%m%d')
        _dl_items = [
            ("bs_workpaper",    "📋 Workpapers",
             f"{_pfx_int}_Workpapers_{_ts_p2}.xlsx",      None),
            ("qc_workbook",     "✅ QC Workbook",
             f"{_pfx_int}_QC_Workbook_{_ts_p2}.xlsx",     None),
            ("exception_report","⚠️ Exception Report",
             f"{_pfx_int}_Exceptions_{_ts_p2}.xlsx",      None),
            ("annotated_bc",    "💬 Budget Comparison",
             f"{_pfx_int}_BC_Internal_{_ts_p2}.xlsx",     None),
            ("audit_trail",     "🔍 Audit Trail",
             f"{_pfx_int}_Audit_Trail_{_ts_p2}.xlsx",     None),
            ("fee_invoice",     "🧾 Management Fee Invoice",
             f"{_inv_pfx}_Invoice_{(result.period or '').replace('-','')}.pdf",
             "application/pdf"),
        ]

        for i, (key, label, fname, mime) in enumerate(_dl_items):
            fpath = p2.get(key)
            if fpath and os.path.exists(fpath):
                with _dl_cols[i % 3]:
                    with open(fpath, "rb") as f:
                        kw = dict(label=label, data=f.read(), file_name=fname,
                                  use_container_width=True)
                        if mime:
                            kw["mime"] = mime
                        st.download_button(**kw)

        # ── Reversing JE ──────────────────────────────────────────────────────
        # Offered in Pass 2 after the close is confirmed — the period is locked in
        # from the final GL, and the JEs are already posted to Yardi.  Use this CSV
        # to manually post reversals if Yardi auto-reversal fails, or to pre-review
        # what Yardi will reverse on the 1st of next month.
        _p2_rev_src = st.session_state.get('pass1_output_files', {}).get('accrual_je_csv')
        with st.expander("🔄 Reversing JE CSV — Next Month Setup", expanded=False):
            if _p2_rev_src and os.path.exists(_p2_rev_src):
                st.caption(
                    "Flips every DR/CR in the Pass 1 accruals CSV and dates them to "
                    "the next period. Yardi auto-reverses BM=-1 entries on the 1st, "
                    "but download this as a backup or to post manually if needed."
                )
                # Auto-fill next period from Pass 2 close period
                _p2_rev_next = ''
                try:
                    _p2_mo_map = dict(Jan=1,Feb=2,Mar=3,Apr=4,May=5,Jun=6,
                                      Jul=7,Aug=8,Sep=9,Oct=10,Nov=11,Dec=12)
                    _p2_mo_names = {v: k for k, v in _p2_mo_map.items()}
                    _p2_m = re.search(
                        r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[- ](\d{4})',
                        result.period or '',
                    )
                    if _p2_m:
                        _p2_mo  = _p2_mo_map[_p2_m.group(1)]
                        _p2_yr  = int(_p2_m.group(2))
                        _p2_nmo = _p2_mo % 12 + 1
                        _p2_nyr = _p2_yr + (1 if _p2_mo == 12 else 0)
                        _p2_rev_next = f"{_p2_mo_names[_p2_nmo]}-{_p2_nyr}"
                except Exception:
                    pass
                _p2_rev_period = st.text_input(
                    "Reversing period",
                    value=_p2_rev_next,
                    help="Period the reversals will post to (auto-filled to next month).",
                    key="p2_rev_period_input",
                )
                if _p2_rev_period:
                    try:
                        from accrual_entry_generator import build_reversing_je_csv as _p2_build_rev
                        _p2_rev_out = os.path.join(
                            st.session_state.temp_dir,
                            f"{_pfx_int}_Reversing_JE_{_p2_rev_period.replace('-', '')}.csv",
                        )
                        _p2_build_rev(
                            source_etl_path=_p2_rev_src,
                            next_period=_p2_rev_period,
                            output_path=_p2_rev_out,
                            property_code=getattr(_active_cfg, 'property_code', '') or '',
                        )
                        with open(_p2_rev_out, 'rb') as _p2_rf:
                            st.download_button(
                                label=f"⬇️ Download Reversing JE — {_p2_rev_period}",
                                data=_p2_rf.read(),
                                file_name=os.path.basename(_p2_rev_out),
                                mime="text/csv",
                                use_container_width=True,
                                key="p2_dl_rev_je",
                            )
                    except Exception as _p2_rev_err:
                        st.error(f"Reversing JE generation failed: {_p2_rev_err}")
            else:
                st.info(
                    "Pass 1 accruals CSV not found in this session. "
                    "Complete Pass 1 first, or upload the Accruals JE CSV below.",
                    icon="ℹ️",
                )
                _p2_rev_manual = st.file_uploader(
                    "Upload Accruals JE CSV (GA_Accruals_JE.csv)",
                    type=["csv"],
                    key="p2_rev_manual_upload",
                )
                if _p2_rev_manual is not None:
                    _p2_rev_manual_path = os.path.join(
                        st.session_state.temp_dir, "p2_rev_manual_accruals.csv"
                    )
                    with open(_p2_rev_manual_path, 'wb') as _p2_wf:
                        _p2_wf.write(_p2_rev_manual.read())
                    st.session_state['p2_rev_manual_csv'] = _p2_rev_manual_path
                    st.rerun()
                # Use manually uploaded file if available
                _p2_rev_src_manual = st.session_state.get('p2_rev_manual_csv')
                if _p2_rev_src_manual and os.path.exists(_p2_rev_src_manual):
                    _p2_rev_period_m = st.text_input(
                        "Reversing period",
                        help="e.g. 'May-2026'",
                        key="p2_rev_period_manual",
                    )
                    if _p2_rev_period_m:
                        try:
                            from accrual_entry_generator import build_reversing_je_csv as _p2_build_rev_m
                            _p2_rev_out_m = os.path.join(
                                st.session_state.temp_dir,
                                f"{_pfx_int}_Reversing_JE_{_p2_rev_period_m.replace('-', '')}.csv",
                            )
                            _p2_build_rev_m(
                                source_etl_path=_p2_rev_src_manual,
                                next_period=_p2_rev_period_m,
                                output_path=_p2_rev_out_m,
                                property_code=getattr(_active_cfg, 'property_code', '') or '',
                            )
                            with open(_p2_rev_out_m, 'rb') as _p2_rf_m:
                                st.download_button(
                                    label=f"⬇️ Download Reversing JE — {_p2_rev_period_m}",
                                    data=_p2_rf_m.read(),
                                    file_name=os.path.basename(_p2_rev_out_m),
                                    mime="text/csv",
                                    use_container_width=True,
                                    key="p2_dl_rev_je_manual",
                                )
                        except Exception as _p2_rev_err_m:
                            st.error(f"Reversing JE generation failed: {_p2_rev_err_m}")

        # ── Sign-off Checklist ─────────────────────────────────────────────────
        st.divider()
        st.markdown("### Sign-off Checklist")
        st.caption(
            "Review each section below and sign off when complete. "
            "Sign-offs are locked for this session. Export the sign-off sheet "
            "before downloading the full package — it will be included automatically."
        )

        _SIGNOFF_REVIEWERS = (_active_cfg.team_members
                              if _active_cfg.team_members
                              else ["[Property Accountant]", "[Property Manager]", "[Accounting Manager/Controller]"])

        for _so_idx, _so_item in enumerate(_SIGNOFF_ITEMS):
            _so_existing = st.session_state.signoff_state.get(_so_idx)
            _col_item, _col_rev, _col_btn, _col_status = st.columns([4, 2, 1.2, 3])
            with _col_item:
                st.markdown(f"**{_so_idx + 1}. {_so_item}**")
            if _so_existing:
                with _col_status:
                    st.markdown(
                        f"<span style='color:#2E7D32;font-weight:600;'>"
                        f"✅ {_so_existing['signed_by']} &nbsp;·&nbsp; {_so_existing['timestamp']}"
                        f"</span>",
                        unsafe_allow_html=True,
                    )
            else:
                with _col_rev:
                    _so_reviewer = st.selectbox(
                        "Reviewer", _SIGNOFF_REVIEWERS,
                        key=f"so_rev_{_so_idx}",
                        label_visibility="collapsed",
                    )
                with _col_btn:
                    if st.button("Sign Off", key=f"so_btn_{_so_idx}",
                                 use_container_width=True):
                        st.session_state.signoff_state[_so_idx] = {
                            "signed_by": _so_reviewer,
                            "timestamp": datetime.now().strftime("%m/%d/%Y %H:%M"),
                        }
                        # Auto-detect Close Tracker Step 7 when all sign-offs complete
                        _so_total = len(_SIGNOFF_ITEMS)
                        _so_done  = len(st.session_state.signoff_state)
                        # +1 because we just added one above
                        if (_so_done >= _so_total and
                                7 not in st.session_state.close_tracker):
                            st.session_state.close_tracker[7] = {
                                "completed_by": _so_reviewer,
                                "timestamp":    datetime.now().strftime("%m/%d/%Y %H:%M"),
                                "auto":         True,
                            }
                            _save_checklist_now()
                            st.session_state.last_completed_step = 7
                        st.rerun()
                with _col_status:
                    st.markdown(
                        "<span style='color:#9E9E9E;'>Pending</span>",
                        unsafe_allow_html=True,
                    )

        # Export sign-off sheet
        st.markdown("")
        _so_exp_col, _ = st.columns([2, 5])
        with _so_exp_col:
            if st.button("📄 Export Sign-off Sheet", use_container_width=True,
                         help="Generates GA_Signoff_Record.xlsx and adds it to the ZIP"):
                try:
                    from signoff_generator import generate_signoff_xlsx as _gen_so
                    _so_path = os.path.join(st.session_state.temp_dir,
                                            "GA_Signoff_Record.xlsx")
                    _gen_so(
                        output_path   = _so_path,
                        signoff_state = st.session_state.signoff_state,
                        items         = _SIGNOFF_ITEMS,
                        period        = result.period or close_period,
                        property_name = result.property_name or _prop_display,
                    )
                    st.session_state.pass2_output_files["signoff_record"] = _so_path
                    st.success("Sign-off sheet exported — included in the ZIP.", icon="✅")
                    st.rerun()
                except Exception as _soe:
                    st.error(f"Sign-off export failed: {_soe}")

        _so_dl_path = p2.get("signoff_record")
        if _so_dl_path and os.path.exists(_so_dl_path):
            with open(_so_dl_path, "rb") as _so_f:
                st.download_button(
                    label="⬇️ Download Sign-off Record",
                    data=_so_f.read(),
                    file_name=f"{_pfx_del}_{period_label}_Signoff_Record.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

    # ── Post-Close Adjustments (always visible in Pass 2 tab) ─────────────────
    st.divider()
    st.markdown("### Closing JE Entries")
    st.caption(
        "After reviewing the QC workbook you may identify corrections, missed re-accruals, "
        "or reclassifications that still need to post. Enter them here and download as a "
        "Yardi-import CSV. "
        "Use **Add JE Lines** to append a pre-numbered pair. "
        "Debits must equal Credits for each **JE #** before export."
    )

    # ── Add JE Lines button ────────────────────────────────────────────────────
    # Use _pcje_latest (saved from the previous render's editor output) so that
    # any account codes / amounts the user already typed are preserved when new
    # rows are appended.  Falls back to post_close_je_df on first render.
    if st.button("➕ Add JE Lines", key="pcje_add_btn"):
        import pandas as _pd_pcje_add
        _existing_pcje = st.session_state.get("_pcje_latest", st.session_state.post_close_je_df)
        _used_jes = (
            _existing_pcje["JE #"]
            .str.strip()
            .replace("", None)
            .dropna()
            .unique()
            .tolist()
        )
        _next_n   = len(_used_jes) + 1
        _next_lbl = f"PC-{_next_n:03d}"
        _new_pair = _pd_pcje_add.DataFrame({
            "JE #":             [_next_lbl, _next_lbl],
            "Description":      ["", ""],
            "Account Code":     ["", ""],
            "Debit ($)":        [0.0, 0.0],
            "Credit ($)":       [0.0, 0.0],
            "Line Description": ["", ""],
        })
        st.session_state.post_close_je_df = _pd_pcje_add.concat(
            [_existing_pcje, _new_pair], ignore_index=True
        )
        st.rerun()

    _pcje_edited = st.data_editor(
        st.session_state.post_close_je_df,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "JE #":             st.column_config.TextColumn("JE #", width="small"),
            "Description":      st.column_config.TextColumn("JE Description", width="medium"),
            "Account Code":     st.column_config.TextColumn("Account Code", width="small"),
            "Debit ($)":        st.column_config.NumberColumn("Debit ($)", format="%.2f", min_value=0.0, width="small"),
            "Credit ($)":       st.column_config.NumberColumn("Credit ($)", format="%.2f", min_value=0.0, width="small"),
            "Line Description": st.column_config.TextColumn("Line Description", width="large"),
        },
        key=f"post_close_je_editor_{st.session_state.get('editor_reset_count', 0)}",
    )
    # Save latest edits for use by Add JE Lines on next render.
    # Do NOT feed _pcje_edited back into post_close_je_df here — doing so changes
    # the base DataFrame on every rerender, which causes Streamlit's data_editor
    # to reset its internal delta state and lose whatever the user just typed.
    st.session_state["_pcje_latest"] = _pcje_edited

    _pcje_valid = _pcje_edited[
        _pcje_edited["Account Code"].fillna("").str.strip().astype(bool) &
        ((_pcje_edited["Debit ($)"] != 0) | (_pcje_edited["Credit ($)"] != 0))
    ]

    if not _pcje_valid.empty:
        # ── Validation: each JE # debits must equal credits ──────────────────
        # Only validate a JE when ALL its rows have both account code and amount
        # filled — skip JEs still being entered to avoid mid-entry errors.
        _pcje_errors = []
        for _jn, _grp in _pcje_valid.groupby("JE #"):
            _all_rows = _pcje_edited[_pcje_edited["JE #"] == _jn]
            _incomplete = _all_rows[
                ~_all_rows["Account Code"].fillna("").str.strip().astype(bool) |
                ((_all_rows["Debit ($)"].fillna(0) == 0) &
                 (_all_rows["Credit ($)"].fillna(0) == 0))
            ]
            if not _incomplete.empty:
                continue  # still being filled in — don't flag yet
            _total_dr = _grp["Debit ($)"].sum()
            _total_cr = _grp["Credit ($)"].sum()
            if abs(_total_dr - _total_cr) > 0.01:
                _pcje_errors.append(
                    f"JE #{_jn}: Debits ${_total_dr:,.2f} ≠ Credits ${_total_cr:,.2f}"
                )

        if _pcje_errors:
            for _err in _pcje_errors:
                st.error(f"❌ {_err}", icon="❌")
        else:
            st.success(
                f"✅ {len(_pcje_valid)} line{'s' if len(_pcje_valid) != 1 else ''} "
                f"across {_pcje_valid['JE #'].nunique()} JE(s) — balanced",
                icon="✅",
            )

            # ── Build JE lines for CSV export ─────────────────────────────────
            _pcje_lines = []
            _pcje_num   = 1
            for _jn, _grp in _pcje_valid.groupby("JE #"):
                _line_seq = 1
                for _, _row in _grp.iterrows():
                    _dr = float(_row["Debit ($)"] or 0)
                    _cr = float(_row["Credit ($)"] or 0)
                    if _dr > 0 or _cr > 0:
                        _pcje_lines.append({
                            'je_number':      f'PCJ-{str(_jn).strip() or _pcje_num:04}',
                            'line':           _line_seq,
                            'date':           '',
                            'account_code':   str(_row["Account Code"]).strip(),
                            'account_name':   str(_row["Line Description"] or _row["Account Code"]).strip(),
                            'description':    str(_row["Description"] or "Post-close adjustment").strip(),
                            'reference':      'POST-CLOSE',
                            'debit':          round(_dr, 2),
                            'credit':         round(_cr, 2),
                            'vendor':         '[Post-Close Adj]',
                            'invoice_number': '',
                            'source':         'post_close',
                            'confidence':     'high',
                        })
                        _line_seq += 1
                _pcje_num += 1

            # ── Generate CSV ───────────────────────────────────────────────────
            _pcje_csv_path = os.path.join(
                st.session_state.temp_dir, "GA_PostClose_JE.csv"
            )
            _p2er = st.session_state.pass2_engine_result
            # Prefer Pass 2 GL period; fall back to the sidebar period selector so
            # the date field is never left blank (which triggers a today's-date fallback).
            _pcje_period = (
                (_p2er.period if _p2er else '') or ''
                or period_key_to_label(
                    st.session_state.get('checklist_period_key', '')
                )
            )
            # ETL PROPERTY field is capped at 8 chars by Yardi.
            # Use yardi_etl_code from config when set; otherwise truncate property_code.
            _pcje_etl_code = (
                getattr(_active_cfg, 'yardi_etl_code', '') or _active_cfg.property_code
            )[:8]
            try:
                from accrual_entry_generator import generate_etl_csv as _gen_etl_pc
                _gen_etl_pc(
                    _pcje_lines,
                    _pcje_csv_path,
                    period=_pcje_period,
                    property_code=_pcje_etl_code,
                    auto_reverse=False,  # post-close JEs are permanent — no reversal
                )
                st.session_state.pass2_output_files["post_close_je_csv"] = _pcje_csv_path
            except Exception as _pcje_err:
                st.warning(f"Post-close JE CSV generation skipped: {_pcje_err}")

            # ── Download button ────────────────────────────────────────────────
            _pcje_path = st.session_state.pass2_output_files.get("post_close_je_csv")
            if _pcje_path and os.path.exists(_pcje_path):
                with open(_pcje_path, "rb") as _f:
                    st.download_button(
                        label="⬇️ Download Post-Close JE CSV (Yardi import)",
                        data=_f.read(),
                        file_name=f"GA_PostClose_JE_{datetime.now().strftime('%Y%m%d')}.csv",
                        mime="text/csv",
                        use_container_width=True,
                    )


# ──────────────────────────────────────────────────────────────
# TAB 3 — HOW TO USE
# ──────────────────────────────────────────────────────────────
with tab3:
    st.markdown("## 📖 Monthly Close Operating Guide")
    st.markdown(
        "This is the full runbook for operating the pipeline — written for anyone running a "
        "close, whether GRP staff or an outsourced accounting team member, with no prior "
        "context assumed. It covers every step, every file, every output, and what to do "
        "when something needs review."
    )
    st.info(
        "**Roles referenced throughout:** the **Property Accountant** runs Pass 1 and Pass 2 "
        "(this may be GRP staff or an outsourced team member). The **Property Manager** "
        "reviews outputs before release. The **Accounting Manager/Controller** is the final reviewer — sees the package "
        "only after the Property Accountant and Property Manager have both signed off. "
        "Actual names for each role are configured per property in the Properties tab.",
        icon="👥",
    )

    # ── Quick-reference flow ──────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### At a Glance")
    st.markdown("""
| Step | Who | Action |
|------|-----|--------|
| 1 | Property Accountant | Export pre-close files from Yardi & pull bank statements |
| 2 | Property Accountant | Upload all Pass 1 files → click **Generate JEs** |
| 3 | Property Accountant | Review the JE list, adjust/exclude as needed → download JE CSVs |
| 4 | Property Accountant | Import CSVs into Yardi → run the final close |
| 5 | Property Accountant | Re-export final GL, TB, BC, Bank Rec, Loan Statements from Yardi |
| 6 | Property Accountant | Upload Pass 2 files (+ raw report overrides) → click **Generate Reports** |
| 7 | Property Accountant | Review QC results and JE posting verification |
| 8 | Property Accountant | Complete Close Tracker steps + sign off on the checklist |
| 9 | Property Manager | Reviews Workpapers, QC Workbook, Audit Trail, and Annotated BC → signs off |
| 10 | Accounting Manager/Controller | Final review of the released package |
""")
    st.markdown(
        "> **Start here every month:** the **Dashboard** tab (leftmost tab) shows the close "
        "checklist for the current period and, once a few months of history exist, trending "
        "charts (NOI, revenue, expenses, cash) across periods. Check it first to see where "
        "this property's close currently stands."
    )

    # ── PASS 1 ────────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("## Pass 1 — Pre-Close")
    with st.expander("📥  Step 1 — What to Upload", expanded=True):
        st.markdown(
            "Upload all files into the **Pass 1 upload zone** at the top of the Pass 1 tab. "
            "The pipeline auto-detects each file type — if it guesses wrong, use the "
            "dropdown next to the filename to correct it."
        )
        st.markdown("#### Core Close Files — export from Yardi before running the close")
        st.markdown("""
| File | Where to get it in Yardi | Notes |
|------|--------------------------|-------|
| **Yardi GL Detail** | Reports → General Ledger → by Property, Period = current month, Book = Accrual | Most important file — drives all accrual logic. **The only required file.** |
| **Yardi Trial Balance** | Reports → Trial Balance → same period & book | Used for GL ↔ TB tie-out |
| **Yardi Budget Comparison** | Reports → Budget Comparison → PTD + YTD columns, same period | Drives historical pattern accruals and variance commentary |
| **12-Month Income Statement** | Reports → 12-Month Statement → trailing 12 months | Used for historical recurring accruals and month-over-month swing detection |
| **Nexus Invoice Detail** | Nexus AP → Export open invoices → .xls format | Open invoices not yet in the GL |
| **Kardin Budget** | Kardin → Export → qryExportData format | Annual budget; used for payroll bonus accruals |
| **Yardi Receivable Detail / Summary** | Reports → Receivable Detail or Summary → current period | Used to calculate management fee on cash received |
| **Yardi AR Detail Aging** | Reports → AR Aging Detail → current period | Used alongside Receivable Detail to exclude prepayments from the fee basis |
""")
        st.markdown("#### Bank Statements")
        st.markdown("""
| File | Where to get it | Notes |
|------|-----------------|-------|
| **Yardi / PNC Bank Rec** | Yardi Reports → Bank Reconciliation → export as PDF | Preferred source — pipeline reads the pre-computed reconciled balance |
| **KeyBank DACA Statement** | KeyBank online → account x5132 → monthly statement PDF | Used as management fee cash-received basis |
| **BofA Development Statement** | BofA online → development account → monthly PDF | Balance only; development account is dormant |
| **Berkadia Loan Statement(s)** | Berkadia portal → monthly loan statements → PDF (all 3 loans) | ⚠️ Upload the statement **due the 7th of the following month** — e.g. for the January close, upload the Feb 7 statement. Interest is paid in arrears: the Feb 7 payment covers January's interest. |
""")
        st.markdown("#### Reference Files")
        st.markdown(f"""
| File | Where to get it | Notes |
|------|-----------------|-------|
| **Prior Month Prepaid Ledger** | Downloaded from last month's Pass 1 run (part of the Carry-Forward Package) | First close for a property: use the seed ledger built for that property's go-live date |
""")
        st.markdown(
            "> **Tip:** You don't need every file every month. The pipeline runs on whatever is "
            "uploaded and flags anything it couldn't calculate. The GL is the only required file."
        )

    # ── One-Off Accruals ──────────────────────────────────────────────────────
    with st.expander("✏️  Step 2 — Fill in the One-Off Accruals Table"):
        st.markdown("""
The **One-Off Accruals** table (Pass 1 tab) is for items the pipeline can't detect automatically —
typically small recurring contracts where no invoice arrives until after close. Which rows are
pre-seeded and their typical amounts are property-specific (configured per property) — review
and adjust every month before generating JEs.

Each row creates a **DR expense / CR 213100 Accrued Expenses** journal entry (labeled `SUP-XXXX`)
that auto-reverses next period. A row entered at **$0** acts as an exclusion flag rather than a JE —
use this to tell the pipeline "don't auto-accrue this account this month" without deleting the row.
""")

    # ── JE Review & editing ────────────────────────────────────────────────────
    with st.expander("🧾  Step 3 — Review the Generated JEs"):
        st.markdown("""
After clicking **Generate JEs**, the results page shows every JE grouped into expandable
sections by **credit account** — e.g. all JEs crediting 213100 Accrued Expenses in one section.
Each JE line has:

- **Include** checkbox — uncheck to exclude a JE from the CSV without deleting it; re-check to
  restore it. This updates the downloadable CSV live, no re-run needed.
- **Description** — editable inline; edits apply immediately.
- **Amount** — editable inline; both the DR and CR leg update together so the JE stays balanced.

**7xxxxx Intercompany Recode** — if the GL has any corporate-expense (7xxxxx) activity miscoded
onto the property, a recode table appears automatically. Enter the correct 6xxxxx or 8xxxxx
target account on each DR row, then click **Re-run Pass 1** (a second copy of this button sits
directly below the recode table, so you don't need to scroll back to the top) to include the
recode JEs in the CSV.

**Add Missed Entries** — for JEs the pipeline didn't generate at all (not a recode, not
covered by One-Off Accruals). Fill in as many rows as you need in the table — DR/CR accounts,
description, amount — then click **Add All Entries** once to submit them all together. These
entries survive a **Re-run Pass 1** click — they are not regenerated by the pipeline, so
re-running only adds to them, never discards them.

> **Re-running Pass 1 after any edit** (table changes, recode entries, new uploads) always
> reflects the current state of everything on the page at the moment you click — you should
> never need to click Re-run more than once for an edit to take effect.
""")

    # ── Pass 1 Outputs ────────────────────────────────────────────────────────
    with st.expander("📄  Step 4 — What Pass 1 Produces"):
        st.markdown(f"""
After clicking **Generate JEs**, download either the full zip package or individual files:

| File | Contents | What to do with it |
|------|----------|--------------------|
| **{_pfx_int}_Accruals_JE.csv** | All accrual entries: Nexus invoices, utility proration, service accruals, historical recurring, management fee, contract supplements, payroll bonus accruals, tenant utility billings, recode entries | **Import into Yardi** as a journal batch |
| **{_pfx_int}_Prepaid_Ledger.xlsx** | Updated prepaid amortization schedule with this month's releases applied | **Save** — part of next month's Carry-Forward Package (see Step 4b) |
| **{_pfx_int}_JE_Cache.json** | Every JE line generated this run, in a compact format | **Save if you plan to run Pass 2 in a different browser session** — upload it in Pass 2 to enable the full Audit Trail and JE posting verification (see Step 8) |

> The results page also shows a **summary table** of every entry generated, grouped by layer
> (Layer 1 Nexus, Layer 2 Proration, Layer 3 Historical, Layer 4 Bonus, etc.) so you can review
> before posting.
""")

    # ── Yardi Upload Step ─────────────────────────────────────────────────────
    with st.expander("⬆️  Step 5 — Post to Yardi & Run the Close"):
        st.markdown(f"""
**In Yardi, before running the final close:**

1. Go to **Journals → Import Journal Entries**
2. Import `{_pfx_int}_Accruals_JE.csv` → review the batch → post
3. Verify the journal batch posted cleanly (no errors)
4. Run the **month-end close** in Yardi (locks the period)

**What Yardi auto-reverses on the 1st of next month:**
All accrual and management fee entries auto-reverse on the first day of the following period.
This is standard accrual accounting — no manual reversal needed.

> **Note:** The Singerman 8-tab monthly report (Balance Sheet, Income Statement, T12,
> Trial Balance MTD/YTD, GL MTD/YTD, Tenancy) is downloaded directly from Yardi
> after the close — it is **not** generated by this pipeline.
""")

    # ── Pass 2 ────────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("## Pass 2 — Post-Close")
    with st.expander("📥  Step 6 — What to Re-Upload", expanded=True):
        st.markdown(
            "After the close runs in Yardi, re-export the **final versions** of these files. "
            "They reflect all journal entries that were posted (including the ones from Pass 1). "
            "Upload them in the **Pass 2 upload zone** at the top of the Pass 2 tab — they override the "
            "Pass 1 versions for the final reports."
        )
        st.markdown("""
| File | Why it needs to be re-exported |
|------|-------------------------------|
| **Yardi GL Detail** | Must include all accrual JEs that were posted — the pre-close GL is missing them |
| **Yardi Trial Balance** | Final balances after all JEs posted — used for GL ↔ TB tie-out |
| **Yardi Budget Comparison** | Actuals update after JEs post — needed for accurate variance commentary |
| **Yardi Bank Rec PDF** | Final reconciliation with outstanding checks as of close |
| **Berkadia Loan Statements** | Same file as Pass 1 (statement due the 7th of the following month) — re-upload or the pipeline reuses the Pass 1 version |

> All other Pass 1 files (Nexus, Kardin, bank statements, T12, etc.) do **not** need to be
> re-exported — the pipeline reuses them automatically from Pass 1.
""")

    # ── Workpaper Raw Report Overrides ─────────────────────────────────────────
    with st.expander("🗂️  Step 6b — Workpaper Raw Report Overrides"):
        st.markdown("""
Six workpaper tabs are sourced from raw Yardi/bank reports rather than computed from the GL:
**111100 PNC Cash, 115100 DACA, 131100 AR Aging, 221100 Prepaid Rent, 211100 AP,** and
**111210 BofA Development**. Upload the corresponding raw file each period in the
**"Workpaper raw report overrides"** section (Pass 2 tab) so these tabs refresh with current
data instead of staying frozen:

| Upload slot | Populates |
|-------------|-----------|
| AR Aging Detail — 133100 AR Control | 131100 AR Aging **and** 221100 Prepaid Rent (one Yardi report covers both — same file, both tabs) |
| AP Aging Detail — 211300 AP Control | 211100 Accounts Payable |
| Bank Rec — 111100 PNC Operating | 111100 PNC Cash |
| Bank Rec — 115100 DACA | 115100 DACA |

> **If a file isn't uploaded for one of these tabs**, that tab shows an explicit
> **"No data uploaded this period"** message instead of silently reusing last month's data —
> upload the missing file and re-run if you see this.

> **Don't have the Excel export?** The two Bank Rec slots also accept a **PDF** — drop the
> same Bank Rec PDF you'd upload elsewhere and assign it to the matching Bank Rec slot; it's
> parsed the same way and takes priority over whatever the main Bank Rec upload produced for
> this period. AR Aging, AP Aging, and the other slots still need the Excel version — there's
> no PDF parser for those yet.
""")

    # ── Pass 2 Outputs ────────────────────────────────────────────────────────
    with st.expander("📊  Step 7 — What Pass 2 Produces"):
        st.markdown(f"""
After clicking **Generate Reports**, download the full package or individual files:

| File | Contents | Audience |
|------|----------|----------|
| **{_pfx_int}_Workpapers.xlsx** | GL ↔ TB tie-out for all balance sheet accounts, bank rec detail, debt service schedule. Grows month-over-month when the prior month file is uploaded. | Property Accountant / Property Manager |
| **{_pfx_int}_QC_Workbook.xlsx** | 7-point QC checklist — see Step 8 below | Property Accountant |
| **{_pfx_int}_Exceptions_Report.xlsx** | All flagged issues with severity (Error / Warning / Info), source, and recommended action | Property Accountant |
| **{_pfx_int}_BC_Internal.xlsx** | Annotated Budget Comparison with variance commentary in columns L/M — GRP internal use only | Property Accountant / Property Manager |
| **{_pfx_int}_Audit_Trail.xlsx** | Every JE's math, the exact Yardi ETL import rows, management fee calculation detail, and QC results in one file — the record an auditor would review | Property Manager / Auditor |
| **{_pfx_int}_Signoff_Record.xlsx** | Who reviewed and approved each section of the close package, and when | Property Manager / Accounting Manager/Controller |
| **{_pfx_int}_Close_Tracker.xlsx** | The 9-step close lifecycle record, from JLL handoff through Accounting Manager/Controller release | Property Manager |
| **{_pfx_int}_Run_Log.csv** | Running history of every close run for this property (timestamp, JE counts, QC results) — not financial detail, just a log | Internal reference |
| Management Fee Invoice (PDF) | JLL/GRP management fee invoice for this period | Property Manager |

> **Before sending to Property Manager:** clear any open Errors in the Exception Report.
> Warnings should be reviewed but may be acceptable.
""")

    # ── QC Results ─────────────────────────────────────────────────────────────
    with st.expander("✅  Step 8 — Reviewing QC Results"):
        st.markdown("""
The QC Workbook runs 7 checks automatically:

| Check | What it verifies |
|-------|-------------------|
| 1. TB → BC Tie-out | Trial Balance and Budget Comparison actuals agree |
| 2. Budget Variances | Tier 1 (≥$5K or 5%) and Tier 2 ($2.5K–$5K) flags against budget |
| 3. Trial Balance Self-Balance | TB debits = credits |
| 4. Month-over-Month Swings | Any account swinging more than the configured threshold vs. last month |
| 5. BS Workpaper Tie-out | Workpaper ending balances agree with GL and TB |
| 6. Accruals vs. Budget Coverage | Whether accrual detection covered everything the budget implies should be accrued |
| 7. Miscellaneous | Insurance, management fee, and other property-specific checks |

Each check shows **PASS**, **FLAG**, or **FAIL**. FLAG/FAIL items need review before the package
is released — check the corresponding row's detail for what to investigate.

**JE Posting Verification** (its own tab inside the QC Workbook) compares every Pass 1 JE against
the final GL to confirm it actually posted:

| Status | Meaning |
|--------|---------|
| VERIFIED | All DR/CR lines found in the GL with matching amounts |
| AMOUNT_MISMATCH | Lines found, but at least one amount differs by more than $0.02 |
| PARTIAL | Only some lines of the JE were found |
| MISSING | No lines found — the JE did not post |

This only runs if the Pass 1 JE data is available — either the same browser session as Pass 1,
or the JE Cache uploaded (see Step 4).
""")

    # ── Post-Close Adjustments ────────────────────────────────────────────────
    with st.expander("🔧  Post-Close Adjustments (Pass 2 JEs)"):
        st.markdown("""
If the review uncovers items that need a correcting entry **after** the close (reclasses,
true-ups, corrections), use the **Post-Close Adjustments** table in the Pass 2 tab.

- Click **➕ Add JE Lines** to add a balanced pair of DR / CR rows
- Each JE pair is auto-numbered (PC-001, PC-002, …)
- Debits must equal Credits per JE# — the pipeline validates before export
- Download `GA_PostClose_JE.csv` and import into Yardi as a separate journal batch

Post-close JEs are **not** auto-reversing — they are permanent adjustments.
""")

    # ── Close Tracker & Sign-off ────────────────────────────────────────────────
    st.markdown("---")
    with st.expander("📋  Close Tracker & Sign-off"):
        st.markdown("""
The **Dashboard** tab tracks the full close lifecycle in 9 steps, from JLL's initial bank rec
through the Accounting Manager/Controller's final release:

`0` JLL Completes Bank Rec & Payments · `1` Pass 1 Files Uploaded & JEs Generated ·
`2` JEs Uploaded to Yardi · `3` Final Close Run in Yardi · `4` Final Files Re-Exported from Yardi ·
`5` Pass 2 Files Uploaded · `6` Reports Generated · `7` QC Review Complete (Property Accountant /
Property Manager) · `8` Final Package Released to Accounting Manager/Controller

Steps 1, 5, and 6 auto-complete when you run Pass 1 / upload Pass 2 files / generate reports.
The rest require clicking **Mark Complete** on the Dashboard as each step actually happens.

**Sign-off checklist** (Pass 2 tab, near the bottom) tracks who reviewed each specific
deliverable — Bank Rec (Operating and DACA), Management Fee Invoice, GL vs TB Workpaper,
Variance Commentary, the QC Checklist, Equity Tabs, and the Exception Report. Each item is
signed by name with a timestamp; click **Export Sign-off Sheet** once everything is checked
off to produce the permanent record.
""")

    # ── Final Deliverables ────────────────────────────────────────────────────
    st.markdown("---")
    with st.expander("📬  Final Deliverables — What Goes Where"):
        st.markdown(f"""
#### To Property Manager
| Item | Source |
|------|--------|
| Workpapers (GL ↔ TB tie-out) | `{_pfx_int}_Workpapers.xlsx` from Pass 2 |
| Annotated Budget Comparison | `{_pfx_int}_BC_Internal.xlsx` from Pass 2 |
| Audit Trail | `{_pfx_int}_Audit_Trail.xlsx` from Pass 2 |
| Singerman 8-Tab Monthly Report | Downloaded directly from Yardi |

#### To Singerman (Capital Partner)
| Item | Source |
|------|--------|
| Monthly Report (BS, IS, T12, TB, GL, Tenancy) | Downloaded directly from Yardi — not from this pipeline |

#### Retained Internally (GRP)
| Item | Purpose |
|------|---------|
| `{_pfx_int}_QC_Workbook.xlsx` | GRP internal QC sign-off |
| `{_pfx_int}_Exceptions_Report.xlsx` | Audit trail of all flagged items |
| `{_pfx_int}_Signoff_Record.xlsx` | Reviewer sign-off record |
| `{_pfx_int}_Close_Tracker.xlsx` | Close lifecycle record |
| `{_pfx_int}_Run_Log.csv` | Historical run log |
""")

    # ── Monthly Carry-Forward Checklist ────────────────────────────────────────
    st.markdown("---")
    with st.expander("➡️  Monthly Carry-Forward — What to Save for Next Month", expanded=True):
        st.markdown(f"""
At the end of Pass 2, download the **Carry-Forward Package** — a small zip with exactly the
2 files the *next* close needs:

| File | Where it goes next month |
|------|---------------------------|
| **{_pfx_int}_Prepaid_Ledger_UPLOAD_NEXT_CLOSE.xlsx** | Upload as the **Prior Month Prepaid Ledger** in Pass 1 |
| **{_pfx_int}_Workpaper_UPLOAD_NEXT_CLOSE.xlsx** | Upload as the **Prior Month Workpaper** in Pass 2 |

This is the single most important handoff between closes — losing either file means the prepaid
schedule or the rolling workpaper history has to be rebuilt manually. If the workpaper isn't
uploaded next month and this property has prior closes on record, the app will warn you rather
than silently starting the workpaper over.
""")

    # ── Troubleshooting ───────────────────────────────────────────────────────
    st.markdown("---")
    with st.expander("🛠️  Common Issues & Tips"):
        st.markdown(f"""
**File uploaded but not recognized**
→ Use the dropdown next to the filename in the upload zone to manually select the file type.

**Management fee shows $0**
→ The pipeline couldn't find cash received. Check that the DACA statement or Receivable Detail
was uploaded. If both are missing, the fee will be $0 and will need a manual One-Off entry.

**Workpaper doesn't include prior months**
→ Upload the prior month's `{_pfx_int}_Workpapers.xlsx` in the Pass 2 upload zone (see the
Carry-Forward section above). Leave blank only for a property's genuinely first close.

**A raw-report workpaper tab (PNC Cash, DACA, AR Aging, Prepaid Rent, AP, BofA Dev) shows
"No data uploaded this period"**
→ Upload the matching file in "Workpaper raw report overrides" (Step 6b) and re-run.

**Prepaid ledger shows 0 active items you know should be there, or "0 released" for the whole
active list**
→ Double-check the **Prior Month Prepaid Ledger** upload slot actually has last month's
`{_pfx_int}_Prepaid_Ledger_UPLOAD_NEXT_CLOSE.xlsx` selected — not last month's Workpaper. If
the wrong file type ends up there, the app now shows a clear error naming the problem instead
of silently loading 0 items, but it's still an easy mix-up if both files are open at once when
picking one to upload.

**JE Verification / Audit Trail JE detail is missing in Pass 2**
→ Pass 1 wasn't run in this session. Upload the JE Cache JSON from Pass 1's download package
(see Step 4).

**RE Tax — what to enter each month**
→ Enter the quarterly bill amount every month (all 3 months in each cycle use the same number).
Payment months (Jan/Apr/Jul/Oct): pipeline defers 2/3 → DR 135120 Prepaid RE Taxes / CR 641110.
Release months (Feb/Mar/May/Jun/Aug/Sep/Nov/Dec): pipeline releases 1/3 → DR 641110 / CR 135120.
Leave $0 only if the RE Tax JE has already been posted manually in Yardi.

**Accrual entry says "REVIEW REQUIRED"**
→ This is a low-confidence entry — the account has a budget but no GL history this year.
Review whether the expense was actually incurred before posting. Uncheck the Include box in
the JE review list if it should not be posted.

**Reset button**
→ Use **Reset All** (sidebar) to clear all uploads and start fresh. Use **Reset Pass 2**
(Pass 2 tab) to clear only the final-close files without losing Pass 1 results.
""")

    st.markdown("---")
    try:
        from version import get_version as _get_guide_version
        st.caption(f"Pipeline built by GRP · {_get_guide_version()}")
    except Exception:
        st.caption("Pipeline built by GRP")


# ──────────────────────────────────────────────────────────────
# TAB 4 — PROPERTIES SETUP
# ──────────────────────────────────────────────────────────────
with tab4:
    from property_writer import (
        build_config_dict, config_to_yaml,
        save_local, save_to_github, github_configured,
    )
    from property_config import discover_properties as _disc_props

    st.markdown("## ⚙️ Property Setup")
    st.markdown(
        "Add or edit properties here. Each property is stored as a YAML config file — "
        "no GitHub access or code changes required."
    )
    with st.expander("📖 New here? How to onboard a property", expanded=False):
        st.markdown(
            "**1. Files first, config second.** The sections above the Add/Edit form "
            "below — Current Year Budget, Tenancy Schedule / Rent Roll, 12-Month GL "
            "History, Bank Statement — read a real document instead of asking you to "
            "type something blind. Drop those in before filling out the form so a few "
            "of its fields (Bank Accounts, Kardin Budget Filename) can be copied "
            "straight from what gets detected.\n\n"
            "**2. Then the Add/Edit form**, section by section — pick "
            "**➕ Create new property** below, or select an existing one to edit.\n\n"
            "**3. Full step-by-step checklist further down this page** — "
            "**📝 First-Close Checklist**, near the bottom — covers exactly what to do "
            "before a property's very first close, split into *Existing GRP Property* "
            "(an acquisition, carrying prior history forward) vs. *New Property* "
            "(ground-up, nothing to carry forward). Start there once the config is saved."
        )

    if github_configured():
        st.success("✅ GitHub connected — saved configs deploy automatically in ~2 min.", icon="🔗")
    else:
        st.warning(
            "⚠️ GitHub not connected. Configs will be saved locally (dev only) and available "
            "to download. To enable auto-deploy: add `[github]` token + repo to Streamlit secrets.",
            icon="⚠️"
        )

    # ── Existing properties ───────────────────────────────────────────────────
    _existing = _disc_props(str(_DATA_DIR))
    if _existing:
        with st.expander(f"📋 Existing properties ({len(_existing)})", expanded=False):
            for _ep in _existing:
                _ec = _ep['cfg']
                _sys_badge = (
                    "🏦 MRI" if getattr(_ec, 'property_system', 'yardi').lower() == 'mri'
                    else "🏢 Yardi"
                )
                st.markdown(
                    f"**{_ec.display()}** &nbsp; `{_ec.property_code}` &nbsp; {_sys_badge} &nbsp;|&nbsp; "
                    f"{_ec.property_address} &nbsp;|&nbsp; "
                    f"{'  ·  '.join(f'{fl.name} {fl.rate:.2%}' for fl in _ec.management_fees)}"
                )
    st.markdown("---")

    # ── Add / Edit property form ──────────────────────────────────────────────
    # Apply a pending reset (set by the Delete flow further down, after a
    # property is removed) BEFORE this widget is instantiated — writing to
    # its session_state value is only safe here, not after it's rendered.
    if st.session_state.pop('_prop_select_reset_pending', False):
        st.session_state['prop_setup_edit_select'] = "➕ Create new property"
    _edit_code = st.selectbox(
        "Edit existing or create new",
        options=["➕ Create new property"] + [p['code'] for p in _existing],
        key="prop_setup_edit_select",
    )
    _is_new = _edit_code == "➕ Create new property"
    _edit_cfg = None if _is_new else next(
        (p['cfg'] for p in _existing if p['code'] == _edit_code), None
    )

    def _ef(field, default=''):
        """Return existing config value or default."""
        if _edit_cfg is None:
            return default
        return getattr(_edit_cfg, field, default) or default

    # ── Property Management System ────────────────────────────────────────────
    _sys_default_idx = (
        0 if (not _edit_cfg or getattr(_edit_cfg, 'property_system', 'yardi').lower() == 'yardi')
        else 1
    )
    _prop_system = st.radio(
        "Property Management System",
        options=["Yardi", "MRI"],
        index=_sys_default_idx,
        horizontal=True,
        key="prop_system_radio",
        help="Yardi: full pipeline support. MRI: onboarding instructions coming soon.",
    )
    _is_mri = (_prop_system == "MRI")

    if _is_mri:
        st.info(
            "**MRI integration is coming soon.**\n\n"
            "MRI Software properties use a different chart of accounts and file export "
            "format from Yardi. Full onboarding instructions and account-mapping tools "
            "will be available in a future pipeline update.\n\n"
            "You can save this property record with system type **MRI** now — it will "
            "appear in the property selector, and MRI-specific configuration fields "
            "will be unlocked when that integration is built.",
            icon="🔜",
        )
        st.markdown("---")

    # ── Onboarding uploaders: clear after a successful save ───────────────────
    # st.file_uploader keeps returning the SAME UploadedFile on every rerun
    # until the user manually removes it from the widget — there's no form
    # here to reset it (file_uploader can't live in st.form). Confirmed bug
    # 2026-08-24: upload a photo/COA/budget for Property A, then switch to
    # Property B without clicking the uploader's own "x" — the next rerun
    # (e.g. Property B's own Save click) re-saves Property A's stale file
    # under Property B's code, since these targets follow whatever _edit_code
    # currently is. Fix: bump this uploader's own generation counter right
    # after a save attempt, which changes its widget key on the next render —
    # a fresh, empty file_uploader, same as bumping key= elsewhere in this
    # app forces a genuine reseed (see Development Notes in CLAUDE.md).
    def _uploader_key(_name: str) -> str:
        return f"{_name}_{st.session_state.get(f'_{_name}_gen', 0)}"

    def _bump_uploader(_name: str) -> None:
        st.session_state[f'_{_name}_gen'] = st.session_state.get(f'_{_name}_gen', 0) + 1

    # Reset per-property onboarding scratch state when the property being
    # edited changes, so a previous property's detected bank accounts don't
    # linger as suggestions for the next one.
    if st.session_state.get('_prop_setup_last_edit_code') != _edit_code:
        st.session_state.prop_bank_detect_rows = []
        st.session_state['_prop_setup_last_edit_code'] = _edit_code

    # ── Building photo upload (outside form — file_uploader can't live in st.form) ──
    st.markdown("### 🏙️ Building Photo")
    st.caption(
        "Upload a photo of the building — displayed in the hero banner for this property. "
        "JPG or PNG, landscape orientation works best (approx 4:1 ratio)."
    )
    _photo_target_code = ('' if _is_new else _edit_code)
    _photo_col1, _photo_col2 = st.columns([2, 1])
    with _photo_col1:
        _hero_upload = st.file_uploader(
            "Building photo (JPG / PNG)",
            type=['jpg', 'jpeg', 'png', 'webp'],
            key=_uploader_key('prop_hero_photo_upload'),
            help="Saved to GitHub as data/{property_code}/hero.jpg — updates hero banner after ~2 min redeploy.",
        )
        if _hero_upload is not None:
            if not _photo_target_code:
                st.warning("Enter the GA Property ID above and save the config first, then re-upload the photo.")
            else:
                from property_writer import save_image_to_github as _save_img_gh, save_image_local as _save_img_loc
                _img_bytes = _hero_upload.read()
                _img_ext   = _hero_upload.name.rsplit('.', 1)[-1].lower()
                _img_fname = f'hero.{_img_ext}'
                _loc_ok, _loc_msg = _save_img_loc(
                    _photo_target_code, _img_bytes, _img_fname, str(_DATA_DIR))
                if github_configured():
                    _gh_ok, _gh_msg = _save_img_gh(_photo_target_code, _img_bytes, _img_fname)
                    if _gh_ok:
                        st.success(f"✅ Photo saved to GitHub — hero banner updates after redeploy (~2 min).")
                    else:
                        st.warning(f"GitHub save failed: {_gh_msg}. Saved locally.")
                else:
                    st.info("Photo saved locally. Set up GitHub secrets to persist to Streamlit Cloud.")
                _bump_uploader('prop_hero_photo_upload')
                st.rerun()
    with _photo_col2:
        # Preview current photo for this property
        if not _is_new and _photo_target_code:
            _prev_src = _prop_hero_src(_photo_target_code)
            if _prev_src:
                st.image(_prev_src, caption="Current photo", use_container_width=True)
            else:
                st.caption("No photo yet")

    # ── Chart of Accounts (outside form) ──────────────────────────────────────
    st.markdown("### 📊 Chart of Accounts")
    st.caption(
        "Powers QC Check 8 — flags any GL account code that isn't on the chart "
        "on file (e.g. a new account Yardi added since the last COA export)."
    )
    _uses_grp_coa = st.radio(
        "Uses the standard GRP Yardi Chart of Accounts?",
        options=["Yes", "No"],
        index=0 if _ef('uses_grp_coa', False) or _is_new else 1,
        horizontal=True,
        key="prop_uses_grp_coa_radio",
        help="Yes: no upload needed — uses data/_shared/GRP_Chart_of_Accounts.xlsx, "
             "shared across every GRP Yardi property. No: upload this property's own "
             "COA below (e.g. a partner running their own Yardi with different codes).",
    ) == "Yes"

    if _uses_grp_coa:
        st.success(
            "✅ Using the shared **GRP Yardi Chart of Accounts** — no upload needed. "
            "Update it once (below) and every GRP-COA property picks it up.",
            icon="📊",
        )
        with st.expander("Update the shared GRP Chart of Accounts", expanded=False):
            st.caption(
                "Replaces `data/_shared/GRP_Chart_of_Accounts.xlsx` — affects every "
                "property with 'Uses the standard GRP Yardi Chart of Accounts?' = Yes."
            )
            _grp_coa_upload = st.file_uploader(
                "GRP Chart of Accounts (Excel)",
                type=['xlsx', 'xls'],
                key=_uploader_key('prop_shared_coa_upload'),
            )
            if _grp_coa_upload is not None:
                from property_writer import save_image_to_github as _save_shared_gh, save_image_local as _save_shared_loc
                _shared_bytes = _grp_coa_upload.read()
                _shared_loc_ok, _shared_loc_msg = _save_shared_loc(
                    '_shared', _shared_bytes, 'GRP_Chart_of_Accounts.xlsx', str(_DATA_DIR))
                if github_configured():
                    _shared_gh_ok, _shared_gh_msg = _save_shared_gh(
                        '_shared', _shared_bytes, 'GRP_Chart_of_Accounts.xlsx')
                    if _shared_gh_ok:
                        st.success("✅ Shared GRP Chart of Accounts saved to GitHub.")
                    else:
                        st.warning(f"GitHub save failed: {_shared_gh_msg}. Saved locally.")
                else:
                    st.info("Shared GRP Chart of Accounts saved locally.")
                _bump_uploader('prop_shared_coa_upload')
                st.rerun()
        st.markdown("---")

    if not _uses_grp_coa:
        st.caption(
            "Upload this property's own Chart of Accounts (Excel or CSV) — e.g. a "
            "partner running their own Yardi with different account codes. Stored as "
            "`data/{property_code}/chart_of_accounts.xlsx`."
        )
        _coa_col1, _coa_col2 = st.columns([2, 1])
        with _coa_col1:
            _coa_upload = st.file_uploader(
                "Chart of Accounts (Excel / CSV)",
                type=['xlsx', 'xls', 'csv'],
                key=_uploader_key('prop_coa_upload'),
                help="Saved to GitHub as data/{property_code}/chart_of_accounts.xlsx",
            )
            if _coa_upload is not None:
                _coa_target = _photo_target_code  # same code as photo (empty for new properties)
                if not _coa_target:
                    st.warning(
                        "Enter the GA Property ID and save the config first, "
                        "then re-upload the Chart of Accounts."
                    )
                else:
                    from property_writer import save_image_to_github as _save_coa_gh, save_image_local as _save_coa_loc
                    _coa_bytes = _coa_upload.read()
                    _coa_ext   = _coa_upload.name.rsplit('.', 1)[-1].lower()
                    _coa_fname = f'chart_of_accounts.{_coa_ext}'
                    _cloc_ok, _cloc_msg = _save_coa_loc(_coa_target, _coa_bytes, _coa_fname, str(_DATA_DIR))
                    if github_configured():
                        _cgh_ok, _cgh_msg = _save_coa_gh(_coa_target, _coa_bytes, _coa_fname)
                        if _cgh_ok:
                            st.success("✅ Chart of Accounts saved to GitHub.")
                        else:
                            st.warning(f"GitHub save failed: {_cgh_msg}. Saved locally.")
                    else:
                        st.info("Chart of Accounts saved locally.")
                    # Show a quick preview if it's an xlsx/csv
                    try:
                        import io as _coa_io
                        import pandas as _coa_pd
                        _coa_bytes_copy = bytes(_coa_bytes)
                        if _coa_ext in ('xlsx', 'xls'):
                            _coa_preview_df = _coa_pd.read_excel(_coa_io.BytesIO(_coa_bytes_copy), nrows=20)
                        else:
                            _coa_preview_df = _coa_pd.read_csv(_coa_io.BytesIO(_coa_bytes_copy), nrows=20)
                        st.dataframe(_coa_preview_df, use_container_width=True, height=250)
                        st.caption(f"Showing first 20 rows of {_coa_upload.name}")
                    except Exception:
                        st.caption(f"Saved: {_coa_upload.name}")
                    _bump_uploader('prop_coa_upload')
                    st.rerun()
        with _coa_col2:
            # Show whether a COA is already on file for this property
            if not _is_new and _photo_target_code:
                _coa_on_disk = False
                for _ext in ('.xlsx', '.xls', '.csv'):
                    if (_DATA_DIR / _photo_target_code / f'chart_of_accounts{_ext}').exists():
                        _coa_on_disk = True
                        st.success(f"✅ COA on file: `chart_of_accounts{_ext}`", icon="📊")
                        break
                if not _coa_on_disk:
                    st.caption("No COA on file yet")

    # ── Current Year Budget (outside form) ────────────────────────────────────
    st.markdown("### 💰 Current Year Budget (Kardin)")
    st.caption(
        "Upload this property's current-year Kardin annual budget. Drives QC "
        "tie-out and the budget-based accrual detection (HVAC, Fire Life "
        "Safety, Snow & Ice). Saved under its own filename — enter that exact "
        "filename in 'Kardin Budget Filename' further down (step 8) once saved."
    )
    _budget_col1, _budget_col2 = st.columns([2, 1])
    with _budget_col1:
        _budget_upload = st.file_uploader(
            "Current Year Budget (Excel)",
            type=['xlsx', 'xls'],
            key=_uploader_key('prop_budget_upload'),
            help="Saved to GitHub as data/{property_code}/<filename you uploaded>",
        )
        if _budget_upload is not None:
            _budget_target = _photo_target_code  # same code as photo (empty for new properties)
            if not _budget_target:
                st.warning(
                    "Enter the GA Property ID and save the config first, "
                    "then re-upload the budget."
                )
            else:
                from property_writer import save_image_to_github as _save_budget_gh, save_image_local as _save_budget_loc
                _budget_bytes = _budget_upload.read()
                _budget_fname = _budget_upload.name
                _bloc_ok, _bloc_msg = _save_budget_loc(_budget_target, _budget_bytes, _budget_fname, str(_DATA_DIR))
                if github_configured():
                    _bgh_ok, _bgh_msg = _save_budget_gh(_budget_target, _budget_bytes, _budget_fname)
                    if _bgh_ok:
                        st.success(f"✅ Budget saved to GitHub as `{_budget_fname}`.")
                    else:
                        st.warning(f"GitHub save failed: {_bgh_msg}. Saved locally.")
                else:
                    st.info(f"Budget saved locally as `{_budget_fname}`.")
                st.caption(
                    f"⬇️ Enter **{_budget_fname}** in 'Kardin Budget Filename' (step 8) "
                    f"below so the pipeline knows to load it automatically."
                )
                _bump_uploader('prop_budget_upload')
                st.rerun()
    with _budget_col2:
        # Show whether a budget file matching the configured filename is on disk
        if not _is_new and _photo_target_code:
            _cur_budget_fname = _ef('kardin_budget_file', 'GA_Kardin_Budget_FY2026.xlsx')
            if (_DATA_DIR / _photo_target_code / _cur_budget_fname).exists():
                st.success(f"✅ On file: `{_cur_budget_fname}`", icon="💰")
            else:
                st.caption(f"No file named `{_cur_budget_fname}` on disk yet")

    # ── 12-Month GL History — onboarding review only (outside form) ───────────
    st.markdown("### 📜 12-Month GL History (Onboarding Review)")
    st.caption(
        "Upload a full year of GL export to see which expense accounts/vendors "
        "bill on a recurring-but-not-monthly cadence (quarterly, semi-annual, "
        "annual) — useful context before your first close. **Informational "
        "only** — nothing here auto-fills any config or accrual table; the One-"
        "Off Accruals table always starts blank, on purpose. Not saved anywhere "
        "— this is a one-time onboarding look, not a monthly upload."
    )
    st.caption(
        "⚠️ Cadence classification is a best guess pending a real 12-month "
        "sample to verify against — treat it as a starting point for review, "
        "not a final answer."
    )
    _gl_hist_upload = st.file_uploader(
        "12-Month GL Export (Excel)",
        type=['xlsx'],
        key='prop_gl_history_upload',
    )
    if _gl_hist_upload is not None:
        try:
            import tempfile as _gh_tempfile
            with _gh_tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as _gh_tmp:
                _gh_tmp.write(_gl_hist_upload.read())
                _gh_tmp_path = _gh_tmp.name
            from parsers.yardi_gl import parse_gl as _parse_gl_hist
            from gl_history_analyzer import analyze_recurring_vendors as _analyze_vendors
            _gl_hist_result = _parse_gl_hist(_gh_tmp_path)
            _vendor_patterns = _analyze_vendors(_gl_hist_result)
            os.remove(_gh_tmp_path)

            if not _vendor_patterns:
                st.info("No expense-account activity found to analyze in this file.")
            else:
                _cadence_order = {'Quarterly': 0, 'Semi-Annual': 1, 'Annual/One-time': 2,
                                   'Irregular': 3, 'Monthly': 4}
                _vendor_patterns.sort(key=lambda p: (_cadence_order.get(p.cadence, 9), p.account_code))
                _vp_df = pd.DataFrame([{
                    'Account Code': p.account_code,
                    'Account Name': p.account_name,
                    'Vendor':       p.vendor,
                    'Cadence':      p.cadence,
                    'Occurrences':  p.occurrences,
                    'Avg Amount':   p.avg_amount,
                    'Months Seen':  ', '.join(p.months_seen),
                } for p in _vendor_patterns])
                _non_monthly = _vp_df[_vp_df['Cadence'] != 'Monthly']
                st.caption(
                    f"{len(_vendor_patterns)} account/vendor pattern(s) found — "
                    f"{len(_non_monthly)} non-monthly (the ones worth a second look)."
                )
                st.dataframe(
                    _vp_df,
                    use_container_width=True,
                    height=min(400, 40 + 35 * len(_vp_df)),
                    column_config={'Avg Amount': st.column_config.NumberColumn(format="$%,.2f")},
                )
        except Exception as _gh_exc:
            st.warning(f"Could not analyze this GL export: {_gh_exc}")

    # ── Bank Statement auto-extract (outside form) ─────────────────────────────
    st.markdown("### 🏦 Bank Statement (Auto-Extract Account)")
    st.caption(
        "Upload a real statement instead of typing the account number blind — "
        "extracts it automatically so you can confirm and copy it into the "
        "Bank Accounts table (step 6) below. Upload one statement per account "
        "(operating, development, DACA) — each appears as its own row here. "
        "Only recognizes **PNC, Bank of America, KeyBank, and Eastern Bank** — a "
        "different bank needs a new parser built first, same as a new lender "
        "does for loan statements."
    )
    if 'prop_bank_detect_rows' not in st.session_state:
        st.session_state.prop_bank_detect_rows = []
    _bank_stmt_upload = st.file_uploader(
        "Bank Statement (PDF)",
        type=['pdf'],
        key=_uploader_key('prop_bank_stmt_upload'),
    )
    if _bank_stmt_upload is not None:
        try:
            import tempfile as _bs_tempfile
            with _bs_tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as _bs_tmp:
                _bs_tmp.write(_bank_stmt_upload.read())
                _bs_tmp_path = _bs_tmp.name
            from bank_statement_detector import detect_and_extract as _detect_bank
            _bs_result = _detect_bank(_bs_tmp_path)
            os.remove(_bs_tmp_path)

            if not _bs_result.recognized:
                st.warning(f"⚠️ {_bs_result._parse_error}")
            elif _bs_result._parse_error:
                st.warning(f"⚠️ {_bs_result._parse_error}")
            else:
                st.success(
                    f"✅ Detected **{_bs_result.bank_label}** — account number "
                    f"`{_bs_result.account_number or '(not found)'}`",
                    icon="🏦",
                )
                _already = any(
                    r['Full Account'] == _bs_result.account_number
                    for r in st.session_state.prop_bank_detect_rows
                )
                if not _already:
                    st.session_state.prop_bank_detect_rows.append({
                        'Slug': _bs_result.suggested_slug,
                        'Bank Name': _bs_result.bank_label,
                        'Full Account': _bs_result.account_number or '',
                    })
        except Exception as _bs_exc:
            st.warning(f"Could not read this statement: {_bs_exc}")
        _bump_uploader('prop_bank_stmt_upload')
        st.rerun()

    if st.session_state.prop_bank_detect_rows:
        st.caption("⬇️ Detected so far — copy into the Bank Accounts table (step 6) below:")
        st.dataframe(pd.DataFrame(st.session_state.prop_bank_detect_rows), use_container_width=True)
        if st.button("Clear detected accounts", key="clear_bank_detect"):
            st.session_state.prop_bank_detect_rows = []
            st.rerun()

    st.markdown("---")

    with st.form("property_setup_form", clear_on_submit=False):
        if _is_mri:
            st.info(
                "**MRI property** — fill in the basic information below and save. "
                "MRI-specific configuration fields (COA mapping, account codes, export formats) "
                "will be available in a future update.",
                icon="🔜",
            )
        st.markdown("### 1 · Basic Information")
        _c1, _c2 = st.columns(2)
        _prop_code    = _c1.text_input("GA Property ID *",
                                        value='' if _is_new else _edit_cfg.property_code,
                                        placeholder="e.g. lexlabspm" if not _is_mri else "e.g. metropark01",
                                        help="This pipeline's OWN internal identifier — not a Yardi property "
                                             "code. It's the storage key for everything in this app (folder "
                                             "name, property selector). For a consolidated property, each "
                                             "building's real Yardi code (e.g. 25hart, 40hart) goes in the "
                                             "Buildings table below instead — not here. Lowercase, no spaces.")
        _display_name = _c2.text_input("Display Name *",
                                        value=_ef('property_display_name'),
                                        placeholder="e.g. Lex Labs")
        _prop_name    = st.text_input("Full Legal Entity Name",
                                       value=_ef('property_name'),
                                       placeholder="e.g. Lex Labs Owner, LLC")
        _address      = st.text_input("Property Address",
                                       value=_ef('property_address'),
                                       placeholder="e.g. 100 Main Street, Boston, MA 02101")
        _c3, _c4 = st.columns(2)
        _prop_type = _c3.text_input("Property Type",
                                     value=_ef('property_type'),
                                     placeholder="e.g. Life Science, Office, Industrial")
        _size_sf   = _c4.number_input("Size (SF)", min_value=0,
                                       value=int(_ef('property_size_sf') or 0), step=1000,
                                       disabled=bool(_ef('consolidated_buildings', [])),
                                       help="Auto-computed as the sum of building rows below "
                                            "once any are entered.")

        st.markdown("#### Buildings")
        st.caption(
            "Leave empty for a single-Yardi-code property (like Revolution Labs). "
            "For a property consolidated from multiple Yardi property codes — one row "
            "per building/entity. Size (SF) here rolls up into the Size field above."
        )
        _default_buildings = [
            {'Building Name': b.name, 'Yardi Property Code': b.yardi_code,
             'Size (SF)': int(b.size_sf or 0)}
            for b in (_ef('consolidated_buildings', []) or [])
        ]
        _buildings_df = pd.DataFrame(
            _default_buildings or
            [{'Building Name': '', 'Yardi Property Code': '', 'Size (SF)': 0}]
        )
        _buildings_edited = st.data_editor(
            _buildings_df,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                'Building Name':        st.column_config.TextColumn("Building Name", width="medium",
                                             help="e.g. '25 Hartwell' or '40 Hartwell'"),
                'Yardi Property Code':  st.column_config.TextColumn("Yardi Property Code", width="small",
                                             help="This building's own Yardi property code."),
                'Size (SF)':            st.column_config.NumberColumn("Size (SF)", min_value=0, step=1000),
            },
            key="prop_buildings_editor",
        )
        _buildings_list = [
            {
                'name':       str(r.get('Building Name', '') or '').strip(),
                'yardi_code': str(r.get('Yardi Property Code', '') or '').strip(),
                'size_sf':    int(r.get('Size (SF)', 0) or 0),
            }
            for _, r in _buildings_edited.iterrows()
            if str(r.get('Building Name', '') or '').strip()
        ]
        if _buildings_list:
            _missing_codes = [b['name'] for b in _buildings_list if not b['yardi_code']]
            if _missing_codes:
                st.warning(
                    f"⚠️ Missing Yardi Property Code for: {', '.join(_missing_codes)}. "
                    f"Every building needs its own Yardi code to be consolidated correctly.",
                    icon="⚠️",
                )
            else:
                _total_sf = sum(b['size_sf'] for b in _buildings_list)
                st.success(
                    f"✅ {len(_buildings_list)} building(s), {', '.join(b['yardi_code'] for b in _buildings_list)} "
                    f"— {_total_sf:,} SF total.",
                    icon="✅",
                )

        _yardi_subset_code = st.text_input(
            "Yardi Report Subset Code",
            value=_ef('yardi_subset_code'),
            placeholder="e.g. .2540hart",
            help="Yardi's own pre-built subset code for a report that already combines "
                 "this property's buildings (e.g. '.2540hart' for 25 & 40 Hartwell) — use "
                 "it directly for any report Yardi can export at the subset level. Leave "
                 "blank for a single-entity property or if no such subset report exists.",
        )

        st.markdown("### 2 · Ownership")
        st.caption("Management Company is always Greatland Realty Partners — not asked here.")
        _c5, _c6 = st.columns(2)
        _investor_legal = _c5.text_input("Investor Legal Entity Name (per W-9)",
                                      value=_ef('investor_legal_name'),
                                      placeholder="e.g. Singerman Real Estate Fund III, LLC")
        _investor   = _c6.text_input("Investor Short Name",
                                      value=_ef('investor_name'),
                                      placeholder="e.g. Singerman Real Estate",
                                      help="Used in the dashboard header and variance commentary tone.")
        # Invoice prefix is no longer manually entered — it defaults to the
        # property code itself (already ends in 'pm' by convention), with
        # month+year appended by mgmt_fee_invoice.py at invoice time (e.g.
        # 'hartwellpm' -> invoice # 'hartwellpm042026' for Apr 2026).
        # An existing property's already-saved prefix (e.g. RevLabs' historical
        # 'RevLabsPM') is preserved as-is — this only supplies a default for a
        # property that's never had one, so past invoice numbering never shifts
        # underneath an already-operating property.
        _inv_prefix = _ef('invoice_prefix') or (_prop_code or '').strip().lower().replace(' ', '')

        st.markdown("### 3 · Team Members")
        st.caption(
            "Everyone who works on this property's monthly close. "
            "These names appear in the Dashboard selector, close tracker, and sign-off sheet. "
            "One name per line."
        )
        _default_members = '\n'.join(
            _edit_cfg.team_members if (_edit_cfg and _edit_cfg.team_members)
            else []
        )
        _team_text = st.text_area(
            "Team members (one per line)",
            value=_default_members,
            height=120,
            placeholder="Jane Smith (Property Accountant)\nJohn Doe (Property Manager)\nAlex Lee (Accounting Manager/Controller)",
            label_visibility="collapsed",
        )

        # Tenants (Utility Billing) — no longer configured here. The tenant
        # list for TUB is now read live from each period's uploaded Tenancy
        # Schedule (rent roll) — see _build_tub_tenants() — so it can't go
        # stale as leases turn over, and the tenant key is Yardi's own
        # tenant code, not a slug someone has to type. Nothing to set up
        # per property; a property's already-saved tenants list (if any,
        # e.g. Revolution Labs pre-dating this change) is preserved as a
        # fallback and just isn't editable from this form anymore.

        # Default One-Off Accruals — removed. Pre-populating rows/amounts here
        # was overkill: most real accruals are already picked up by the 4-layer
        # detection engine, and a genuine one-off accrual is something the PM
        # or accounting team decides and types fresh each month — not something
        # the config should be guessing at in advance. Confirmed with Ryan
        # 2026-08-23. The One-Off Accruals table in Pass 1 now always starts
        # blank; account-name auto-fill (when someone types a code) still
        # works from the GL/Budget Comparison, just not from this config.

        st.markdown("### 4 · Building / Allocation Splits (Multi-Building Properties)")
        st.caption(
            "Leave empty for single-building properties. "
            "For multi-building properties, add one row per building per schedule. "
            "**Schedule Name** groups rows into separate allocation pools — "
            "each schedule must total 100% independently. "
            "Examples: *'2-Bldg'* for a two-way split, *'4-Bldg'* for a four-way split. "
            "Set Yardi Code only if each building has its own separate Yardi property code."
        )
        _default_splits = [
            {
                'Schedule Name':  bs.schedule,
                'Building Name':  bs.name,
                'Yardi Code':     bs.yardi_code,
                'Share %':        round(bs.share_pct * 100, 4),
                'Notes':          bs.notes,
            }
            for bs in (_edit_cfg.building_splits if _edit_cfg else [])
        ]
        # No splits configured yet, but the Buildings list (step 1) already
        # has 2+ real entries — auto-suggest an Equal-share schedule and an
        # SF-proportional schedule from that data instead of a blank row the
        # user has to build (and do the SF percentage math for) by hand.
        # Generalizes to any number of buildings, not just a pair —
        # correct for Equal at 2 (50/50) and computed the same way for 3-5.
        # Remainder from rounding is assigned to the last row so each
        # schedule sums to exactly 100.00%, not 99.99/100.01.
        if not _default_splits and len(_buildings_list) >= 2:
            _n_bldgs = len(_buildings_list)
            _equal_pct = round(100.0 / _n_bldgs, 4)
            for _i, _b in enumerate(_buildings_list):
                _pct = _equal_pct if _i < _n_bldgs - 1 else round(100.0 - _equal_pct * (_n_bldgs - 1), 4)
                _default_splits.append({
                    'Schedule Name': 'Equal', 'Building Name': _b['name'],
                    'Yardi Code': _b['yardi_code'], 'Share %': _pct, 'Notes': '',
                })
            _total_sf = sum(_b['size_sf'] for _b in _buildings_list)
            if _total_sf > 0:
                _running = 0.0
                for _i, _b in enumerate(_buildings_list):
                    if _i < _n_bldgs - 1:
                        _pct = round(_b['size_sf'] / _total_sf * 100, 4)
                        _running += _pct
                    else:
                        _pct = round(100.0 - _running, 4)
                    _default_splits.append({
                        'Schedule Name': 'By-SF', 'Building Name': _b['name'],
                        'Yardi Code': _b['yardi_code'], 'Share %': _pct,
                        'Notes': f"{_b['size_sf']:,} SF",
                    })

        # Plain widgets (text_input/number_input) per row, not st.data_editor.
        # Confirmed with Ryan 2026-08-24: a typed custom Share % (e.g. 67/33
        # overriding the SF-prorated default) was silently lost on save,
        # reverting to the auto-computed By-SF percentages every time.
        # st.data_editor already has two other confirmed failure modes in
        # this app (see Development Notes in CLAUDE.md) — same root-cause
        # class, fixed the same way elsewhere (One-Off Accruals, Intercompany
        # Recode): independently-keyed plain widgets per row instead of a
        # canvas grid, which never re-merges/discards edits against a
        # freshly-computed default on rerun. Since this table lives inside
        # st.form (which disallows a plain st.button() for add/remove), row
        # count is a number_input instead of an Add/Remove row button.
        _SPLITS_NROWS_KEY = "splits_nrows"

        def _splits_seed_widget(_i: int, _seed: dict) -> None:
            st.session_state[f"splits_sched_{_i}"] = _seed['Schedule Name']
            st.session_state[f"splits_name_{_i}"]  = _seed['Building Name']
            st.session_state[f"splits_code_{_i}"]  = _seed['Yardi Code']
            st.session_state[f"splits_pct_{_i}"]   = _seed['Share %']
            st.session_state[f"splits_notes_{_i}"] = _seed['Notes']

        # (Re)seed row widgets only when this property hasn't been seeded
        # yet in this session (switch, first load, Reset All) — never on an
        # ordinary rerun, so this block's own widgets below never clobber
        # a user's in-progress edit.
        if st.session_state.get("_splits_seeded_for") != _edit_code:
            _seed_rows = _default_splits or [
                {'Schedule Name': '', 'Building Name': '', 'Yardi Code': '', 'Share %': 0.0, 'Notes': ''}
            ]
            for _i, _seed in enumerate(_seed_rows):
                _splits_seed_widget(_i, _seed)
            st.session_state[_SPLITS_NROWS_KEY] = len(_seed_rows)
            st.session_state["_splits_seeded_for"] = _edit_code

        _splits_nrows = st.number_input(
            "Number of split rows", min_value=1, max_value=50,
            key=_SPLITS_NROWS_KEY, step=1,
            help="Increase to add more rows — e.g. 2 schedules × 2 buildings = 4 rows; "
                 "add more for a 3-5 building property or an extra schedule.",
        )
        _sh0, _sh1, _sh2, _sh3, _sh4 = st.columns([1.3, 2.0, 1.3, 1.0, 2.0])
        _sh0.markdown("**Schedule Name**")
        _sh1.markdown("**Building Name**")
        _sh2.markdown("**Yardi Code**")
        _sh3.markdown("**Share %**")
        _sh4.markdown("**Notes**")
        for _i in range(int(_splits_nrows)):
            _sr0, _sr1, _sr2, _sr3, _sr4 = st.columns([1.3, 2.0, 1.3, 1.0, 2.0])
            _sr0.text_input("Schedule Name", key=f"splits_sched_{_i}", label_visibility="collapsed")
            _sr1.text_input("Building Name", key=f"splits_name_{_i}", label_visibility="collapsed")
            _sr2.text_input("Yardi Code", key=f"splits_code_{_i}", label_visibility="collapsed")
            _sr3.number_input("Share %", key=f"splits_pct_{_i}", label_visibility="collapsed",
                               min_value=0.0, max_value=100.0, step=0.01, format="%.2f")
            _sr4.text_input("Notes", key=f"splits_notes_{_i}", label_visibility="collapsed")

        def _splits_current_rows() -> list:
            _rows = []
            for _i in range(int(st.session_state.get(_SPLITS_NROWS_KEY, 0))):
                _bname = str(st.session_state.get(f"splits_name_{_i}", '') or '').strip()
                if not _bname:
                    continue
                _rows.append({
                    'Schedule Name': str(st.session_state.get(f"splits_sched_{_i}", '') or '').strip(),
                    'Building Name': _bname,
                    'Yardi Code':    str(st.session_state.get(f"splits_code_{_i}", '') or '').strip(),
                    'Share %':       float(st.session_state.get(f"splits_pct_{_i}", 0.0) or 0.0),
                    'Notes':         str(st.session_state.get(f"splits_notes_{_i}", '') or '').strip(),
                })
            return _rows

        _splits_edited_rows = _splits_current_rows()

        # Live validation — check each schedule group sums to 100%
        _split_rows_filled = _splits_edited_rows
        if _split_rows_filled:
            # Group by schedule name
            _sched_totals: dict = {}
            _sched_counts: dict = {}
            for _sr in _split_rows_filled:
                _sn = str(_sr.get('Schedule Name', '') or '').strip() or 'default'
                _sched_totals[_sn] = _sched_totals.get(_sn, 0.0) + float(_sr.get('Share %', 0) or 0)
                _sched_counts[_sn] = _sched_counts.get(_sn, 0) + 1
            _all_ok = all(abs(v - 100.0) <= 0.01 for v in _sched_totals.values())
            if _all_ok:
                _scheds_summary = ', '.join(
                    f"**{k}** ({_sched_counts[k]} bldgs)" for k in _sched_totals
                )
                st.success(f"✅ All schedules total 100% — {_scheds_summary}.", icon="✅")
            else:
                for _sn, _total in _sched_totals.items():
                    if abs(_total - 100.0) > 0.01:
                        st.warning(
                            f"⚠️ Schedule **'{_sn}'** totals **{_total:.2f}%** — must be 100%.",
                            icon="⚠️",
                        )

        # Default split schedule selector (only shown when splits are defined)
        _avail_schedules = list(dict.fromkeys(
            str(r.get('Schedule Name', '') or '').strip()
            for r in _splits_edited_rows
            if str(r.get('Schedule Name', '') or '').strip()
        ))
        _cur_default_sch = _ef('default_split_schedule', '')
        if _avail_schedules:
            _default_sch_options = ['(none — no automatic splitting)'] + _avail_schedules
            _default_sch_idx = (
                _avail_schedules.index(_cur_default_sch) + 1
                if _cur_default_sch in _avail_schedules else 0
            )
            _default_sch_sel = st.selectbox(
                "Default Split Schedule",
                _default_sch_options,
                index=_default_sch_idx,
                help="Applied automatically to all auto-detected accruals (Nexus, historical, mgmt fee, etc.). "
                     "Per-line overrides in the one-off accruals table take precedence.",
                key="prop_default_split_sch",
            )
            _default_split_schedule = (
                '' if _default_sch_sel.startswith('(none') else _default_sch_sel
            )
        else:
            _default_split_schedule = ''

        st.markdown("### 5 · Management Fee Lines")
        st.caption("One row per PM agreement line. Leave Name blank to skip a row.")
        _default_fees = [
            {'Name': fl.name, 'Rate (decimal)': fl.rate, 'Minimum ($)': fl.minimum,
             'DR Account': fl.dr_account, 'CR Account': fl.cr_account, 'Ref Prefix': fl.ref_prefix}
            for fl in (_edit_cfg.management_fees if _edit_cfg else [])
        ] or [
            {'Name': 'PM', 'Rate (decimal)': 0.03, 'Minimum ($)': 0.0,
             'DR Account': '637130', 'CR Account': '213100', 'Ref Prefix': 'MGMT-FEE-PM'},
        ]
        _fees_df = pd.DataFrame(_default_fees)
        _fees_edited = st.data_editor(
            _fees_df,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                'Name':           st.column_config.TextColumn("PM Name", width="small"),
                'Rate (decimal)': st.column_config.NumberColumn("Rate", format="%.4f", min_value=0.0, max_value=1.0),
                'Minimum ($)':    st.column_config.NumberColumn("Min ($)", format="$%.0f", min_value=0.0),
                'DR Account':     st.column_config.TextColumn("DR Acct", width="small"),
                'CR Account':     st.column_config.TextColumn("CR Acct", width="small"),
                'Ref Prefix':     st.column_config.TextColumn("Ref Prefix"),
            },
            key="prop_fees_editor",
        )

        st.markdown("### 6 · Bank Accounts")
        st.caption(
            "One row per bank account. **Slug** = unique key (lowercase, underscores). "
            "**Bank Name** = text that appears in PDF statements. "
            "Account type detected from slug: contains `operat` → operating, `dev` → development, `daca` → DACA."
        )
        _default_banks = [
            {'Slug': slug,
             'Label': ba.label, 'Bank Name': ba.bank_name,
             'Last 4': ba.last4, 'Full Account': ba.full_account, 'GL Account': ba.gl_account}
            for slug, ba in (_edit_cfg.bank_accounts.items() if _edit_cfg else {}.items())
        ] or [
            {'Slug': 'pnc_operating',    'Label': 'PNC Operating',    'Bank Name': 'PNC',           'Last 4': '', 'Full Account': '', 'GL Account': '111100'},
            {'Slug': 'bofa_development', 'Label': 'BofA Development', 'Bank Name': 'Bank of America','Last 4': '', 'Full Account': '', 'GL Account': '111210'},
            {'Slug': 'keybank_daca',     'Label': 'KeyBank DACA',     'Bank Name': 'KeyBank',        'Last 4': '', 'Full Account': '', 'GL Account': '115100'},
        ]
        _banks_edited = st.data_editor(
            pd.DataFrame(_default_banks),
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                'Slug':         st.column_config.TextColumn("Slug *", width="medium"),
                'Label':        st.column_config.TextColumn("Label"),
                'Bank Name':    st.column_config.TextColumn("Bank Name (PDF match)"),
                'Last 4':       st.column_config.TextColumn("Last 4", width="small"),
                'Full Account': st.column_config.TextColumn("Full Account #"),
                'GL Account':   st.column_config.TextColumn("GL Acct", width="small"),
            },
            key="prop_banks_editor",
        )

        st.markdown("### 7 · Payment Instructions (Invoice PDF)")
        _ca, _cb = st.columns(2)
        with _ca:
            st.markdown("**ACH / Wire**")
            _ach = _ef('payment_ach') or {}
            _ach_acct_name  = st.text_input("Account Name",  value=_ach.get('account_name', ''), key="ach_acct_name")
            _ach_bank       = st.text_input("Bank Name",      value=_ach.get('bank_name', ''),    key="ach_bank")
            _ach_acct_num   = st.text_input("Account Number", value=_ach.get('account_number', ''), key="ach_acct_num")
            _ach_routing    = st.text_input("Routing (ABA)",  value=_ach.get('routing_number', ''), key="ach_routing")
            _ach_addr       = st.text_input("Bank Address",   value=_ach.get('bank_address', ''), key="ach_addr")
        with _cb:
            st.markdown("**Check**")
            _chk = _ef('payment_check') or {}
            _chk_payable    = st.text_input("Payable To",      value=_chk.get('payable_to', ''),    key="chk_payable")
            _chk_addr1      = st.text_input("Address Line 1",  value=_chk.get('address_line1', ''), key="chk_addr1")
            _chk_addr2      = st.text_input("Address Line 2",  value=_chk.get('address_line2', ''), key="chk_addr2")
            _chk_attn       = st.text_input("Attention",       value=_chk.get('attention', ''),     key="chk_attn")

        st.markdown("### 8 · RE Tax & Other")
        _c9, _c10 = st.columns(2)
        _retax_months_str = _c9.text_input(
            "RE Tax Payment Months (comma-separated)",
            value=', '.join(str(m) for m in (_edit_cfg.re_tax_payment_months if _edit_cfg else [1, 4, 7, 10])),
            help="Months (1-12) when the quarterly RE tax bill is paid. Typically Jan/Apr/Jul/Oct."
        )
        _parcel_str = _c10.text_input(
            "Parcel IDs (comma-separated, optional)",
            value=', '.join(_edit_cfg.parcel_ids if _edit_cfg else []),
        )
        _c11, _c12 = st.columns(2)
        _kardin_file = _c11.text_input("Kardin Budget Filename",
                                        value=_ef('kardin_budget_file', 'GA_Kardin_Budget_FY2026.xlsx'))
        _file_pfx_del = _c12.text_input("Deliverable File Prefix",
                                          value=_ef('file_prefix_deliverable'),
                                          placeholder="e.g. LexLabs  → LexLabs_Jan2026_Workpapers.xlsx",
                                          help="Leave blank to auto-derive from display name.")
        # Internal File Prefix is always "GA" (GRP's own internal-file branding,
        # not property-specific) — no longer asked here, matching Management
        # Company. Kept as a real config field/param since output filenames
        # (GA_Accruals_JE.csv, etc.) still read it — just not user-editable.
        _file_pfx_int = 'GA'

        st.markdown("---")
        _submitted = st.form_submit_button("💾 Save Property Config", type="primary",
                                            use_container_width=True)

    # ── Handle form submission ────────────────────────────────────────────────
    if _submitted:
        _prop_code = (_prop_code or '').strip().lower().replace(' ', '')
        # Guard against silently overwriting a DIFFERENT property that happens to
        # already use this code — e.g. entering '12&24 Hartwell' but leaving/typing
        # the code from an already-saved '25&40 Hartwell' by mistake. Editing the
        # SAME property without changing its code is unaffected (_prop_code == _edit_code).
        _code_collision = (
            _prop_code in {p['code'] for p in _existing} and _prop_code != _edit_code
        )
        if not _prop_code:
            st.error("GA Property ID is required.")
        elif not _display_name:
            st.error("Display Name is required.")
        elif _code_collision:
            _collision_name = next(
                (p['display_name'] for p in _existing if p['code'] == _prop_code), _prop_code
            )
            st.error(
                f"⚠️ GA Property ID '{_prop_code}' is already used by **{_collision_name}**. "
                f"Saving would overwrite that property's config. Choose a different code."
            )
        else:
            # Parse fee rows
            _fee_list = []
            for _, _frow in _fees_edited.iterrows():
                _fname = str(_frow.get('Name', '') or '').strip()
                if not _fname:
                    continue
                _fee_list.append({
                    'name':       _fname,
                    'rate':       float(_frow.get('Rate (decimal)', 0) or 0),
                    'minimum':    float(_frow.get('Minimum ($)', 0) or 0),
                    'dr_account': str(_frow.get('DR Account', '637130') or '637130'),
                    'cr_account': str(_frow.get('CR Account', '213100') or '213100'),
                    'ref_prefix': str(_frow.get('Ref Prefix', '') or ''),
                })

            # Parse bank account rows
            _bank_list = []
            for _, _brow in _banks_edited.iterrows():
                _bslug = str(_brow.get('Slug', '') or '').strip().lower().replace(' ', '_')
                if not _bslug:
                    continue
                _bank_list.append({
                    'slug':         _bslug,
                    'label':        str(_brow.get('Label', '') or ''),
                    'bank_name':    str(_brow.get('Bank Name', '') or ''),
                    'last4':        str(_brow.get('Last 4', '') or ''),
                    'full_account': str(_brow.get('Full Account', '') or ''),
                    'gl_account':   str(_brow.get('GL Account', '') or ''),
                })

            # Parse RE tax months
            try:
                _retax_months = [int(m.strip()) for m in _retax_months_str.split(',') if m.strip()]
            except Exception:
                _retax_months = [1, 4, 7, 10]

            # Parse parcel IDs
            _parcels = [p.strip() for p in _parcel_str.split(',') if p.strip()]

            # Build payment dicts
            _payment_ach   = {k: v for k, v in {
                'account_name':   _ach_acct_name,
                'bank_name':      _ach_bank,
                'account_number': _ach_acct_num,
                'routing_number': _ach_routing,
                'bank_address':   _ach_addr,
            }.items() if v}
            _payment_check = {k: v for k, v in {
                'payable_to':    _chk_payable,
                'address_line1': _chk_addr1,
                'address_line2': _chk_addr2,
                'attention':     _chk_attn,
            }.items() if v}

            # Build config dict and render YAML
            # Parse team members from text area (one per line)
            _team_members_parsed = [
                m.strip() for m in _team_text.splitlines() if m.strip()
            ]  # empty list is valid; config will have no team_members

            # Tenants — not edited from this form (see the "Tenants (Utility
            # Billing)" removal note above); preserve whatever this property
            # already had saved, if anything, rather than wiping it out just
            # because there's no UI for it here.
            _tenants_list = list(getattr(_edit_cfg, 'tenants', []) or [])

            # Default accruals — removed (see the "Default One-Off Accruals"
            # removal note above); always empty going forward, including on
            # re-save of a property that
            # had rows from before this change.
            _daccruals_list = []

            # Parse building splits
            _splits_list = [
                {
                    'schedule':   _srow['Schedule Name'] or 'default',
                    'name':       _srow['Building Name'],
                    'yardi_code': _srow['Yardi Code'],
                    'share_pct':  _srow['Share %'] / 100.0,
                    'notes':      _srow['Notes'],
                }
                for _srow in _splits_edited_rows
            ]

            _cfg_dict = build_config_dict(
                property_code          = _prop_code,
                property_name          = _prop_name,
                property_display_name  = _display_name,
                property_address       = _address,
                property_type          = _prop_type,
                property_size_sf       = (sum(b['size_sf'] for b in _buildings_list)
                                           if _buildings_list else (int(_size_sf) if _size_sf else None)),
                consolidated_buildings = _buildings_list,
                yardi_subset_code      = _yardi_subset_code,
                investor_name          = _investor,
                investor_legal_name    = _investor_legal,
                management_company     = 'Greatland Realty Partners',
                invoice_prefix         = _inv_prefix,
                team_members           = _team_members_parsed,
                tenants                = _tenants_list,
                default_accruals       = _daccruals_list,
                building_splits        = _splits_list,
                default_split_schedule = _default_split_schedule if '_default_split_schedule' in dir() else '',
                management_fees        = _fee_list,
                gl_accounts            = {},
                bank_accounts          = _bank_list,
                payment_ach            = _payment_ach,
                payment_check          = _payment_check,
                re_tax_payment_months  = _retax_months,
                parcel_ids             = _parcels,
                kardin_budget_file     = _kardin_file,
                fiscal_year_start_month = 1,
                file_prefix_internal   = _file_pfx_int or _pfx_int,
                file_prefix_deliverable = _file_pfx_del,
                active                 = getattr(_edit_cfg, 'active', True) if not _is_new else True,
                property_system        = _prop_system.lower(),
                uses_grp_coa           = _uses_grp_coa,
            )
            _yaml_str = config_to_yaml(_cfg_dict)

            # Save local
            _loc_ok, _loc_msg = save_local(_prop_code, _yaml_str, str(_DATA_DIR))

            # Save to GitHub
            _gh_ok, _gh_msg = False, 'GitHub not configured'
            if github_configured():
                _action = 'Update' if not _is_new else 'Add'
                _gh_ok, _gh_msg = save_to_github(
                    _prop_code, _yaml_str,
                    commit_message=f'{_action} property config: {_prop_code} ({_display_name})'
                )

            # Results
            if _loc_ok:
                st.success(f"✅ Saved locally: `{_loc_msg}`")
            if _gh_ok:
                st.success(f"✅ {_gh_msg}")
            elif github_configured():
                st.error(f"GitHub save failed: {_gh_msg}")

            # Tell the user how to switch to the new property
            if _is_new:
                if _gh_ok:
                    st.info(
                        f"**{_display_name}** will appear in the **sidebar property selector** "
                        f"after Streamlit finishes redeploying (~2 min). "
                        f"Refresh the page and it will be ready to use.",
                        icon="🏢",
                    )
                elif _loc_ok:
                    st.info(
                        f"**{_display_name}** is now available — select it from the "
                        f"**sidebar property selector** at the top of the sidebar. "
                        f"If you don't see it yet, click the refresh button in your browser.",
                        icon="🏢",
                    )
            else:
                st.info(
                    f"**{_display_name}** config updated. "
                    f"{'Reload in ~2 min after Streamlit redeploys.' if _gh_ok else 'Changes are live locally.'}",
                    icon="✅",
                )

            # Always offer download
            st.download_button(
                label="⬇️ Download config.yaml",
                data=_yaml_str.encode('utf-8'),
                file_name=f"{_prop_code}_config.yaml",
                mime="text/yaml",
                help="Upload this file to data/{property_code}/ in GitHub if auto-save failed.",
            )

            # Preview
            with st.expander("📄 Preview generated config.yaml"):
                st.code(_yaml_str, language="yaml")

    # ── Deactivate / Reactivate property (outside form, existing properties only) ──
    from property_writer import deactivate_property as _deactivate, reactivate_property as _reactivate
    if not _is_new and _edit_cfg is not None:
        st.divider()
        _prop_is_active = getattr(_edit_cfg, 'active', True)

        if _prop_is_active:
            st.markdown("#### 🗃️ Deactivate Property")
            st.caption(
                "Hides this property from the selector and pipeline without deleting any data. "
                "All configs, checklists, and workpapers are preserved. "
                "You can reactivate it at any time."
            )
            if "confirm_deactivate" not in st.session_state:
                st.session_state.confirm_deactivate = False

            if not st.session_state.confirm_deactivate:
                if st.button(f"Deactivate {_edit_cfg.display()}", type="secondary"):
                    st.session_state.confirm_deactivate = True
                    st.rerun()
            else:
                st.warning(
                    f"⚠️ This will hide **{_edit_cfg.display()}** from the property selector. "
                    f"No data will be deleted. Reactivate any time from this tab."
                )
                _dc1, _dc2 = st.columns(2)
                if _dc1.button("✅ Yes, Deactivate", use_container_width=True):
                    _dok, _dmsg = _deactivate(_edit_code, str(_DATA_DIR))
                    st.session_state.confirm_deactivate = False
                    if _dok:
                        st.success(f"Property deactivated. {_dmsg}")
                    else:
                        st.error(f"Deactivation failed: {_dmsg}")
                if _dc2.button("❌ Cancel", use_container_width=True):
                    st.session_state.confirm_deactivate = False
                    st.rerun()
        else:
            st.markdown("#### 🔄 Reactivate Property")
            st.caption("This property is currently deactivated. Reactivate to restore it to the selector.")
            if st.button(f"Reactivate {_edit_cfg.display()}", type="primary"):
                _rok, _rmsg = _reactivate(_edit_code, str(_DATA_DIR))
                if _rok:
                    st.success(f"Property reactivated. {_rmsg}")
                else:
                    st.error(f"Reactivation failed: {_rmsg}")

        # ── Permanently delete property (irreversible — for a property entered
        # by mistake, not a real one being wound down; use Deactivate for that) ──
        from property_writer import delete_property as _delete_prop
        st.markdown("#### 🗑️ Permanently Delete Property")
        st.caption(
            "Removes this property's entire config, workpaper template, budget, and "
            "photo — locally and from GitHub. **This cannot be undone.** Use this only "
            "if the property was entered by mistake; use Deactivate above for a real "
            "property that's just no longer active."
        )
        if "confirm_delete_code" not in st.session_state:
            st.session_state.confirm_delete_code = ''

        if not st.session_state.confirm_delete_code:
            if st.button(f"Delete {_edit_cfg.display()}", type="secondary"):
                st.session_state.confirm_delete_code = _edit_code
                st.rerun()
        else:
            st.error(
                f"⚠️ **This permanently deletes {_edit_cfg.display()} (`{_edit_code}`) "
                f"and everything in its data folder — config, template, budget, photo.** "
                f"Type the property code below to confirm."
            )
            _del_confirm_text = st.text_input(
                f"Type `{_edit_code}` to confirm",
                key="delete_confirm_input",
            )
            _dc3, _dc4 = st.columns(2)
            _del_ready = _del_confirm_text.strip() == _edit_code
            if _dc3.button("🗑️ Permanently Delete", use_container_width=True,
                           type="primary", disabled=not _del_ready):
                _delok, _delmsg = _delete_prop(_edit_code, str(_DATA_DIR))
                st.session_state.confirm_delete_code = ''
                if _delok:
                    st.success(f"Property permanently deleted. {_delmsg}")
                    # The "Edit existing or create new" selectbox still holds
                    # the just-deleted code, which no longer exists in its
                    # options list once _existing is rebuilt from disk on
                    # rerun. Can't reset its session_state value directly
                    # here — Streamlit forbids writing to a widget-backed key
                    # in the SAME script run after that widget has already
                    # been drawn (the selectbox renders long before this
                    # button's handler, even though st.rerun() follows right
                    # away). Confirmed as a real crash 2026-08-24. Instead,
                    # set a plain (non-widget) pending flag and apply the
                    # reset at the top of the next run, before the selectbox
                    # is instantiated — see the check right above where that
                    # selectbox is created.
                    st.session_state._prop_select_reset_pending = True
                    st.rerun()
                else:
                    st.error(f"Deletion failed: {_delmsg}")
            if _dc4.button("❌ Cancel", use_container_width=True, key="cancel_delete"):
                st.session_state.confirm_delete_code = ''
                st.rerun()

    # ── Archived properties ───────────────────────────────────────────────────
    from property_config import discover_all_properties as _disc_all
    _all_with_inactive = _disc_all(str(_DATA_DIR))
    _inactive = [p for p in _all_with_inactive if not p.get('active', True)]
    if _inactive:
        with st.expander(f"🗃️ Archived Properties ({len(_inactive)})", expanded=False):
            for _ip in _inactive:
                _ic1, _ic2 = st.columns([4, 2])
                _ic1.markdown(f"**{_ip['display_name']}** &nbsp; `{_ip['code']}`")
                if _ic2.button("Reactivate", key=f"reactivate_{_ip['code']}",
                               use_container_width=True):
                    _rok2, _rmsg2 = _reactivate(_ip['code'], str(_DATA_DIR))
                    if _rok2:
                        st.success(f"Reactivated {_ip['display_name']}. {_rmsg2}")
                    else:
                        st.error(f"Failed: {_rmsg2}")

    # ══════════════════════════════════════════════════════════════════════════
    # PREPAID LEDGER SEED BUILDER
    # ══════════════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("## 📋 Prepaid Ledger Seed Builder")
    st.markdown(
        "Use this to create the **prior-month prepaid ledger seed** for a new or "
        "acquired property. Enter every active prepaid item (service contracts, "
        "software subscriptions, prepaid maintenance, etc.) and download the "
        "correctly formatted `.xlsx` file to upload as the 'prior month ledger' "
        "on the first close.\n\n"
        "> **Note:** Insurance (639110/639120) and RE Tax (641110) are excluded "
        "automatically — those are handled by dedicated amortization functions "
        "driven by `config.yaml → insurance_policies` and the GL."
    )

    _seed_col1, _seed_col2 = st.columns([2, 1])
    with _seed_col1:
        _seed_period = st.text_input(
            "As-Of Period (the period BEFORE your first close)",
            value="Dec-2025",
            help="e.g. 'Dec-2025' if your first pipeline close is January 2026. "
                 "This becomes the 'first_added_period' for all items.",
            key="seed_as_of_period",
        )
    with _seed_col2:
        _seed_scenario = st.radio(
            "Scenario",
            options=["Existing GRP Property", "New Property"],
            index=0,
            key="seed_scenario",
            help="Existing GRP Property: set 'Months Amortized' to reflect months already released. "
                 "New Property: leave at 0.",
        )

    _is_acquisition = _seed_scenario == "Existing GRP Property"

    if _is_acquisition:
        st.info(
            "**Existing GRP Property:** Set *Months Already Amortized* for each item to "
            "reflect how many months have already been released. "
            "The pipeline will pick up from the correct remaining balance.",
            icon="🏢",
        )
    else:
        st.info(
            "**New Property:** Leave *Months Already Amortized* at 0 for all items. "
            "The pipeline will start amortizing from the beginning on the first close.",
            icon="🏗️",
        )

    import pandas as _pd_seed

    _SEED_COLUMNS = [
        "Vendor", "Description", "GL Account #", "GL Account Name",
        "Total Amount", "Monthly Amount", "Service Start", "Service End",
        "Months Amortized", "Invoice #", "Invoice Date",
    ]

    # ── Download a blank template / upload it back filled in ──────────────────
    # For a property with many prepaid items, filling this out in Excel (better
    # copy-paste, no per-row web typing) and uploading it back is easier than
    # typing every row into the table below. The table stays fully editable
    # either way — this just changes how it gets pre-filled.
    _seed_dl_col, _seed_ul_col = st.columns(2)
    with _seed_dl_col:
        import io as _seed_io
        from openpyxl import Workbook as _SeedWorkbook
        _blank_wb = _SeedWorkbook()
        _blank_ws = _blank_wb.active
        _blank_ws.title = 'Prepaid Seed'
        for _ci, _colname in enumerate(_SEED_COLUMNS, start=1):
            _blank_ws.cell(row=1, column=_ci, value=_colname)
        for _ci in range(1, len(_SEED_COLUMNS) + 1):
            _blank_ws.column_dimensions[_blank_ws.cell(row=1, column=_ci).column_letter].width = 18
        _blank_buf = _seed_io.BytesIO()
        _blank_wb.save(_blank_buf)
        st.download_button(
            "⬇️ Download Blank Template",
            data=_blank_buf.getvalue(),
            file_name="GA_Prepaid_Seed_Template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with _seed_ul_col:
        _seed_upload = st.file_uploader(
            "⬆️ Upload Filled-In Template", type=['xlsx'], key='prepaid_seed_upload',
        )

    if "_prepaid_seed_gen" not in st.session_state:
        st.session_state._prepaid_seed_gen = 0
    if "_prepaid_seed_uploaded_rows" not in st.session_state:
        st.session_state._prepaid_seed_uploaded_rows = None

    if _seed_upload is not None:
        try:
            _seed_upload_df = _pd_seed.read_excel(_seed_upload)
            _missing_cols = [c for c in _SEED_COLUMNS if c not in _seed_upload_df.columns]
            if _missing_cols:
                st.warning(
                    f"⚠️ Uploaded file is missing column(s): {', '.join(_missing_cols)}. "
                    f"Use the downloaded template's exact headers — nothing loaded."
                )
            else:
                _uploaded_rows = _seed_upload_df[_SEED_COLUMNS].to_dict('records')
                _uploaded_rows = [r for r in _uploaded_rows if str(r.get('Vendor', '') or '').strip()]
                if _uploaded_rows != st.session_state._prepaid_seed_uploaded_rows:
                    st.session_state._prepaid_seed_uploaded_rows = _uploaded_rows
                    st.session_state._prepaid_seed_gen += 1
                st.success(f"✅ Loaded {len(_uploaded_rows)} row(s) from the uploaded template below.")
        except Exception as _seed_ul_exc:
            st.warning(f"Could not read this file: {_seed_ul_exc}")

    _seed_default_rows = st.session_state._prepaid_seed_uploaded_rows or [
        {"Vendor": "", "Description": "", "GL Account #": "", "GL Account Name": "",
         "Total Amount": 0.0, "Monthly Amount": 0.0,
         "Service Start": None, "Service End": None,
         "Months Amortized": 0, "Invoice #": "", "Invoice Date": None},
    ]
    _seed_df = st.data_editor(
        _pd_seed.DataFrame(_seed_default_rows),
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Vendor":            st.column_config.TextColumn("Vendor *", width="medium"),
            "Description":       st.column_config.TextColumn("Description *", width="large"),
            "GL Account #":      st.column_config.TextColumn("GL Account # *", width="small",
                                    help="6-digit GL code, e.g. '619120'. "
                                         "DO NOT enter 639110/639120/641110 — those are auto-handled."),
            "GL Account Name":   st.column_config.TextColumn("GL Account Name", width="medium"),
            "Total Amount":      st.column_config.NumberColumn("Total Amount ($)", format="$%.2f", min_value=0.0),
            "Monthly Amount":    st.column_config.NumberColumn("Monthly Amount ($)", format="$%.2f", min_value=0.0,
                                    help="Leave 0 to auto-compute: Total ÷ contract months."),
            "Service Start":     st.column_config.DateColumn("Service Start *", format="MM/DD/YYYY"),
            "Service End":       st.column_config.DateColumn("Service End *",   format="MM/DD/YYYY"),
            "Months Amortized":  st.column_config.NumberColumn("Months Amortized",
                                    help="Months already released by prior management. 0 = not yet started.",
                                    min_value=0, step=1, format="%d"),
            "Invoice #":         st.column_config.TextColumn("Invoice #", width="small"),
            "Invoice Date":      st.column_config.DateColumn("Invoice Date", format="MM/DD/YYYY"),
        },
        key=f"prepaid_seed_editor_{st.session_state._prepaid_seed_gen}",
    )

    # Preview calculated values
    _seed_preview_rows = []
    for _, _srow in _seed_df.iterrows():
        _sv = str(_srow.get("Vendor", "") or "").strip()
        _ss = _srow.get("Service Start")
        _se = _srow.get("Service End")
        if not _sv or _ss is None or _se is None:
            continue
        try:
            from dateutil.relativedelta import relativedelta as _rd
            from datetime import date as _date
            if hasattr(_ss, 'date'):
                _ss = _ss.date()
            if hasattr(_se, 'date'):
                _se = _se.date()
            _rd_val = _rd((_se + _rd(days=1)), _ss)
            _tm = max(1, _rd_val.years * 12 + _rd_val.months)
        except Exception:
            _tm = 1
        _ma = max(0, int(_srow.get("Months Amortized", 0) or 0))
        _rm = max(0, _tm - _ma)
        _total = float(_srow.get("Total Amount", 0) or 0)
        _mo_amt = float(_srow.get("Monthly Amount", 0) or 0) or (round(_total / _tm, 2) if _tm else 0)
        _gl = str(_srow.get("GL Account #", "") or "").strip()
        _excluded = _gl in {"639110", "639120", "641110"}
        _seed_preview_rows.append({
            "Vendor": _sv,
            "GL": _gl + (" ⚠️ excluded" if _excluded else ""),
            "Total Mo.": _tm,
            "Amortized": _ma,
            "Remaining": _rm,
            "Monthly $": f"${_mo_amt:,.2f}",
            "Remaining $": f"${_rm * _mo_amt:,.2f}",
        })

    if _seed_preview_rows:
        st.markdown("**Preview** — calculated fields:")
        st.dataframe(_pd_seed.DataFrame(_seed_preview_rows), use_container_width=True, hide_index=True)

    _seed_items_valid = [
        {
            "vendor":           str(r.get("Vendor", "") or "").strip(),
            "description":      str(r.get("Description", "") or "").strip(),
            "gl_account_number":str(r.get("GL Account #", "") or "").strip(),
            "gl_account":       str(r.get("GL Account Name", "") or "").strip(),
            "total_amount":     float(r.get("Total Amount", 0) or 0),
            "monthly_amount":   float(r.get("Monthly Amount", 0) or 0),
            "service_start":    r.get("Service Start"),
            "service_end":      r.get("Service End"),
            "months_amortized": int(r.get("Months Amortized", 0) or 0),
            "invoice_number":   str(r.get("Invoice #", "") or "").strip(),
            "invoice_date":     r.get("Invoice Date"),
        }
        for _, r in _seed_df.iterrows()
        if str(r.get("Vendor", "") or "").strip()
    ]

    if _seed_items_valid and _seed_period:
        try:
            from prepaid_ledger import generate_seed as _gen_seed
            _seed_bytes = _gen_seed(_seed_items_valid, _seed_period)
            _seed_prop_code = (_edit_code if not _is_new else "property") if '_edit_code' in dir() else "property"
            st.download_button(
                label="⬇️ Download Prepaid Ledger Seed",
                data=_seed_bytes,
                file_name=f"GA_Prepaid_Ledger_Seed_{_seed_period.replace('-', '')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                help="Upload this as the 'Prior Month Prepaid Ledger' on the first close.",
            )
            st.caption(
                f"✅ {len(_seed_items_valid)} item(s) included "
                f"(any 639110/639120/641110 accounts were excluded automatically)."
            )
        except Exception as _se_err:
            st.error(f"Seed generation failed: {_se_err}")
    else:
        st.caption("Add at least one item with Vendor, Service Start, and Service End to generate the seed file.")

    # ══════════════════════════════════════════════════════════════════════════
    # FIRST-CLOSE CHECKLIST
    # ══════════════════════════════════════════════════════════════════════════
    st.divider()
    st.markdown("## 📝 First-Close Checklist")

    _chk_scenario = st.radio(
        "Property type",
        options=["Existing GRP Property", "New Property"],
        horizontal=True,
        key="first_close_checklist_type",
    )

    if _chk_scenario == "Existing GRP Property":
        st.markdown("""
### Existing GRP Property — Before First Close
(an acquisition, or a property moving from JLL/manual close onto this pipeline)

**1. Property Config — ⚙️ Properties → Add/Edit, section by section**
- [ ] **Basic Information**: GA Property ID (e.g. `2540hartwellpm`), Display Name, Address, Type, Size
- [ ] If this property consolidates 2+ Yardi property codes (e.g. two buildings, one workpaper): fill in the **Buildings** table (name, Yardi code, SF per building) — Size (SF) above auto-sums from it
- [ ] **Ownership**: Investor Legal Name (per W-9) + Short Name, Invoice Prefix (leave blank to auto-derive from Display Name)
- [ ] **Team Members**: everyone reviewing this property's close
- [ ] **Chart of Accounts**: "Uses the standard GRP Yardi COA?" — Yes for any GRP-managed Yardi property (no upload needed); No only for a partner running their own Yardi with different codes
- [ ] **Building / Allocation Splits**: only if this property allocates shared costs across buildings — auto-suggested (Equal + By-SF schedules) once the Buildings table above has 2+ rows
- [ ] **Management Fee Lines**: one row per PM agreement (e.g. JLL 1.25% + GRP 1.75%, matching Rev Labs)
- [ ] **Bank Accounts**: one row per account (operating/development/DACA) — Bank Name + Account Number drive monthly auto-classification of uploaded statements
- [ ] **Payment Instructions**: ACH and/or check details for the management fee invoice
- [ ] **RE Tax & Other**: `re_tax_payment_months` for this jurisdiction (typically Jan/Apr/Jul/Oct), Parcel IDs if relevant

**2. Files to drop in at onboarding (Property Setup page, above the form)**
- [ ] **Current Year Budget (Kardin)** — upload the annual budget; enter the saved filename in `Kardin Budget Filename` (step 8)
- [ ] **Tenancy Schedule / Rent Roll** — upload the current rent roll once to confirm Tenant Utility Billing picks up the right tenants (it re-reads this fresh every period going forward, so nothing to configure — just confirm it looks right)
- [ ] **12-Month GL History** (if available) — informational only, flags which vendors bill quarterly/semi-annually so nothing gets missed as a one-off accrual later
- [ ] **Bank Statement(s)** — upload one real statement per account to auto-extract the account number into step 6, instead of typing it blind (PNC / Bank of America / KeyBank only — a different bank needs a new parser first)

**3. Prepaid Ledger Seed** ← most critical for acquisitions
- [ ] Gather all active prepaid schedules from prior management (insurance and RE tax are excluded automatically; focus on service contracts, subscriptions, maintenance agreements)
- [ ] Either type directly into the **Prepaid Ledger Seed Builder**, or download its blank template, fill it in Excel, and upload it back
- [ ] Set the correct **Months Amortized** per item (= months already released by prior manager) — this is what makes an acquisition's carry-forward balance correct
- [ ] Download the generated seed file and upload it as the "Prior Month Prepaid Ledger" in the Pass 1 sidebar on the first close
- [ ] Verify the **Remaining** column in the preview — it determines how many more months the pipeline will amortize

**4. T12 for Layer 3 Historical Accruals**
- [ ] Source the T12 from prior management in Yardi Income Statement format (`.xlsx`)
- [ ] ⚠️ If T12 is unavailable, Layer 3 will rely on BC YTD ÷ months elapsed — expect over/under accruals on the first 1–2 closes while history builds
- [ ] Upload the T12 in the Pass 1 sidebar under "Reference" files

**5. Bank Rec Starting Balance**
- [ ] Confirm prior period outstanding checks with the prior manager (if any)
- [ ] On the first close, upload the Yardi Bank Rec PDF as the bank rec source — it carries the correct reconciled balance forward automatically

**6. On the First Close — Review Carefully**
- [ ] **Layer 3 accruals** may fire for accounts that prior management handled differently — review all "historical" source entries before uploading to Yardi
- [ ] **Pass 2 → Accrual Check tab** will be empty on the first close (no prior pipeline J-type JEs to compare against) — this is expected
- [ ] **Prior accrual check warning** in Pass 2 is expected to show $0 — not a bug
- [ ] Review the prepaid ledger `Remaining` column after the first close to confirm the amortization schedule looks correct going forward
""")
    else:
        st.markdown("""
### New Property — Before First Close
(new construction / lease-up, no prior manager to carry anything forward from)

**1. Property Config — ⚙️ Properties → Add/Edit, section by section**
- [ ] **Basic Information**: GA Property ID, Display Name, Address, Type, Size
- [ ] If this property consolidates 2+ Yardi property codes: fill in the **Buildings** table (name, Yardi code, SF)
- [ ] **Ownership**: Investor Legal Name (per W-9) + Short Name, Invoice Prefix (leave blank to auto-derive)
- [ ] **Team Members**
- [ ] **Chart of Accounts**: "Uses the standard GRP Yardi COA?" — Yes for a GRP-managed Yardi property
- [ ] **Building / Allocation Splits**: only if multi-building — auto-suggested once the Buildings table has 2+ rows
- [ ] **Management Fee Lines**, **Bank Accounts**, **Payment Instructions**, **RE Tax & Other** (re_tax_payment_months for this jurisdiction)

**2. Files to drop in at onboarding (Property Setup page, above the form)**
- [ ] **Current Year Budget (Kardin)** — upload the annual budget; enter the saved filename in `Kardin Budget Filename` (step 8)
- [ ] **Tenancy Schedule / Rent Roll** — upload once any leases exist, so Tenant Utility Billing has tenants to bill; safe to skip if there are none yet
- [ ] **12-Month GL History** — not applicable yet, there's no history
- [ ] **Bank Statement(s)** — upload one per account to auto-extract the account number into step 6, instead of typing it blind

**3. Prepaid Ledger Seed**
- [ ] If there are NO active prepaids at open: skip — the ledger starts empty
- [ ] If prepaids exist from day 1 (e.g. insurance paid at inception): use the **Prepaid Ledger Seed Builder** with `Months Amortized = 0`

**4. On the First Close**
- [ ] T12 upload is optional — there is no history, so Layer 3 will be silent (expected for month 1)
- [ ] January Layer 3 uses the annual budget ÷ 12 fallback for accounts with zero GL activity — review these accruals carefully
- [ ] **Pass 2 → Accrual Check tab** will be empty on the first close — expected
- [ ] The BS Workpaper starts fresh — do not upload a prior workpaper (leave that slot blank)
""")

    # Workpaper Seed Builder removed 2026-08-24 — workpapers are never part of
    # property config; the real BS Workpaper is generated by the pipeline
    # itself and carried forward by uploading the prior month's file as
    # "Prior Month Workpaper" in the Pass 2 sidebar each period. This tool's
    # output (pipeline/bs_workpaper_generator.py: generate_workpaper_seed(),
    # _seed_period_sort()) is left in place in case it's useful again later,
    # just no longer surfaced here.

# ── Feedback Inbox (bottom of Properties tab) ──────────────────────────────────
with tab4:
    st.divider()
    st.markdown("## 🐛 Feedback Inbox")
    st.caption(
        "Issues submitted by the team via the 'Report an Issue' sidebar form. "
        "Mark items resolved once fixed — the health check agent will skip resolved items."
    )

    try:
        import json as _fb_inbox_json
        from pathlib import Path as _FbInboxPath
        _fb_log_path = _FbInboxPath(__file__).parent / 'data' / 'feedback_log.jsonl'

        if not _fb_log_path.exists():
            st.info("No feedback submitted yet.", icon="ℹ️")
        else:
            # Load all items
            _fb_all = []
            for _fb_line in _fb_log_path.read_text(encoding='utf-8').splitlines():
                _fb_line = _fb_line.strip()
                if not _fb_line:
                    continue
                try:
                    _fb_all.append(_fb_inbox_json.loads(_fb_line))
                except Exception:
                    continue

            if not _fb_all:
                st.info("No feedback items found.", icon="ℹ️")
            else:
                _sev_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
                _sev_icon  = {'critical': '🔴', 'high': '🟠', 'medium': '🟡', 'low': '🟢'}
                _status_icon = {'open': '⬜', 'acknowledged': '🔵', 'resolved': '✅'}

                # Filter controls
                _fb_filter_col1, _fb_filter_col2 = st.columns(2)
                with _fb_filter_col1:
                    _fb_show_status = st.multiselect(
                        "Show status",
                        ['open', 'acknowledged', 'resolved'],
                        default=['open', 'acknowledged'],
                        key="fb_inbox_status_filter",
                    )
                with _fb_filter_col2:
                    _fb_show_sev = st.multiselect(
                        "Show severity",
                        ['critical', 'high', 'medium', 'low'],
                        default=['critical', 'high', 'medium', 'low'],
                        key="fb_inbox_sev_filter",
                    )

                _fb_filtered = [
                    item for item in _fb_all
                    if item.get('status', 'open') in _fb_show_status
                    and item.get('severity', 'medium') in _fb_show_sev
                ]
                _fb_filtered.sort(
                    key=lambda x: (_sev_order.get(x.get('severity', 'medium'), 2),
                                   x.get('submitted_at', ''))
                )

                st.caption(f"Showing {len(_fb_filtered)} of {len(_fb_all)} item(s)")

                _fb_changed = False
                for _fbi, _fb_item in enumerate(_fb_filtered):
                    _fb_sev    = _fb_item.get('severity', 'medium')
                    _fb_status = _fb_item.get('status', 'open')
                    _fb_si     = _sev_icon.get(_fb_sev, '🟡')
                    _fb_sti    = _status_icon.get(_fb_status, '⬜')

                    with st.container():
                        _fbc1, _fbc2, _fbc3 = st.columns([5, 2, 2])
                        with _fbc1:
                            st.markdown(
                                f"**{_fb_si} [{_fb_sev.upper()}]** &nbsp; "
                                f"{_fb_item.get('submitted_at', '')} &nbsp;·&nbsp; "
                                f"{_fb_item.get('reporter', '?')} "
                                f"({_fb_item.get('property_code', '')} / "
                                f"{_fb_item.get('period', '')})"
                            )
                            st.markdown(
                                f"<div style='padding:4px 0 8px 0;color:#212121;'>"
                                f"{_fb_item.get('description', '')}</div>",
                                unsafe_allow_html=True,
                            )
                        with _fbc2:
                            _new_status = st.selectbox(
                                "Status",
                                ['open', 'acknowledged', 'resolved'],
                                index=['open', 'acknowledged', 'resolved'].index(_fb_status),
                                key=f"fb_status_{_fbi}",
                                label_visibility="collapsed",
                            )
                        with _fbc3:
                            if st.button("Save", key=f"fb_save_{_fbi}",
                                         use_container_width=True):
                                # Update this item's status in the full list
                                _orig_idx = next(
                                    (i for i, x in enumerate(_fb_all)
                                     if x.get('submitted_at') == _fb_item.get('submitted_at')
                                     and x.get('description') == _fb_item.get('description')),
                                    None,
                                )
                                if _orig_idx is not None:
                                    _fb_all[_orig_idx]['status'] = _new_status
                                    _fb_changed = True

                    st.divider()

                if _fb_changed:
                    # Rewrite the log with updated statuses
                    with open(_fb_log_path, 'w', encoding='utf-8') as _fb_wf:
                        for _fb_row in _fb_all:
                            _fb_wf.write(_fb_inbox_json.dumps(_fb_row) + '\n')
                    st.success("Status updated.")
                    st.rerun()

    except Exception as _fb_inbox_err:
        st.error(f"Could not load feedback log: {_fb_inbox_err}")

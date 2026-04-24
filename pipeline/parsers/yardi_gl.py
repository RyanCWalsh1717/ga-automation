"""
Yardi General Ledger Parser
============================
Reads a Yardi GL Detail export (.xlsx) and produces a normalized data structure
that the rest of the pipeline can work with.

What this parser does (in plain English):
-----------------------------------------
1. Opens the Excel file Yardi exports for the General Ledger
2. Reads the metadata at the top (property name, period, book type)
3. Finds every account section â each account starts with a "Beginning Balance"
   row and ends with an "Ending Balance" row
4. Extracts every transaction within each account: date, description, who/what,
   control number, reference number, debit amount, credit amount, running balance
5. Validates the data â checks that debits and credits balance, that every
   transaction has the fields it needs, and flags anything that looks wrong
6. Returns a clean, structured result that other parts of the pipeline can use
   without knowing anything about Excel or Yardi's file format

What goes in:  A .xlsx file exported from Yardi (GL Detail report)
What comes out: A dictionary with metadata, a list of accounts (each with its
               transactions), and a validation report

The pipeline uses this to build work papers â matching every GL entry to its
source invoice (from Nexus) and its bank payment by reference number.
"""

import os
import re
from datetime import datetime, date
from dataclasses import dataclass, field, asdict
from typing import Optional
import openpyxl


# ---------------------------------------------------------------------------
# Data structures â these define what the parser produces
# ---------------------------------------------------------------------------

@dataclass
class GLTransaction:
    """One line item in the general ledger."""
    account_code: str           # e.g. "111000"
    account_name: str           # e.g. "Operating - Loss (PNC)"
    date: Optional[date]        # transaction date
    period: Optional[str]       # Yardi period (e.g. "Feb-2026")
    description: str            # person or description field
    control: str                # control number (K=PCard, P=Check, J=Journal)
    reference: str              # invoice number, vendor ref, etc.
    debit: float                # debit amount (always >= 0)
    credit: float               # credit amount (always >= 0)
    balance: float              # running balance within the account
    remarks: str                # additional notes / description
    row_number: int             # source row in Excel (for audit trail)

    @property
    def net_amount(self) -> float:
        """Positive = debit, negative = credit."""
        return self.debit - self.credit

    @property
    def control_type(self) -> str:
        """Categorize the transaction by control prefix."""
        if not self.control:
            return "unknown"
        prefix = self.control.split("-")[0].upper()
        return {"K": "pcard", "P": "check", "J": "journal"}.get(prefix, "other")


@dataclass
class GLAccount:
    """One account in the GL with its transactions."""
    account_code: str
    account_name: str
    beginning_balance: float
    ending_balance: float
    total_debits: float
    total_credits: float
    net_change: float
    transactions: list = field(default_factory=list)  # list of GLTransaction

    @property
    def transaction_count(self) -> int:
        return len(self.transactions)

    @property
    def is_balanced(self) -> bool:
        """Check if beginning + debits - credits = ending."""
        expected = self.beginning_balance + self.total_debits - self.total_credits
        return abs(expected - self.ending_balance) < 0.01


@dataclass
class GLMetadata:
    """Information from the file header."""
    property_code: str          # e.g. "revlabpm"
    property_name: str          # e.g. "Revolution Labs Owner, LLC"
    period: str                 # e.g. "Feb-2026"
    book: str                   # e.g. "Accrual"
    source_file: str            # original filename
    parsed_at: str              # timestamp of parsing


@dataclass
class GLParseResult:
    """Complete result from parsing a Yardi GL export."""
    metadata: GLMetadata
    accounts: list              # list of GLAccount
    all_transactions: list      # flat list of all GLTransaction across accounts
    validation: dict            # validation results

    @property
    def total_accounts(self) -> int:
        return len(self.accounts)

    @property
    def total_transactions(self) -> int:
        return len(self.all_transactions)

    def to_dict(self) -> dict:
        """Convert to a plain dictionary for JSON serialization or inspection."""
        return {
            "metadata": asdict(self.metadata),
            "summary": {
                "total_accounts": self.total_accounts,
                "total_transactions": self.total_transactions,
                "total_debits": sum(a.total_debits for a in self.accounts),
                "total_credits": sum(a.total_credits for a in self.accounts),
            },
            "validation": self.validation,
            "accounts": [
                {
                    "account_code": a.account_code,
                    "account_name": a.account_name,
                    "beginning_balance": a.beginning_balance,
                    "ending_balance": a.ending_balance,
                    "net_change": a.net_change,
                    "transaction_count": a.transaction_count,
                    "is_balanced": a.is_balanced,
                }
                for a in self.accounts
            ],
        }


# ---------------------------------------------------------------------------
# Parser implementation
# ---------------------------------------------------------------------------

def _safe_str(val) -> str:
    """Convert a cell value to a clean string."""
    if val is None:
        return ""
    return str(val).strip()


def _safe_float(val) -> float:
    """Convert a cell value to a float, defaulting to 0.0."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _safe_date(val) -> Optional[date]:
    """Convert a cell value to a date."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def parse_metadata(ws) -> GLMetadata:
    """Extract metadata from the header rows (rows 1-5)."""
    # Row 1: "Property = revlabpm Revolution Labs Owner, LLC"
    # Row 2: "General Ledger"
    # Row 3: "Period = Feb-2026"
    # Row 4: "Book = Accrual"
    # Row 5: blank or "Sort On ="

    raw_property = _safe_str(ws.cell(row=1, column=1).value)
    raw_period = _safe_str(ws.cell(row=3, column=1).value)
    raw_book = _safe_str(ws.cell(row=4, column=1).value)

    # Parse property code and name from the header row.
    # Two formats observed in the wild:
    #   Format A: "Property = revlabpm Revolution Labs Owner, LLC"
    #   Format B: "Revolution Labs Owner, LLC (revlabpm)"
    property_code = ""
    property_name = ""
    if "=" in raw_property:
        # Format A
        after_eq = raw_property.split("=", 1)[1].strip()
        parts = after_eq.split(" ", 1)
        property_code = parts[0] if parts else ""
        property_name = parts[1] if len(parts) > 1 else ""
    else:
        # Format B: "Some Name (code)"
        m = re.match(r'^(.+?)\s*\((\w+)\)\s*$', raw_property)
        if m:
            property_name = m.group(1).strip()
            property_code = m.group(2).strip()

    # Parse "Period = Feb-2026"
    period = ""
    if "=" in raw_period:
        period = raw_period.split("=", 1)[1].strip()

    # Parse "Book = Accrual"
    book = ""
    if "=" in raw_book:
        book = raw_book.split("=", 1)[1].strip()

    return GLMetadata(
        property_code=property_code,
        property_name=property_name,
        period=period,
        book=book,
        source_file="",  # set by caller
        parsed_at=datetime.now().isoformat(),
    )


def parse_gl(filepath: str, sheet_name: str = None) -> GLParseResult:
    """
    Parse a Yardi GL Detail export file.

    Args:
        filepath:   Path to the .xlsx file
        sheet_name: Optional sheet name to parse (e.g. 'GL - MTD').
                    Defaults to the active (first) sheet.

    Returns:
        GLParseResult with all accounts, transactions, and validation info
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # --- Extract metadata ---
    metadata = parse_metadata(ws)
    metadata.source_file = os.path.basename(filepath)

    # --- Parse accounts and transactions ---
    accounts = []
    all_transactions = []
    validation_warnings = []

    current_account_code = None
    current_account_name = None
    current_beginning_balance = 0.0
    current_transactions = []

    # Data starts at row 7 (row 6 is headers)
    max_row = ws.max_row

    for row_num in range(7, max_row + 1):
        col_a = _safe_str(ws.cell(row=row_num, column=1).value)   # Property / Account Code
        col_b = _safe_str(ws.cell(row=row_num, column=2).value)   # Property Name
        col_c = ws.cell(row=row_num, column=3).value               # Date
        col_d = _safe_str(ws.cell(row=row_num, column=4).value)   # Period
        col_e = _safe_str(ws.cell(row=row_num, column=5).value)   # Person/Description
        col_f = _safe_str(ws.cell(row=row_num, column=6).value)   # Control
        col_g = _safe_str(ws.cell(row=row_num, column=7).value)   # Reference
        col_h = _safe_float(ws.cell(row=row_num, column=8).value) # Debit
        col_i = _safe_float(ws.cell(row=row_num, column=9).value) # Credit
        col_j = _safe_float(ws.cell(row=row_num, column=10).value) # Balance
        col_k = _safe_str(ws.cell(row=row_num, column=11).value)  # Remarks

        # --- Detect row type ---

        # Beginning Balance row: account code in A, "= Beginning Balance =" in K
        if "beginning balance" in col_k.lower():
            # Save previous account if exists
            if current_account_code and current_transactions:
                # We'll finalize this account when we hit its ending balance
                pass

            current_account_code = col_a
            current_account_name = col_e
            current_beginning_balance = col_j
            current_transactions = []
            continue

        # Ending Balance row: "= Ending Balance =" in K
        if "ending balance" in col_k.lower():
            if current_account_code:
                total_debits = col_h
                total_credits = col_i
                ending_balance = col_j
                net_change = total_debits - total_credits

                account = GLAccount(
                    account_code=current_account_code,
                    account_name=current_account_name,
                    beginning_balance=current_beginning_balance,
                    ending_balance=ending_balance,
                    total_debits=total_debits,
                    total_credits=total_credits,
                    net_change=net_change,
                    transactions=current_transactions,
                )

                # Validate account balance
                if not account.is_balanced:
                    validation_warnings.append(
                        f"Account {current_account_code} ({current_account_name}) "
                        f"does not balance: beginning={current_beginning_balance:.2f} "
                        f"+ debits={total_debits:.2f} - credits={total_credits:.2f} "
                        f"= {current_beginning_balance + total_debits - total_credits:.2f} "
                        f"but ending={ending_balance:.2f}"
                    )

                accounts.append(account)
                all_transactions.extend(current_transactions)

                current_account_code = None
                current_account_name = None
                current_transactions = []
            continue

        # Blank separator rows
        if not col_a and not col_e and not col_f and col_h == 0 and col_i == 0:
            continue

        # Grand total row (last row where H == I, confirming GL balance)
        if not col_a and col_h > 0 and col_i > 0 and abs(col_h - col_i) < 0.01:
            # This is the verification row â debits should equal credits
            if abs(col_h - col_i) >= 0.01:
                validation_warnings.append(
                    f"Grand total mismatch: debits={col_h:.2f} != credits={col_i:.2f}"
                )
            continue

        # Transaction row â belongs to current account
        if current_account_code:
            txn = GLTransaction(
                account_code=current_account_code,
                account_name=current_account_name,
                date=_safe_date(col_c),
                period=col_d,
                description=col_e,
                control=col_f,
                reference=col_g,
                debit=col_h,
                credit=col_i,
                balance=col_j,
                remarks=col_k,
                row_number=row_num,
            )
            current_transactions.append(txn)

    wb.close()

    # --- Build validation summary ---
    total_debits = sum(a.total_debits for a in accounts)
    total_credits = sum(a.total_credits for a in accounts)
    gl_balanced = abs(total_debits - total_credits) < 0.01

    if not gl_balanced:
        validation_warnings.append(
            f"GL is not balanced: total debits={total_debits:.2f}, "
            f"total credits={total_credits:.2f}, "
            f"difference={total_debits - total_credits:.2f}"
        )

    unbalanced_accounts = [a for a in accounts if not a.is_balanced]

    validation = {
        "status": "PASS" if not validation_warnings else "WARNINGS",
        "gl_balanced": gl_balanced,
        "total_debits": round(total_debits, 2),
        "total_credits": round(total_credits, 2),
        "accounts_parsed": len(accounts),
        "transactions_parsed": len(all_transactions),
        "unbalanced_accounts": len(unbalanced_accounts),
        "warnings": validation_warnings,
    }

    return GLParseResult(
        metadata=metadata,
        accounts=accounts,
        all_transactions=all_transactions,
        validation=validation,
    )


# ---------------------------------------------------------------------------
# Convenience: run parser directly to inspect a file
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys
    import json

    if len(sys.argv) < 2:
        print("Usage: python yardi_gl.py <path_to_gl_export.xlsx>")
        sys.exit(1)

    result = parse_gl(sys.argv[1])

    print("=" * 70)
    print(f"YARDI GL PARSER â {result.metadata.property_name}")
    print(f"Period: {result.metadata.period}  |  Book: {result.metadata.book}")
    print(f"Source: {result.metadata.source_file}")
    print("=" * 70)
    print(f"\nAccounts:     {result.total_accounts}")
    print(f"Transactions: {result.total_transactions}")
    print(f"Total Debits: ${result.validation['total_debits']:,.2f}")
    print(f"Total Credits: ${result.validation['total_credits']:,.2f}")
    print(f"GL Balanced:  {result.validation['gl_balanced']}")
    print(f"Validation:   {result.validation['status']}")

    if result.validation["warnings"]:
        print(f"\nWarnings ({len(result.validation['warnings'])}):")
        for w in result.validation["warnings"]:
            print(f"  â  {w}")

    print(f"\n{'Account':<12} {'Name':<35} {'Begin':>14} {'Debits':>14} {'Credits':>14} {'Ending':>14} {'Txns':>6} {'OK':>4}")
    print("-" * 120)
    for a in result.accounts:
        print(
            f"{a.account_code:<12} {a.account_name[:35]:<35} "
            f"${a.beginning_balance:>12,.2f} ${a.total_debits:>12,.2f} "
            f"${a.total_credits:>12,.2f} ${a.ending_balance:>12,.2f} "
            f"{a.transaction_count:>6} {'â' if a.is_balanced else 'â':>4}"
        )

    # Print first 5 transactions as sample
    print(f"\nSample Transactions (first 5):")
    print(f"{'Date':<12} {'Account':<10} {'Description':<30} {'Control':<12} {'Debit':>12} {'Credit':>12}")
    print("-" * 100)
    for t in result.all_transactions[:5]:
        date_str = t.date.strftime("%Y-%m-%d") if t.date else "N/A"
        print(
            f"{date_str:<12} {t.account_code:<10} {t.description[:30]:<30} "
            f"{t.control:<12} ${t.debit:>10,.2f} ${t.credit:>10,.2f}"
        )

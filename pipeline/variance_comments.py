"""
Variance Comment Generator — GRP Standards
==========================================
Applies Greatland Realty Partners' 3-tier variance threshold and generates
narrative commentary conforming to the Variance Commentary Standards document.

Tier classification (first-match, no overlap):
  Tier 3  abs < $2,500                           → No action required
  Tier 1  abs ≥ $5,000  OR  pct ≥ 5%            → Full 1–2 sentence comment
  Tier 2  $2,500 ≤ abs < $5,000  AND  pct < 5%  → Flag phrase only

Note: first-match precedence is enforced in classify_tier(). A variance that
hits the Tier 1 pct condition (≥5%) will never be classified Tier 2 even if
its dollar amount falls in the Tier 2 range ($2,500–$4,999).

Output is written directly into the budget comparison Excel file:
  Column L = MTD Variance Notes  (Tahoma 10, wrap_text, vertical top)
  Column M = YTD Variance Notes

Never comments on Total / Subtotal / Section Header / NOI / Net Income rows.
"""

import json
import os
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from property_config import is_revenue_account


# ── Thresholds ────────────────────────────────────────────────
TIER1_ABS = 5_000.0   # ≥ $5,000 absolute  OR
TIER1_PCT = 0.05      # ≥ 5 % of budget    (and abs ≥ $2,500 floor)
TIER2_MIN = 2_500.0   # $2,500 floor — below this: Tier 3 (no action)

MONTH_MAP = {
    'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4,
    'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8,
    'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12,
}

# ── Account-specific behavioral context (Rev Labs) ───────────
ACCOUNT_CONTEXT: Dict[str, str] = {
    '440100': (
        'Recovery-Operating Expense is a CAM billback to tenants. '
        'A structural variance occurs if the budget included non-recoverable items '
        '(e.g., management fees not permitted under the lease). '
        'Flag as ongoing and recommend a budget revision if so.'
    ),
    '613110': (
        'Utilities-Electricity: compare to Kardin monthly estimate by sub-meter. '
        'Large variances may relate to tenant sub-metering reimbursements — '
        'check accounts 613115 and 440500 for offsetting billback activity.'
    ),
    '613210': (
        'Utilities-Gas is a seasonal account: high in Jan–Mar and Nov–Dec, '
        'low in summer. Kardin has three meter lines: 49789-85790 HVAC (main driver), '
        '49789-42240 STYGEN, and 49789-20610 EMGEN. State which meter is driving the variance.'
    ),
    '613310': (
        'Utilities-Water/Sewer: Newton municipality bills semi-annually. '
        'Large accrual variances against the flat monthly budget are timing differences. '
        'State the billing period covered by the accrual and confirm '
        'full-year cost is within annual budget.'
    ),
    '617110': (
        'HVAC Contract Services: semi-annual preventive maintenance invoices '
        '(spring Q2 / fall Q4) create favorable MTD/YTD variance until invoiced. '
        'Name the specific contractors outstanding and the expected quarters for posting.'
    ),
    '617120': (
        'HVAC Repairs: variable account. Compare actual to prior-year run rate '
        'and remaining annual budget. Name the specific system or repair event.'
    ),
    '635110': (
        'Snow & Ice Removal: highly weather-dependent. '
        'Name the vendor (Landscape America if applicable), '
        'reference specific storm events visible in GL, '
        'state remaining annual budget and risk of full-year overage.'
    ),
    '637130': (
        'Admin-Management Fees: verify the accrual matches the calculated fee '
        '(cash received × applicable rate). Flag discrepancy for Finance review '
        'if accrual does not tie to the calculation.'
    ),
    '637290': (
        'Admin-Telephone: watch for duplicate PCard charges (two charges per property '
        'per month). Flag for Finance review if duplicates are detected rather than '
        'writing a normalized comment.'
    ),
    '639110': (
        'Insurance-Property: monthly amortization of prepaid premium. '
        'If charges exceed budget, note the policy period being amortized '
        'and expected monthly run rate going forward.'
    ),
    '639120': (
        'Insurance-General Liability: if $0 actual, GL/auto insurance amortization '
        'is likely being coded to 639110 instead. Flag the account mapping issue '
        'for Finance rather than treating as a favorable variance.'
    ),
    '801110': (
        'Interest Expense: compare actual mortgage payment to the GRP amortization '
        'schedule budget. If a scheduled step-down does not match the actual payment, '
        'flag for Finance to reconcile the amortization schedule.'
    ),
}

# ── Rows that must never receive comments ────────────────────
_SKIP_NAMES = {
    'INCOME', 'REVENUE', 'TOTAL REVENUE', 'RENTAL REVENUE', 'TOTAL RENTAL REVENUE',
    'BASE RENT', 'TOTAL BASE RENT', 'RECOVERY INCOME', 'TOTAL RECOVERY INCOME',
    'OTHER TENANT INCOME', 'TOTAL OTHER TENANT INCOME',
    'TENANT SERVICE REVENUE', 'TOTAL TENANT SERVICE REVENUE',
    'OTHER INCOME', 'TOTAL OTHER INCOME',
    'OPERATING EXPENSES - RECOVERABLE', 'TOTAL OPERATING EXPENSES - RECOVERABLE',
    'OPERATING EXPENSES - NON RECOVERABLE', 'TOTAL OPERATING EXPENSES - NON RECOVERABLE',
    'TOTAL EXPENSES', 'NET OPERATING INCOME', 'NET INCOME',
    'INTEREST EXPENSE', 'CLEANING / JANITORIAL', 'UTILITIES',
    'GENERAL REPAIRS & MAINTENANCE', 'HVAC MAINTENANCE',
    'SECURITY / FIRE / LIFE SAFETY', 'LANDSCAPING',
    'PARKING & SNOW REMOVAL', 'ADMINISTRATIVE', 'INSURANCE', 'REAL ESTATE TAXES',
}


# ══════════════════════════════════════════════════════════════
# 1. TIER CLASSIFICATION
# ══════════════════════════════════════════════════════════════

def classify_tier(actual: float, budget: float) -> Tuple[str, float, float]:
    """
    Apply GRP 3-tier threshold logic with explicit first-match precedence.

    Rules (evaluated in order — first match wins, no overlap):
      Tier 3  abs_var < $2,500
      Tier 1  abs_var ≥ $5,000  OR  |pct_var| ≥ 5 %
      Tier 2  $2,500 ≤ abs_var < $5,000  AND  |pct_var| < 5 %

    A $3,000 variance at 10% → Tier 1 (pct condition fires first).
    A $3,000 variance at 3%  → Tier 2 (dollar in range, pct below threshold).

    Returns:
        (tier, abs_variance, pct_variance)
        tier: 'tier_1' | 'tier_2' | 'tier_3'
    """
    abs_var = actual - budget
    abs_var_dollar = abs(abs_var)
    pct_var = (abs_var / abs(budget) * 100) if budget and budget != 0 else 0.0
    abs_pct = abs(pct_var)

    # ── Step 1: below floor → Tier 3 (no action) ──
    if abs_var_dollar < TIER2_MIN:
        return 'tier_3', abs_var, pct_var

    # ── Step 2: Tier 1 — large dollar OR significant pct ──
    if abs_var_dollar >= TIER1_ABS or abs_pct >= (TIER1_PCT * 100):
        return 'tier_1', abs_var, pct_var

    # ── Step 3: Tier 2 — mid-range dollar AND sub-threshold pct ──
    # Reached only when: $2,500 ≤ abs < $5,000 AND pct < 5%
    return 'tier_2', abs_var, pct_var


def _is_skip_row(account_code: Any, account_name: Any) -> bool:
    """Return True if this row should never receive a comment."""
    code = str(account_code or '').strip()
    name = str(account_name or '').strip().upper()

    if not code:
        return True
    if code.endswith('000') or code.endswith('999'):
        return True
    if name in _SKIP_NAMES:
        return True
    if any(kw in name for kw in ('TOTAL', 'SUBTOTAL', 'NET OPERATING', 'NET INCOME')):
        return True
    return False


# ══════════════════════════════════════════════════════════════
# 2. KARDIN ENRICHMENT
# ══════════════════════════════════════════════════════════════

def build_kardin_enrichment(kardin_records: List[dict], account_code: str,
                             period_month: int) -> dict:
    """
    Aggregate all Kardin sub-lines for an account into enrichment context.

    Returns dict with:
      descriptions       — list of budget line descriptions (vendor/contract intent)
      month_budget       — total Kardin budget for the reporting month
      annual_budget      — sum of MTotal across all sub-lines
      monthly_pattern    — {1..12: amount} summed across sub-lines
      is_seasonal        — True if variance across months is high (σ/mean > 0.5)
      low_months         — list of month numbers where budget < 50% of monthly avg
      high_months        — list of month numbers where budget > 150% of monthly avg
      seasonality_note   — human-readable summary of seasonal pattern
    """
    code = str(account_code).strip()
    rows = [r for r in kardin_records if str(r.get('account_code', '')).strip() == code]

    if not rows:
        return {}

    # Collect descriptions (skip generic linked descriptions)
    descs = []
    for r in rows:
        d = str(r.get('description', '')).strip()
        if d and 'Linked to' not in d and d not in descs:
            descs.append(d)

    # Sum monthly amounts
    monthly: Dict[int, float] = {}
    annual_total = 0.0
    for m in range(1, 13):
        key = f'M{m}'
        val = sum(float(r.get(key, 0) or 0) for r in rows)
        monthly[m] = val
        annual_total += val

    month_budget = monthly.get(period_month, 0.0)

    # Seasonality: compare monthly values
    nonzero = [v for v in monthly.values() if v != 0]
    is_seasonal = False
    low_months: List[int] = []
    high_months: List[int] = []
    seasonality_note = ''

    if len(nonzero) >= 3:
        avg = sum(nonzero) / len(nonzero)
        if avg > 0:
            deviations = [abs(v - avg) / avg for v in nonzero]
            is_seasonal = (sum(deviations) / len(deviations)) > 0.40

            if is_seasonal:
                low_months = [m for m in range(1, 13)
                              if monthly[m] < avg * 0.50 and monthly[m] >= 0]
                high_months = [m for m in range(1, 13) if monthly[m] > avg * 1.50]

                if high_months:
                    hi_names = [_month_name(m) for m in high_months]
                    lo_names = [_month_name(m) for m in low_months] if low_months else ['other months']
                    seasonality_note = (
                        f"Budget is concentrated in {', '.join(hi_names)} "
                        f"with lower spend in {', '.join(lo_names)}."
                    )

    return {
        'descriptions': descs,
        'month_budget': month_budget,
        'annual_budget': annual_total,
        'monthly_pattern': monthly,
        'is_seasonal': is_seasonal,
        'low_months': low_months,
        'high_months': high_months,
        'seasonality_note': seasonality_note,
    }


def _month_name(m: int) -> str:
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    return months[m - 1] if 1 <= m <= 12 else str(m)


# ══════════════════════════════════════════════════════════════
# 3. GL TRANSACTION CONTEXT
# ══════════════════════════════════════════════════════════════

def build_gl_context(gl_parsed, account_code: str) -> dict:
    """
    Pull transaction-level detail from parsed GL for a specific account.

    Returns dict with:
      transactions    — list of {date, description, debit, credit, net, remarks}
      vendor_summary  — {vendor_key: {total, count}} sorted by abs(total)
      net_activity    — total net debit/credit for the period
      has_reversals   — True if any transaction description contains ':Reversal'
    """
    code = str(account_code).strip()
    result = {
        'transactions': [],
        'vendor_summary': {},
        'net_activity': 0.0,
        'has_reversals': False,
    }

    if not gl_parsed:
        return result

    # Support both object-style (with .accounts) and dict-style parsed data
    accounts_list = None
    if hasattr(gl_parsed, 'accounts'):
        accounts_list = gl_parsed.accounts
    elif isinstance(gl_parsed, dict):
        accounts_list = gl_parsed.get('accounts', [])
    elif isinstance(gl_parsed, list):
        accounts_list = gl_parsed

    if not accounts_list:
        return result

    for acct in accounts_list:
        acct_code = str(getattr(acct, 'account_code', '') or
                        (acct.get('account_code', '') if isinstance(acct, dict) else '')).strip()
        if acct_code != code:
            continue

        txns_raw = (getattr(acct, 'transactions', None) or
                    (acct.get('transactions', []) if isinstance(acct, dict) else []))

        net_total = 0.0
        for txn in txns_raw:
            if isinstance(txn, dict):
                debit = float(txn.get('debit', 0) or 0)
                credit = float(txn.get('credit', 0) or 0)
                desc = str(txn.get('description', '') or '')
                date_val = txn.get('date', '')
                remarks = str(txn.get('remarks', '') or '')
            else:
                debit = float(getattr(txn, 'debit', 0) or 0)
                credit = float(getattr(txn, 'credit', 0) or 0)
                desc = str(getattr(txn, 'description', '') or '')
                date_val = getattr(txn, 'date', '')
                remarks = str(getattr(txn, 'remarks', '') or '')

            net = debit - credit
            net_total += net

            is_reversal = ':Reversal' in desc or ':Reversal' in remarks
            if is_reversal:
                result['has_reversals'] = True

            # Format date
            date_str = ''
            if date_val:
                try:
                    if hasattr(date_val, 'strftime'):
                        date_str = date_val.strftime('%m/%d/%Y')
                    else:
                        date_str = str(date_val)[:10]
                except Exception:
                    date_str = str(date_val)

            t = {
                'date': date_str,
                'description': desc,
                'debit': debit,
                'credit': credit,
                'net': net,
                'remarks': remarks,
                'is_reversal': is_reversal,
            }
            result['transactions'].append(t)

            # Vendor summary — strip Yardi internal codes like (v0000073) (t0000011)
            import re as _re
            clean_desc = _re.sub(r'\s*\([vt]\d+\)\s*', '', desc).strip()
            vendor_key = (clean_desc or desc)[:50].strip() or '(no description)'
            vs = result['vendor_summary']
            if vendor_key not in vs:
                vs[vendor_key] = {'total': 0.0, 'count': 0}
            vs[vendor_key]['total'] += net
            vs[vendor_key]['count'] += 1

        result['net_activity'] = net_total
        break

    return result


# ══════════════════════════════════════════════════════════════
# 4. API PROMPT — GRP STANDARD
# ══════════════════════════════════════════════════════════════

_SYSTEM_PROMPT = """\
You are a CRE finance analyst at Greatland Realty Partners (GRP) writing monthly \
variance commentary for the Revolution Labs property (1050 Waltham St, Lexington MA). \
The audience is GRP management and the institutional investor Singerman Real Estate.

COMMENTARY STANDARDS — follow exactly:

FORMAT: Start with "Variance due to [specific cause]." — not with the account name or dollar amount.

REQUIRED ELEMENTS IN EVERY TIER 1 COMMENT:
1. CAUSE — Name the specific vendor, invoice, event, pay period, billing period, or GL entry. Never say "costs increased."
2. DOLLAR AMOUNT — Use K notation: $10K not $10,000. Include the specific amount driving the variance.
3. TIMING — Use M.YY date format: "3.26" = March 2026, "1.26-3.26" = Jan–Mar 2026. State whether the expense is a reversal, timing difference, or genuine overage.
4. OUTLOOK — If under-budget due to timing, say "expected to be paid in [M.YY]" or "expected to hit GL next period." For ongoing variances, flag for budget revision.

LENGTH: 1 sentence preferred. 2 sentences max.

TIER 2 ACCOUNTS: Write a short flag phrase only (5–10 words). Example: "Timing — invoices expected 4.26."

NEVER:
- Start a comment with the account name or a number.
- Use "favorable" or "unfavorable" — the reader sees the sign already.
- Say "activity via accrual" or echo an accrual entry description as the cause.
- Comment on Total, Subtotal, NOI, or Net Income rows.
- Speculate beyond what GL data shows. If unclear, write: "Cause requires Finance review — see GL detail."

REAL EXAMPLES (match this style exactly):
- "Variance due to over accrual for 2.26 gas bills reversing in 3.26."
- "High accrual per most recent water and sewer bill for 3.25-9.25."
- "Variance due to large RM OT payment in 3.13 pay period leading to a high accrual. Additionally, 2025 RM Bonus accrual was increased by $10K in alignment with actuals which will hit GL next period."
- "Variance due to mechanical service contract for $8.4K budgeted in 3.26 and expected to be paid in 4.26."
- "Variance is due to a series of unbudgeted critical HVAC repairs. These stemmed from freezing temperatures freezing pipes and cracking cooling tower."
- "High snow removal costs and accruals in 1.26-3.26 due to several snowstorms hitting the region."
- "Variance due to $17K reduction in 2025 admin bonus accrual to reflect actuals which will hit the GL next period."
- "Variance due to $30K mgmt. fee and subsidy accrual for Craft Food Hall that has been dropped as it is no longer paid."
- "RET lower than budget due to successful appeal to Town of Lexington by GRP."
- "Variance due to $4K budgeted for Garage Cameras which have not been purchased."
- "Variance due to expense for JLL MU expected to hit GL next period and budgeted in 3.26 for $5.6K."

DIRECTION CONVENTION (for your internal reasoning — do NOT write these words):
- Revenue over budget = favorable to NOI
- Expense over budget = unfavorable to NOI
- Revenue under budget = unfavorable to NOI
- Expense under budget = favorable to NOI
"""


def _build_api_prompt(accounts_data: List[dict], period: str, property_name: str) -> str:
    """Build the user-turn prompt with all variance data for the API call."""
    lines = [
        f"Property: {property_name}",
        f"Period: {period}",
        "",
        "Generate MTD and YTD variance comments for the accounts listed below.",
        "For each account, return a JSON object with keys:",
        '  "account_code", "mtd_comment", "ytd_comment"',
        "",
        "Rules:",
        "- Tier 1 accounts: write a full 1–2 sentence comment per the standards.",
        "- Tier 2 accounts: write a short flag phrase only (5–10 words).",
        "- If MTD and YTD tell the same story, the comments may be similar but must",
        "  each stand alone (different readers may see only one column).",
        "- If YTD variance is smaller % than MTD, note it as a timing difference.",
        "",
        "Return ONLY a JSON array. No markdown fences. No extra text.",
        "",
    ]

    for a in accounts_data:
        code = a['account_code']
        name = a['account_name']
        mtd_tier = a.get('mtd_tier', 'tier_3')
        ytd_tier = a.get('ytd_tier', 'tier_3')

        lines.append(f"--- ACCOUNT {code}: {name} ---")
        lines.append(f"MTD Tier: {mtd_tier.upper()}")
        lines.append(
            f"MTD Actual: ${a.get('mtd_actual', 0):,.2f}  "
            f"Budget: ${a.get('mtd_budget', 0):,.2f}  "
            f"Variance: ${a.get('mtd_var', 0):+,.2f}  "
            f"({a.get('mtd_pct', 0):+.1f}%)  "
            f"[{a.get('mtd_noi', '')}]"
        )
        lines.append(f"YTD Tier: {ytd_tier.upper()}")
        lines.append(
            f"YTD Actual: ${a.get('ytd_actual', 0):,.2f}  "
            f"Budget: ${a.get('ytd_budget', 0):,.2f}  "
            f"Variance: ${a.get('ytd_var', 0):+,.2f}  "
            f"({a.get('ytd_pct', 0):+.1f}%)"
        )

        annual = a.get('annual_budget')
        if annual:
            lines.append(f"Annual Budget: ${annual:,.2f}")

        # Kardin context
        kardin = a.get('kardin', {})
        if kardin.get('descriptions'):
            lines.append("Budget intent (from Kardin):")
            for d in kardin['descriptions'][:5]:
                m_key = f"M{a.get('period_month', 1)}"
                # Try to find the matching monthly amount
                lines.append(f"  · {d}")
        if kardin.get('seasonality_note'):
            lines.append(f"Seasonality: {kardin['seasonality_note']}")

        # Account-specific behavioral context
        behavior = ACCOUNT_CONTEXT.get(code)
        if behavior:
            lines.append(f"Known behavior: {behavior}")

        # GL transactions (top 12 by absolute net)
        gl = a.get('gl', {})
        txns = gl.get('transactions', [])
        if txns:
            sorted_txns = sorted(txns, key=lambda t: abs(t['net']), reverse=True)
            lines.append(f"GL transactions ({len(txns)} total, top shown):")
            for t in sorted_txns[:12]:
                rev_flag = " [REVERSAL]" if t.get('is_reversal') else ''
                lines.append(
                    f"  {t['date']}  {t['description'][:55]}  "
                    f"${t['net']:+,.2f}{rev_flag}"
                )
            if gl.get('has_reversals'):
                lines.append("  ^ Contains accrual reversals — treat as timing, not true spend.")
        else:
            lines.append("GL transactions: none found for this period.")

        lines.append("")

    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════
# 5. DATA-DRIVEN FALLBACK COMMENT
# ══════════════════════════════════════════════════════════════

# Prefixes that identify pipeline-generated accrual entries (not real vendor spend)
_ACCRUAL_DESC_PREFIXES = (
    'accrual', 'budget gap', 'invoice proration', 'historical pattern',
    'recurring pattern', 'prepaid', 'manual je', 'je-', 'journal entry',
    'sup-', 'mgt-', 'catch-up', 'payroll bonus',
)


def _is_accrual_entry(desc: str) -> bool:
    """Return True if this GL description looks like a pipeline-generated accrual."""
    d = desc.lower().strip()
    if ':reversal' in d or 'auto-reversal' in d or 'auto reversal' in d:
        return True
    return any(d.startswith(p) for p in _ACCRUAL_DESC_PREFIXES)


def _period_code(period: str) -> str:
    """Convert 'Mar-2026' or 'Mar 2026' to '3.26'."""
    if not period:
        return ''
    for sep in ('-', ' '):
        parts = period.strip().split(sep)
        if len(parts) >= 2:
            month_num = MONTH_MAP.get(parts[0][:3].title(), 0)
            year = parts[-1].strip()
            if month_num and len(year) == 4 and year.isdigit():
                return f"{month_num}.{year[2:]}"
    return period.strip()


def _next_period_code(period: str) -> str:
    """Return the next month as M.YY (e.g. 'Mar 2026' → '4.26')."""
    if not period:
        return ''
    for sep in ('-', ' '):
        parts = period.strip().split(sep)
        if len(parts) >= 2:
            month_num = MONTH_MAP.get(parts[0][:3].title(), 0)
            year = parts[-1].strip()
            if month_num and len(year) == 4 and year.isdigit():
                nm = month_num + 1
                ny = int(year)
                if nm > 12:
                    nm = 1
                    ny += 1
                return f"{nm}.{str(ny)[2:]}"
    return ''


def _prev_period_code(period: str) -> str:
    """Return the previous month as M.YY (e.g. 'Mar 2026' → '2.26')."""
    if not period:
        return ''
    for sep in ('-', ' '):
        parts = period.strip().split(sep)
        if len(parts) >= 2:
            month_num = MONTH_MAP.get(parts[0][:3].title(), 0)
            year = parts[-1].strip()
            if month_num and len(year) == 4 and year.isdigit():
                pm = month_num - 1
                py = int(year)
                if pm < 1:
                    pm = 12
                    py -= 1
                return f"{pm}.{str(py)[2:]}"
    return ''


def _k_fmt(amount: float) -> str:
    """Format a dollar amount in K notation: '$8.4K', '$10K', '$125K'."""
    a = abs(float(amount))
    if a < 1_000:
        return f"${a:,.0f}"
    k = a / 1_000.0
    rounded = round(k, 1)
    if rounded == int(rounded):
        return f"${int(rounded)}K"
    return f"${rounded}K"


def _data_driven_comment(account_name: str, var_dollar: float, var_pct: float,
                          tier: str, gl: dict, kardin: dict,
                          is_revenue: bool, period: str) -> str:
    """
    Generate a GRP-style variance comment without the API.

    Format: "Variance due to [specific cause]." with M.YY dates and $K amounts.
    Matches the style of human-reviewed GRP comments:
      - "Variance due to over accrual for 2.26 gas bills reversing in 3.26."
      - "Variance due to mechanical service contract for $8.4K budgeted in 3.26
         and expected to be paid in 4.26."
    """
    cur_code = _period_code(period)
    nxt_code = _next_period_code(period)
    prv_code = _prev_period_code(period)
    abs_var = abs(var_dollar)

    # Directional flags
    expense_over  = (not is_revenue) and (var_dollar > 0)  # actual > budget
    expense_under = (not is_revenue) and (var_dollar < 0)  # actual < budget

    # ── Tier 2: short flag phrase only (5–10 words) ───────────
    if tier == 'tier_2':
        if gl.get('has_reversals'):
            return f"Timing — prior-period accrual reversal."
        if expense_under and nxt_code:
            return f"Timing — expected to be paid in {nxt_code}."
        kardin_descs = kardin.get('descriptions') or []
        kd = next((d.strip()[:40] for d in kardin_descs if d.strip()), '')
        if kd and expense_under:
            return f"Timing — {kd} expected {nxt_code or 'next period'}."
        return "See GL detail for variance support."

    # ── Classify GL transactions ───────────────────────────────
    txns = gl.get('transactions', [])
    reversal_txns = [t for t in txns if t.get('is_reversal')]
    real_txns     = [t for t in txns if not t.get('is_reversal')]

    reversal_gross = sum(abs(t['net']) for t in reversal_txns)
    real_gross     = sum(abs(t['net']) for t in real_txns)
    total_gross    = reversal_gross + real_gross

    # Pure reversal: reversal transactions dominate and real spend is small
    pure_reversal = (
        gl.get('has_reversals')
        and total_gross > 0
        and (reversal_gross / total_gross) > 0.60
        and real_gross < abs_var * 0.50
    )

    # Filter vendor summary — strip pipeline-generated accrual entries
    real_vendors: List[Tuple[str, float]] = []
    for k, v in gl.get('vendor_summary', {}).items():
        if not _is_accrual_entry(k):
            real_vendors.append((k.strip()[:60], abs(v['total'])))
    real_vendors.sort(key=lambda x: x[1], reverse=True)

    top_vendor     = real_vendors[0][0] if real_vendors else None
    top_vendor_amt = real_vendors[0][1] if real_vendors else 0.0

    # First meaningful Kardin description (skip accrual-type entries)
    kardin_desc: Optional[str] = None
    for d in (kardin.get('descriptions') or []):
        d = d.strip()
        if d and len(d) > 3 and not _is_accrual_entry(d):
            kardin_desc = d[:60]
            break

    # ── PATTERN 1: Prior-period accrual reversing this month ──
    if pure_reversal:
        subj = kardin_desc.lower() if kardin_desc else account_name.lower()
        if prv_code and cur_code:
            return (f"Variance due to over accrual for {prv_code} "
                    f"{subj} reversing in {cur_code}.")
        return (f"Variance due to prior-period accrual reversal "
                f"for {account_name.lower()}.")

    # ── PATTERN 2: Expense under budget — timing / not yet incurred ──
    if expense_under:
        amt_str = _k_fmt(abs_var)
        if top_vendor and top_vendor_amt >= 500:
            if cur_code and nxt_code:
                return (f"Variance due to {top_vendor} ({amt_str}) budgeted in "
                        f"{cur_code} and expected to be paid in {nxt_code}.")
            return f"Variance due to {top_vendor} of {amt_str} not yet incurred."
        if kardin_desc:
            if cur_code and nxt_code:
                return (f"Variance due to {kardin_desc} of {amt_str} budgeted in "
                        f"{cur_code} and expected to be paid in {nxt_code}.")
            return f"Variance due to {kardin_desc} of {amt_str} not yet incurred."
        if cur_code and nxt_code:
            return (f"Variance due to expense budgeted in {cur_code} "
                    f"not yet incurred; expected to be paid in {nxt_code}.")
        return "Variance due to timing — expense not yet incurred this period."

    # ── PATTERN 3: Over budget — GL vendor identified ──────────
    if top_vendor and top_vendor_amt >= 500:
        amt_str = _k_fmt(top_vendor_amt)
        if cur_code:
            return f"Variance due to {top_vendor} of {amt_str} in {cur_code}."
        return f"Variance due to {top_vendor} of {amt_str}."

    # ── PATTERN 4: Over budget — fall back to Kardin intent ───
    if kardin_desc:
        amt_str = _k_fmt(abs_var)
        if cur_code:
            return f"Variance due to {kardin_desc} of {amt_str} in {cur_code}."
        return f"Variance due to {kardin_desc} of {amt_str}."

    # ── PATTERN 5: Seasonal note ───────────────────────────────
    if kardin.get('is_seasonal') and kardin.get('seasonality_note'):
        return kardin['seasonality_note']

    # ── PATTERN 6: Finance review required ────────────────────
    return "Cause requires Finance review — see GL detail."


# ══════════════════════════════════════════════════════════════
# 6. NOI DIRECTION HELPER
# ══════════════════════════════════════════════════════════════

def _noi_direction(account_code: str, variance: float) -> str:
    """Return 'favorable' or 'unfavorable' relative to NOI."""
    if _is_revenue(account_code):
        return 'favorable' if variance > 0 else 'unfavorable'
    return 'unfavorable' if variance > 0 else 'favorable'


def _is_revenue(account_code: str) -> bool:
    return is_revenue_account(account_code)


# ══════════════════════════════════════════════════════════════
# 7. MAIN ENTRY POINT — GRP VARIANCE COMMENTS
# ══════════════════════════════════════════════════════════════

def generate_variance_comments_grp(
    budget_rows: List[dict],
    gl_parsed,
    kardin_records: List[dict],
    period: str = '',
    property_name: str = 'Revolution Labs Owner, LLC',
    api_key: Optional[str] = None,
    je_adjustments: Optional[Dict[str, float]] = None,
) -> Dict[str, dict]:
    """
    Generate MTD and YTD variance comments for all budget comparison rows
    that meet GRP's tier thresholds.

    Args:
        budget_rows:     Parsed budget comparison rows (from yardi_budget_comparison parser).
                         Each row must have: account_code, account_name,
                         ptd_actual, ptd_budget, ytd_actual, ytd_budget,
                         annual (optional), ptd_variance, ptd_variance_pct,
                         ytd_variance, ytd_variance_pct.
        gl_parsed:       Parsed GL data (from yardi_gl parser).
        kardin_records:  Parsed Kardin budget records (from kardin_budget parser).
        period:          Period string (e.g. "Apr 2026").
        property_name:   Property display name.
        api_key:         Anthropic API key. If omitted, uses data-driven fallback.
        je_adjustments:  Optional dict {account_code: signed_delta} representing the
                         net effect of all pipeline JEs on each income-statement account's
                         PTD actual. Revenue accounts: positive = more revenue earned.
                         Expense accounts: positive = more expense incurred.
                         When provided, actuals are adjusted before tier classification so
                         comments reflect the projected final-close position, not the
                         pre-close Yardi snapshot.

    Returns:
        Dict keyed by account_code:
          {
            'account_name': str,
            'mtd_tier': 'tier_1' | 'tier_2' | 'tier_3',
            'ytd_tier': 'tier_1' | 'tier_2' | 'tier_3',
            'mtd_comment': str,
            'ytd_comment': str,
            'mtd_actual': float,  'mtd_budget': float,
            'ytd_actual': float,  'ytd_budget': float,
          }
    """
    # Pass 2 safety guard — GL is already final; je_adjustments must not be used.
    if je_adjustments is not None:
        raise ValueError(
            "je_adjustments must not be passed to generate_variance_comments_grp() in Pass 2. "
            "The GL is already final after the close — read actuals directly from GL."
        )

    # Determine reporting month from period string (e.g. "Apr 2026" → 4)
    period_month = 1
    if period:
        for name, num in MONTH_MAP.items():
            if name in period:
                period_month = num
                break

    # ── Pass 1: classify tiers and build enrichment context ──
    accounts_data: List[dict] = []
    all_results: Dict[str, dict] = {}

    for row in budget_rows:
        code = str(row.get('account_code', '') or '').strip()
        name = str(row.get('account_name', '') or '').strip()

        if _is_skip_row(code, name):
            continue

        mtd_actual = float(row.get('ptd_actual', 0) or 0)
        mtd_budget = float(row.get('ptd_budget', 0) or 0)
        ytd_actual = float(row.get('ytd_actual', 0) or 0)
        ytd_budget = float(row.get('ytd_budget', 0) or 0)
        annual = row.get('annual')
        annual = float(annual) if annual else None

        # Apply pro-forma JE adjustments — shift actuals to projected final-close
        # position so comments are written against the numbers that will be posted,
        # not the mid-close Yardi snapshot.
        if je_adjustments and code in je_adjustments:
            _delta = je_adjustments[code]
            mtd_actual += _delta
            ytd_actual += _delta   # current-period JEs affect MTD and YTD equally

        mtd_tier, mtd_var, mtd_pct = classify_tier(mtd_actual, mtd_budget)
        ytd_tier, ytd_var, ytd_pct = classify_tier(ytd_actual, ytd_budget)

        # Only process accounts where at least one period is Tier 1 or Tier 2
        if mtd_tier == 'tier_3' and ytd_tier == 'tier_3':
            all_results[code] = {
                'account_name': name,
                'mtd_tier': 'tier_3', 'ytd_tier': 'tier_3',
                'mtd_comment': '', 'ytd_comment': '',
                'mtd_actual': mtd_actual, 'mtd_budget': mtd_budget,
                'ytd_actual': ytd_actual, 'ytd_budget': ytd_budget,
            }
            continue

        # Build enrichment context
        kardin = build_kardin_enrichment(kardin_records, code, period_month)
        gl = build_gl_context(gl_parsed, code)

        entry = {
            'account_code': code,
            'account_name': name,
            'period_month': period_month,
            'mtd_tier': mtd_tier,
            'ytd_tier': ytd_tier,
            'mtd_actual': mtd_actual,
            'mtd_budget': mtd_budget,
            'mtd_var': mtd_var,
            'mtd_pct': mtd_pct,
            'mtd_noi': _noi_direction(code, mtd_var),
            'ytd_actual': ytd_actual,
            'ytd_budget': ytd_budget,
            'ytd_var': ytd_var,
            'ytd_pct': ytd_pct,
            'annual_budget': annual or kardin.get('annual_budget'),
            'kardin': kardin,
            'gl': gl,
            'is_revenue': _is_revenue(code),
        }
        accounts_data.append(entry)

    if not accounts_data:
        return all_results

    # ── Pass 2: generate comments ─────────────────────────────
    api_fallback_reason: Optional[str] = None   # set if API was requested but failed
    if api_key:
        comments_map, api_fallback_reason = _call_api(
            accounts_data, period, property_name, api_key
        )
    else:
        comments_map = _generate_data_driven(accounts_data, period)

    # ── Pass 3: assemble final results ────────────────────────
    for entry in accounts_data:
        code = entry['account_code']
        api_result = comments_map.get(code, {})

        all_results[code] = {
            'account_name': entry['account_name'],
            'mtd_tier': entry['mtd_tier'],
            'ytd_tier': entry['ytd_tier'],
            'mtd_comment': api_result.get('mtd_comment', ''),
            'ytd_comment': api_result.get('ytd_comment', ''),
            'mtd_actual': entry['mtd_actual'],
            'mtd_budget': entry['mtd_budget'],
            'ytd_actual': entry['ytd_actual'],
            'ytd_budget': entry['ytd_budget'],
            # Visible signal for downstream consumers (app.py, qc_engine.py)
            '_api_fallback': api_fallback_reason,
        }

    return all_results


def _call_api(
    accounts_data: List[dict],
    period: str,
    property_name: str,
    api_key: str,
) -> Tuple[Dict[str, dict], Optional[str]]:
    """
    Call Claude API and parse JSON response.

    Returns:
        (comments_map, fallback_reason)
        fallback_reason is None on success; a human-readable string when the API
        was requested but failed and data-driven mode was used instead.
    """
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        prompt = _build_api_prompt(accounts_data, period, property_name)

        message = client.messages.create(
            model='claude-sonnet-4-6',
            max_tokens=4096,
            system=_SYSTEM_PROMPT,
            messages=[{'role': 'user', 'content': prompt}],
        )

        raw = message.content[0].text.strip()

        # Strip markdown fences if present
        if raw.startswith('```'):
            raw = raw.split('\n', 1)[1] if '\n' in raw else raw[3:]
            if raw.endswith('```'):
                raw = raw[:-3]
            raw = raw.strip()

        items = json.loads(raw)
        return (
            {
                item['account_code']: {
                    'mtd_comment': item.get('mtd_comment', ''),
                    'ytd_comment': item.get('ytd_comment', ''),
                }
                for item in items
                if 'account_code' in item
            },
            None,   # success — no fallback
        )

    except ImportError:
        reason = (
            'Anthropic SDK not installed — variance commentary generated from '
            'data-driven templates, not AI. Install `anthropic` to enable API mode.'
        )
    except Exception as exc:
        reason = (
            f'Anthropic API call failed ({type(exc).__name__}: {exc}) — '
            f'variance commentary generated from data-driven templates, not AI. '
            f'Review API key and connectivity before sign-off.'
        )

    return _generate_data_driven(accounts_data, period), reason


def _generate_data_driven(accounts_data: List[dict],
                           period: str) -> Dict[str, dict]:
    """Data-driven fallback — no API call."""
    result: Dict[str, dict] = {}
    for entry in accounts_data:
        code = entry['account_code']
        name = entry['account_name']
        rev = entry['is_revenue']

        mtd_comment = ''
        if entry['mtd_tier'] != 'tier_3':
            mtd_comment = _data_driven_comment(
                name, entry['mtd_var'], entry['mtd_pct'],
                entry['mtd_tier'], entry['gl'], entry['kardin'],
                rev, period,
            )

        ytd_comment = ''
        if entry['ytd_tier'] != 'tier_3':
            ytd_comment = _data_driven_comment(
                name, entry['ytd_var'], entry['ytd_pct'],
                entry['ytd_tier'], entry['gl'], entry['kardin'],
                rev, period,
            )

        result[code] = {
            'mtd_comment': mtd_comment,
            'ytd_comment': ytd_comment,
        }
    return result


# ══════════════════════════════════════════════════════════════
# 8. EXCEL WRITE-BACK
# ══════════════════════════════════════════════════════════════

def write_comments_to_budget_comparison(
    input_path: str,
    output_path: str,
    comments: Dict[str, dict],
    mtd_col: int = 12,   # Excel column L (1-indexed)
    ytd_col: int = 13,   # Excel column M (1-indexed)
    data_start_row: int = 6,
    account_col: int = 1,  # Excel column A
) -> None:
    """
    Write variance comments into the budget comparison Excel file.

    Args:
        input_path:      Path to the source budget comparison .xlsx
        output_path:     Path to write the annotated output file
        comments:        Dict from generate_variance_comments_grp()
        mtd_col:         Excel column number for MTD notes (default 12 = col L)
        ytd_col:         Excel column number for YTD notes (default 13 = col M)
        data_start_row:  First data row (default 6, after 5 header rows)
        account_col:     Column containing account codes (default 1 = col A)
    """
    _comment_font = Font(name='Tahoma', size=10)
    _comment_align = Alignment(wrap_text=True, vertical='top')

    wb = load_workbook(input_path)
    ws = wb.active

    for row_num in range(data_start_row, ws.max_row + 1):
        code_cell = ws.cell(row=row_num, column=account_col)
        code = str(code_cell.value or '').strip()

        if not code or code not in comments:
            continue

        entry = comments[code]

        mtd_text = entry.get('mtd_comment', '')
        ytd_text = entry.get('ytd_comment', '')

        # Skip Tier 3 (empty strings already handle this, but be explicit)
        mtd_tier = entry.get('mtd_tier', 'tier_3')
        ytd_tier = entry.get('ytd_tier', 'tier_3')

        if mtd_text or mtd_tier in ('tier_1', 'tier_2'):
            cell = ws.cell(row=row_num, column=mtd_col)
            cell.value = mtd_text
            cell.font = _comment_font
            cell.alignment = _comment_align

        if ytd_text or ytd_tier in ('tier_1', 'tier_2'):
            cell = ws.cell(row=row_num, column=ytd_col)
            cell.value = ytd_text
            cell.font = _comment_font
            cell.alignment = _comment_align

    wb.save(output_path)


# ══════════════════════════════════════════════════════════════
# 9. BACKWARDS-COMPATIBLE WRAPPER (existing pipeline interface)
# ══════════════════════════════════════════════════════════════

def generate_variance_comments(engine_result, api_key: Optional[str] = None) -> List[dict]:
    """
    Backwards-compatible wrapper for the existing pipeline.

    Returns list of dicts with keys:
      account_code, account_name, ptd_actual, ptd_budget,
      variance_amount, variance_pct, comment, method
    """
    gl_data = engine_result.parsed.get('gl')
    budget_data = engine_result.parsed.get('budget_comparison')
    kardin_data = engine_result.parsed.get('kardin_budget') or []
    period = getattr(engine_result, 'period', '') or ''
    prop = getattr(engine_result, 'property_name', 'Revolution Labs Owner, LLC') or ''

    # Normalize budget_data to list of dicts
    budget_rows: List[dict] = []
    if budget_data:
        if isinstance(budget_data, list):
            budget_rows = budget_data
        elif hasattr(budget_data, 'line_items'):
            budget_rows = [
                {
                    'account_code': getattr(item, 'account_code', ''),
                    'account_name': getattr(item, 'account_name', ''),
                    'ptd_actual': getattr(item, 'ptd_actual', 0),
                    'ptd_budget': getattr(item, 'ptd_budget', 0),
                    'ytd_actual': getattr(item, 'ytd_actual', 0),
                    'ytd_budget': getattr(item, 'ytd_budget', 0),
                    'annual': getattr(item, 'annual', None),
                    'ptd_variance': getattr(item, 'ptd_variance', 0),
                    'ptd_variance_pct': getattr(item, 'ptd_variance_pct', 0),
                    'ytd_variance': getattr(item, 'ytd_variance', 0),
                    'ytd_variance_pct': getattr(item, 'ytd_variance_pct', 0),
                }
                for item in budget_data.line_items
            ]

    if not budget_rows:
        # Fall back to engine_result.budget_variances if no budget comparison
        budget_rows = [
            {
                'account_code': v.get('account_code', ''),
                'account_name': v.get('account_name', ''),
                'ptd_actual': v.get('ptd_actual', 0),
                'ptd_budget': v.get('ptd_budget', 0),
                'ytd_actual': v.get('ytd_actual', 0),
                'ytd_budget': v.get('ytd_budget', 0),
                'annual': v.get('annual'),
            }
            for v in (engine_result.budget_variances or [])
        ]

    comments_map = generate_variance_comments_grp(
        budget_rows=budget_rows,
        gl_parsed=gl_data,
        kardin_records=kardin_data,
        period=period,
        property_name=prop,
        api_key=api_key,
    )

    # Detect whether API was requested but silently fell back to data-driven
    fallback_reasons = {
        entry.get('_api_fallback')
        for entry in comments_map.values()
        if entry.get('_api_fallback')
    }
    api_fallback_reason: Optional[str] = next(iter(fallback_reasons), None)

    if api_key and api_fallback_reason:
        # API was requested but failed — mark method so UI and reports surface it
        method = 'data-driven (API FALLBACK)'
    elif api_key:
        method = 'api'
    else:
        method = 'data-driven'

    results = []
    for code, entry in comments_map.items():
        if entry.get('mtd_tier') == 'tier_3' and entry.get('ytd_tier') == 'tier_3':
            continue
        results.append({
            'account_code': code,
            'account_name': entry['account_name'],
            'ptd_actual': entry['mtd_actual'],
            'ptd_budget': entry['mtd_budget'],
            'variance_amount': entry['mtd_actual'] - entry['mtd_budget'],
            'variance_pct': (
                (entry['mtd_actual'] - entry['mtd_budget']) / abs(entry['mtd_budget']) * 100
                if entry['mtd_budget'] else 0
            ),
            'comment': entry.get('mtd_comment', ''),
            'ytd_comment': entry.get('ytd_comment', ''),
            'mtd_tier': entry.get('mtd_tier', 'tier_3'),
            'ytd_tier': entry.get('ytd_tier', 'tier_3'),
            'method': method,
            # Carry the fallback reason so app.py / qc_engine can surface it
            'api_fallback_reason': api_fallback_reason,
        })

    return results

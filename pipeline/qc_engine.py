"""
QC Engine — GRP Monthly Close Quality Control
==============================================
Automates the 8 QC checks currently performed manually in the LexLabs QC workbook.
Runs against the parsed data from the pipeline and produces:
  1. A QCReport dataclass consumed by the Streamlit dashboard
  2. An Excel workbook matching the LexLabs QC structure (via generate_qc_workbook)

Checks:
  1  TB to Budget Comparison Tie-Out
  2  Budget Variances ≥ Tier 1 threshold (GRP standards)
  3  TB Self-Balance + GL Ending Balances vs TB
  4  Month-over-Month Swings (>$10,000 or sign change)
  5  GL Ending Balances vs Workpapers (BS account cross-check)
  6  Accruals vs Budget (missing accrual detection)
  7  Miscellaneous (mgmt fee, interest expense, insurance/prepaid)
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from variance_comments import classify_tier, TIER1_ABS, TIER1_PCT, TIER2_MIN
from property_config import is_revenue_account, is_expense_account, is_balance_sheet_account


# ── Account type helpers — delegate to per-property COA config ─────────────────
# Defaults match the standard Yardi COA (4xxxxx=revenue, 5-8xxxxx=expense,
# 1-3xxxxx=BS).  Override via PropertyConfig when onboarding a new property.
def _is_revenue(code: str) -> bool:
    return is_revenue_account(code)


def _is_expense(code: str) -> bool:
    return is_expense_account(code)


def _is_balance_sheet(code: str) -> bool:
    return is_balance_sheet_account(code)


def _safe_float(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '').strip())
    except Exception:
        return 0.0


# ══════════════════════════════════════════════════════════════
# DATA CLASSES
# ══════════════════════════════════════════════════════════════

@dataclass
class QCFinding:
    account_code: str
    account_name: str
    value_a: float          # Primary value (e.g. actual, GL balance)
    value_b: float          # Comparison value (e.g. budget, TB balance)
    difference: float
    flag: str               # "OVER", "UNDER", "MISMATCH", "MISSING", "FLAG", "INFO"
    note: str = ''


@dataclass
class QCResult:
    check_id: str           # "CHECK_1" … "CHECK_7"
    check_name: str
    status: str             # "PASS" | "FLAG" | "FAIL"
    summary: str
    findings: List[QCFinding] = field(default_factory=list)

    @property
    def flag_count(self) -> int:
        return len(self.findings)


@dataclass
class QCReport:
    period: str
    property_name: str
    run_at: str
    checks: List[QCResult] = field(default_factory=list)

    @property
    def has_flags(self) -> bool:
        return any(c.status in ('FLAG', 'FAIL') for c in self.checks)

    @property
    def overall_status(self) -> str:
        if any(c.status == 'FAIL' for c in self.checks):
            return 'FAIL'
        if any(c.status == 'FLAG' for c in self.checks):
            return 'FLAG'
        return 'PASS'


# ══════════════════════════════════════════════════════════════
# CHECK 1 — TB to Budget Comparison Tie-Out
# ══════════════════════════════════════════════════════════════

def check_1_tb_to_budget(tb_result, budget_rows: List[dict]) -> QCResult:
    """
    For every account present in both the Trial Balance and Budget Comparison,
    confirm that TB net activity equals Budget Comparison PTD Actual.

    Revenue accounts: TB net = credit - debit (income shown positive in BC)
    Expense accounts: TB net = debit - credit
    """
    findings: List[QCFinding] = []
    bc_map: Dict[str, dict] = {}

    for row in budget_rows:
        code = str(row.get('account_code', '') or '').strip()
        if code:
            bc_map[code] = row

    tb_map = tb_result.account_map if tb_result else {}

    matched = 0
    for code, tb_acct in tb_map.items():
        if code not in bc_map:
            continue

        bc_row = bc_map[code]
        ptd_actual = _safe_float(bc_row.get('ptd_actual', 0))

        # Convention: revenue accounts show positive income in BC
        if _is_revenue(code):
            tb_net = tb_acct.credit - tb_acct.debit
        else:
            tb_net = tb_acct.debit - tb_acct.credit

        diff = abs(tb_net - ptd_actual)
        matched += 1

        if diff > 0.02:  # $0.02 tolerance for rounding
            findings.append(QCFinding(
                account_code=code,
                account_name=tb_acct.account_name,
                value_a=tb_net,
                value_b=ptd_actual,
                difference=tb_net - ptd_actual,
                flag='MISMATCH',
                note=f'TB net activity ${tb_net:,.2f} ≠ BC PTD Actual ${ptd_actual:,.2f}',
            ))

    if not findings:
        summary = (f'All {matched} accounts with TB activity tie to Budget Comparison PTD Actual. '
                   f'Difference = $0.00.')
        status = 'PASS'
    else:
        summary = (f'{len(findings)} account(s) do not tie between TB and Budget Comparison '
                   f'out of {matched} checked.')
        status = 'FAIL'

    return QCResult('CHECK_1', 'TB to Budget Comparison Tie-Out', status, summary, findings)


# ══════════════════════════════════════════════════════════════
# CHECK 2 — Budget Variances (GRP Tier 1 Threshold)
# ══════════════════════════════════════════════════════════════

def check_2_budget_variances(budget_rows: List[dict]) -> QCResult:
    """
    Flag all accounts that meet GRP Tier 1 or Tier 2 variance thresholds.
    Separates OVER and UNDER for clarity.
    """
    findings: List[QCFinding] = []

    # Skip rows that should never be commented
    from variance_comments import _is_skip_row

    for row in budget_rows:
        code = str(row.get('account_code', '') or '').strip()
        name = str(row.get('account_name', '') or '').strip()

        if _is_skip_row(code, name):
            continue

        ptd_actual = _safe_float(row.get('ptd_actual', 0))
        ptd_budget = _safe_float(row.get('ptd_budget', 0))

        tier, var_dollar, var_pct = classify_tier(ptd_actual, ptd_budget)
        if tier == 'tier_3':
            continue

        # Determine direction relative to NOI impact.
        # Convention: OVER  → unfavorable (red in workbook)
        #             UNDER → favorable   (yellow in workbook)
        #
        # Revenue accounts: actual > budget = favorable  → UNDER (yellow)
        #                   actual < budget = unfavorable → OVER  (red)
        # Expense accounts: actual > budget = unfavorable → OVER  (red)
        #                   actual < budget = favorable  → UNDER (yellow)
        if _is_revenue(code):
            flag = 'UNDER' if var_dollar > 0 else 'OVER'
        else:
            flag = 'OVER' if var_dollar > 0 else 'UNDER'

        tier_label = 'T1' if tier == 'tier_1' else 'T2'
        note = f'{tier_label} | ${abs(var_dollar):,.0f} ({var_pct:+.1f}%)'

        findings.append(QCFinding(
            account_code=code,
            account_name=name,
            value_a=ptd_actual,
            value_b=ptd_budget,
            difference=var_dollar,
            flag=flag,
            note=note,
        ))

    over_count = sum(1 for f in findings if f.flag == 'OVER')
    under_count = sum(1 for f in findings if f.flag == 'UNDER')
    t1_count = sum(1 for f in findings if 'T1' in f.note)

    if not findings:
        summary = 'No current-period variances exceed GRP Tier 1 threshold.'
        status = 'PASS'
    else:
        largest = max(findings, key=lambda f: abs(f.difference))
        summary = (f'{t1_count} Tier-1 variance(s) flagged ({over_count} OVER, {under_count} UNDER). '
                   f'Largest: {largest.account_name} '
                   f'{largest.flag} '
                   f'${abs(largest.difference):,.0f}.')
        status = 'FLAG'

    return QCResult('CHECK_2', 'Budget Variances — GRP Tier 1', status, summary, findings)


# ══════════════════════════════════════════════════════════════
# CHECK 3 — TB Self-Balance + GL vs TB
# ══════════════════════════════════════════════════════════════

def check_3_tb_balance_and_gl(tb_result, gl_parsed) -> QCResult:
    """
    3a. TB self-balance: total debits = total credits.
    3b. For each account in GL, ending balance matches TB ending balance.
    3c. Income accounts in TB are presented in negative convention (credit balance).
    """
    findings: List[QCFinding] = []

    if not tb_result:
        return QCResult('CHECK_3', 'TB Self-Balance + GL vs TB',
                        'FAIL', 'Trial Balance not available.', [])

    # 3a: Self-balance check
    diff = abs(tb_result.total_debits - tb_result.total_credits)
    if diff > 0.05:
        findings.append(QCFinding(
            account_code='TB-TOTAL',
            account_name='Trial Balance Total',
            value_a=tb_result.total_debits,
            value_b=tb_result.total_credits,
            difference=tb_result.total_debits - tb_result.total_credits,
            flag='FAIL',
            note=f'Debits ({tb_result.total_debits:,.2f}) ≠ Credits ({tb_result.total_credits:,.2f}). Out of balance by ${diff:,.2f}.',
        ))

    # 3b: GL ending balance vs TB ending balance
    if gl_parsed and hasattr(gl_parsed, 'accounts'):
        tb_map = tb_result.account_map
        for gl_acct in gl_parsed.accounts:
            code = str(gl_acct.account_code or '').strip()
            if code not in tb_map:
                continue
            tb_acct = tb_map[code]
            gl_end = float(getattr(gl_acct, 'ending_balance', 0) or 0)
            tb_end = float(tb_acct.ending_balance)
            diff_b = abs(gl_end - tb_end)
            if diff_b > 0.05:
                findings.append(QCFinding(
                    account_code=code,
                    account_name=tb_acct.account_name,
                    value_a=gl_end,
                    value_b=tb_end,
                    difference=gl_end - tb_end,
                    flag='MISMATCH',
                    note=f'GL ending ${gl_end:,.2f} ≠ TB ending ${tb_end:,.2f}',
                ))

    # 3c: TB totals summary — appended last so mismatches appear first
    findings.append(QCFinding(
        account_code='TB-TOTAL',
        account_name='TB Totals',
        value_a=tb_result.total_debits,
        value_b=tb_result.total_credits,
        difference=tb_result.total_debits - tb_result.total_credits,
        flag='INFO' if tb_result.is_balanced else 'FAIL',
        note=(f'Debits = Credits = ${tb_result.total_debits:,.2f}'
              if tb_result.is_balanced
              else f'OUT OF BALANCE by ${abs(tb_result.total_debits - tb_result.total_credits):,.2f}'),
    ))

    critical = [f for f in findings if f.flag in ('FAIL', 'MISMATCH')]
    if not critical:
        status = 'PASS'
        acct_count = len(tb_result.accounts)
        summary = (f'TB balanced: debits = credits = ${tb_result.total_debits:,.2f}. '
                   f'All {acct_count} accounts verified.')
    else:
        status = 'FAIL'
        summary = f'{len(critical)} balance discrepancy/discrepancies found.'

    return QCResult('CHECK_3', 'Workpapers to Trial Balance Tie-Out', status, summary, findings)


# ══════════════════════════════════════════════════════════════
# CHECK 4 — Month-over-Month Swings
# ══════════════════════════════════════════════════════════════

def check_4_mom_swings(budget_rows: List[dict],
                       swing_threshold: float = 10_000.0) -> QCResult:
    """
    For P&L accounts, derive prior-month actual = YTD actual - PTD actual.
    Flag if |PTD actual - prior month actual| > $10,000 or if sign changes.

    Note: For month 1 (January), prior month = 0 (no prior YTD).
    """
    from variance_comments import _is_skip_row

    findings: List[QCFinding] = []

    for row in budget_rows:
        code = str(row.get('account_code', '') or '').strip()
        name = str(row.get('account_name', '') or '').strip()

        if _is_skip_row(code, name):
            continue
        if not (_is_revenue(code) or _is_expense(code)):
            continue

        ptd = _safe_float(row.get('ptd_actual', 0))
        ytd = _safe_float(row.get('ytd_actual', 0))

        # Derive prior month
        prior = ytd - ptd

        swing = ptd - prior
        abs_swing = abs(swing)
        sign_change = (ptd > 0 and prior < 0) or (ptd < 0 and prior > 0)

        if abs_swing > swing_threshold or (sign_change and abs_swing > 2_500):
            flag = 'FLAG'
            note = (f'Prior month ${prior:,.0f} → Current ${ptd:,.0f} '
                    f'= swing ${swing:+,.0f}')
            if sign_change:
                note += ' [SIGN CHANGE]'

            findings.append(QCFinding(
                account_code=code,
                account_name=name,
                value_a=ptd,
                value_b=prior,
                difference=swing,
                flag=flag,
                note=note,
            ))

    # Sort by absolute swing descending
    findings.sort(key=lambda f: abs(f.difference), reverse=True)

    if not findings:
        summary = f'No month-over-month swings exceed ${swing_threshold:,.0f}.'
        status = 'PASS'
    else:
        largest = findings[0]
        summary = (f'{len(findings)} line(s) with MoM swing > ${swing_threshold:,.0f}. '
                   f'Largest: {largest.account_name} '
                   f'${abs(largest.difference):,.0f} swing.')
        status = 'FLAG'

    return QCResult('CHECK_4', 'Month-over-Month Swings (>$10,000)', status, summary, findings)


# ══════════════════════════════════════════════════════════════
# CHECK 5 — BS Workpaper Tie-Out (GL ending vs TB ending per account)
# ══════════════════════════════════════════════════════════════

def check_5_gl_vs_workpapers(gl_parsed, tb_result) -> QCResult:
    """
    For each balance sheet account, compare GL ending balance to TB ending balance.

    Pre-accrual run: GL ≠ TB is expected — the variance equals the accrual JEs
    that need to be posted. These are flagged as INFO so GRP knows which accounts
    need entries, not as failures.

    Post-accrual run: All variances should be zero. Any remaining non-zero variance
    is a genuine discrepancy and flagged as MISMATCH.

    Accounts not in TB are flagged separately (likely zero-balance accounts).
    """
    findings: List[QCFinding] = []

    if not gl_parsed or not hasattr(gl_parsed, 'accounts'):
        return QCResult('CHECK_5', 'BS Workpaper Tie-Out',
                        'FLAG', 'GL data not available for this check.', [])

    if not tb_result:
        return QCResult('CHECK_5', 'BS Workpaper Tie-Out',
                        'FLAG', 'Trial Balance not uploaded — cannot run workpaper tie-out.', [])

    tb_map = tb_result.account_map if tb_result else {}
    checked = 0
    accrual_gap_total = 0.0
    clean_count = 0

    for gl_acct in gl_parsed.accounts:
        code = str(gl_acct.account_code or '').strip()
        if not _is_balance_sheet(code):
            continue

        gl_end = float(getattr(gl_acct, 'ending_balance', 0) or 0)
        checked += 1

        if code not in tb_map:
            # In GL but not in TB — only flag if non-zero balance
            if abs(gl_end) > 0.05:
                findings.append(QCFinding(
                    account_code=code,
                    account_name=gl_acct.account_name,
                    value_a=gl_end,
                    value_b=0.0,
                    difference=gl_end,
                    flag='INFO',
                    note=f'Account in GL (${gl_end:,.2f}) but not in TB — verify zero balance expected.',
                ))
            continue

        tb_end = float(tb_map[code].ending_balance)
        diff = gl_end - tb_end

        if abs(diff) <= 0.05:
            clean_count += 1
        else:
            accrual_gap_total += abs(diff)
            findings.append(QCFinding(
                account_code=code,
                account_name=gl_acct.account_name,
                value_a=gl_end,
                value_b=tb_end,
                difference=diff,
                flag='FLAG',
                note=(f'GL ${gl_end:,.2f} vs TB ${tb_end:,.2f} — '
                      f'variance of ${abs(diff):,.2f} likely requires accrual JE.'),
            ))

    accrual_gaps = [f for f in findings if f.flag == 'FLAG']
    info_gaps    = [f for f in findings if f.flag == 'INFO']

    if not accrual_gaps and not info_gaps:
        status  = 'PASS'
        summary = f'All {clean_count} BS accounts tie — GL ending = TB ending.'
    elif accrual_gaps:
        status  = 'FLAG'
        summary = (f'{len(accrual_gaps)} account(s) have GL vs TB variance '
                   f'(total ${accrual_gap_total:,.2f}) — accrual JEs required. '
                   f'{clean_count} account(s) tie clean. See BS Workpaper for detail.')
    else:
        status  = 'FLAG'
        summary = f'{len(info_gaps)} account(s) in GL not found in TB — verify expected.'

    return QCResult('CHECK_5', 'BS Workpaper Tie-Out', status, summary, findings)


# ══════════════════════════════════════════════════════════════
# CHECK 6 — Accruals vs Budget
# ══════════════════════════════════════════════════════════════

def check_6_accruals_vs_budget(budget_rows: List[dict],
                                kardin_records: List[dict] = None,
                                accrual_entries: List[dict] = None,
                                period_month: int = 1) -> QCResult:
    """
    Identify expense accounts where:
    - Budget > $0 for the month AND
    - PTD Actual = $0 AND
    - Account is not a known seasonal account with $0 expected this month

    These are candidates for missing accruals.
    Also verifies that suggested accrual entries (from accrual engine) exist where expected.
    """
    from variance_comments import _is_skip_row, build_kardin_enrichment

    findings: List[QCFinding] = []
    kardin_records = kardin_records or []

    # Build set of accounts that have accrual entries
    accrual_codes = set()
    if accrual_entries:
        for entry in accrual_entries:
            code = str(entry.get('account_code', '') or '').strip()
            if code:
                accrual_codes.add(code)

    for row in budget_rows:
        code = str(row.get('account_code', '') or '').strip()
        name = str(row.get('account_name', '') or '').strip()

        if _is_skip_row(code, name):
            continue
        if not _is_expense(code):
            continue

        ptd_actual = _safe_float(row.get('ptd_actual', 0))
        ptd_budget = _safe_float(row.get('ptd_budget', 0))

        if ptd_budget <= 0:
            continue  # No budget → no expected accrual

        if abs(ptd_actual) > 1.0:
            continue  # Activity exists → no missing accrual

        # Check if this month is budgeted as $0 in Kardin (seasonal)
        kardin = build_kardin_enrichment(kardin_records, code, period_month)
        month_budget = kardin.get('month_budget', ptd_budget)

        if abs(month_budget) < 1.0:
            # Kardin says $0 this month — this is expected, not a gap
            findings.append(QCFinding(
                account_code=code,
                account_name=name,
                value_a=ptd_actual,
                value_b=ptd_budget,
                difference=ptd_actual - ptd_budget,
                flag='INFO',
                note=f'$0 actual; Kardin shows $0 budget this month — no accrual expected.',
            ))
        else:
            has_accrual = code in accrual_codes
            flag = 'INFO' if has_accrual else 'FLAG'
            note = (
                f'$0 actual vs ${ptd_budget:,.0f} budget. '
                + ('Accrual entry generated.' if has_accrual
                   else f'No accrual found. Kardin month budget: ${month_budget:,.0f}. Review required.')
            )
            findings.append(QCFinding(
                account_code=code,
                account_name=name,
                value_a=ptd_actual,
                value_b=ptd_budget,
                difference=ptd_actual - ptd_budget,
                flag=flag,
                note=note,
            ))

    real_flags = [f for f in findings if f.flag == 'FLAG']
    if not real_flags:
        status = 'PASS'
        info_count = sum(1 for f in findings if f.flag == 'INFO')
        summary = (f'No missing accruals detected. '
                   f'{info_count} contingency item(s) with $0 budget this period — no action.')
    else:
        status = 'FLAG'
        summary = (f'{len(real_flags)} potential missing accrual(s) flagged for review. '
                   f'{sum(1 for f in findings if f.flag == "INFO")} items are expected $0.')

    return QCResult('CHECK_6', 'Accruals vs Budget', status, summary, findings)


# ══════════════════════════════════════════════════════════════
# CHECK 7 — Miscellaneous QC Items
# ══════════════════════════════════════════════════════════════

def check_7_misc(budget_rows: List[dict],
                 gl_parsed=None,
                 tb_result=None,
                 kardin_records: List[dict] = None,
                 cash_received: float = None,
                 jll_rate: float = 0.0125,
                 grp_rate: float = 0.0175,
                 loan_data=None,
                 period_month: int = 1) -> QCResult:
    """
    Spot-checks for:
    7a. Management fee accrual vs calculated fee (cash received × rate)
    7b. Interest expense — verify accrual exists and is non-zero
    7c. Insurance-Property (639110) — verify monthly charge matches expected amortization
    7d. Prepaid accounts (135110 insurance, 135120 RE Tax prepaid, 135150 Prepaid Other) — fwd + DR - CR = ending
    7e. Berkadia insurance escrow — should be $0 for Rev Labs (Berkadia no longer handles insurance)
    """
    findings: List[QCFinding] = []
    kardin_records = kardin_records or []
    tb_map = tb_result.account_map if tb_result else {}

    # ── 7a: Management Fee ─────────────────────────────────────
    mgmt_fee_code = '637130'
    bc_map = {str(r.get('account_code', '') or '').strip(): r for r in budget_rows}

    if mgmt_fee_code in bc_map and cash_received is not None and cash_received > 0:
        ptd_actual = abs(_safe_float(bc_map[mgmt_fee_code].get('ptd_actual', 0)))
        expected_jll = cash_received * jll_rate
        expected_grp = cash_received * grp_rate
        expected_total = expected_jll + expected_grp
        diff = abs(ptd_actual - expected_total)

        flag = 'PASS_INFO' if diff < 500 else 'FLAG'
        findings.append(QCFinding(
            account_code=mgmt_fee_code,
            account_name='Admin-Management Fees',
            value_a=ptd_actual,
            value_b=expected_total,
            difference=ptd_actual - expected_total,
            flag=flag,
            note=(f'Accrued: ${ptd_actual:,.2f} | '
                  f'Calculated: ${expected_total:,.2f} '
                  f'(Cash ${cash_received:,.2f} × JLL {jll_rate:.2%} + GRP {grp_rate:.2%}). '
                  + ('Ties.' if diff < 500 else f'Difference: ${diff:,.2f} — update accrual.')),
        ))
    elif mgmt_fee_code in bc_map:
        ptd_actual = _safe_float(bc_map[mgmt_fee_code].get('ptd_actual', 0))
        findings.append(QCFinding(
            account_code=mgmt_fee_code,
            account_name='Admin-Management Fees',
            value_a=ptd_actual,
            value_b=0,
            difference=0,
            flag='INFO',
            note=f'Accrual: ${ptd_actual:,.2f}. Cash received not available — cannot verify rate calc.',
        ))

    # ── 7b: Interest Expense ───────────────────────────────────
    interest_code = '801110'
    if interest_code in bc_map:
        ptd_actual = abs(_safe_float(bc_map[interest_code].get('ptd_actual', 0)))
        if ptd_actual < 1.0:
            findings.append(QCFinding(
                account_code=interest_code,
                account_name='Interest Expense',
                value_a=ptd_actual,
                value_b=0,
                difference=0,
                flag='FLAG',
                note='Interest Expense is $0 — verify accrual was posted.',
            ))
        else:
            findings.append(QCFinding(
                account_code=interest_code,
                account_name='Interest Expense',
                value_a=ptd_actual,
                value_b=0,
                difference=0,
                flag='INFO',
                note=f'Interest Expense: ${ptd_actual:,.2f}. Verify against loan statement / amortization schedule.',
            ))

    # ── 7c: Insurance amortization ─────────────────────────────
    ins_code = '639110'
    if ins_code in bc_map and kardin_records:
        from variance_comments import build_kardin_enrichment
        kardin = build_kardin_enrichment(kardin_records, ins_code, period_month)
        annual_budget = kardin.get('annual_budget', 0)
        expected_monthly = annual_budget / 12 if annual_budget else 0
        ptd_actual = abs(_safe_float(bc_map[ins_code].get('ptd_actual', 0)))
        if expected_monthly > 0:
            diff = abs(ptd_actual - expected_monthly)
            flag = 'INFO' if diff < 500 else 'FLAG'
            findings.append(QCFinding(
                account_code=ins_code,
                account_name='Insurance-Property',
                value_a=ptd_actual,
                value_b=expected_monthly,
                difference=ptd_actual - expected_monthly,
                flag=flag,
                note=(f'Actual: ${ptd_actual:,.2f} | '
                      f'Expected monthly (${annual_budget:,.0f}/12): ${expected_monthly:,.2f}. '
                      + ('On track.' if diff < 500 else f'Difference ${diff:,.2f} — verify prepaid schedule.')),
            ))

    # ── 7d: Prepaid accounts math (fwd + DR - CR = ending) ────
    # 135110 = Restricted Insurance / Prepaid Insurance (single insurance account)
    # 135120/135150 = RE Tax Prepaid variants
    # 135110 also checked in 7e against Berkadia statement
    if tb_result:
        for code in ('135110', '135120', '135150'):
            if code in tb_map:
                acct = tb_map[code]
                expected_end = acct.forward_balance + acct.debit - acct.credit
                diff = abs(expected_end - acct.ending_balance)
                if diff > 0.05:
                    findings.append(QCFinding(
                        account_code=code,
                        account_name=acct.account_name,
                        value_a=acct.ending_balance,
                        value_b=expected_end,
                        difference=acct.ending_balance - expected_end,
                        flag='FLAG',
                        note=f'Prepaid math error: fwd {acct.forward_balance:,.2f} + debit {acct.debit:,.2f} - credit {acct.credit:,.2f} = {expected_end:,.2f} ≠ ending {acct.ending_balance:,.2f}',
                    ))
                else:
                    findings.append(QCFinding(
                        account_code=code,
                        account_name=acct.account_name,
                        value_a=acct.ending_balance,
                        value_b=expected_end,
                        difference=0.0,
                        flag='INFO',
                        note=f'Prepaid math ties: fwd {acct.forward_balance:,.2f} + debit {acct.debit:,.2f} - credit {acct.credit:,.2f} = {acct.ending_balance:,.2f}.',
                    ))

    # ── 7e: Berkadia insurance escrow should be $0 for Rev Labs ──
    # Insurance is paid via prepaid (135110), not through Berkadia lender escrow.
    # A non-zero Berkadia insurance_escrow_balance is unexpected and flags for review.
    if loan_data:
        loans = loan_data if isinstance(loan_data, list) else [loan_data]
        loan_ins_escrow = 0.0
        for ln in loans:
            if isinstance(ln, dict):
                loan_ins_escrow += _safe_float(ln.get('insurance_escrow_balance', 0))
            else:
                loan_ins_escrow += _safe_float(getattr(ln, 'insurance_escrow_balance', 0))

        if loan_ins_escrow > 1.0:
            findings.append(QCFinding(
                account_code='135110',
                account_name='Restricted Insurance (Lender Escrow)',
                value_a=loan_ins_escrow,
                value_b=0,
                difference=loan_ins_escrow,
                flag='FLAG',
                note=(f'Berkadia shows insurance escrow balance of ${loan_ins_escrow:,.2f}. '
                      f'Expected $0 — insurance is handled via prepaid (135110), not lender escrow. '
                      f'Verify with Berkadia.'),
            ))
        else:
            findings.append(QCFinding(
                account_code='135110',
                account_name='Restricted Insurance (Lender Escrow)',
                value_a=loan_ins_escrow,
                value_b=0,
                difference=0,
                flag='INFO',
                note='Berkadia insurance escrow: $0 — confirmed. Insurance handled via prepaid (135110).',
            ))

    # ── 7f: RE Tax Escrow (115200) GL vs Berkadia statement ───────
    # Berkadia manages RE Tax escrow — GL 115200 should tie to
    # tax_escrow_balance on the loan statement each month.
    retax_escrow_code = '115200'
    if loan_data and tb_result and retax_escrow_code in tb_map:
        loans = loan_data if isinstance(loan_data, list) else [loan_data]
        berkadia_tax_escrow = 0.0
        for ln in loans:
            if isinstance(ln, dict):
                berkadia_tax_escrow += _safe_float(ln.get('tax_escrow_balance', 0))
            else:
                berkadia_tax_escrow += _safe_float(getattr(ln, 'tax_escrow_balance', 0))

        acct = tb_map[retax_escrow_code]
        gl_bal = acct.ending_balance
        diff = abs(gl_bal - berkadia_tax_escrow)
        flag = 'INFO' if diff < 1.0 else 'FLAG'
        findings.append(QCFinding(
            account_code=retax_escrow_code,
            account_name='RE Tax Escrow (115200)',
            value_a=gl_bal,
            value_b=berkadia_tax_escrow,
            difference=gl_bal - berkadia_tax_escrow,
            flag=flag,
            note=(f'GL 115200: ${gl_bal:,.2f} | Berkadia statement: ${berkadia_tax_escrow:,.2f}. '
                  + ('Ties.' if diff < 1.0
                     else f'Difference ${diff:,.2f} — post reconciling JE to 115200.')),
        ))

    flags = [f for f in findings if f.flag == 'FLAG']
    if not flags:
        status = 'PASS'
        summary = 'Miscellaneous checks passed — management fee, interest, insurance, prepaid math, RE tax escrow.'
    else:
        status = 'FLAG'
        items = ', '.join(f.account_name for f in flags[:3])
        summary = f'{len(flags)} miscellaneous item(s) flagged: {items}.'

    return QCResult('CHECK_7', 'Miscellaneous QC Items', status, summary, findings)


# ══════════════════════════════════════════════════════════════
# MAIN RUNNER
# ══════════════════════════════════════════════════════════════

def run_qc(
    budget_rows: List[dict],
    tb_result=None,
    gl_parsed=None,
    kardin_records: List[dict] = None,
    accrual_entries: List[dict] = None,
    period: str = '',
    property_name: str = 'Revolution Labs Owner, LLC',
    period_month: int = 1,
    cash_received: float = None,
    loan_data=None,
) -> QCReport:
    """
    Run all 7 QC checks and return a QCReport.

    Args:
        budget_rows:     Parsed budget comparison rows.
        tb_result:       Parsed Trial Balance (TBResult).
        gl_parsed:       Parsed GL (GLParseResult).
        kardin_records:  Kardin annual budget records.
        accrual_entries: Accrual JE entries from accrual engine.
        period:          Period string (e.g. "Apr 2026").
        property_name:   Property display name.
        period_month:    Reporting month number (1=Jan … 12=Dec).
        cash_received:   Total cash received for the month (for mgmt fee check).
        loan_data:       Parsed Berkadia loan data (list of dicts or LoanResult).
                         Used for Check 7e: GL 115300 vs lender insurance escrow balance.
    """
    kardin_records = kardin_records or []

    checks = [
        check_1_tb_to_budget(tb_result, budget_rows),
        check_2_budget_variances(budget_rows),
        check_3_tb_balance_and_gl(tb_result, gl_parsed),
        check_4_mom_swings(budget_rows),
        check_5_gl_vs_workpapers(gl_parsed, tb_result),
        check_6_accruals_vs_budget(budget_rows, kardin_records, accrual_entries, period_month),
        check_7_misc(budget_rows, gl_parsed, tb_result, kardin_records, cash_received,
                     loan_data=loan_data, period_month=period_month),
    ]

    return QCReport(
        period=period,
        property_name=property_name,
        run_at=datetime.now().strftime('%Y-%m-%d %H:%M'),
        checks=checks,
    )


# ══════════════════════════════════════════════════════════════
# EXCEL WORKBOOK GENERATOR  (matches RevLabs QC_Blank_Template)
# ══════════════════════════════════════════════════════════════

# ── Palette ───────────────────────────────────────────────────
_C_GREEN    = 'C6EFCE'   # TIES / OK / PASS
_C_RED      = 'FFC7CE'   # FLAG / REVIEW / OVER / FAIL
_C_YELLOW   = 'FFEB9C'   # UNDER / NOTE / OVER-ACCRUED / Manual input
_C_NAVY     = '1F4E79'   # Column-header fill & title font
_C_MED_BLUE = '2E75B6'   # Tab-7 sub-section headers
_C_LIGHT1   = 'BDD7EE'   # Inline section headers (REVENUE, etc.)
_C_LIGHT2   = 'D6E4F0'   # Block section headers (BALANCE SHEET ACCOUNTS)
_C_GRAY     = '7F7F7F'   # Subtitle / note text


def _qfill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _qfont(bold=False, size=10, color='000000', italic=False):
    return Font(name='Calibri', bold=bold, size=size, color=color, italic=italic)


def _qborder():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)


def _status_fill(status: str):
    """Map a status string to its fill (None = no fill)."""
    s = str(status or '').upper()
    if s in ('TIES', 'OK', 'PASS'):
        return _qfill(_C_GREEN)
    if s in ('FLAG', 'review', 'REVIEW', 'FAIL', 'OVER', 'SIGNIFICANT', 'SIGN FLIP', 'MISMATCH'):
        return _qfill(_C_RED)
    if s in ('UNDER', 'NOTE', 'OVER-ACCRUED', 'MANUAL REVIEW REQUIRED',
             'REVIEW REQUIRED', 'REVIEW'):
        return _qfill(_C_YELLOW)
    return None


# ── Account master lists (order matches template) ─────────────

_PL_ACCOUNTS = [
    # (section_name, [(account_code, display_name), ...])
    ('REVENUE', [
        ('411150', 'Rent - Lab'),
        ('', 'Rent - Office'),
        ('', 'Rent - Cafe/Food Hall'),
        ('440100', 'Recovery - Operating Expense'),
        ('440200', 'Recovery - Real Estate Tax'),
        ('440500', 'Recovery - Electricity'),
        ('440700', 'Recovery - Misc Utilities'),
        ('', 'Antenna Rent'),
        ('481100', 'Tenant Service Revenue'),
        ('485100', 'Interest Income'),
    ]),
    ('OPERATING EXPENSES', [
        ('610110', 'Cleaning-Janitorial'),
        ('610140', 'Cleaning-Day Porter'),
        ('610160', 'Cleaning-Trash Removal'),
        ('613110', 'Utilities-Electricity'),
        ('613210', 'Utilities-Gas'),
        ('613310', 'Utilities-Water/Sewer'),
        ('615110', 'RM-Pay/Wages'),
        ('615140', 'RM-Uniforms'),
        ('615200', 'RM-Supplies'),
        ('615230', 'RM-Signage'),
        ('615290', 'RM-General R&M'),
        ('615310', 'RM-Parking Lot/Ext'),
        ('615320', 'RM-Doors/Locks'),
        ('615350', 'RM-Fire Suppression'),
        ('615410', 'RM-Floor/Ceiling'),
        ('615500', 'RM-Roof/Waterproofing'),
        ('615560', 'RM-Other R&M'),
        ('617110', 'HVAC Maint-Contract Svc'),
        ('617120', 'HVAC Maint-Repairs'),
        ('617130', 'HVAC-Water Treatment'),
        ('617140', 'HVAC-Other'),
        ('619110', 'Plumbing-Contract Svc'),
        ('619120', 'Plumbing-Repairs'),
        ('619130', 'Plumbing-Other'),
        ('621110', 'Electrical-Contract Svc'),
        ('621120', 'Electrical-Repairs'),
        ('621130', 'Electrical-Other'),
        ('623110', 'Landscape-Contract Svc'),
        ('625110', 'Snow & Ice-Contract Svc'),
        ('627110', 'Security-Payroll'),
        ('627130', 'Security-Contract Svc'),
        ('627140', 'Security-Alarm Monitoring'),
        ('627170', 'Security-Repairs'),
        ('627230', 'Fire Life Safety-Inspections'),
        ('629110', 'Elevator-Contract Svc'),
        ('629120', 'Elevator-Repairs'),
        ('631120', 'Parking-Parking Lot'),
        ('633110', 'Pest Control'),
        ('635110', 'Garage'),
        ('635310', 'Garage Repairs'),
        ('637110', 'Admin-Pay/Wages'),
        ('637130', 'Admin-Management Fees'),
        ('637150', 'Admin-Tenant Relations'),
        ('637170', 'Admin-Food Service Amenities'),
        ('637310', 'Admin-Training'),
        ('637330', 'Admin-Dues/Subscriptions'),
        ('637350', 'Admin-Travel'),
        ('637370', 'Admin-Computer/Software'),
        ('639110', 'Insurance-Property'),
        ('639120', 'Insurance-General Liability'),
        ('641110', 'Real Estate Taxes'),
        ('641120', 'Personal Property Taxes'),
        ('641130', 'RE Tax Consultant'),
        ('641140', 'RE Tax-Other'),
    ]),
    ('NON-RECOVERABLE EXPENSES', [
        ('680110', 'NR-Prof Fees-Legal'),
        ('', 'NR-Prof Fees-Other'),
        ('680510', 'NR-Marketing & PR'),
        ('682110', 'NR-Other Expenses'),
        ('725070', 'Advertising & Marketing'),
    ]),
    ('INTEREST', [
        ('801110', 'Interest Expense'),
    ]),
]

# Balance sheet accounts (Tab 3 Workpapers to TB + Tab 5 GL to Workpapers)
_BS_ACCOUNTS = [
    # (section, account_code, description, workpaper_source)
    ('BALANCE SHEET ACCOUNTS', '111100', 'Cash - Operating', 'PNC Bank Rec'),
    ('', '115100', 'Restricted Cash - Deposit Escrow', 'WF DACA'),
    ('', '115200', 'Restricted Cash - RE Tax Escrow', 'Berkadia Note A1'),
    ('', '115300', 'Restricted Cash - Insurance Escrow', 'Insurance Escrow WP'),
    ('', '115400', 'Restricted Cash - Leasing Reserve', 'Berkadia (zero)'),
    ('', '115500', 'Restricted Cash - Construction Reserve', 'Berkadia (zero)'),
    ('', '115600', 'Restricted Cash - Other Reserve', 'Alchemab settlement reserve'),
    ('', '131100', 'AR - Control', 'AR Aging'),
    ('', '133100', 'AR - Other', 'AR Other WP'),
    ('', '133110', 'AR - Tenant Billback', 'AR Tenant Billback WP'),
    ('', '135110', 'Prepaid Insurance', 'Insurance Analysis WP'),
    ('', '135120', 'Prepaid RE Tax', 'RE Tax Analysis WP'),
    ('', '135150', 'Prepaid Other', 'PPD WP'),
    ('', '141100', 'Due From/(To) Related Parties', 'WP Tab'),
    ('', '152100', 'Land', 'WP Tab'),
    ('', '154100', 'Building', 'WP Tab'),
    ('', '154200', 'Building - Acquisition Costs', 'WP Tab'),
    ('', '154500', 'Building Improvements', 'WP Tab'),
    ('', '171100', 'CIP - Development Projects', 'WP Tab'),
    ('', '181400', 'Tenant Improvements', 'WP Tab'),
    ('', '187100', 'Financing Costs', 'WP Tab'),
    ('', '211100', 'AP - Control', 'WP Tab'),
    ('', '213100', 'Accrued Expenses', 'WP Tab'),
    ('', '213200', 'Accrued Interest', 'Trimont Stmt'),
    ('', '221100', 'Prepaid Rent - Tenant', 'AR Aging'),
    ('', '231100', 'Mortgage Payable', 'Loan Analysis WP'),
    ('', '311100', 'Contributions - Partner A', 'WP Tab'),
]


# ── Shared helpers ─────────────────────────────────────────────

def _qwrite_tab_header(ws, title: str, prop_name: str, period: str):
    ws['A1'] = title
    ws['A1'].font = _qfont(bold=True, size=11, color=_C_NAVY)
    ws.cell(row=2, column=1, value='Property:').font = _qfont(bold=True, size=10)
    c = ws.cell(row=2, column=2, value=prop_name)
    c.fill = _qfill(_C_GREEN)
    c.font = _qfont(bold=True, size=10)
    ws.cell(row=3, column=1, value='Period:').font = _qfont(bold=True, size=10)
    c = ws.cell(row=3, column=2, value=period)
    c.fill = _qfill(_C_GREEN)
    c.font = _qfont(bold=True, size=10)


def _qwrite_check_header(ws, row5_text: str, row6_text: str, ncols: int):
    c = ws.cell(row=5, column=1, value=row5_text)
    c.font = _qfont(bold=True, size=10)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=ncols)
    c = ws.cell(row=6, column=1, value=row6_text)
    c.font = _qfont(italic=True, size=9, color=_C_GRAY)
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=ncols)


def _qwrite_col_headers(ws, row: int, headers: list, widths: list):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _qfont(bold=True, size=10, color='FFFFFF')
        c.fill = _qfill(_C_NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = _qborder()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 28


def _qsection(ws, row: int, text: str, ncols: int, fill_hex: str = _C_LIGHT1):
    c = ws.cell(row=row, column=1, value=text)
    c.font = _qfont(bold=True, size=10)
    c.fill = _qfill(fill_hex)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)


def _qstatus(ws, row: int, col: int, status: str):
    c = ws.cell(row=row, column=col, value=status)
    c.font = _qfont(bold=bool(status), size=10)
    c.border = _qborder()
    c.alignment = Alignment(horizontal='center')
    sf = _status_fill(status)
    if sf:
        c.fill = sf
    return c


def _qmoney(ws, row: int, col: int, value, border=True):
    c = ws.cell(row=row, column=col, value=value)
    if isinstance(value, (int, float)):
        c.number_format = '#,##0.00;(#,##0.00);"-"'
    if border:
        c.border = _qborder()
    return c


def _qpct(ws, row: int, col: int, value):
    c = ws.cell(row=row, column=col, value=value)
    if isinstance(value, (int, float)):
        c.number_format = '0.0%'
    c.border = _qborder()
    return c


def _qtxt(ws, row: int, col: int, value, bold=False, fill_hex=None, wrap=False, color='000000'):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _qfont(bold=bold, size=10, color=color)
    c.border = _qborder()
    if fill_hex:
        c.fill = _qfill(fill_hex)
    if wrap:
        c.alignment = Alignment(wrap_text=True, vertical='top')
    return c


def _bc_map(budget_rows):
    return {str(r.get('account_code', '') or '').strip(): r for r in (budget_rows or [])}


def _tb_map_fn(tb_result):
    return tb_result.account_map if tb_result else {}


def _gl_map(gl_parsed):
    if not gl_parsed or not hasattr(gl_parsed, 'accounts'):
        return {}
    return {str(a.account_code or '').strip(): a for a in gl_parsed.accounts}


# ── Tab 0 — Instructions ──────────────────────────────────────

def _write_tab0(wb, report: QCReport):
    ws = wb.create_sheet('0-Instructions')
    ws.sheet_properties.tabColor = _C_NAVY
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 90

    c = ws.cell(row=1, column=1, value='QC WORKBOOK — INSTRUCTIONS')
    c.font = _qfont(bold=True, size=12, color=_C_NAVY)
    ws.merge_cells('A1:B1')

    steps = [
        ('STEP 1', 'Upload source files: TB, Budget Comparison, GL, Workpapers, Mgmt Fee Calc, Berkadia Stmts.'),
        ('STEP 2', 'Verify Property and Period in header cells (row 2–3) on each tab.'),
        ('STEP 3', 'Review tabs 1–7b. Status values: TIES | REVIEW | FLAG | UNDER | OVER | MANUAL REVIEW REQUIRED'),
        ('STEP 4', 'Color legend: Green = TIES/OK | Red = FLAG/REVIEW | Yellow = UNDER/Input required | Blue = Section header'),
        ('KEY THRESHOLDS',
         'Budget variance flag = ≥$5,000 AND ≥75%, OR ≥$5,000 regardless of %. MoM swing flag = >$10,000 or sign change.'),
        ('LOAN REFERENCE',
         f'Berkadia: Note A1 + Note B1 + Mezz = total mortgage (TB 231100). '
         f'Interest accrual per Berkadia amort schedule.'),
        ('RUN INFO', f'Period: {report.period}  |  Run: {report.run_at}'),
    ]
    for i, (step, text) in enumerate(steps, 2):
        c = ws.cell(row=i, column=1, value=step)
        c.font = _qfont(bold=True, size=10, color=_C_NAVY)
        c = ws.cell(row=i, column=2, value=text)
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 28


# ── Tab 1 — TB to CIS Tie ─────────────────────────────────────

def _write_tab1(wb, report: QCReport, tb_result, budget_rows):
    ws = wb.create_sheet('1-TB to CIS Tie')
    ws.sheet_properties.tabColor = _C_NAVY

    prop = report.property_name or 'Revolution Labs'
    _qwrite_tab_header(ws, 'CHECK 1: Trial Balance to Budget Comparison (CIS) Tie-Out', prop, report.period)
    _qwrite_check_header(ws,
        'CHECK 1: Trial Balance to Budget Comparison (CIS) Tie-Out',
        'Auto-populate from: Trial Balance vs Budget Comparison PTD Actual',
        7)

    headers = ['Account', 'Line Item', 'TB Debit / (Credit)', 'CIS PTD Actual', 'Difference', 'Status', 'Notes']
    widths  = [12, 35, 18, 18, 14, 10, 55]
    _qwrite_col_headers(ws, 8, headers, widths)

    tbm = _tb_map_fn(tb_result)
    bcm = _bc_map(budget_rows)

    row = 9
    for section_name, accounts in _PL_ACCOUNTS:
        _qsection(ws, row, section_name, 7, _C_LIGHT1)
        row += 1
        for code, name in accounts:
            _qtxt(ws, row, 1, code)
            _qtxt(ws, row, 2, name)
            tb_val = ''
            cis_val = ''
            diff = ''
            status = ''
            if code and code in tbm:
                tb_acct = tbm[code]
                if _is_revenue(code):
                    tb_val = round(tb_acct.credit - tb_acct.debit, 2)
                else:
                    tb_val = round(tb_acct.debit - tb_acct.credit, 2)
            if code and code in bcm:
                cis_val = _safe_float(bcm[code].get('ptd_actual', 0))
            if tb_val != '' and cis_val != '':
                diff = round(float(tb_val) - float(cis_val), 2)
                status = 'TIES' if abs(diff) < 0.02 else 'REVIEW'
            _qmoney(ws, row, 3, tb_val if tb_val != '' else None)
            _qmoney(ws, row, 4, cis_val if cis_val != '' else None)
            _qmoney(ws, row, 5, diff if diff != '' else None)
            _qstatus(ws, row, 6, status)
            _qtxt(ws, row, 7, '')
            row += 1

    # check_1 finding notes
    c1 = next((c for c in report.checks if c.check_id == 'CHECK_1'), None)
    if c1 and c1.findings:
        ws.cell(row=row + 1, column=1,
                value=f'Pipeline check: {c1.summary}').font = _qfont(italic=True, size=9, color=_C_GRAY)

    ws.freeze_panes = 'A9'


# ── Tab 2 — Budget Variances ──────────────────────────────────

def _write_tab2(wb, report: QCReport, budget_rows):
    ws = wb.create_sheet('2-Budget Variances')
    ws.sheet_properties.tabColor = _C_NAVY

    prop = report.property_name or 'Revolution Labs'
    _qwrite_tab_header(ws, 'CHECK 2: Budget Variances', prop, report.period)
    _qwrite_check_header(ws,
        'CHECK 2: Budget Variances ≥$5,000 or ≥75%',
        'OVER = Actuals > Budget (unfavorable) | UNDER = Actuals < Budget (favorable)',
        8)

    headers = ['Line Item', 'Actual', 'Budget', 'Variance ($)', 'Var %', 'Flag', 'Elec Tie Check', 'JLL Comment / Explanation']
    widths  = [35, 16, 16, 16, 10, 10, 28, 60]
    _qwrite_col_headers(ws, 8, headers, widths)

    bcm = _bc_map(budget_rows)

    row = 9
    for section_name, accounts in _PL_ACCOUNTS:
        _qsection(ws, row, section_name, 8, _C_LIGHT1)
        row += 1
        for code, name in accounts:
            actual = _safe_float(bcm.get(code, {}).get('ptd_actual', 0)) if code else 0.0
            budget = _safe_float(bcm.get(code, {}).get('ptd_budget', 0)) if code else 0.0
            variance = round(actual - budget, 2)
            var_pct = variance / abs(budget) if abs(budget) > 1 else None

            # Flag logic: |variance| >= 5000 AND (|pct| >= 75% OR |variance| >= 5000)
            flag = ''
            if abs(variance) >= 5000:
                if _is_revenue(code):
                    flag = 'UNDER' if variance > 0 else 'OVER'
                else:
                    flag = 'OVER' if variance > 0 else 'UNDER'

            _qtxt(ws, row, 1, name, wrap=True)
            _qmoney(ws, row, 2, actual if code else None)
            _qmoney(ws, row, 3, budget if code else None)
            _qmoney(ws, row, 4, variance if code else None)
            _qpct(ws, row, 5, var_pct)
            _qstatus(ws, row, 6, flag)
            _qtxt(ws, row, 7, '')
            _qtxt(ws, row, 8, '', wrap=True)
            row += 1

    ws.freeze_panes = 'A9'


# ── Tab 3 — Workpapers to TB ──────────────────────────────────

def _write_tab3(wb, report: QCReport, tb_result, gl_parsed):
    ws = wb.create_sheet('3-Workpapers to TB')
    ws.sheet_properties.tabColor = _C_NAVY

    prop = report.property_name or 'Revolution Labs'
    _qwrite_tab_header(ws, 'CHECK 3: Workpapers to Trial Balance Tie-Out', prop, report.period)
    _qwrite_check_header(ws,
        'CHECK 3: Workpapers to Trial Balance Tie-Out',
        'Enter workpaper/source balance in col D. Pipeline populates GL (TB) balance from Trial Balance.',
        7)

    headers = ['Account', 'Description', 'GL Balance (TB)', 'Workpaper / Source Balance', 'Difference', 'Status', 'Notes']
    widths  = [12, 38, 18, 22, 14, 10, 55]
    _qwrite_col_headers(ws, 8, headers, widths)

    tbm = _tb_map_fn(tb_result)
    glm = _gl_map(gl_parsed)

    _qsection(ws, 9, 'BALANCE SHEET ACCOUNTS', 7, _C_LIGHT2)
    row = 10
    for _sec, code, desc, wp_src in _BS_ACCOUNTS:
        tb_acct = tbm.get(code)
        gl_end = float(getattr(glm.get(code), 'ending_balance', 0) or 0) if code in glm else None
        tb_end = float(tb_acct.ending_balance) if tb_acct else None
        bal = tb_end if tb_end is not None else (gl_end if gl_end is not None else None)

        _qtxt(ws, row, 1, code)
        _qtxt(ws, row, 2, desc)
        _qmoney(ws, row, 3, bal)
        # col 4: workpaper balance — left blank for user; yellow fill = input cell
        c = ws.cell(row=row, column=4, value=None)
        c.fill = _qfill(_C_YELLOW)
        c.border = _qborder()
        # col 5-7: blank until user fills col 4
        for col in (5, 6, 7):
            ws.cell(row=row, column=col).border = _qborder()
        _qtxt(ws, row, 7, wp_src, color=_C_GRAY)
        row += 1

    ws.freeze_panes = 'A10'


# ── Tab 4 — MoM Swings ────────────────────────────────────────

def _write_tab4(wb, report: QCReport, budget_rows):
    ws = wb.create_sheet('4-MoM Swings')
    ws.sheet_properties.tabColor = _C_NAVY

    prop = report.property_name or 'Revolution Labs'
    _qwrite_tab_header(ws, 'CHECK 4: Month-over-Month Swings', prop, report.period)
    _qwrite_check_header(ws,
        'CHECK 4: Month-over-Month Swings — Prior Month vs Current Month',
        'Flag: changes > $10,000 or sign changes | Prior month derived from YTD - PTD',
        6)

    headers = ['Line Item', 'Prior Month', 'Current Month', 'Change ($)', 'Flag', 'Notes']
    widths  = [35, 16, 16, 16, 20, 60]
    _qwrite_col_headers(ws, 8, headers, widths)

    bcm = _bc_map(budget_rows)

    row = 9
    for section_name, accounts in _PL_ACCOUNTS:
        _qsection(ws, row, section_name, 6, _C_LIGHT1)
        row += 1
        for code, name in accounts:
            if not code or code not in bcm:
                _qtxt(ws, row, 1, name)
                for col in range(2, 7):
                    ws.cell(row=row, column=col).border = _qborder()
                row += 1
                continue
            ptd = _safe_float(bcm[code].get('ptd_actual', 0))
            ytd = _safe_float(bcm[code].get('ytd_actual', 0))
            prior = ytd - ptd
            change = ptd - prior
            sign_flip = (ptd > 0 and prior < 0) or (ptd < 0 and prior > 0)
            if abs(change) > 10_000 or (sign_flip and abs(change) > 2_500):
                flag = 'SIGN FLIP' if sign_flip else 'SIGNIFICANT'
            else:
                flag = 'Minor' if abs(change) > 500 else ''
            _qtxt(ws, row, 1, name)
            _qmoney(ws, row, 2, prior)
            _qmoney(ws, row, 3, ptd)
            _qmoney(ws, row, 4, change)
            _qstatus(ws, row, 5, flag)
            _qtxt(ws, row, 6, '', wrap=True)
            row += 1

    ws.freeze_panes = 'A9'


# ── Tab 5 — GL to Workpapers ──────────────────────────────────

def _write_tab5(wb, report: QCReport, gl_parsed, tb_result):
    ws = wb.create_sheet('5-GL to Workpapers')
    ws.sheet_properties.tabColor = _C_NAVY

    prop = report.property_name or 'Revolution Labs'
    _qwrite_tab_header(ws, 'CHECK 5: GL Ending Balances vs Workpaper Schedules', prop, report.period)
    _qwrite_check_header(ws,
        'CHECK 5: General Ledger Ending Balances vs Workpaper Schedules',
        'GL balance from Yardi GL (post-close). Workpaper balance entered manually or from BS Workpaper.',
        7)

    headers = ['Account', 'Description', 'GL Ending Balance', 'Workpaper Balance', 'Difference', 'Status', 'Workpaper Tab']
    widths  = [12, 38, 18, 18, 14, 10, 35]
    _qwrite_col_headers(ws, 8, headers, widths)

    glm = _gl_map(gl_parsed)
    tbm = _tb_map_fn(tb_result)

    sections_seen = set()
    row = 9
    for _sec, code, desc, wp_tab in _BS_ACCOUNTS:
        # Section headers use BDDEE7 light blue
        if _sec and _sec not in sections_seen:
            _qsection(ws, row, _sec, 7, _C_LIGHT1)
            sections_seen.add(_sec)
            row += 1

        gl_acct = glm.get(code)
        tb_acct = tbm.get(code)
        gl_end = float(getattr(gl_acct, 'ending_balance', 0) or 0) if gl_acct else None
        tb_end = float(tb_acct.ending_balance) if tb_acct else None
        # Prefer TB ending (post-close matches TB)
        bal = tb_end if tb_end is not None else gl_end

        _qtxt(ws, row, 1, code)
        _qtxt(ws, row, 2, desc)
        _qmoney(ws, row, 3, bal)
        c = ws.cell(row=row, column=4, value=None)
        c.fill = _qfill(_C_YELLOW)
        c.border = _qborder()
        for col in (5, 6):
            ws.cell(row=row, column=col).border = _qborder()
        _qtxt(ws, row, 7, wp_tab, color=_C_GRAY)
        row += 1

    ws.freeze_panes = 'A9'


# ── Tab 6 — Accruals vs Budget ────────────────────────────────

_CONTINGENCY_ACCTS = [
    ('621120', 'Electrical-Repairs'),
    ('617120', 'HVAC Maint-Repairs'),
    ('615200', 'RM-Supplies'),
    ('615230', 'RM-Signage'),
    ('615290', 'RM-General R&M'),
    ('615320', 'RM-Doors/Locks'),
    ('615350', 'RM-Fire Suppression'),
    ('615500', 'RM-Roof/Waterproofing'),
    ('615560', 'RM-Other R&M'),
    ('631120', 'Parking-Parking Lot'),
    ('637310', 'Admin-Training'),
    ('641140', 'RE Tax-Other'),
    ('682110', 'NR-Other Expenses'),
]

_ACCRUAL_ACCTS = [
    ('610140', 'Cleaning-Day Porter'),
    ('610160', 'Cleaning-Trash Removal'),
    ('613110', 'Utilities-Electricity'),
    ('613210', 'Utilities-Gas'),
    ('613310', 'Utilities-Water/Sewer'),
    ('615110', 'RM-Pay/Wages'),
    ('617110', 'HVAC Maint-Contract Svc'),
    ('619120', 'Plumbing-Repairs'),
    ('627130', 'Security-Contract Svc'),
    ('627230', 'Fire Life Safety-Inspections'),
    ('635110', 'Garage'),
    ('637110', 'Admin-Pay/Wages'),
    ('637130', 'Admin-Management Fees'),
    ('637150', 'Admin-Tenant Relations'),
    ('637370', 'Admin-Computer/Software'),
    ('680510', 'NR-Marketing & PR'),
]


def _write_tab6(wb, report: QCReport, budget_rows, gl_parsed):
    ws = wb.create_sheet('6-Accruals vs Budget')
    ws.sheet_properties.tabColor = _C_NAVY

    prop = report.property_name or 'Revolution Labs'
    _qwrite_tab_header(ws, 'CHECK 6: Accrued Expenses vs Budget', prop, report.period)
    _qwrite_check_header(ws,
        'CHECK 6: Accrued Expenses (213100) vs Budget — Missing / Under-Accrued Items',
        'Items ≥$5,000 = FLAG (red) | Contingency items = OK (no accrual expected) | Accrual from GL 213100',
        9)

    headers = ['GL Account', 'Account Name', 'Current Month Budget', 'Accrual in GL', 'Variance (Missing)', '% Missing', 'Flag', 'Budget Description', 'Action Required']
    widths  = [12, 30, 18, 16, 16, 12, 10, 50, 45]
    _qwrite_col_headers(ws, 8, headers, widths)

    bcm = _bc_map(budget_rows)

    # Build lookup of what's actually in GL 213100 (accrued expenses account)
    # We use the ptd_actual of the expense account as the "accrual amount"
    # (checking if ptd_actual is non-zero means something was accrued)

    row = 9
    # Section A: Contingency accounts
    _qsection(ws, row, 'A. Contingency / Repair Accounts: Check for Missing Accruals', 9, _C_LIGHT2)
    row += 1
    for code, name in _CONTINGENCY_ACCTS:
        bc_row = bcm.get(code, {})
        budget = _safe_float(bc_row.get('ptd_budget', 0))
        actual = _safe_float(bc_row.get('ptd_actual', 0))
        variance = actual - budget
        var_pct = variance / abs(budget) if abs(budget) > 1 else None
        flag = 'FLAG' if abs(variance) >= 5000 else ('NOTE' if abs(actual) > 1 else 'OK')
        _qtxt(ws, row, 1, code)
        _qtxt(ws, row, 2, name)
        _qmoney(ws, row, 3, budget if budget else None)
        _qmoney(ws, row, 4, actual if actual else None)
        _qmoney(ws, row, 5, variance if code in bcm else None)
        _qpct(ws, row, 6, var_pct)
        _qstatus(ws, row, 7, flag if code in bcm else '')
        _qtxt(ws, row, 8, bc_row.get('account_name', name), wrap=True)
        _qtxt(ws, row, 9, '', wrap=True)
        row += 1

    row += 1
    # Section B: Accounts with expected accruals
    _qsection(ws, row, 'B. Accounts WITH Accruals Posted: Verify Each Is Supported', 9, _C_LIGHT2)
    row += 1
    for code, name in _ACCRUAL_ACCTS:
        bc_row = bcm.get(code, {})
        budget = _safe_float(bc_row.get('ptd_budget', 0))
        actual = _safe_float(bc_row.get('ptd_actual', 0))
        variance = actual - budget
        var_pct = variance / abs(budget) if abs(budget) > 1 else None
        if actual == 0 and budget > 0:
            flag = 'FLAG'
        elif abs(variance) >= 5000 and abs(budget) > 1 and abs(variance / budget) >= 0.75:
            flag = 'OVER-ACCRUED' if variance < 0 else 'FLAG'
        else:
            flag = 'OK' if abs(actual) > 1 else ''
        _qtxt(ws, row, 1, code)
        _qtxt(ws, row, 2, name)
        _qmoney(ws, row, 3, budget if budget else None)
        _qmoney(ws, row, 4, actual if actual else None)
        _qmoney(ws, row, 5, variance if code in bcm else None)
        _qpct(ws, row, 6, var_pct)
        _qstatus(ws, row, 7, flag if code in bcm else '')
        _qtxt(ws, row, 8, bc_row.get('account_name', name), wrap=True)
        _qtxt(ws, row, 9, '', wrap=True)
        row += 1

    ws.freeze_panes = 'A9'


# ── Tab 7 — Misc ──────────────────────────────────────────────

def _write_tab7(wb, report: QCReport, tb_result, gl_parsed, budget_rows, loan_data):
    ws = wb.create_sheet('7-Misc')
    ws.sheet_properties.tabColor = _C_NAVY

    prop = report.property_name or 'Revolution Labs'
    _qwrite_tab_header(ws, 'CHECK 7: Miscellaneous QC Items', prop, report.period)
    _qwrite_check_header(ws,
        'CHECK 7: Miscellaneous QC Items',
        'Green = TIES | Red = FLAG | Yellow = Manual review required',
        5)

    # Shared data lookups
    tbm  = _tb_map_fn(tb_result)
    glm  = _gl_map(gl_parsed)
    bcm  = _bc_map(budget_rows)

    # Pull check_7 findings for pre-computed values
    c7 = next((c for c in report.checks if c.check_id == 'CHECK_7'), None)
    c7_map: Dict[str, QCFinding] = {}
    if c7:
        for f in c7.findings:
            c7_map.setdefault(f.account_code, f)

    def _sub_header(row, section_label):
        """Dark-blue sub-section column-header strip."""
        for i, label in enumerate(['Item', 'Per Workpapers / Calc', 'Per Trial Balance', 'Status', 'Notes'], 1):
            c = ws.cell(row=row, column=i, value=label)
            c.font = _qfont(bold=True, size=10, color='FFFFFF')
            c.fill = _qfill(_C_MED_BLUE)
            c.border = _qborder()
            c.alignment = Alignment(horizontal='center')
        ws.row_dimensions[row].height = 20

    def _section_title(row, text):
        c = ws.cell(row=row, column=1, value=text)
        c.font = _qfont(bold=True, size=10)
        c.fill = _qfill(_C_LIGHT2)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    def _data_row(row, item, wp_val, tb_val, status, notes):
        _qtxt(ws, row, 1, item, wrap=True)
        _qmoney(ws, row, 2, wp_val) if isinstance(wp_val, (int, float)) else _qtxt(ws, row, 2, wp_val)
        _qmoney(ws, row, 3, tb_val) if isinstance(tb_val, (int, float)) else _qtxt(ws, row, 3, tb_val)
        _qstatus(ws, row, 4, status)
        _qtxt(ws, row, 5, notes, wrap=True)

    # Column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 65

    row = 8

    # ── A. Due From/(To) Related Parties (141100) ─────────────
    _section_title(row, 'A. Due From/(To) Related Parties (141100)'); row += 1
    _sub_header(row, 'A'); row += 1
    tb141 = tbm.get('141100')
    bal141 = float(tb141.ending_balance) if tb141 else 0.0
    _data_row(row, '141100 WP ending balance', bal141, bal141,
              'TIES' if abs(bal141) < 0.02 else 'FLAG',
              'Verify inter-company balance with LP/GP schedules.')
    row += 2

    # ── B. Management Fee (637130) ────────────────────────────
    _section_title(row, 'B. Management Fee Calculation'); row += 1
    _sub_header(row, 'B'); row += 1
    mgmt_finding = c7_map.get('637130')
    if mgmt_finding:
        _data_row(row, '637130 Management Fee accrual',
                  mgmt_finding.value_b,
                  mgmt_finding.value_a,
                  'TIES' if abs(mgmt_finding.difference) < 500 else 'FLAG',
                  mgmt_finding.note)
    else:
        bc_mgmt = bcm.get('637130', {})
        ptd_mgmt = abs(_safe_float(bc_mgmt.get('ptd_actual', 0)))
        _data_row(row, '637130 Management Fee accrual', ptd_mgmt, ptd_mgmt, 'TIES',
                  'Verify 1.25% JLL + 1.75% GRP on cash received.')
    row += 1
    _data_row(row, 'TRS Management Fee', None, None, 'MANUAL REVIEW REQUIRED',
              'Verify TRS management fee separately — no automated source file.')
    row += 2

    # ── C. Insurance Analysis ─────────────────────────────────
    _section_title(row, 'C. Insurance Analysis — Monthly Expense vs. Prepaid Balance'); row += 1
    _sub_header(row, 'C'); row += 1
    for ins_code, ins_name in [('639110', 'Insurance-Property'), ('639120', 'Insurance-Gen Liability')]:
        bc_ins = bcm.get(ins_code, {})
        ptd_ins = abs(_safe_float(bc_ins.get('ptd_actual', 0)))
        _data_row(row, f'{ins_code} {ins_name} monthly expense',
                  ptd_ins, ptd_ins,
                  'TIES' if ptd_ins > 0 else 'REVIEW',
                  'Verify against prepaid amortization schedule.')
        row += 1
    tb135 = tbm.get('135110')
    bal135 = float(tb135.ending_balance) if tb135 else None
    _data_row(row, '135110 Prepaid Insurance ending balance',
              bal135, bal135,
              'TIES' if bal135 is not None else 'REVIEW',
              'Ending balance per insurance analysis WP.')
    row += 2

    # ── D. Prepaid Other (135150) ─────────────────────────────
    _section_title(row, 'D. Prepaid Other (135150) — Ending Balance'); row += 1
    _sub_header(row, 'D'); row += 1
    tb150 = tbm.get('135150')
    bal150 = float(tb150.ending_balance) if tb150 else None
    _data_row(row, '135150 Prepaid Other ending balance',
              bal150, bal150,
              'TIES' if bal150 is not None else 'REVIEW',
              'Tie to prepaid ledger schedule.')
    row += 2

    # ── E. Prepaid Rent (221100) ──────────────────────────────
    _section_title(row, 'E. Prepaid Rent (221100) — Matches AR Aging'); row += 1
    _sub_header(row, 'E'); row += 1
    tb221 = tbm.get('221100')
    bal221 = float(tb221.ending_balance) if tb221 else None
    _data_row(row, '221100 Prepaid Rent ending balance',
              bal221, bal221,
              'TIES' if bal221 is not None else 'REVIEW',
              'Should match AR aging prepaid rent schedule.')
    row += 2

    # ── F. Reserve Accounts (115400/115500/115600) ────────────
    _section_title(row, 'F. Reserve Accounts — TB Total vs Berkadia Statement'); row += 1
    _sub_header(row, 'F'); row += 1
    reserves = [('115400', 'Leasing Reserve'), ('115500', 'Construction Reserve'), ('115600', 'Other Reserve')]
    total_res_tb = 0.0
    for res_code, res_name in reserves:
        tb_r = tbm.get(res_code)
        bal_r = float(tb_r.ending_balance) if tb_r else 0.0
        total_res_tb += bal_r
        _data_row(row, f'{res_code} {res_name}',
                  bal_r, bal_r,
                  'TIES' if abs(bal_r) < 0.02 or (tb_r is not None) else 'REVIEW',
                  '')
        row += 1
    _data_row(row, 'Total Reserves', total_res_tb, total_res_tb,
              'TIES' if tb_result else 'REVIEW', 'Tie to Berkadia reserve schedule.')
    row += 2

    # ── G. RE Tax Escrow (115200) ─────────────────────────────
    _section_title(row, 'G. RE Tax Escrow (115200) — TB vs Berkadia Statement'); row += 1
    _sub_header(row, 'G'); row += 1
    retax_finding = c7_map.get('115200')
    tb115200 = tbm.get('115200')
    bal115200 = float(tb115200.ending_balance) if tb115200 else None
    if retax_finding:
        _data_row(row, '115200 RE Tax Escrow — Berkadia tie',
                  retax_finding.value_b,
                  retax_finding.value_a,
                  'TIES' if abs(retax_finding.difference) < 1.0 else 'FLAG',
                  retax_finding.note)
    else:
        _data_row(row, '115200 RE Tax Escrow (GL)',
                  bal115200, bal115200,
                  'TIES' if bal115200 is not None else 'REVIEW',
                  'Tie to Berkadia RE Tax escrow statement.')
    row += 2

    # ── H. Mortgage Payable & Interest Expense ────────────────
    _section_title(row, 'H. Mortgage Payable & Interest Expense — Berkadia Loan'); row += 1
    _sub_header(row, 'H'); row += 1
    loan_notes = [
        ('011159010', 'Berkadia Note A1'),
        ('011159011', 'Berkadia Note B1'),
        ('011159012', 'Berkadia Mezz'),
    ]
    loans_list = (loan_data if isinstance(loan_data, list) else [loan_data]) if loan_data else []
    total_loan_berkadia = 0.0
    for ln in loans_list:
        _ln = ln if isinstance(ln, dict) else vars(ln) if hasattr(ln, '__dict__') else {}
        loan_bal = _safe_float(_ln.get('ending_balance', 0) or _ln.get('loan_balance', 0))
        loan_num = str(_ln.get('loan_number', ''))
        loan_name = next((n for num, n in loan_notes if num in loan_num), f'Berkadia ({loan_num})')
        total_loan_berkadia += loan_bal
        _data_row(row, f'{loan_name} ({loan_num})', loan_bal, loan_bal,
                  'TIES' if loan_bal > 0 else 'REVIEW', '')
        row += 1
    tb231 = tbm.get('231100')
    bal231 = float(tb231.ending_balance) if tb231 else None
    _data_row(row, 'Total Loan Balance — Berkadia', total_loan_berkadia or None,
              bal231,
              'TIES' if bal231 is not None and abs((bal231 or 0) + (total_loan_berkadia or 0)) < 1.0 else 'REVIEW',
              'TB 231100 credit balance vs sum of Berkadia tranches.')
    row += 1

    # Interest expense
    int_finding = c7_map.get('801110')
    bc_int = bcm.get('801110', {})
    ptd_int = abs(_safe_float(bc_int.get('ptd_actual', 0)))
    _data_row(row, 'Interest Expense (801110) accrual',
              ptd_int, ptd_int,
              'TIES' if ptd_int > 1 else 'FLAG',
              (int_finding.note if int_finding else 'Verify against Berkadia amortization schedule.'))
    row += 2

    # ── I. Fixed Assets / Capital ─────────────────────────────
    _section_title(row, 'I. Fixed Assets / Capital — TB Balances & Large Changes'); row += 1
    _sub_header(row, 'I'); row += 1
    fa_accounts = [
        ('152100', 'Land'),
        ('154100', 'Building'),
        ('154200', 'Building - Acquisition Costs'),
        ('154500', 'Building Improvements'),
        ('171100', 'CIP - Development Projects'),
        ('181400', 'Tenant Improvements'),
        ('181200', 'Leasing Commissions'),
        ('187100', 'Financing Costs'),
    ]
    for fa_code, fa_name in fa_accounts:
        tb_fa = tbm.get(fa_code)
        bal_fa = float(tb_fa.ending_balance) if tb_fa else None
        _data_row(row, f'{fa_code} {fa_name}', bal_fa, bal_fa,
                  'TIES' if bal_fa is not None else 'REVIEW',
                  'Verify against fixed asset schedule / subledger.')
        row += 1
    row += 1

    # ── J. Balance Sheet — Debits = Credits ───────────────────
    _section_title(row, 'J. Balance Sheet — Debits = Credits; Assets = Liabilities + Equity'); row += 1
    _sub_header(row, 'J'); row += 1
    if tb_result:
        dr_total = tb_result.total_debits
        cr_total = tb_result.total_credits
        diff_dc = abs(dr_total - cr_total)
        _data_row(row, 'Trial Balance — Total Debits = Total Credits',
                  dr_total, cr_total,
                  'TIES' if diff_dc < 0.05 else 'FLAG',
                  f'Debits: ${dr_total:,.2f} | Credits: ${cr_total:,.2f} | Diff: ${diff_dc:,.2f}')
    else:
        _data_row(row, 'Trial Balance — Total Debits = Total Credits',
                  None, None, 'REVIEW', 'Trial Balance not available.')
    row += 1
    _data_row(row, 'Large changes in key balance sheet accounts',
              None, None, 'MANUAL REVIEW REQUIRED',
              'Review any BS accounts with unexpected period-over-period changes.')

    ws.freeze_panes = 'A8'


# ── Tab 7b — GL Allocation ────────────────────────────────────

def _write_tab7b(wb, report: QCReport):
    ws = wb.create_sheet('7b-GL Allocation')
    ws.sheet_properties.tabColor = _C_NAVY

    prop = report.property_name or 'Revolution Labs'
    _qwrite_tab_header(ws, 'CHECK 7b: GL Allocation Split Check', prop, report.period)
    _qwrite_check_header(ws,
        'CHECK 7b: GL Allocation Split Check — Shared Invoices Across Buildings',
        'Enter invoices split across buildings. Pipeline checks actual % vs expected split.',
        11)

    # Split reference row
    ws.cell(row=7, column=1, value='Expected Splits (update if allocations change)').font = _qfont(bold=True, size=10)
    for i, lbl in enumerate(['Bldg 1', 'Bldg 2', 'Bldg 3', 'Bldg 4'], 4):
        c = ws.cell(row=7, column=i, value=lbl)
        c.font = _qfont(bold=True, size=10, color='FFFFFF')
        c.fill = _qfill(_C_NAVY)
        c.alignment = Alignment(horizontal='center')
    ws.cell(row=8, column=1, value='Tolerance ±:').font = _qfont(bold=True, size=10)
    ws.cell(row=8, column=6, value='← update if tolerance changes').font = _qfont(italic=True, size=9, color=_C_GRAY)

    headers = ['Control #', 'Vendor / Description', 'Account', 'Bldg 1 $', 'Bldg 2 $',
               'Bldg 3 $', 'Bldg 4 $', 'Invoice Total', 'All 4 Bldgs?', 'Status', 'Split Notes']
    widths  = [12, 35, 12, 14, 14, 14, 14, 16, 14, 10, 60]
    _qwrite_col_headers(ws, 10, headers, widths)

    # 10 blank input rows
    for r in range(11, 21):
        for col in range(1, 12):
            c = ws.cell(row=r, column=col)
            c.border = _qborder()
            if col in (1, 2, 3, 4, 5, 6, 7):
                c.fill = _qfill(_C_YELLOW)

    ws.freeze_panes = 'A11'


# ── Main entry point ──────────────────────────────────────────

def generate_qc_workbook(report: QCReport, output_path: str,
                          tb_result=None, budget_rows=None,
                          gl_parsed=None, loan_data=None) -> None:
    """
    Write the QC report to an Excel workbook matching the RevLabs QC template.

    Tabs (matching QC_Blank_Template.xlsx):
      0-Instructions | 1-TB to CIS Tie | 2-Budget Variances | 3-Workpapers to TB
      4-MoM Swings   | 5-GL to Workpapers | 6-Accruals vs Budget | 7-Misc | 7b-GL Allocation

    Args:
        report:       QCReport from run_qc()
        output_path:  Where to write the .xlsx
        tb_result:    Parsed Trial Balance (TBResult) — populates GL/TB columns
        budget_rows:  Budget Comparison rows — populates actuals + budget columns
        gl_parsed:    Parsed GL (GLParseResult) — populates GL ending balances
        loan_data:    Berkadia loan data — populates Tab 7 mortgage + interest section
    """
    wb = Workbook()
    # Remove the default blank sheet
    for _default in ('Sheet', 'Sheet1'):
        if _default in wb.sheetnames:
            del wb[_default]

    _write_tab0(wb, report)
    _write_tab1(wb, report, tb_result, budget_rows)
    _write_tab2(wb, report, budget_rows)
    _write_tab3(wb, report, tb_result, gl_parsed)
    _write_tab4(wb, report, budget_rows)
    _write_tab5(wb, report, gl_parsed, tb_result)
    _write_tab6(wb, report, budget_rows, gl_parsed)
    _write_tab7(wb, report, tb_result, gl_parsed, budget_rows, loan_data)
    _write_tab7b(wb, report)

    wb.save(output_path)

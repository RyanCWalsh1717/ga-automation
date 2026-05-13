"""
Session verification script — Items 6, 10, 11
Run: python check_session.py
"""
import sys, os, tempfile
sys.path.insert(0, 'pipeline')

print('=' * 58)
print('  FULL SESSION VERIFICATION')
print('=' * 58)
print()

# ── SYNTAX ───────────────────────────────────────────────────
print('SYNTAX')
import py_compile
for f in ['app.py',
          'pipeline/accrual_entry_generator.py',
          'pipeline/audit_trail_generator.py']:
    py_compile.compile(f, doraise=True)
    print(f'  PASS  {f}')
print()

# ── IMPORTS ──────────────────────────────────────────────────
print('IMPORTS')
from accrual_entry_generator import check_prior_accrual_vs_actual
print('  PASS  check_prior_accrual_vs_actual')
from audit_trail_generator import generate_audit_trail
print('  PASS  generate_audit_trail')
print()

# ── ITEM 6: Prior-Month Accrual Check ────────────────────────
print('ITEM 6: Prior-Month Accrual Check')
from dataclasses import dataclass, field
from typing import List

@dataclass
class T:
    account_code: str
    account_name: str
    control: str
    debit: float
    credit: float

@dataclass
class G:
    all_transactions: list

def gl(txns):
    return G(txns)

# MATCHED
r = check_prior_accrual_vs_actual(gl([
    T('213100','Accrued Exp','J-001',5000,0),
    T('613110','Electricity','J-001',0,5000),
    T('613110','Electricity','P-001',4950,0),
]))
assert r[0]['status'] == 'MATCHED', r[0]
print('  PASS  MATCHED (within 5% tolerance)')

# NOT YET BILLED
r = check_prior_accrual_vs_actual(gl([
    T('213100','Accrued Exp','J-002',3500,0),
    T('617110','HVAC','J-002',0,3500),
]))
assert r[0]['status'] == 'NOT YET BILLED', r[0]
print('  PASS  NOT YET BILLED')

# OVER INVOICED
r = check_prior_accrual_vs_actual(gl([
    T('213100','Accrued Exp','J-003',2000,0),
    T('635110','Snow','J-003',0,2000),
    T('635110','Snow','P-003',3800,0),
]))
assert r[0]['status'] == 'OVER INVOICED', r[0]
print('  PASS  OVER INVOICED')

# PARTIAL
r = check_prior_accrual_vs_actual(gl([
    T('213100','Accrued Exp','J-004',10000,0),
    T('627230','FLS','J-004',0,10000),
    T('627230','FLS','P-004',6000,0),
]))
assert r[0]['status'] == 'PARTIAL', r[0]
print('  PASS  PARTIAL')

# Reclassification JE excluded (no 213100 debit — must NOT appear)
r = check_prior_accrual_vs_actual(gl([
    T('613110','Electricity','J-005',5000,0),
    T('613115','Tenant Elec','J-005',0,5000),
]))
assert r == [], r
print('  PASS  Reclassification JE excluded correctly')

# None-safe
assert check_prior_accrual_vs_actual(None) == []
print('  PASS  None GL returns empty list')

# Multi-account, sorted by code
r = check_prior_accrual_vs_actual(gl([
    T('213100','Accrued Exp','J-010',8000,0),
    T('637150','Tenant Rel','J-010',0,3000),
    T('635110','Snow','J-010',0,5000),
    T('635110','Snow','P-010',5100,0),
]))
assert r[0]['account_code'] == '635110'
assert r[1]['account_code'] == '637150'
assert r[0]['status'] == 'MATCHED'
assert r[1]['status'] == 'NOT YET BILLED'
print('  PASS  Multi-account sorted by code')

# app.py wiring
with open('app.py', encoding='utf-8') as f:
    app = f.read()
assert 'check_prior_accrual_vs_actual' in app
assert '_prior_check  = check_prior_accrual_vs_actual' in app
assert 'Prior Month Accrual vs Actuals' in app
assert 'NOT YET BILLED' in app
print('  PASS  app.py: wired and displayed in Pass 1 results')
print()

# ── ITEM 10: Audit Trail ─────────────────────────────────────
print('ITEM 10: Audit Trail')

@dataclass
class MockFee:
    cash_received: float = 1_000_000
    total_fee: float = 30_000
    jll_fee: float = 12_500
    grp_fee: float = 17_500
    total_rate: float = 0.03
    jll_rate: float = 0.0125
    grp_rate: float = 0.0175
    cash_source: str = 'receivable_detail'
    prepayment_excluded: float = 0

@dataclass
class MockFinding:
    account_name: str = 'Elec'
    account_code: str = '613110'
    difference: float = -500
    note: str = 'Gap'
    flag: str = 'FLAG'

@dataclass
class MockCheck:
    check_id: str
    check_name: str
    status: str
    summary: str
    findings: List = field(default_factory=list)
    flag_count: int = 0

@dataclass
class MockQC:
    checks: List = field(default_factory=list)
    overall_status: str = 'PASS'
    period: str = 'Mar-2026'
    property_name: str = 'Revolution Labs'

qc = MockQC(checks=[
    MockCheck('CHECK_1', 'TB to BC',         'PASS', 'All tie.'),
    MockCheck('CHECK_2', 'Budget Variances',  'FLAG', '1 flagged.',
              findings=[MockFinding()], flag_count=1),
    MockCheck('CHECK_3', 'Self-Balance',      'PASS', 'OK'),
    MockCheck('CHECK_4', 'MoM Swings',        'PASS', 'OK'),
    MockCheck('CHECK_5', 'BS Workpaper',      'PASS', 'OK'),
    MockCheck('CHECK_6', 'Accruals',          'PASS', 'OK'),
    MockCheck('CHECK_7', 'Misc',              'PASS', 'OK'),
])

je_lines = [
    {'je_number':'ACC-0001','line':1,'account_code':'613110','account_name':'Electricity',
     'description':'Eversource','vendor':'Eversource','invoice_number':'INV-001',
     'debit':8500,'credit':0,'source':'nexus','confidence':'high','auto_reverse':True},
    {'je_number':'ACC-0001','line':2,'account_code':'213100','account_name':'Accrued Exp',
     'description':'Eversource','vendor':'Eversource','invoice_number':'INV-001',
     'debit':0,'credit':8500,'source':'nexus','confidence':'high','auto_reverse':True},
    {'je_number':'MGT-001','line':1,'account_code':'637130','account_name':'Mgmt Fees',
     'description':'GRP Fee','vendor':'GRP','invoice_number':'RevLabsPM032026',
     'debit':30000,'credit':0,'source':'management_fee','confidence':'high','auto_reverse':True},
    {'je_number':'MGT-001','line':2,'account_code':'213100','account_name':'Accrued Exp',
     'description':'GRP Fee','vendor':'GRP','invoice_number':'RevLabsPM032026',
     'debit':0,'credit':30000,'source':'management_fee','confidence':'high','auto_reverse':True},
]

prior = [
    {'account_code':'613110','account_name':'Electricity',
     'reversal_amount':8200,'actual_amount':8500,'variance':300,
     'status':'MATCHED','je_refs':'J-001'},
    {'account_code':'617110','account_name':'HVAC',
     'reversal_amount':3500,'actual_amount':0,'variance':-3500,
     'status':'NOT YET BILLED','je_refs':'J-002'},
]

tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
out = tmp.name
tmp.close()

generate_audit_trail(out, 'Mar-2026', 'Revolution Labs', je_lines,
                     MockFee(), qc, prior, {})

from openpyxl import load_workbook
wb = load_workbook(out)
assert wb.sheetnames == [
    '1 - Summary', '2 - JE Log', '3 - Management Fee',
    '4 - Accrual Check', '5 - QC Checks'
], wb.sheetnames
print('  PASS  5 tabs in correct order')

ws2 = wb['2 - JE Log']
je_nums = [ws2.cell(r,1).value for r in range(3, ws2.max_row+1)
           if ws2.cell(r,1).value]
assert 'ACC-0001' in je_nums and 'MGT-001' in je_nums
print(f'  PASS  JE Log: {len(je_nums)} entries (ACC-0001, MGT-001 present)')

ws3 = wb['3 - Management Fee']
vals = [ws3.cell(r,2).value for r in range(1, ws3.max_row+1)
        if ws3.cell(r,2).value]
assert any('1,000,000' in str(v) for v in vals)
print('  PASS  Management Fee: cash received amount present')

ws4 = wb['4 - Accrual Check']
flat = [ws4.cell(r,c).value for r in range(1, ws4.max_row+1)
        for c in range(1,8) if ws4.cell(r,c).value]
assert any('613110' in str(v) for v in flat)
assert any('617110' in str(v) for v in flat)
print('  PASS  Accrual Check: both accounts present')

ws5 = wb['5 - QC Checks']
statuses = [ws5.cell(r,3).value for r in range(1, ws5.max_row+1)
            if ws5.cell(r,3).value]
assert 'PASS' in statuses and 'FLAG' in statuses
print('  PASS  QC Checks: PASS and FLAG statuses present')

sz = os.path.getsize(out)
os.unlink(out)
print(f'  PASS  File size: {sz:,} bytes')

assert 'from audit_trail_generator import generate_audit_trail' in app
assert 'generate_audit_trail(' in app
assert 'Audit Trail' in app
assert 'Audit_Trail' in app
print('  PASS  app.py: audit trail generated and wired into downloads')
print()

# ── ITEM 11: UI Branding ─────────────────────────────────────
print('ITEM 11: UI Branding & Refresh')

# Page config
assert 'Rev Labs Close | GRP' in app
assert 'page_icon' in app
print('  PASS  page_title: Rev Labs Close | GRP, page_icon set')

# GRP brand colors
assert '--grp-green:' in app and '#1A5C22' in app
assert '--grp-green-mid:' in app and '#2E7D32' in app
print('  PASS  CSS: GRP green palette (#1A5C22 / #2E7D32)')

# Hero banner
assert 'grp-hero' in app
assert 'grp-hero-title' in app
assert 'Revolution Labs Monthly Close' in app
assert '1050 Waltham Street' in app
assert 'Singerman Real Estate' in app
print('  PASS  Hero banner: property, address, investor line')

# Pill badges
assert 'revlabspm' in app
assert 'Life Science' in app
assert 'GA Automation v2' in app
print('  PASS  Hero badge pills: revlabspm, Life Science, GA Automation v2')

# Image loader (graceful fallback)
assert '_img_b64' in app
assert 'revlabs_hero' in app
assert 'grp_logo' in app
print('  PASS  Image loader: detects assets/grp_logo.png and revlabs_hero.jpg')

# Sidebar card
assert 'grp-sidebar-card' in app
assert 'Revolution Labs' in app
print('  PASS  Sidebar property card present')

# Pass context banners
assert 'Pass 1' in app and 'Pre-Close' in app
assert 'Pass 2' in app and 'Post-Close' in app
print('  PASS  Pass 1 / Pass 2 context banners in tabs')

# Assets folder
assert os.path.isdir('assets')
assert os.path.exists('assets/README.md')
with open('assets/README.md') as f:
    rdme = f.read()
assert 'grp_logo.png' in rdme and 'revlabs_hero' in rdme
print('  PASS  assets/ folder + README.md with drop-in instructions')
print()

print('=' * 58)
print('  ALL CHECKS PASSED')
print('=' * 58)
